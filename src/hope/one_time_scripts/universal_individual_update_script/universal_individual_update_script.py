from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.core import BusinessArea
from models.geo import Country
from hope.apps.household.documents import HouseholdDocument, get_individual_doc
from models.household import Document, DocumentType, Household, Individual
from hope.apps.payment.models import Account, AccountType
from models.program import Program
from hope.apps.registration_datahub.tasks.deduplicate import (
    DeduplicateTask,
    HardDocumentDeduplication,
)
from hope.apps.utils.elasticsearch_utils import populate_index


class UniversalIndividualUpdateScript:
    def __init__(  # noqa
        self,
        business_area: BusinessArea,
        program: Program,
        file_path: str,
        household_fields: dict[str, tuple[str, Any, Any]] | None = None,
        individual_fields: dict[str, tuple[str, Any, Any]] | None = None,
        individual_flex_fields: dict[str, tuple[str, Any, Any]] | None = None,
        document_fields: list[tuple[str, str]] | None = None,
        deliver_mechanism_data_fields: dict[str, tuple[list[str], ...]] | None = None,
        ignore_empty_values: bool = True,
        deduplicate_es: bool = True,
        deduplicate_documents: bool = True,
        batch_size: int = 100,
    ) -> None:
        self.business_area = business_area
        self.program = program
        self.file_path = file_path
        self.household_fields = household_fields
        self.individual_fields = individual_fields
        self.individual_flex_fields = individual_flex_fields
        self.document_fields = document_fields
        self.deliver_mechanism_data_fields = deliver_mechanism_data_fields
        self.ignore_empty_values = ignore_empty_values
        self.deduplicate_es = deduplicate_es
        self.deduplicate_documents = deduplicate_documents
        document_types = DocumentType.objects.filter()
        self.countries = {country.name: country for country in Country.objects.all()}
        self.document_types = {
            f"{document_type.key}_no_i_c": document_type
            for document_type in document_types
        }
        self.delivery_mechanisms_account_types = {}
        if deliver_mechanism_data_fields is not None:
            self.delivery_mechanisms_account_types = {
                account_type.key: account_type
                for account_type in AccountType.objects.filter(
                    key__in=self.deliver_mechanism_data_fields.keys()
                )
            }
        self.batch_size = batch_size

    def validate_household_fields(
        self, row: tuple[Any, ...], headers: list[str], household: Any, row_index: int
    ) -> list[str]:
        if self.household_fields is None:
            return []
        errors = []
        for field, (name, validator, _handler) in self.household_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, row, self.business_area, self.program)
            if error:
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_individual_fields(
        self,
        row: tuple[Any, ...],
        headers: list[str],
        individual: Individual,
        row_index: int,
    ) -> list[str]:
        if self.individual_fields is None:
            return []
        errors = []
        for field, (name, validator, _handler) in self.individual_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, row, self.business_area, self.program)
            if error:  # pragma: no cover
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_individual_flex_fields(
        self,
        row: tuple[Any, ...],
        headers: list[str],
        individual: Individual,
        row_index: int,
    ) -> list[str]:
        errors = []
        if self.individual_flex_fields is None:
            return []
        for field, (name, validator, _handler) in self.individual_flex_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, row, self.business_area, self.program)
            if error:  # pragma: no cover
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_documents(
        self,
        row: tuple[Any, ...],
        headers: list[str],
        individual: Individual,
        row_index: int,
    ) -> list[str]:
        if self.document_fields is None:
            return []
        errors = []
        for number_column_name, country_column_name in self.document_fields:
            document_type = self.document_types.get(number_column_name)
            number_text = row[headers.index(number_column_name)]
            country_text = row[headers.index(country_column_name)]
            if country_text is None and number_text is None:
                continue
            country = Country.objects.filter(name=country_text).first()
            if country is None:
                errors.append(
                    f"Row: {row_index} - Country not found for field {country_column_name} and value {country_text}"
                )
            if document_type is None:
                errors.append(
                    f"Row: {row_index} - Document type not found for field {number_column_name}"
                )
            if (
                individual.documents.filter(type=document_type).count() > 1
            ):  # pragma: no cover
                errors.append(
                    f"Row: {row_index} - Multiple documents with document type {document_type} found"
                )
        return errors

    def validate(self, sheet: Worksheet, headers: list[str]) -> list[str]:
        errors = []
        row_index = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_index += 1
            if (row_index - 2) % self.batch_size == 0:
                print(
                    f"Validating row {row_index - 2} to {row_index - 2 + self.batch_size} Indivduals"
                )
            unicef_id = row[headers.index("unicef_id")]
            individuals_queryset = Individual.objects.filter(
                unicef_id=unicef_id,
                business_area=self.business_area,
                program=self.program,
            )
            if not individuals_queryset.exists():  # pragma: no cover
                errors.append(
                    f"Row: {row_index} - Individual with unicef_id {unicef_id} not found"
                )
                continue
            if individuals_queryset.count() > 1:  # pragma: no cover
                errors.append(
                    f"Row: {row_index} - Multiple individuals with unicef_id {unicef_id} found"
                )
                continue
            individual: Individual = individuals_queryset.first()  # type: ignore
            household = individual.household
            if household is None:  # pragma: no cover
                errors.append(
                    f"Row: {row_index} - Household not found for individual with unicef_id {unicef_id}"
                )
                continue
            errors.extend(
                self.validate_household_fields(row, headers, household, row_index)
            )
            errors.extend(
                self.validate_individual_fields(row, headers, individual, row_index)
            )
            errors.extend(
                self.validate_individual_flex_fields(
                    row, headers, individual, row_index
                )
            )
            errors.extend(self.validate_documents(row, headers, individual, row_index))
        return errors

    def handle_household_update(
        self, row: tuple[Any, ...], headers: list[str], household: Any
    ) -> None:
        if self.household_fields is None:
            return
        for field, (_name, _validator, handler) in self.household_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(
                value, field, household, self.business_area, self.program
            )
            if self.ignore_empty_values and (
                handled_value is None or handled_value == ""
            ):  # pragma: no cover
                continue
            setattr(household, _name, handled_value)

    def handle_individual_update(
        self, row: tuple[Any, ...], headers: list[str], individual: Individual
    ) -> None:
        if self.individual_fields is None:
            return
        for field, (_name, _validator, handler) in self.individual_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(
                value, field, individual, self.business_area, self.program
            )
            if self.ignore_empty_values and (
                handled_value is None or handled_value == ""
            ):  # pragma: no cover
                continue
            setattr(individual, _name, handled_value)

    def handle_individual_flex_update(
        self, row: tuple[Any, ...], headers: list[str], individual: Individual
    ) -> None:
        if self.individual_flex_fields is None:
            return
        for field, (name, _validator, handler) in self.individual_flex_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(
                value, field, individual, self.business_area, self.program
            )
            if self.ignore_empty_values and (
                handled_value is None or handled_value == ""
            ):  # pragma: no cover
                continue
            individual.flex_fields[name] = handled_value

    def handle_documents_update(
        self, row: tuple[Any, ...], headers: list[str], individual: Individual
    ) -> tuple[list, list]:
        documents_to_update = []
        documents_to_create = []
        if self.document_fields is None:
            return documents_to_update, documents_to_create
        for number_column_name, country_column_name in self.document_fields:
            document_type = self.document_types.get(number_column_name)
            document_number = row[headers.index(number_column_name)]
            document_country = row[headers.index(country_column_name)]
            if self.ignore_empty_values and (
                document_number is None or document_number == ""
            ):  # pragma: no cover
                continue
            country = self.countries[document_country]
            document = None
            for doc in individual.documents.all():
                if doc.type == document_type:
                    document = doc
                    break
            if document:
                document.document_number = document_number
                document.status = Document.STATUS_PENDING
                document.country = country
                documents_to_update.append(document)
            else:
                documents_to_create.append(
                    Document(
                        individual=individual,
                        type=document_type,
                        document_number=document_number,
                        country=country,
                        rdi_merge_status="MERGED",
                    )
                )
        return documents_to_update, documents_to_create

    def handle_delivery_mechanism_data_update(
        self, row: tuple[Any, ...], headers: list[str], individual: Individual
    ) -> None:
        if self.deliver_mechanism_data_fields is None:
            return
        individual_accounts = individual.accounts.all()
        for (
            account_type,
            delivery_mechanism_columns_mapping,
        ) in self.deliver_mechanism_data_fields.items():
            account_type_instance = self.delivery_mechanisms_account_types.get(
                account_type
            )
            single_data_object = next(
                (
                    d
                    for d in individual_accounts
                    if str(d.account_type.key) == str(account_type)
                ),
                None,
            )

            for column_name, field_name in delivery_mechanism_columns_mapping:
                if column_name in headers:
                    value = row[headers.index(column_name)]
                    if self.ignore_empty_values and (value is None or value == ""):
                        continue
                    if single_data_object is None:
                        single_data_object = Account(
                            individual=individual,
                            account_type=account_type_instance,
                            rdi_merge_status=Account.MERGED,
                        )
                    single_data_object.data[field_name] = value
            if single_data_object:
                single_data_object.save()
                single_data_object.update_unique_field()

    def handle_update(self, sheet: Worksheet, headers: list[str]) -> list[str]:
        row_index = 1
        individual_ids = []
        household_fields_to_update = ["flex_fields"]
        individual_fields_to_update = ["flex_fields"]
        document_fields_to_update = ["document_number", "status", "country"]
        if self.household_fields:
            household_fields_to_update.extend(
                [field for _, (field, _, _) in self.household_fields.items()]
            )
        if self.individual_fields:
            individual_fields_to_update.extend(
                [field for _, (field, _, _) in self.individual_fields.items()]
            )
        individuals_to_update = []
        households_to_update = []
        documents_to_update = []
        documents_to_create = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_index += 1
            if (row_index - 2) % self.batch_size == 0:
                print(
                    f"Updating row {row_index - 2} to {row_index - 2 + self.batch_size} Individuals"
                )
            unicef_id = row[headers.index("unicef_id")]
            individual = (
                Individual.objects.select_related("household")
                .prefetch_related("documents", "accounts")
                .get(
                    unicef_id=unicef_id,
                    business_area=self.business_area,
                    program=self.program,
                )
            )
            individual_ids.append(str(individual.id))
            household = individual.household
            self.handle_household_update(row, headers, household)
            self.handle_individual_update(row, headers, individual)
            self.handle_individual_flex_update(row, headers, individual)
            (
                documents_to_update_part,
                documents_to_create_part,
            ) = self.handle_documents_update(row, headers, individual)
            documents_to_update.extend(documents_to_update_part)
            documents_to_create.extend(documents_to_create_part)
            self.handle_delivery_mechanism_data_update(row, headers, individual)
            households_to_update.append(household)
            individuals_to_update.append(individual)
            if len(individuals_to_update) == self.batch_size:
                self.batch_update(
                    document_fields_to_update,
                    documents_to_create,
                    documents_to_update,
                    household_fields_to_update,
                    households_to_update,
                    individual_fields_to_update,
                    individuals_to_update,
                )
        self.batch_update(
            document_fields_to_update,
            documents_to_create,
            documents_to_update,
            household_fields_to_update,
            households_to_update,
            individual_fields_to_update,
            individuals_to_update,
        )
        return individual_ids

    def batch_update(  # noqa
        self,
        document_fields_to_update: list,
        documents_to_create: list,
        documents_to_update: list,
        household_fields_to_update: list,
        households_to_update: list,
        individual_fields_to_update: list,
        individuals_to_update: list,
    ) -> None:
        Document.objects.bulk_update(documents_to_update, document_fields_to_update)
        Document.objects.bulk_create(documents_to_create)
        Household.objects.bulk_update(households_to_update, household_fields_to_update)
        Individual.objects.bulk_update(
            individuals_to_update, individual_fields_to_update
        )
        populate_index(
            Individual.objects.filter(
                id__in=[individual.id for individual in individuals_to_update]
            ),
            get_individual_doc(self.business_area.slug),
        )
        populate_index(
            Household.objects.filter(
                id__in=[household.id for household in households_to_update]
            ),
            HouseholdDocument,
        )
        documents_to_update.clear()
        documents_to_create.clear()
        households_to_update.clear()
        individuals_to_update.clear()

    def execute(self) -> None:
        workbook = load_workbook(filename=self.file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        errors = self.validate(sheet, headers)
        if errors:
            print("Validation failed")
            for error in errors:
                print(error)
            return
        print("Validation successful")
        processed_individuals_ids = self.handle_update(sheet, headers)
        if self.deduplicate_es:
            print("Deduplicating individuals Elasticsearch")
            DeduplicateTask(
                self.business_area.slug, self.program.id
            ).deduplicate_individuals_against_population(
                Individual.objects.filter(id__in=processed_individuals_ids)
            )
        if self.deduplicate_documents:
            print("Deduplicating documents")
            HardDocumentDeduplication().deduplicate(
                Document.objects.filter(
                    individual__id__in=processed_individuals_ids,
                    status=Document.STATUS_PENDING,
                ),
                program=self.program,
            )

        print("Update successful")
