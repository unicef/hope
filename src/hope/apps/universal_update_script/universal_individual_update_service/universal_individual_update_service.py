import logging
from io import BytesIO
from typing import Any

import openpyxl
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hope.models.geo import Country
from hope.apps.household.documents import HouseholdDocument, get_individual_doc
from hope.models.household import Document, DocumentType, Household, Individual
from hope.apps.payment.models import Account, AccountType
from hope.apps.registration_datahub.tasks.deduplicate import (
    DeduplicateTask,
    HardDocumentDeduplication,
)
from hope.models.universal_update_script import UniversalUpdate
from hope.apps.universal_update_script.universal_individual_update_service.all_updatable_fields import (
    get_account_fields,
    get_document_fields,
    get_household_flex_fields,
    get_individual_flex_fields,
    household_fields,
    individual_fields,
)
from hope.apps.universal_update_script.universal_individual_update_service.validator_and_handlers import (
    get_generator_handler,
)
from hope.apps.utils.elasticsearch_utils import populate_index

logger = logging.getLogger(__name__)


class UniversalIndividualUpdateService:
    def __init__(
        self,
        universal_update: UniversalUpdate,
        ignore_empty_values: bool = True,
        deduplicate_es: bool = True,
        deduplicate_documents: bool = True,
        batch_size: int = 100,
    ) -> None:
        self.universal_update = universal_update
        self.business_area = universal_update.program.business_area
        self.program = universal_update.program
        self.file_path = universal_update.update_file
        document_types = DocumentType.objects.filter()
        self.document_types = {f"{document_type.key}_no_i_c": document_type for document_type in document_types}
        self.household_fields = {
            column_name: data
            for column_name, data in household_fields.items()
            if data[0] in universal_update.household_fields
        }
        self.individual_fields = {
            column_name: data
            for column_name, data in individual_fields.items()
            if data[0] in universal_update.individual_fields
        }
        self.individual_flex_fields = {
            column_name: data
            for column_name, data in get_individual_flex_fields().items()
            if data[0] in universal_update.individual_flex_fields_fields
        }
        self.household_flex_fields = {
            column_name: data
            for column_name, data in get_household_flex_fields().items()
            if data[0] in universal_update.household_flex_fields_fields
        }
        self.document_fields = []
        for document_no_column_name, _ in get_document_fields():
            key = self.document_types[document_no_column_name].key
            if key in universal_update.document_types.values_list("key", flat=True):
                self.document_fields.append((document_no_column_name, f"{key}_country_i_c"))
        self.account_data_fields = {
            account_type: data
            for account_type, data in get_account_fields().items()
            if account_type in universal_update.account_types.values_list("key", flat=True)
        }
        self.ignore_empty_values = ignore_empty_values
        self.deduplicate_es = deduplicate_es
        self.deduplicate_documents = deduplicate_documents
        self.countries = {country.name: country for country in Country.objects.all()}
        self.delivery_mechanisms_account_types = {at.key: at for at in AccountType.objects.all()}
        self.batch_size = batch_size

    def print_message(self, text: str) -> None:
        self.universal_update.save_logs(text)
        logger.info(text)

    def validate_household_fields(
        self, row: tuple[Any, ...], headers: list[str], household: Any, row_index: int
    ) -> list[str]:
        if self.household_fields is None:
            return []  # pragma: no cover
        errors = []
        for field, (name, validator, _handler) in self.household_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, Household, self.business_area, self.program)
            if error:
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_individual_fields(self, row: tuple[Any, ...], headers: list[str], row_index: int) -> list[str]:
        errors = []
        for field, (name, validator, _handler) in self.individual_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, Individual, self.business_area, self.program)
            if error:  # pragma: no cover
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_individual_flex_fields(self, row: tuple[Any, ...], headers: list[str], row_index: int) -> list[str]:
        errors = []
        for field, (name, validator, _handler) in self.individual_flex_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, row, self.business_area, self.program)
            if error:  # pragma: no cover
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_household_flex_fields(self, row: tuple[Any, ...], headers: list[str], row_index: int) -> list[str]:
        errors = []
        for field, (name, validator, _handler) in self.household_flex_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, row, self.business_area, self.program)
            if error:  # pragma: no cover
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_documents(
        self,
        row: tuple[Any, ...],
        headers: list[str],
        individual: Individual,
        row_index: int,
    ) -> list[str]:
        errors = []
        for number_column_name, country_column_name in self.document_fields:
            document_type = self.document_types.get(number_column_name)
            number_text = row[headers.index(number_column_name)]
            country_text = row[headers.index(country_column_name)]
            if country_text is None and number_text is None:
                continue
            country = Country.objects.filter(name=country_text).first()
            if country is None:
                errors.append(
                    f"Row: {row_index} - Country not found for field {country_column_name} and value {country_text}"
                )
            if document_type is None:
                errors.append(
                    f"Row: {row_index} - Document type not found for field {number_column_name}"
                )  # pragma: no cover
            if individual.documents.filter(type=document_type).count() > 1:  # pragma: no cover
                errors.append(f"Row: {row_index} - Multiple documents with document type {document_type} found")
        return errors

    def validate_accounts(
        self,
        row: tuple[Any, ...],
        headers: list[str],
        individual: Individual,
        row_index: int,
    ) -> list[str]:
        errors = []
        for account_type, account_columns_mapping in self.account_data_fields.items():
            columns_from_header = self.get_all_account_columns(
                account_type, headers, [x for x, y in account_columns_mapping]
            )  # this gets columns starting with account__{account_type}__ and not in the mapping
            all_account_fields = []
            all_account_fields.extend(account_columns_mapping)
            all_account_fields.extend(columns_from_header)
            updating_anything = False
            had_financial_institution = False
            for column_name, field_name in all_account_fields:
                if column_name == f"account__{account_type}__*":
                    continue
                value = row[headers.index(column_name)]
                if not (value is None or value == ""):
                    updating_anything = True
                    if field_name == "financial_institution":
                        if not isinstance(value, int):
                            errors.append(
                                f"Row: {row_index} - Financial institution ID must be a number for field {column_name}"
                            )
                        had_financial_institution = True
            if updating_anything and not had_financial_institution:
                errors.append(
                    f"Row: {row_index} - Financial institution ID must be provided for account type {account_type}"
                    f" if any other field is updated"
                )
        return errors

    def validate(self, sheet: Worksheet, headers: list[str]) -> list[str]:
        errors = []
        row_index = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_index += 1
            if (row_index - 2) % self.batch_size == 0:
                self.print_message(
                    f"Validating row {row_index - 2} to {min(row_index - 2 + self.batch_size, sheet.max_row - 1)}"
                    f" Indivduals"
                )
            unicef_id = row[headers.index("unicef_id")]
            individuals_queryset = Individual.objects.filter(
                unicef_id=unicef_id,
                business_area=self.business_area,
                program=self.program,
            )
            if not individuals_queryset.exists():  # pragma: no cover
                errors.append(f"Row: {row_index} - Individual with unicef_id {unicef_id} not found")
                continue
            if individuals_queryset.count() > 1:  # pragma: no cover
                errors.append(f"Row: {row_index} - Multiple individuals with unicef_id {unicef_id} found")
                continue
            individual: Individual = individuals_queryset.first() or Individual()
            household = individual.household
            if household is None:  # pragma: no cover
                errors.append(f"Row: {row_index} - Household not found for individual with unicef_id {unicef_id}")
                continue
            errors.extend(self.validate_household_fields(row, headers, household, row_index))
            errors.extend(self.validate_individual_fields(row, headers, row_index))
            errors.extend(self.validate_individual_flex_fields(row, headers, row_index))
            errors.extend(self.validate_household_flex_fields(row, headers, row_index))
            errors.extend(self.validate_documents(row, headers, individual, row_index))
            errors.extend(self.validate_accounts(row, headers, individual, row_index))
        return errors

    def handle_household_update(self, row: tuple[Any, ...], headers: list[str], household: Any) -> None:
        for field, (_name, _validator, handler) in self.household_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(value, field, household, self.business_area, self.program)
            if self.ignore_empty_values and (handled_value is None or handled_value == ""):
                continue
            setattr(household, _name, handled_value)

    def handle_individual_update(self, row: tuple[Any, ...], headers: list[str], individual: Individual) -> None:
        for field, (_name, _validator, handler) in self.individual_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(value, field, individual, self.business_area, self.program)
            if self.ignore_empty_values and (handled_value is None or handled_value == ""):
                continue
            setattr(individual, _name, handled_value)

    def handle_individual_flex_update(self, row: tuple[Any, ...], headers: list[str], individual: Individual) -> None:
        for field, (name, _validator, handler) in self.individual_flex_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(value, field, individual, self.business_area, self.program)
            if self.ignore_empty_values and (handled_value is None or handled_value == ""):
                continue
            individual.flex_fields[name] = handled_value

    def handle_household_flex_update(self, row: tuple[Any, ...], headers: list[str], household: Household) -> None:
        for field, (name, _validator, handler) in self.household_flex_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(value, field, household, self.business_area, self.program)
            if self.ignore_empty_values and (handled_value is None or handled_value == ""):
                continue
            household.flex_fields[name] = handled_value

    def handle_documents_update(
        self, row: tuple[Any, ...], headers: list[str], individual: Individual
    ) -> tuple[list, list]:
        documents_to_update = []
        documents_to_create = []
        if self.document_fields is None:
            return documents_to_update, documents_to_create
        for number_column_name, country_column_name in self.document_fields:
            document_type = self.document_types.get(number_column_name)
            document_number = row[headers.index(number_column_name)]
            document_country = row[headers.index(country_column_name)]
            if self.ignore_empty_values and (document_number is None or document_number == ""):
                continue
            country = self.countries[document_country]
            document = None
            for doc in individual.documents.all():
                if doc.type == document_type:
                    document = doc
                    break
            if document:
                document.document_number = document_number
                document.status = Document.STATUS_PENDING
                document.country = country
                documents_to_update.append(document)
            else:
                documents_to_create.append(
                    Document(
                        individual=individual,
                        type=document_type,
                        document_number=document_number,
                        country=country,
                        rdi_merge_status="MERGED",
                    )
                )
        return documents_to_update, documents_to_create

    def get_all_account_columns(
        self, account_type: str, headers: list[str], columns_to_ignore: list[str]
    ) -> list[tuple[str, str]]:
        prefix = f"account__{account_type}__"
        prefix_len = len(prefix)
        return [(x, x[prefix_len:]) for x in headers if x.startswith(prefix) and x not in columns_to_ignore]

    def handle_account_update(self, row: tuple[Any, ...], headers: list[str], individual: Individual) -> None:
        individual_accounts = individual.accounts.all()
        for account_type, account_columns_mapping in self.account_data_fields.items():
            account_type_instance = self.delivery_mechanisms_account_types.get(account_type)

            single_data_object = next(
                (d for d in individual_accounts if str(d.account_type.key) == str(account_type)),
                None,
            )
            columns_from_header = self.get_all_account_columns(
                account_type, headers, [x for x, y in account_columns_mapping]
            )  # this gets columns starting with account__{account_type}__ and not in the mapping
            all_account_fields = []
            all_account_fields.extend(account_columns_mapping)
            all_account_fields.extend(columns_from_header)
            for column_name, field_name in all_account_fields:
                if column_name == f"account__{account_type}__*":
                    continue
                if column_name in headers:
                    value = row[headers.index(column_name)]
                    if self.ignore_empty_values and (value is None or value == ""):
                        continue
                    if single_data_object is None:
                        single_data_object = Account(
                            individual=individual,
                            account_type=account_type_instance,
                            rdi_merge_status=Account.MERGED,
                        )
                    if field_name == "number":
                        single_data_object.number = value
                    if field_name == "financial_institution":
                        single_data_object.financial_institution_id = value
                        continue
                    single_data_object.data[field_name] = value
            if single_data_object:
                single_data_object.save()
                single_data_object.update_unique_field()

    def handle_update(self, sheet: Worksheet, headers: list[str]) -> list[str]:
        row_index = 1
        individual_ids = []
        household_fields_to_update = ["flex_fields"]
        individual_fields_to_update = ["flex_fields"]
        document_fields_to_update = ["document_number", "status", "country"]
        if self.household_fields:
            household_fields_to_update.extend([field for _, (field, _, _) in self.household_fields.items()])
        if self.individual_fields:
            individual_fields_to_update.extend([field for _, (field, _, _) in self.individual_fields.items()])
        individuals_to_update = []
        households_to_update = []
        documents_to_update = []
        documents_to_create = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_index += 1
            if (row_index - 2) % self.batch_size == 0:
                self.print_message(
                    f"Updating row {row_index - 2} to {min(row_index - 2 + self.batch_size, sheet.max_row - 1)}"
                    f" Individuals"
                )
            unicef_id = row[headers.index("unicef_id")]
            individual = (
                Individual.objects.select_related("household")
                .prefetch_related("documents", "accounts")
                .get(
                    unicef_id=unicef_id,
                    business_area=self.business_area,
                    program=self.program,
                )
            )
            individual_ids.append(str(individual.id))
            household = individual.household
            self.handle_household_update(row, headers, household)
            self.handle_individual_update(row, headers, individual)
            self.handle_individual_flex_update(row, headers, individual)
            self.handle_household_flex_update(row, headers, household)
            (
                documents_to_update_part,
                documents_to_create_part,
            ) = self.handle_documents_update(row, headers, individual)
            documents_to_update.extend(documents_to_update_part)
            documents_to_create.extend(documents_to_create_part)
            self.handle_account_update(row, headers, individual)
            households_to_update.append(household)
            individuals_to_update.append(individual)
            if len(individuals_to_update) == self.batch_size:
                self.batch_update(
                    document_fields_to_update,
                    documents_to_create,
                    documents_to_update,
                    household_fields_to_update,
                    households_to_update,
                    individual_fields_to_update,
                    individuals_to_update,
                )
        self.batch_update(
            document_fields_to_update,
            documents_to_create,
            documents_to_update,
            household_fields_to_update,
            households_to_update,
            individual_fields_to_update,
            individuals_to_update,
        )
        return individual_ids

    def batch_update(
        self,
        document_fields_to_update: list,
        documents_to_create: list,
        documents_to_update: list,
        household_fields_to_update: list,
        households_to_update: list,
        individual_fields_to_update: list,
        individuals_to_update: list,
    ) -> None:
        Document.objects.bulk_update(documents_to_update, document_fields_to_update)
        Document.objects.bulk_create(documents_to_create)
        Household.objects.bulk_update(households_to_update, household_fields_to_update)
        Individual.objects.bulk_update(individuals_to_update, individual_fields_to_update)
        populate_index(
            Individual.objects.filter(id__in=[individual.id for individual in individuals_to_update]),
            get_individual_doc(self.business_area.slug),
        )
        populate_index(
            Household.objects.filter(id__in=[household.id for household in households_to_update]),
            HouseholdDocument,
        )
        documents_to_update.clear()
        documents_to_create.clear()
        households_to_update.clear()
        individuals_to_update.clear()

    def get_excel_value(self, value: Any) -> Any:
        return get_generator_handler(value)(value)

    def get_individual_row(self, individual: Individual) -> list[Any]:
        row = [individual.unicef_id]
        household = individual.household
        row += [
            self.get_excel_value(getattr(individual, field_data[0])) for field_data in self.individual_fields.values()
        ]
        row += [
            self.get_excel_value(individual.flex_fields.get(field_data[0]))
            for field_data in self.individual_flex_fields.values()
        ]
        row += [
            self.get_excel_value(getattr(household, field_data[0])) for field_data in self.household_fields.values()
        ]
        row += [
            self.get_excel_value(household.flex_fields.get(field_data[0]))
            for field_data in self.household_flex_fields.values()
        ]
        all_documents = individual.documents.all()
        for document_no_column, _ in self.document_fields:
            document = [x for x in all_documents if x.type_id == self.document_types[document_no_column].id]
            if len(document) > 1:
                raise ValueError("Multiple documents found")
            if len(document) == 1:
                row.append(document[0].document_number)
                row.append(document[0].country.name)
            else:
                row.append(None)
                row.append(None)
        all_wallets = individual.accounts.all()
        for account_type, data_fields in self.account_data_fields.items():
            wallet = [x for x in all_wallets if x.account_type.key == account_type]
            if len(wallet) > 1:
                raise ValueError("Multiple wallets found")
            if len(wallet) == 1:
                for _, field_name in data_fields:
                    value = wallet[0].data.get(field_name)
                    if field_name == "financial_institution":
                        value = wallet[0].financial_institution_id
                    row.append(self.get_excel_value(value))
            else:
                for _ in data_fields:
                    row.append(None)
        return row

    def generate_xlsx_template(self) -> BytesIO:
        columns = ["unicef_id"]

        columns += list(self.individual_fields)
        columns += list(self.individual_flex_fields)
        columns += list(self.household_fields)
        columns += list(self.household_flex_fields)

        for col_pair in self.document_fields:
            columns.extend(col_pair)
        for lists in self.account_data_fields.values():
            for pair in lists:
                columns.append(pair[0])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(columns)

        individuals = Individual.objects.filter(
            program=self.universal_update.program,
            unicef_id__in=self.universal_update.unicef_ids.split(","),
        )
        individuals_length = len(individuals)

        for index, individual in enumerate(individuals):
            if (index) % self.batch_size == 0:
                self.print_message(f"Generating row {index} to {min(index + self.batch_size, individuals_length)}")
            row_data = self.get_individual_row(individual)
            ws.append(row_data)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        self.print_message("Generating Finished")
        return output

    @transaction.atomic
    def execute(self) -> None:
        workbook = load_workbook(filename=self.file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        errors = self.validate(sheet, headers)
        if errors:
            self.print_message("Validation failed")
            for error in errors:
                self.print_message(error)
            return
        self.print_message("Validation successful")
        processed_individuals_ids = self.handle_update(sheet, headers)
        if self.deduplicate_es:
            self.print_message("Deduplicating individuals Elasticsearch")
            DeduplicateTask(self.business_area.slug, self.program.id).deduplicate_individuals_against_population(
                Individual.objects.filter(id__in=processed_individuals_ids)
            )
        if self.deduplicate_documents:
            self.print_message("Deduplicating documents")
            HardDocumentDeduplication().deduplicate(
                Document.objects.filter(
                    individual__id__in=processed_individuals_ids,
                    status=Document.STATUS_PENDING,
                ),
                program=self.program,
            )

        self.print_message("Update successful")
