import json
import operator

from django.db import transaction
import openpyxl

from hope.apps.registration_data.validators import UploadXLSXInstanceValidator
from hope.models import ImportData, Program


class ValidateXlsxImport:
    @staticmethod
    def _count_non_empty_rows(sheet) -> int:
        count = 0
        for row in sheet.iter_rows(min_row=3):
            if any(cell.value for cell in row):
                count += 1
        return count

    @transaction.atomic()
    def execute(self, import_data: ImportData, program: Program) -> dict:
        import_data.status = ImportData.STATUS_RUNNING
        import_data.save()
        errors = UploadXLSXInstanceValidator(program).validate_everything(
            import_data.file, import_data.business_area_slug
        )
        if errors:
            errors.sort(key=operator.itemgetter("row_number", "header"))
            import_data.status = ImportData.STATUS_VALIDATION_ERROR
            import_data.validation_errors = json.dumps(errors)
        else:
            import_data.status = ImportData.STATUS_FINISHED

        wb = openpyxl.load_workbook(import_data.file)

        number_of_households = 0
        number_of_individuals = 0
        hh_sheet = wb["Households"] if "Households" in wb.sheetnames else None
        ind_sheet = wb["Individuals"] if "Individuals" in wb.sheetnames else None
        people_sheet = wb["People"] if "People" in wb.sheetnames else None

        if not program.is_social_worker_program:
            if hh_sheet:
                number_of_households = self._count_non_empty_rows(hh_sheet)
            if ind_sheet:
                number_of_individuals = self._count_non_empty_rows(ind_sheet)
        elif people_sheet:
            number_of_individuals = self._count_non_empty_rows(people_sheet)

        import_data.number_of_households = number_of_households
        import_data.number_of_individuals = number_of_individuals
        import_data.save()
        return {"import_data_id": import_data.id}
