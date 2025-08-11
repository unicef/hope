from openpyxl import Workbook
from openpyxl.utils import get_column_letter

COLUMNS = (
    "FIRST NAME",
    "SECOND NAME",
    "THIRD NAME",
    "FOURTH NAME",
    "DATE OF BIRTH",
)


def get_template_file() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sanction List Check"
    ws.append(COLUMNS)

    for i in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    return wb
