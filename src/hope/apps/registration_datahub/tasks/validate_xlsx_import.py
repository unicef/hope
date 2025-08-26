import json
import operator

import openpyxl
from django.db import transaction

from hope.models.program import Program
from hope.models.registration_data import ImportData
from hope.apps.registration_datahub.validators import UploadXLSXInstanceValidator


class ValidateXlsxImport:
    @transaction.atomic()
    def execute(self, import_data: ImportData, program: Program) -> dict:
        import_data.status = ImportData.STATUS_RUNNING
        import_data.save()
        errors = UploadXLSXInstanceValidator(program).validate_everything(
            import_data.file, import_data.business_area_slug
        )
        if errors:
            errors.sort(key=operator.itemgetter("row_number", "header"))
            import_data.status = ImportData.STATUS_VALIDATION_ERROR
            import_data.validation_errors = json.dumps(errors)
        else:
            import_data.status = ImportData.STATUS_FINISHED

        wb = openpyxl.load_workbook(import_data.file)

        number_of_households = 0
        number_of_individuals = 0
        hh_sheet = wb["Households"] if "Households" in wb.sheetnames else None
        ind_sheet = wb["Individuals"] if "Individuals" in wb.sheetnames else None
        people_sheet = wb["People"] if "People" in wb.sheetnames else None

        # Could just return max_row if openpyxl won't count empty rows too
        if not program.is_social_worker_program:
            if hh_sheet:
                for row in hh_sheet.iter_rows(min_row=3):
                    if not any(cell.value for cell in row):
                        continue
                    number_of_households += 1

            if ind_sheet:
                for row in ind_sheet.iter_rows(min_row=3):
                    if not any(cell.value for cell in row):
                        continue
                    number_of_individuals += 1
        elif people_sheet:
            for row in people_sheet.iter_rows(min_row=3):
                if not any(cell.value for cell in row):
                    continue
                number_of_individuals += 1
        import_data.number_of_households = number_of_households
        import_data.number_of_individuals = number_of_individuals
        import_data.save()
        return {"import_data_id": import_data.id}
