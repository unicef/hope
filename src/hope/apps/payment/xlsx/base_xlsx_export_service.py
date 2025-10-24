from datetime import datetime
import decimal
import logging
from typing import TYPE_CHECKING, Any

from django.urls import reverse
import openpyxl
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

from hope.apps.payment.utils import get_link

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from hope.apps.account.models import User


logger = logging.getLogger(__name__)


class XlsxExportBaseService:
    text_template = "payment/xlsx_file_generated_email.txt"
    html_template = "payment/xlsx_file_generated_email.html"

    def _create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_active = wb.active
        ws_active.title = self.TITLE
        self.wb = wb
        self.ws_export_list = ws_active
        return wb

    def _add_headers(self) -> None:
        self.ws_export_list.append(self.HEADERS)

    def generate_workbook(self) -> openpyxl.Workbook:
        self._create_workbook()
        self._add_headers()
        # add export items and what you need
        return self.wb

    def generate_file(self, filename: str) -> None:
        self.generate_workbook()
        self.wb.save(filename=filename)

    @staticmethod
    def _adjust_column_width_from_col(ws: "Worksheet") -> None:
        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

        ws.column_dimensions = dim_holder

    def _add_col_bgcolor(
        self,
        col: list | None = None,
        hex_code: str = "A0FDB0",
    ) -> None:
        for row_index in col or []:
            fill = PatternFill(bgColor=hex_code, fgColor=hex_code, fill_type="lightUp")
            bd = Side(style="thin", color="999999")
            for y in range(
                1,
                self.ws_export_list.max_row + 1,
            ):
                cell = self.ws_export_list.cell(row=y, column=row_index)
                cell.fill = fill
                cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    def get_email_context(self, user: "User") -> dict:
        payment_plan_id = str(self.payment_plan.id)
        path_name = "download-payment-plan-payment-list"
        link = get_link(reverse(path_name, args=[payment_plan_id]))

        msg = "Payment Plan Payment List xlsx file(s) were generated and below You have the link to download this file."

        return {
            "first_name": getattr(user, "first_name", ""),
            "last_name": getattr(user, "last_name", ""),
            "email": getattr(user, "email", ""),
            "message": msg,
            "link": link,
            "title": "Payment Plan Payment List files generated",
        }

    def right_format_for_xlsx(self, value: Any) -> Any:
        # this function will return something that excel will accept
        if value is None:
            return ""
        if isinstance(value, str | int | float | decimal.Decimal | datetime):
            return value
        return str(value)
