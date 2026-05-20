from decimal import Decimal, InvalidOperation
import io
from typing import IO
from zipfile import BadZipFile

import openpyxl

from hope.apps.payment.utils import to_decimal
from hope.apps.payment.xlsx.base_xlsx_import_service import XlsxImportBaseService
from hope.apps.payment.xlsx.xlsx_error import XlsxError
from hope.apps.payment.xlsx.xlsx_payment_plan_base_service import XlsxPaymentPlanBaseService
from hope.models import PaymentPlan


class XlsxTopUpCreateImportService(XlsxPaymentPlanBaseService, XlsxImportBaseService):
    """Parse an XLSX top-up template uploaded by ops and return per-beneficiary amounts.

    The XLSX must have a header row including ``household_unicef_id`` and
    ``entitlement_quantity``. Each data row maps one eligible household to a top-up
    amount. Rows with ``entitlement_quantity == 0`` are dropped silently (the
    beneficiary is skipped). Unknown households, negative amounts, missing columns or
    completely-zero files are reported as ``XlsxError`` entries.
    """

    KEY_COLUMN = "household_unicef_id"
    TITLE = "Top Up - Template"

    def __init__(self, source_payment_plan: PaymentPlan, file: io.BytesIO | IO[bytes]) -> None:
        self.source_payment_plan = source_payment_plan
        self.file = file
        self.errors: list[XlsxError] = []
        self.header_to_index: dict[str, int] = {}
        self.eligible_hh_unicef_ids: set[str] = set(
            source_payment_plan.eligible_payments_for_top_up().values_list("household__unicef_id", flat=True)
        )

    def parse_and_validate(self) -> dict[str, Decimal]:
        """Run open + validate + extract. Returns mapping or empty dict if errors."""
        try:
            self._open_workbook()
        except (openpyxl.utils.exceptions.InvalidFileException, BadZipFile, KeyError, ValueError, OSError) as exc:
            self.errors.append(XlsxError(sheet=self.TITLE, coordinates=None, message=f"Unable to open workbook: {exc}"))
            return {}

        self._resolve_header_positions()
        self._validate_required_columns()
        if self.errors:
            return {}

        amounts = self._extract_amounts()
        if self.errors:
            return {}

        if not any(value > 0 for value in amounts.values()):
            self.errors.append(
                XlsxError(
                    sheet=self.TITLE,
                    coordinates=None,
                    message="At least one beneficiary must receive a positive top-up amount",
                )
            )
            return {}

        return {hh_id: amount for hh_id, amount in amounts.items() if amount > 0}

    def _open_workbook(self) -> None:
        self.wb = openpyxl.load_workbook(self.file, data_only=True)
        sheet_name = self.TITLE if self.TITLE in self.wb.sheetnames else self.wb.sheetnames[0]
        self.ws = self.wb[sheet_name]

    def _resolve_header_positions(self) -> None:
        header_row = next(self.ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        self.header_to_index = {
            (str(value).strip() if value is not None else ""): idx
            for idx, value in enumerate(header_row)
            if value is not None and str(value).strip()
        }

    def _validate_required_columns(self) -> None:
        for required in (self.KEY_COLUMN, self.COLUMN_ENTITLEMENT_QUANTITY):
            if required not in self.header_to_index:
                self.errors.append(
                    XlsxError(
                        sheet=self.ws.title,
                        coordinates="1",
                        message=f"Missing required column '{required}'",
                    )
                )

    def _extract_amounts(self) -> dict[str, Decimal]:
        amounts: dict[str, Decimal] = {}
        key_idx = self.header_to_index[self.KEY_COLUMN]
        amount_idx = self.header_to_index[self.COLUMN_ENTITLEMENT_QUANTITY]

        for row in self.ws.iter_rows(min_row=2):
            if not any(cell.value for cell in row):
                continue

            key_cell = row[key_idx]
            amount_cell = row[amount_idx]

            household_unicef_id = self._cell_str(key_cell)
            if not household_unicef_id:
                self.errors.append(
                    XlsxError(
                        sheet=self.ws.title,
                        coordinates=key_cell.coordinate,
                        message=f"Missing {self.KEY_COLUMN}",
                    )
                )
                continue

            if household_unicef_id not in self.eligible_hh_unicef_ids:
                self.errors.append(
                    XlsxError(
                        sheet=self.ws.title,
                        coordinates=key_cell.coordinate,
                        message=f"Household {household_unicef_id} is not eligible for top-up",
                    )
                )
                continue

            if household_unicef_id in amounts:
                self.errors.append(
                    XlsxError(
                        sheet=self.ws.title,
                        coordinates=key_cell.coordinate,
                        message=f"Duplicate row for household {household_unicef_id}",
                    )
                )
                continue

            amount = self._coerce_amount(amount_cell)
            if amount is None:
                continue
            amounts[household_unicef_id] = amount

        return amounts

    def _coerce_amount(self, cell) -> Decimal | None:
        raw = cell.value
        if raw is None or raw == "":
            return Decimal(0)
        try:
            amount = to_decimal(raw) if isinstance(raw, str) else Decimal(str(raw))
        except (InvalidOperation, ValueError):
            self.errors.append(
                XlsxError(
                    sheet=self.ws.title,
                    coordinates=cell.coordinate,
                    message=f"Invalid amount '{raw}'",
                )
            )
            return None
        if amount < 0:
            self.errors.append(
                XlsxError(
                    sheet=self.ws.title,
                    coordinates=cell.coordinate,
                    message=f"Amount must be non-negative, got {amount}",
                )
            )
            return None
        return amount

    @staticmethod
    def _cell_str(cell) -> str:
        return str(cell.value).strip() if cell.value is not None else ""
