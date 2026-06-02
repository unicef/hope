from __future__ import annotations

from io import BytesIO
import logging
from typing import IO, TYPE_CHECKING

from django.db import transaction
import openpyxl

from hope.apps.payment.flows import PaymentPlanFlow
from hope.apps.payment.services.payment_plan_services import PaymentPlanService
from hope.apps.payment.xlsx.xlsx_error import XlsxError
from hope.apps.payment.xlsx.xlsx_payment_plan_delivery_import_service import XlsxPaymentPlanDeliveryImportService
from hope.models import (
    Payment,
    PaymentPlan,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from hope.models import PaymentPlanGroup

logger = logging.getLogger(__name__)


class XlsxPaymentPlanGroupDeliveryImportService:
    """Import a single-sheet reconciliation xlsx covering the plans of a single-FSP group.

    The group is bound to one FSP, so the file has a single flat header (the FSP template's
    columns). A group may still hold several payment plans sharing that FSP, so rows are routed
    to their owning plan by Payment.unicef_id; each plan's rows are handed to an unmodified
    single-plan XlsxPaymentPlanDeliveryImportService.
    """

    REQUIRED_COLUMNS = ("payment_id", "delivered_quantity")

    def __init__(self, payment_plan_group: "PaymentPlanGroup", file: IO[bytes]) -> None:
        self.payment_plan_group = payment_plan_group
        self.file = file
        self.errors: list[XlsxError] = []
        self.payment_plans: list[PaymentPlan] = list(
            payment_plan_group.payment_plans.filter(
                status__in=[PaymentPlan.Status.ACCEPTED, PaymentPlan.Status.FINISHED],
                plan_type=PaymentPlan.PlanType.REGULAR,
            ).order_by("unicef_id")
        )
        self.eligible_plans: list[PaymentPlan] = []
        self.payment_to_plan: dict[str, PaymentPlan] = {}
        self.payment_gateway_payment_ids: set[str] = set()
        self.headers: list[str] = []
        self.sheetname: str = ""
        self.ws: Worksheet | None = None
        self.wb: openpyxl.Workbook | None = None
        self.per_plan_services: dict[str, XlsxPaymentPlanDeliveryImportService] = {}

    def _prepare_eligible_plans(self) -> None:
        payment_gateway_plans: list[PaymentPlan] = []
        for payment_plan in self.payment_plans:
            if payment_plan.is_payment_gateway:
                logger.warning(
                    f"Skipping Payment Plan {payment_plan.unicef_id}: uses payment gateway, "
                    f"manual reconciliation is not allowed."
                )
                payment_gateway_plans.append(payment_plan)
                continue
            self.eligible_plans.append(payment_plan)
        if payment_gateway_plans:
            self.payment_gateway_payment_ids = {
                str(unicef_id)
                for unicef_id in Payment.objects.filter(parent__in=payment_gateway_plans)
                .eligible()
                .values_list("unicef_id", flat=True)
            }

    def _build_payment_index(self) -> None:
        payments = (
            Payment.objects.filter(parent__in=self.eligible_plans).eligible().values_list("unicef_id", "parent_id")
        )
        payment_plan_by_id = {str(payment_plan.id): payment_plan for payment_plan in self.eligible_plans}
        for unicef_id, parent_id in payments:
            payment_plan = payment_plan_by_id.get(str(parent_id))
            if payment_plan is not None:
                self.payment_to_plan[str(unicef_id)] = payment_plan

    def open_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.load_workbook(self.file, data_only=True)
        self.wb = wb
        self.ws = wb[wb.sheetnames[0]]
        self.sheetname = wb.sheetnames[0]
        self.headers = [cell.value for cell in self.ws[1]]
        self._prepare_eligible_plans()
        self._build_payment_index()
        return wb

    def _validate_required_headers(self) -> bool:
        missing = [col for col in self.REQUIRED_COLUMNS if col not in self.headers]
        if missing:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    None,
                    f"Provided headers {self.headers} do not match expected headers. "
                    f"{list(self.REQUIRED_COLUMNS)} are required headers.",
                )
            )
            return False
        return True

    def _validate_row_payment_ids(self) -> None:
        """Emit XlsxErrors for unknown, payment-gateway, or duplicated payment_id values."""
        if self.ws is None:
            return
        seen_ids: set[str] = set()
        payment_id_idx = self.headers.index("payment_id")
        for row in self.ws.iter_rows(min_row=2):
            if not any(cell.value for cell in row):
                continue
            id_cell = row[payment_id_idx]
            if id_cell.value is None:
                continue
            payment_id = str(id_cell.value)
            if payment_id in self.payment_gateway_payment_ids:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        id_cell.coordinate,
                        f"Payment id {payment_id} belongs to a payment plan that uses payment gateway "
                        f"and cannot be manually reconciled.",
                    )
                )
                continue
            if payment_id not in self.payment_to_plan:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        id_cell.coordinate,
                        f"Payment id {payment_id} does not belong to any payment plan in this group.",
                    )
                )
                continue
            if payment_id in seen_ids:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        id_cell.coordinate,
                        f"Payment id {payment_id} appears multiple times in the import file",
                    )
                )
                continue
            seen_ids.add(payment_id)

    def _row_groups_by_plan(self) -> dict[str, list[tuple]]:
        """Group valid rows by owning plan id.

        Rows with unknown or duplicate payment_id values are skipped silently — validation of
        those cases is done separately by _validate_row_payment_ids.
        """
        if self.ws is None:
            return {}
        rows_by_plan: dict[str, list[tuple]] = {}
        seen_ids: set[str] = set()
        payment_id_idx = self.headers.index("payment_id")
        for row in self.ws.iter_rows(min_row=2):
            if not any(cell.value for cell in row):
                continue
            payment_id = row[payment_id_idx].value
            if payment_id is None:
                continue
            payment_id = str(payment_id)
            payment_plan = self.payment_to_plan.get(payment_id)
            if payment_plan is None or payment_id in seen_ids:
                continue
            seen_ids.add(payment_id)
            rows_by_plan.setdefault(str(payment_plan.id), []).append(tuple(cell.value for cell in row))
        return rows_by_plan

    def _build_per_plan_workbook(self, rows: list[tuple]) -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheetname or "Sheet"
        ws.append(self.headers)
        for row_values in rows:
            ws.append(list(row_values))
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _build_per_plan_services(self) -> None:
        rows_by_plan = self._row_groups_by_plan()
        for payment_plan in self.eligible_plans:
            rows = rows_by_plan.get(str(payment_plan.id), [])
            if not rows:
                continue
            sub_workbook_file = self._build_per_plan_workbook(rows)
            service = XlsxPaymentPlanDeliveryImportService(payment_plan, sub_workbook_file)
            service.open_workbook()
            self.per_plan_services[str(payment_plan.id)] = service

    def validate(self) -> None:
        if not self._validate_required_headers():
            return
        self._validate_row_payment_ids()
        self._build_per_plan_services()
        any_updated = False
        for service in self.per_plan_services.values():
            service._validate_headers()
            if service.errors:
                self.errors.extend(service.errors)
                service.errors = []
                continue
            service._validate_rows()
            if service.is_updated:
                any_updated = True
            self.errors.extend(service.errors)
            service.errors = []
        if not self.errors and not any_updated and self.per_plan_services:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    None,
                    "There aren't any updates in imported file, please add changes and try again",
                )
            )

    def import_payment_list(self) -> None:
        if not self.per_plan_services:
            self._build_per_plan_services()
        with transaction.atomic():
            for payment_plan_id, service in self.per_plan_services.items():
                payment_plan = service.payment_plan
                service.import_payment_list()
                payment_plan.remove_export_files()
                flow = PaymentPlanFlow(payment_plan)
                flow.background_action_status_none()
                payment_plan.update_money_fields()
                if payment_plan.is_reconciled and payment_plan.status == PaymentPlan.Status.ACCEPTED:
                    flow.status_finished()
                payment_plan.save()
                # invalidate cache for program cycle list
                payment_plan.program_cycle.save()
                logger.info(f"Scheduled update payments signature for payment plan {payment_plan_id}")
                PaymentPlanService(payment_plan).recalculate_signatures_in_batch()
