from __future__ import annotations

import datetime
from decimal import Decimal
from io import BytesIO
import logging
from typing import IO, TYPE_CHECKING, Any, cast

from django.db.models import Prefetch
from django.utils import timezone
import openpyxl

from hope.apps.activity_log.utils import copy_model_object
from hope.apps.payment.flows import PaymentPlanFlow
from hope.apps.payment.services.handle_total_cash_in_households import (
    handle_total_cash_in_specific_households,
)
from hope.apps.payment.utils import (
    bulk_log_payment_changes,
    calculate_counts,
    get_payment_delivered_quantity_status_and_value,
    get_quantity_in_usd,
    log_payment_plan_change,
    to_decimal,
)
from hope.apps.payment.xlsx.base_xlsx_import_service import XlsxImportBaseService
from hope.apps.payment.xlsx.xlsx_error import XlsxError
from hope.models import FollowUpInstruction, Payment, PaymentPlan, PaymentVerification, User

if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell


class XlsxFollowUpInstructionReconciliationImportService(XlsxImportBaseService):
    logger = logging.getLogger(__name__)
    HOUSEHOLD_IDENTIFIER_HEADER = "household_id"

    def __init__(self, instruction: FollowUpInstruction, file: IO[bytes]) -> None:
        self.instruction = instruction
        self.file = file
        self.errors: list[XlsxError] = []
        self.required_columns = [self.HOUSEHOLD_IDENTIFIER_HEADER, "delivered_quantity"]
        self.xlsx_headers: list[str] = []
        self.household_ids_from_xlsx: list[str] = []
        self.household_updates: dict[str, Decimal] = {}
        self.payments_to_save: list[Payment] = []
        self.old_payments: dict = {}
        self.payment_verifications_to_save: list[PaymentVerification] = []
        self.payment_plans_to_update: dict[str, PaymentPlan] = {}
        self.payment_verification_plans_to_update: dict[str, Any] = {}
        self.is_updated = False

        payment_list = (
            Payment.objects.filter(parent__follow_up_instruction=instruction)
            .eligible()
            .select_related("currency", "household", "parent")
            .prefetch_related(
                Prefetch(
                    "payment_verifications",
                    queryset=PaymentVerification.objects.select_related("payment_verification_plan"),
                )
            )
            .order_by("parent__created_at", "created_at", "id")
        )
        self.payments_by_household_unicef_id: dict[str, list[Payment]] = {}
        self.current_delivered_by_household: dict[str, Decimal] = {}
        self.total_entitled_by_household: dict[str, Decimal] = {}
        for payment in payment_list:
            household_unicef_id = payment.household.unicef_id
            self.payments_by_household_unicef_id.setdefault(household_unicef_id, []).append(payment)
            self.current_delivered_by_household[household_unicef_id] = self.current_delivered_by_household.get(
                household_unicef_id, Decimal(0)
            ) + (payment.delivered_quantity or Decimal(0))
            self.total_entitled_by_household[household_unicef_id] = self.total_entitled_by_household.get(
                household_unicef_id, Decimal(0)
            ) + (payment.entitlement_quantity or Decimal(0))

    def open_workbook(self) -> openpyxl.Workbook:
        self.logger.info(f"Opening workbook for follow-up instruction: {self.instruction.id}")
        self.file.seek(0)
        workbook = openpyxl.load_workbook(BytesIO(self.file.read()), data_only=True)
        self.wb = workbook
        self.ws_payments = workbook[workbook.sheetnames[0]]
        self.sheetname = workbook.sheetnames[0]
        self.xlsx_headers = [header.value for header in self.ws_payments[1]]
        return workbook

    def _validate_headers(self) -> None:
        for required_column in self.required_columns:
            if required_column not in self.xlsx_headers:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        None,
                        f"Provided headers {self.xlsx_headers} do not match expected headers. "
                        f"{self.required_columns} are required headers.",
                    )
                )
                return

    def _validate_household_unicef_id(self, row: tuple[Cell, ...]) -> str | None:
        household_unicef_id = row[self.xlsx_headers.index(self.HOUSEHOLD_IDENTIFIER_HEADER)].value
        cell = row[self.xlsx_headers.index(self.HOUSEHOLD_IDENTIFIER_HEADER)]
        if household_unicef_id is None or str(household_unicef_id).strip() == "":
            self.errors.append(XlsxError(self.sheetname, cell.coordinate, "Household ID is required."))
            return None
        household_unicef_id = str(household_unicef_id).strip()
        if household_unicef_id not in self.payments_by_household_unicef_id:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"Household ID {household_unicef_id} is not part of this Follow Up Instruction.",
                )
            )
            return None
        if household_unicef_id in self.household_ids_from_xlsx:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"Household ID {household_unicef_id} appears multiple times in the import file.",
                )
            )
            return None
        self.household_ids_from_xlsx.append(household_unicef_id)
        return household_unicef_id

    def _should_skip_household_row(self, household_unicef_id: str | None) -> bool:
        if household_unicef_id is None:
            return True
        return any(
            payment.delivered_quantity is not None and payment.status not in Payment.PENDING_STATUSES
            for payment in self.payments_by_household_unicef_id.get(household_unicef_id, [])
        )

    def _validate_delivered_quantity(self, row: tuple[Cell, ...], household_unicef_id: str | None) -> None:
        if household_unicef_id is None:
            return
        cell = row[self.xlsx_headers.index("delivered_quantity")]
        delivered_quantity = cell.value
        if delivered_quantity is None or str(delivered_quantity).strip() == "":
            return
        if isinstance(delivered_quantity, datetime.date):
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"Delivered quantity {delivered_quantity} is not a valid number.",
                )
            )
            return
        delivered_quantity_decimal = to_decimal(delivered_quantity)
        if delivered_quantity_decimal is None:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"Delivered quantity {delivered_quantity} is not a valid number.",
                )
            )
            return
        if delivered_quantity_decimal < 0:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"Household {household_unicef_id}: Delivered quantity cannot be below zero.",
                )
            )
            return
        total_entitled_quantity = self.total_entitled_by_household[household_unicef_id]
        if delivered_quantity_decimal > total_entitled_quantity:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"Household {household_unicef_id}: Delivered quantity {delivered_quantity_decimal} is bigger than "
                    f"Entitlement quantity {total_entitled_quantity}",
                )
            )
            return
        self.household_updates[household_unicef_id] = delivered_quantity_decimal
        if delivered_quantity_decimal != self.current_delivered_by_household.get(household_unicef_id, Decimal(0)):
            self.is_updated = True

    def _validate_rows(self) -> None:
        for row in self.ws_payments.iter_rows(min_row=2):
            if not any(cell.value for cell in row):
                continue
            household_unicef_id = self._validate_household_unicef_id(row)
            if self._should_skip_household_row(household_unicef_id):
                continue
            self._validate_delivered_quantity(row, household_unicef_id)

    def validate(self) -> None:
        self.logger.info("Starting instruction reconciliation validation")
        self._validate_headers()
        if not self.errors:
            self._validate_rows()
            if not self.errors and not self.is_updated:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        None,
                        "There aren't any updates in imported file, please add changes and try again",
                    )
                )
        self.logger.info("Finished instruction reconciliation validation")

    def _update_payment_verification(self, payment: Payment, delivered_quantity: Decimal | None) -> None:
        payment_verification = next(iter(payment.payment_verifications.all()), None)
        if payment_verification and payment_verification.status != PaymentVerification.STATUS_PENDING:
            if payment_verification.received_amount == delivered_quantity:
                verification_status = PaymentVerification.STATUS_RECEIVED
            elif delivered_quantity == 0 or delivered_quantity is None:
                verification_status = PaymentVerification.STATUS_NOT_RECEIVED
            else:
                verification_status = PaymentVerification.STATUS_RECEIVED_WITH_ISSUES

            if payment_verification.status != verification_status:
                payment_verification.status = verification_status
                payment_verification.status_date = timezone.now()
                self.payment_verifications_to_save.append(payment_verification)
                self.payment_verification_plans_to_update[str(payment_verification.payment_verification_plan_id)] = (
                    payment_verification.payment_verification_plan
                )

    @staticmethod
    def _normalize_delivery_date(payment: Payment, delivered_quantity: Decimal | None) -> datetime.datetime | None:
        if delivered_quantity and delivered_quantity > 0:
            return payment.delivery_date or timezone.now()
        return None

    def _import_household(self, household_unicef_id: str, delivered_quantity: Decimal) -> None:
        remaining_quantity = delivered_quantity
        for payment in self.payments_by_household_unicef_id[household_unicef_id]:
            entitlement_quantity = payment.entitlement_quantity or Decimal(0)
            allocated_quantity = min(remaining_quantity, entitlement_quantity)
            status, delivered_quantity_value = get_payment_delivered_quantity_status_and_value(
                str(allocated_quantity), entitlement_quantity
            )
            delivery_date = self._normalize_delivery_date(payment, delivered_quantity_value)
            delivered_quantity_usd = get_quantity_in_usd(
                amount=delivered_quantity_value,
                currency=payment.currency,
                exchange_rate=Decimal(payment.parent.exchange_rate or 0),
                currency_exchange_date=payment.parent.currency_exchange_date,
            )
            if (
                payment.delivered_quantity != delivered_quantity_value
                or payment.delivered_quantity_usd != delivered_quantity_usd
                or payment.status != status
                or payment.delivery_date != delivery_date
            ):
                self.old_payments[payment.pk] = copy_model_object(payment)
                payment.delivered_quantity = delivered_quantity_value
                payment.delivered_quantity_usd = delivered_quantity_usd
                payment.status = status
                payment.delivery_date = delivery_date
                self.payments_to_save.append(payment)
                self.payment_plans_to_update[str(payment.parent_id)] = payment.parent
                self._update_payment_verification(payment, delivered_quantity_value)
            remaining_quantity -= allocated_quantity

    def import_payment_list(self, user_id: str | None = None) -> None:
        self.logger.info("Starting instruction reconciliation import")
        for household_unicef_id, delivered_quantity in self.household_updates.items():
            self._import_household(household_unicef_id, delivered_quantity)

        Payment.objects.bulk_update(
            self.payments_to_save,
            ("delivered_quantity", "delivered_quantity_usd", "status", "delivery_date"),
            batch_size=500,
        )
        user = User.objects.filter(pk=user_id).first() if user_id else None
        bulk_log_payment_changes(
            [(self.old_payments[payment.pk], payment) for payment in self.payments_to_save],
            user,
        )
        handle_total_cash_in_specific_households([payment.household_id for payment in self.payments_to_save])
        PaymentVerification.objects.bulk_update(self.payment_verifications_to_save, ("status", "status_date"))
        for payment_verification_plan in self.payment_verification_plans_to_update.values():
            calculate_counts(payment_verification_plan)
            payment_verification_plan.save()
        for payment_plan in self.payment_plans_to_update.values():
            old_payment_plan = cast("PaymentPlan", copy_model_object(payment_plan))
            payment_plan.update_money_fields()
            if payment_plan.is_reconciled and payment_plan.status == PaymentPlan.Status.ACCEPTED:
                flow = PaymentPlanFlow(payment_plan)
                flow.status_finished()
                payment_plan.save()
            payment_plan.program_cycle.save()
            log_payment_plan_change(payment_plan, old_payment_plan, user_id)
        self.logger.info("Finished instruction reconciliation import")
