import datetime
from decimal import Decimal
import io
import logging
from typing import TYPE_CHECKING

from dateutil.parser import parse
from django.utils import timezone
import openpyxl
import pytz
from xlwt import Row

from hope.apps.payment.models import Payment, PaymentVerification
from hope.apps.payment.services.handle_total_cash_in_households import (
    handle_total_cash_in_specific_households,
)
from hope.apps.payment.utils import (
    calculate_counts,
    get_payment_delivered_quantity_status_and_value,
    get_quantity_in_usd,
    to_decimal,
)
from hope.apps.payment.xlsx.base_xlsx_import_service import XlsxImportBaseService
from hope.apps.payment.xlsx.xlsx_error import XlsxError

if TYPE_CHECKING:
    from django.db.models import QuerySet

    from hope.apps.payment.models import PaymentPlan


class XlsxPaymentPlanImportPerFspService(XlsxImportBaseService):
    logger = logging.getLogger(__name__)

    class XlsxPaymentPlanImportPerFspServiceError(Exception):
        pass

    def __init__(self, payment_plan: "PaymentPlan", file: io.BytesIO) -> None:
        self.payment_plan = payment_plan
        self.payment_list: QuerySet["Payment"] = payment_plan.eligible_payments
        self.file = file
        self.errors: list[XlsxError] = []
        self.payments_dict: dict = {str(x.unicef_id): x for x in self.payment_list}
        self.payment_ids: list = list(self.payments_dict.keys())
        self.payments_to_save: list = []
        self.payment_verifications_to_save: list = []
        self.required_columns: list[str] = ["payment_id", "delivered_quantity"]
        self.xlsx_headers = []
        self.is_updated: bool = False

    def open_workbook(self) -> openpyxl.Workbook:
        self.logger.info(f"Opening workbook for payment plan: {self.payment_plan.id}")
        wb = openpyxl.load_workbook(self.file, data_only=True)
        self.wb = wb
        self.ws_payments = wb[wb.sheetnames[0]]
        self.sheetname = wb.sheetnames[0]
        self.logger.info("Generating headers")
        self.xlsx_headers = [header.value for header in self.ws_payments[1]]
        return wb

    def _validate_headers(self) -> None:
        # need check only if "payment_id" and "delivered_quantity" exists
        for required_column in self.required_columns:
            if required_column not in self.xlsx_headers:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        None,
                        f"Provided headers {self.xlsx_headers}"
                        f" do not match expected headers. {self.required_columns} "
                        f"are required headers.",
                    )
                )
                return

    def _validate_payment_id(self, row: Row) -> None:
        cell = row[self.xlsx_headers.index("payment_id")]
        if cell.value not in self.payment_ids:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"This payment id {cell.value} is not in Payment Plan Payment List",
                )
            )

    def _validate_delivered_quantity(self, row: Row) -> None:
        """Define when possible for a user to upload a file.

        * Fully Delivered (entitled quantity = delivered quantity) [float]
        * Partially Delivered (entitled quantity > delivered quantity > 0) [float]
        * Not Delivered (0 = delivered quantity) [0.0]
        * Unsuccessful (failed at the delivery processing level) [-1.0]
        * Pending (no information) [None]

        The validation should not pass when:

        * delivered quantity > entitled quantity
        """
        payment_id = row[self.xlsx_headers.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return

        cell = row[self.xlsx_headers.index("delivered_quantity")]
        delivered_quantity = cell.value

        if delivered_quantity is not None and delivered_quantity != "":
            delivered_quantity = to_decimal(delivered_quantity)
            if delivered_quantity != payment.delivered_quantity:  # update value
                entitlement_quantity = payment.entitlement_quantity or Decimal(0)
                if delivered_quantity > entitlement_quantity:
                    self.errors.append(
                        XlsxError(
                            self.sheetname,
                            cell.coordinate,
                            f"Payment {payment_id}: Delivered quantity {delivered_quantity} is bigger than "
                            f"Entitlement quantity {entitlement_quantity}",
                        )
                    )
                else:
                    self.is_updated = True

    def _validate_reason_for_unsuccessful_payment(self, row: Row) -> None:
        payment_id = row[self.xlsx_headers.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return

        if "reason_for_unsuccessful_payment" in self.xlsx_headers:
            reason_for_unsuccessful_payment = row[self.xlsx_headers.index("reason_for_unsuccessful_payment")].value
            if reason_for_unsuccessful_payment != payment.reason_for_unsuccessful_payment:
                self.is_updated = True

    def _validate_delivery_date(self, row: Row) -> None:
        payment_id = row[self.xlsx_headers.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return

        if "delivery_date" in self.xlsx_headers:
            cell = row[self.xlsx_headers.index("delivery_date")]
            delivery_date = cell.value
        else:
            delivery_date = None

        if delivery_date is None:
            self.is_updated = True  # Update Payment item with current datetime
            return

        try:
            if not isinstance(delivery_date, datetime.datetime):
                delivery_date = parse(delivery_date)

            if not delivery_date.tzinfo:
                delivery_date = pytz.utc.localize(delivery_date)

            if delivery_date != payment.delivery_date:
                self.is_updated = True
            # convert to date
            delivery_date = delivery_date.date() if isinstance(delivery_date, datetime.datetime) else delivery_date
            if delivery_date > datetime.date.today() or delivery_date < self.payment_plan.program.start_date:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        cell.coordinate,
                        f"Payment {payment_id}: Delivery date ({delivery_date}) cannot be greater than today's date,"
                        f" and cannot be before Programme's start date",
                    )
                )
        except ValueError:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"Payment {payment_id}: Delivered date {delivery_date} is not a datetime",
                )
            )

    def _validate_reference_id(self, row: Row) -> None:
        payment_id = row[self.xlsx_headers.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return

        if "reference_id" in self.xlsx_headers:
            reference_id = row[self.xlsx_headers.index("reference_id")].value
            if reference_id != payment.transaction_reference_id:
                self.is_updated = True

    def _validate_rows(self) -> None:
        for row in self.ws_payments.iter_rows(min_row=2):
            if not any(cell.value for cell in row):
                continue

            self._validate_payment_id(row)
            self._validate_delivered_quantity(row)
            self._validate_delivery_date(row)
            self._validate_reason_for_unsuccessful_payment(row)
            self._validate_reference_id(row)

    def _validate_imported_file(self) -> None:
        if not self.is_updated:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    None,
                    "There aren't any updates in imported file, please add changes and try again",
                )
            )

    def validate(self) -> None:
        self.logger.info("Starting validation")
        self.logger.info("Validating headers")
        self._validate_headers()
        if not self.errors:
            self.logger.info("Validating rows")
            self._validate_rows()

            self.logger.info("Validating if the file was changed")
            self._validate_imported_file()
        self.logger.info("Finished validation")

    def import_payment_list(self) -> None:
        self.logger.info("Starting importing payment list")
        exchange_rate = self.payment_plan.get_exchange_rate()

        for row in self.ws_payments.iter_rows(min_row=2):
            self._import_row(row, exchange_rate)

        self.logger.info("Updating payments")
        Payment.objects.bulk_update(
            self.payments_to_save,
            (
                "delivered_quantity",
                "delivered_quantity_usd",
                "status",
                "delivery_date",
                "transaction_reference_id",
                "reason_for_unsuccessful_payment",
                "additional_collector_name",
                "additional_document_type",
                "additional_document_number",
                "transaction_status_blockchain_link",
            ),
            batch_size=500,
        )
        self.logger.info("Update total cash in households")
        handle_total_cash_in_specific_households([payment.household_id for payment in self.payments_to_save])
        self.logger.info("Updating status and status date in payment verifications")
        PaymentVerification.objects.bulk_update(self.payment_verifications_to_save, ("status", "status_date"))
        self.logger.info("Finished import payment list")

    def _get_delivered_quantity_status_and_value(
        self,
        delivered_quantity: int | float | str,
        entitlement_quantity: Decimal,
        payment_id: str,
    ) -> tuple[str, Decimal | None]:
        try:
            status, quantity = get_payment_delivered_quantity_status_and_value(delivered_quantity, entitlement_quantity)
        except ValueError:
            raise self.XlsxPaymentPlanImportPerFspServiceError(
                f"Invalid delivered_quantity {delivered_quantity} provided for payment_id {payment_id}"
            )

        return status, quantity

    def _import_row(self, row: Row, exchange_rate: float) -> None:
        payment_id = row[self.xlsx_headers.index("payment_id")].value
        if payment_id is None:
            return  # safety check
        payment = self.payments_dict[payment_id]
        self.logger.info(f"Importing row for payment {payment_id}")
        delivered_quantity = row[self.xlsx_headers.index("delivered_quantity")].value

        if "delivery_date" in self.xlsx_headers:
            delivery_date = row[self.xlsx_headers.index("delivery_date")].value
        else:
            delivery_date = None

        if "reference_id" in self.xlsx_headers:
            reference_id = row[self.xlsx_headers.index("reference_id")].value
        else:
            reference_id = None

        if "reason_for_unsuccessful_payment" in self.xlsx_headers:
            reason_for_unsuccessful_payment = row[self.xlsx_headers.index("reason_for_unsuccessful_payment")].value
        else:
            reason_for_unsuccessful_payment = None

        if "additional_collector_name" in self.xlsx_headers:
            additional_collector_name = row[self.xlsx_headers.index("additional_collector_name")].value
        else:
            additional_collector_name = None

        if "additional_document_type" in self.xlsx_headers:
            additional_document_type = row[self.xlsx_headers.index("additional_document_type")].value
        else:
            additional_document_type = None

        if "additional_document_number" in self.xlsx_headers:
            additional_document_number = row[self.xlsx_headers.index("additional_document_number")].value
        else:
            additional_document_number = None

        if "transaction_status_blockchain_link" in self.xlsx_headers:
            transaction_status_blockchain_link = row[
                self.xlsx_headers.index("transaction_status_blockchain_link")
            ].value
        else:
            transaction_status_blockchain_link = None

        if isinstance(delivery_date, str):
            delivery_date = parse(delivery_date)

        if delivery_date and delivery_date.tzinfo is None:
            delivery_date = pytz.utc.localize(delivery_date)

        if payment_delivery_date := payment.delivery_date:
            payment_delivery_date = payment.delivery_date.replace(tzinfo=None)

        # convert to date
        delivery_date = delivery_date.date() if isinstance(delivery_date, datetime.datetime) else delivery_date

        if (
            delivery_date
            and delivery_date > datetime.date.today()
            or delivery_date
            and delivery_date < self.payment_plan.program.start_date
        ):
            # validate and skip update the date
            delivery_date = payment_delivery_date

        if delivered_quantity is not None and str(delivered_quantity).strip() != "":
            status, delivered_quantity = self._get_delivered_quantity_status_and_value(
                delivered_quantity, payment.entitlement_quantity, payment_id
            )

            if (
                (delivered_quantity != payment.delivered_quantity)
                or (status != payment.status)
                or (delivery_date != payment_delivery_date)
                or (reason_for_unsuccessful_payment != payment.reason_for_unsuccessful_payment)
                or (additional_collector_name != payment.additional_collector_name)
                or (additional_document_type != payment.additional_document_type)
                or (additional_document_number != payment.additional_document_number)
                or (reference_id != payment.transaction_reference_id)
                or (transaction_status_blockchain_link != payment.transaction_status_blockchain_link)
            ):
                payment.delivered_quantity = delivered_quantity
                payment.delivered_quantity_usd = get_quantity_in_usd(
                    amount=delivered_quantity,
                    currency=self.payment_plan.currency,
                    exchange_rate=Decimal(exchange_rate),
                    currency_exchange_date=self.payment_plan.currency_exchange_date,
                )
                payment.status = status
                if delivery_date:
                    payment.delivery_date = delivery_date
                elif payment.delivered_quantity and not payment.delivery_date:
                    payment.delivery_date = timezone.now()
                elif not payment.delivered_quantity:
                    payment.delivery_date = None
                payment.reason_for_unsuccessful_payment = reason_for_unsuccessful_payment
                payment.additional_collector_name = additional_collector_name
                payment.additional_document_type = additional_document_type
                payment.additional_document_number = additional_document_number
                payment.transaction_reference_id = reference_id
                payment.transaction_status_blockchain_link = transaction_status_blockchain_link

                self.payments_to_save.append(payment)
                # update PaymentVerification status
                payment_verification = payment.payment_verifications.first()
                if payment_verification and payment_verification.status != PaymentVerification.STATUS_PENDING:
                    if payment_verification.received_amount == delivered_quantity:
                        pv_status = PaymentVerification.STATUS_RECEIVED
                    elif delivered_quantity == 0 or delivered_quantity is None:
                        pv_status = PaymentVerification.STATUS_NOT_RECEIVED
                    else:
                        pv_status = PaymentVerification.STATUS_RECEIVED_WITH_ISSUES

                    payment_verification.status = pv_status
                    payment_verification.status_date = timezone.now()
                    self.payment_verifications_to_save.append(payment_verification)

                    payment_verification_plan = payment_verification.payment_verification_plan
                    self.logger.info(f"Calculating counts for payment verification plan {payment_verification_plan.id}")
                    calculate_counts(payment_verification_plan)
                    payment_verification_plan.save()
