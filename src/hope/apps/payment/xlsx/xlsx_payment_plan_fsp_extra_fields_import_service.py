from __future__ import annotations

import datetime
from decimal import Decimal
from typing import IO, TYPE_CHECKING, Any, cast

import openpyxl

from hope.apps.activity_log.utils import copy_model_object
from hope.apps.payment.utils import bulk_log_payment_changes
from hope.apps.payment.xlsx.base_xlsx_import_service import XlsxImportBaseService
from hope.apps.payment.xlsx.xlsx_error import XlsxError
from hope.apps.payment.xlsx.xlsx_payment_plan_delivery_export_service import (
    XlsxPaymentPlanDeliveryExportService,
)
from hope.models import (
    Payment,
    PaymentPlan,
    User,
)

if TYPE_CHECKING:
    from openpyxl.cell import Cell

Row = tuple["Cell", ...]


class XlsxPaymentPlanFspExtraFieldsImportService(XlsxImportBaseService):
    TITLE = "Payment Plan - FSP Extra Fields"
    PAYMENT_ID_COLUMN = "payment_id"
    BATCH_SIZE = 500

    def __init__(self, payment_plan: PaymentPlan, file: IO[bytes]) -> None:
        self.payment_plan = payment_plan
        self.file = file
        self.errors: list[XlsxError] = []
        self.headers: list[Any] = []
        self.header_to_index: dict[str, int] = {}
        self.custom_headers: list[str] = []
        self.payments = {
            str(payment.unicef_id): payment
            for payment in payment_plan.eligible_payments.select_related(
                "currency",
                "delivery_type",
                "household_snapshot",
            )
        }
        self.is_updated = False

    def open_workbook(self) -> openpyxl.Workbook:
        self.workbook = openpyxl.load_workbook(self.file, data_only=True)
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]
        self.sheetname = str(self.worksheet.title)
        self.headers = [cell.value for cell in self.worksheet[1]]
        self._resolve_headers()
        return self.workbook

    def _resolve_headers(self) -> None:
        self.header_to_index = {
            header: index for index, header in enumerate(self.headers) if isinstance(header, str) and header
        }
        self.custom_headers = [
            header for header in self.headers if isinstance(header, str) and header and header != self.PAYMENT_ID_COLUMN
        ]

    def validate(self) -> None:
        self._validate_headers()
        if self.errors:
            return
        self._validate_rows()
        if not self.errors and not self.is_updated:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    None,
                    "There aren't any updates in imported file, please add changes and try again",
                )
            )

    def _validate_headers(self) -> None:
        if self.headers.count(self.PAYMENT_ID_COLUMN) != 1:
            self.errors.append(
                XlsxError(self.sheetname, None, f"Header {self.PAYMENT_ID_COLUMN} is required exactly once")
            )
        invalid_headers = [header for header in self.headers if not isinstance(header, str) or not header.strip()]
        if invalid_headers:
            self.errors.append(XlsxError(self.sheetname, None, "All XLSX columns must have a non-empty text header"))
        string_headers = [header for header in self.headers if isinstance(header, str)]
        duplicate_headers = sorted({header for header in string_headers if string_headers.count(header) > 1})
        if duplicate_headers:
            self.errors.append(
                XlsxError(self.sheetname, None, f"Duplicate headers are not allowed: {duplicate_headers}")
            )
        reserved_headers = sorted(set(self.custom_headers) & self._system_controlled_headers())
        if reserved_headers:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    None,
                    f"System-controlled headers cannot be imported: {reserved_headers}",
                )
            )

    def _system_controlled_headers(self) -> frozenset[str]:
        return XlsxPaymentPlanDeliveryExportService.get_system_controlled_headers(
            self.payment_plan,
        )

    def _validate_rows(self) -> None:
        seen_payment_ids: set[str] = set()
        for row in self.worksheet.iter_rows(min_row=2):
            if not any(cell.value not in (None, "") for cell in row):
                continue
            payment_id_cell = self._cell(row, self.PAYMENT_ID_COLUMN)
            payment_id = str(payment_id_cell.value) if payment_id_cell.value not in (None, "") else ""
            if not payment_id:
                self.errors.append(XlsxError(self.sheetname, payment_id_cell.coordinate, "Payment id is required"))
                continue
            payment = self.payments.get(payment_id)
            if payment is None:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        payment_id_cell.coordinate,
                        f"This payment id {payment_id} is not in Payment Plan Payment List",
                    )
                )
                continue
            if payment_id in seen_payment_ids:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        payment_id_cell.coordinate,
                        f"Payment id {payment_id} appears multiple times in the import file",
                    )
                )
                continue
            seen_payment_ids.add(payment_id)
            updates = self._fsp_extra_fields_for_row(row)
            if any(payment.fsp_extra_fields.get(key) != value for key, value in updates.items()):
                self.is_updated = True

    def import_payment_list(self, user_id: str | None = None) -> int:
        payments_to_update: list[Payment] = []
        log_pairs: list[tuple[Payment | None, Payment]] = []
        seen_payment_ids: set[str] = set()

        for row in self.worksheet.iter_rows(min_row=2):
            if not any(cell.value not in (None, "") for cell in row):
                continue
            payment_id_value = self._cell(row, self.PAYMENT_ID_COLUMN).value
            payment_id = str(payment_id_value) if payment_id_value not in (None, "") else ""
            payment = self.payments.get(payment_id)
            if payment is None or payment_id in seen_payment_ids:
                continue
            seen_payment_ids.add(payment_id)
            updates = self._fsp_extra_fields_for_row(row)
            if not updates or not any(payment.fsp_extra_fields.get(key) != value for key, value in updates.items()):
                continue
            old_payment = cast("Payment", copy_model_object(payment))
            payment.extras = {
                **payment.extras,
                Payment.FSP_EXTRA_FIELDS_KEY: {
                    **payment.fsp_extra_fields,
                    **updates,
                },
            }
            payment.update_signature_hash()
            payments_to_update.append(payment)
            log_pairs.append((old_payment, payment))

        if payments_to_update:
            Payment.objects.bulk_update(
                payments_to_update,
                fields=("extras", "signature_hash"),
                batch_size=self.BATCH_SIZE,
            )
            user = User.objects.filter(pk=user_id).first() if user_id else None
            bulk_log_payment_changes(log_pairs, user)
        return len(payments_to_update)

    def _cell(self, row: Row, header: str) -> Cell:
        return row[self.header_to_index[header]]

    def _fsp_extra_fields_for_row(self, row: Row) -> dict[str, object]:
        values: dict[str, object] = {}
        for header in self.custom_headers:
            value = self._cell(row, header).value
            if value in (None, ""):
                continue
            values[header] = self._normalize_value(value)
        return values

    @staticmethod
    def _normalize_value(value: Any) -> object:
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, datetime.datetime | datetime.date | datetime.time):
            return value.isoformat()
        if isinstance(value, int | float | bool | str):
            return value
        return str(value)
