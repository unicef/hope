from decimal import Decimal
import io
from typing import TYPE_CHECKING, Union

from django.contrib.admin.options import get_content_type_for_model
from django.utils import timezone
import openpyxl
from openpyxl.cell import Cell

from hope.apps.core.models import FileTemp
from hope.apps.payment.models import Payment, PaymentPlan
from hope.apps.payment.utils import get_quantity_in_usd, to_decimal
from hope.apps.payment.xlsx.base_xlsx_import_service import XlsxImportBaseService
from hope.apps.payment.xlsx.xlsx_error import XlsxError
from hope.apps.payment.xlsx.xlsx_payment_plan_base_service import (
    XlsxPaymentPlanBaseService,
)

if TYPE_CHECKING:
    from django.contrib.auth.base_user import AbstractBaseUser
    from django.contrib.auth.models import AnonymousUser

    from hope.apps.account.models import User

Row = tuple[Cell]


class XlsxPaymentPlanImportService(XlsxPaymentPlanBaseService, XlsxImportBaseService):
    BATCH_SIZE = 1000

    def __init__(self, payment_plan: PaymentPlan, file: io.BytesIO) -> None:
        self.payment_plan = payment_plan
        self.file = file
        self.payments_dict: dict[str, Payment] = {
            str(x.unicef_id): x for x in payment_plan.eligible_payments.only("unicef_id", "entitlement_quantity")
        }
        self.errors: list[XlsxError] = []
        self.is_updated = False
        self.headers = list(self.HEADERS)
        if self.payment_plan.is_social_worker_program:
            self.headers.remove("household_size")
            self.headers.remove("household_id")

    def open_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.load_workbook(self.file, data_only=True)
        self.wb = wb
        self.ws_payments = wb[self.TITLE]
        return wb

    def validate(self) -> None:
        self._validate_headers()
        if self.errors:
            self._validate_rows()
            self._validate_imported_file()

    def import_payment_list(self) -> None:
        exchange_rate = self.payment_plan.get_exchange_rate()
        payments_to_save = []

        for row in self.ws_payments.iter_rows(min_row=2):
            if not any(cell.value for cell in row):
                continue
            if payment := self._import_row(row, exchange_rate):
                payments_to_save.append(payment)

            if len(payments_to_save) == self.BATCH_SIZE:
                self._save_payments(payments_to_save)
                payments_to_save = []

        if payments_to_save:
            self._save_payments(payments_to_save)

    def _save_payments(self, payments_to_save: list[Payment]) -> None:
        Payment.objects.bulk_update(
            payments_to_save,
            fields=(
                "entitlement_quantity",
                "entitlement_quantity_usd",
                "entitlement_date",
            ),
        )
        payments_ids = [payment.id for payment in payments_to_save]
        payments = Payment.objects.filter(id__in=payments_ids).select_related("household_snapshot")
        for payment in payments:
            payment.update_signature_hash()
        Payment.objects.bulk_update(payments, fields=("signature_hash",))

    def _validate_headers(self) -> None:
        all_headers = [header.value for header in self.ws_payments[1]]

        for required_header in ["payment_id", "entitlement_quantity"]:
            if required_header not in all_headers:
                self.errors.append(
                    XlsxError(
                        self.TITLE,
                        None,
                        f"Header {required_header} is required",
                    )
                )

    def _validate_payment_id(self, row: Row, payments_ids: list[str]) -> None:
        cell = row[self.headers.index("payment_id")]
        if cell.value not in payments_ids:
            self.errors.append(
                XlsxError(
                    self.TITLE,
                    cell.coordinate,
                    f"This payment id {cell.value} is not in Payment Plan Payment List",
                )
            )

    def _validate_entitlement(self, row: Row) -> None:
        payment_id = row[self.headers.index("payment_id")].value
        payment = self.payments_dict.get(str(payment_id))
        if payment is None:
            return
        entitlement_amount = row[self.headers.index("entitlement_quantity")].value
        if entitlement_amount is not None and entitlement_amount != "":
            converted_entitlement_amount = to_decimal(str(entitlement_amount)) or Decimal(0.0)
            if converted_entitlement_amount != payment.entitlement_quantity:
                self.is_updated = True

    def _validate_imported_file(self) -> None:
        if not self.is_updated:
            self.errors.append(
                XlsxError(
                    self.TITLE,
                    None,
                    "There aren't any updates in imported file, please add changes and try again",
                )
            )

    def _validate_rows(self) -> None:
        payments_ids = list(self.payments_dict.keys())
        for row in self.ws_payments.iter_rows(min_row=2):
            if not any(cell.value for cell in row):
                continue
            self._validate_payment_id(row, payments_ids)
            self._validate_entitlement(row)

    def _import_row(self, row: Row, exchange_rate: float) -> Payment | None:
        payment_id = row[self.headers.index("payment_id")].value
        entitlement_amount = row[self.headers.index("entitlement_quantity")].value

        payment = self.payments_dict.get(str(payment_id))

        if payment is None:
            return None

        if entitlement_amount is not None and entitlement_amount != "":
            converted_entitlement_amount = to_decimal(str(entitlement_amount)) or Decimal(0.0)
            if converted_entitlement_amount != payment.entitlement_quantity:
                entitlement_date = timezone.now()
                entitlement_quantity_usd = get_quantity_in_usd(
                    amount=converted_entitlement_amount,
                    currency=self.payment_plan.currency,
                    exchange_rate=Decimal(exchange_rate) if exchange_rate is not None else 1,
                    currency_exchange_date=self.payment_plan.currency_exchange_date,
                )
                return Payment(
                    id=payment.id,
                    entitlement_quantity=converted_entitlement_amount,
                    entitlement_date=entitlement_date,
                    entitlement_quantity_usd=entitlement_quantity_usd,
                )
        return None

    def create_import_xlsx_file(self, user: Union["User", "AbstractBaseUser", "AnonymousUser"]) -> PaymentPlan:
        # remove old imported file
        self.payment_plan.remove_imported_file()

        # create new imported xlsx file
        xlsx_file = FileTemp.objects.create(
            object_id=self.payment_plan.pk,
            content_type=get_content_type_for_model(self.payment_plan),
            created_by=user,
            file=self.file,
        )

        self.payment_plan.imported_file = xlsx_file
        self.payment_plan.save()

        return self.payment_plan
