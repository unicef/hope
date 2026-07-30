from decimal import Decimal
from typing import IO, TYPE_CHECKING
from zipfile import BadZipFile

from django.core.exceptions import ValidationError
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from hope.apps.payment.utils import to_decimal
from hope.apps.payment.xlsx.xlsx_payment_plan_base_service import XlsxPaymentPlanBaseService
from hope.apps.payment.xlsx.xlsx_payment_plan_export_service import XlsxPaymentPlanExportService

if TYPE_CHECKING:
    from hope.models import Payment, PaymentPlan


class TopUpAmountTemplateService(XlsxPaymentPlanExportService):
    """Blank amount template listing the payments eligible for this plan's child plan.

    Used for both a Top-Up and a Top-Up Amendment. Same sheet shape as the entitlement export
    operators already know, with two differences: the row set is ``eligible_payments_for_child_plan()``
    rather than every eligible payment, and the entitlement columns ship empty. Empty is deliberate —
    prefilling the source amount invites the operator to submit it unchanged, which would silently
    repeat the original payment instead of topping it up.
    """

    def _add_payment_list(self) -> None:
        qs = (
            self.payment_plan.eligible_payments_for_child_plan()
            .select_related("household_snapshot", "currency", "delivery_type", "financial_service_provider")
            .order_by("unicef_id")
        )
        for payment in qs.iterator(chunk_size=self.batch_size):
            self._add_payment_row(payment)

    def _payment_row(self, payment: "Payment") -> list:
        # Blank the amounts before the row reaches the sheet: reading them back out of openpyxl
        # costs a full-sheet scan per row, which turns a 10k export into minutes.
        row = super()._payment_row(payment)
        for column_name in (self.COLUMN_ENTITLEMENT_QUANTITY, "entitlement_quantity_usd"):
            row[self.headers.index(column_name)] = None
        return row


def _open_amount_sheet(file: IO[bytes]) -> Worksheet:
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
    except (InvalidFileException, BadZipFile, OSError):
        raise ValidationError("The uploaded file could not be read as an XLSX workbook.")
    if XlsxPaymentPlanBaseService.TITLE not in workbook.sheetnames:
        raise ValidationError(f"Sheet '{XlsxPaymentPlanBaseService.TITLE}' not found in the uploaded file.")
    return workbook[XlsxPaymentPlanBaseService.TITLE]


def _row_amount(raw_amount: object, payment_id: str) -> Decimal | None:
    """Amount for one row, or ``None`` when the row funds nobody (blank or zero)."""
    if raw_amount is None or str(raw_amount).strip() == "":
        return None
    amount = to_decimal(str(raw_amount))  # returns None for anything non-numeric
    if amount is None:
        raise ValidationError(f"Invalid amount '{raw_amount}' for payment {payment_id}.")
    if amount < 0:
        raise ValidationError(f"Negative amount '{raw_amount}' for payment {payment_id}.")
    return amount or None


def parse_top_up_amount_file(source_payment_plan: "PaymentPlan", file: IO[bytes]) -> dict[str, Decimal]:
    """Read a filled-in Top-Up amount template into ``{payment unicef_id: amount}``.

    Only funded rows are returned: a row left empty or set to zero means the beneficiary is not
    part of the Top-Up at all, which is how the operator narrows it down. Raises when the file is
    unreadable, references payments outside the source plan's eligible set, carries a negative or
    non-numeric amount, or funds nobody.
    """
    worksheet = _open_amount_sheet(file)
    headers = [cell.value for cell in worksheet[1]]
    required_columns = (
        XlsxPaymentPlanBaseService.COLUMN_PAYMENT_ID,
        XlsxPaymentPlanBaseService.COLUMN_ENTITLEMENT_QUANTITY,
    )
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise ValidationError(f"Column '{missing[0]}' is required in the amount file.")

    payment_id_index = headers.index(XlsxPaymentPlanBaseService.COLUMN_PAYMENT_ID)
    amount_index = headers.index(XlsxPaymentPlanBaseService.COLUMN_ENTITLEMENT_QUANTITY)
    eligible_ids = set(source_payment_plan.eligible_payments_for_child_plan().values_list("unicef_id", flat=True))

    amounts: dict[str, Decimal] = {}
    # Tracked separately from ``amounts``, which a blank or zero row never reaches: checking there
    # would let a duplicate through or not depending on which of its rows carries the amount.
    seen_payment_ids: set[str] = set()
    for row in worksheet.iter_rows(min_row=2):
        raw_payment_id = row[payment_id_index].value
        if raw_payment_id is None:
            continue
        payment_id = str(raw_payment_id).strip()
        if payment_id not in eligible_ids:
            raise ValidationError(f"Payment {payment_id} is not eligible for a Top-Up of this Payment Plan.")
        if payment_id in seen_payment_ids:
            raise ValidationError(f"Payment {payment_id} appears more than once in the amount file.")
        seen_payment_ids.add(payment_id)
        if (amount := _row_amount(row[amount_index].value, payment_id)) is not None:
            amounts[payment_id] = amount

    if not amounts:
        raise ValidationError("The amount file funds nobody — set an amount above zero for at least one beneficiary.")
    return amounts
