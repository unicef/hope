from __future__ import annotations

from collections import defaultdict
import csv
import dataclasses
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from enum import IntEnum
import io
import logging
import re
from tempfile import NamedTemporaryFile
from typing import IO
import zipfile

from django.contrib.admin.options import get_content_type_for_model
from django.core.exceptions import ObjectDoesNotExist
from django.core.files import File
from django.core.files.base import ContentFile
from django.db import DatabaseError, transaction
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font
from pypdf import PdfReader

from hope.apps.account.models import User
from hope.apps.account.permissions import Permissions
from hope.apps.payment.celery_tasks import send_qcf_report_email_notifications_async_task
from hope.apps.payment.services.western_union_ftp import WesternUnionFTPClient
from hope.apps.payment.utils import get_link
from hope.models import (
    FileTemp,
    Payment,
    PaymentPlan,
    WesternUnionData,
    WesternUnionInvoice,
    WesternUnionInvoicePayment,
    WesternUnionPaymentPlanReport,
)

logger = logging.getLogger(__name__)


@dataclasses.dataclass
class InvoiceParseResult:
    date: date
    advice_name: str
    net_amount: Decimal
    charges: Decimal


@dataclasses.dataclass
class DataParseResult:
    date: date
    amount: Decimal


@dataclasses.dataclass
class MatchedDataPaymentRow:
    payment: Payment
    mtcn: str
    transaction_status: str
    principal_amount: Decimal
    charges_amount: Decimal
    fee_amount: Decimal


@dataclasses.dataclass
class QCFReportPaymentRowData:
    payment_unicef_id: str | None
    mtcn: str
    mtcns_match: bool
    hope_mtcn: str | None
    payment_plan_unicef_id: str | None
    programme: str
    fc: str
    principal_amount: float
    charges_amount: float
    fee_amount: float
    ad_name_value: str


@dataclasses.dataclass
class QCFReportPaymentPlanData:
    payment_plan_unicef_id: str | None
    principal_total: float = dataclasses.field(init=False)
    charges_total: float = dataclasses.field(init=False)
    refunds_total: float = dataclasses.field(init=False)
    fees_total: float = dataclasses.field(init=False)
    payments_data: list[QCFReportPaymentRowData]

    def __post_init__(self) -> None:
        self.principal_total = sum(p.principal_amount for p in self.payments_data if p.principal_amount >= 0) or 0
        self.charges_total = sum(p.charges_amount for p in self.payments_data) or 0
        self.refunds_total = sum(p.principal_amount for p in self.payments_data if p.principal_amount < 0) or 0
        self.fees_total = sum(p.fee_amount for p in self.payments_data) or 0


class QCFReportsService:
    UNMATCHED_INVOICE_ERROR_GRACE_PERIOD = timedelta(days=31)

    class QCFReportsServiceError(Exception):
        pass

    class DataFieldIndex(IntEnum):
        RECORD_TYPE = 0
        MTCN = 1
        PAYMENT_UNICEF_ID = 9
        PRINCIPAL_AMOUNT = 10
        CHARGES_AMOUNT = 11
        FEE_AMOUNT = 12
        TRANSACTION_STATUS = 18

        @classmethod
        def last_required_index(cls) -> int:
            return max(member.value for member in cls)

    class DataRecordType:
        REGULAR = "1"
        ADJUSTMENT_REVERSAL = "2"
        ADJUSTMENT_NEW_RESULT = "3"
        FOOTER = "9"
        DETAIL_TYPES = {REGULAR, ADJUSTMENT_REVERSAL, ADJUSTMENT_NEW_RESULT}

    class DataFooterFieldIndex(IntEnum):
        RECORD_TYPE = 0
        TOTAL_NUMBER_OF_TRANSACTIONS = 1
        TOTAL_PRINCIPAL = 2
        TOTAL_CHARGES = 3
        TOTAL_TRANSACTION_FEE = 4

        @classmethod
        def last_required_index(cls) -> int:
            return max(member.value for member in cls)

    FILENAME_DATE_FORMAT = "%Y%m%d"
    FILENAME_DATE_PATTERN = re.compile(r"(\d{8})(?=[^0-9]*\.[^.]+$)")
    DATA_PDF_PRINCIPAL_PATTERN = re.compile(r"Principal\s+([\d,]+\.\d{2})", re.MULTILINE)
    DATA_PDF_CHARGES_PATTERN = re.compile(r"Charges\s+([\d,]+\.\d{2})", re.MULTILINE)

    def __init__(self) -> None:
        self._ftp_client: WesternUnionFTPClient | None = None

    @property
    def ftp_client(self) -> WesternUnionFTPClient:
        if self._ftp_client is None:
            self._ftp_client = WesternUnionFTPClient()
        return self._ftp_client

    def process_files_since(self, date_from: datetime) -> None:
        self.import_invoice_files_since(date_from)
        self.import_data_files_since(date_from)
        self.reconcile_pending_records()

    def import_invoice_files_since(self, date_from: datetime) -> None:
        files = self.ftp_client.get_invoice_files_since(date_from)
        for filename, file_like in files:
            if WesternUnionInvoice.objects.filter(name=filename).exists():
                continue

            try:
                self.import_invoice_file(filename, file_like)
            except (
                QCFReportsService.QCFReportsServiceError,
                DatabaseError,
                ObjectDoesNotExist,
                OSError,
                ValueError,
                zipfile.BadZipFile,
            ):
                logger.exception("Failed to import Western Union invoice file %s", filename)

    def import_data_files_since(self, date_from: datetime) -> None:
        files = self.ftp_client.get_data_files_since(date_from)
        for filename, file_like in files:
            if WesternUnionData.objects.filter(name=filename).exists():
                continue

            try:
                self.import_data_file(filename, file_like)
            except (
                QCFReportsService.QCFReportsServiceError,
                DatabaseError,
                ObjectDoesNotExist,
                OSError,
                ValueError,
                zipfile.BadZipFile,
            ):
                logger.exception("Failed to import Western Union data file %s", filename)

    def import_invoice_file(self, filename: str, file_like: io.BytesIO) -> None:
        invoice = WesternUnionInvoice.objects.create(name=filename)

        try:
            self.attach_file(invoice, filename, file_like)
            file_like.seek(0)
            parse_result = self.parse_invoice_file(filename, file_like)
            invoice.date = parse_result.date
            invoice.advice_name = parse_result.advice_name
            invoice.net_amount = parse_result.net_amount
            invoice.charges = parse_result.charges
            invoice.status = WesternUnionInvoice.STATUS_PENDING
            invoice.error_msg = ""
            invoice.save(update_fields=["date", "advice_name", "net_amount", "charges", "status", "error_msg"])
        except (
            QCFReportsService.QCFReportsServiceError,
            DatabaseError,
            ObjectDoesNotExist,
            OSError,
            ValueError,
            zipfile.BadZipFile,
        ) as exc:
            self.mark_record_error(invoice, exc)

    def import_data_file(self, filename: str, file_like: io.BytesIO) -> None:
        data_file = WesternUnionData.objects.create(name=filename)

        try:
            self.attach_file(data_file, filename, file_like)
            file_like.seek(0)
            parse_result = self.parse_data_file(filename, file_like)
            data_file.date = parse_result.date
            data_file.amount = parse_result.amount
            data_file.status = WesternUnionData.STATUS_PENDING
            data_file.error_msg = ""
            data_file.save(update_fields=["date", "amount", "status", "error_msg"])
        except (
            QCFReportsService.QCFReportsServiceError,
            DatabaseError,
            ObjectDoesNotExist,
            OSError,
            ValueError,
            zipfile.BadZipFile,
        ) as exc:
            self.mark_record_error(data_file, exc)

    def reconcile_pending_records(self) -> None:
        pending_invoices = WesternUnionInvoice.objects.filter(
            status=WesternUnionInvoice.STATUS_PENDING,
            is_legacy=False,
            advice_name__isnull=False,
            net_amount__isnull=False,
        ).order_by("date", "id")

        for invoice in pending_invoices:
            try:
                self.reconcile_invoice(invoice)
            except (
                QCFReportsService.QCFReportsServiceError,
                DatabaseError,
                ObjectDoesNotExist,
                OSError,
                ValueError,
                zipfile.BadZipFile,
            ):
                logger.exception("Failed to reconcile Western Union invoice %s", invoice.name)

    def reconcile_invoice(self, invoice: WesternUnionInvoice, send_notifications: bool = True) -> None:
        matched_data = self.get_matching_data_file(invoice)
        if matched_data is None:
            if self.should_mark_unmatched_invoice_as_error(invoice):
                self.mark_record_error(invoice, self.QCFReportsServiceError("No matching data file found"))
            return

        mark_matched_data_error = False
        try:
            with transaction.atomic():
                try:
                    payment_plan_id_payments_map = self.extract_data_payment_rows(matched_data)
                    self.link_payments_from_data_rows(invoice, payment_plan_id_payments_map)
                except (
                    QCFReportsService.QCFReportsServiceError,
                    DatabaseError,
                    ObjectDoesNotExist,
                    OSError,
                    ValueError,
                    zipfile.BadZipFile,
                ):
                    mark_matched_data_error = True
                    raise

                self.generate_reports_for_invoice(
                    invoice,
                    payment_plan_id_payments_map,
                    send_notifications=send_notifications,
                )
                invoice.matched_data = matched_data
                invoice.status = WesternUnionInvoice.STATUS_COMPLETED
                invoice.error_msg = ""
                invoice.save(update_fields=["matched_data", "status", "error_msg"])
                matched_data.status = WesternUnionData.STATUS_COMPLETED
                matched_data.error_msg = ""
                matched_data.save(update_fields=["status", "error_msg"])
        except (
            QCFReportsService.QCFReportsServiceError,
            DatabaseError,
            ObjectDoesNotExist,
            OSError,
            ValueError,
            zipfile.BadZipFile,
        ) as exc:
            if mark_matched_data_error:
                self.mark_record_error(matched_data, exc)
            self.mark_record_error(invoice, exc)

    def should_mark_unmatched_invoice_as_error(self, invoice: WesternUnionInvoice) -> bool:
        if invoice.date is None:
            return False

        return timezone.now().date() - invoice.date > self.UNMATCHED_INVOICE_ERROR_GRACE_PERIOD

    def get_matching_data_candidates(
        self,
        invoice: WesternUnionInvoice,
        include_completed: bool = False,
    ) -> list[WesternUnionData]:
        queryset = WesternUnionData.objects.exclude(amount__isnull=True).exclude(status=WesternUnionData.STATUS_ERROR)
        if include_completed:
            queryset = queryset.filter(amount=invoice.net_amount)
        else:
            queryset = queryset.filter(status=WesternUnionData.STATUS_PENDING, amount=invoice.net_amount)
        return list(queryset)

    def get_matching_data_file(
        self,
        invoice: WesternUnionInvoice,
        include_completed: bool = False,
    ) -> WesternUnionData | None:
        candidates = self.get_matching_data_candidates(invoice, include_completed=include_completed)
        if not candidates:
            return None
        return self.pick_best_data_match(invoice, candidates)

    def pick_best_data_match(
        self,
        invoice: WesternUnionInvoice,
        candidates: list[WesternUnionData],
    ) -> WesternUnionData | None:
        if len(candidates) == 1:
            return candidates[0]

        return min(
            candidates,
            key=lambda candidate: (
                abs((invoice.date - candidate.date).days) if invoice.date and candidate.date else 10**9,
                candidate.date or date.max,
                candidate.id,
            ),
        )

    def link_payments_from_data_rows(
        self,
        invoice: WesternUnionInvoice,
        payment_plan_id_payments_map: dict[str, list[MatchedDataPaymentRow]],
    ) -> None:
        WesternUnionInvoicePayment.objects.filter(western_union_invoice=invoice).delete()

        payments_to_link = [
            WesternUnionInvoicePayment(
                western_union_invoice=invoice,
                payment=row.payment,
                transaction_status=row.transaction_status,
            )
            for payment_rows in payment_plan_id_payments_map.values()
            for row in payment_rows
        ]
        if not payments_to_link:
            raise self.QCFReportsServiceError("No valid payments found in matching data file")

        WesternUnionInvoicePayment.objects.bulk_create(payments_to_link)

    def generate_reports_for_invoice(
        self,
        invoice: WesternUnionInvoice,
        payment_plan_id_payments_map: dict[str, list[MatchedDataPaymentRow]],
        send_notifications: bool = True,
    ) -> None:
        ad_name_value = invoice.advice_name or ""
        for payment_plan_id, payment_rows in payment_plan_id_payments_map.items():
            payment_plan = PaymentPlan.objects.get(id=payment_plan_id)
            report_name, report = self.generate_report(payment_plan, payment_rows, ad_name_value)

            wu_qcf_report = WesternUnionPaymentPlanReport.objects.create(
                qcf_file=invoice,
                payment_plan=payment_plan,
            )

            with NamedTemporaryFile() as tmp:
                file_temp = FileTemp.objects.create(
                    object_id=str(wu_qcf_report.pk),
                    content_type=get_content_type_for_model(wu_qcf_report),
                )
                report.save(tmp.name)
                tmp.seek(0)
                file_temp.file.save(report_name, File(tmp))
                wu_qcf_report.report_file = file_temp
                wu_qcf_report.save()

            if send_notifications:
                transaction.on_commit(
                    lambda wu_qcf_report_id=str(wu_qcf_report.id): send_qcf_report_email_notifications_async_task(
                        wu_qcf_report_id
                    )
                )

    def extract_data_payment_rows(self, data_file: WesternUnionData) -> dict[str, list[MatchedDataPaymentRow]]:
        if data_file.file is None:
            raise self.QCFReportsServiceError("Matching data file has no file")

        payment_plan_id_payments_map: dict[str, list[MatchedDataPaymentRow]] = defaultdict(list)

        data_file.file.file.open("rb")
        try:
            lines = self.read_rows_from_file_like(data_file.name, data_file.file.file)
        finally:
            data_file.file.file.close()

        grouped_rows: dict[str, list[list[str]]] = defaultdict(list)
        for fields in self.extract_data_rows(lines, data_file.name):
            payment_unicef_id = self.normalize_payment_unicef_id(fields[self.DataFieldIndex.PAYMENT_UNICEF_ID])
            grouped_rows[payment_unicef_id].append(fields)

        for payment_unicef_id, payment_rows in grouped_rows.items():
            try:
                payment = Payment.objects.get(unicef_id=payment_unicef_id)
            except Payment.DoesNotExist:
                logger.error("%s: payment %s does not exist", data_file.name, payment_unicef_id)
                continue

            matched_row = self.aggregate_data_payment_row(payment_unicef_id, payment, payment_rows)
            payment_plan_id_payments_map[str(payment.parent_id)].append(matched_row)

        return payment_plan_id_payments_map

    def aggregate_data_payment_row(
        self,
        payment_unicef_id: str,
        payment: Payment,
        payment_rows: list[list[str]],
    ) -> MatchedDataPaymentRow:
        mtcns = {self.clean_field(row[self.DataFieldIndex.MTCN]) for row in payment_rows}
        if len(mtcns) != 1:
            raise self.QCFReportsServiceError(f"Inconsistent MTCN values found for payment {payment_unicef_id}")

        representative_row = max(
            payment_rows,
            key=lambda row: self.record_type_rank(self.clean_field(row[self.DataFieldIndex.RECORD_TYPE])),
        )

        return MatchedDataPaymentRow(
            payment=payment,
            mtcn=self.clean_field(representative_row[self.DataFieldIndex.MTCN]),
            transaction_status=self.clean_field(representative_row[self.DataFieldIndex.TRANSACTION_STATUS]),
            principal_amount=sum(
                (self.to_decimal(row[self.DataFieldIndex.PRINCIPAL_AMOUNT]) for row in payment_rows),
                start=Decimal(0),
            ),
            charges_amount=sum(
                (self.to_decimal(row[self.DataFieldIndex.CHARGES_AMOUNT]) for row in payment_rows),
                start=Decimal(0),
            ),
            fee_amount=sum(
                (self.to_decimal(row[self.DataFieldIndex.FEE_AMOUNT]) for row in payment_rows),
                start=Decimal(0),
            ),
        )

    def record_type_rank(self, record_type: str) -> int:
        rank_map = {
            self.DataRecordType.ADJUSTMENT_NEW_RESULT: 3,
            self.DataRecordType.REGULAR: 2,
            self.DataRecordType.ADJUSTMENT_REVERSAL: 1,
        }
        return rank_map.get(record_type, 0)

    def attach_file(self, record: WesternUnionInvoice | WesternUnionData, filename: str, file_like: io.BytesIO) -> None:
        content_file = ContentFile(file_like.read(), name=filename)
        file_temp = FileTemp.objects.create(
            object_id=str(record.pk),
            content_type=get_content_type_for_model(record),
        )
        file_temp.file.save(filename, content_file)
        record.file = file_temp
        record.save(update_fields=["file"])

    def mark_record_error(self, record: WesternUnionInvoice | WesternUnionData, exc: Exception) -> None:
        record.status = record.STATUS_ERROR
        record.error_msg = str(exc)
        record.save(update_fields=["status", "error_msg"])
        logger.exception("Western Union file processing failed for %s", record.name, exc_info=exc)

    def parse_invoice_file(self, filename: str, file_like: IO[bytes]) -> InvoiceParseResult:
        parsed_date = self.parse_date_from_filename(filename)
        advice_name, pdf_bytes = self.extract_pdf_from_archive(filename, file_like.read())
        pdf_text = self.extract_text_from_pdf(filename, pdf_bytes)
        net_amount = self.parse_principal_amount_from_pdf_text(pdf_text, filename)
        charges = self.parse_charges_amount_from_pdf_text(pdf_text)
        if not advice_name:
            raise self.QCFReportsServiceError(f"No advice filename found in invoice file {filename}")

        return InvoiceParseResult(
            date=parsed_date,
            advice_name=advice_name,
            net_amount=net_amount,
            charges=charges,
        )

    def parse_data_file(self, filename: str, file_like: IO[bytes]) -> DataParseResult:
        rows = self.read_rows_from_file_like(filename, file_like)
        parsed_date = self.parse_date_from_filename(filename)
        footer_row = self.extract_footer_row(rows, filename)

        amount = self.to_decimal(footer_row[self.DataFooterFieldIndex.TOTAL_PRINCIPAL])
        return DataParseResult(date=parsed_date, amount=amount)

    def parse_date_from_filename(self, filename: str) -> date:
        match = self.FILENAME_DATE_PATTERN.search(filename)
        if not match:
            raise self.QCFReportsServiceError(f"Could not parse date from filename {filename}")

        return datetime.strptime(match.group(1), self.FILENAME_DATE_FORMAT).date()

    def read_rows_from_file_like(self, filename: str, file_like: IO[bytes]) -> list[list[str]]:
        file_bytes = file_like.read()
        if not file_bytes:
            raise self.QCFReportsServiceError(f"File {filename} is empty")

        if zipfile.is_zipfile(io.BytesIO(file_bytes)):
            return self.read_rows_from_zip(filename, file_bytes)

        return self.decode_rows(filename, file_bytes)

    def read_rows_from_zip(self, filename: str, file_bytes: bytes) -> list[list[str]]:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            candidate_files = [
                zip_info
                for zip_info in archive.infolist()
                if not zip_info.is_dir() and zip_info.filename.lower().endswith(".cf")
            ]

            if not candidate_files:
                raise self.QCFReportsServiceError(f"No supported text files found in archive {filename}")

            for zip_info in candidate_files:
                with archive.open(zip_info) as extracted_file:
                    try:
                        return self.decode_rows(zip_info.filename, extracted_file.read())
                    except self.QCFReportsServiceError:
                        logger.warning("Skipping unsupported archive member %s in %s", zip_info.filename, filename)

        raise self.QCFReportsServiceError(f"Could not parse any supported archive member in {filename}")

    def extract_pdf_from_archive(self, filename: str, file_bytes: bytes) -> tuple[str, bytes]:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            candidate_files = [
                zip_info
                for zip_info in archive.infolist()
                if not zip_info.is_dir() and zip_info.filename.lower().endswith(".pdf")
            ]
            if not candidate_files:
                raise self.QCFReportsServiceError(f"No PDF advice file found in archive {filename}")
            if len(candidate_files) > 1:
                raise self.QCFReportsServiceError(f"More than one PDF advice file found in archive {filename}")

            zip_info = candidate_files[0]
            advice_name = zip_info.filename.rsplit("/", 1)[-1].rsplit(".", 1)[0]
            return advice_name, archive.read(zip_info)

    def extract_text_from_pdf(self, filename: str, pdf_bytes: bytes) -> str:
        try:
            pages = PdfReader(io.BytesIO(pdf_bytes)).pages
        except (OSError, TypeError, ValueError) as exc:
            raise self.QCFReportsServiceError(
                f"Could not extract text from Western Union AD PDF in {filename}: {exc}"
            ) from exc

        pdf_text = "\n".join(page.extract_text() or "" for page in pages)
        if not pdf_text.strip():
            raise self.QCFReportsServiceError(f"No text extracted from Western Union AD PDF in {filename}")
        return pdf_text

    def decode_rows(self, filename: str, file_bytes: bytes) -> list[list[str]]:
        try:
            decoded_content = file_bytes.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise self.QCFReportsServiceError(f"Could not decode file {filename} as UTF-8") from exc

        return list(csv.reader(io.StringIO(decoded_content)))

    def extract_data_rows(self, rows: list[list[str]], filename: str) -> list[list[str]]:
        if not rows:
            raise self.QCFReportsServiceError(f"Data file {filename} has no rows")

        data_rows = [
            row
            for row in rows[1:]
            if row and self.clean_field(row[self.DataFieldIndex.RECORD_TYPE]) in self.DataRecordType.DETAIL_TYPES
        ]
        if not data_rows:
            raise self.QCFReportsServiceError(f"Data file {filename} has no data rows")

        if any(len(row) <= self.DataFieldIndex.last_required_index() for row in data_rows):
            raise self.QCFReportsServiceError(f"Invalid data row length in file {filename}")

        return data_rows

    def extract_footer_row(self, rows: list[list[str]], filename: str) -> list[str]:
        footer_rows = [row for row in rows[1:] if row and self.clean_field(row[0]) == self.DataRecordType.FOOTER]
        if not footer_rows:
            raise self.QCFReportsServiceError(f"Data file {filename} has no footer row")
        if len(footer_rows) > 1:
            raise self.QCFReportsServiceError(f"Data file {filename} has more than one footer row")

        footer_row = footer_rows[0]
        if len(footer_row) <= self.DataFooterFieldIndex.last_required_index():
            raise self.QCFReportsServiceError(f"Invalid data footer row length in file {filename}")

        return footer_row

    def clean_field(self, value: str) -> str:
        return str(value).strip().strip('"')

    def normalize_payment_unicef_id(self, payment_reference: str) -> str:
        normalized_reference = self.clean_field(payment_reference)
        if normalized_reference.startswith("RC"):
            return normalized_reference
        if normalized_reference.startswith("PT-"):
            return f"RC{normalized_reference}"
        raise self.QCFReportsServiceError(f"Unsupported payment reference {payment_reference}")

    def to_decimal(self, value: str) -> Decimal:
        try:
            return Decimal(self.clean_field(value))
        except InvalidOperation as exc:
            raise self.QCFReportsServiceError(f"Could not parse decimal value {value}") from exc

    def parse_principal_amount_from_pdf_text(self, pdf_text: str, filename: str) -> Decimal:
        match = self.DATA_PDF_PRINCIPAL_PATTERN.search(pdf_text)
        if not match:
            raise self.QCFReportsServiceError(f"Could not parse principal amount from AD PDF in {filename}")
        return self.to_decimal(match.group(1).replace(",", ""))

    def parse_charges_amount_from_pdf_text(self, pdf_text: str) -> Decimal:
        match = self.DATA_PDF_CHARGES_PATTERN.search(pdf_text)
        if not match:
            return Decimal("0.00")
        return self.to_decimal(match.group(1).replace(",", ""))

    def generate_report(
        self,
        payment_plan: PaymentPlan,
        payment_rows: list[MatchedDataPaymentRow],
        ad_name_value: str | None = None,
    ) -> tuple[str, openpyxl.Workbook]:
        no_of_qcf_reports_for_pp = WesternUnionPaymentPlanReport.objects.filter(payment_plan=payment_plan).count()
        report_filename = (
            f"QCF_{payment_plan.business_area.slug}_{payment_plan.unicef_id}_{no_of_qcf_reports_for_pp + 1}.xlsx"
        )

        funds_commitments_str = ", ".join(
            f"{number}/{item}"
            for number, item in payment_plan.funds_commitments.all().values_list(
                "funds_commitment_group__funds_commitment_number", "funds_commitment_item"
            )
        )
        payments_data: list[QCFReportPaymentRowData] = []
        for row in payment_rows:
            mtcn_match = str(row.mtcn) == str(row.payment.fsp_auth_code)

            payments_data.append(
                QCFReportPaymentRowData(
                    payment_unicef_id=row.payment.unicef_id,
                    mtcn=row.mtcn,
                    mtcns_match=mtcn_match,
                    hope_mtcn=row.payment.fsp_auth_code,
                    payment_plan_unicef_id=payment_plan.unicef_id,
                    programme=payment_plan.program_cycle.program.name,
                    fc=funds_commitments_str,
                    principal_amount=float(row.principal_amount),
                    charges_amount=float(row.charges_amount),
                    fee_amount=float(row.fee_amount),
                    ad_name_value=ad_name_value or "",
                )
            )

        report_data = QCFReportPaymentPlanData(
            payment_plan_unicef_id=payment_plan.unicef_id,
            payments_data=payments_data,
        )

        return report_filename, self.generate_report_xlsx(report_data)

    def generate_report_xlsx(self, report_data: QCFReportPaymentPlanData) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"QCF Report - {report_data.payment_plan_unicef_id}"

        headers = [
            "Payment ID",
            "MTCN",
            "MTCNs match",
            "HOPE MTCN",
            "Payment Plan ID",
            "Programme",
            "FC",
            "Principal Amount",
            "Charges Amount",
            "Fee Amount",
            "Advice Filename",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for payment_data in report_data.payments_data:
            ws.append(
                [
                    payment_data.payment_unicef_id,
                    payment_data.mtcn,
                    payment_data.mtcns_match,
                    payment_data.hope_mtcn,
                    payment_data.payment_plan_unicef_id,
                    payment_data.programme,
                    payment_data.fc,
                    payment_data.principal_amount,
                    payment_data.charges_amount,
                    payment_data.fee_amount,
                    payment_data.ad_name_value,
                ]
            )

        ws.append([])

        ws.append(["Principal Total", "", "", "", report_data.principal_total])
        ws.append(["Charges Total", "", "", "", report_data.charges_total])
        ws.append(["Refunds Total", "", "", "", report_data.refunds_total])

        last_row = ws.max_row
        for row_number in range(last_row - 3, last_row + 1):
            ws[f"A{row_number}"].font = Font(bold=True)

        return wb

    def send_notification_emails(self, report: WesternUnionPaymentPlanReport) -> None:
        business_area = report.payment_plan.business_area
        users = [
            user for user in User.objects.all() if user.has_perm(Permissions.RECEIVE_PARSED_WU_QCF.name, business_area)
        ]
        if users:
            text_template = "payment/qcf_report_email.txt"
            html_template = "payment/qcf_report_email.html"

            path_name = "download-payment-plan-invoice-report-pdf"
            payment_plan = report.payment_plan

            report_id = str(report.id)
            payment_plan_id = str(payment_plan.id)
            program_code = report.payment_plan.program.code

            payment_plan_link = get_link(
                f"/{payment_plan.business_area.slug}/programs/{program_code}/payment-module/payment-plans/{payment_plan_id}"
            )
            download_link = get_link(reverse(path_name, args=[report_id]))

        for user in users:
            context = {
                "first_name": getattr(user, "first_name", ""),
                "last_name": getattr(user, "last_name", ""),
                "email": getattr(user, "email", ""),
                "message": f"Payment Plan: {payment_plan_link}",
                "title": f"Payment Plan {report.report_file.file.name} Western Union QCF Report",
                "link": f"Western Union QCF Report file: {download_link}",
            }
            user.email_user(
                subject=context["title"],
                html_body=render_to_string(html_template, context=context),
                text_body=render_to_string(text_template, context=context),
            )
