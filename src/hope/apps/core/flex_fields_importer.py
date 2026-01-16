from collections import defaultdict
from io import BytesIO
import logging
from os.path import isfile
from typing import Any

from django.core.exceptions import ValidationError
from django.core.files import File
from django.db import transaction
from django.utils.html import strip_tags
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from xlwt import Row

from hope.apps.core.field_attributes.fields_types import (
    TYPE_DATE,
    TYPE_DECIMAL,
    TYPE_IMAGE,
    TYPE_INTEGER,
    TYPE_SELECT_MANY,
    TYPE_SELECT_ONE,
    TYPE_STRING,
)
from hope.models import FlexibleAttribute, FlexibleAttributeChoice, FlexibleAttributeGroup

logger = logging.getLogger(__name__)


class FlexibleAttributeImporter:
    TYPE_CHOICE_MAP = {
        "note": TYPE_STRING,
        "image": TYPE_IMAGE,
        "select_one": TYPE_SELECT_ONE,
        "text": TYPE_STRING,
        "integer": TYPE_INTEGER,
        "decimal": TYPE_DECIMAL,
        "date": TYPE_DATE,
        "select_multiple": TYPE_SELECT_MANY,
    }
    CALCULATE_TYPE_CHOICE_MAP = {
        "text": TYPE_STRING,
        "integer": TYPE_INTEGER,
        "decimal": TYPE_DECIMAL,
        "date": TYPE_DATE,
    }

    # Constants for xls import
    ATTRIBUTE_MODEL_FIELDS = [field.name for field in FlexibleAttribute._meta.get_fields()]

    GROUP_MODEL_FIELDS = [field.name for field in FlexibleAttributeGroup._meta.get_fields()]

    CHOICE_MODEL_FIELDS = [field.name for field in FlexibleAttributeChoice._meta.get_fields()]

    CORE_FIELD_SUFFIXES = (
        "_h_c",
        "_i_c",
    )

    FLEX_FIELD_SUFFIXES = (
        "_h_f",
        "_i_f",
    )

    JSON_MODEL_FIELDS = (
        "label",
        "hint",
    )

    EXCLUDED_MODEL_FIELDS = (
        "start",
        "end",
        "deviceid",
    )

    def _get_model_fields(self, object_type_to_add: Any) -> list[str] | None:
        return {
            "attribute": self.ATTRIBUTE_MODEL_FIELDS,
            "group": self.GROUP_MODEL_FIELDS,
            "choice": self.CHOICE_MODEL_FIELDS,
        }.get(object_type_to_add)

    def _handle_calculate_field(
        self,
        object_type_to_add: str,
        header_name: str,
        row,
        row_number: int,
        value: Any,
    ) -> None:
        is_calculate_field = (
            object_type_to_add == "attribute"
            and header_name == "calculated_result_field_type"
            and str(row[0].value) == "calculate"
            and any(self.object_fields_to_create["name"].endswith(suffix) for suffix in self.FLEX_FIELD_SUFFIXES)
        )
        if not is_calculate_field:
            return

        choice_key = str(value).strip() if value else None
        if choice_key is None:
            msg = (
                f"Survey Sheet: Row {row_number}: "
                "Calculated result field type must be provided for calculate type fields"
            )
            logger.warning(msg)
            raise ValidationError(msg)
        if choice_key not in self.CALCULATE_TYPE_CHOICE_MAP:
            msg = (
                f"Survey Sheet: Row {row_number}: "
                f"Invalid type: {choice_key} for calculate field, valid choices are "
                f"{', '.join(self.CALCULATE_TYPE_CHOICE_MAP.keys())}"
            )
            logger.warning(msg)
            raise ValidationError(msg)
        self.object_fields_to_create["type"] = self.CALCULATE_TYPE_CHOICE_MAP[choice_key]

    def _assign_field_values(
        self,
        value: Any,
        header_name: str,
        object_type_to_add: str,
        row,  # openpyxl Row object (tuple of Cells)
        row_number: int,  # should match Excel row number
    ) -> None:
        # Get model fields for this object type
        model_fields = self._get_model_fields(object_type_to_add)
        if not model_fields:
            return

        # Handle JSON fields (like label::English(EN))
        if any(header_name.startswith(i) for i in self.JSON_MODEL_FIELDS):
            if "::" in header_name:
                label, language = header_name.split("::")
            else:
                label, language = header_name.split(":")
            if label in model_fields:
                cleared_value = strip_tags(str(value)).replace("#", "").strip() if value else ""

                # English label validation
                if label == "label" and language == "English(EN)":
                    cell_name = str(row[1].value) if row[1].value else ""
                    is_index_field = cell_name.endswith("_index")

                    if object_type_to_add == "attribute":
                        field_suffix = cell_name[-4:]
                        is_empty_and_not_index_field = not value and not is_index_field
                        is_core_or_flex_field = (
                            field_suffix in self.CORE_FIELD_SUFFIXES or field_suffix in self.FLEX_FIELD_SUFFIXES
                        )
                        if is_empty_and_not_index_field and is_core_or_flex_field:
                            logger.warning(f"Survey Sheet: Row {row_number}: English label cannot be empty")
                            raise ValidationError(f"Survey Sheet: Row {row_number}: English label cannot be empty")

                    if object_type_to_add == "choice" and not value:
                        logger.warning(f"Choices Sheet: Row {row_number}: English label cannot be empty")
                        raise ValidationError(f"Choices Sheet: Row {row_number}: English label cannot be empty")

                self.json_fields_to_create[label].update({language: cleared_value if value else ""})
            return

        # Handle required field
        if header_name == "required":
            self.object_fields_to_create[header_name] = value == "true"
            return

        # Handle normal fields
        if header_name in model_fields:
            if header_name == "type":
                if not value:
                    logger.warning(f"Survey Sheet: Row {row_number}: Type is required")
                    raise ValidationError(f"Survey Sheet: Row {row_number}: Type is required")

                choice_key = str(value).split(" ")[0]
                if choice_key == "calculate":
                    self.object_fields_to_create["type"] = "calculate"
                elif choice_key in self.TYPE_CHOICE_MAP:
                    self.object_fields_to_create["type"] = self.TYPE_CHOICE_MAP.get(choice_key)
            else:
                is_attribute_name_empty = header_name == "name" and value in (None, "")
                is_choice_list_name_empty = (
                    header_name == "list_name" and object_type_to_add == "choice"
                ) and not value

                if is_attribute_name_empty:
                    logger.warning(f"Survey Sheet: Row {row_number}: Name is required")
                    raise ValidationError(f"Survey Sheet: Row {row_number}: Name is required")

                if is_choice_list_name_empty:
                    logger.warning(f"Survey Sheet: Row {row_number}: List Name is required")
                    raise ValidationError(f"Survey Sheet: Row {row_number}: List Name is required")

                self.object_fields_to_create[header_name] = str(value) if value else ""

        self._handle_calculate_field(object_type_to_add, header_name, row, row_number, value)

    def _handle_group_object(self, repeatable: bool, parent: FlexibleAttributeGroup | None) -> FlexibleAttributeGroup:
        obj = FlexibleAttributeGroup.all_objects.filter(
            name=self.object_fields_to_create["name"],
        ).first()

        if obj:
            obj.label = self.json_fields_to_create["label"]
            obj.hint = self.json_fields_to_create["hint"]
            obj.repeatable = repeatable
            obj.parent = parent
            obj.is_removed = False
            obj.save()
            group = obj
        else:
            group = FlexibleAttributeGroup.objects.create(
                **self.object_fields_to_create,
                **self.json_fields_to_create,
                repeatable=repeatable,
                parent=parent,
            )

        self.current_group_tree.append(group)
        FlexibleAttributeGroup.objects.rebuild()
        return group

    def _can_add_row(self, row: Row) -> bool:
        cell_0 = str(row[0].value) if row[0].value is not None else ""
        cell_1 = str(row[1].value) if row[1].value is not None else ""

        # Core field check
        is_core_field = any(cell_1.endswith(suffix) for suffix in self.CORE_FIELD_SUFFIXES) and not cell_0.endswith(
            "_group"
        )

        # Excluded fields
        is_in_excluded = cell_0 in self.EXCLUDED_MODEL_FIELDS

        # Version field
        is_version_field = cell_1 == "__version__"

        # End of group/repeat
        is_end_info = cell_0 in ("end_repeat", "end_group")
        if is_end_info:
            if self.current_group_tree:
                self.current_group_tree.pop()
            else:
                self.current_group_tree = []
            return False

        return not (is_core_field or is_in_excluded or is_version_field)

    def _get_list_of_field_choices(self, sheet: Worksheet) -> set:
        fields_with_choices = []
        for row_number in range(2, sheet.max_row + 1):  # skip header row
            row = sheet[row_number]  # returns a tuple of Cell objects
            cell_value = row[0].value  # first column (A)
            if cell_value and str(cell_value).startswith("select_"):
                fields_with_choices.append(row)

        # Extract the choice name after the first space
        return {row[0].value.split(" ")[1] for row in fields_with_choices if row[0].value and " " in row[0].value}

    def _get_field_choice_name(self, row: Row) -> str | None:
        has_choice = row[0].value.startswith("select_")
        if has_choice:
            return row[0].value.split(" ")[1]
        return None

    def _reset_model_fields_variables(self) -> None:
        self.json_fields_to_create = defaultdict(dict)
        self.object_fields_to_create = {}

    def _handle_choices(self, sheets: dict | Worksheet) -> None:
        choices_assigned_to_fields = self._get_list_of_field_choices(sheets["survey"])
        choices_from_db = FlexibleAttributeChoice.objects.all()

        # First row of "choices" sheet for headers
        choices_first_row = sheets["choices"][1]  # openpyxl 1-indexed; first row
        choices_headers_map = [cell.value for cell in choices_first_row]

        to_create_choices = []
        updated_choices = []

        # Loop through rows, skipping header
        for row_number in range(2, sheets["choices"].max_row + 1):  # start at row 2
            row = sheets["choices"][row_number]
            self._reset_model_fields_variables()

            # Skip empty rows
            if all(cell.value is None for cell in row):
                continue

            # Skip rows not assigned to fields
            if row[0].value not in choices_assigned_to_fields:
                continue

            # Assign field values
            for cell, header_name in zip(row, choices_headers_map, strict=True):
                cell_value = cell.value
                if isinstance(cell_value, float) and cell_value.is_integer():
                    cell_value = str(int(cell_value))
                self._assign_field_values(
                    cell_value,
                    header_name,
                    "choice",
                    row,
                    row_number - 1,  # match previous 0-indexing if needed
                )

            # Update existing or prepare new choice
            obj = FlexibleAttributeChoice.all_objects.filter(
                list_name=self.object_fields_to_create["list_name"],
                name=self.object_fields_to_create["name"],
            ).first()

            if obj:
                obj.label = self.json_fields_to_create["label"]
                obj.is_removed = False
                obj.save()
                updated_choices.append(obj)
            else:
                choice = FlexibleAttributeChoice(
                    **self.object_fields_to_create,
                    **self.json_fields_to_create,
                )
                to_create_choices.append(choice)

        # Bulk create new choices
        created_choices = FlexibleAttributeChoice.objects.bulk_create(to_create_choices)

        # Delete choices no longer present
        choices_to_delete = set(choices_from_db).difference(set(created_choices + updated_choices))
        for choice in choices_to_delete:
            choice.delete()

    def _handle_groups_and_fields(self, sheet: Worksheet) -> None:
        groups_from_db, attrs_from_db = (
            FlexibleAttributeGroup.objects.all(),
            FlexibleAttribute.objects.all(),
        )

        # openpyxl first row (header) is 1-indexed
        first_row = sheet[1]
        headers_map = [cell.value for cell in first_row]

        all_attrs = []
        all_groups = []

        for row_number in range(2, sheet.max_row + 1):  # skip header row
            row = sheet[row_number]

            # Skip empty rows
            if all(cell.value is None for cell in row):
                continue

            # Determine type of object
            object_type_to_add = "group" if row[0].value in ("begin_group", "begin_repeat") else "attribute"
            repeatable = bool(row[0].value == "begin_repeat")
            self._reset_model_fields_variables()

            if not self._can_add_row(row):
                continue

            # Assign values for each column
            for cell, header_name in zip(row, headers_map, strict=True):
                value = cell.value
                self._assign_field_values(
                    value,
                    header_name,
                    object_type_to_add,
                    row,
                    row_number - 1,  # keep previous 0-indexed row logic
                )

            # Determine if flexible field
            is_flex_field = any(
                self.object_fields_to_create["name"].endswith(suffix) for suffix in self.FLEX_FIELD_SUFFIXES
            )

            parent = self.current_group_tree[-1] if self.current_group_tree else None

            # Handle group objects
            if object_type_to_add == "group":
                group = self._handle_group_object(repeatable, parent)
                all_groups.append(group)

            # Handle attribute objects
            elif object_type_to_add == "attribute" and is_flex_field:
                choice_name = self._get_field_choice_name(row)
                obj = FlexibleAttribute.all_objects.filter(
                    name=self.object_fields_to_create["name"],
                ).first()

                if obj:
                    if obj.type != self.object_fields_to_create["type"] and not obj.is_removed:
                        logger.warning(f"Survey Sheet: Row {row_number}: Type of the attribute cannot be changed!")
                        raise ValidationError(
                            f"Survey Sheet: Row {row_number}: Type of the attribute cannot be changed!"
                        )
                    obj.type = self.object_fields_to_create["type"]
                    obj.name = self.object_fields_to_create["name"]
                    obj.required = self.object_fields_to_create["required"]
                    obj.label = self.json_fields_to_create["label"]
                    obj.hint = self.json_fields_to_create["hint"]
                    obj.is_removed = False
                    obj.save()
                    field = obj
                else:
                    attribute_suffix = self.object_fields_to_create["name"][-4:]
                    field = FlexibleAttribute.objects.create(
                        group=parent,
                        associated_with=(0 if attribute_suffix == "_h_f" else 1),
                        **self.object_fields_to_create,
                        **self.json_fields_to_create,
                    )

                if choice_name:
                    choices = FlexibleAttributeChoice.objects.filter(list_name=choice_name)
                    field.choices.set(choices)

                all_attrs.append(field)

        # Delete removed groups and attributes
        groups_to_delete = set(groups_from_db).difference(all_groups)
        for group in groups_to_delete:
            group.delete()

        attrs_to_delete = set(attrs_from_db).difference(all_attrs)
        for attr in attrs_to_delete:
            attr.delete()

    # Variables initialized for model creation
    current_group_tree = None
    json_fields_to_create = defaultdict(dict)
    object_fields_to_create = {}
    can_add_flag = True

    @transaction.atomic
    def import_xls(self, xls_file: File | str) -> None:
        self.current_group_tree = [None]
        if isinstance(xls_file, str) and isfile(xls_file):
            wb = load_workbook(filename=xls_file, data_only=True)
        else:  # noqa
            xls_file.seek(0)
            wb = load_workbook(filename=BytesIO(xls_file.read()), data_only=True)
        sheets = {
            "survey": wb["survey"],
            "choices": wb["choices"],
        }
        self._handle_choices(sheets)
        self._handle_groups_and_fields(sheets["survey"])
