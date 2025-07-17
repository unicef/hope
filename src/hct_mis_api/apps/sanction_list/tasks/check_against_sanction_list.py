import base64
import io
from datetime import date, datetime
from itertools import permutations
from typing import TYPE_CHECKING

from django.conf import settings
from django.db.models import Q
from django.template.loader import render_to_string
from django.utils import timezone

import dateutil.parser
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from hct_mis_api.apps.sanction_list.models import (
    SanctionListIndividual,
    UploadedXLSXFile,
)
from hct_mis_api.apps.utils.mailjet import MailjetClient

if TYPE_CHECKING:
    from uuid import UUID


class CheckAgainstSanctionListTask:
    def execute(self, uploaded_file_id: "UUID", original_file_name: str) -> None:
        today = timezone.now()
        uploaded_file = UploadedXLSXFile.objects.get(id=uploaded_file_id)
        wb = load_workbook(uploaded_file.file, data_only=True)
        sheet = wb.worksheets[0]
        headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
        results_dict = {}

        for row in sheet.iter_rows(min_row=2):
            if not any([cell.value for cell in row]):
                continue

            filter_values = {
                "first_name": "",
                "second_name": "",
                "third_name": "",
                "fourth_name": "",
                "date_of_birth": "",
            }

            row_number = 1
            for cell, header in zip(row, headers.keys()):
                value = cell.value
                row_number = cell.row
                header_as_key = header.replace(" ", "_").lower().strip()
                if header_as_key == "date_of_birth":
                    if not isinstance(value, (datetime, date)):
                        try:
                            value = dateutil.parser.parse(value)
                        except Exception:
                            pass
                if value:
                    filter_values[header_as_key] = value

            dob = filter_values.pop("date_of_birth", "")
            names = [str(n).capitalize().strip() for n in filter_values.values() if n]

            name_permutations = permutations(names)
            full_name_permutations = [" ".join(permutation).title() for permutation in name_permutations]

            if isinstance(dob, datetime):
                dob_query = (
                    Q(dates_of_birth__date=dob.date())
                    | Q(dates_of_birth__date__year=dob.year)
                    # to return something when full_name matches but dob not
                    | Q(first_name__isnull=False)
                )
            else:
                dob_query = Q(full_name__isnull=False)

            if len(names) == 0:
                continue
            elif len(names) == 1:
                name_query = Q(full_name__in=full_name_permutations) | Q(first_name__iexact=names[0])
            else:
                name_query = Q(full_name__in=full_name_permutations) | (
                    Q(full_name__icontains=names[0]) & Q(full_name__icontains=names[1])
                )

            qs = SanctionListIndividual.objects.filter(name_query & dob_query).first()

            if qs:
                results_dict[row_number] = qs

        # MAIL SENDING
        context = {
            "results": results_dict,
            "results_count": len(results_dict),
            "file_name": original_file_name,
            "today_date": timezone.now(),
            "title": "Sanction List Check",
        }
        text_body = render_to_string("sanction_list/check_results.txt", context)
        html_body = render_to_string("sanction_list/check_results.html", context)
        subject = f"Sanction List Check - file: {original_file_name}, " f"date: {today.strftime('%Y-%m-%d %I:%M %p')}"

        attachment_wb = Workbook()
        attachment_ws = attachment_wb.active
        attachment_ws.title = "Sanction List Check Results"

        header_row_names = (
            "FIRST NAME",
            "SECOND NAME",
            "THIRD NAME",
            "FOURTH NAME",
            "DATE OF BIRTH",
            "ORIGINAL FILE ROW NUMBER",
        )
        attachment_ws.append(header_row_names)

        for row_number, individual in results_dict.items():
            attachment_ws.append(
                (
                    individual.first_name,
                    individual.second_name,
                    individual.third_name,
                    individual.fourth_name,
                    ", ".join(d.strftime("%Y-%m-%d") for d in individual.dates_of_birth.values_list("date", flat=True)),
                    row_number,
                )
            )
        for i in range(1, len(header_row_names) + 1):
            attachment_ws.column_dimensions[get_column_letter(i)].width = 30

        buffer = io.BytesIO()
        attachment_wb.save(buffer)
        buffer.seek(0)

        attachment_content = buffer.getvalue()
        base64_encoded_content = base64.b64encode(attachment_content).decode("utf-8")

        email = MailjetClient(
            subject=subject,
            recipients=[uploaded_file.associated_email],
            html_body=html_body,
            text_body=text_body,
            ccs=[settings.SANCTION_LIST_CC_MAIL],
        )
        email.attach_file(
            attachment=base64_encoded_content,
            filename=f"{subject}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        email.send_email()
