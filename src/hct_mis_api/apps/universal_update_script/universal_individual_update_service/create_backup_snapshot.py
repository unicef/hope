import json
import tempfile
from typing import Callable, Optional

from django.core.files.base import ContentFile

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from hct_mis_api.apps.core.utils import chunks
from hct_mis_api.apps.household.models import Household, Individual
from hct_mis_api.apps.payment.services.payment_household_snapshot_service import (
    get_household_snapshot,
)
from hct_mis_api.apps.universal_update_script.models import UniversalUpdate


def _get_unicef_ids_from_sheet(ws: Worksheet) -> list[str]:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        col_index = header.index("unicef_id") + 1
    except ValueError:
        raise ValueError("The column 'unicef_id' was not found in the header row.")
    return [row[col_index - 1] for row in ws.iter_rows(min_row=2, values_only=True)]


def _get_unicef_ids_from_workbook(workbook: Workbook, sheet_name: Optional[str] = None) -> list[str]:
    ws = workbook[sheet_name] if sheet_name else workbook.active
    return _get_unicef_ids_from_sheet(ws)


def create_and_save_snapshot_chunked(universal_update: UniversalUpdate) -> None:
    universal_update.update_file.open("rb")
    workbook = load_workbook(universal_update.update_file, data_only=True)
    unicef_ids = _get_unicef_ids_from_workbook(workbook)
    log_message: Callable[[str], None] = lambda message_log: universal_update.save_logs(message_log)
    program_id = universal_update.program_id
    content = create_snapshot_content(log_message, str(program_id), unicef_ids)
    content_bytes = content.encode("utf-8")
    universal_update.backup_snapshot.save("snapshot.json", ContentFile(content_bytes))
    universal_update.save()


def create_snapshot_content(log_message: Callable[[str], None], program_id: str, unicef_ids: [str]) -> str:
    db_count = Individual.objects.filter(unicef_id__in=unicef_ids, program_id=program_id).count()
    if db_count != len(unicef_ids):
        unicef_ids_from_db = list(
            Individual.objects.filter(unicef_id__in=unicef_ids, program_id=program_id).values_list(
                "unicef_id", flat=True
            )
        )
        diff = set(unicef_ids) - set(unicef_ids_from_db)
        raise Exception("Some unicef ids are not in the program: " + str(diff))
    household_ids = list(
        Individual.objects.filter(unicef_id__in=unicef_ids, program_id=program_id)
        .distinct("household_id")
        .values_list("household_id", flat=True)
    )
    with tempfile.NamedTemporaryFile(mode="w+", delete=False, suffix=".json") as tmp_file:
        tmp_file.write("[\n")
        first_item = True
        chunk_size = 10
        length = len(household_ids)
        index = 0
        for id_chunk in chunks(household_ids, chunk_size):
            message = (
                f"Creating backup snapshot for  records {index} to {min(index + chunk_size, length)} out of {length}"
            )
            log_message(message)
            index += 1
            households = (
                Household.objects.filter(id__in=id_chunk)
                .select_related("head_of_household")
                .prefetch_related("individuals", "individuals_and_roles")
            )
            for household in households:
                snapshot = get_household_snapshot(household)
                if not first_item:
                    tmp_file.write(",\n")
                else:
                    first_item = False
                tmp_file.write(json.dumps(snapshot))
        tmp_file.write("\n]")
        tmp_file.flush()
        tmp_file.seek(0)
        content = tmp_file.read()
    return content
