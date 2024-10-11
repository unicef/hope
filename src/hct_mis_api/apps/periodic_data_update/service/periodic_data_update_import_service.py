import json
from typing import Any, Optional, Union

from django import forms
from django.core.exceptions import ValidationError
from django.core.files import File
from django.db import transaction

import openpyxl

from hct_mis_api.apps.core.models import FlexibleAttribute, PeriodicFieldData
from hct_mis_api.apps.household.models import Individual
from hct_mis_api.apps.periodic_data_update.models import (
    PeriodicDataUpdateTemplate,
    PeriodicDataUpdateUpload,
)
from hct_mis_api.apps.periodic_data_update.service.periodic_data_update_export_template_service import (
    PeriodicDataUpdateExportTemplateService,
)


class PeriodicDataUpdateBaseForm(forms.Form):
    individual__uuid = forms.UUIDField()
    individual_unicef_id = forms.CharField()
    first_name = forms.CharField(required=False)
    last_name = forms.CharField(required=False)


class StrictBooleanField(forms.Field):
    def to_python(self, value: Union[Optional[str], bool]) -> Optional[bool]:
        if value is None or value == "":
            return None
        if value in (True, "True", "true", "TRUE", "1"):
            return True
        elif value in (False, "False", "false", "FALSE", "0"):
            return False
        else:
            raise ValidationError("Invalid boolean value", code="invalid")


class RowValidationError(ValidationError):
    pass


def validation_error_to_json(
    error: Union[ValidationError, list[ValidationError]], seen: Optional[set[int]] = None
) -> Any:  # pragma: no cover
    """
    Recursively convert a Django ValidationError into a JSON-serializable format.
    Handles nested ValidationError instances within dicts and lists, avoiding infinite recursion by tracking seen objects.

    Args:
        error (ValidationError): The ValidationError instance to convert.
        seen (set): Set of id's of already processed ValidationError objects to avoid infinite recursion.

    Returns:
        dict or list: A JSON-serializable representation of the errors.
    """
    if seen is None:
        seen = set()

    # Check if this instance is already processed to prevent infinite recursion
    if id(error) in seen:
        return error.message

    seen.add(id(error))
    # if isinstance(error, RowValidationError):

    if hasattr(error, "error_dict"):
        # Handle ValidationError with error_dict (errors from ModelForms, for example)
        return {key: validation_error_to_json(value, seen) for key, value in error.error_dict.items()}
    elif hasattr(error, "error_list"):
        # Handle ValidationError with error_list (non-field errors and formset errors)
        return [validation_error_to_json(e, seen) for e in error.error_list]
    else:
        # Handle simple ValidationError instances
        return error.messages if hasattr(error, "messages") else str(error)


class PeriodicDataUpdateImportService:
    def __init__(self, periodic_data_update_upload: PeriodicDataUpdateUpload) -> None:
        self.periodic_data_update_upload = periodic_data_update_upload
        self.periodic_data_update_template = self.periodic_data_update_upload.template
        self.file = self.periodic_data_update_upload.file

    def import_data(self) -> None:
        try:
            self.periodic_data_update_upload.status = PeriodicDataUpdateUpload.Status.PROCESSING
            self.periodic_data_update_upload.save()
            with transaction.atomic():
                self._open_workbook()
                self._read_flexible_attributes()
                cleaned_data_list, form_errors = self._read_rows()
                if form_errors:
                    self.periodic_data_update_upload.status = PeriodicDataUpdateUpload.Status.FAILED
                    self.periodic_data_update_upload.error_message = json.dumps(
                        {
                            "form_errors": form_errors,
                            "non_form_errors": None,
                        }
                    )
                    self.periodic_data_update_upload.save()
                    return
                self._update_individuals(cleaned_data_list)
                self.file.close()
                self.periodic_data_update_upload.status = PeriodicDataUpdateUpload.Status.SUCCESSFUL
                self.periodic_data_update_upload.save()
        except ValidationError as e:
            self.periodic_data_update_upload.status = PeriodicDataUpdateUpload.Status.FAILED
            self.periodic_data_update_upload.error_message = json.dumps(
                {
                    "non_form_errors": validation_error_to_json(e),
                    "form_errors": [],
                }
            )
            self.periodic_data_update_upload.save()
            return

    def _open_workbook(self) -> None:
        self.wb = openpyxl.load_workbook(self.file)
        self.ws_pdu = self.wb[PeriodicDataUpdateExportTemplateService.PDU_SHEET]
        self.ws_meta = self.wb[PeriodicDataUpdateExportTemplateService.META_SHEET]
        self.periodic_data_update_template_id: Optional[PeriodicDataUpdateTemplate] = None
        self.flexible_attributes_dict: Optional[dict[str, FlexibleAttribute]] = None

    @classmethod
    def read_periodic_data_update_template_object(cls, file: File) -> PeriodicDataUpdateTemplate:
        wb = openpyxl.load_workbook(file)  # type: ignore
        ws_meta = wb[PeriodicDataUpdateExportTemplateService.META_SHEET]
        try:
            periodic_data_update_template_id = wb.custom_doc_props[
                PeriodicDataUpdateExportTemplateService.PROPERTY_ID_NAME
            ]
        except KeyError:
            periodic_data_update_template_id = None
        if periodic_data_update_template_id:
            periodic_data_update_template_id = periodic_data_update_template_id.value
        if not periodic_data_update_template_id:
            periodic_data_update_template_id = ws_meta[PeriodicDataUpdateExportTemplateService.META_ID_ADDRESS].value
        if not periodic_data_update_template_id:
            raise ValidationError("Periodic Data Update Template ID is missing in the file")
        try:
            if isinstance(periodic_data_update_template_id, str):
                periodic_data_update_template_id = periodic_data_update_template_id.strip()
                periodic_data_update_template_id = int(periodic_data_update_template_id)
        except ValueError:
            raise ValidationError("Periodic Data Update Template ID must be a number")
        if type(periodic_data_update_template_id) is not int:
            raise ValidationError("Periodic Data Update Template ID must be an integer")

        periodic_data_update_template = PeriodicDataUpdateTemplate.objects.filter(
            id=periodic_data_update_template_id
        ).first()
        if not periodic_data_update_template:
            raise ValidationError(f"Periodic Data Update Template with ID {periodic_data_update_template_id} not found")
        return periodic_data_update_template

    def _read_flexible_attributes(self) -> None:
        rounds_data = self.periodic_data_update_template.rounds_data
        fields_name_list = [field["field"] for field in rounds_data]
        fields_name = set(fields_name_list)
        flexible_attributes = FlexibleAttribute.objects.filter(
            name__in=fields_name, type=FlexibleAttribute.PDU, program=self.periodic_data_update_template.program
        )
        if len(flexible_attributes) != len(fields_name):
            raise ValidationError("Some fields are missing in the flexible attributes")
        self.flexible_attributes_dict = {field.name: field for field in flexible_attributes}

    def _read_header(self) -> list[str]:
        header = [cell.value for cell in self.ws_pdu[1]]
        return header

    def _read_rows(self) -> tuple[list[dict], list]:
        header = self._read_header()
        errors = []
        cleaned_data_list = []
        for index, row in enumerate(self.ws_pdu.iter_rows(min_row=2, values_only=True)):
            row_number = index + 2
            cleaned_data = self._read_row(errors, header, row, row_number)
            if not cleaned_data:
                continue
            cleaned_data_list.append(cleaned_data)
        return cleaned_data_list, errors

    def _read_row(self, errors: list, header: list, row: list, row_number: int) -> Optional[dict]:
        row_empty_values = []
        for value in row:
            if value == "-":
                row_empty_values.append(None)
            else:
                row_empty_values.append(value)
        data = dict(zip(header, row_empty_values))
        form = self._build_form()(data=data)
        if not form.is_valid():
            form_errors = json.loads(form.errors.as_json())
            errors.append(
                {
                    "row": row_number,
                    "errors": form_errors,
                }
            )
            return None
        cleaned_data = form.cleaned_data
        return cleaned_data

    def _update_individuals(self, cleaned_data_list: list[dict]) -> None:
        individuals = []
        for cleaned_data in cleaned_data_list:
            individual = self._import_cleaned_data(cleaned_data)
            individuals.append(individual)
        Individual.objects.bulk_update(individuals, ["flex_fields"])

    def _import_cleaned_data(self, cleaned_data: dict) -> Individual:
        individual_uuid = cleaned_data["individual__uuid"]
        individual_unicef_id = cleaned_data["individual_unicef_id"]
        individual = Individual.objects.filter(id=individual_uuid).first()
        for round in self.periodic_data_update_template.rounds_data:
            field_name = round["field"]
            round_number = round["round"]
            round_number_from_xlsx = cleaned_data[f"{field_name}__round_number"]
            value_from_xlsx = cleaned_data[f"{field_name}__round_value"]
            collection_date_from_xlsx = cleaned_data[f"{field_name}__collection_date"]
            if value_from_xlsx is None:  # pragma: no cover
                continue
            if value_from_xlsx == "":  # pragma: no cover
                continue
            if round_number_from_xlsx != round_number:
                raise ValidationError(
                    f"Round number mismatch for field {field_name} and individual {individual_unicef_id}"
                )
            if not individual:
                raise ValidationError(f"Individual not found for {individual_unicef_id} ")
            current_value = self._get_round_value(individual, field_name, round_number)
            if current_value and value_from_xlsx:
                raise ValidationError(
                    f"Value already exists for field {field_name} for round {round_number} and individual {individual_unicef_id}"
                )
            self.set_round_value(
                individual,
                field_name,
                round_number,
                value_from_xlsx,
                (
                    collection_date_from_xlsx
                    if collection_date_from_xlsx
                    else self.periodic_data_update_template.created_at.date()
                ),
            )
        return individual

    def _get_round_value(
        self, individual: Individual, pdu_field_name: str, round_number: int
    ) -> Optional[Union[str, int, float, bool]]:
        flex_fields_data = individual.flex_fields
        field_data = flex_fields_data.get(pdu_field_name)
        if field_data:
            round_data = field_data.get(str(round_number))
            if round_data:
                return round_data.get("value")
        return None

    @staticmethod
    def set_round_value(
        individual: Individual, pdu_field_name: str, round_number: int, value: Any, collection_date: Any
    ) -> None:
        flex_fields_data = individual.flex_fields
        field_data = flex_fields_data[pdu_field_name]
        round_data = field_data.get(str(round_number))
        round_data["value"] = value
        round_data["collection_date"] = collection_date

    def _build_form(self) -> type[forms.Form]:
        form_fields_dict: dict[str, forms.Field] = {}
        for round in self.periodic_data_update_template.rounds_data:
            flexible_attribute = self.flexible_attributes_dict.get(round["field"])
            if not flexible_attribute:
                raise ValidationError(f"Flexible Attribute for field {round['field']} not found")
            form_fields_dict[f"{round['field']}__round_number"] = forms.IntegerField()
            form_fields_dict[f"{round['field']}__round_name"] = forms.CharField(required=False)
            form_fields_dict[f"{round['field']}__round_value"] = self._get_form_field_for_value(flexible_attribute)
            form_fields_dict[f"{round['field']}__collection_date"] = forms.DateField(required=False)

        return type("PeriodicDataUpdateForm", (PeriodicDataUpdateBaseForm,), form_fields_dict)

    def _get_form_field_for_value(self, flexible_attribute: FlexibleAttribute) -> forms.Field:
        if flexible_attribute.pdu_data.subtype == PeriodicFieldData.STRING:
            return forms.CharField(required=False)
        elif flexible_attribute.pdu_data.subtype == PeriodicFieldData.DECIMAL:
            return forms.FloatField(required=False)
        elif flexible_attribute.pdu_data.subtype == PeriodicFieldData.BOOL:
            return StrictBooleanField(required=False)
        elif flexible_attribute.pdu_data.subtype == PeriodicFieldData.DATE:
            return forms.DateField(required=False)
        raise ValidationError(f"Invalid subtype for field {flexible_attribute.name}")
