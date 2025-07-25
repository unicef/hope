from functools import cached_property
from typing import Any

from django.db.models import Q, QuerySet

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hct_mis_api.apps.core.utils import nested_getattr
from hct_mis_api.apps.household.models import Document, Individual
from hct_mis_api.apps.payment.models import PaymentPlan


class XlsxExportTargetingService:
    # TODO: should we refactor this service to import PaymentPlan?
    INDIVIDUALS_SHEET = "Individuals"
    META_SHEET = "Meta"
    VERSION_CELL_NAME_COORDINATES = "A1"
    VERSION_CELL_COORDINATES = "B1"
    VERSION_CELL_NAME = "FILE_TEMPLATE_VERSION"
    VERSION = "1.0"

    def __init__(self, payment_plan: PaymentPlan) -> None:
        self.payment_plan = payment_plan
        self.documents_columns_dict = {}
        self.current_header_column_index = 0
        self.COLUMNS_MAPPING_DICT = {
            "Household unicef_id": "household.unicef_id",
            "unicef_id": "unicef_id",
            "Linked Households": self._render_all_linked_households,
            "Accounts information": self._accounts_info,
        }

    @cached_property
    def households(self) -> Any:
        if self.payment_plan.status == PaymentPlan.Status.TP_OPEN:
            return self.payment_plan.household_list

        filters = {}
        if self.payment_plan.vulnerability_score_max is not None:
            filters["vulnerability_score__lte"] = self.payment_plan.vulnerability_score_max
        if self.payment_plan.vulnerability_score_min is not None:
            filters["vulnerability_score__gte"] = self.payment_plan.vulnerability_score_min
        hh_ids = list(self.payment_plan.payment_items.filter(**filters).values_list("household_id", flat=True))
        return self.payment_plan.household_list.filter(id__in=hh_ids)

    @cached_property
    def individuals(self) -> QuerySet[Individual]:
        return (
            Individual.objects.filter(
                Q(household__in=self.households, withdrawn=False, duplicate=False)
                | Q(households_and_roles__household__in=self.households)
            )
            .select_related("household")
            .prefetch_related("accounts")
            .order_by("household__unicef_id")
            .distinct()
        )

    def generate_workbook(self) -> Workbook:
        self._create_workbook()
        self._add_version()
        self._add_standard_columns_headers()
        self._add_individuals_rows()
        self._adjust_column_width_from_col(self.ws_individuals, 1, 1, self.current_header_column_index)
        return self.workbook

    def _add_version(self) -> None:
        self.ws_meta[XlsxExportTargetingService.VERSION_CELL_NAME_COORDINATES] = (
            XlsxExportTargetingService.VERSION_CELL_NAME
        )
        self.ws_meta[XlsxExportTargetingService.VERSION_CELL_COORDINATES] = XlsxExportTargetingService.VERSION

    def _create_workbook(self) -> openpyxl.Workbook:
        workbook = openpyxl.Workbook()
        self.ws_individuals = workbook.active
        self.ws_individuals.title = XlsxExportTargetingService.INDIVIDUALS_SHEET
        self.workbook = workbook
        self.ws_meta = workbook.create_sheet(XlsxExportTargetingService.META_SHEET)
        return workbook

    def _add_standard_columns_headers(self) -> None:
        standard_columns_names = list(self.COLUMNS_MAPPING_DICT.keys())
        self.ws_individuals.append(standard_columns_names)
        self.current_header_column_index += len(standard_columns_names)

    def _add_individual_row(self, individual: Individual) -> None:
        individual_row = {}
        for index, field in enumerate(self.COLUMNS_MAPPING_DICT.values()):
            if callable(field):
                value = field(individual)
            else:
                try:
                    value = nested_getattr(individual, field)
                except AttributeError:
                    value = None
            individual_row[index + 1] = value
        self._add_individual_documents_to_row(individual, individual_row)
        self.ws_individuals.append(individual_row)

    def _add_individual_documents_to_row(self, individual: Individual, row: dict) -> None:
        document: Document
        for document in individual.documents.all():
            column_index = self._add_document_column_header(document)
            row[column_index + 1] = document.document_number

    def _add_document_column_header(self, document: Document) -> int:
        type_string = str(document.type)
        if type_string in self.documents_columns_dict:
            return self.documents_columns_dict[type_string]
        self.documents_columns_dict[type_string] = self.current_header_column_index
        self.ws_individuals.cell(1, self.current_header_column_index + 1, type_string)
        old_header_column_index = self.current_header_column_index
        self.current_header_column_index += 1
        return old_header_column_index

    def _add_individuals_rows(self) -> None:
        for individual in self.individuals:
            self._add_individual_row(individual)

    def _render_all_linked_households(self, individual: Individual) -> str:
        roles_string_list = [
            f"{role.household.unicef_id} - {role.role}"
            for role in individual.households_and_roles.filter(household__in=self.households)
        ]
        return ",".join(roles_string_list)

    @staticmethod
    def _accounts_info(individual: Individual) -> str:
        if individual.accounts.exists():
            return ", ".join([str(account.account_data) for account in individual.accounts.all()])
        return ""

    def _adjust_column_width_from_col(self, ws: Worksheet, min_row: int, min_col: int, max_col: int) -> None:
        column_widths = []

        for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
            for cell in col:
                value = cell.value
                if value:
                    if isinstance(value, str) is False:
                        value = str(value)

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i in range(len(column_widths)):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            ws.column_dimensions[col_name].width = value
