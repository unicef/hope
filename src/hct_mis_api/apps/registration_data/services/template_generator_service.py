from typing import Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from hct_mis_api.apps.core.field_attributes.core_fields_attributes import FieldFactory
from hct_mis_api.apps.core.field_attributes.fields_types import Scope
from hct_mis_api.apps.core.models import FlexibleAttribute
from hct_mis_api.apps.core.utils import serialize_flex_attributes
from hct_mis_api.apps.geo.models import Area
from hct_mis_api.apps.program.models import Program


class TemplateFileGeneratorService:
    def __init__(self, program: Program):
        self.program = program
        self.business_area = program.business_area
        self.flex_fields = serialize_flex_attributes()
        self.pdu_attributes = FlexibleAttribute.objects.filter(
            type=FlexibleAttribute.PDU, program=program
        ).select_related("pdu_data")
        self.core_fields = FieldFactory.from_scopes(
            [Scope.GLOBAL, Scope.XLSX, Scope.HOUSEHOLD_ID, Scope.COLLECTOR]
        ).apply_business_area(business_area_slug=self.business_area.slug)

    def create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        self.import_helper_ws = wb.active
        self.import_helper_ws.title = "Import helper"
        self.households_ws = wb.create_sheet("Households")
        self.individuals_ws = wb.create_sheet("Individuals")
        self.people_ws = wb.create_sheet("People")
        self.choices_ws = wb.create_sheet("Choices")
        self.wb = wb
        self._add_import_helper()
        self._add_households_columns()
        self._add_individuals_columns()
        self._add_people_columns()
        self._handle_choices({**self.households_fields, **self.individuals_fields})
        return wb

    def _add_households_columns(self) -> None:
        self.households_fields = {
            **self.core_fields.associated_with_household().to_dict_by("xlsx_field"),
            **self.flex_fields["households"],
        }
        households_rows = self._handle_name_and_label_row(self.households_fields)
        self._append_rows(self.households_ws, households_rows)

    def _add_individuals_columns(self) -> None:
        self.individuals_fields = {
            **self.core_fields.associated_with_individual().to_dict_by("xlsx_field"),
            **self.flex_fields["individuals"],
        }
        individuals_rows = self._handle_name_and_label_row(self.individuals_fields)
        pdu_names, pdu_labels = self._get_pdu_columns()
        individuals_rows[0].extend(pdu_names)
        individuals_rows[1].extend(pdu_labels)
        self._append_rows(self.individuals_ws, individuals_rows)

    def _add_people_columns(self) -> None:
        people_fields = {
            **FieldFactory.from_scope(Scope.XLSX_PEOPLE)
            .apply_business_area(business_area_slug=self.business_area.slug)
            .to_dict_by("xlsx_field"),
            **self.flex_fields["individuals"],
        }
        people_rows = self._handle_name_and_label_row(people_fields)
        pdu_names, pdu_labels = self._get_pdu_columns()
        people_rows[0].extend(pdu_names)
        people_rows[1].extend(pdu_labels)
        self._append_rows(self.people_ws, people_rows)

    def _add_import_helper(self) -> None:
        default_helper_text = """
        Sheets and their purposes:
        - Households: Use this sheet to enter details about the households you want to import.
        - Individuals: Use this sheet to enter information about the individuals within the households you want to import.
        - People: Use this sheet to enter data about individuals without households.\n
        Please note that you must decide whether to fill out the data in the Households and Individuals sheets or the People sheet, as these options are mutually exclusive.
        """
        self.import_helper_ws.append([default_helper_text])

    def _get_pdu_columns(self) -> tuple[list[str], list[str]]:
        name_row = []
        label_row = []
        for flexible_attribute in self.pdu_attributes:
            name_row.append(f"{flexible_attribute.name}_round_1_value")
            name_row.append(f"{flexible_attribute.name}_round_1_collection_date")
            label_row.append(
                f"{self._get_label(flexible_attribute)} - First round value - {flexible_attribute.pdu_data.subtype}"
            )
            label_row.append(f"{self._get_label(flexible_attribute)} - First round collection date - DATE")
        return name_row, label_row

    def _get_label(self, flexible_attribute: FlexibleAttribute) -> str:
        if flexible_attribute.label:
            if "English(EN)" in flexible_attribute.label:
                return flexible_attribute.label["English(EN)"]
        return flexible_attribute.name

    def _handle_choices(self, fields: dict) -> None:
        self.choices_ws.append(["Field Name", "Label", "Value to be used in template"])
        for field_name, field_value in fields.items():
            is_admin_level = field_name in ("admin1_h_c", "admin2_h_c")
            choices = field_value["choices"]
            if is_admin_level:
                choices = Area.get_admin_areas_as_choices(field_name[-5])
            if choices:
                for choice in field_value["choices"]:
                    row = [
                        field_name,
                        str(choice["label"]["English(EN)"]),
                        choice["value"],
                    ]
                    self.choices_ws.append(row)

    def _handle_name_and_label_row(self, fields: dict) -> tuple[list[str], list[str]]:
        names: list[str] = []
        labels: list[str] = []

        for field_name, field_value in fields.items():
            names.append(field_name)
            label = (
                f"{field_value['label']['English(EN)']}"
                f" - {field_value['type']}"
                f"{' - required' if field_value['required'] else ''}"
            )
            labels.append(label)

        return names, labels

    def _append_rows(self, ws: Worksheet, rows: Iterable[Iterable[str]]) -> None:
        for row in rows:
            ws.append(row)
