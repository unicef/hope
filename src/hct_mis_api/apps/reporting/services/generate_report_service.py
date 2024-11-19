import copy
import logging
from datetime import datetime, timedelta
from tempfile import NamedTemporaryFile
from typing import TYPE_CHECKING, List, Tuple

from django.conf import settings
from django.contrib.postgres.aggregates.general import ArrayAgg
from django.contrib.postgres.fields import ArrayField
from django.core.files import File
from django.db import models
from django.db.models import (
    Case,
    Count,
    DateTimeField,
    DecimalField,
    F,
    IntegerField,
    Max,
    Min,
    Q,
    QuerySet,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce, Concat, Greatest, Least
from django.template.loader import render_to_string

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hct_mis_api.apps.core.utils import (
    decode_id_string,
    encode_id_base64,
    timezone_datetime,
)
from hct_mis_api.apps.geo.models import Area
from hct_mis_api.apps.grievance.models import GrievanceTicket
from hct_mis_api.apps.household.models import (
    NONE,
    WORK_STATUS_CHOICE,
    Household,
    Individual,
)
from hct_mis_api.apps.payment.delivery_mechanisms import DeliveryMechanismChoices
from hct_mis_api.apps.payment.models import (
    CashPlan,
    PaymentPlan,
    PaymentRecord,
    PaymentVerification,
    PaymentVerificationPlan,
)
from hct_mis_api.apps.reporting.models import Report
from hct_mis_api.apps.utils.mailjet import MailjetClient

if TYPE_CHECKING:
    from hct_mis_api.apps.account.models import User


logger = logging.getLogger(__name__)


class GenerateReportContentHelpers:
    @staticmethod
    def get_individuals(report: Report) -> QuerySet[Individual]:
        filter_vars = {
            "household__business_area": report.business_area,
            # "household__program": report.program # TODO Uncomment after add program to household
            "withdrawn": False,
            "duplicate": False,
            "last_registration_date__gte": report.date_from,
            "last_registration_date__lte": report.date_to,
        }
        if report.admin_area.all().exists():
            filter_vars["household__admin_area__in"] = report.admin_area.all()
        return Individual.objects.filter(**filter_vars)

    @classmethod
    def format_individual_row(self, individual: Individual) -> tuple:
        return (
            individual.household.id,
            individual.household.country_origin.name if individual.household.country_origin else "",
            individual.household.admin_area.name if individual.household.admin_area else "",
            individual.birth_date,
            individual.estimated_birth_date,
            individual.sex,
            individual.marital_status,
            len(individual.observed_disability) >= 1 and NONE not in individual.observed_disability,
            individual.observed_disability,
            individual.comms_disability,
            individual.hearing_disability,
            individual.memory_disability,
            individual.physical_disability,
            individual.seeing_disability,
            individual.selfcare_disability,
            individual.pregnant,
            individual.relationship,
            self._to_values_list(individual.households_and_roles.all(), "role"),
            dict(WORK_STATUS_CHOICE).get(individual.work_status, ""),
            individual.sanction_list_possible_match,
            individual.deduplication_batch_status,
            individual.deduplication_golden_record_status,
            (
                individual.deduplication_golden_record_results.get("duplicates", "")
                if individual.deduplication_golden_record_results
                else ""
            ),
            (
                individual.deduplication_golden_record_results.get("possible_duplicates", "")
                if individual.deduplication_golden_record_results
                else ""
            ),
            self._format_date(individual.first_registration_date),
            self._format_date(individual.last_registration_date),
        )

    @staticmethod
    def get_households(report: Report) -> QuerySet:
        filter_vars = {
            "business_area": report.business_area,
            "withdrawn": False,
            "last_registration_date__gte": timezone_datetime(report.date_from),
            "last_registration_date__lte": timezone_datetime(report.date_to),
        }
        if report.admin_area.all().exists():
            filter_vars["admin_area__in"] = report.admin_area.all()
        if report.program:
            filter_vars["program"] = report.program
        return Household.objects.filter(**filter_vars)

    @classmethod
    def format_household_row(self, household: Household) -> tuple:
        row = [
            household.id,
            household.country_origin.name if household.country_origin else "",
            household.admin_area.name if household.admin_area else "",
            household.size,
            household.geopoint[0] if household.geopoint else "",
            household.geopoint[1] if household.geopoint else "",
            household.residence_status,
            household.returnee,
            household.status,
            household.village,
            household.female_age_group_0_5_count,
            household.female_age_group_0_5_disabled_count,
            household.female_age_group_6_11_count,
            household.female_age_group_6_11_disabled_count,
            household.female_age_group_12_17_count,
            household.female_age_group_12_17_disabled_count,
            household.female_age_group_18_59_count,
            household.female_age_group_18_59_disabled_count,
            household.female_age_group_60_count,
            household.female_age_group_60_disabled_count,
            household.pregnant_count,
            household.male_age_group_0_5_count,
            household.male_age_group_0_5_disabled_count,
            household.male_age_group_6_11_count,
            household.male_age_group_6_11_disabled_count,
            household.male_age_group_12_17_count,
            household.male_age_group_12_17_disabled_count,
            household.male_age_group_18_59_count,
            household.male_age_group_18_59_disabled_count,
            household.male_age_group_60_count,
            household.male_age_group_60_disabled_count,
            self._format_date(household.first_registration_date),
            self._format_date(household.last_registration_date),
            household.org_name_enumerator,
        ]
        for program in household.programs.all():
            row.append(program.name)
        return tuple(row)

    @staticmethod
    def get_cash_plan_verifications(report: Report) -> QuerySet:
        pp_business_area_ids = list(
            CashPlan.objects.filter(business_area=report.business_area).values_list("id", flat=True)
        )
        filter_vars = {
            "payment_plan_object_id__in": pp_business_area_ids,
            "completion_date__isnull": False,
            "completion_date__gte": report.date_from,
            "completion_date__lte": report.date_to,
        }
        if report.program:
            pp_program_ids = list(CashPlan.objects.filter(program=report.program).values_list("id", flat=True))
            filter_vars["payment_plan_object_id__in"] = pp_program_ids
        return PaymentVerificationPlan.objects.filter(**filter_vars)

    @staticmethod
    def _map_admin_area_names_from_ids(admin_areas_ids: list) -> str:
        if not admin_areas_ids:
            return ""
        result = []
        for admin_area_id in admin_areas_ids:
            admin_area_id = decode_id_string(admin_area_id)
            if admin_area := Area.objects.filter(id=admin_area_id).first():
                result.append(admin_area.name)
        return ", ".join(result)

    @classmethod
    def format_cash_plan_verification_row(cls, verification: PaymentVerificationPlan) -> tuple:
        return (
            verification.id,
            verification.payment_plan_obj.get_unicef_id,
            verification.payment_plan_obj.program.name,
            cls._format_date(verification.activation_date),
            verification.status,
            verification.verification_channel,
            cls._format_date(verification.completion_date),
            verification.sample_size,
            verification.responded_count,
            verification.received_count,
            verification.received_with_problems_count,
            verification.not_received_count,
            verification.sampling,
            verification.sex_filter,
            cls._map_admin_area_names_from_ids(verification.excluded_admin_areas_filter),
            verification.age_filter,
        )

    @staticmethod
    def get_payments(report: Report) -> QuerySet:
        filter_vars = {
            "business_area": report.business_area,
            "delivery_date__date__range": (report.date_from, report.date_to),
            "parent__program": report.program,
        }
        if report.admin_area.all().exists():
            filter_vars["household__admin_area__in"] = report.admin_area.all()
        return PaymentRecord.objects.filter(**filter_vars)

    @classmethod
    def format_payment_row(cls, payment: PaymentRecord) -> tuple:
        cash_or_voucher = ""
        if payment.delivery_type:
            if payment.delivery_type in [
                DeliveryMechanismChoices.DELIVERY_TYPE_CASH,
                DeliveryMechanismChoices.DELIVERY_TYPE_DEPOSIT_TO_CARD,
                DeliveryMechanismChoices.DELIVERY_TYPE_TRANSFER,
            ]:
                cash_or_voucher = "cash"
            else:
                # TODO: check if this is even an option, I don't see the delivery_type being anything else but above three options
                # but following the spreadsheet here
                cash_or_voucher = "voucher"

        return (
            payment.ca_id,
            payment.parent.ca_id if payment.parent else "",
            payment.status,
            payment.currency,
            payment.delivered_quantity,
            payment.delivered_quantity_usd or payment.delivered_quantity,
            cls._format_date(payment.delivery_date),
            payment.delivery_type,
            payment.distribution_modality,
            payment.entitlement_quantity,
            payment.target_population.id,
            payment.target_population.name,
            cash_or_voucher,
            payment.household.id,
        )

    @staticmethod
    def get_payment_verifications(report: Report) -> QuerySet:
        pp_business_area_ids = list(
            PaymentPlan.objects.filter(business_area=report.business_area).values_list("id", flat=True)
        )
        filter_vars = {
            "payment_verification_plan__payment_plan_object_id__in": pp_business_area_ids,
            "payment_verification_plan__completion_date__isnull": False,
            "payment_verification_plan__completion_date__date__range": (report.date_from, report.date_to),
        }
        if report.program:
            pp_program_ids = list(PaymentPlan.objects.filter(program=report.program).values_list("id", flat=True))
            filter_vars["payment_verification_plan__payment_plan_object_id__in"] = pp_program_ids
        return PaymentVerification.objects.filter(**filter_vars)

    @classmethod
    def format_payment_verification_row(cls, payment_verification: PaymentVerification) -> tuple:
        return (
            payment_verification.payment_verification_plan.id,
            payment_verification.payment_obj.unicef_id,
            payment_verification.payment_verification_plan.get_payment_plan.get_unicef_id,
            cls._format_date(payment_verification.payment_verification_plan.completion_date),
            payment_verification.received_amount,
            payment_verification.status,
            payment_verification.status_date,
        )

    @staticmethod
    def get_payment_plans(report: Report) -> QuerySet[PaymentPlan]:
        filter_vars = {
            "business_area": report.business_area,
            "start_date__gte": report.date_from,
            "end_date__lte": report.date_to,
        }
        if report.program:
            filter_vars["program_cycle__program"] = report.program
        return PaymentPlan.objects.filter(**filter_vars)

    @classmethod
    def format_payment_plan_row(cls, payment_plan: PaymentPlan) -> tuple:
        return (
            payment_plan.get_unicef_id,
            payment_plan.get_status_display(),
            payment_plan.total_households_count,
            payment_plan.get_currency_display(),
            payment_plan.total_entitled_quantity,
            payment_plan.total_delivered_quantity,
            payment_plan.total_undelivered_quantity,
            cls._format_date(payment_plan.dispersion_start_date),
            cls._format_date(payment_plan.dispersion_end_date),
        )

    @staticmethod
    def get_cash_plans(report: Report) -> QuerySet:
        filter_vars = {
            "business_area": report.business_area,
            "end_date__gte": report.date_from,
            "end_date__lte": report.date_to,
        }
        if report.program:
            filter_vars["program"] = report.program
        return CashPlan.objects.filter(**filter_vars)

    @classmethod
    def format_cash_plan_row(cls, cash_plan: CashPlan) -> tuple:
        return (
            cash_plan.ca_id,
            cash_plan.name,
            cls._format_date(cash_plan.start_date),
            cls._format_date(cash_plan.end_date),
            cash_plan.program.name,
            cash_plan.funds_commitment,
            cash_plan.assistance_measurement,
            cash_plan.assistance_through,
            cash_plan.delivery_type,
            cls._format_date(cash_plan.dispersion_date),
            cash_plan.down_payment,
            cash_plan.total_delivered_quantity,
            cash_plan.total_undelivered_quantity,
            cash_plan.total_entitled_quantity,
            cash_plan.total_entitled_quantity_revised,
            cash_plan.total_persons_covered,
            cash_plan.total_persons_covered_revised,
            cash_plan.status,
            cls._format_date(cash_plan.status_date),
            cash_plan.vision_id,
            cash_plan.validation_alerts_count,
            # cash_plan.verification_status,
        )

    @staticmethod
    def get_payments_for_individuals(report: Report) -> QuerySet:
        if isinstance(report.date_to, str):
            report.date_to = datetime.strptime(report.date_to, "%Y-%m-%d").date()

        date_to_time = datetime.fromordinal(report.date_to.toordinal())
        date_to_time += timedelta(days=1)
        filter_q = Q(
            #  Q(household__program=report.program) & # TODO Uncomment after add program to household
            Q(
                Q(household__paymentrecord__business_area=report.business_area)
                | Q(household__payment__business_area=report.business_area)
            )
            & Q(
                Q(household__paymentrecord__delivery_date__gte=report.date_from)
                | Q(household__payment__delivery_date__gte=report.date_from)
            )
            & Q(
                Q(household__paymentrecord__delivery_date__lt=date_to_time)
                | Q(household__payment__delivery_date__lt=date_to_time)
            )
        )
        if report.admin_area.all().exists():
            filter_q &= Q(household__admin_area__in=report.admin_area.all())
        if report.program:
            filter_q &= Q(
                Q(household__paymentrecord__parent__program=report.program)
                | Q(household__payment__parent__program=report.program)
            )

        return (
            Individual.objects.filter(filter_q)
            .annotate(first_delivery_date_paymentrecord=Min("household__paymentrecord__delivery_date"))
            .annotate(first_delivery_date_payment=Min("household__payment__delivery_date"))
            .annotate(
                first_delivery_date=Least(
                    F("first_delivery_date_paymentrecord"),
                    F("first_delivery_date_payment"),
                    output_field=DateTimeField(),
                )
            )
            .annotate(last_delivery_date_paymentrecord=Max("household__paymentrecord__delivery_date"))
            .annotate(last_delivery_date_payment=Max("household__payment__delivery_date"))
            .annotate(
                last_delivery_date=Greatest(
                    F("last_delivery_date_paymentrecord"),
                    F("last_delivery_date_payment"),
                    output_field=DateTimeField(),
                )
            )
            .annotate(
                payments_made=Count(
                    Case(
                        When(
                            Q(household__paymentrecord__delivered_quantity__gte=0),
                            then=F("household__paymentrecord__id"),
                        ),
                        When(
                            Q(household__payment__delivered_quantity__gte=0),
                            then=F("household__payment__id"),
                        ),
                        output_field=IntegerField(),
                    )
                )
            )
            .annotate(
                payment_currency=ArrayAgg(
                    Concat(
                        "household__paymentrecord__currency",
                        Value(" "),
                        "household__payment__currency",
                        output_field=ArrayField(models.CharField()),
                    )
                )
            )
            .annotate(
                total_delivered_quantity_local=Coalesce(
                    Sum("household__paymentrecord__delivered_quantity"), Value(0), output_field=DecimalField()
                )
                + Coalesce(Sum("household__payment__delivered_quantity"), Value(0), output_field=DecimalField())
            )
            .annotate(
                total_delivered_quantity_usd=Coalesce(
                    Sum("household__paymentrecord__delivered_quantity_usd"), Value(0), output_field=DecimalField()
                )
                + Coalesce(Sum("household__payment__delivered_quantity_usd"), Value(0), output_field=DecimalField())
            )
            .order_by("household__id")
            .distinct()
        )

    @classmethod
    def format_payments_for_individuals_row(cls, individual: Individual) -> tuple:
        return (
            individual.household.unicef_id,
            individual.household.country_origin.name if individual.household.country_origin else "",
            individual.household.admin1.name if individual.household.admin1 else "",
            individual.household.admin1.p_code if individual.household.admin1 else "",
            individual.household.admin2.name if individual.household.admin2 else "",
            individual.household.admin2.p_code if individual.household.admin2 else "",
            individual.household.admin3.name if individual.household.admin3 else "",
            individual.household.admin3.p_code if individual.household.admin3 else "",
            cls._format_date(individual.first_delivery_date),
            cls._format_date(individual.last_delivery_date),
            individual.payments_made,
            " ".join(individual.payment_currency).strip().replace("  ", " ").replace(" ", ", "),
            individual.total_delivered_quantity_local,
            individual.total_delivered_quantity_usd or individual.total_delivered_quantity_local,
            individual.birth_date,
            individual.estimated_birth_date,
            individual.sex,
            individual.marital_status,
            len(individual.observed_disability) >= 1 and NONE not in individual.observed_disability,
            individual.observed_disability,
            individual.comms_disability,
            individual.hearing_disability,
            individual.memory_disability,
            individual.physical_disability,
            individual.seeing_disability,
            individual.selfcare_disability,
            individual.pregnant,
            individual.relationship,
            cls._to_values_list(individual.households_and_roles.all(), "role"),
            dict(WORK_STATUS_CHOICE).get(individual.work_status, ""),
            individual.sanction_list_possible_match,
            individual.deduplication_batch_status,
            individual.deduplication_golden_record_status,
            (
                individual.deduplication_golden_record_results.get("duplicates", "")
                if individual.deduplication_golden_record_results
                else ""
            ),
            (
                individual.deduplication_golden_record_results.get("possible_duplicates", "")
                if individual.deduplication_golden_record_results
                else ""
            ),
        )

    @staticmethod
    def get_grievance_tickets(report: Report) -> QuerySet:
        # TODO filter tickets based on program
        filter_vars = {
            "business_area": report.business_area,
            "created_at__gte": report.date_from,
            "created_at__lte": report.date_to,
        }

        return GrievanceTicket.objects.filter(**filter_vars).select_related("admin2", "created_by", "assigned_to")

    @classmethod
    def format_grievance_tickets_row(cls, grievance_ticket: GrievanceTicket) -> tuple:
        def get_full_name(user: "User") -> str:
            if not user:
                return ""
            return " ".join(filter(None, [user.first_name, user.last_name]))

        def get_username(user: "User") -> str:
            if not user:
                return ""
            return user.username

        return (
            grievance_ticket.unicef_id,
            grievance_ticket.created_at,
            grievance_ticket.updated_at,
            grievance_ticket.get_status_display(),
            grievance_ticket.get_category_display(),
            grievance_ticket.get_issue_type(),
            getattr(grievance_ticket.admin2, "name", ""),
            getattr(grievance_ticket.admin2, "p_code", ""),
            get_username(grievance_ticket.created_by),
            get_full_name(grievance_ticket.created_by),
            get_username(grievance_ticket.assigned_to),
            get_full_name(grievance_ticket.assigned_to),
        )

    @staticmethod
    def _to_values_list(instances: List, field_name: str) -> str:
        values_list = list(instances.values_list(field_name, flat=True))
        return ", ".join([str(value) for value in values_list])

    @staticmethod
    def _format_date(date: datetime) -> str:
        if not date:
            return ""
        return date.strftime("%Y-%m-%d")


class GenerateReportService:
    HEADERS = {
        Report.INDIVIDUALS: (
            "household id",  # 8e8ea94a-2ca5-4b76-b055-e098bc24eee8
            "country of origin",  # South Sudan
            "administrative area 2",  # Juba
            "birth date",  # 2000-06-24
            "estimated birth date",  # TRUE
            "gender",  # FEMALE,
            "marital status",  # MARRIED
            "disability",  # TRUE
            "observed disability",
            "communication disability",
            "hearing disability",  # LOT_DIFFICULTY
            "remembering disability",
            "physical disability",
            "seeing disability",
            "self-care disability",
            "pregnant",  # TRUE
            "relationship to hoh",  # WIFE
            "role",  # PRIMARY
            "work status",  # NOT_PROVIDED
            "sanction list possible match",  # FALSE
            "dedupe in batch status",  # UNIQUE_IN_BATCH
            "dedupe in Pop. status",  # DUPLICATE
            "dedupe in Pop.duplicates",
            "dedupe in Pop. possible duplicates",
            "first registration date",  # 2000-06-24
            "last registration date",  # 2000-06-24
        ),
        Report.HOUSEHOLD_DEMOGRAPHICS: (
            "household id",
            "country of origin",  # South Sudan
            "administrative area 2",  # Juba
            "household size",  # 4
            "latitude",  # 54,367759
            "longitude",  # 60,964675
            "residence status",  # HOST
            "returnee",  # FALSE
            "status",  # ACTIVE
            "village",  # Mendika
            "females 0-5",  # 0
            "females 0-5 w/ disability",  # 0
            "females 6-11",  # 1
            "females 6-11 w/ disability",  # 1
            "females 12-17",  # 1
            "females 12-17 w/ disability",  # 0
            "females 18-59",  # 1
            "females 18-59 w/ disability",  # 0
            "females 60+",  # 0
            "females 60+ w/ disability",  # 0
            "pregnant females",  # 0
            "males 0-5",  # 0
            "males 0-5 w/ disability",  # 0
            "males 6-11",  # 1
            "males 6-11 w/ disability",  # 1
            "males 12-17",  # 1
            "males 12-17 w/ disability",  # 0
            "males 18-59",  # 1
            "males 18-59 w/ disability",  # 0
            "males 60+",  # 0
            "males 60+ w/ disability",  # 0
            "first registration date",  # 2020-08-25
            "last registration date",  # 2020-08-25
            "organization name enumerator",
        ),
        Report.CASH_PLAN_VERIFICATION: (
            "cash plan verification ID",
            "cash plan ID",  # ANT-21-CSH-00001
            "programme",  # Winterization 2020
            "activation date",
            "status",
            "verification method",
            "completion date",
            "sample size",  # 500
            "responded",  # 340
            "received",  # 320
            "received with issues",  # 12
            "not received",  # 8
            "sampling",  # FULL_LIST or RANDOM
            "gender filter",  # FEMALE
            "excluded admin areas",  # Juba
            "age filter",  # {'max': 100, 'min': 0}
        ),
        Report.PAYMENTS: (
            "payment record ID",  # ANT-21-CSH-00001-0000002
            "cash plan ID",  # ANT-21-CSH-00001
            "status",  # Transaction successful
            "currency",
            "delivered quantity (local)",  # 999,00
            "delivered quantity (USD)",  # 235,99
            "delivery date",  # 2020-11-02 07:50:18+00
            "delivery type",  # deposit to card
            "distribution modality",  # 10K AFN per hh
            "entitlement quantity",  # 1000,00
            "TP ID",
            "TP name",
            "cash or voucher",  # if voucher or e-voucher -> voucher, else -> cash,
            "household id",  # 145aacc4-160a-493e-9d36-4f7f981284c7
        ),
        Report.PAYMENT_VERIFICATION: (
            "plan verification ID",
            "payment record ID",  # ANT-21-CSH-00001-0000002
            "plan ID",  # ANT-21-CSH-00001
            "verification completion date",
            "received amount",  # 30,00
            "status",  # RECEIVED_WITH_ISSUES
            "status date",
        ),
        Report.PAYMENT_PLAN: (
            "payment plan ID",
            "status",
            "no. of households",
            "currency",
            "total entitled quantity",
            "total delivered quantity",
            "total undelivered quantity",
            "dispersion start date",
            "dispersion end date",
        ),
        Report.CASH_PLAN: (
            "cash plan ID",  # ANT-21-CSH-00001
            "cash plan name",
            "start date",
            "end date",
            "programme",
            "funds commitment",  # 234567
            "assistance measurement",  # Euro
            "assistance through",  # Cairo Amman Bank
            "delivery type",  # DEPOSIT_TO_CARD
            "dispersion date",
            "down payment",
            "total delivered quantity",  # 220,00
            "total undelivered quantity",  # 10,00
            "total entitled quantity",  # 230,00
            "total entitled quantity revised",  # 230,00
            "total persons covered",  # 12
            "total persons covered revised",  # 12
            "status",  # DISTRIBUTION_COMPLETED_WITH_ERRORS
            "status date",
            "VISION ID",  # 2345253423
            "validation alerts count",  # 2
            # "cash plan verification status",  # FINISHED
        ),
        Report.INDIVIDUALS_AND_PAYMENT: (
            "household id",
            "country of origin",
            "admin1",
            "admin1 P code",
            "admin2",
            "admin2 P code",
            "admin3",
            "admin3 P code",
            "first delivery date",
            "last delivery date",
            "payments made",
            "currency",
            "total delivered quantity (local)",
            "total delivered quantity (USD)",
            "birth date",  # 2000-06-24
            "estimated birth date",  # TRUE
            "gender",  # FEMALE,
            "marital status",  # MARRIED
            "disability",  # TRUE
            "observed disability",
            "communication disability",
            "hearing disability",  # LOT_DIFFICULTY
            "remembering disability",
            "physical disability",
            "seeing disability",
            "self-care disability",
            "pregnant",  # TRUE
            "relationship to hoh",  # WIFE
            "role",  # PRIMARY
            "work status",  # NOT_PROVIDED
            "sanction list possible match",  # FALSE
            "dedupe in batch status",  # UNIQUE_IN_BATCH
            "dedupe in Pop. status",  # DUPLICATE
            "dedupe in Pop.duplicates",
            "dedupe in Pop. possible duplicates",
        ),
        Report.GRIEVANCES: (
            "ticket number",
            "date created",
            "date updated",
            "status",
            "category",
            "issue type",
            "district",
            "admin2 p-code",
            "created by (username)",
            "created by (name)",
            "assigned to (username)",
            "assigned to (name)",
        ),
    }
    OPTIONAL_HEADERS = {Report.HOUSEHOLD_DEMOGRAPHICS: "programme enrolled"}
    TIMEFRAME_CELL_LABELS = {
        Report.INDIVIDUALS: ("Last Registration Date From", "Last Registration Date To"),
        Report.HOUSEHOLD_DEMOGRAPHICS: ("Last Registration Date From", "Last Registration Date To"),
        Report.CASH_PLAN_VERIFICATION: ("Completion Date From", "Completion Date To"),
        Report.PAYMENT_VERIFICATION: ("Completion Date From", "Completion Date To"),
        Report.PAYMENTS: ("Delivery Date From", "Delivery Date To"),
        Report.PAYMENT_PLAN: ("Dispersion Start Date", "Dispersion End Date"),
        Report.INDIVIDUALS_AND_PAYMENT: ("Delivery Date From", "Delivery Date To"),
        Report.CASH_PLAN: ("End Date From", "End Date To"),
        Report.GRIEVANCES: ("End Date From", "End Date To"),
    }
    ROW_CONTENT_METHODS = {
        Report.INDIVIDUALS: (
            GenerateReportContentHelpers.get_individuals,
            GenerateReportContentHelpers.format_individual_row,
        ),
        Report.HOUSEHOLD_DEMOGRAPHICS: (
            GenerateReportContentHelpers.get_households,
            GenerateReportContentHelpers.format_household_row,
        ),
        Report.CASH_PLAN_VERIFICATION: (
            GenerateReportContentHelpers.get_cash_plan_verifications,
            GenerateReportContentHelpers.format_cash_plan_verification_row,
        ),
        Report.PAYMENTS: (GenerateReportContentHelpers.get_payments, GenerateReportContentHelpers.format_payment_row),
        Report.PAYMENT_VERIFICATION: (
            GenerateReportContentHelpers.get_payment_verifications,
            GenerateReportContentHelpers.format_payment_verification_row,
        ),
        Report.PAYMENT_PLAN: (
            GenerateReportContentHelpers.get_payment_plans,
            GenerateReportContentHelpers.format_payment_plan_row,
        ),
        Report.CASH_PLAN: (
            GenerateReportContentHelpers.get_cash_plans,
            GenerateReportContentHelpers.format_cash_plan_row,
        ),
        Report.INDIVIDUALS_AND_PAYMENT: (
            GenerateReportContentHelpers.get_payments_for_individuals,
            GenerateReportContentHelpers.format_payments_for_individuals_row,
        ),
        Report.GRIEVANCES: (
            GenerateReportContentHelpers.get_grievance_tickets,
            GenerateReportContentHelpers.format_grievance_tickets_row,
        ),
    }
    FILTERS_SHEET = "Meta"
    MAX_COL_WIDTH = 75

    def __init__(self, report: Report) -> None:
        self.report = report
        self.business_area = report.business_area

    def _create_workbook(self) -> None:
        self.wb = openpyxl.Workbook()
        ws_report = self.wb.active
        ws_report.title = f"{self.report.get_report_type_display()} Report"
        self.ws_report = ws_report
        self.ws_filters = self.wb.create_sheet(GenerateReportService.FILTERS_SHEET)

    def _add_filters_info(self) -> None:
        filter_rows = [
            ("Report type", str(self.report.get_report_type_display())),
            ("Business area", self.business_area.name),
            (GenerateReportService.TIMEFRAME_CELL_LABELS[self.report.report_type][0], str(self.report.date_from)),
            (GenerateReportService.TIMEFRAME_CELL_LABELS[self.report.report_type][0], str(self.report.date_to)),
        ]

        if self.report.admin_area.all().exists():
            filter_rows.append(
                (
                    "Administrative area 2",
                    GenerateReportContentHelpers._to_values_list(self.report.admin_area.all(), "name"),
                )
            )
        if self.report.program:
            filter_rows.append(("Program", self.report.program.name))

        for filter_row in filter_rows:
            self.ws_filters.append(filter_row)

    def _add_headers(self) -> None:
        headers_row = GenerateReportService.HEADERS[self.report.report_type]
        self.ws_report.append(headers_row)

    def _add_rows(self) -> int:
        get_row_methods: Tuple = GenerateReportService.ROW_CONTENT_METHODS[self.report.report_type]
        all_instances = get_row_methods[0](self.report)
        self.report.number_of_records = all_instances.count()
        number_of_columns_based_on_set_headers = len(GenerateReportService.HEADERS[self.report.report_type])
        col_instances_len = 0
        for instance in all_instances:
            row = get_row_methods[1](instance)
            str_row = self._stringify_all_values(row)
            if len(str_row) > col_instances_len:
                col_instances_len = len(str_row)
            self.ws_report.append(str_row)
        if col_instances_len > number_of_columns_based_on_set_headers:
            # to cover bases when we create extra columns for reverse foreign key instances and we don't know in advance how many columns there will be
            self._add_missing_headers(
                self.ws_report,
                number_of_columns_based_on_set_headers + 1,
                col_instances_len,
                GenerateReportService.OPTIONAL_HEADERS.get(self.report.report_type, ""),
            )
        return col_instances_len

    def generate_workbook(self) -> openpyxl.Workbook:
        self._create_workbook()
        self._add_filters_info()
        self._add_headers()
        number_of_columns = self._add_rows()
        self._adjust_column_width_from_col(self.ws_filters, 1, 2, 1)
        self._adjust_column_width_from_col(self.ws_report, 1, number_of_columns, 0)
        return self.wb

    def save_wb_file_in_db(self) -> None:
        with NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            self.report.file.save(
                f"{self.report.get_report_type_display()}-{GenerateReportContentHelpers._format_date(self.report.created_at)}.xlsx",
                File(tmp),
                save=False,
            )

    def generate_report(self) -> None:
        try:
            self.generate_workbook()
            self.save_wb_file_in_db()
            self.report.status = Report.COMPLETED
        except Exception as e:
            logger.exception(e)
            self.report.status = Report.FAILED
        self.report.save()

        if self.report.file and self.report.business_area.enable_email_notification:
            self._send_email()

    def _send_email(self) -> None:
        context = {
            "report_type": self.report.get_report_type_display(),
            "created_at": GenerateReportContentHelpers._format_date(self.report.created_at),
            "report_url": f'https://{settings.FRONTEND_HOST}/{self.business_area.slug}/programs/all/reporting/{encode_id_base64(self.report.id, "Report")}',
            "title": "Report",
        }
        text_body = render_to_string("report.txt", context=context)
        html_body = render_to_string("report.html", context=context)
        email = MailjetClient(
            subject="HOPE report generated",
            recipients=[self.report.created_by.email],
            html_body=html_body,
            text_body=text_body,
        )
        email.send_email()

    def _add_missing_headers(self, ws: Worksheet, column_to_start: int, column_to_finish: int, label: str) -> None:
        for x in range(column_to_start, column_to_finish + 1):
            col_letter = get_column_letter(x)
            ws[f"{col_letter}1"] = label

    def _adjust_column_width_from_col(self, ws: Worksheet, min_col: int, max_col: int, min_row: int) -> None:
        column_widths = []

        for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
            for cell in col:
                value = cell.value

                if value is not None:
                    if isinstance(value, str) is False:
                        value = str(value)

                    if len(value) > GenerateReportService.MAX_COL_WIDTH:
                        alignment = copy.copy(cell.alignment)
                        alignment.wrapText = True
                        cell.alignment = alignment

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i in range(len(column_widths)):
            col_name = get_column_letter(min_col + i)
            ws.column_dimensions[col_name].width = min(column_widths[i] + 2, GenerateReportService.MAX_COL_WIDTH)

    def _stringify_all_values(self, row: tuple) -> tuple:
        str_row = []
        for value in row:
            str_row.append(str(value if value is not None else ""))
        return tuple(str_row)
