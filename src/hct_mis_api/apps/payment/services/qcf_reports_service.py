import csv
import dataclasses
import io
import logging
import zipfile
from collections import defaultdict
from datetime import datetime
from tempfile import NamedTemporaryFile
from typing import IO, List, Tuple

from django.contrib.admin.options import get_content_type_for_model
from django.core.files import File
from django.core.files.base import ContentFile
from django.db import transaction
from django.template.loader import render_to_string

import openpyxl
from openpyxl.styles import Font

from hct_mis_api.apps.account.models import User
from hct_mis_api.apps.account.permissions import Permissions
from hct_mis_api.apps.core.models import FileTemp
from hct_mis_api.apps.payment.celery_tasks import send_qcf_report_email_notifications
from hct_mis_api.apps.payment.models import Payment, PaymentPlan
from hct_mis_api.apps.payment.models.payment import (
    WesternUnionInvoice,
    WesternUnionInvoicePayment,
    WesternUnionPaymentPlanReport,
)
from hct_mis_api.apps.payment.services.western_union_ftp import WesternUnionFTPClient

logger = logging.getLogger(__name__)


@dataclasses.dataclass
class QCFReportPaymentRowData:
    payment_unicef_id: str
    mtcn: str
    mtcns_match: bool
    hope_mtcn: str
    payment_plan_unicef_id: str
    programme: str
    fc: str
    principal_amount: float
    charges_amount: float
    fee_amount: float


@dataclasses.dataclass
class QCFReportPaymentPlanData:
    payment_plan_unicef_id: str
    principal_total: float = dataclasses.field(init=False)
    charges_total: float = dataclasses.field(init=False)
    refunds_total: float = dataclasses.field(init=False)
    fees_total: float = dataclasses.field(init=False)
    payments_data: List[QCFReportPaymentRowData]

    def __post_init__(self) -> None:
        self.principal_total = sum(p.principal_amount for p in self.payments_data if p.principal_amount >= 0) or 0
        self.charges_total = sum(p.charges_amount for p in self.payments_data) or 0
        self.refunds_total = sum(p.principal_amount for p in self.payments_data if p.principal_amount < 0) or 0
        self.fees_total = sum(p.fee_amount for p in self.payments_data) or 0


class QCFReportsService:
    class QCFReportsServiceException(Exception):
        pass

    def process_zip_file(self, file_content: io.BytesIO, wu_qcf_file: WesternUnionInvoice) -> None:
        with zipfile.ZipFile(file_content) as z:
            for zip_info in z.infolist():
                if zip_info.filename.endswith(".cf") or zip_info.filename.endswith(".CF"):
                    with z.open(zip_info) as txt_file:
                        self.process_qcf_report_file(zip_info.filename, txt_file, wu_qcf_file)

    def process_files_since(self, date_from: datetime) -> None:
        ftp_client = WesternUnionFTPClient()
        files: List[Tuple[str, io.BytesIO]] = ftp_client.get_files_since(date_from)
        for filename, file_like in files:
            if not WesternUnionInvoice.objects.filter(name=filename).exists():
                wu_qcf_file = self.store_qcf_file(filename, file_like)
                file_like.seek(0)  # rewind
                self.process_zip_file(file_like, wu_qcf_file)

    def test_process_file(self, filename: str, file_like: io.BytesIO) -> None:
        # TODO remove after testing
        wu_qcf_file = self.store_qcf_file(filename, file_like)
        self.process_zip_file(file_like, wu_qcf_file)

    def store_qcf_file(self, filename: str, file_like: io.BytesIO) -> WesternUnionInvoice:
        content_file = ContentFile(file_like.read(), name=filename)
        wu_qcf_file, created = WesternUnionInvoice.objects.get_or_create(
            name=filename,
        )
        if created:
            file_temp = FileTemp.objects.create(
                object_id=wu_qcf_file.pk,
                content_type=get_content_type_for_model(wu_qcf_file),
                file=content_file,
            )
            file_temp.file.save(filename, content_file)
            wu_qcf_file.file = file_temp
            wu_qcf_file.save()

        return wu_qcf_file

    def link_payments_with_invoice(
        self, wu_qcf_file: WesternUnionInvoice, payment_raw_data: List[Tuple[Payment, list]]
    ) -> None:
        objs = []
        for payment, fields in payment_raw_data:
            transaction_status = fields[18]
            objs.append(
                WesternUnionInvoicePayment(
                    western_union_invoice=wu_qcf_file,
                    payment=payment,
                    transaction_status=transaction_status,
                )
            )
        WesternUnionInvoicePayment.objects.bulk_create(objs)

    def process_qcf_report_file(self, filename: str, txt_file: IO[bytes], wu_qcf_file: WesternUnionInvoice) -> None:
        file_content = txt_file.read()
        payment_plan_id_payments_lines_map = defaultdict(list)

        try:
            decoded_content = file_content.decode("utf-8")
        except UnicodeDecodeError:
            raise self.QCFReportsServiceException(f"Could not decode QCF report file {filename}")

        lines = list(csv.reader(io.StringIO(decoded_content)))
        for fields in lines[1:-1]:
            payment_unicef_id = f"RC{fields[9].strip('\"')}"

            try:
                payment = Payment.objects.get(unicef_id=payment_unicef_id)
                payment_plan_id_payments_lines_map[payment.parent_id].append((payment, fields))
            except Payment.DoesNotExist:
                logger.error(f"{wu_qcf_file}: File:{filename}, Payment {payment_unicef_id} does not exist")
                continue

        for payment_plan_id, payment_raw_data in payment_plan_id_payments_lines_map.items():
            payment_plan = PaymentPlan.objects.get(id=payment_plan_id)

            with transaction.atomic():
                self.link_payments_with_invoice(wu_qcf_file, payment_raw_data)
                report_name, report = self.generate_report(payment_plan, payment_raw_data)

                wu_qcf_report = WesternUnionPaymentPlanReport.objects.create(
                    qcf_file=wu_qcf_file,
                    payment_plan=payment_plan,
                )

                with NamedTemporaryFile() as tmp:
                    file_temp = FileTemp.objects.create(
                        object_id=str(wu_qcf_report.pk),
                        content_type=get_content_type_for_model(wu_qcf_report),
                    )
                    report.save(tmp.name)
                    tmp.seek(0)
                    file_temp.file.save(report_name, File(tmp))
                    wu_qcf_report.report_file = file_temp
                    wu_qcf_report.save()

                transaction.on_commit(
                    lambda wu_qcf_report_id=wu_qcf_report.id: send_qcf_report_email_notifications.delay(
                        wu_qcf_report_id
                    )
                )

    def generate_report(
        self, payment_plan: PaymentPlan, payment_raw_data: List[Tuple[Payment, list]]
    ) -> Tuple[str, openpyxl.Workbook]:
        no_of_qcf_reports_for_pp = WesternUnionPaymentPlanReport.objects.filter(payment_plan=payment_plan).count()
        report_filename = (
            f"QCF_{payment_plan.business_area.slug}_{payment_plan.unicef_id}_{no_of_qcf_reports_for_pp + 1}.xlsx"
        )

        funds_commitments_str = ", ".join(
            f"{number}/{item}"
            for number, item in payment_plan.funds_commitments.all().values_list(
                "funds_commitment_group__funds_commitment_number", "funds_commitment_item"
            )
        )
        payments_data: List[QCFReportPaymentRowData] = []
        for payment, fields in payment_raw_data:
            report_mtcn = fields[1]
            mtcn_match: bool = str(report_mtcn) == str(payment.fsp_auth_code)

            payments_data.append(
                QCFReportPaymentRowData(
                    payment_unicef_id=payment.unicef_id,
                    mtcn=fields[1],
                    mtcns_match=mtcn_match,
                    hope_mtcn=payment.fsp_auth_code,
                    payment_plan_unicef_id=payment_plan.unicef_id,
                    programme=payment_plan.program_cycle.program.name,
                    fc=funds_commitments_str,
                    principal_amount=float(fields[10]),
                    charges_amount=float(fields[11]),
                    fee_amount=float(fields[12]),
                )
            )

        report_data = QCFReportPaymentPlanData(
            payment_plan_unicef_id=payment_plan.unicef_id, payments_data=payments_data
        )

        return report_filename, self.generate_report_xlsx(report_data)

    def generate_report_xlsx(self, report_data: QCFReportPaymentPlanData) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"QCF Report - {report_data.payment_plan_unicef_id}"

        headers = [
            "Payment ID",
            "MTCN",
            "MTCNs match",
            "HOPE MTCN",
            "Payment Plan ID",
            "Programme",
            "FC",
            "Principal Amount",
            "Charges Amount",
            "Fee Amount",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for p in report_data.payments_data:
            ws.append(
                [
                    p.payment_unicef_id,
                    p.mtcn,
                    p.mtcns_match,
                    p.hope_mtcn,
                    p.payment_plan_unicef_id,
                    p.programme,
                    p.fc,
                    p.principal_amount,
                    p.charges_amount,
                    p.fee_amount,
                ]
            )

        # empty row before totals
        ws.append([])

        ws.append(["Principal Total", "", "", "", report_data.principal_total])
        ws.append(["Charges Total", "", "", "", report_data.charges_total])
        ws.append(["Refunds Total", "", "", "", report_data.refunds_total])

        last_row = ws.max_row
        for row in range(last_row - 3, last_row + 1):
            ws[f"A{row}"].font = Font(bold=True)

        return wb

    def send_notification_emails(self, report: WesternUnionPaymentPlanReport) -> None:
        """
        # TODO refactor to 'dev' new perms
        role_assignments = RoleAssignment.objects.filter(
            role__permissions__contains=[Permissions.RECEIVE_PARSED_WU_QCF.name],
            business_area=business_area,
        ).exclude(expiry_date__lt=timezone.now())
        users = User.objects.filter(
            Q(role_assignments__in=role_assignments) |
            Q(partner__role_assignments__in=role_assignments)
        ).distinct()
        """
        business_area = report.payment_plan.business_area
        users = [
            user
            for user in User.objects.all()
            if user.has_permission(Permissions.RECEIVE_PARSED_WU_QCF.name, business_area)
        ]
        print(users)
        for user in users:
            self.send_report_email_to_user(user, report)

    def send_report_email_to_user(self, user: User, report: WesternUnionPaymentPlanReport) -> None:
        text_template = "payment/qcf_report_email.txt"
        html_template = "payment/qcf_report_email.html"

        context = {
            "first_name": getattr(user, "first_name", ""),
            "last_name": getattr(user, "last_name", ""),
            "email": getattr(user, "email", ""),
            "message": f"Payment Plan {report.payment_plan.unicef_id} Payment List export file's Passwords.",
            "title": f"Payment Plan {report.report_file.file.name} Western Union QCF Report",
            "link": f"Western Union QCF Report file: {report.report_file.file.url}",
        }

        user.email_user(
            subject=context["title"],
            html_body=render_to_string(html_template, context=context),
            text_body=render_to_string(text_template, context=context),
        )
