from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hct_mis_api.apps.core.models import BusinessArea
from hct_mis_api.apps.geo.models import Country
from hct_mis_api.apps.household.models import Document, DocumentType, Individual
from hct_mis_api.apps.program.models import Program
from hct_mis_api.apps.registration_datahub.tasks.deduplicate import (
    DeduplicateTask,
    HardDocumentDeduplication,
)


class UniversalIndividualUpdateScript:
    def __init__(
        self,
        business_area: BusinessArea,
        program: Program,
        file_path: str,
        household_fields: Dict[str, Tuple[str, Any, Any]],
        individual_fields: Dict[str, Tuple[str, Any, Any]],
        individual_flex_fields: Dict[str, Tuple[str, Any, Any]],
        document_fields: List[Tuple[str, str]],
        ignore_empty_values: bool = True,
        deduplicate_es: bool = True,
        deduplicate_documents: bool = True,
    ) -> None:
        self.business_area = business_area
        self.program = program
        self.file_path = file_path
        self.household_fields = household_fields
        self.individual_fields = individual_fields
        self.individual_flex_fields = individual_flex_fields
        self.document_fields = document_fields
        self.ignore_empty_values = ignore_empty_values
        self.deduplicate_es = deduplicate_es
        self.deduplicate_documents = deduplicate_documents
        document_types = DocumentType.objects.filter()
        self.document_types = {f"{document_type.key}_no_i_c": document_type for document_type in document_types}

    def validate_household_fields(
        self, row: Tuple[Any, ...], headers: List[str], household: Any, row_index: int
    ) -> List[str]:
        errors = []
        for field, (name, validator, _handler) in self.household_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, row, self.business_area, self.program)
            if error:
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_individual_fields(
        self, row: Tuple[Any, ...], headers: List[str], individual: Individual, row_index: int
    ) -> List[str]:
        errors = []
        for field, (name, validator, _handler) in self.individual_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, row, self.business_area, self.program)
            if error:  # pragma: no cover
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_individual_flex_fields(
        self, row: Tuple[Any, ...], headers: List[str], individual: Individual, row_index: int
    ) -> List[str]:
        errors = []
        for field, (name, validator, _handler) in self.individual_flex_fields.items():
            value = row[headers.index(field)]
            error = validator(value, name, row, self.business_area, self.program)
            if error:  # pragma: no cover
                errors.append(f"Row: {row_index} - {error}")
        return errors

    def validate_documents(
        self, row: Tuple[Any, ...], headers: List[str], individual: Individual, row_index: int
    ) -> List[str]:
        errors = []
        for number_column_name, country_column_name in self.document_fields:
            document_type = self.document_types.get(number_column_name)
            country_text = row[headers.index(country_column_name)]
            country = Country.objects.filter(name=country_text).first()
            if country is None:
                errors.append(
                    f"Row: {row_index} - Country not found for field {country_column_name} and value {country_text}"
                )
            if document_type is None:
                errors.append(f"Row: {row_index} - Document type not found for field {number_column_name}")
            if individual.documents.filter(type=document_type).count() > 1:  # pragma: no cover
                errors.append(f"Row: {row_index} - Multiple documents with document type {document_type} found")
        return errors

    def validate(self, sheet: Worksheet, headers: List[str]) -> List[str]:
        errors = []
        row_index = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_index += 1
            if (row_index - 2) % 1000 == 0:
                print(f"Validating row {row_index - 2} to {row_index - 2 + 100} Indivduals")
            unicef_id = row[headers.index("unicef_id")]
            individuals_queryset = Individual.objects.filter(
                unicef_id=unicef_id, business_area=self.business_area, program=self.program
            )
            if not individuals_queryset.exists():  # pragma: no cover
                errors.append(f"Row: {row_index} - Individual with unicef_id {unicef_id} not found")
                continue
            if individuals_queryset.count() > 1:  # pragma: no cover
                errors.append(f"Row: {row_index} - Multiple individuals with unicef_id {unicef_id} found")
                continue
            individual = individuals_queryset.first()
            household = individual.household
            if household is None:  # pragma: no cover
                errors.append(f"Row: {row_index} - Household not found for individual with unicef_id {unicef_id}")
                continue
            errors.extend(self.validate_household_fields(row, headers, household, row_index))
            errors.extend(self.validate_individual_fields(row, headers, individual, row_index))
            errors.extend(self.validate_individual_flex_fields(row, headers, individual, row_index))
            errors.extend(self.validate_documents(row, headers, individual, row_index))
        return errors

    def handle_household_update(self, row: Tuple[Any, ...], headers: List[str], household: Any) -> None:
        for field, (_name, _validator, handler) in self.household_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(value, field, household, self.business_area, self.program)
            if self.ignore_empty_values and (handled_value is None or handled_value == ""):  # pragma: no cover
                continue
            setattr(household, _name, handled_value)

    def handle_individual_update(self, row: Tuple[Any, ...], headers: List[str], individual: Individual) -> None:
        for field, (_name, _validator, handler) in self.individual_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(value, field, individual, self.business_area, self.program)
            if self.ignore_empty_values and (handled_value is None or handled_value == ""):  # pragma: no cover
                continue
            setattr(individual, _name, handled_value)

    def handle_individual_flex_update(self, row: Tuple[Any, ...], headers: List[str], individual: Individual) -> None:
        for field, (name, _validator, handler) in self.individual_flex_fields.items():
            value = row[headers.index(field)]
            handled_value = handler(value, field, individual, self.business_area, self.program)
            if self.ignore_empty_values and (handled_value is None or handled_value == ""):  # pragma: no cover
                continue
            individual.flex_fields[name] = handled_value

    def handle_documents_update(self, row: Tuple[Any, ...], headers: List[str], individual: Individual) -> None:
        for number_column_name, country_column_name in self.document_fields:
            document_type = self.document_types.get(number_column_name)
            document_number = row[headers.index(number_column_name)]
            document_country = row[headers.index(country_column_name)]
            country = Country.objects.filter(name=document_country).first()
            if self.ignore_empty_values and (document_number is None or document_number == ""):  # pragma: no cover
                continue
            document = individual.documents.filter(type=document_type).first()
            if document:
                document.document_number = document_number
                document.status = Document.STATUS_PENDING
                document.save()
            else:
                Document.objects.create(
                    individual=individual,
                    type=document_type,
                    document_number=document_number,
                    country=country,
                    rdi_merge_status="MERGED",
                )

    def handle_update(self, sheet: Worksheet, headers: List[str]) -> List[str]:
        row_index = 1
        individual_ids = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_index += 1
            if (row_index - 2) % 1000 == 0:
                print(f"Updating row {row_index - 2} to {row_index - 2 + 100} Individuals")
            unicef_id = row[headers.index("unicef_id")]
            individual = Individual.objects.filter(
                unicef_id=unicef_id, business_area=self.business_area, program=self.program
            ).first()
            individual_ids.append(str(individual.id))
            household = individual.household
            self.handle_household_update(row, headers, household)
            self.handle_individual_update(row, headers, individual)
            self.handle_individual_flex_update(row, headers, individual)
            self.handle_documents_update(row, headers, individual)
            household.save()
            individual.save()
        return individual_ids

    def execute(self) -> None:
        workbook = load_workbook(filename=self.file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        errors = self.validate(sheet, headers)
        if errors:
            print("Validation failed")
            for error in errors:
                print(error)
            return
        print("Validation successful")
        processed_individuals_ids = self.handle_update(sheet, headers)
        if self.deduplicate_es:
            print("Deduplicating individuals Elasticsearch")
            DeduplicateTask(self.business_area.slug, self.program.id).deduplicate_individuals_against_population(
                Individual.objects.filter(id__in=processed_individuals_ids)
            )
        if self.deduplicate_documents:
            print("Deduplicating documents")
            HardDocumentDeduplication().deduplicate(
                Document.objects.filter(individual__id__in=processed_individuals_ids, status=Document.STATUS_PENDING),
                program=self.program,
            )

        print("Update successful")
