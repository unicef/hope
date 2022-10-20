import copy
import datetime
import functools
import io
import logging
from itertools import chain
from types import FunctionType
from typing import Dict, List, Tuple

from django.conf import settings
from django.contrib.sites.models import Site
from django.core.mail import EmailMultiAlternatives
from django.db.models import Count, DecimalField, F, Q, Sum
from django.template.loader import render_to_string
from django.urls import reverse

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from hct_mis_api.apps.account.models import User
from hct_mis_api.apps.core.models import BusinessArea
from hct_mis_api.apps.geo.models import Area
from hct_mis_api.apps.grievance.models import GrievanceTicket
from hct_mis_api.apps.household.models import Household
from hct_mis_api.apps.payment.models import PaymentRecord, PaymentVerification
from hct_mis_api.apps.program.models import Program
from hct_mis_api.apps.reporting.models import DashboardReport

logger = logging.getLogger(__name__)


class GenerateDashboardReportContentHelpers:
    @classmethod
    def get_beneficiaries(cls, report: DashboardReport):

        children_count_fields = [
            "female_age_group_0_5_count",
            "female_age_group_6_11_count",
            "female_age_group_12_17_count",
            "male_age_group_0_5_count",
            "male_age_group_6_11_count",
            "male_age_group_12_17_count",
        ]
        individual_count_fields = cls._get_all_individual_count_fields()
        valid_payment_records = cls._get_payment_records_for_report(report)

        (
            instances,
            valid_payment_records_in_instance_filter_key,
        ) = cls._get_business_areas_or_programs(report, valid_payment_records)

        for instance in instances:
            valid_payment_records_in_instance = valid_payment_records.filter(
                **{valid_payment_records_in_instance_filter_key: instance["id"]}
            )
            valid_households = Household.objects.filter(
                payment_records__in=valid_payment_records_in_instance
            ).distinct()
            households_aggr = cls._aggregate_instances_sum(
                valid_households,
                individual_count_fields,
            )
            instance["total_children"] = cls._reduce_aggregate(households_aggr, children_count_fields)
            instance["total_individuals"] = cls._reduce_aggregate(households_aggr, individual_count_fields)
            instance["num_households"] = valid_households.count()

        # get total distincts (can't use the sum of column since some households might belong to multiple programs)
        households = Household.objects.filter(payment_records__in=valid_payment_records).distinct()
        households_aggr = cls._aggregate_instances_sum(households, individual_count_fields)
        totals = {
            "num_households": households.count(),
            "total_individuals": cls._reduce_aggregate(households_aggr, individual_count_fields),
            "total_children": cls._reduce_aggregate(households_aggr, children_count_fields),
        }
        # return instances for rows and totals row info
        return instances, totals

    @classmethod
    def get_individuals(cls, report: DashboardReport):

        valid_payment_records = cls._get_payment_records_for_report(report)
        individual_count_fields = cls._get_all_with_disabled_individual_count_fields()
        (
            instances,
            valid_payment_records_in_instance_filter_key,
        ) = cls._get_business_areas_or_programs(report, valid_payment_records)

        for instance in instances:
            valid_payment_records_in_instance = valid_payment_records.filter(
                **{valid_payment_records_in_instance_filter_key: instance["id"]}
            )
            households_aggr = cls._aggregate_instances_sum(
                Household.objects.filter(payment_records__in=valid_payment_records_in_instance).distinct(),
                individual_count_fields,
            )
            instance.update(households_aggr)

        # get total distincts (can't use the sum of column since some households might belong to multiple programs)
        households_aggr = cls._aggregate_instances_sum(
            Household.objects.filter(payment_records__in=valid_payment_records).distinct(),
            individual_count_fields,
        )
        # return instances for rows and totals row info
        return instances, households_aggr

    @classmethod
    def get_volumes_by_delivery(cls, report: DashboardReport):

        valid_payment_records = cls._get_payment_records_for_report(report)
        (
            instances,
            valid_payment_records_in_instance_filter_key,
        ) = cls._get_business_areas_or_programs(report, valid_payment_records)

        def aggregate_by_delivery_type(payment_records):
            result = dict()
            for delivery_type in PaymentRecord.DELIVERY_TYPE_CHOICE:
                value = delivery_type[0]
                result[value] = (
                    payment_records.filter(delivery_type=value)
                    .aggregate(Sum("delivered_quantity_usd", output_field=DecimalField()))
                    .get("delivered_quantity_usd__sum")
                )
            return result

        for instance in instances:
            valid_payment_records_in_instance = valid_payment_records.filter(
                **{valid_payment_records_in_instance_filter_key: instance["id"]}
            )
            aggregated_by_delivery_type = aggregate_by_delivery_type(valid_payment_records_in_instance)
            instance.update(aggregated_by_delivery_type)

        totals = aggregate_by_delivery_type(valid_payment_records)
        return instances, totals

    @classmethod
    def get_programs(cls, report: DashboardReport):

        filter_vars = cls._format_filters(
            report,
            {},
            "cash_plans__payment_records__delivery_date",
            "admin_areas",
            "id",
            "business_area",
        )
        months_labels = cls.get_all_months()

        def get_filter_query(cash: bool, month: int):
            if cash:
                return Q(
                    cash_plans__payment_records__delivery_type__in=PaymentRecord.DELIVERY_TYPES_IN_CASH,
                    cash_plans__payment_records__delivery_date__month=month,
                )
            else:
                return Q(
                    cash_plans__payment_records__delivery_type__in=PaymentRecord.DELIVERY_TYPES_IN_VOUCHER,
                    cash_plans__payment_records__delivery_date__month=month,
                )

        def get_annotation(index_number: int, cash=True):
            key_label = months_labels[index_number]
            label = f"{key_label}_cash" if cash else f"{key_label}_voucher"
            return {
                label: Sum(
                    "cash_plans__payment_records__delivered_quantity_usd",
                    filter=get_filter_query(cash, index_number + 1),
                    output_field=DecimalField(),
                )
            }

        programs = (
            Program.objects.filter(**filter_vars)
            .distinct()
            .annotate(
                successful_payments=Count(
                    "cash_plans__payment_records",
                    filter=Q(cash_plans__payment_records__delivered_quantity_usd__gt=0),
                )
            )
            .annotate(
                unsuccessful_payments=Count(
                    "cash_plans__payment_records",
                    filter=Q(cash_plans__payment_records__delivered_quantity_usd=0),
                )
            )
        )
        for index_number in range(0, len(months_labels)):
            programs = programs.annotate(**get_annotation(index_number, True))
            programs = programs.annotate(**get_annotation(index_number, False))

        return programs, None

    @classmethod
    def get_grievances(cls, report: DashboardReport):
        filter_vars = {
            "created_at__year": report.year,
        }
        if report.admin_area:
            filter_vars["admin2"] = report.admin_area

        if not cls._is_report_global(report):
            filter_vars["business_area"] = report.business_area

        valid_grievances = GrievanceTicket.objects.filter(**filter_vars)
        days_30_from_now = datetime.date.today() - datetime.timedelta(days=30)
        days_60_from_now = datetime.date.today() - datetime.timedelta(days=60)

        feedback_categories = [
            GrievanceTicket.CATEGORY_POSITIVE_FEEDBACK,
            GrievanceTicket.CATEGORY_NEGATIVE_FEEDBACK,
        ]
        status_closed_query = Q(tickets__status=GrievanceTicket.STATUS_CLOSED)
        status_open_query = ~Q(tickets__status=GrievanceTicket.STATUS_CLOSED)
        instances = (
            BusinessArea.objects.filter(tickets__in=valid_grievances)
            .distinct()
            .annotate(total_grievances=Count("tickets", filter=~Q(tickets__category__in=feedback_categories)))
            .annotate(total_feedback=Count("tickets", filter=Q(tickets__category__in=feedback_categories)))
            .annotate(total_resolved=Count("tickets", filter=status_closed_query))
            .annotate(
                total_unresolved_lte_30=Count(
                    "tickets",
                    filter=Q(status_open_query, tickets__created_at__gte=days_30_from_now),
                )
            )
            .annotate(
                total_unresolved_30=Count(
                    "tickets",
                    filter=Q(
                        status_open_query,
                        tickets__created_at__lt=days_30_from_now,
                        tickets__created_at__gte=days_60_from_now,
                    ),
                )
            )
            .annotate(
                total_unresolved_60=Count(
                    "tickets",
                    filter=Q(status_open_query, tickets__created_at__lt=days_60_from_now),
                )
            )
            .annotate(
                total_open_sensitive=Count(
                    "tickets",
                    filter=Q(
                        status_open_query,
                        tickets__category=GrievanceTicket.CATEGORY_SENSITIVE_GRIEVANCE,
                    ),
                )
            )
        )

        totals = instances.aggregate(
            Sum("total_grievances"),
            Sum("total_feedback"),
            Sum("total_resolved"),
            Sum("total_unresolved_lte_30"),
            Sum("total_unresolved_30"),
            Sum("total_unresolved_60"),
            Sum("total_open_sensitive"),
        )

        return instances, totals

    @classmethod
    def get_payment_verifications(cls, report: DashboardReport):
        filter_vars = {"payment_record__delivery_date__year": report.year}
        if report.admin_area:
            filter_vars["payment_record__household__admin_area"] = report.admin_area
        if report.program:
            filter_vars["payment_record__cash_plan__program"] = report.program
        if not cls._is_report_global(report):
            filter_vars["payment_record__business_area"] = report.business_area
        valid_verifications = PaymentVerification.objects.filter(**filter_vars)
        path_to_payment_record_verifications = "cash_plans__verifications__payment_record_verifications"

        def format_status_filter(status):
            return Q(**{f"{path_to_payment_record_verifications}__status": status})

        programs = (
            Program.objects.filter(**{f"{path_to_payment_record_verifications}__in": valid_verifications})
            .distinct()
            .annotate(total_cash_plan_verifications=Count("cash_plans__verifications", distinct=True))
            .annotate(
                total_households=Count(
                    f"{path_to_payment_record_verifications}__payment_record__household",
                    distinct=True,
                )
            )
            .annotate(
                total_payment_records=Count(
                    "cash_plans__payment_records",
                    distinct=True,
                )
            )
            .annotate(
                all_possible_payment_records=Count(
                    "cash_plans__payment_records",
                    distinct=True,
                    filter=Q(
                        cash_plans__verifications__isnull=False,
                        cash_plans__payment_records__status=PaymentRecord.STATUS_SUCCESS,
                        cash_plans__payment_records__delivered_quantity__gt=0,
                    ),
                )
            )
            .annotate(total_verifications_done=Count(path_to_payment_record_verifications, distinct=True))
            .annotate(
                received=Count(
                    path_to_payment_record_verifications,
                    filter=format_status_filter(PaymentVerification.STATUS_RECEIVED),
                    distinct=True,
                )
            )
            .annotate(
                not_received=Count(
                    path_to_payment_record_verifications,
                    filter=format_status_filter(PaymentVerification.STATUS_NOT_RECEIVED),
                    distinct=True,
                )
            )
            .annotate(
                received_with_issues=Count(
                    path_to_payment_record_verifications,
                    filter=format_status_filter(PaymentVerification.STATUS_RECEIVED_WITH_ISSUES),
                    distinct=True,
                )
            )
            .annotate(
                not_responded=Count(
                    path_to_payment_record_verifications,
                    distinct=True,
                    filter=format_status_filter(PaymentVerification.STATUS_PENDING),
                )
            )
        )

        return programs, None

    @classmethod
    def get_total_transferred_by_country(cls, report: DashboardReport):
        # only for HQ dashboard
        business_areas = (
            BusinessArea.objects.filter(
                paymentrecord__delivered_quantity_usd__gt=0,
                paymentrecord__delivery_date__year=report.year,
            )
            .annotate(
                total_cash=Sum(
                    "paymentrecord__delivered_quantity_usd",
                    filter=Q(paymentrecord__delivery_type__in=PaymentRecord.DELIVERY_TYPES_IN_CASH),
                    output_field=DecimalField(),
                )
            )
            .annotate(
                total_voucher=Sum(
                    "paymentrecord__delivered_quantity_usd",
                    filter=Q(paymentrecord__delivery_type__in=PaymentRecord.DELIVERY_TYPES_IN_VOUCHER),
                    output_field=DecimalField(),
                )
            )
        )

        totals = business_areas.aggregate(
            Sum("total_cash", output_field=DecimalField()), Sum("total_voucher", output_field=DecimalField())
        )

        return business_areas, totals

    @classmethod
    def get_total_transferred_by_admin_area(cls, report: DashboardReport):
        # only for country dashboard
        valid_payment_records = cls._get_payment_records_for_report(report)
        admin_areas = (
            Area.objects.filter(
                area_type__area_level=2,
                household__payment_records__in=valid_payment_records,
            )
            .distinct()
            .annotate(
                total_transferred=Sum("household__payment_records__delivered_quantity_usd", output_field=DecimalField())
            )
            .annotate(num_households=Count("household", distinct=True))
        )

        totals = admin_areas.aggregate(
            Sum("total_transferred", output_field=DecimalField()), Sum("num_households", output_field=DecimalField())
        )
        admin_areas = admin_areas.values("id", "name", "p_code", "num_households", "total_transferred")

        individual_count_fields = cls._get_all_individual_count_fields()

        for admin_area in admin_areas:
            valid_payment_records_in_instance = valid_payment_records.filter(household__admin_area=admin_area["id"])
            households_aggr = cls._aggregate_instances_sum(
                Household.objects.filter(payment_records__in=valid_payment_records_in_instance).distinct(),
                individual_count_fields,
            )
            admin_area.update(households_aggr)

        totals.update(
            cls._aggregate_instances_sum(
                Household.objects.filter(payment_records__in=valid_payment_records).distinct(),
                individual_count_fields,
            )
        )

        return admin_areas, totals

    @staticmethod
    def format_beneficiaries_row(instance: Dict, is_totals: bool, *args) -> Tuple:
        return (
            instance.get("business_area_code", "") if not is_totals else "",
            instance.get("name", "") if not is_totals else "Total Distinct",
            instance.get("num_households", ""),
            instance.get("total_individuals", ""),
            instance.get("total_children", ""),
        )

    @classmethod
    def format_individuals_row(cls, instance: Dict, is_totals: bool, *args) -> Tuple:
        all_count_fields = cls._get_all_with_disabled_individual_count_fields()
        result = [
            instance.get("business_area_code", "") if not is_totals else "",
            instance.get("name", "") if not is_totals else "Total Distinct",
        ]
        for field in all_count_fields:
            result.append(instance.get(f"{field}__sum", 0))
        return tuple(result)

    @staticmethod
    def format_volumes_by_delivery_row(instance: Dict, is_totals: bool, *args):
        result = [
            instance.get("business_area_code", "") if not is_totals else "",
            instance.get("name", "") if not is_totals else "Total",
        ]
        for choice in PaymentRecord.DELIVERY_TYPE_CHOICE:
            result.append(instance.get(choice[0]))

        return tuple(result)

    @classmethod
    def format_programs_row(cls, instance: Program, *args) -> tuple:
        result = (
            instance.business_area.code,
            instance.business_area.name,
            instance.name,
            instance.sector,
            instance.cash_plus,
            instance.frequency_of_payments,
            instance.unsuccessful_payments,
            instance.successful_payments,
        )
        months = cls.get_all_months()
        for month in months:
            result += (
                getattr(instance, f"{month}_cash", 0),
                getattr(instance, f"{month}_voucher", 0),
            )
        return result

    @staticmethod
    def format_total_transferred_by_country(instance: BusinessArea, is_totals: bool, *args) -> tuple:
        if is_totals:
            return (
                "",
                "Total",
                instance.get("total_cash__sum") or 0,
                instance.get("total_voucher__sum") or 0,
            )
        else:
            return (
                instance.code,
                instance.name,
                instance.total_cash or 0,
                instance.total_voucher or 0,
            )

    @staticmethod
    def format_grievances_row(instance, is_totals: bool, is_hq: bool):
        if is_totals and not is_hq:
            # no totals row for country report
            return ()
        elif is_totals:
            return (
                "",
                "Total",
                instance.get("total_grievances__sum", 0),
                instance.get("total_feedback__sum", 0),
                instance.get("total_resolved__sum", 0),
                instance.get("total_unresolved_lte_30__sum", 0),
                instance.get("total_unresolved_30__sum", 0),
                instance.get("total_unresolved_60__sum", 0),
                instance.get("total_open_sensitive__sum", 0),
            )
        else:
            shared_cells = (
                instance.total_grievances,
                instance.total_feedback,
                instance.total_resolved,
                instance.total_unresolved_lte_30,
                instance.total_unresolved_30,
                instance.total_unresolved_60,
                instance.total_open_sensitive,
            )
            if is_hq:
                return (instance.code, instance.name) + shared_cells
            else:
                return (instance.code,) + shared_cells

    @staticmethod
    def format_payment_verifications_row(instance: Program, *args):
        return (
            instance.business_area.code,
            instance.business_area.name,
            instance.name,
            instance.total_cash_plan_verifications,
            instance.total_households,
            round((instance.total_verifications_done / instance.all_possible_payment_records) * 100)
            if instance.total_payment_records
            else 0,
            instance.received,
            instance.not_received,
            instance.received_with_issues,
            instance.not_responded,
        )

    @classmethod
    def format_total_transferred_by_admin_area_row(cls, instance, is_totals: bool, *args):
        fields_list = cls._get_all_individual_count_fields()

        shared_cells = tuple(instance.get(f"{field_name}__sum", 0) for field_name in fields_list)

        if is_totals:
            return (
                "",
                "Total",
                instance.get("total_transferred__sum", 0),
                instance.get("num_households__sum", 0),
            ) + shared_cells
        else:
            return (
                instance.get("name", ""),
                instance.get("p_code", ""),
                instance.get("total_transferred", 0),
                instance.get("num_households", 0),
            ) + shared_cells

    @staticmethod
    def _is_report_global(report: DashboardReport) -> bool:
        return report.business_area.slug == "global"

    @classmethod
    def _format_filters(
        cls,
        report: DashboardReport,
        custom_filters: Dict,
        date_path: str,
        admin_area_path: str,
        program_path: str,
        business_area_path: str,
    ) -> Dict:
        filter_vars = custom_filters or {}
        if date_path:
            filter_vars.update({f"{date_path}__year": report.year})
        if admin_area_path and report.admin_area:
            filter_vars.update({admin_area_path: report.admin_area, f"{admin_area_path}__level": 2})
        if program_path and report.program:
            filter_vars.update({program_path: report.program})
        if not cls._is_report_global(report) and business_area_path:
            filter_vars.update({business_area_path: report.business_area})
        return filter_vars

    @classmethod
    def _format_filters_for_payment_records(self, report: DashboardReport):
        return self._format_filters(
            report,
            {"delivered_quantity_usd__gt": 0},
            "delivery_date",
            "household__admin_area",
            "cash_plan__program",
            "business_area",
        )

    @classmethod
    def _get_payment_records_for_report(self, report):
        filter_vars = self._format_filters_for_payment_records(report)
        return PaymentRecord.objects.filter(**filter_vars)

    @classmethod
    def _get_business_areas_or_programs(cls, report, valid_payment_records):
        if cls._is_report_global(report):
            business_area_code_path = "code"
            instances = BusinessArea.objects.filter(paymentrecord__in=valid_payment_records)
            valid_payment_records_in_instance_filter_key = "business_area"
        else:
            business_area_code_path = "business_area__code"
            instances = Program.objects.filter(cash_plans__payment_records__in=valid_payment_records)
            valid_payment_records_in_instance_filter_key = "cash_plan__program"

        instances = (
            instances.distinct()
            .annotate(business_area_code=F(business_area_code_path))
            .values("id", "name", "business_area_code")
        )
        return instances, valid_payment_records_in_instance_filter_key

    @staticmethod
    def _aggregate_instances_sum(instances, field_list: List) -> Dict:
        aggregation_list = [Sum(field_name) for field_name in field_list]
        return instances.aggregate(*aggregation_list)

    @staticmethod
    def _reduce_aggregate(aggregate: Dict, fields_list: List) -> int:
        return functools.reduce(
            lambda a, b: a + aggregate[f"{b}__sum"] if aggregate[f"{b}__sum"] else a,
            fields_list,
            0,
        )

    @staticmethod
    def _get_all_with_disabled_individual_count_fields():
        return [
            "female_age_group_0_5_count",
            "female_age_group_0_5_disabled_count",
            "female_age_group_6_11_count",
            "female_age_group_6_11_disabled_count",
            "female_age_group_12_17_count",
            "female_age_group_12_17_disabled_count",
            "female_age_group_18_59_count",
            "female_age_group_18_59_disabled_count",
            "female_age_group_60_count",
            "female_age_group_60_disabled_count",
            "male_age_group_0_5_count",
            "male_age_group_0_5_disabled_count",
            "male_age_group_6_11_count",
            "male_age_group_6_11_disabled_count",
            "male_age_group_12_17_count",
            "male_age_group_12_17_disabled_count",
            "male_age_group_18_59_count",
            "male_age_group_18_59_disabled_count",
            "male_age_group_60_count",
            "male_age_group_60_disabled_count",
        ]

    @staticmethod
    def _get_all_individual_count_fields():
        return [
            "female_age_group_0_5_count",
            "female_age_group_6_11_count",
            "female_age_group_12_17_count",
            "female_age_group_18_59_count",
            "female_age_group_60_count",
            "male_age_group_0_5_count",
            "male_age_group_6_11_count",
            "male_age_group_12_17_count",
            "male_age_group_18_59_count",
            "male_age_group_60_count",
        ]

    @staticmethod
    def get_all_months():
        return [
            "january",
            "february",
            "march",
            "april",
            "may",
            "june",
            "july",
            "august",
            "september",
            "october",
            "november",
            "december",
        ]


class GenerateDashboardReportService:
    HQ = 1
    COUNTRY = 2
    SHARED = 3
    HEADERS = {
        DashboardReport.BENEFICIARIES_REACHED: {
            HQ: ("business area", "country"),
            COUNTRY: ("business area", "programme"),
            SHARED: ("households reached", "individuals reached", "children reached"),
        },
        DashboardReport.TOTAL_TRANSFERRED_BY_ADMIN_AREA: {
            HQ: (),
            COUNTRY: (
                "Admin Level 2",
                "Admin Code",
                "Total tranferred (USD)",
                "Households reached",
                "Female 0-5 Reached",
                "Female 6-11 Reached",
                "Female 12-17 Reached",
                "Female 18-59 Reached",
                "Female 60+ Reached",
                "Male 0-5 Reached",
                "Male 6-11 Reached",
                "Male 12-17 Reached",
                "Male 18-59 Reached",
                "Male 60+ Reached",
            ),
            SHARED: (),
        },
        DashboardReport.PAYMENT_VERIFICATION: {
            HQ: (),
            COUNTRY: (),
            SHARED: (
                "business area",
                "country",
                "programme",
                "cash plan verifications",
                "Households Contacted",
                "average sampling",
                "Received",
                "Not Received",
                "Received with issues",
                "Not Responded",
            ),
        },
        DashboardReport.GRIEVANCES_AND_FEEDBACK: {
            HQ: (
                "business area",
                "country",
            ),
            COUNTRY: ("business area",),
            SHARED: (
                "grievance tickets",
                "feedback tickets",
                "resolved tickets",
                "Unresolved =<30 days",
                "Unresolved >30 days",
                "Unresolved >60 days",
                "open sensitive grievances",
            ),
        },
        DashboardReport.TOTAL_TRANSFERRED_BY_COUNTRY: {
            HQ: (
                "business area",
                "country",
                "actual cash transferred",
                "actual voucher transferred",
            ),
            COUNTRY: (),
            SHARED: (),
        },
        DashboardReport.PROGRAMS: {
            HQ: (),
            COUNTRY: (),
            SHARED: (
                "business area",
                "country",
                "programme",
                "sector",
                "cash+",
                "frequency",
                "unsuccessful payment",
                "successful payment",
            )
            + tuple(
                chain.from_iterable(
                    [
                        [f"{month.capitalize()} cash", f"{month.capitalize()} voucher"]
                        for month in GenerateDashboardReportContentHelpers.get_all_months()
                    ]
                )
            ),
        },
        DashboardReport.VOLUME_BY_DELIVERY_MECHANISM: {
            HQ: (
                "business area",
                "country",
            ),
            COUNTRY: (
                "business area",
                "programme",
            ),
            SHARED: tuple(choice[1] for choice in PaymentRecord.DELIVERY_TYPE_CHOICE),
        },
        DashboardReport.INDIVIDUALS_REACHED: {
            HQ: (
                "business area",
                "country",
            ),
            COUNTRY: (
                "business area",
                "programme",
            ),
            SHARED: (
                "females 0-5 reached",
                "females 0-5 w/ disability reached",
                "females 6-11 reached",
                "females 6-11 w/ disability reached",
                "females 12-17 reached",
                "females 12-17 w/ disability reached",
                "females 18-59 reached",
                "females 18-59 w/ disability reached",
                "females 60+ reached",
                "females 60+ w/ disability reached",
                "males 0-5 reached",
                "males 0-5 w/ disability reached",
                "males 6-11 reached",
                "males 6-11 w/ disability reached",
                "males 12-17 reached",
                "males 12-17 w/ disability reached",
                "males 18-59 reached",
                "males 18-59 w/ disability reached",
                "males 60+ reached",
                "males 60+ w/ disability reached",
            ),
        },
    }
    ROW_CONTENT_METHODS = {
        DashboardReport.BENEFICIARIES_REACHED: (
            GenerateDashboardReportContentHelpers.get_beneficiaries,
            GenerateDashboardReportContentHelpers.format_beneficiaries_row,
        ),
        DashboardReport.INDIVIDUALS_REACHED: (
            GenerateDashboardReportContentHelpers.get_individuals,
            GenerateDashboardReportContentHelpers.format_individuals_row,
        ),
        DashboardReport.VOLUME_BY_DELIVERY_MECHANISM: (
            GenerateDashboardReportContentHelpers.get_volumes_by_delivery,
            GenerateDashboardReportContentHelpers.format_volumes_by_delivery_row,
        ),
        DashboardReport.PROGRAMS: (
            GenerateDashboardReportContentHelpers.get_programs,
            GenerateDashboardReportContentHelpers.format_programs_row,
        ),
        DashboardReport.TOTAL_TRANSFERRED_BY_COUNTRY: (
            GenerateDashboardReportContentHelpers.get_total_transferred_by_country,
            GenerateDashboardReportContentHelpers.format_total_transferred_by_country,
        ),
        DashboardReport.GRIEVANCES_AND_FEEDBACK: (
            GenerateDashboardReportContentHelpers.get_grievances,
            GenerateDashboardReportContentHelpers.format_grievances_row,
        ),
        DashboardReport.PAYMENT_VERIFICATION: (
            GenerateDashboardReportContentHelpers.get_payment_verifications,
            GenerateDashboardReportContentHelpers.format_payment_verifications_row,
        ),
        DashboardReport.TOTAL_TRANSFERRED_BY_ADMIN_AREA: (
            GenerateDashboardReportContentHelpers.get_total_transferred_by_admin_area,
            GenerateDashboardReportContentHelpers.format_total_transferred_by_admin_area_row,
        ),
    }
    META_HEADERS = (
        "report type",
        "creation date",
        "created by",
        "business area",
        "report year",
    )
    REMOVE_EMPTY_COLUMNS = {
        DashboardReport.VOLUME_BY_DELIVERY_MECHANISM: (
            3,
            len(PaymentRecord.DELIVERY_TYPE_CHOICE) + 3,
        )
    }
    META_SHEET = "Meta data"
    MAX_COL_WIDTH = 75

    def __init__(self, report: DashboardReport):
        self.report = report
        self.report_types = report.report_type
        self.business_area = report.business_area
        self.hq_or_country = self.HQ if report.business_area.slug == "global" else self.COUNTRY

    def _create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_meta = wb.active
        ws_meta.title = self.META_SHEET
        self.wb = wb
        self.ws_meta = ws_meta
        return wb

    def _format_meta_tab(self):
        self.ws_meta.append(self.META_HEADERS)
        info_row = (
            self._report_types_to_joined_str(),
            self._format_date(self.report.created_at),
            self._format_user_name(self.report.created_by),
            self.business_area.name,
            str(self.report.year),
        )
        self.ws_meta.append(info_row)

    def _add_headers(self, active_sheet, report_type) -> int:
        headers_row = self.HEADERS[report_type][self.hq_or_country] + self.HEADERS[report_type][self.SHARED]
        headers_row = self._stringify_all_values(headers_row)
        active_sheet.append(headers_row)
        return len(headers_row)

    def _add_rows(self, active_sheet, report_type):
        is_hq_report = self.hq_or_country == self.HQ
        get_row_methods: List[FunctionType] = self.ROW_CONTENT_METHODS[report_type]
        all_instances, totals = get_row_methods[0](self.report)
        for instance in all_instances:
            row = get_row_methods[1](instance, False, is_hq_report)
            str_row = self._stringify_all_values(row)
            active_sheet.append(str_row)
            # active_sheet.append(str_row)
        # append totals row
        if totals:
            row = get_row_methods[1](totals, True, is_hq_report)
            str_row = self._stringify_all_values(row)
            active_sheet.append(str_row)
        return len(all_instances)

    def generate_workbook(self) -> openpyxl.Workbook:
        self._create_workbook()
        self._format_meta_tab()
        self._add_font_style_to_sheet(self.ws_meta)
        self._adjust_column_width_from_col(self.ws_meta, 1, 5, 1)

        # loop through all selected report types and add sheet for each
        for report_type in self.report_types:
            sheet_title = self._report_type_to_str(report_type)
            active_sheet = self.wb.create_sheet(sheet_title, -1)
            number_of_columns = self._add_headers(active_sheet, report_type)
            number_of_rows = self._add_rows(active_sheet, report_type)
            self._add_font_style_to_sheet(active_sheet, number_of_rows + 2)
            remove_empty_columns_values = self.REMOVE_EMPTY_COLUMNS.get(report_type)
            if remove_empty_columns_values:
                self._remove_empty_columns(
                    active_sheet,
                    number_of_rows + 2,
                    remove_empty_columns_values[0],
                    remove_empty_columns_values[1],
                )
            self._adjust_column_width_from_col(active_sheet, 1, number_of_columns, 1)

        return self.wb

    def generate_report(self):
        try:
            self.generate_workbook()
            file_name = (
                self._report_type_to_str(self.report_types[0]) if len(self.report_types) == 1 else "Multiple reports"
            )
            self.report.file.save(
                f"{file_name}-{self._format_date(self.report.created_at)}.xlsx",
                io.BytesIO(save_virtual_workbook(self.wb)),
                save=False,
            )
            self.report.status = DashboardReport.COMPLETED
        except Exception as e:
            logger.exception(e)
            self.report.status = DashboardReport.FAILED
        self.report.save()

        if self.report.file:
            self._send_email()

    def _send_email(self):
        path = reverse("dashboard_report", kwargs={"report_id": self.report.id})
        protocol = "http" if settings.IS_DEV else "https"
        context = {
            "report_type": self._report_types_to_joined_str(),
            "created_at": self._format_date(self.report.created_at),
            "report_url": f"{protocol}://{Site.objects.first()}{path}",
        }
        text_body = render_to_string("dashboard_report.txt", context=context)
        html_body = render_to_string("dashboard_report.html", context=context)
        msg = EmailMultiAlternatives(
            subject="HOPE report generated",
            from_email=settings.EMAIL_HOST_USER,
            to=[self.report.created_by.email],
            body=text_body,
        )
        msg.attach_alternative(html_body, "text/html")
        msg.send()

    @staticmethod
    def _adjust_column_width_from_col(ws, min_col, max_col, min_row):
        column_widths = []
        for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
            for cell in col:
                value = cell.value
                if value is not None:
                    if isinstance(value, str) is False:
                        value = str(value)
                    if len(value) > GenerateDashboardReportService.MAX_COL_WIDTH:
                        alignment = copy.copy(cell.alignment)
                        alignment.wrapText = True
                        cell.alignment = alignment
                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i in range(len(column_widths)):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            value = (
                GenerateDashboardReportService.MAX_COL_WIDTH
                if value > GenerateDashboardReportService.MAX_COL_WIDTH
                else value
            )
            ws.column_dimensions[col_name].width = value

    @staticmethod
    def _add_font_style_to_sheet(ws, totals_row=None):
        bold_font = Font(bold=True)
        for cell in ws["1:1"]:
            cell.font = bold_font
        if totals_row:
            ws[f"B{totals_row}"].font = bold_font

    @staticmethod
    def _report_type_to_str(report_type) -> str:
        types_dict = dict(DashboardReport.REPORT_TYPES)
        label = str(types_dict.get(report_type, ""))
        return label[:31]

    def _report_types_to_joined_str(self) -> str:
        return ", ".join([self._report_type_to_str(report_type) for report_type in self.report_types])

    @staticmethod
    def _stringify_all_values(row: tuple) -> tuple:
        str_row = []
        for value in row:
            if isinstance(value, (str, int, float)):
                str_row.append(value)
            else:
                str_row.append(str(value if value is not None else ""))
        return tuple(str_row)

    @staticmethod
    def _format_date(date) -> str:
        return date.strftime("%Y-%m-%d") if date else ""

    @staticmethod
    def _format_user_name(user: User) -> str:
        return (
            f"{user.first_name} {user.last_name}" if user.first_name or user.last_name else user.email or user.username
        )

    @staticmethod
    def _remove_empty_columns(ws, totals_row, min_col=1, max_col=2):
        to_remove_columns = []
        for col_idx in range(min_col, max_col):
            col_letter = get_column_letter(col_idx)
            if not ws[f"{col_letter}{totals_row}"].value:
                to_remove_columns.append(col_idx)
        columns_removed = 0
        for column in to_remove_columns:
            ws.delete_cols(column - columns_removed)
            columns_removed += 1
        return len(to_remove_columns)
