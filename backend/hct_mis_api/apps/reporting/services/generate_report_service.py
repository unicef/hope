import copy
import logging
import openpyxl

from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from tempfile import NamedTemporaryFile

from django.conf import settings
from django.contrib.postgres.aggregates.general import ArrayAgg
from django.core.files import File
from django.core.mail import EmailMultiAlternatives
from django.db.models import Count, DecimalField, Max, Min, Q, Sum
from django.template.loader import render_to_string

from hct_mis_api.apps.core.utils import decode_id_string, encode_id_base64
from hct_mis_api.apps.geo.models import Area
from hct_mis_api.apps.grievance.models import GrievanceTicket
from hct_mis_api.apps.household.models import (
    NONE,
    WORK_STATUS_CHOICE,
    Household,
    Individual,
)
from hct_mis_api.apps.payment.models import (
    CashPlan,
    PaymentPlan,
    PaymentRecord,
    PaymentVerification,
    PaymentVerificationPlan,
)
from hct_mis_api.apps.program.models import Program
from hct_mis_api.apps.reporting.models import Report

logger = logging.getLogger(__name__)


class GenerateReportContentHelpers:
    @staticmethod
    def get_individuals(report: Report):
        filter_vars = {
            "household__business_area": report.business_area,
            "withdrawn": False,
            "duplicate": False,
            "last_registration_date__gte": report.date_from,
            "last_registration_date__lte": report.date_to,
        }
        if report.admin_area.all().exists():
            filter_vars["household__admin_area__in"] = report.admin_area.all()
        return Individual.objects.filter(**filter_vars)

    @classmethod
    def format_individual_row(self, individual: Individual) -> tuple:

        return (
            individual.household.id,
            individual.household.country_origin.name if individual.household.country_origin else "",
            individual.household.admin_area.name if individual.household.admin_area else "",
            individual.birth_date,
            individual.estimated_birth_date,
            individual.sex,
            individual.marital_status,
            len(individual.observed_disability) >= 1 and NONE not in individual.observed_disability,
            individual.observed_disability,
            individual.comms_disability,
            individual.hearing_disability,
            individual.memory_disability,
            individual.physical_disability,
            individual.seeing_disability,
            individual.selfcare_disability,
            individual.pregnant,
            individual.relationship,
            self._to_values_list(individual.households_and_roles.all(), "role"),
            dict(WORK_STATUS_CHOICE).get(individual.work_status, ""),
            individual.sanction_list_possible_match,
            individual.deduplication_batch_status,
            individual.deduplication_golden_record_status,
            individual.deduplication_golden_record_results.get("duplicates", "")
            if individual.deduplication_golden_record_results
            else "",
            individual.deduplication_golden_record_results.get("possible_duplicates", "")
            if individual.deduplication_golden_record_results
            else "",
            self._format_date(individual.first_registration_date),
            self._format_date(individual.last_registration_date),
        )

    @staticmethod
    def get_households(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "withdrawn": False,
            "last_registration_date__gte": report.date_from,
            "last_registration_date__lte": report.date_to,
        }
        if report.admin_area.all().exists():
            filter_vars["admin_area__in"] = report.admin_area.all()
        return Household.objects.filter(**filter_vars)

    @classmethod
    def format_household_row(self, household: Household) -> tuple:
        row = [
            household.id,
            household.country_origin.name if household.country_origin else "",
            household.admin_area.name if household.admin_area else "",
            household.size,
            household.geopoint[0] if household.geopoint else "",
            household.geopoint[1] if household.geopoint else "",
            household.residence_status,
            household.returnee,
            household.status,
            household.village,
            household.female_age_group_0_5_count,
            household.female_age_group_0_5_disabled_count,
            household.female_age_group_6_11_count,
            household.female_age_group_6_11_disabled_count,
            household.female_age_group_12_17_count,
            household.female_age_group_12_17_disabled_count,
            household.female_age_group_18_59_count,
            household.female_age_group_18_59_disabled_count,
            household.female_age_group_60_count,
            household.female_age_group_60_disabled_count,
            household.pregnant_count,
            household.male_age_group_0_5_count,
            household.male_age_group_0_5_disabled_count,
            household.male_age_group_6_11_count,
            household.male_age_group_6_11_disabled_count,
            household.male_age_group_12_17_count,
            household.male_age_group_12_17_disabled_count,
            household.male_age_group_18_59_count,
            household.male_age_group_18_59_disabled_count,
            household.male_age_group_60_count,
            household.male_age_group_60_disabled_count,
            self._format_date(household.first_registration_date),
            self._format_date(household.last_registration_date),
            household.org_name_enumerator,
        ]
        for program in household.programs.all():
            row.append(program.name)
        return tuple(row)

    @staticmethod
    def get_cash_plan_verifications(report: Report):
        pp_business_area_ids = list(
            CashPlan.objects
            .filter(business_area=report.business_area)
            .values_list("id", flat=True)
        )
        filter_vars = {
            "payment_plan_object_id__in": pp_business_area_ids,
            "completion_date__isnull": False,
            "completion_date__gte": report.date_from,
            "completion_date__lte": report.date_to,
        }
        if report.program:
            pp_program_ids = list(
                CashPlan.objects
                .filter(program=report.program)
                .values_list("id", flat=True)
            )
            filter_vars["payment_plan_object_id__in"] = pp_program_ids
        return PaymentVerificationPlan.objects.filter(**filter_vars)

    @staticmethod
    def _map_admin_area_names_from_ids(admin_areas_ids: list) -> str:
        if not admin_areas_ids:
            return ""
        result = []
        for admin_area_id in admin_areas_ids:
            admin_area_id = decode_id_string(admin_area_id)
            if admin_area := Area.objects.filter(id=admin_area_id).first():
                result.append(admin_area.name)
        return ", ".join(result)

    @classmethod
    def format_cash_plan_verification_row(self, verification: PaymentVerificationPlan) -> tuple:
        return (
            verification.id,
            verification.payment_plan.unicef_id,
            verification.payment_plan.program.name,
            self._format_date(verification.activation_date),
            verification.status,
            verification.verification_channel,
            self._format_date(verification.completion_date),
            verification.sample_size,
            verification.responded_count,
            verification.received_count,
            verification.received_with_problems_count,
            verification.not_received_count,
            verification.sampling,
            verification.sex_filter,
            self._map_admin_area_names_from_ids(verification.excluded_admin_areas_filter),
            verification.age_filter,
        )

    @staticmethod
    def get_payments(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "delivery_date__date__range": (report.date_from, report.date_to),
        }
        if report.admin_area.all().exists():
            filter_vars["household__admin_area__in"] = report.admin_area.all()
        return PaymentRecord.objects.filter(**filter_vars)

    @classmethod
    def format_payment_row(self, payment: PaymentRecord) -> tuple:
        cash_or_voucher = ""
        if payment.delivery_type:
            if payment.delivery_type in [
                PaymentRecord.DELIVERY_TYPE_CASH,
                PaymentRecord.DELIVERY_TYPE_DEPOSIT_TO_CARD,
                PaymentRecord.DELIVERY_TYPE_TRANSFER,
            ]:
                cash_or_voucher = "cash"
            else:
                # TODO: check if this is even an option, I don't see the delivery_type being anything else but above three options
                # but following the spreadsheet here
                cash_or_voucher = "voucher"

        return (
            payment.ca_id,
            payment.parent.ca_id if payment.parent else "",
            payment.status,
            payment.currency,
            payment.delivered_quantity,
            payment.delivered_quantity_usd or payment.delivered_quantity,
            self._format_date(payment.delivery_date),
            payment.delivery_type,
            payment.distribution_modality,
            payment.entitlement_quantity,
            payment.target_population.id,
            payment.target_population.name,
            cash_or_voucher,
            payment.household.id,
        )

    @staticmethod
    def get_payment_verifications(report: Report):
        pp_business_area_ids = list(
            PaymentPlan.objects
            .filter(business_area=report.business_area)
            .values_list("id", flat=True)
        )
        filter_vars = {
            "payment_verification_plan__payment_plan_object_id__in": pp_business_area_ids,
            "payment_verification_plan__completion_date__isnull": False,
            "payment_verification_plan__completion_date__date__range": (report.date_from, report.date_to),
        }
        if report.program:
            pp_program_ids = list(
                PaymentPlan.objects
                .filter(program=report.program)
                .values_list("id", flat=True)
            )
            filter_vars["payment_verification_plan__payment_plan_object_id__in"] = pp_program_ids
        return PaymentVerification.objects.filter(**filter_vars)

    @classmethod
    def format_payment_verification_row(self, payment_verification: PaymentVerification) -> tuple:
        return (
            payment_verification.payment_verification_plan.id,
            payment_verification.payment_record.ca_id,
            payment_verification.payment_verification_plan.payment_plan.unicef_id,
            self._format_date(payment_verification.payment_verification_plan.completion_date),
            payment_verification.received_amount,
            payment_verification.status,
            payment_verification.status_date,
        )

    @staticmethod
    def get_payment_plans(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "dispersion_start_date__gte": report.date_from,
            "dispersion_end_date__lte": report.date_to,
        }
        return PaymentPlan.objects.filter(**filter_vars)

    @classmethod
    def format_payment_plan_row(self, payment_plan: PaymentPlan) -> tuple:
        return (
            payment_plan.unicef_id,
            payment_plan.get_status_display(),
            payment_plan.total_households_count,
            payment_plan.get_currency_display(),
            payment_plan.total_entitled_quantity,
            payment_plan.total_delivered_quantity,
            payment_plan.total_undelivered_quantity,
            self._format_date(payment_plan.dispersion_start_date),
            self._format_date(payment_plan.dispersion_end_date),
        )

    @staticmethod
    def get_cash_plans(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "end_date__gte": report.date_from,
            "end_date__lte": report.date_to,
        }
        if report.program:
            filter_vars["program"] = report.program
        return CashPlan.objects.filter(**filter_vars)

    @classmethod
    def format_cash_plan_row(self, cash_plan: CashPlan) -> tuple:
        return (
            cash_plan.ca_id,
            cash_plan.name,
            self._format_date(cash_plan.start_date),
            self._format_date(cash_plan.end_date),
            cash_plan.program.name,
            cash_plan.funds_commitment,
            cash_plan.assistance_measurement,
            cash_plan.assistance_through,
            cash_plan.delivery_type,
            self._format_date(cash_plan.dispersion_date),
            cash_plan.down_payment,
            cash_plan.total_delivered_quantity,
            cash_plan.total_undelivered_quantity,
            cash_plan.total_entitled_quantity,
            cash_plan.total_entitled_quantity_revised,
            cash_plan.total_persons_covered,
            cash_plan.total_persons_covered_revised,
            cash_plan.status,
            self._format_date(cash_plan.status_date),
            cash_plan.vision_id,
            cash_plan.validation_alerts_count,
            # cash_plan.verification_status,
        )

    @staticmethod
    def get_programs(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "end_date__gte": report.date_from,
            "end_date__lte": report.date_to,
        }
        return Program.objects.filter(**filter_vars)

    @classmethod
    def format_program_row(self, program: Program) -> tuple:
        return (
            program.id,
            program.name,
            program.scope,
            program.sector,
            program.status,
            program.start_date,
            program.end_date,
            program.cash_plus,
            program.description,
            program.budget,
            program.frequency_of_payments,
            program.administrative_areas_of_implementation,
            program.population_goal,
            program.total_number_of_households,
        )

    @staticmethod
    def get_payments_for_individuals(report: Report):
        if isinstance(report.date_to, str):
            report.date_to = datetime.strptime(report.date_to, "%Y-%m-%d").date()

        date_to_time = datetime.fromordinal(report.date_to.toordinal())
        date_to_time += timedelta(days=1)
        filter_vars = {
            "household__paymentrecord__business_area": report.business_area,
            "household__paymentrecord__delivery_date__gte": report.date_from,
            "household__paymentrecord__delivery_date__lt": date_to_time,
        }
        if report.admin_area.all().exists():
            filter_vars["household__admin_area__in"] = report.admin_area.all()
        if report.program:
            filter_vars["household__paymentrecord__parent__program"] = report.program

        return (
            Individual.objects.filter(**filter_vars)
            .annotate(first_delivery_date=Min("household__paymentrecord__delivery_date"))
            .annotate(last_delivery_date=Max("household__paymentrecord__delivery_date"))
            .annotate(
                payments_made=Count(
                    "household__paymentrecord",
                    filter=Q(household__paymentrecord__delivered_quantity__gte=0),
                )
            )
            .annotate(payment_currency=ArrayAgg("household__paymentrecord__currency"))
            .annotate(
                total_delivered_quantity_local=Sum(
                    "household__paymentrecord__delivered_quantity", output_field=DecimalField()
                )
            )
            .annotate(
                total_delivered_quantity_usd=Sum(
                    "household__paymentrecord__delivered_quantity_usd", output_field=DecimalField()
                )
            )
            .order_by("household__id")
        )

    @classmethod
    def format_payments_for_individuals_row(self, individual: Individual) -> tuple:
        return (
            individual.household.id,
            individual.household.country_origin.name if individual.household.country_origin else "",
            individual.household.admin_area.name if individual.household.admin_area else "",
            self._format_date(individual.first_delivery_date),
            self._format_date(individual.last_delivery_date),
            individual.payments_made,
            ", ".join(individual.payment_currency),
            individual.total_delivered_quantity_local,
            individual.total_delivered_quantity_usd or individual.total_delivered_quantity_local,
            individual.birth_date,
            individual.estimated_birth_date,
            individual.sex,
            individual.marital_status,
            len(individual.observed_disability) >= 1 and NONE not in individual.observed_disability,
            individual.observed_disability,
            individual.comms_disability,
            individual.hearing_disability,
            individual.memory_disability,
            individual.physical_disability,
            individual.seeing_disability,
            individual.selfcare_disability,
            individual.pregnant,
            individual.relationship,
            self._to_values_list(individual.households_and_roles.all(), "role"),
            dict(WORK_STATUS_CHOICE).get(individual.work_status, ""),
            individual.sanction_list_possible_match,
            individual.deduplication_batch_status,
            individual.deduplication_golden_record_status,
            individual.deduplication_golden_record_results.get("duplicates", "")
            if individual.deduplication_golden_record_results
            else "",
            individual.deduplication_golden_record_results.get("possible_duplicates", "")
            if individual.deduplication_golden_record_results
            else "",
        )

    @staticmethod
    def get_grievance_tickets(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "created_at__gte": report.date_from,
            "created_at__lte": report.date_to,
        }

        return GrievanceTicket.objects.filter(**filter_vars).select_related("admin2", "created_by", "assigned_to")

    @classmethod
    def format_grievance_tickets_row(cls, grievance_ticket: GrievanceTicket) -> tuple:
        def get_full_name(user):
            if not user:
                return ""
            return " ".join(filter(None, [user.first_name, user.last_name]))

        def get_username(user):
            if not user:
                return ""
            return user.username

        return (
            grievance_ticket.unicef_id,
            grievance_ticket.created_at,
            grievance_ticket.updated_at,
            grievance_ticket.get_status_display(),
            grievance_ticket.get_category_display(),
            grievance_ticket.get_issue_type(),
            grievance_ticket.admin2.name,
            grievance_ticket.admin2.p_code,
            get_username(grievance_ticket.created_by),
            get_full_name(grievance_ticket.created_by),
            get_username(grievance_ticket.assigned_to),
            get_full_name(grievance_ticket.assigned_to),
        )

    @staticmethod
    def _to_values_list(instances, field_name: str) -> str:
        values_list = list(instances.values_list(field_name, flat=True))
        return ", ".join([str(value) for value in values_list])

    @staticmethod
    def _format_date(date) -> str:
        if not date:
            return ""
        return date.strftime("%Y-%m-%d")


class GenerateReportService:
    HEADERS = {
        Report.INDIVIDUALS: (
            "household id",  # 8e8ea94a-2ca5-4b76-b055-e098bc24eee8
            "country of origin",  # South Sudan
            "administrative area 2",  # Juba
            "birth date",  # 2000-06-24
            "estimated birth date",  # TRUE
            "gender",  # FEMALE,
            "marital status",  # MARRIED
            "disability",  # TRUE
            "observed disability",
            "communication disability",
            "hearing disability",  # LOT_DIFFICULTY
            "remembering disability",
            "physical disability",
            "seeing disability",
            "self-care disability",
            "pregnant",  # TRUE
            "relationship to hoh",  # WIFE
            "role",  # PRIMARY
            "work status",  # NOT_PROVIDED
            "sanction list possible match",  # FALSE
            "dedupe in batch status",  # UNIQUE_IN_BATCH
            "dedupe in Pop. status",  # DUPLICATE
            "dedupe in Pop.duplicates",
            "dedupe in Pop. possible duplicates",
            "first registration date",  # 2000-06-24
            "last registration date",  # 2000-06-24
        ),
        Report.HOUSEHOLD_DEMOGRAPHICS: (
            "household id",
            "country of origin",  # South Sudan
            "administrative area 2",  # Juba
            "household size",  # 4
            "latitude",  # 54,367759
            "longitude",  # 60,964675
            "residence status",  # HOST
            "returnee",  # FALSE
            "status",  # ACTIVE
            "village",  # Mendika
            "females 0-5",  # 0
            "females 0-5 w/ disability",  # 0
            "females 6-11",  # 1
            "females 6-11 w/ disability",  # 1
            "females 12-17",  # 1
            "females 12-17 w/ disability",  # 0
            "females 18-59",  # 1
            "females 18-59 w/ disability",  # 0
            "females 60+",  # 0
            "females 60+ w/ disability",  # 0
            "pregnant females",  # 0
            "males 0-5",  # 0
            "males 0-5 w/ disability",  # 0
            "males 6-11",  # 1
            "males 6-11 w/ disability",  # 1
            "males 12-17",  # 1
            "males 12-17 w/ disability",  # 0
            "males 18-59",  # 1
            "males 18-59 w/ disability",  # 0
            "males 60+",  # 0
            "males 60+ w/ disability",  # 0
            "first registration date",  # 2020-08-25
            "last registration date",  # 2020-08-25
            "organization name enumerator",
        ),
        Report.CASH_PLAN_VERIFICATION: (
            "cash plan verification ID",
            "cash plan ID",  # ANT-21-CSH-00001
            "programme",  # Winterization 2020
            "activation date",
            "status",
            "verification method",
            "completion date",
            "sample size",  # 500
            "responded",  # 340
            "received",  # 320
            "received with issues",  # 12
            "not received",  # 8
            "sampling",  # FULL_LIST or RANDOM
            "gender filter",  # FEMALE
            "excluded admin areas",  # Juba
            "age filter",  # {'max': 100, 'min': 0}
        ),
        Report.PAYMENTS: (
            "payment record ID",  # ANT-21-CSH-00001-0000002
            "cash plan ID",  # ANT-21-CSH-00001
            "status",  # Transaction successful
            "currency",
            "delivered quantity (local)",  # 999,00
            "delivered quantity (USD)",  # 235,99
            "delivery date",  # 2020-11-02 07:50:18+00
            "delivery type",  # deposit to card
            "distribution modality",  # 10K AFN per hh
            "entitlement quantity",  # 1000,00
            "TP ID",
            "TP name",
            "cash or voucher",  # if voucher or e-voucher -> voucher, else -> cash,
            "household id",  # 145aacc4-160a-493e-9d36-4f7f981284c7
        ),
        Report.PAYMENT_VERIFICATION: (
            "plan verification ID",
            "payment record ID",  # ANT-21-CSH-00001-0000002
            "plan ID",  # ANT-21-CSH-00001
            "verification completion date",
            "received amount",  # 30,00
            "status",  # RECEIVED_WITH_ISSUES
            "status date",
        ),
        Report.PAYMENT_PLAN: (
            "payment plan ID",
            "status",
            "no. of households",
            "currency",
            "total entitled quantity",
            "total delivered quantity",
            "total undelivered quantity",
            "dispersion start date",
            "dispersion end date",
        ),
        Report.CASH_PLAN: (
            "cash plan ID",  # ANT-21-CSH-00001
            "cash plan name",
            "start date",
            "end date",
            "programme",
            "funds commitment",  # 234567
            "assistance measurement",  # Euro
            "assistance through",  # Cairo Amman Bank
            "delivery type",  # DEPOSIT_TO_CARD
            "dispersion date",
            "down payment",
            "total delivered quantity",  # 220,00
            "total undelivered quantity",  # 10,00
            "total entitled quantity",  # 230,00
            "total entitled quantity revised",  # 230,00
            "total persons covered",  # 12
            "total persons covered revised",  # 12
            "status",  # DISTRIBUTION_COMPLETED_WITH_ERRORS
            "status date",
            "VISION ID",  # 2345253423
            "validation alerts count",  # 2
            # "cash plan verification status",  # FINISHED
        ),
        Report.PROGRAM: (
            "programme ID",  # e46064c4-d5e2-4990-bb9b-f5cc2dde96f9
            "name",  # Winterization 2020
            "scope",  # UNICEF
            "sector",  # EDUCATION
            "status",  # ACTIVE
            "start date",  # 2020-10-13
            "end date",  # 2020-11-17
            "cash plus",  # False
            "description",  # Description goes here
            "budget in USD",  # 10000.00
            "frequency of payments",  # REGULAR
            "administrative areas of implementation",  # Juba, Morobo, Xyz
            "individual population goal",  # 50
            "total number of households",  # 4356
        ),
        Report.INDIVIDUALS_AND_PAYMENT: (
            "household id",
            "country of origin",
            "administrative area 2",
            "first delivery date",
            "last delivery date",
            "payments made",
            "currency",
            "total delivered quantity (local)",
            "total delivered quantity (USD)",
            "birth date",  # 2000-06-24
            "estimated birth date",  # TRUE
            "gender",  # FEMALE,
            "marital status",  # MARRIED
            "disability",  # TRUE
            "observed disability",
            "communication disability",
            "hearing disability",  # LOT_DIFFICULTY
            "remembering disability",
            "physical disability",
            "seeing disability",
            "self-care disability",
            "pregnant",  # TRUE
            "relationship to hoh",  # WIFE
            "role",  # PRIMARY
            "work status",  # NOT_PROVIDED
            "sanction list possible match",  # FALSE
            "dedupe in batch status",  # UNIQUE_IN_BATCH
            "dedupe in Pop. status",  # DUPLICATE
            "dedupe in Pop.duplicates",
            "dedupe in Pop. possible duplicates",
        ),
        Report.GRIEVANCES: (
            "ticket number",
            "date created",
            "date updated",
            "status",
            "category",
            "issue type",
            "district",
            "admin2 p-code",
            "created by (username)",
            "created by (name)",
            "assigned to (username)",
            "assigned to (name)",
        ),
    }
    OPTIONAL_HEADERS = {Report.HOUSEHOLD_DEMOGRAPHICS: "programme enrolled"}
    TIMEFRAME_CELL_LABELS = {
        Report.INDIVIDUALS: ("Last Registration Date From", "Last Registration Date To"),
        Report.HOUSEHOLD_DEMOGRAPHICS: ("Last Registration Date From", "Last Registration Date To"),
        Report.CASH_PLAN_VERIFICATION: ("Completion Date From", "Completion Date To"),
        Report.PAYMENT_VERIFICATION: ("Completion Date From", "Completion Date To"),
        Report.PAYMENTS: ("Delivery Date From", "Delivery Date To"),
        Report.PAYMENT_PLAN: ("Dispersion Start Date", "Dispersion End Date"),
        Report.INDIVIDUALS_AND_PAYMENT: ("Delivery Date From", "Delivery Date To"),
        Report.CASH_PLAN: ("End Date From", "End Date To"),
        Report.PROGRAM: ("End Date From", "End Date To"),
        Report.GRIEVANCES: ("End Date From", "End Date To"),
    }
    ROW_CONTENT_METHODS = {
        Report.INDIVIDUALS: (
            GenerateReportContentHelpers.get_individuals,
            GenerateReportContentHelpers.format_individual_row,
        ),
        Report.HOUSEHOLD_DEMOGRAPHICS: (
            GenerateReportContentHelpers.get_households,
            GenerateReportContentHelpers.format_household_row,
        ),
        Report.CASH_PLAN_VERIFICATION: (
            GenerateReportContentHelpers.get_cash_plan_verifications,
            GenerateReportContentHelpers.format_cash_plan_verification_row,
        ),
        Report.PAYMENTS: (GenerateReportContentHelpers.get_payments, GenerateReportContentHelpers.format_payment_row),
        Report.PAYMENT_VERIFICATION: (
            GenerateReportContentHelpers.get_payment_verifications,
            GenerateReportContentHelpers.format_payment_verification_row,
        ),
        Report.PAYMENT_PLAN: (
            GenerateReportContentHelpers.get_payment_plans,
            GenerateReportContentHelpers.format_payment_plan_row,
        ),
        Report.CASH_PLAN: (
            GenerateReportContentHelpers.get_cash_plans,
            GenerateReportContentHelpers.format_cash_plan_row,
        ),
        Report.PROGRAM: (GenerateReportContentHelpers.get_programs, GenerateReportContentHelpers.format_program_row),
        Report.INDIVIDUALS_AND_PAYMENT: (
            GenerateReportContentHelpers.get_payments_for_individuals,
            GenerateReportContentHelpers.format_payments_for_individuals_row,
        ),
        Report.GRIEVANCES: (
            GenerateReportContentHelpers.get_grievance_tickets,
            GenerateReportContentHelpers.format_grievance_tickets_row,
        ),
    }
    FILTERS_SHEET = "Meta"
    MAX_COL_WIDTH = 75

    def __init__(self, report: Report):
        self.report = report
        self.report_type = report.report_type
        self.business_area = report.business_area

    def _create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_report = wb.active
        ws_report.title = f"{self._report_type_to_str()} Report"
        self.wb = wb
        self.ws_report = ws_report
        self.ws_filters = wb.create_sheet(GenerateReportService.FILTERS_SHEET)
        return wb

    def _add_filters_info(self):
        filter_rows = [
            ("Report type", str(self._report_type_to_str())),
            ("Business area", self.business_area.name),
            (GenerateReportService.TIMEFRAME_CELL_LABELS[self.report_type][0], str(self.report.date_from)),
            (GenerateReportService.TIMEFRAME_CELL_LABELS[self.report_type][0], str(self.report.date_to)),
        ]

        if self.report.admin_area.all().exists():
            filter_rows.append(
                (
                    "Administrative area 2",
                    GenerateReportContentHelpers._to_values_list(self.report.admin_area.all(), "name"),
                )
            )
        if self.report.program:
            filter_rows.append(("Program", self.report.program.name))

        for filter_row in filter_rows:
            self.ws_filters.append(filter_row)

    def _add_headers(self):
        headers_row = GenerateReportService.HEADERS[self.report_type]
        self.ws_report.append(headers_row)

    def _add_rows(self) -> int:
        get_row_methods = GenerateReportService.ROW_CONTENT_METHODS[self.report_type]
        all_instances = get_row_methods[0](self.report)
        self.report.number_of_records = all_instances.count()
        number_of_columns_based_on_set_headers = len(GenerateReportService.HEADERS[self.report_type])
        col_instances_len = 0
        for instance in all_instances:
            row = get_row_methods[1](instance)
            str_row = self._stringify_all_values(row)
            if len(str_row) > col_instances_len:
                col_instances_len = len(str_row)
            self.ws_report.append(str_row)
        if col_instances_len > number_of_columns_based_on_set_headers:
            # to cover bases when we create extra columns for reverse foreign key instances and we don't know in advance how many columns there will be
            self._add_missing_headers(
                self.ws_report,
                number_of_columns_based_on_set_headers + 1,
                col_instances_len,
                GenerateReportService.OPTIONAL_HEADERS.get(self.report_type, ""),
            )
        return col_instances_len

    def generate_workbook(self) -> openpyxl.Workbook:
        self._create_workbook()
        self._add_filters_info()
        self._add_headers()
        number_of_columns = self._add_rows()
        self._adjust_column_width_from_col(self.ws_filters, 1, 2, 1)
        self._adjust_column_width_from_col(self.ws_report, 1, number_of_columns, 0)
        return self.wb

    def generate_report(self):
        try:
            self.generate_workbook()
            with NamedTemporaryFile() as tmp:
                self.wb.save(tmp.name)
                tmp.seek(0)
                self.report.file.save(
                    f"{self._report_type_to_str()}-{GenerateReportContentHelpers._format_date(self.report.created_at)}.xlsx",
                    File(tmp),
                    save=False,
                )
                self.report.status = Report.COMPLETED
        except Exception as e:
            logger.exception(e)
            self.report.status = Report.FAILED
        self.report.save()

        if self.report.file:
            self._send_email()

    def _send_email(self):
        context = {
            "report_type": self._report_type_to_str(),
            "created_at": GenerateReportContentHelpers._format_date(self.report.created_at),
            "report_url": f'https://{settings.FRONTEND_HOST}/{self.business_area.slug}/reporting/{encode_id_base64(self.report.id, "Report")}',
        }
        text_body = render_to_string("report.txt", context=context)
        html_body = render_to_string("report.html", context=context)
        msg = EmailMultiAlternatives(
            subject="HOPE report generated",
            from_email=settings.EMAIL_HOST_USER,
            to=[self.report.created_by.email],
            body=text_body,
        )
        msg.attach_alternative(html_body, "text/html")
        msg.send()

    def _add_missing_headers(self, ws, column_to_start, column_to_finish, label):
        for x in range(column_to_start, column_to_finish + 1):
            col_letter = get_column_letter(x)
            ws[f"{col_letter}1"] = label

    def _adjust_column_width_from_col(self, ws, min_col, max_col, min_row):
        column_widths = []

        for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
            for cell in col:
                value = cell.value

                if value is not None:
                    if isinstance(value, str) is False:
                        value = str(value)

                    if len(value) > GenerateReportService.MAX_COL_WIDTH:
                        alignment = copy.copy(cell.alignment)
                        alignment.wrapText = True
                        cell.alignment = alignment

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i, width in enumerate(column_widths):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            value = GenerateReportService.MAX_COL_WIDTH if value > GenerateReportService.MAX_COL_WIDTH else value
            ws.column_dimensions[col_name].width = value

    def _report_type_to_str(self) -> str:
        return [name for value, name in Report.REPORT_TYPES if value == self.report_type][0]

    def _stringify_all_values(self, row: tuple) -> tuple:
        str_row = []
        for value in row:
            str_row.append(str(value if value is not None else ""))
        return tuple(str_row)
