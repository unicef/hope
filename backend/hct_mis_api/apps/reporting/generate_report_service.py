import openpyxl
from django.core.files import File
from openpyxl.utils import get_column_letter

# from openpyxl.worksheet.datavalidation import DataValidation
from tempfile import NamedTemporaryFile
from reporting.models import Report
from household.models import Individual, Household
from program.models import CashPlanPaymentVerification, CashPlan, Program
from payment.models import PaymentRecord, PaymentVerification


class GenerateReportService:
    HEADERS = {
        Report.INDIVIDUALS: (
            "admin_area_id",
            "business_area_id",
            "country",
            "document_id",  # 8e8ea94a-2ca5-4b76-b055-e098bc24eee8
            "household_country_origin",  # TM
            "birth_date",  # 2000-06-24
            "comms_disability",
            "deduplication_batch_results",  # {"duplicates": [], "possible_duplicates": []}
            "deduplication_golden_record_results",  # {"duplicates": [], "possible_duplicates": []}
            "deduplication_golden_record_status",  # UNIQUE_IN_BATCH
            "disability",
            "estimated_birth_date",  # False
            "hearing_disability",
            "marital_status",  # MARRIED
            "memory_disability",
            "observed_disability",  # NONE
            "physical_disability",
            "pregnant",  # False
            "relationship",
            "sanction_list_possible_match",  # False
            "seeing_disability",
            "selfcare_disability",
            "sex",  # MALE
            "work_status",  # NOT_PROVIDED
            "role_in_household",  # PRIMARY
        ),
        Report.HOUSEHOLD_DEMOGRAPHICS: (
            "admin_area_id",
            "business_area_id",
            "country",
            "unicef_id",  # HH-20-0000.0368
            "country_origin",  # TM
            "female_adults_count",  # 0
            "female_adults_disabled_count",  # 0
            "female_age_group_0_5_count",  # 0
            "female_age_group_0_5_disabled_count",  # 0
            "female_age_group_12_17_count",  # 0
            "female_age_group_12_17_disabled_count",  # 0
            "female_age_group_6_11_count",  # 0
            "female_age_group_6_11_disabled_count",  # 0
            "first_registration_date",  # 2020-08-25
            "geopoint",
            "last_registration_date",  # 2020-08-25
            "male_adults_count",  # 0
            "male_adults_disabled_count",  # 0
            "male_age_group_0_5_count",  # 0
            "male_age_group_0_5_disabled_count",  # 0
            "male_age_group_12_17_count",  # 0
            "male_age_group_12_17_disabled_count",  # 0
            "male_age_group_6_11_count",  # 0
            "male_age_group_6_11_disabed_count",  # 0
            "org_name_enumerator",
            "pregnant_count",
            "residence_status",  # HOST
            "returnee",  # False
            "size",  # 4
            "status",  # ACTIVE
            "village",
            "program_id",
        ),
        Report.CASH_PLAN_VERIFICATION: (
            "program_name",  # ?
            "activation_date",
            "cash_plan_id",
            "completion_date",
            "not_received_count",
            "received_count",
            "received_with_problems_count",
            "responded_count",
            "sample_size",
            "sampling",
            "sex_filter",
            "status",
            "verification_method",
        ),
        Report.PAYMENTS: (
            "business_area_id",
            "admin_area_id",
            "ca_hash_id",
            "ca_id",
            "currency",
            "delivered_quantity",
            "delivery_date",
            "delivery_type",
            "distribution_modality",
            "entitlement_quantity",
            "status",
            "target_population_cash_assist_id",
            "cash_plan_id",
            "cash_or_voucher",
        ),
        Report.PAYMENT_VERIFICATION: (
            "business_area_id",
            "program_name",
            "cash_plan_payment_verification_id",
            "payment_record_id",
            "received_amount",
            "status",
            "status_date",
        ),
        Report.CASH_PLAN: (
            "program_name",
            "assistance_measurement",
            "assistance_through",
            "business_area_id",
            "ca_hash_id",
            "delivery_type",
            "dispersion_date",
            "down_payment",
            "end_date",
            "funds_commitment",
            "name",
            "program_id",
            "start_date",
            "status",
            "status_date",
            "total_delivered_quantity",
            "total_entitled_quantity",
            "total_entitled_quantity_revised",
            "total_persons_covered",
            "total_persons_covered_revised",
            "total_undelivered_quantity",
            "validation_alerts_count",
            "verification_status",
            "vision_id",  # 54
        ),
        Report.PROGRAM: (
            "business_area_id",
            "administrative_areas_of_implementation",  # Test
            "budget",  # 10000.00
            "cash_plus",  # False
            "description",  # Description goes here
            "end_date",  # 2020-11-17
            "frequency_of_payments",  # REGULAR
            "id",  # e46064c4-d5e2-4990-bb9b-f5cc2dde96f9
            "name",  # Programme 13/10/2020 04:43:28
            "population_goal",  # 50
            "scope",  # UNICEF
            "sector",  # EDUCATION
            "start_date",  # 2020-10-13
            "status",  # ACTIVE
            "total_number_of_households",  # Payment records with delivered amount  > 0 to distinct households
        ),
        Report.INDIVIDUALS_AND_PAYMENT: (
            "admin_area_id",
            "business_area_id",
            "country",
            "program_name",
            "household_unicef_id",  # HH-20-0000.0368
            "household_country_origin",  # TM
            "birth_date",  # 2000-06-24
            "comms_disability",
            "deduplication_batch_results",  # {"duplicates": [], "possible_duplicates": []}
            "deduplication_golden_record_results",  # {"duplicates": [], "possible_duplicates": []}
            "deduplication_golden_record_status",  # UNIQUE
            "disability",
            "estimated_birth_date",  # False
            "hearing_disability",
            "marital_status",
            "memory_disability",
            "observed_disability",  # NONE
            "physical_disability",
            "pregnant",  # False
            "relationship",  # NON_BENEFICIARY
            "sanction_list_possible_match",  # False
            "seeing_disability",
            "selfcare_disability",
            "sex",  # Retrieving data. Wait a few seconds and try to cut or copy again.
            "work_status",  # NOT_PROVIDED
            "role",  # PRIMARY
            "currency",
            "delivered_quantity",  # Sum
            "delivered_quantity_usd",
        ),
    }
    FILTERS_SHEET = "Filters"
    TRUE_FALSE_MAPPING = {True: "YES", False: "NO"}
    MAX_COL_WIDTH = 50

    def __init__(self, report):
        self.report = report
        self.report_type = report.report_type
        self.business_area = report.business_area
        self.filter_vars = {"created_at__gte": report.date_from, "created_at__lte": report.date_to}

    def _report_type_to_str(self):
        return [name for value, name in Report.REPORT_TYPES if value == self.report_type][0]

    def _to_values_list(self, instances, field_name):
        values_list = list(instances.values_list(field_name, flat=True))
        return ", ".join([str(value) for value in values_list])

    def _sum_values(self, *values):
        total = 0
        for value in values:
            total = total + value if value else total
        return total

    def _stringify_all_values(self, row):
        str_row = []
        for value in row:
            str_row.append(str(value if value is not None else ""))
        return tuple(str_row)

    def _create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_report = wb.active
        ws_report.title = f"{self._report_type_to_str()} Report"
        self.wb = wb
        self.ws_report = ws_report
        self.ws_filters = wb.create_sheet(GenerateReportService.FILTERS_SHEET)
        return wb

    def _add_filters_info(self):
        filter_rows = [
            ("timeframe", f"{str(self.report.date_from)} - {str(self.report.date_to)}"),
            ("business_area", self.business_area.slug),
        ]

        if self.report.country:
            filter_rows.append(("country", str(self.report.country)))
        if self.report.admin_area:
            filter_rows.append(("admin_area", self.report.admin_area.title))
        if self.report.program:
            filter_rows.append(("program", self.program.name))

        for filter_row in filter_rows:
            self.ws_filters.append(filter_row)

    def _add_headers(self):
        headers_row = GenerateReportService.HEADERS[self.report_type]
        self.ws_report.append(headers_row)

    def _add_rows(self):
        report_rows_methods = {
            Report.INDIVIDUALS: (self._get_individuals, self._format_individual_row),
            Report.HOUSEHOLD_DEMOGRAPHICS: (self._get_households, self._format_household_row),
            Report.CASH_PLAN_VERIFICATION: (self._get_cash_plan_verifications, self._format_cash_plan_verification_row),
            Report.PAYMENTS: (self._get_payments, self._format_payment_row),
            Report.PAYMENT_VERIFICATION: (self._get_payment_verifications, self._format_payment_verification_row),
            Report.CASH_PLAN: (self._get_cash_plans, self._format_cash_plan_row),
            Report.PROGRAM: (self._get_programs, self._format_program_row),
        }
        type_methods = report_rows_methods[self.report_type]
        all_instances = type_methods[0]()
        for instance in all_instances:
            row = type_methods[1](instance)
            str_row = self._stringify_all_values(row)
            self.ws_report.append(str_row)

    # def _to_received_column(self, payment_record_verification):
    #     status = payment_record_verification.status
    #     if payment_record_verification.status == PaymentVerification.STATUS_PENDING:
    #         return None
    #     if status == PaymentVerification.STATUS_NOT_RECEIVED:
    #         return XlsxVerificationExportService.TRUE_FALSE_MAPPING[False]
    #     return XlsxVerificationExportService.TRUE_FALSE_MAPPING[True]

    def _get_individuals(self):
        self.filter_vars["business_area"] = self.business_area
        if self.report.country:
            self.filter_vars["household__country"] = self.report.country
        if self.report.admin_area:
            self.filter_vars["household__admin_area"] = self.report.admin_area
        return Individual.objects.filter(**self.filter_vars)

    def _format_individual_row(self, individual: Individual):

        return (
            individual.household.admin_area.id if individual.household and individual.household.admin_area else "",
            individual.business_area.id,
            individual.household.country if individual.household else "",
            self._to_values_list(individual.documents.all(), "id"),
            individual.household.country_origin if individual.household else "",
            individual.birth_date,
            individual.comms_disability,
            individual.deduplication_batch_results,
            individual.deduplication_golden_record_results,
            individual.deduplication_golden_record_status,
            individual.disability,
            individual.estimated_birth_date,
            individual.hearing_disability,
            individual.marital_status,
            individual.memory_disability,
            individual.observed_disability,
            individual.physical_disability,
            individual.pregnant,
            individual.relationship,
            individual.sanction_list_possible_match,
            individual.seeing_disability,
            individual.selfcare_disability,
            individual.sex,
            individual.work_status,
            self._to_values_list(individual.households_and_roles.all(), "role"),
        )

    def _get_households(self):
        self.filter_vars["business_area"] = self.business_area
        if self.report.country:
            self.filter_vars["country"] = self.report.country
        if self.report.admin_area:
            self.filter_vars["admin_area"] = self.report.admin_area
        return Household.objects.filter(**self.filter_vars)

    def _format_household_row(self, household: Household):
        return (
            household.admin_area.id if household.admin_area else "",
            household.business_area.id,
            household.country,
            household.unicef_id,
            household.country_origin,
            # TODO: check if adults_count should be a sum of these two fields
            self._sum_values(household.female_age_group_18_59_count, household.female_age_group_60_count),
            self._sum_values(
                household.female_age_group_18_59_disabled_count, household.female_age_group_18_59_disabled_count
            ),
            household.female_age_group_0_5_count,
            household.female_age_group_0_5_disabled_count,
            household.female_age_group_12_17_count,
            household.female_age_group_12_17_disabled_count,
            household.female_age_group_6_11_count,
            household.female_age_group_6_11_disabled_count,
            household.first_registration_date,
            household.geopoint,
            household.last_registration_date,
            self._sum_values(household.male_age_group_18_59_count, household.male_age_group_60_count),
            self._sum_values(household.male_age_group_18_59_disabled_count, household.male_age_group_60_disabled_count),
            household.male_age_group_0_5_count,
            household.male_age_group_0_5_disabled_count,
            household.male_age_group_12_17_count,
            household.male_age_group_12_17_disabled_count,
            household.male_age_group_6_11_count,
            household.male_age_group_6_11_disabled_count,
            household.org_name_enumerator,
            household.pregnant_count,
            household.residence_status,
            household.returnee,
            household.size,
            household.status,
            household.village,
            self._to_values_list(household.programs.all(), "id"),
        )

    def _get_cash_plan_verifications(self):
        self.filter_vars["cash_plan__business_area"] = self.business_area
        if self.report.program:
            self.filter_vars["cash_plan__program"] = self.report.program
        return CashPlanPaymentVerification.objects.filter(**self.filter_vars)

    def _format_cash_plan_verification_row(self, verification: CashPlanPaymentVerification):
        return (
            verification.cash_plan.program.name,
            verification.activation_date,
            verification.cash_plan.id,
            verification.completion_date,
            verification.not_received_count,
            verification.received_count,
            verification.received_with_problems_count,
            verification.responded_count,
            verification.sample_size,
            verification.sampling,
            verification.sex_filter,
            verification.status,
            verification.verification_method,
        )

    def _get_payments(self):
        self.filter_vars["business_area"] = self.business_area
        if self.report.admin_area:
            self.filter_vars["household__admin_area"] = self.report.admin_area
        return PaymentRecord.objects.filter(**self.filter_vars)

    def _format_payment_row(self, payment: PaymentRecord):
        return (
            payment.business_area.id,
            payment.household.admin_area.id if payment.household.admin_area else "",
            payment.ca_hash_id,
            payment.ca_id,
            payment.currency,
            payment.delivered_quantity,
            payment.delivery_date,
            payment.delivery_type,
            payment.distribution_modality,
            payment.entitlement_quantity,
            payment.status,
            payment.target_population_cash_assist_id,
            payment.cash_plan.id if payment.cash_plan else "",
        )

    def _get_payment_verifications(self):
        self.filter_vars["cash_plan_payment_verification__cash_plan__business_area"] = self.business_area
        if self.report.program:
            self.filter_vars["cash_plan_payment_verification__cash_plan__program"] = self.report.program
        return PaymentVerification.objects.filter(**self.filter_vars)

    def _format_payment_verification_row(self, payment_verification: PaymentVerification):
        return (
            payment_verification.cash_plan_payment_verification.cash_plan.business_area.id,
            payment_verification.cash_plan_payment_verification.cash_plan.program.name,
            payment_verification.cash_plan_payment_verification.id,
            payment_verification.payment_record.id,
            payment_verification.received_amount,
            payment_verification.status,
            payment_verification.status_date,
        )

    def _get_cash_plans(self):
        self.filter_vars["business_area"] = self.business_area
        if self.report.program:
            self.filter_vars["cash_plan__program"] = self.report.program
        return CashPlan.objects.filter(**self.filter_vars)

    def _format_cash_plan_row(self, cash_plan: CashPlan):
        return (
            cash_plan.program.name,
            cash_plan.assistance_measurement,
            cash_plan.assistance_through,
            cash_plan.business_area.id,
            cash_plan.ca_hash_id,
            cash_plan.delivery_type,
            cash_plan.dispersion_date,
            cash_plan.down_payment,
            cash_plan.end_date,
            cash_plan.funds_commitment,
            cash_plan.name,
            cash_plan.program.id,
            cash_plan.start_date,
            cash_plan.status,
            cash_plan.status_date,
            cash_plan.total_delivered_quantity,
            cash_plan.total_entitled_quantity,
            cash_plan.total_entitled_quantity_revised,
            cash_plan.total_persons_covered,
            cash_plan.total_persons_covered_revised,
            cash_plan.total_undelivered_quantity,
            cash_plan.validation_alerts_count,
            cash_plan.verification_status,
            cash_plan.vision_id,
        )

    def _get_programs(self):
        self.filter_vars["business_area"] = self.business_area
        return Program.objects.filter(**self.filter_vars)

    def _format_program_row(self, program: Program):
        return (
            program.business_area.id,
            program.administrative_areas_of_implementation,
            program.budget,
            program.cash_plus,
            program.description,
            program.end_date,
            program.frequency_of_payments,
            program.id,
            program.name,
            program.population_goal,
            program.scope,
            program.sector,
            program.start_date,
            program.status,
            program.total_number_of_households,
        )

    # def _add_data_validation(self):
    #     self.dv_received = DataValidation(type="list", formula1=f'"YES,NO"', allow_blank=False)
    #     self.dv_received.add(f"B2:B{len(self.ws_verifications['B'])}")
    #     self.ws_verifications.add_data_validation(self.dv_received)
    #     cell_range = self.ws_verifications["B2":f"B{len(self.ws_verifications['B'])}"]

    def generate_workbook(self):
        self._create_workbook()
        self._add_filters_info()
        self._add_headers()
        self._add_rows()
        self._adjust_column_width_from_col(self.ws_filters, 1, 2, 1)
        self._adjust_column_width_from_col(self.ws_report, 1, len(GenerateReportService.HEADERS[self.report_type]), 0)
        return self.wb

    def generate_file(self):
        try:
            self.generate_workbook()
            with NamedTemporaryFile() as tmp:
                self.wb.save(tmp.name)
                tmp.seek(0)
                self.report.file.save(
                    f"Report:_{self._report_type_to_str()}_{str(self.report.created_at)}.xlsx", File(tmp)
                )
                self.report.status = Report.COMPLETED
        except Exception as e:
            print("ERROR", e)
            self.report.status = Report.FAILED
        self.report.save()

    def _adjust_column_width_from_col(self, ws, min_col, max_col, min_row):

        column_widths = []

        for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):

            for cell in col:
                value = cell.value
                if value is not None:

                    if isinstance(value, str) is False:
                        value = str(value)

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i, width in enumerate(column_widths):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            value = GenerateReportService.MAX_COL_WIDTH if value > GenerateReportService.MAX_COL_WIDTH else value
            ws.column_dimensions[col_name].width = value
