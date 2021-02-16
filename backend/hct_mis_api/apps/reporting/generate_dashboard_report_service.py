import openpyxl
import copy
import functools
from django.core.files import File
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.db.models import Sum, Count, F
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from tempfile import NamedTemporaryFile

from hct_mis_api.apps.core.models import BusinessArea
from hct_mis_api.apps.reporting.models import DashboardReport
from hct_mis_api.apps.household.models import Household
from hct_mis_api.apps.program.models import Program
from hct_mis_api.apps.payment.models import PaymentRecord
from hct_mis_api.apps.account.models import User


class GenerateDashboardReportContentHelpers:
    @staticmethod
    def _to_values_list(instances, field_name: str) -> str:
        values_list = list(instances.values_list(field_name, flat=True))
        return ", ".join([str(value) for value in values_list])

    @staticmethod
    def _format_date(date) -> str:
        if not date:
            return ""
        return date.strftime("%Y-%m-%d")

    @staticmethod
    def _is_report_global(report: DashboardReport):
        return report.business_area.slug == "global"

    @classmethod
    def _format_filters_for_payment_records(self, report: DashboardReport):
        filter_vars = {
            "delivery_date__year": report.year,
            "delivered_quantity__gt": 0,
        }
        if report.admin_area:
            filter_vars["household__admin_area"] = report.admin_area
            filter_vars["household__admin_area_level__admin_level"] = 2
        if report.program:
            filter_vars["cash_plan__program"] = report.program
        if not self._is_report_global(report):
            filter_vars["business_area"] = report.business_area

        return filter_vars

    @classmethod
    def get_beneficiaries(self, report: DashboardReport):
        filter_vars = self._format_filters_for_payment_records(report)

        tot_individual_count_fields = [
            "total_female_0_5",
            "total_female_6_11",
            "total_female_12_17",
            "total_female_18_59",
            "total_female_60",
            "total_male_0_5",
            "total_male_6_11",
            "total_male_12_17",
            "total_male_18_59",
            "total_male_60",
        ]
        tot_children_count_fields = [
            "total_female_0_5",
            "total_female_6_11",
            "total_female_12_17",
            "total_male_0_5",
            "total_male_6_11",
            "total_male_12_17",
        ]

        valid_payment_records = PaymentRecord.objects.filter(**filter_vars)
        instances = None
        valid_payment_records_in_instance_filter_key = None

        if self._is_report_global(report):
            instances = (
                BusinessArea.objects.filter(paymentrecord__in=valid_payment_records)
                .annotate(business_area_code=F("code"))
                .annotate(num_households=Count("paymentrecord__household", distinct=True))
                .values("name", "id", "num_households", "business_area_code")
            )
            valid_payment_records_in_instance_filter_key = "business_area"
        else:
            instances = (
                Program.objects.filter(cash_plans__payment_records__in=valid_payment_records)
                .annotate(num_households=Count("cash_plans__payment_records__household", distinct=True))
                .annotate(business_area_code=F("business_area__code"))
                .values("id", "name", "num_households", "business_area_code")
            )
            valid_payment_records_in_instance_filter_key = "cash_plan__program"

        def aggregate_households(households):
            return households.aggregate(
                total_female_0_5=Sum("female_age_group_0_5_count"),
                total_female_6_11=Sum("female_age_group_6_11_count"),
                total_female_12_17=Sum("female_age_group_12_17_count"),
                total_female_18_59=Sum("female_age_group_18_59_count"),
                total_female_60=Sum("female_age_group_60_count"),
                total_male_0_5=Sum("male_age_group_0_5_count"),
                total_male_6_11=Sum("male_age_group_6_11_count"),
                total_male_12_17=Sum("male_age_group_12_17_count"),
                total_male_18_59=Sum("male_age_group_18_59_count"),
                total_male_60=Sum("male_age_group_60_count"),
            )

        for instance in instances:
            valid_payment_records_in_instance = valid_payment_records.filter(
                **{valid_payment_records_in_instance_filter_key: instance["id"]}
            )
            households = aggregate_households(
                Household.objects.filter(payment_records__in=valid_payment_records_in_instance).distinct()
            )
            instance["total_children"] = functools.reduce(
                lambda a, b: a + households[b] if households[b] else a, tot_children_count_fields, 0
            )
            instance["total_individuals"] = functools.reduce(
                lambda a, b: a + households[b] if households[b] else a, tot_individual_count_fields, 0
            )
        # get total distincts (can't use the sum of column since some households might belong to multiple programs)
        households = Household.objects.filter(payment_records__in=valid_payment_records).distinct()
        total_households = households.count()
        households_aggr = aggregate_households(households)
        # return instances for rows and totals row info
        return instances, {
            "num_households": total_households,
            "total_individuals": functools.reduce(
                lambda a, b: a + households_aggr[b] if households_aggr[b] else a, tot_individual_count_fields, 0
            ),
            "total_children": functools.reduce(
                lambda a, b: a + households_aggr[b] if households_aggr[b] else a, tot_children_count_fields, 0
            ),
        }

    @staticmethod
    def format_beneficiaries_row(instance: dict, is_totals: bool) -> tuple:
        return (
            instance.get("business_area_code", "") if not is_totals else "",
            instance.get("name", "") if not is_totals else "Total Distinct",
            instance.get("num_households", ""),
            instance.get("total_individuals", ""),
            instance.get("total_children", ""),
        )

    @classmethod
    def get_individuals(self, report: DashboardReport):
        filter_vars = self._format_filters_for_payment_records(report)

        valid_payment_records = PaymentRecord.objects.filter(**filter_vars)
        instances = None
        valid_payment_records_in_instance_filter_key = None

        if self._is_report_global(report):
            instances = (
                BusinessArea.objects.filter(paymentrecord__in=valid_payment_records)
                .distinct()
                .annotate(business_area_code=F("code"))
                .values("name", "id", "business_area_code")
            )
            valid_payment_records_in_instance_filter_key = "business_area"
        else:
            instances = (
                Program.objects.filter(cash_plans__payment_records__in=valid_payment_records)
                .distinct()
                .annotate(business_area_code=F("business_area__code"))
                .values("id", "name", "business_area_code")
            )
            valid_payment_records_in_instance_filter_key = "cash_plan__program"

        def aggregate_households(households):
            return households.aggregate(
                total_female_0_5=Sum("female_age_group_0_5_count"),
                total_female_0_5_disabled=Sum("female_age_group_0_5_disabled_count"),
                total_female_6_11=Sum("female_age_group_6_11_count"),
                total_female_6_11_disabled=Sum("female_age_group_6_11_disabled_count"),
                total_female_12_17=Sum("female_age_group_12_17_count"),
                total_female_12_17_disabled=Sum("female_age_group_12_17_disabled_count"),
                total_female_18_59=Sum("female_age_group_18_59_count"),
                total_female_18_59_disabled=Sum("female_age_group_18_59_disabled_count"),
                total_female_60=Sum("female_age_group_60_count"),
                total_female_60_disabled=Sum("female_age_group_60_disabled_count"),
                total_male_0_5=Sum("male_age_group_0_5_count"),
                total_male_0_5_disabled=Sum("male_age_group_0_5_disabled_count"),
                total_male_6_11=Sum("male_age_group_6_11_count"),
                total_male_6_11_disabled=Sum("male_age_group_6_11_disabled_count"),
                total_male_12_17=Sum("male_age_group_12_17_count"),
                total_male_12_17_disabled=Sum("male_age_group_12_17_disabled_count"),
                total_male_18_59=Sum("male_age_group_18_59_count"),
                total_male_18_59_disabled=Sum("male_age_group_18_59_disabled_count"),
                total_male_60=Sum("male_age_group_60_count"),
                total_male_60_disabled=Sum("male_age_group_60_disabled_count"),
            )

        for instance in instances:
            valid_payment_records_in_instance = valid_payment_records.filter(
                **{valid_payment_records_in_instance_filter_key: instance["id"]}
            )
            households = aggregate_households(
                Household.objects.filter(payment_records__in=valid_payment_records_in_instance).distinct()
            )

            for key, value in households.items():
                instance[key] = value

        # get total distincts (can't use the sum of column since some households might belong to multiple programs)
        households_aggr = aggregate_households(
            Household.objects.filter(payment_records__in=valid_payment_records).distinct()
        )
        # return instances for rows and totals row info
        return instances, households_aggr

    @staticmethod
    def format_individuals_row(instance: dict, is_totals: bool) -> tuple:
        return (
            instance.get("business_area_code", "") if not is_totals else "",
            instance.get("name", "") if not is_totals else "Total Distinct",
            instance.get("total_female_0_5", ""),
            instance.get("total_female_0_5_disabled", ""),
            instance.get("total_female_6_11", ""),
            instance.get("total_female_6_11_disabled", ""),
            instance.get("total_female_12_17", ""),
            instance.get("total_female_12_17_disabled", ""),
            instance.get("total_female_18_59", ""),
            instance.get("total_female_18_59_disabled", ""),
            instance.get("total_female_60", ""),
            instance.get("total_female_60_disabled", ""),
            instance.get("total_male_0_5", ""),
            instance.get("total_male_0_5_disabled", ""),
            instance.get("total_male_6_11", ""),
            instance.get("total_male_6_11_disabled", ""),
            instance.get("total_male_12_17", ""),
            instance.get("total_male_12_17_disabled", ""),
            instance.get("total_male_18_59", ""),
            instance.get("total_male_18_59_disabled", ""),
            instance.get("total_male_60", ""),
            instance.get("total_male_60_disabled", ""),
        )

    @classmethod
    def get_volumes_by_delivery(self, report: DashboardReport):
        filter_vars = self._format_filters_for_payment_records(report)

        valid_payment_records = PaymentRecord.objects.filter(**filter_vars)
        instances = None
        valid_payment_records_in_instance_filter_key = None
        if self._is_report_global(report):
            instances = (
                BusinessArea.objects.filter(paymentrecord__in=valid_payment_records)
                .distinct()
                .annotate(business_area_code=F("code"))
                .values("name", "id", "business_area_code")
            )
            valid_payment_records_in_instance_filter_key = "business_area"
        else:
            instances = (
                Program.objects.filter(cash_plans__payment_records__in=valid_payment_records)
                .distinct()
                .annotate(business_area_code=F("business_area__code"))
                .values("id", "name", "business_area_code")
            )
            valid_payment_records_in_instance_filter_key = "cash_plan__program"

        def aggregate_by_delivery_type(payment_records):
            result = dict()
            for delivery_type in PaymentRecord.DELIVERY_TYPE_CHOICE:
                value = delivery_type[0]
                result[value] = (
                    payment_records.filter(delivery_type=value)
                    .aggregate(Sum("delivered_quantity_usd"))
                    .get("delivered_quantity_usd__sum")
                )
            return result

        for instance in instances:
            valid_payment_records_in_instance = valid_payment_records.filter(
                **{valid_payment_records_in_instance_filter_key: instance["id"]}
            )
            aggregated_by_delivery_type = aggregate_by_delivery_type(valid_payment_records_in_instance)
            instance.update(aggregated_by_delivery_type)

        totals = aggregate_by_delivery_type(valid_payment_records)
        return instances, totals

    @staticmethod
    def format_volumes_by_delivery_row(instance: dict, is_totals: bool):
        result = [
            instance.get("business_area_code", "") if not is_totals else "",
            instance.get("name", "") if not is_totals else "Total",
        ]
        for choice in PaymentRecord.DELIVERY_TYPE_CHOICE:
            result.append(instance.get(choice[0]))

        return tuple(result)


class GenerateDashboardReportService:
    HQ = 1
    COUNTRY = 2
    SHARED = 3
    HEADERS = {
        DashboardReport.BENEFICIARIES_REACHED: {
            HQ: ("business area", "country"),
            COUNTRY: ("business area", "programme"),
            SHARED: ("households reached", "individuals reached", "children reached"),
        },
        DashboardReport.TOTAL_TRANSFERRED_BY_ADMIN_AREA: {
            HQ: (),
            COUNTRY: (
                "Admin Level 2",
                "Admin Code",
                "Total tranferred (USD)",
                "Households reached",
                "Female 0-5 Reached",
                "Female 6-11 Reached",
                "Female 12-17 Reached",
                "Female 18-59 Reached",
                "Female 60+ Reached",
                "Male 0-5 Reached",
                "Male 6-11 Reached",
                "Male 12-17 Reached",
                "Male 18-59 Reached",
                "Male 60+ Reached",
            ),
            SHARED: (),
        },
        DashboardReport.PAYMENT_VERIFICATION: {
            HQ: (),
            COUNTRY: (),
            SHARED: (
                "business area",
                "country",
                "programme",
                "cash plan verifications",
                "Households Contacted",
                "average sampling",
                "Received",
                "Not Received",
                "Received with issues",
                "Not Responded",
            ),
        },
        DashboardReport.GRIEVANCES_AND_FEEDBACK: {
            HQ: (
                "business area",
                "country",
            ),
            COUNTRY: (
                "business area",
                "programme",
            ),
            SHARED: (
                "grievance tickets",
                "feedback tickets",
                "resolved tickets",
                "Unresolved >30 days",
                "Unresolved >60 days",
                "open sensitive grievances",
            ),
        },
        DashboardReport.TOTAL_TRANSFERRED_BY_COUNTRY: {
            HQ: ("business area", "country", "actual cash transferred", "actual voucher transferred"),
            COUNTRY: (),
            SHARED: (),
        },
        DashboardReport.PROGRAMS: {
            HQ: (),
            COUNTRY: (),
            SHARED: (
                "business area",
                "country",
                "programme",
                "sector",
                "cash+",
                "frequency",
                "unsuccessful payment",
                "January cash",
                "January voucher",
                "February cash",
                "February voucher",
                "March cash",
                "March voucher",
                "April cash",
                "April voucher",
                "May cash",
                "May voucher",
                "June cash",
                "June voucher",
                "July cash",
                "July voucher",
                "August cash",
                "August voucher",
                "September cash",
                "September voucher",
                "October cash",
                "October voucher",
                "November cash",
                "November voucher",
                "December cash",
                "December voucher",
            ),
        },
        DashboardReport.VOLUME_BY_DELIVERY_MECHANISM: {
            HQ: (
                "business area",
                "country",
            ),
            COUNTRY: (
                "business area",
                "programme",
            ),
            SHARED: tuple([choice[1] for choice in PaymentRecord.DELIVERY_TYPE_CHOICE]),
        },
        DashboardReport.INDIVIDUALS_REACHED: {
            HQ: (
                "business area",
                "country",
            ),
            COUNTRY: (
                "business area",
                "programme",
            ),
            SHARED: (
                "females 0-5 reached",
                "females 0-5 w/ disability reached",
                "females 6-11 reached",
                "females 6-11 w/ disability reached",
                "females 12-17 reached",
                "females 12-17 w/ disability reached",
                "females 18-59 reached",
                "females 18-59 w/ disability reached",
                "females 60+ reached",
                "females 60+ w/ disability reached",
                "males 0-5 reached",
                "males 0-5 w/ disability reached",
                "males 6-11 reached",
                "males 6-11 w/ disability reached",
                "males 12-17 reached",
                "males 12-17 w/ disability reached",
                "males 18-59 reached",
                "males 18-59 w/ disability reached",
                "males 60+ reached",
                "males 60+ w/ disability reached",
            ),
        },
    }
    ROW_CONTENT_METHODS = {
        DashboardReport.BENEFICIARIES_REACHED: (
            GenerateDashboardReportContentHelpers.get_beneficiaries,
            GenerateDashboardReportContentHelpers.format_beneficiaries_row,
        ),
        DashboardReport.INDIVIDUALS_REACHED: (
            GenerateDashboardReportContentHelpers.get_individuals,
            GenerateDashboardReportContentHelpers.format_individuals_row,
        ),
        DashboardReport.VOLUME_BY_DELIVERY_MECHANISM: (
            GenerateDashboardReportContentHelpers.get_volumes_by_delivery,
            GenerateDashboardReportContentHelpers.format_volumes_by_delivery_row,
        )
        # TODO: add the rest of the methods
    }
    META_HEADERS = ("report type", "creation date", "created by", "business area", "report year")
    META_SHEET = "Meta data"
    MAX_COL_WIDTH = 75

    def __init__(self, report: DashboardReport):
        self.report = report
        self.report_types = report.report_type
        self.business_area = report.business_area
        self.hq_or_country = self.HQ if report.business_area.slug == "global" else self.COUNTRY

    def _create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_meta = wb.active
        ws_meta.title = self.META_SHEET
        self.wb = wb
        self.ws_meta = ws_meta
        return wb

    def _format_meta_tab(self):
        self.ws_meta.append(self.META_HEADERS)
        info_row = (
            self._report_types_to_joined_str(),
            self._format_date(self.report.created_at),
            self._format_user_name(self.report.created_by),
            self.business_area.name,
            str(self.report.year),
        )
        self.ws_meta.append(info_row)

    def _add_headers(self, active_sheet, report_type) -> int:
        headers_row = self.HEADERS[report_type][self.hq_or_country] + self.HEADERS[report_type][self.SHARED]
        headers_row = self._stringify_all_values(headers_row)
        active_sheet.append(headers_row)
        return len(headers_row)

    # def _append_totals_row(self, active_sheet, report_type, totals):
    #     label = "Total"
    #     if report_type in [DashboardReport.BENEFICIARIES_REACHED, DashboardReport.INDIVIDUALS_REACHED]:
    #         label = "Total distinct"
    #     totals_row = ("", label) + totals
    #     str_totals_row = self._stringify_all_values(totals_row)
    #     active_sheet.append(str_totals_row)

    def _add_rows(self, active_sheet, report_type):
        get_row_methods = self.ROW_CONTENT_METHODS[report_type]
        all_instances, totals = get_row_methods[0](self.report)
        for instance in all_instances:
            row = get_row_methods[1](instance, False)
            str_row = self._stringify_all_values(row)
            active_sheet.append(str_row)
        # append totals row
        print("RIGHT BEFORE TOTALS")
        row = get_row_methods[1](totals, True)
        str_row = self._stringify_all_values(row)
        active_sheet.append(str_row)
        return len(all_instances)

    def generate_workbook(self) -> openpyxl.Workbook:
        self._create_workbook()
        self._format_meta_tab()
        self._add_font_style_to_sheet(self.ws_meta)
        self._adjust_column_width_from_col(self.ws_meta, 1, 5, 1)

        # loop through all selected report types and add sheet for each
        for report_type in self.report_types:
            print("IN FOR LOOP", report_type)
            sheet_title = self._report_type_to_str(report_type)
            print("SHEET TITLE", sheet_title)
            active_sheet = self.wb.create_sheet(sheet_title, -1)
            print("CREATED ACTIVE SHEET")
            number_of_columns = self._add_headers(active_sheet, report_type)
            print("ADDED HEADERS")
            number_of_rows = self._add_rows(active_sheet, report_type)
            print("ADDED ROWS")
            self._adjust_column_width_from_col(active_sheet, 1, number_of_columns, 1)
            self._add_font_style_to_sheet(active_sheet, number_of_rows + 2)
            print("ADJUTED WIDTH")
        return self.wb

    def generate_report(self):
        try:
            self.generate_workbook()
            print(self.wb)
            with NamedTemporaryFile() as tmp:
                self.wb.save(tmp.name)
                tmp.seek(0)
                self.report.file.save(
                    f"{self._report_types_to_joined_str()}-{GenerateDashboardReportContentHelpers._format_date(self.report.created_at)}.xlsx",
                    File(tmp),
                    save=False,
                )
                self.report.status = DashboardReport.COMPLETED
        except Exception as e:
            print("ERROR", e)
            self.report.status = DashboardReport.FAILED
        self.report.save()

        if self.report.file:
            self._send_email()

    def _send_email(self):
        pass
        # context = {
        #     "report_type": self._report_types_to_joined_str(),
        #     "created_at": self._format_date(self.report.created_at),
        #     "report_url": self.report.file.url,
        # }
        # text_body = render_to_string("dashboard_report.txt", context=context)
        # html_body = render_to_string("dashboard_report.html", context=context)
        # msg = EmailMultiAlternatives(
        #     subject="HOPE report generated",
        #     from_email=settings.EMAIL_HOST_USER,
        #     to=[self.report.created_by.email],
        #     body=text_body,
        # )
        # msg.attach_alternative(html_body, "text/html")
        # msg.send()

    def _adjust_column_width_from_col(self, ws, min_col, max_col, min_row):

        column_widths = []

        for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):

            for cell in col:
                value = cell.value

                if value is not None:

                    if isinstance(value, str) is False:
                        value = str(value)

                    if len(value) > GenerateDashboardReportService.MAX_COL_WIDTH:
                        alignment = copy.copy(cell.alignment)
                        alignment.wrapText = True
                        cell.alignment = alignment

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i, width in enumerate(column_widths):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            value = (
                GenerateDashboardReportService.MAX_COL_WIDTH
                if value > GenerateDashboardReportService.MAX_COL_WIDTH
                else value
            )
            ws.column_dimensions[col_name].width = value

    def _add_font_style_to_sheet(self, ws, totals_row=None):
        bold_font = Font(bold=True)
        for cell in ws["1:1"]:
            cell.font = bold_font
        if totals_row:
            ws[f"B{totals_row}"].font = bold_font

    def _report_type_to_str(self, report_type) -> str:
        return str([name for value, name in DashboardReport.REPORT_TYPES if value == report_type][0])

    def _report_types_to_joined_str(self) -> str:
        return ", ".join([self._report_type_to_str(report_type) for report_type in self.report_types])

    def _stringify_all_values(self, row: tuple) -> tuple:
        str_row = []
        for value in row:
            str_row.append(str(value if value is not None else ""))
        return tuple(str_row)

    def _format_date(self, date) -> str:
        if not date:
            return ""
        return date.strftime("%Y-%m-%d")

    def _format_user_name(self, user: User) -> str:
        return (
            f"{user.first_name} {user.last_name}" if user.first_name or user.last_name else user.email or user.username
        )
