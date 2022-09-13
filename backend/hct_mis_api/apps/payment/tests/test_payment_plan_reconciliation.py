import os
import tempfile

from zipfile import ZipFile
from datetime import timedelta
from unittest.mock import patch
from openpyxl import load_workbook

from django.utils import timezone
from django.core.files.uploadedfile import SimpleUploadedFile

from hct_mis_api.apps.payment.celery_tasks import (
    create_payment_plan_payment_list_xlsx_per_fsp,
    import_payment_plan_payment_list_per_fsp_from_xlsx,
)
from hct_mis_api.apps.payment.models import PaymentPlan
from hct_mis_api.apps.core.utils import decode_id_string
from hct_mis_api.apps.payment.fixtures import PaymentFactory
from hct_mis_api.apps.account.fixtures import UserFactory
from hct_mis_api.apps.account.permissions import Permissions
from hct_mis_api.apps.core.base_test_case import APITestCase
from hct_mis_api.apps.core.fixtures import create_afghanistan
from hct_mis_api.apps.core.models import BusinessArea
from hct_mis_api.apps.core.utils import encode_id_base64
from hct_mis_api.apps.household.fixtures import create_household_and_individuals, IndividualRoleInHouseholdFactory
from hct_mis_api.apps.household.models import ROLE_PRIMARY
from hct_mis_api.apps.payment.celery_tasks import payment_plan_apply_steficon
from hct_mis_api.apps.payment.fixtures import FinancialServiceProviderFactory, PaymentChannelFactory
from hct_mis_api.apps.payment.models import GenericPayment
from hct_mis_api.apps.registration_data.fixtures import RegistrationDataImportFactory
from hct_mis_api.apps.steficon.fixtures import RuleFactory, RuleCommitFactory


CREATE_PROGRAMME_MUTATION = """
mutation CreateProgram($programData: CreateProgramInput!) {
  createProgram(programData: $programData) {
    program {
      id
    }
  }
}
"""

UPDATE_PROGRAM_MUTATION = """
mutation UpdateProgram($programData: UpdateProgramInput!) {
  updateProgram(programData: $programData) {
    program {
      id
    }
  }
}
"""


CREATE_TARGET_POPULATION_MUTATION = """
mutation CreateTP($input: CreateTargetPopulationInput!) {
  createTargetPopulation(input: $input) {
    targetPopulation {
      id
      status
    }
  }
}
"""


CREATE_PAYMENT_PLAN_MUTATION = """
mutation CreatePaymentPlan($input: CreatePaymentPlanInput!) {
    createPaymentPlan(input: $input) {
        paymentPlan {
            id
        }
    }
}
"""

LOCK_TARGET_POPULATION_MUTATION = """
mutation LockTP($id: ID!) {
    lockTargetPopulation(id: $id) {
        targetPopulation {
            id
        }
    }
}
"""

FINALIZE_TARGET_POPULATION_MUTATION = """
mutation FinalizeTP($id: ID!) {
    finalizeTargetPopulation(id: $id) {
        targetPopulation {
            id
            status
        }
    }
}
"""

SET_STEFICON_RULE_MUTATION = """
mutation SetSteficonRuleOnPaymentPlanPaymentList($paymentPlanId: ID!, $steficonRuleId: ID!) {
    setSteficonRuleOnPaymentPlanPaymentList(paymentPlanId: $paymentPlanId, steficonRuleId: $steficonRuleId) {
        paymentPlan {
            id
        }
    }
}
"""


PAYMENT_PLAN_ACTION_MUTATION = """
mutation ActionPaymentPlanMutation($input: ActionPaymentPlanInput!) {
    actionPaymentPlanMutation(input: $input) {
        paymentPlan {
            status
            id
        }
    }
}"""

CHOOSE_DELIVERY_MECHANISMS_MUTATION = """
mutation ChooseDeliveryMechanismsForPaymentPlan($input: ChooseDeliveryMechanismsForPaymentPlanInput!) {
    chooseDeliveryMechanismsForPaymentPlan(input: $input) {
        paymentPlan {
            id
            deliveryMechanisms {
                order
                name
            }
        }
    }
}
"""

AVAILABLE_FSPS_FOR_DELIVERY_MECHANISMS_QUERY = """
query PaymentPlan($paymentPlanId: ID!, $fspChoices: [FspSelection!]!) {
    paymentPlan(id: $paymentPlanId) {
        id
        availableFspsForDeliveryMechanisms(fspChoices: $fspChoices) {
            deliveryMechanism
            fsps {
                id
                name
            }
        }
    }
}
"""

ASSIGN_FSPS_MUTATION = """
mutation AssignFspToDeliveryMechanism($paymentPlanId: ID!, $mappings: [FSPToDeliveryMechanismMappingInput!]!) {
    assignFspToDeliveryMechanism(input: {
        paymentPlanId: $paymentPlanId,
        mappings: $mappings
    }) {
        paymentPlan {
            id
            deliveryMechanisms {
                name
                order
                fsp {
                    id
                }
            }
        }
    }
}
"""

EXPORT_XLSX_PER_FSP_MUTATION = """
mutation ExportXlsxPaymentPlanPaymentListPerFsp($paymentPlanId: ID!) {
    exportXlsxPaymentPlanPaymentListPerFsp(paymentPlanId: $paymentPlanId) {
        paymentPlan {
            id
        }
    }
}
"""

IMPORT_XLSX_PER_FSP_MUTATION = """
mutation ImportXlsxPaymentPlanPaymentListPerFsp($paymentPlanId: ID!, $file: Upload!) {
    importXlsxPaymentPlanPaymentListPerFsp(paymentPlanId: $paymentPlanId, file: $file) {
        paymentPlan {
            id
        }
    }
}
"""


class TestPaymentPlanReconciliation(APITestCase):
    @classmethod
    def create_household_and_individual(cls):
        household, individuals = create_household_and_individuals(
            household_data={
                "registration_data_import": cls.registration_data_import,
                "business_area": cls.business_area,
            },
            individuals_data=[{}],
        )
        IndividualRoleInHouseholdFactory(household=household, individual=individuals[0], role=ROLE_PRIMARY)
        return household, individuals[0]

    @classmethod
    def setUpTestData(cls):
        create_afghanistan(
            is_payment_plan_applicable=True,
            approval_number_required=1,
            authorization_number_required=1,
            finance_review_number_required=1,
        )
        cls.business_area = BusinessArea.objects.get(slug="afghanistan")
        cls.user = UserFactory.create()
        cls.all_necessary_permissions = [
            Permissions.PAYMENT_MODULE_CREATE,
            Permissions.PAYMENT_MODULE_VIEW_DETAILS,
            Permissions.PAYMENT_MODULE_VIEW_LIST,
            Permissions.PROGRAMME_CREATE,
            Permissions.PROGRAMME_ACTIVATE,
            Permissions.TARGETING_CREATE,
            Permissions.TARGETING_LOCK,
            Permissions.TARGETING_SEND,
        ]
        cls.create_user_role_with_permissions(
            cls.user,
            cls.all_necessary_permissions,
            cls.business_area,
        )

        cls.registration_data_import = RegistrationDataImportFactory(business_area=cls.business_area)

        cls.household_1, cls.individual_1 = cls.create_household_and_individual()
        cls.payment_channel_1_cash = PaymentChannelFactory(
            individual=cls.individual_1,
            delivery_mechanism=GenericPayment.DELIVERY_TYPE_CASH,
        )

        cls.household_2, cls.individual_2 = cls.create_household_and_individual()
        cls.payment_channel_2_cash = PaymentChannelFactory(
            individual=cls.individual_2,
            delivery_mechanism=GenericPayment.DELIVERY_TYPE_CASH,
        )

        cls.household_3, cls.individual_3 = cls.create_household_and_individual()
        cls.payment_channel_3_cash = PaymentChannelFactory(
            individual=cls.individual_3,
            delivery_mechanism=GenericPayment.DELIVERY_TYPE_CASH,
        )

    @patch("hct_mis_api.apps.payment.models.PaymentPlan.get_exchange_rate", return_value=2.0)
    def test_receiving_reconciliations_from_fsp(self, mock_get_exchange_rate):
        create_programme_response = self.graphql_request(
            request_string=CREATE_PROGRAMME_MUTATION,
            context={"user": self.user},
            variables={
                "programData": {
                    "name": "NName",
                    "scope": "UNICEF",
                    "startDate": "2022-08-24",
                    "endDate": "2022-08-31",
                    "description": "desc",
                    "budget": "0.00",
                    "administrativeAreasOfImplementation": "",
                    "populationGoal": 0,
                    "frequencyOfPayments": "REGULAR",
                    "sector": "MULTI_PURPOSE",
                    "cashPlus": True,
                    "individualDataNeeded": False,
                    "businessAreaSlug": self.business_area.slug,
                }
            },
        )
        program_id = create_programme_response["data"]["createProgram"]["program"]["id"]

        self.graphql_request(
            request_string=UPDATE_PROGRAM_MUTATION,
            context={"user": self.user},
            variables={
                "programData": {
                    "id": program_id,
                    "status": "ACTIVE",
                },
            },
        )

        create_target_population_response = self.graphql_request(
            request_string=CREATE_TARGET_POPULATION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "programId": program_id,
                    "name": "TargP",
                    "excludedIds": "",
                    "exclusionReason": "",
                    "businessAreaSlug": self.business_area.slug,
                    "targetingCriteria": {
                        "rules": [
                            {
                                "filters": [
                                    {
                                        "comparisonMethod": "EQUALS",
                                        "arguments": ["True"],
                                        "fieldName": "consent",
                                        "isFlexField": False,
                                    }
                                ],
                                "individualsFiltersBlocks": [],
                            }
                        ]
                    },
                }
            },
        )
        target_population_id = create_target_population_response["data"]["createTargetPopulation"]["targetPopulation"][
            "id"
        ]

        self.graphql_request(
            request_string=LOCK_TARGET_POPULATION_MUTATION,
            context={"user": self.user},
            variables={
                "id": target_population_id,
            },
        )

        finalize_tp_response = self.graphql_request(
            request_string=FINALIZE_TARGET_POPULATION_MUTATION,
            context={"user": self.user},
            variables={
                "id": target_population_id,
            },
        )
        status = finalize_tp_response["data"]["finalizeTargetPopulation"]["targetPopulation"]["status"]
        self.assertEqual(status, "READY_FOR_PAYMENT_MODULE")

        # TODO: naive datetime
        create_payment_plan_response = self.graphql_request(
            request_string=CREATE_PAYMENT_PLAN_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "businessAreaSlug": self.business_area.slug,
                    "targetingId": target_population_id,
                    "startDate": "2022-08-24",
                    "endDate": "2022-08-31",
                    "dispersionStartDate": (timezone.now() - timedelta(days=1)).strftime("%Y-%m-%d"),
                    "dispersionEndDate": (timezone.now() + timedelta(days=1)).strftime("%Y-%m-%d"),
                    "currency": "USD",
                }
            },
        )
        assert "errors" not in create_payment_plan_response, create_payment_plan_response
        encoded_payment_plan_id = create_payment_plan_response["data"]["createPaymentPlan"]["paymentPlan"]["id"]
        payment_plan_id = decode_id_string(encoded_payment_plan_id)

        santander_fsp = FinancialServiceProviderFactory(
            name="Santander",
            delivery_mechanisms=[GenericPayment.DELIVERY_TYPE_CASH, GenericPayment.DELIVERY_TYPE_TRANSFER],
            distribution_limit=None,
        )
        encoded_santander_fsp_id = encode_id_base64(santander_fsp.id, "FinancialServiceProvider")

        payment = PaymentFactory(
            payment_plan=PaymentPlan.objects.get(id=payment_plan_id),
            business_area=self.business_area,
            household=self.household_1,
            collector=self.individual_1,
            delivery_type=GenericPayment.DELIVERY_TYPE_CASH,
            entitlement_quantity=1000,
            entitlement_quantity_usd=100,
            delivered_quantity=None,
            delivered_quantity_usd=None,
            financial_service_provider=santander_fsp,
            assigned_payment_channel=self.payment_channel_1_cash,
            excluded=False,
        )
        self.assertEqual(payment.entitlement_quantity, 1000)

        lock_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "LOCK",
                }
            },
        )
        assert "errors" not in lock_payment_plan_response, lock_payment_plan_response
        assert lock_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"] == "LOCKED"

        rule = RuleFactory(name="Rule")
        RuleCommitFactory(definition="result.value=Decimal('500')", rule=rule)

        payment_plan = PaymentPlan.objects.get(id=payment_plan_id)
        self.assertEqual(payment_plan.background_action_status, None)

        with patch("hct_mis_api.apps.payment.mutations.payment_plan_apply_steficon") as mock:
            set_steficon_response = self.graphql_request(
                request_string=SET_STEFICON_RULE_MUTATION,
                context={"user": self.user},
                variables={
                    "paymentPlanId": encoded_payment_plan_id,
                    "steficonRuleId": encode_id_base64(rule.id, "Rule"),
                },
            )
            assert "errors" not in set_steficon_response, set_steficon_response
            assert mock.delay.call_count == 1
            call_args = mock.delay.call_args[0]
            payment_plan_apply_steficon(*call_args)

        payment_plan.refresh_from_db()
        self.assertEqual(payment_plan.background_action_status, None)

        payment.refresh_from_db()
        self.assertEqual(payment.entitlement_quantity, 500)

        choose_dms_response = self.graphql_request(
            request_string=CHOOSE_DELIVERY_MECHANISMS_MUTATION,
            context={"user": self.user},
            variables=dict(
                input=dict(
                    paymentPlanId=encoded_payment_plan_id,
                    deliveryMechanisms=[
                        GenericPayment.DELIVERY_TYPE_CASH,
                    ],
                )
            ),
        )
        assert "errors" not in choose_dms_response, choose_dms_response

        available_fsps_query_response = self.graphql_request(
            request_string=AVAILABLE_FSPS_FOR_DELIVERY_MECHANISMS_QUERY,
            context={"user": self.user},
            variables={
                "paymentPlanId": encoded_payment_plan_id,
                "fspChoices": [],
            },
        )
        assert "errors" not in available_fsps_query_response, available_fsps_query_response
        available_fsps_data = available_fsps_query_response["data"]["paymentPlan"]["availableFspsForDeliveryMechanisms"]
        print("available_fsps_data", available_fsps_data)
        assert len(available_fsps_data) == 1
        fsps = available_fsps_data[0]["fsps"]
        assert len(fsps) > 0
        assert fsps[0]["name"] == santander_fsp.name

        assign_fsp_mutation_response = self.graphql_request(
            request_string=ASSIGN_FSPS_MUTATION,
            context={"user": self.user},
            variables={
                "paymentPlanId": encoded_payment_plan_id,
                "mappings": [
                    {
                        "deliveryMechanism": GenericPayment.DELIVERY_TYPE_CASH,
                        "fspId": encoded_santander_fsp_id,
                        "order": 1,
                    }
                ],
            },
        )
        assert "errors" not in assign_fsp_mutation_response, assign_fsp_mutation_response

        lock_fsp_in_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "LOCK_FSP",
                }
            },
        )
        assert "errors" not in lock_fsp_in_payment_plan_response, lock_fsp_in_payment_plan_response
        self.assertEqual(
            lock_fsp_in_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "LOCKED_FSP",
        )
        # TODO: observe that payments have received amounts set

        send_for_approval_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "SEND_FOR_APPROVAL",
                }
            },
        )
        assert "errors" not in send_for_approval_payment_plan_response, send_for_approval_payment_plan_response
        self.assertEqual(
            send_for_approval_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "IN_APPROVAL",
        )

        approve_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "APPROVE",
                }
            },
        )
        assert "errors" not in approve_payment_plan_response, approve_payment_plan_response
        self.assertEqual(
            approve_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "IN_AUTHORIZATION",
        )

        authorize_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "AUTHORIZE",
                }
            },
        )
        assert "errors" not in authorize_payment_plan_response, authorize_payment_plan_response
        self.assertEqual(
            authorize_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "IN_REVIEW",
        )

        review_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "REVIEW",
                }
            },
        )
        assert "errors" not in review_payment_plan_response, review_payment_plan_response
        self.assertEqual(
            review_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "ACCEPTED",
        )

        payment_plan.refresh_from_db()
        self.assertEqual(payment_plan.background_action_status, None)

        with patch(
            "hct_mis_api.apps.payment.services.payment_plan_services.create_payment_plan_payment_list_xlsx_per_fsp"
        ) as mock_export:
            export_file_mutation = self.graphql_request(
                request_string=EXPORT_XLSX_PER_FSP_MUTATION,
                context={"user": self.user},
                variables={
                    "paymentPlanId": encoded_payment_plan_id,
                },
            )
            assert "errors" not in export_file_mutation, export_file_mutation
            assert mock_export.delay.call_count == 1
            call_args = mock_export.delay.call_args[0]
            create_payment_plan_payment_list_xlsx_per_fsp(*call_args)

        payment_plan.refresh_from_db()
        zip_file = payment_plan.export_per_fsp_zip_file
        assert zip_file is not None

        with tempfile.TemporaryDirectory() as temp_dir:
            self.assertEqual(len(os.listdir(temp_dir)), 0)

            assert zip_file.file is not None
            with ZipFile(zip_file.file, "r") as zip_ref:
                self.assertEqual(len(zip_ref.namelist()), 1)
                zip_ref.extractall(temp_dir)

            self.assertEqual(len(os.listdir(temp_dir)), 1)

            file_name = os.listdir(temp_dir)[0]
            assert file_name.endswith(".xlsx")
            file_path = os.path.join(temp_dir, file_name)

            workbook = load_workbook(file_path)
            assert workbook.sheetnames == ["Santander"], workbook.sheetnames

            sheet = workbook["Santander"]
            assert sheet.max_row == 2, sheet.max_row

            self.assertEqual(sheet.cell(row=1, column=1).value, "payment_id")
            assert payment_plan.payments.count() == 1
            payment = payment_plan.payments.first()
            self.assertEqual(sheet.cell(row=2, column=1).value, payment.unicef_id)  # unintuitive

            self.assertEqual(payment.assigned_payment_channel.delivery_mechanism, "Cash")

            self.assertEqual(sheet.cell(row=1, column=5).value, "payment_channel")
            self.assertEqual(sheet.cell(row=2, column=5).value, "Cash")

            self.assertEqual(sheet.cell(row=1, column=7).value, "entitlement_quantity")
            self.assertEqual(sheet.cell(row=2, column=7).value, 500)

            self.assertEqual(sheet.cell(row=1, column=8).value, "delivered_quantity")
            self.assertEqual(sheet.cell(row=2, column=8).value, None)

            payment.refresh_from_db()
            self.assertEqual(payment.entitlement_quantity, 500)
            self.assertEqual(payment.delivered_quantity, None)
            self.assertEqual(payment.is_reconciled, False)

            sheet.cell(row=2, column=8).value = 500
            filled_file_name = "filled.xlsx"
            filled_file_path = os.path.join(temp_dir, filled_file_name)
            workbook.save(filled_file_path)

            with open(filled_file_path, "rb") as file:
                uploaded_file = SimpleUploadedFile(filled_file_name, file.read())
                with patch(
                    "hct_mis_api.apps.payment.services.payment_plan_services.import_payment_plan_payment_list_per_fsp_from_xlsx"
                ) as mock_import:
                    import_mutation_response = self.graphql_request(
                        request_string=IMPORT_XLSX_PER_FSP_MUTATION,
                        context={"user": self.user},
                        variables={
                            "paymentPlanId": encoded_payment_plan_id,
                            "file": uploaded_file,
                        },
                    )
                    assert "errors" not in import_mutation_response, import_mutation_response
                    assert mock_import.delay.call_count == 1
                    call_args = mock_import.delay.call_args[0]
                    import_payment_plan_payment_list_per_fsp_from_xlsx(*call_args)

            payment.refresh_from_db()
            self.assertEqual(payment.entitlement_quantity, 500)
            self.assertEqual(payment.delivered_quantity, 500)
            self.assertEqual(payment.is_reconciled, True)
