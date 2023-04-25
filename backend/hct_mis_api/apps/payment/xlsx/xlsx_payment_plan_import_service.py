import io
from decimal import Decimal
from typing import TYPE_CHECKING, List

from django.contrib.admin.options import get_content_type_for_model
from django.utils import timezone

import openpyxl
from xlwt import Row

from hct_mis_api.apps.core.models import FileTemp
from hct_mis_api.apps.payment.models import Payment, PaymentPlan
from hct_mis_api.apps.payment.utils import get_quantity_in_usd, to_decimal
from hct_mis_api.apps.payment.xlsx.base_xlsx_import_service import XlsxImportBaseService
from hct_mis_api.apps.payment.xlsx.xlsx_error import XlsxError
from hct_mis_api.apps.payment.xlsx.xlsx_payment_plan_base_service import (
    XlsxPaymentPlanBaseService,
)

if TYPE_CHECKING:
    from hct_mis_api.apps.account.models import User


class XlsxPaymentPlanImportService(XlsxPaymentPlanBaseService, XlsxImportBaseService):
    COLUMNS_TYPES = ("s", "s", "n", "s", "s", "s", "s", "s", "n", "n", "n")

    def __init__(self, payment_plan: PaymentPlan, file: io.BytesIO) -> None:
        self.payment_plan = payment_plan
        self.payment_list = payment_plan.eligible_payments
        self.file = file
        self.errors: List[XlsxError] = []
        self.payments_dict = {str(x.unicef_id): x for x in self.payment_list}
        self.payment_ids = list(self.payments_dict.keys())
        self.payments_to_save = []
        self.is_updated = False

    def open_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.load_workbook(self.file, data_only=True)
        self.wb = wb
        self.ws_payments = wb[self.TITLE]
        return wb

    def validate(self) -> None:
        self._validate_headers()
        self._validate_rows()
        self._validate_imported_file()

    def import_payment_list(self) -> None:
        exchange_rate = self.payment_plan.get_exchange_rate()

        for row in self.ws_payments.iter_rows(min_row=2):
            if not any(cell.value for cell in row):
                continue
            self._import_row(row, exchange_rate)

        Payment.objects.bulk_update(
            self.payments_to_save, ("entitlement_quantity", "entitlement_quantity_usd", "entitlement_date")
        )

    def _validate_headers(self) -> None:
        headers_row = self.ws_payments[1]
        accepted_headers = self.HEADERS
        if len(headers_row) != len(accepted_headers):
            self.errors.append(
                XlsxError(
                    self.TITLE,
                    None,
                    f"Different count of headers. Acceptable headers are: [{accepted_headers}]",
                )
            )
        column = 0
        for header in headers_row:
            if column >= len(accepted_headers):
                self.errors.append(
                    XlsxError(
                        self.TITLE,
                        header.coordinate,
                        f"Unexpected header {header.value}",
                    )
                )
            elif header.value != accepted_headers[column]:
                self.errors.append(
                    XlsxError(
                        self.TITLE,
                        header.coordinate,
                        f"Unexpected header {header.value} expected {accepted_headers[column]}",
                    )
                )
            column += 1

    def _validate_row_types(self, row: Row) -> None:
        column = 0
        for cell in row:
            if cell.value is None:
                column += 1
                continue
            if cell.data_type != self.COLUMNS_TYPES[column]:
                readable_cell_error = self.TYPES_READABLE_MAPPING[self.COLUMNS_TYPES[column]]
                self.errors.append(
                    XlsxError(
                        self.TITLE,
                        cell.coordinate,
                        f"Wrong type off cell {readable_cell_error} "
                        f"expected, {self.TYPES_READABLE_MAPPING[cell.data_type]} given.",
                    )
                )
            column += 1

    def _validate_payment_id(self, row: Row) -> None:
        cell = row[self.HEADERS.index("payment_id")]
        if cell.value not in self.payment_ids:
            self.errors.append(
                XlsxError(
                    self.TITLE,
                    cell.coordinate,
                    f"This payment id {cell.value} is not in Payment Plan Payment List",
                )
            )

    def _validate_entitlement(self, row: Row) -> None:
        payment_id = row[self.HEADERS.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return
        entitlement_amount = row[self.HEADERS.index("entitlement_quantity")].value
        if entitlement_amount is not None and entitlement_amount != "":
            entitlement_amount = to_decimal(entitlement_amount)
            if entitlement_amount != payment.entitlement_quantity:
                self.is_updated = True

    def _validate_imported_file(self) -> None:
        if not self.is_updated:
            self.errors.append(
                XlsxError(
                    self.TITLE,
                    None,
                    "There aren't any updates in imported file, please add changes and try again",
                )
            )

    def _validate_rows(self) -> None:
        for row in self.ws_payments.iter_rows(min_row=2):
            if not any(cell.value for cell in row):
                continue
            self._validate_row_types(row)
            self._validate_payment_id(row)
            self._validate_entitlement(row)

    def _import_row(self, row: Row, exchange_rate: float) -> None:
        payment_id = row[self.HEADERS.index("payment_id")].value
        entitlement_amount = row[self.HEADERS.index("entitlement_quantity")].value

        payment = self.payments_dict.get(payment_id)

        if payment is None:
            return

        if entitlement_amount is not None and entitlement_amount != "":
            entitlement_amount = to_decimal(entitlement_amount)
            if entitlement_amount != payment.entitlement_quantity:
                payment.entitlement_quantity = entitlement_amount
                payment.entitlement_date = timezone.now()
                payment.entitlement_quantity_usd = get_quantity_in_usd(
                    amount=entitlement_amount,
                    currency=self.payment_plan.currency,
                    exchange_rate=Decimal(exchange_rate) if exchange_rate is not None else 1,
                    currency_exchange_date=self.payment_plan.currency_exchange_date,
                )
                self.payments_to_save.append(payment)

    def create_import_xlsx_file(self, user: "User") -> PaymentPlan:
        # remove old imported file
        self.payment_plan.remove_imported_file()

        # create new imported xlsx file
        xlsx_file = FileTemp.objects.create(
            object_id=self.payment_plan.pk,
            content_type=get_content_type_for_model(self.payment_plan),
            created_by=user,
            file=self.file,
        )

        self.payment_plan.imported_file = xlsx_file
        self.payment_plan.save()

        return self.payment_plan
