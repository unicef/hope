import openpyxl

from hct_mis_api.apps.payment.xlsx.BaseXlsxImportService import XlsxImportBaseService
from hct_mis_api.apps.payment.models import Payment
from hct_mis_api.apps.payment.utils import float_to_decimal, get_quantity_in_usd
from hct_mis_api.apps.payment.xlsx.XlsxPaymentPlanExportService import XlsxPaymentPlanExportService


class XlsxPaymentPlanImportPerFspService(XlsxImportBaseService):
    def __init__(self, payment_plan, file):
        self.payment_plan = payment_plan
        self.payment_list = payment_plan.not_excluded_payments
        self.file = file
        self.errors = []
        self.payments_dict = {str(x.unicef_id): x for x in self.payment_list}
        self.payment_ids = list(self.payments_dict.keys())
        self.payments_to_save = []
        self.fsp = None
        self.expected_columns = []
        self.is_updated = False

    def _set_fsp_expected_columns(self):
        first_payment_row = self.ws_payments[2]
        payment_id = first_payment_row[XlsxPaymentPlanExportService.HEADERS.index("payment_id")].value
        payment = self.payments_dict[payment_id]

        self.fsp = payment.financial_service_provider
        self.expected_columns = (
            self.fsp.fsp_xlsx_template and self.fsp.fsp_xlsx_template.columns
        ) or XlsxPaymentPlanExportService.HEADERS

    def open_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.load_workbook(self.file, data_only=True)
        self.wb = wb
        self.ws_payments = wb[wb.sheetnames[0]]
        self._set_fsp_expected_columns()

        return wb

    def _validate_headers(self):
        headers_row = self.ws_payments[1]

        if len(headers_row) != len(self.expected_columns):
            self.errors.append(
                (
                    self.fsp.name,
                    None,
                    f"Provided headers {[header.value for header in headers_row]}"
                    f" do not match expected headers {self.expected_columns}, "
                    f"please use exported Template File to import data.",
                )
            )
            return

        for header, expected_column in zip(headers_row, self.expected_columns):
            if header.value != expected_column:
                self.errors.append(
                    (
                        self.fsp.name,
                        header.coordinate,
                        f"Unexpected header {header.value}, expected {expected_column}, "
                        f"please use exported Template File to import data.",
                    )
                )

    def _validate_payment_id(self, row):
        cell = row[self.expected_columns.index("payment_id")]
        if cell.value not in self.payment_ids:
            self.errors.append(
                (
                    self.fsp.name,
                    cell.coordinate,
                    f"This payment id {cell.value} is not in Payment Plan Payment List",
                )
            )

    def _validate_fsp(self, row):
        cell = row[self.expected_columns.index("payment_id")]
        if cell.value not in self.payment_ids:
            self.errors.append(
                (
                    self.fsp.name,
                    cell.coordinate,
                    f"This payment id {cell.value} is not in Payment Plan Payment List",
                )
            )

    def _validate_delivered_quantity(self, row):
        payment_id = row[self.expected_columns.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return
        delivered_quantity = row[self.expected_columns.index("delivered_quantity")].value
        if delivered_quantity is not None and delivered_quantity != "":
            delivered_quantity = float_to_decimal(delivered_quantity)
            if delivered_quantity != payment.delivered_quantity:
                self.is_updated = True

    def _validate_rows(self):
        for row in self.ws_payments.iter_rows(min_row=2):
            if not any([cell.value for cell in row]):
                continue

            self._validate_payment_id(row)
            self._validate_fsp(row)
            self._validate_delivered_quantity(row)

    def _validate_imported_file(self):
        if not self.is_updated:
            self.errors.append(
                (
                    self.fsp.name,
                    None,
                    "There aren't any updates in imported file, please add changes and try again",
                )
            )

    def validate(self):
        self._validate_headers()
        self._validate_rows()
        self._validate_imported_file()

    def import_payment_list(self):
        exchange_rate = self.payment_plan.get_exchange_rate()

        for row in self.ws_payments.iter_rows(min_row=2):
            self._import_row(row, exchange_rate)

        Payment.objects.bulk_update(self.payments_to_save, ("delivered_quantity", "status"))

    def _import_row(self, row, exchange_rate):
        payment_id = row[self.expected_columns.index("payment_id")].value
        payment = self.payments_dict[payment_id]
        delivered_quantity = row[self.expected_columns.index("delivered_quantity")].value

        if delivered_quantity is not None and delivered_quantity != "":
            delivered_quantity = float_to_decimal(delivered_quantity)
            if delivered_quantity != payment.delivered_quantity:
                payment.delivered_quantity = delivered_quantity
                payment.entitlement_quantity_usd = get_quantity_in_usd(
                    amount=delivered_quantity,
                    currency=self.payment_plan.currency,
                    exchange_rate=exchange_rate,
                    currency_exchange_date=self.payment_plan.currency_exchange_date,
                )

                payment.status = Payment.STATUS_DISTRIBUTION_SUCCESS
                self.payments_to_save.append(payment)
