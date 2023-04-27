import datetime
import io
from decimal import Decimal
from typing import TYPE_CHECKING, Dict, List, Optional, Tuple

from django.db.models import QuerySet
from django.utils import timezone

import openpyxl
import pytz
from dateutil.parser import parse
from xlwt import Row

from hct_mis_api.apps.payment.models import Payment, PaymentVerification
from hct_mis_api.apps.payment.services.handle_total_cash_in_households import (
    handle_total_cash_in_specific_households,
)
from hct_mis_api.apps.payment.utils import (
    calculate_counts,
    get_quantity_in_usd,
    to_decimal,
)
from hct_mis_api.apps.payment.xlsx.base_xlsx_import_service import XlsxImportBaseService
from hct_mis_api.apps.payment.xlsx.xlsx_error import XlsxError

if TYPE_CHECKING:
    from hct_mis_api.apps.payment.models import PaymentPlan


class XlsxPaymentPlanImportPerFspService(XlsxImportBaseService):
    class XlsxPaymentPlanImportPerFspServiceException(Exception):
        pass

    def __init__(self, payment_plan: "PaymentPlan", file: io.BytesIO) -> None:
        self.payment_plan = payment_plan
        self.payment_list: QuerySet["Payment"] = payment_plan.not_excluded_payments
        self.file = file
        self.errors: List[XlsxError] = []
        self.payments_dict: Dict = {str(x.unicef_id): x for x in self.payment_list}
        self.payment_ids: List = list(self.payments_dict.keys())
        self.payments_to_save: List = []
        self.payment_verifications_to_save: List = []
        self.required_columns: List[str] = ["payment_id", "delivered_quantity"]
        self.xlsx_headers = []
        self.is_updated: bool = False

    def open_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.load_workbook(self.file, data_only=True)
        self.wb = wb
        self.ws_payments = wb[wb.sheetnames[0]]
        self.sheetname = wb.sheetnames[0]
        self.xlsx_headers = [header.value for header in self.ws_payments[1]]
        return wb

    def _validate_headers(self) -> None:
        # need check only if "payment_id" and "delivered_quantity" exists
        for required_column in self.required_columns:
            if required_column not in self.xlsx_headers:
                self.errors.append(
                    XlsxError(
                        self.sheetname,
                        None,
                        f"Provided headers {self.xlsx_headers}"
                        f" do not match expected headers. {self.required_columns} "
                        f"are required headers.",
                    )
                )
                return

    def _validate_payment_id(self, row: Row) -> None:
        cell = row[self.xlsx_headers.index("payment_id")]
        if cell.value not in self.payment_ids:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"This payment id {cell.value} is not in Payment Plan Payment List",
                )
            )

    def _validate_delivered_quantity(self, row: Row) -> None:
        """
        It should be possible for a user to upload a file when:

        * Fully Delivered (entitled quantity = delivered quantity) [float]
        * Partially Delivered (entitled quantity > delivered quantity > 0) [float]
        * Not Delivered (0 = delivered quantity) [0.0]
        * Unsuccessful (failed at the delivery processing level) [-1.0]
        * Pending (no information) [None]

        The validation should not pass when:

        * delivered quantity > entitled quantity
        """

        payment_id = row[self.xlsx_headers.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return

        cell = row[self.xlsx_headers.index("delivered_quantity")]
        delivered_quantity = cell.value

        if delivered_quantity is not None and delivered_quantity != "":
            delivered_quantity = to_decimal(delivered_quantity)
            if delivered_quantity != payment.delivered_quantity:  # update value
                if delivered_quantity > payment.entitlement_quantity:
                    self.errors.append(
                        XlsxError(
                            self.sheetname,
                            cell.coordinate,
                            f"Payment {payment_id}: Delivered quantity {delivered_quantity} is bigger than "
                            f"Entitlement quantity {payment.entitlement_quantity}",
                        )
                    )
                else:
                    self.is_updated = True

    def _validate_delivery_date(self, row: Row) -> None:
        payment_id = row[self.xlsx_headers.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return

        cell = row[self.xlsx_headers.index("delivery_date")]
        delivery_date = cell.value

        if delivery_date is None:
            self.is_updated = True  # Update Payment item with current datetime
            return

        try:
            if not isinstance(delivery_date, datetime.datetime):
                delivery_date = parse(delivery_date)
            if delivery_date != payment.delivery_date.replace(tzinfo=None):
                self.is_updated = True
        except Exception:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    cell.coordinate,
                    f"Payment {payment_id}: Delivered date {delivery_date} is not a datetime",
                )
            )

    def _validate_rows(self) -> None:
        for row in self.ws_payments.iter_rows(min_row=2):
            if not any([cell.value for cell in row]):
                continue

            self._validate_payment_id(row)
            self._validate_delivered_quantity(row)
            self._validate_delivery_date(row)

    def _validate_imported_file(self) -> None:
        if not self.is_updated:
            self.errors.append(
                XlsxError(
                    self.sheetname,
                    None,
                    "There aren't any updates in imported file, please add changes and try again",
                )
            )

    def validate(self) -> None:
        self._validate_headers()
        if not self.errors:
            self._validate_rows()
            self._validate_imported_file()

    def import_payment_list(self) -> None:
        exchange_rate = self.payment_plan.get_exchange_rate()

        for row in self.ws_payments.iter_rows(min_row=2):
            self._import_row(row, exchange_rate)

        Payment.objects.bulk_update(
            self.payments_to_save, ("delivered_quantity", "delivered_quantity_usd", "status", "delivery_date")
        )
        handle_total_cash_in_specific_households([payment.household_id for payment in self.payments_to_save])
        PaymentVerification.objects.bulk_update(self.payment_verifications_to_save, ("status", "status_date"))

    def _get_delivered_quantity_status_and_value(
        self, delivered_quantity: float, entitlement_quantity: Decimal, payment_id: str
    ) -> Tuple[str, Optional[Decimal]]:
        """
        * Fully Delivered (entitled quantity = delivered quantity) [float]
        * Partially Delivered (entitled quantity > delivered quantity > 0) [float]
        * Not Delivered (0 = delivered quantity) [0.0]
        * Unsuccessful (failed at the delivery processing level) [-1.0]
        """

        if delivered_quantity < 0:
            return Payment.STATUS_ERROR, None

        elif delivered_quantity == 0:
            return Payment.STATUS_NOT_DISTRIBUTED, to_decimal(delivered_quantity)

        elif delivered_quantity < entitlement_quantity:
            return Payment.STATUS_DISTRIBUTION_PARTIAL, to_decimal(delivered_quantity)

        elif delivered_quantity == entitlement_quantity:
            return Payment.STATUS_DISTRIBUTION_SUCCESS, to_decimal(delivered_quantity)

        else:
            raise self.XlsxPaymentPlanImportPerFspServiceException(
                f"Invalid delivered_quantity {delivered_quantity} provided for payment_id {payment_id}"
            )

    def _import_row(self, row: Row, exchange_rate: float) -> None:
        payment_id = row[self.xlsx_headers.index("payment_id")].value
        if payment_id is None:
            return  # safety check
        payment = self.payments_dict[payment_id]
        delivered_quantity = row[self.xlsx_headers.index("delivered_quantity")].value
        delivery_date = row[self.xlsx_headers.index("delivery_date")].value

        if delivery_date is None:
            delivery_date = timezone.now()
            print("delivery_date", delivery_date)
        elif isinstance(delivery_date, str):
            delivery_date = parse(delivery_date)

        if delivery_date.tzinfo is None:
            delivery_date = pytz.utc.localize(delivery_date)

        # payment_delivery_date = payment.delivery_date
        if payment_delivery_date := payment.delivery_date:
            payment_delivery_date = payment.delivery_date.replace(tzinfo=None)

        if delivered_quantity is not None and delivered_quantity != "":
            status, delivered_quantity = self._get_delivered_quantity_status_and_value(
                delivered_quantity, payment.entitlement_quantity, payment_id
            )

            if (
                (delivered_quantity != payment.delivered_quantity)
                or (status != payment.status)
                or (delivery_date != payment_delivery_date)
            ):
                payment.delivered_quantity = delivered_quantity
                payment.delivered_quantity_usd = get_quantity_in_usd(
                    amount=delivered_quantity,
                    currency=self.payment_plan.currency,
                    exchange_rate=Decimal(exchange_rate),
                    currency_exchange_date=self.payment_plan.currency_exchange_date,
                )
                payment.status = status
                payment.delivery_date = delivery_date
                self.payments_to_save.append(payment)
                # update PaymentVerification status
                if payment.payment_verification.exists():
                    payment_verification = payment.payment_verification.first()

                    if payment_verification.status != PaymentVerification.STATUS_PENDING:
                        if payment_verification.received_amount == delivered_quantity:
                            pv_status = PaymentVerification.STATUS_RECEIVED
                        elif delivered_quantity == 0 or delivered_quantity is None:
                            pv_status = PaymentVerification.STATUS_NOT_RECEIVED
                        else:
                            pv_status = PaymentVerification.STATUS_RECEIVED_WITH_ISSUES

                        payment_verification.status = pv_status
                        payment_verification.status_date = timezone.now()
                        self.payment_verifications_to_save.append(payment_verification)

                        calculate_counts(payment_verification.payment_verification_plan)
                        payment_verification.payment_verification_plan.save()
