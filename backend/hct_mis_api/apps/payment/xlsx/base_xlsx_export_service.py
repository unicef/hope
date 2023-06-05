import decimal
import logging
from datetime import datetime
from typing import TYPE_CHECKING, Any, Dict, List, Optional

from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.urls import reverse

import openpyxl
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

from hct_mis_api.apps.core.utils import encode_id_base64

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from hct_mis_api.apps.account.models import User


logger = logging.getLogger(__name__)


class XlsxExportBaseService:
    def _create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_active = wb.active
        ws_active.title = self.TITLE
        self.wb = wb
        self.ws_export_list = ws_active
        return wb

    def _add_headers(self) -> None:
        self.ws_export_list.append(self.HEADERS)

    def generate_workbook(self) -> openpyxl.Workbook:
        self._create_workbook()
        self._add_headers()
        # add export items and what you need
        return self.wb

    def generate_file(self, filename: str) -> None:
        self.generate_workbook()
        self.wb.save(filename=filename)

    @staticmethod
    def _adjust_column_width_from_col(ws: "Worksheet") -> None:
        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

        ws.column_dimensions = dim_holder

    def _add_col_bgcolor(
        self, col: Optional[List] = None, hex_code: str = "A0FDB0", no_of_columns: Optional[int] = None
    ) -> None:
        for row_index in col or []:
            fill = PatternFill(bgColor=hex_code, fgColor=hex_code, fill_type="lightUp")
            bd = Side(style="thin", color="999999")
            for y in range(
                1,
                (self.ws_export_list.max_column if no_of_columns is None else no_of_columns) + 1,
            ):
                cell = self.ws_export_list.cell(row=y, column=row_index)
                cell.fill = fill
                cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    @staticmethod
    def get_link(api_url: Optional[str] = None) -> str:
        protocol = "https" if settings.SOCIAL_AUTH_REDIRECT_IS_HTTPS else "http"
        link = f"{protocol}://{settings.FRONTEND_HOST}{api_url}"
        if api_url:
            return link
        return ""

    def send_email(self, context: Dict) -> None:
        text_body = render_to_string("payment/xlsx_file_generated_email.txt", context=context)
        html_body = render_to_string("payment/xlsx_file_generated_email.html", context=context)

        email = EmailMultiAlternatives(
            subject=context["title"],
            from_email=settings.EMAIL_HOST_USER,
            to=[context["email"]],
            body=text_body,
        )
        email.attach_alternative(html_body, "text/html")
        result = email.send()
        if not result:
            logger.error(f"Email couldn't be send to {context['email']}")

    def get_email_context(self, user: "User") -> Dict:
        payment_verification_id = encode_id_base64(self.payment_plan.id, "PaymentPlan")
        path_name = "download-payment-plan-payment-list"
        link = self.get_link(reverse(path_name, args=[payment_verification_id]))

        msg = "Payment Plan Payment List xlsx file(s) were generated and below You have the link to download this file."

        context = {
            "first_name": getattr(user, "first_name", ""),
            "last_name": getattr(user, "last_name", ""),
            "email": getattr(user, "email", ""),
            "message": msg,
            "link": link,
            "title": "Payment Plan Payment List files generated",
        }

        return context

    def right_format_for_xlsx(self, value: Any) -> Any:
        # this function will return something that excel will accept
        if value is None:
            return ""
        if isinstance(value, (str, int, float, decimal.Decimal, datetime)):
            return value
        return str(value)
