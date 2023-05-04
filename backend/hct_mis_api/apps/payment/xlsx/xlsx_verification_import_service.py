import io
from decimal import Decimal
from typing import Dict, List

import openpyxl
from graphql import GraphQLError
from openpyxl.utils import get_column_letter
from xlwt import Row, Worksheet

from hct_mis_api.apps.payment.models import PaymentVerification, PaymentVerificationPlan
from hct_mis_api.apps.payment.utils import from_received_yes_no_to_status, to_decimal
from hct_mis_api.apps.payment.xlsx.base_xlsx_import_service import XlsxImportBaseService
from hct_mis_api.apps.payment.xlsx.xlsx_error import XlsxError
from hct_mis_api.apps.payment.xlsx.xlsx_verification_export_service import (
    XlsxVerificationExportService,
)


class XlsxVerificationImportService(XlsxImportBaseService):
    def __init__(self, cashplan_payment_verification: PaymentVerificationPlan, file: io.BytesIO) -> None:
        self.file = file
        self.cashplan_payment_verification = cashplan_payment_verification
        self.payment_record_verifications = cashplan_payment_verification.payment_record_verifications.all()
        self.current_row = 0
        self.errors: List[XlsxError] = []
        payment_record_verification_obj = self.cashplan_payment_verification.payment_record_verifications
        self.payment_record_verifications = payment_record_verification_obj.all()  # .prefetch_related("payment")
        self.payment_record_ids = [str(x.payment_object_id) for x in self.payment_record_verifications]
        self.payment_record_verifications_dict = {
            str(x.payment_object_id): x for x in self.payment_record_verifications
        }
        self.payment_records_dict = {str(x.payment_object_id): x.payment_obj for x in self.payment_record_verifications}
        self.payment_verifications_to_save = []
        self.was_validation_run = False

        # Get mandatory column indices as order of columns might be different in each ws
        self.PAYMENT_RECORD_ID_COLUMN_INDEX = 0
        self.RECEIVED_COLUMN_INDEX = 0
        self.RECEIVED_AMOUNT_COLUMN_INDEX = 0

        self.COLUMN_DICT_TYPES = {}

    @staticmethod
    def get_columns_from_worksheet(ws: Worksheet) -> Dict[str, Dict[str, int]]:
        return {
            cell.value: {"letter": get_column_letter(cell.column), "number": cell.column - 1}
            for cell in ws[1]
            if cell.value
        }

    def open_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.load_workbook(self.file, data_only=True)
        self.wb = wb
        self.ws_verifications = wb[XlsxVerificationExportService.VERIFICATION_SHEET]
        return wb

    def validate(self) -> None:
        self._check_version()
        self._validate_headers()
        self._validate_rows()
        self.was_validation_run = True

    def import_verifications(self) -> None:
        if len(self.errors):
            raise GraphQLError("You can't import verifications with errors.")
        if not self.was_validation_run:
            raise GraphQLError("Run validation before import.")
        for row in self.ws_verifications.iter_rows(min_row=2):
            if not any([cell.value for cell in row]):
                continue
            self._import_row(row)
        PaymentVerification.objects.bulk_update(self.payment_verifications_to_save, ("status", "received_amount"))

    def _check_version(self) -> None:
        ws_meta = self.wb[XlsxVerificationExportService.META_SHEET]
        version_cell_name = ws_meta[XlsxVerificationExportService.VERSION_CELL_NAME_COORDINATES].value
        version = ws_meta[XlsxVerificationExportService.VERSION_CELL_COORDINATES].value

        if (
            version_cell_name != XlsxVerificationExportService.VERSION_CELL_NAME
            or version != XlsxVerificationExportService.VERSION
        ):
            raise GraphQLError(
                f"Unsupported file version ({version}). "
                f"Only version: {XlsxVerificationExportService.VERSION} is supported"
            )

    def _validate_headers(self) -> None:
        headers = self.get_columns_from_worksheet(self.ws_verifications)

        for header_name, header_attrs in headers.items():
            if header_name == "payment_record_id":
                self.PAYMENT_RECORD_ID_COLUMN_INDEX = header_attrs["number"]
            elif header_name == "received":
                self.RECEIVED_COLUMN_INDEX = header_attrs["number"]
            elif header_name == "received_amount":
                self.RECEIVED_AMOUNT_COLUMN_INDEX = header_attrs["number"]

        missing_headers = []

        if self.PAYMENT_RECORD_ID_COLUMN_INDEX is None:
            missing_headers.append("payment_record_id")
        if self.RECEIVED_COLUMN_INDEX is None:
            missing_headers.append("received")
        if self.RECEIVED_AMOUNT_COLUMN_INDEX is None:
            missing_headers.append("received_amount")

        if missing_headers:
            self.errors.append(
                XlsxError(
                    sheet="Payment Verifications",
                    message=f"Missing mandatory headers: {' '.join(missing_headers)}",
                )
            )
        else:
            self.COLUMN_DICT_TYPES = {
                self.PAYMENT_RECORD_ID_COLUMN_INDEX: "s",
                self.RECEIVED_COLUMN_INDEX: "s",
                self.RECEIVED_AMOUNT_COLUMN_INDEX: "n",
            }

    def _validate_mandatory_row_types(self, row: Row) -> None:
        mandatory_column_ids = self.COLUMN_DICT_TYPES.keys()

        column = 0
        for cell in row:
            if column in mandatory_column_ids:
                if cell.value is None:
                    self.errors.append(XlsxError("Payment Verifications", cell.coordinate, "Cell value cannot be null"))
                    continue

                if cell.data_type != self.COLUMN_DICT_TYPES[column]:
                    readable_cell_error = XlsxVerificationImportService.TYPES_READABLE_MAPPING[
                        self.COLUMN_DICT_TYPES[column]
                    ]
                    self.errors.append(
                        XlsxError(
                            "Payment Verifications",
                            cell.coordinate,
                            f"Wrong type off cell {readable_cell_error} "
                            f"expected, {XlsxVerificationImportService.TYPES_READABLE_MAPPING[cell.data_type]} given.",
                        )
                    )
            column += 1

    def _validate_payment_record_id(self, row: Row) -> None:
        cell = row[self.PAYMENT_RECORD_ID_COLUMN_INDEX]
        if cell.value not in self.payment_record_ids:
            self.errors.append(
                XlsxError(
                    "Payment Verifications",
                    cell.coordinate,
                    f"This payment record id {cell.value} is not in Cash Plan Payment Record Verification",
                )
            )

    def _validate_row_received(self, row: Row) -> None:
        valid_received = (None, "YES", "NO")
        cell = row[self.RECEIVED_COLUMN_INDEX]
        if cell.value not in valid_received:
            self.errors.append(
                XlsxError(
                    "Payment Verifications",
                    cell.coordinate,
                    f"The received of this payment verification is not correct: "
                    f"{cell.value} should be one of: {valid_received}",
                )
            )

    def _validate_received_to_received_amount(self, row: Row) -> None:
        payment_record_id = row[self.PAYMENT_RECORD_ID_COLUMN_INDEX].value
        payment_record = self.payment_records_dict.get(payment_record_id)
        if payment_record is None:
            return
        received_amount = row[self.RECEIVED_AMOUNT_COLUMN_INDEX].value
        if received_amount is not None:
            if not isinstance(received_amount, (int, float)):
                return
            received_amount = Decimal(format(round(received_amount, 2), ".2f"))
        received_cell = row[self.RECEIVED_COLUMN_INDEX]
        received = received_cell.value
        if received is None and received_amount is not None and received_amount == 0:
            self.errors.append(
                XlsxError(
                    "Payment Verifications",
                    received_cell.coordinate,
                    f"You can't set received_amount {received_amount} and not set received to NO",
                )
            )
        elif received is None and received_amount is not None:
            self.errors.append(
                XlsxError(
                    "Payment Verifications",
                    received_cell.coordinate,
                    f"You can't set received_amount {received_amount} and not set received to YES",
                )
            )
        elif received_amount == 0 and received != "NO":
            self.errors.append(
                XlsxError(
                    "Payment Verifications",
                    received_cell.coordinate,
                    "If received_amount is 0, you should set received to NO",
                )
            )
        elif received_amount is not None and received_amount != 0 and received != "YES":
            self.errors.append(
                XlsxError(
                    "Payment Verifications",
                    received_cell.coordinate,
                    f"If received_amount({received_amount}) is not 0, you should set received to YES",
                )
            )

    def _validate_rows(self) -> None:
        for row in self.ws_verifications.iter_rows(min_row=2):
            if not any([cell.value for cell in row]):
                continue
            self._validate_mandatory_row_types(row)
            self._validate_payment_record_id(row)
            self._validate_row_received(row)
            self._validate_received_to_received_amount(row)

    def _import_row(self, row: Row) -> None:
        payment_record_id = row[self.PAYMENT_RECORD_ID_COLUMN_INDEX].value
        received = row[self.RECEIVED_COLUMN_INDEX].value
        received_amount = row[self.RECEIVED_AMOUNT_COLUMN_INDEX].value

        payment_verification = self.payment_record_verifications_dict[payment_record_id]
        payment_record = self.payment_records_dict.get(payment_record_id)
        delivered_amount = payment_record.delivered_quantity

        payment_verification.status = from_received_yes_no_to_status(received, received_amount, delivered_amount)
        if received_amount is not None and received_amount != "":
            payment_verification.received_amount = to_decimal(received_amount)
        else:
            payment_verification.received_amount = None
        self.payment_verifications_to_save.append(payment_verification)
