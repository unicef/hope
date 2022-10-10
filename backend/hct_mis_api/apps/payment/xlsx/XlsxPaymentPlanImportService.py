from django.contrib.admin.options import get_content_type_for_model
from django.utils import timezone

import openpyxl

from hct_mis_api.apps.core.models import FileTemp
from hct_mis_api.apps.payment.models import (
    GenericPayment,
    Payment,
    PaymentChannel,
    PaymentPlan,
)
from hct_mis_api.apps.payment.utils import float_to_decimal, get_quantity_in_usd
from hct_mis_api.apps.payment.xlsx.BaseXlsxImportService import XlsxImportBaseService
from hct_mis_api.apps.payment.xlsx.XlsxPaymentPlanExportService import (
    XlsxPaymentPlanExportService,
)


class XlsxPaymentPlanImportService(XlsxImportBaseService):
    COLUMNS_TYPES = ("s", "s", "n", "s", "s", "s", "s", "s", "n", "n")

    def __init__(self, payment_plan: PaymentPlan, file):
        self.payment_plan = payment_plan
        self.payment_list = payment_plan.not_excluded_payments
        self.file = file
        self.errors = []
        self.payments_dict = {str(x.unicef_id): x for x in self.payment_list}
        self.payment_ids = list(self.payments_dict.keys())
        self.payments_to_save = []
        self.is_updated = False

    def open_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.load_workbook(self.file, data_only=True)
        self.wb = wb
        self.ws_payments = wb[XlsxPaymentPlanExportService.TITLE]
        return wb

    def validate(self):
        self._validate_headers()
        self._validate_rows()
        self._validate_imported_file()

    def import_payment_list(self):
        exchange_rate = self.payment_plan.get_exchange_rate()

        for row in self.ws_payments.iter_rows(min_row=2):
            if not any([cell.value for cell in row]):
                continue
            self._import_row(row, exchange_rate)

        Payment.objects.bulk_update(
            self.payments_to_save, ("entitlement_quantity", "entitlement_quantity_usd", "entitlement_date")
        )

    def _validate_headers(self):
        headers_row = self.ws_payments[1]
        accepted_headers = XlsxPaymentPlanExportService.HEADERS
        if len(headers_row) != len(accepted_headers):
            self.errors.append(
                (
                    XlsxPaymentPlanExportService.TITLE,
                    None,
                    f"Different count of headers. Acceptable headers are: [{accepted_headers}]",
                )
            )
        column = 0
        for header in headers_row:
            if column >= len(accepted_headers):
                self.errors.append(
                    (
                        XlsxPaymentPlanExportService.TITLE,
                        header.coordinate,
                        f"Unexpected header {header.value}",
                    )
                )
            elif header.value != accepted_headers[column]:
                self.errors.append(
                    (
                        XlsxPaymentPlanExportService.TITLE,
                        header.coordinate,
                        f"Unexpected header {header.value} expected {accepted_headers[column]}",
                    )
                )
            column += 1

    def _validate_row_types(self, row):
        column = 0
        for cell in row:
            if cell.value is None:
                column += 1
                continue
            if cell.data_type != XlsxPaymentPlanImportService.COLUMNS_TYPES[column]:
                readable_cell_error = XlsxPaymentPlanImportService.TYPES_READABLE_MAPPING[
                    XlsxPaymentPlanImportService.COLUMNS_TYPES[column]
                ]
                self.errors.append(
                    (
                        XlsxPaymentPlanExportService.TITLE,
                        cell.coordinate,
                        f"Wrong type off cell {readable_cell_error} "
                        f"expected, {XlsxPaymentPlanImportService.TYPES_READABLE_MAPPING[cell.data_type]} given.",
                    )
                )
            column += 1

    def _validate_payment_id(self, row):
        cell = row[XlsxPaymentPlanExportService.HEADERS.index("payment_id")]
        if cell.value not in self.payment_ids:
            self.errors.append(
                (
                    XlsxPaymentPlanExportService.TITLE,
                    cell.coordinate,
                    f"This payment id {cell.value} is not in Payment Plan Payment List",
                )
            )

    def _validate_entitlement(self, row):
        payment_id = row[XlsxPaymentPlanExportService.HEADERS.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return
        entitlement_amount = row[XlsxPaymentPlanExportService.HEADERS.index("entitlement_quantity")].value
        if entitlement_amount is not None and entitlement_amount != "":
            entitlement_amount = float_to_decimal(entitlement_amount)
            if entitlement_amount != payment.entitlement_quantity:
                self.is_updated = True

    def _validate_payment_channel(self, row):
        payment_id = row[XlsxPaymentPlanExportService.HEADERS.index("payment_id")].value
        payment = self.payments_dict.get(payment_id)
        if payment is None:
            return
        payment_channels_list = row[XlsxPaymentPlanExportService.HEADERS.index("payment_channel")].value.split(", ")
        payment_channel_cell = row[XlsxPaymentPlanExportService.HEADERS.index("payment_channel")]
        for payment_channel in payment_channels_list:
            if payment_channel not in [x[0] for x in GenericPayment.DELIVERY_TYPE_CHOICE]:
                self.errors.append(
                    (
                        XlsxPaymentPlanExportService.TITLE,
                        payment_channel_cell.coordinate,
                        f"Payment_channel should be one of {[x[0] for x in GenericPayment.DELIVERY_TYPE_CHOICE]} "
                        f"but received {payment_channel}",
                    )
                )
            collectors_payment_channels = list(
                payment.collector.payment_channels.all()
                .distinct("delivery_mechanism")
                .values_list("delivery_mechanism", flat=True)
            )
            if payment.collector.payment_channels.exists() and payment_channel not in collectors_payment_channels:
                self.errors.append(
                    (
                        XlsxPaymentPlanExportService.TITLE,
                        payment_channel_cell.coordinate,
                        f"You can't set payment_channel {payment_channel} for Collector with already assigned payment "
                        f"channel(s): {', '.join(collectors_payment_channels)}",
                    )
                )
            if not payment.collector.payment_channels.exists() and payment_channel:
                self.is_updated = True

    def _validate_imported_file(self):
        if not self.is_updated:
            self.errors.append(
                (
                    XlsxPaymentPlanExportService.TITLE,
                    None,
                    "There are no any updates in imported file, please add changes and try again",
                )
            )

    def _validate_rows(self):
        for row in self.ws_payments.iter_rows(min_row=2):
            if not any([cell.value for cell in row]):
                continue
            self._validate_row_types(row)
            self._validate_payment_id(row)
            self._validate_entitlement(row)
            self._validate_payment_channel(row)

    def _import_row(self, row, exchange_rate):
        payment_id = row[XlsxPaymentPlanExportService.HEADERS.index("payment_id")].value
        entitlement_amount = row[XlsxPaymentPlanExportService.HEADERS.index("entitlement_quantity")].value
        payment_channels_list = row[XlsxPaymentPlanExportService.HEADERS.index("payment_channel")].value.split(", ")

        payment = self.payments_dict.get(payment_id)

        if payment is None:
            return

        if not payment.collector.payment_channels.exists():
            for payment_channel in payment_channels_list:
                if payment_channel is not None and payment_channel != "":
                    # TODO handle delivery data
                    PaymentChannel.objects.get_or_create(
                        individual=payment.collector,
                        delivery_mechanism=payment_channel,
                    )

        if entitlement_amount is not None and entitlement_amount != "":
            entitlement_amount = float_to_decimal(entitlement_amount)
            if entitlement_amount != payment.entitlement_quantity:
                payment.entitlement_quantity = entitlement_amount
                payment.entitlement_date = timezone.now()
                payment.entitlement_quantity_usd = get_quantity_in_usd(
                    amount=entitlement_amount,
                    currency=self.payment_plan.currency,
                    exchange_rate=exchange_rate,
                    currency_exchange_date=self.payment_plan.currency_exchange_date,
                )

                self.payments_to_save.append(payment)

    def create_import_xlsx_file(self, user):
        # remove old imported file
        self.payment_plan.remove_imported_file()

        # create new imported xlsx file
        xlsx_file = FileTemp.objects.create(
            object_id=self.payment_plan.pk,
            content_type=get_content_type_for_model(self.payment_plan),
            created_by=user,
            file=self.file,
        )

        self.payment_plan.imported_file = xlsx_file
        self.payment_plan.save()
