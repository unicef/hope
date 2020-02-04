from collections import defaultdict

from django.contrib import admin
from django.forms import forms, ModelChoiceField
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.html import strip_tags
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

from core.models import Location, FlexibleField

from core.models import BusinessArea


class XLSImportForm(forms.Form):
    location = ModelChoiceField(queryset=Location.objects.all())
    xls_file = forms.FileField()


@admin.register(BusinessArea)
class BusinessAreaAdmin(admin.ModelAdmin):
    list_display = ("name", "slug")


@admin.register(FlexibleField)
class FlexibleFieldAdmin(admin.ModelAdmin):
    change_list_template = "core/flexible_fields_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path("import-xls/", self.import_xls),
        ]
        return my_urls + urls

    def import_xls(self, request):
        if request.method == "POST":
            xls_file = request.FILES["xls_file"]
            location_id = request.POST.get("location")
            location = Location.objects.get(id=location_id)

            model_fields = [
                field.name for field in FlexibleField._meta.get_fields()
            ]

            core_field_suffixes = (
                "_h_c",
                "_h_f",
                "_i_c",
                "_i_f",
            )

            language_fields = (
                "label",
                "hint",
            )

            boolean_fields = (
                "required",
                "read_only",
            )

            wb = load_workbook(xls_file)

            sheet = wb.get_sheet_by_name("survey")

            headers_index_map = {}

            first_row = tuple(sheet.rows)[0]
            for col in first_row:
                if col.value is not None:
                    headers_index_map[col.value] = column_index_from_string(
                        coordinate_from_string(col.coordinate)[0]
                    )

            all_objects = []
            for row in sheet.iter_rows(min_row=2):
                json_fields = defaultdict(dict)
                fields = {}
                add_flag = True
                for cell, header_name in zip(row, headers_index_map.keys()):
                    value = cell.value
                    is_end_repeat_info = (
                        value == "end_repeat" and header_name == "type"
                    )

                    is_core_field = isinstance(value, str) and any(
                        value.endswith(i) for i in core_field_suffixes
                    )

                    if is_end_repeat_info or is_core_field:
                        add_flag = False
                        break

                    if any(header_name.startswith(i) for i in language_fields):
                        field_name, language = header_name.split(":")
                        cleared_value = (
                            strip_tags(value).replace("#", "").strip()
                        )
                        json_fields[field_name].update(
                            {language: cleared_value if value else ""}
                        )
                    elif header_name in model_fields:
                        if header_name in boolean_fields:
                            if value == "true":
                                fields[header_name] = True
                            else:
                                fields[header_name] = False
                        else:
                            fields[header_name] = value if value else ""
                if add_flag:
                    all_objects.append(
                        FlexibleField(
                            location=location, **fields, **json_fields,
                        )
                    )
            FlexibleField.objects.bulk_create(all_objects)

            self.message_user(
                request,
                "Your xls file has been imported, "
                f"Created {len(all_objects)} FlexibleField objects",
            )
            return redirect("..")

        form = XLSImportForm()
        payload = {"form": form}

        return render(request, "core/xls_form.html", payload)
