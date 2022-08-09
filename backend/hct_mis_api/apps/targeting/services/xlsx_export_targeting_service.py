from functools import cached_property

from django.db.models import Q

import openpyxl
from openpyxl.utils import get_column_letter

from hct_mis_api.apps.core.utils import nested_getattr
from hct_mis_api.apps.household.models import Document, Individual
from hct_mis_api.apps.targeting.models import TargetPopulation


class XlsxExportTargetingService:
    INDIVIDUALS_SHEET = "Individuals"
    META_SHEET = "Meta"
    VERSION_CELL_NAME_COORDINATES = "A1"
    VERSION_CELL_COORDINATES = "B1"
    VERSION_CELL_NAME = "FILE_TEMPLATE_VERSION"
    VERSION = "1.0"

    def __init__(self, target_population: TargetPopulation):
        self.target_population = target_population
        self.documents_columns_dict = {}
        self.current_header_column_index = 0
        self.COLUMNS_MAPPING_DICT = {
            "Household unicef_id": "household.unicef_id",
            "unicef_id": "unicef_id",
            "Linked Households": self._render_all_linked_households,
            "Bank account information": self._bank_account_info,
        }

    @cached_property
    def households(self):
        if self.target_population.status == TargetPopulation.STATUS_DRAFT:
            return self.target_population.candidate_list
        return self.target_population.vulnerability_score_filtered_households

    @cached_property
    def individuals(self):
        return (
            Individual.objects.filter(
                Q(household__in=self.households, withdrawn=False, duplicate=False)
                | Q(households_and_roles__household__in=self.households)
            )
            .select_related("household")
            .prefetch_related("bank_account_info")
            .order_by("household__unicef_id")
            .distinct()
        )

    def generate_workbook(self):
        self._create_workbook()
        self._add_version()
        self._add_standard_columns_headers()
        self._add_individuals_rows()
        self._adjust_column_width_from_col(self.ws_individuals, 1, 1, self.current_header_column_index)
        return self.workbook

    def _add_version(self):
        self.ws_meta[
            XlsxExportTargetingService.VERSION_CELL_NAME_COORDINATES
        ] = XlsxExportTargetingService.VERSION_CELL_NAME
        self.ws_meta[XlsxExportTargetingService.VERSION_CELL_COORDINATES] = XlsxExportTargetingService.VERSION

    def _create_workbook(self) -> openpyxl.Workbook:
        workbook = openpyxl.Workbook()
        self.ws_individuals = workbook.active
        self.ws_individuals.title = XlsxExportTargetingService.INDIVIDUALS_SHEET
        self.workbook = workbook
        self.ws_meta = workbook.create_sheet(XlsxExportTargetingService.META_SHEET)
        return workbook

    def _add_standard_columns_headers(self):
        standard_columns_names = list(self.COLUMNS_MAPPING_DICT.keys())
        self.ws_individuals.append(standard_columns_names)
        self.current_header_column_index += len(standard_columns_names)

    def _add_individual_row(self, individual: Individual):
        individual_row = {}
        for index, field in enumerate(self.COLUMNS_MAPPING_DICT.values()):
            if callable(field):
                value = field(individual)
            else:
                try:
                    value = nested_getattr(individual, field)
                except AttributeError:
                    value = None
            individual_row[index + 1] = value
        self._add_individual_documents_to_row(individual, individual_row)
        self.ws_individuals.append(individual_row)

    def _add_individual_documents_to_row(self, individual: Individual, row: dict):
        document: Document
        for document in individual.documents.all():
            column_index = self._add_document_column_header(document)
            row[column_index + 1] = document.document_number

    def _add_document_column_header(self, document):
        type_string = str(document.type)
        if type_string in self.documents_columns_dict:
            return self.documents_columns_dict[type_string]
        self.documents_columns_dict[type_string] = self.current_header_column_index
        self.ws_individuals.cell(1, self.current_header_column_index + 1, type_string)
        old_header_column_index = self.current_header_column_index
        self.current_header_column_index += 1
        return old_header_column_index

    def _add_individuals_rows(self):
        for individual in self.individuals:
            self._add_individual_row(individual)

    def _render_all_linked_households(self, individual):
        roles_string_list = [
            f"{role.household.unicef_id} - {role.role}"
            for role in individual.households_and_roles.filter(household__in=self.households)
        ]
        return ",".join(roles_string_list)

    @staticmethod
    def _bank_account_info(individual):
        if individual.bank_account_info.exists():
            return ", ".join([str(bank_info) for bank_info in individual.bank_account_info.all()])
        return ""

    def _adjust_column_width_from_col(self, ws, min_row, min_col, max_col):
        column_widths = []

        for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
            for cell in col:
                value = cell.value
                if value:

                    if isinstance(value, str) is False:
                        value = str(value)

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i, width in enumerate(column_widths):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            ws.column_dimensions[col_name].width = value
