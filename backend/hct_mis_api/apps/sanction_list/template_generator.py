from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class TemplateFileGenerator:
    COLUMNS = (
        "FIRST NAME",
        "SECOND NAME",
        "THIRD NAME",
        "FOURTH NAME",
        "DATE OF BIRTH",
    )

    @classmethod
    def get_template_file(cls) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sanction List Check"
        ws.append(cls.COLUMNS)

        for i in range(1, len(cls.COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        return wb
