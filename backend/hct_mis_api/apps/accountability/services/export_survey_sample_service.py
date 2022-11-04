import logging

from django.conf import settings
from django.core.files import File
from django.core.files.temp import NamedTemporaryFile
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.urls import reverse

import openpyxl
from openpyxl.utils import get_column_letter

from hct_mis_api.apps.accountability.models import Survey
from hct_mis_api.apps.core.utils import encode_id_base64

logger = logging.getLogger(__name__)


class ExportSurveySampleService:
    WORKBOOK_TITLE = "Survey Sample"
    HEADERS = (
        "household_unicef_id",
        "head_of_household",
        "household_size",
        "admin2",
        "residence_status",
        "registration_date",
    )

    def __init__(self, survey: Survey, user):
        self.survey = survey
        self.user = user

    def export_sample(self):
        filename = f"survey_sample_{self.survey.unicef_id}.xlsx"

        wb = self._generate_file()
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            self.survey.store_sample_file(filename, File(tmp))

    def send_email(self):
        protocol = "http" if settings.IS_DEV else "https"
        survey_id = encode_id_base64(self.survey.id, "Survey")
        api = reverse("download-survey-sample", args=[survey_id])
        link = f"{protocol}://{settings.FRONTEND_HOST}{api}"

        msg = "Survey sample xlsx file was generated and below You have the link to download this file."
        context = {
            "first_name": self.user.first_name,
            "last_name": self.user.last_name,
            "email": self.user.email,
            "message": msg,
            "link": link,
        }
        text_body = render_to_string("survey/survey_sample_xlsx_file_generated_email.txt", context=context)
        html_body = render_to_string("survey/survey_sample_xlsx_file_generated_email.html", context=context)

        email = EmailMultiAlternatives(
            subject="Survey sample XLSX file generated",
            from_email=settings.EMAIL_HOST_USER,
            to=[context["email"]],
            body=text_body,
        )
        email.attach_alternative(html_body, "text/html")
        result = email.send()
        if not result:
            logger.error(f"Email couldn't be send to {context['email']}")

    def _generate_file(self):
        wb = self._create_workbook()
        self._generate_headers()
        self._add_records()
        self._adjust_columns_width()

        return wb

    def _create_workbook(self):
        wb = openpyxl.Workbook()
        active_wb = wb.active
        active_wb.title = self.WORKBOOK_TITLE
        self.active_wb = active_wb
        return wb

    def _generate_headers(self):
        self.active_wb.append(self.HEADERS)

    def _add_records(self):
        for recipient in self.survey.recipients.all():
            self._get_record(recipient)

    def _get_record(self, recipient):
        row = (
            str(recipient.unicef_id),
            str(recipient.head_of_household.full_name),
            recipient.size,
            str(recipient.admin2.name),
            str(recipient.residence_status),
            str(recipient.last_registration_date),
        )
        self.active_wb.append(row)

    def _adjust_columns_width(self):
        min_row = 0
        min_col = 1
        max_col = 8
        column_widths = []

        for i, col in enumerate(self.active_wb.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
            for cell in col:
                value = cell.value
                if value is None:
                    continue

                if not isinstance(value, str):
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

        for i in range(len(column_widths)):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            self.active_wb.column_dimensions[col_name].width = value
