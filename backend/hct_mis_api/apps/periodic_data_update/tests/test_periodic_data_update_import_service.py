import datetime
import json
import uuid
from tempfile import NamedTemporaryFile, _TemporaryFileWrapper
from typing import Any

from django.core.exceptions import ValidationError
from django.core.files import File
from django.test import TestCase

import openpyxl

from hct_mis_api.apps.core.fixtures import (
    create_afghanistan,
    create_pdu_flexible_attribute,
)
from hct_mis_api.apps.core.models import PeriodicFieldData
from hct_mis_api.apps.household.fixtures import create_household_and_individuals
from hct_mis_api.apps.periodic_data_update.models import (
    PeriodicDataUpdateTemplate,
    PeriodicDataUpdateUpload,
)
from hct_mis_api.apps.periodic_data_update.service.periodic_data_update_export_template_service import (
    PeriodicDataUpdateExportTemplateService,
)
from hct_mis_api.apps.periodic_data_update.service.periodic_data_update_import_service import (
    PeriodicDataUpdateImportService,
)
from hct_mis_api.apps.periodic_data_update.utils import populate_pdu_with_null_values
from hct_mis_api.apps.program.fixtures import ProgramFactory
from hct_mis_api.apps.program.models import Program
from hct_mis_api.apps.registration_data.fixtures import RegistrationDataImportFactory


def add_pdu_data_to_xlsx(
    periodic_data_update_template: PeriodicDataUpdateTemplate, rows: list[list[Any]]
) -> _TemporaryFileWrapper:
    wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
    ws_pdu = wb[PeriodicDataUpdateExportTemplateService.PDU_SHEET]
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            ws_pdu.cell(row=row_index + 2, column=col_index + 7, value=value)
    tmp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)
    tmp_file.seek(0)
    return tmp_file


class TestPeriodicDataUpdateImportService(TestCase):
    rdi = None
    business_area = None
    program = None

    @classmethod
    def setUpTestData(cls) -> None:
        cls.business_area = create_afghanistan()
        cls.program = ProgramFactory(name="Test Program", status=Program.DRAFT, business_area=cls.business_area)
        cls.rdi = RegistrationDataImportFactory()
        cls.household, cls.individuals = create_household_and_individuals(
            household_data={
                "business_area": cls.business_area,
                "program_id": cls.program.pk,
                "registration_data_import": cls.rdi,
            },
            individuals_data=[
                {
                    "business_area": cls.business_area,
                    "program_id": cls.program.pk,
                    "registration_data_import": cls.rdi,
                },
            ],
        )
        cls.individual = cls.individuals[0]

        cls.string_attribute = create_pdu_flexible_attribute(
            label="String Attribute",
            subtype=PeriodicFieldData.STRING,
            number_of_rounds=1,
            rounds_names=["May"],
            program=cls.program,
        )
        cls.decimal_attribute = create_pdu_flexible_attribute(
            label="Decimal Attribute",
            subtype=PeriodicFieldData.DECIMAL,
            number_of_rounds=1,
            rounds_names=["May"],
            program=cls.program,
        )
        cls.boolean_attribute = create_pdu_flexible_attribute(
            label="Boolean Attribute",
            subtype=PeriodicFieldData.BOOL,
            number_of_rounds=1,
            rounds_names=["May"],
            program=cls.program,
        )
        cls.date_attribute = create_pdu_flexible_attribute(
            label="Date Attribute",
            subtype=PeriodicFieldData.DATE,
            number_of_rounds=1,
            rounds_names=["May"],
            program=cls.program,
        )

        populate_pdu_with_null_values(cls.program, cls.individual.flex_fields)
        cls.individual.save()

    def prepare_test_data(self, rounds_data: list, rows: list) -> tuple:
        periodic_data_update_template = PeriodicDataUpdateTemplate.objects.create(
            program=self.program,
            business_area=self.business_area,
            filters=dict(),
            rounds_data=rounds_data,
        )
        service = PeriodicDataUpdateExportTemplateService(periodic_data_update_template)
        service.generate_workbook()
        service.save_xlsx_file()
        tmp_file = add_pdu_data_to_xlsx(periodic_data_update_template, rows)
        periodic_data_update_upload = PeriodicDataUpdateUpload(
            template=periodic_data_update_template,
            created_by=periodic_data_update_template.created_by,
        )
        tmp_file.seek(0)
        periodic_data_update_upload.file.save("test.xlsx", File(tmp_file))
        periodic_data_update_upload.save()
        tmp_file.close()
        return periodic_data_update_template, periodic_data_update_upload

    def test_import_data_string(self) -> None:
        flexible_attribute = self.string_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["Test Value", "2021-05-02"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service.import_data()
        self.assertEqual(periodic_data_update_upload.status, PeriodicDataUpdateUpload.Status.SUCCESSFUL)
        self.assertEqual(periodic_data_update_upload.error_message, None)
        self.individual.refresh_from_db()
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["value"], "Test Value")
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["collection_date"], "2021-05-02")

    def test_import_data_decimal(self) -> None:
        flexible_attribute = self.decimal_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["20.456", "2021-05-02"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service.import_data()
        self.assertEqual(periodic_data_update_upload.status, PeriodicDataUpdateUpload.Status.SUCCESSFUL)
        self.assertEqual(periodic_data_update_upload.error_message, None)
        self.individual.refresh_from_db()
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["value"], 20.456)
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["collection_date"], "2021-05-02")

    def test_import_data_boolean(self) -> None:
        flexible_attribute = self.boolean_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [[True, "2021-05-02"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service.import_data()
        self.assertEqual(periodic_data_update_upload.status, PeriodicDataUpdateUpload.Status.SUCCESSFUL)
        self.assertEqual(periodic_data_update_upload.error_message, None)
        self.individual.refresh_from_db()
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["value"], True)
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["collection_date"], "2021-05-02")

    def test_import_data_date(self) -> None:
        flexible_attribute = self.date_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["1996-06-21", "2021-05-02"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service.import_data()
        self.assertEqual(periodic_data_update_upload.status, PeriodicDataUpdateUpload.Status.SUCCESSFUL)
        self.assertEqual(periodic_data_update_upload.error_message, None)
        self.individual.refresh_from_db()
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["value"], "1996-06-21")
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["collection_date"], "2021-05-02")

    def test_import_data_date_no_collection_date(self) -> None:
        flexible_attribute = self.date_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["1996-06-21"]],
        )
        periodic_data_update_template.created_at = datetime.datetime(2021, 3, 7)
        periodic_data_update_template.save()
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service.import_data()
        self.assertEqual(periodic_data_update_upload.status, PeriodicDataUpdateUpload.Status.SUCCESSFUL)
        self.assertEqual(periodic_data_update_upload.error_message, None)
        self.individual.refresh_from_db()
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["value"], "1996-06-21")
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["collection_date"], "2021-03-07")

    def test_import_data_date_fail(self) -> None:
        flexible_attribute = self.date_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["bla bla", "2021-05-02"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service.import_data()
        self.assertEqual(periodic_data_update_upload.status, PeriodicDataUpdateUpload.Status.FAILED)
        errors = {
            "form_errors": [
                {
                    "row": 2,
                    "errors": {"date_attribute__round_value": [{"message": "Enter a valid date.", "code": "invalid"}]},
                }
            ],
            "non_form_errors": None,
        }
        self.assertEqual(
            json.loads(periodic_data_update_upload.error_message),
            errors,
        )

    def test_read_periodic_data_update_non_form_errors(self) -> None:
        flexible_attribute = self.date_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["2021-05-02", "2021-05-02"]],
        )
        service = PeriodicDataUpdateExportTemplateService(periodic_data_update_template)
        service.generate_workbook()
        service.save_xlsx_file()
        flexible_attribute.delete()
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service.import_data()
        self.assertEqual(periodic_data_update_upload.status, PeriodicDataUpdateUpload.Status.FAILED)
        errors = {
            "form_errors": [],
            "non_form_errors": ["Some fields are missing in the flexible attributes"],
        }
        self.assertEqual(
            json.loads(periodic_data_update_upload.error_message),
            errors,
        )

    def test_read_periodic_data_update_template_object(self) -> None:
        periodic_data_update_template = PeriodicDataUpdateTemplate.objects.create(
            program=self.program,
            business_area=self.business_area,
            filters=dict(),
            rounds_data=[
                {
                    "field": self.string_attribute.name,
                    "round": 1,
                    "round_name": self.string_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
        )
        service = PeriodicDataUpdateExportTemplateService(periodic_data_update_template)
        service.generate_workbook()
        service.save_xlsx_file()
        periodic_data_update_template_from_xlsx = (
            PeriodicDataUpdateImportService.read_periodic_data_update_template_object(
                periodic_data_update_template.file.file
            )
        )
        self.assertEqual(periodic_data_update_template_from_xlsx.pk, periodic_data_update_template.pk)
        wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
        del wb.custom_doc_props[PeriodicDataUpdateExportTemplateService.PROPERTY_ID_NAME]
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            periodic_data_update_template_from_xlsx = (
                PeriodicDataUpdateImportService.read_periodic_data_update_template_object(tmp_file)
            )
            self.assertEqual(periodic_data_update_template_from_xlsx.pk, periodic_data_update_template.pk)
        wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
        del wb.custom_doc_props[PeriodicDataUpdateExportTemplateService.PROPERTY_ID_NAME]
        ws_meta = wb[PeriodicDataUpdateExportTemplateService.META_SHEET]
        ws_meta[PeriodicDataUpdateExportTemplateService.META_ID_ADDRESS] = ""
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            with self.assertRaisesMessage(ValidationError, "Periodic Data Update Template ID is missing in the file"):
                PeriodicDataUpdateImportService.read_periodic_data_update_template_object(tmp_file)

        wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
        del wb.custom_doc_props[PeriodicDataUpdateExportTemplateService.PROPERTY_ID_NAME]
        ws_meta = wb[PeriodicDataUpdateExportTemplateService.META_SHEET]
        ws_meta[PeriodicDataUpdateExportTemplateService.META_ID_ADDRESS] = "abc"
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            with self.assertRaisesMessage(ValidationError, "Periodic Data Update Template ID must be a number"):
                PeriodicDataUpdateImportService.read_periodic_data_update_template_object(tmp_file)

        wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
        del wb.custom_doc_props[PeriodicDataUpdateExportTemplateService.PROPERTY_ID_NAME]
        ws_meta = wb[PeriodicDataUpdateExportTemplateService.META_SHEET]
        ws_meta[PeriodicDataUpdateExportTemplateService.META_ID_ADDRESS] = True
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            with self.assertRaisesMessage(ValidationError, "Periodic Data Update Template ID must be an integer"):
                PeriodicDataUpdateImportService.read_periodic_data_update_template_object(tmp_file)

        wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
        del wb.custom_doc_props[PeriodicDataUpdateExportTemplateService.PROPERTY_ID_NAME]
        ws_meta = wb[PeriodicDataUpdateExportTemplateService.META_SHEET]
        ws_meta[PeriodicDataUpdateExportTemplateService.META_ID_ADDRESS] = -1
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            with self.assertRaisesMessage(ValidationError, "Periodic Data Update Template with ID -1 not found"):
                PeriodicDataUpdateImportService.read_periodic_data_update_template_object(tmp_file)

    def test_read_flexible_attributes(self) -> None:
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": self.string_attribute.name,
                    "round": 1,
                    "round_name": self.string_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                },
                {
                    "field": "Not existing field",
                    "round": 1,
                    "round_name": self.string_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                },
            ],
            [["1996-06-21", "2021-05-02"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service._open_workbook()
        with self.assertRaisesMessage(ValidationError, "Some fields are missing in the flexible attributes"):
            service._read_flexible_attributes()

    def test_read_row(self) -> None:
        flexible_attribute = self.date_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["-", "-"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service._open_workbook()
        service._read_flexible_attributes()
        service._read_rows()

    def test_import_cleaned_data(self) -> None:
        flexible_attribute = self.date_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["-", "-"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service._open_workbook()
        service._read_flexible_attributes()
        cleaned_data = {
            "individual__uuid": self.individual.id,
            "individual_unicef_id": self.individual.unicef_id,
            "first_name": "Debra",
            "last_name": "Taylor",
            "date_attribute__round_number": 2,
            "date_attribute__round_name": "May",
            "date_attribute__round_value": datetime.date(1996, 6, 21),
            "date_attribute__collection_date": datetime.date(2021, 5, 2),
        }
        with self.assertRaisesMessage(
            ValidationError,
            f"Round number mismatch for field date_attribute and individual {self.individual.unicef_id}",
        ):
            service._import_cleaned_data(cleaned_data)
        not_existing_individual_id = uuid.uuid4()
        cleaned_data = {
            "individual__uuid": not_existing_individual_id,
            "individual_unicef_id": self.individual.unicef_id,
            "first_name": "Debra",
            "last_name": "Taylor",
            "date_attribute__round_number": 1,
            "date_attribute__round_name": "May",
            "date_attribute__round_value": datetime.date(1996, 6, 21),
            "date_attribute__collection_date": datetime.date(2021, 5, 2),
        }
        with self.assertRaisesMessage(
            ValidationError,
            f"Individual not found for {self.individual.unicef_id}",
        ):
            service._import_cleaned_data(cleaned_data)
        self.individual.flex_fields = {
            "date_attribute": {"1": {"value": "1996-06-21", "collection_date": "2021-05-02"}}
        }
        self.individual.save()
        cleaned_data = {
            "individual__uuid": self.individual.id,
            "individual_unicef_id": self.individual.unicef_id,
            "first_name": "Debra",
            "last_name": "Taylor",
            "date_attribute__round_number": 1,
            "date_attribute__round_name": "May",
            "date_attribute__round_value": datetime.date(1996, 6, 21),
            "date_attribute__collection_date": datetime.date(2021, 5, 2),
        }
        with self.assertRaisesMessage(
            ValidationError,
            f"Value already exists for field date_attribute for round 1 and individual {self.individual.unicef_id}",
        ):
            service._import_cleaned_data(cleaned_data)

    def test_set_round_value(self) -> None:
        flexible_attribute = self.date_attribute
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["-", "-"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        service._open_workbook()
        service._read_flexible_attributes()
        service.set_round_value(self.individual, flexible_attribute.name, 1, "1996-06-21", "2021-05-02")
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["value"], "1996-06-21")
        self.assertEqual(self.individual.flex_fields[flexible_attribute.name]["1"]["collection_date"], "2021-05-02")

    def test_get_form_field_for_value(self) -> None:
        flexible_attribute = self.date_attribute
        pdu_data = flexible_attribute.pdu_data
        PeriodicFieldData.objects.filter(id=pdu_data.id).update(subtype="INVALID")
        flexible_attribute.refresh_from_db()
        periodic_data_update_template, periodic_data_update_upload = self.prepare_test_data(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["-", "-"]],
        )
        service = PeriodicDataUpdateImportService(periodic_data_update_upload)
        with self.assertRaisesMessage(
            ValidationError,
            f"Invalid subtype for field {flexible_attribute.name}",
        ):
            service._get_form_field_for_value(flexible_attribute)
