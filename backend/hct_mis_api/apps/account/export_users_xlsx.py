from collections import OrderedDict
from typing import Optional

from django.db import transaction

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from hct_mis_api.apps.account.models import User


class GenericField:
    def __init__(self, name: str, column_name: str) -> None:
        self.name = name
        self.column_name = column_name

    def value(self, user: User, business_area: str) -> str:
        return getattr(user, self.name)


class PartnerField(GenericField):
    def value(self, user: User, business_area: str) -> str:
        if not user.partner:
            return ""
        return user.partner.name


class UserRoleField(GenericField):
    def value(self, user: User, business_area: str) -> str:
        all_roles = user.user_roles.filter(business_area__slug=business_area)
        return ", ".join([f"{role.business_area.name}-{role.role.name}" for role in all_roles])


class ExportUsersXlsx:
    FIELDS_TO_COLUMNS_MAPPING = OrderedDict(
        {
            "first_name": GenericField("first_name", "FIRST NAME"),
            "last_name": GenericField("last_name", "LAST NAME"),
            "email": GenericField("email", "E-MAIL"),
            "status": GenericField("status", "ACCOUNT STATUS"),
            "partner": PartnerField("partner", "PARTNER"),
            "user_roles": UserRoleField("user_roles", "USER ROLES"),
        }
    )

    def __init__(self, business_area_slug) -> None:
        self.business_area_slug = business_area_slug
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Exported Users to Cash Assist"
        self._add_headers()

    def _add_headers(self) -> None:
        self.ws.append([field.column_name for field in self.FIELDS_TO_COLUMNS_MAPPING.values()])
        for i in range(1, len(self.FIELDS_TO_COLUMNS_MAPPING) + 1):
            self.ws.column_dimensions[get_column_letter(i)].width = 20

    @transaction.atomic(using="default")
    def get_exported_users_file(self) -> Optional[Workbook]:
        fields = self.FIELDS_TO_COLUMNS_MAPPING.values()
        users = (
            User.objects.prefetch_related("user_roles")
            .select_related("partner")
            .filter(is_superuser=False, user_roles__business_area__slug=self.business_area_slug)
        )
        if users.exists() is False:
            return

        for user in users.iterator(chunk_size=2000):
            self.ws.append([field.value(user, self.business_area_slug) for field in fields])

        return self.wb
