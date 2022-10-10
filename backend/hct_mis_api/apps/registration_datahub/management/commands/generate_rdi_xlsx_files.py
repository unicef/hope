import os
import random
import shutil

from django.conf import settings
from django.core.management import BaseCommand

import openpyxl
from faker import Faker

faker = Faker()

random_number = lambda: random.randint(1, 2**31)
date = lambda: faker.date_between(start_date="-30y", end_date="today")
name = lambda: faker.name()
phone_number = lambda: faker.phone_number()

household_ids = []

household_header_mapping = {
    "A": ("household_id", random_number),
    "B": ("residence_status_h_c", "REFUGEE"),
    "C": ("consent_h_c", "TRUE"),
    "D": ("consent_sign_h_c", None),
    "E": ("consent_origin_h_c", "POL"),
    "F": ("country_h_c", "POL"),
    "G": ("address_h_c", ""),  # SPECIAL
    "H": ("admin1_h_c", "AF11"),
    "I": ("admin2_h_c", "AF1115"),
    "J": ("hh_geopoint_h_c", "70.210209, 172.085021"),
    "K": ("unhcr_hh_id_h_c", "something"),
    "L": ("returnee_h_c", False),
    "M": ("size_h_c", 1),
    "N": ("first_registration_date_h_c", date),
    "O": ("pregnant_member_h_c", 0),
    "P": ("f_0_5_age_group_h_c", 0),
    "Q": ("f_6_11_age_group_h_c", 0),
    "R": ("f_12_17_age_group_h_c", 0),
    "S": ("f_adults_h_c", 0),
    "T": ("f_pregnant_h_c", 0),
    "U": ("m_0_5_age_group_h_c", 0),
    "V": ("m_6_11_age_group_h_c", 0),
    "W": ("m_12_17_age_group_h_c", 0),
    "X": ("m_adults_h_c", 1),
    "Y": ("f_0_5_disability_h_c", 0),
    "Z": ("f_6_11_disability_h_c", 0),
    "AA": ("f_12_17_disability_h_c", 0),
    "AB": ("f_adults_disability_h_c", 0),
    "AC": ("m_0_5_disability_h_c", 0),
    "AD": ("m_6_11_disability_h_c", 0),
    "AE": ("m_12_17_disability_h_c", 0),
    "AF": ("m_adults_disability_h_c", 0),
    "AG": ("fchild_hoh_h_c", False),
    "AH": ("child_hoh_h_c", False),
    "AI": ("village_h_c", None),
    "AJ": ("start_h_c", date),
    "AK": ("end_h_c", date),
    "AL": ("deviceid_h_c", None),
    "AM": ("name_enumerator_h_c", name),
    "AN": ("org_enumerator_h_c", "UNICEF"),
    "AO": ("consent_sharing_h_c", "HUMANITARIAN_PARTNER"),
    "AP": ("org_name_enumerator_h_c", "UNICEF"),
    "AQ": ("unaccompanied_child_h_f", None),
    "AR": ("recent_illness_child_h_f", None),
    "AS": ("difficulty_breathing_h_f", None),
    "AT": ("treatment_h_f", None),
    "AU": ("treatment_facility_h_f", None),
    "AV": ("other_treatment_facility_h_f", None),
    "AW": ("living_situation_h_f", None),
    "AX": ("number_of_rooms_h_f", None),
    "AY": ("total_dwellers_h_f", None),
    "AZ": ("one_room_dwellers_h_f", None),
    "BA": ("total_households_h_f", None),
    "BB": ("water_source_h_f", None),
    "BC": ("sufficient_water_h_f", None),
    "BD": ("latrine_h_f", None),
    "BE": ("meals_yesterday_h_f", None),
    "BF": ("food_consumption_h_f", None),
    "BG": ("cereals_h_f", None),
    "BH": ("tubers_roots_h_f", None),
    "BI": ("vegetables_h_f", None),
    "BJ": ("fruits_h_f", None),
    "BK": ("meat_fish_h_f", None),
    "BL": ("pulses_h_f", None),
    "BM": ("dairy_h_f", None),
    "BN": ("months_displaced_h_f", 6),
    "BO": ("oilfat_h_f", None),
    "BP": ("sugarsweet_h_f", None),
    "BQ": ("condiments_h_f", None),
    "BR": ("assistance_type_h_f", None),
    "BS": ("assistance_source_h_f", None),
    "BT": ("collect_individual_data_h_c", "1"),  # 1 == YES == FULL
}

individual_header_mapping = {
    "A": ("household_id", ""),  # SPECIAL
    "B": ("age", random_number),
    "C": ("relationship_i_c", "HEAD"),
    "D": ("full_name_i_c", name),
    "E": ("given_name_i_c", None),
    "F": ("middle_name_i_c", None),
    "G": ("family_name_i_c", name),
    "H": ("gender_i_c", "MALE"),
    "I": ("birth_date_i_c", date),
    "J": ("first_registration_date_i_c", date),
    "K": ("estimated_birth_date_i_c", False),
    "L": ("photo_i_c", None),
    "M": ("marital_status_i_c", "SINGLE"),
    "N": ("phone_no_i_c", phone_number),
    "O": ("phone_no_alternative_i_c", None),
    "P": ("birth_certificate_no_i_c", None),
    "Q": ("birth_certificate_issuer_i_c", None),
    "R": ("birth_certificate_photo_i_c", None),
    "S": ("drivers_license_no_i_c", None),
    "T": ("drivers_license_issuer_i_c", None),
    "U": ("drivers_license_photo_i_c", None),
    "V": ("electoral_card_no_i_c", None),
    "W": ("electoral_card_issuer_i_c", None),
    "X": ("electoral_card_photo_i_c", None),
    "Y": ("unhcr_id_no_i_c", None),
    "Z": ("unhcr_id_issuer_i_c", None),
    "AA": ("unhcr_id_photo_i_c", None),
    "AB": ("national_passport_i_c", None),
    "AC": ("national_passport_issuer_i_c", None),
    "AD": ("national_passport_photo_i_c", None),
    "AE": ("national_id_no_i_c", random_number),
    "AF": ("national_id_issuer_i_c", "POL"),
    "AG": ("national_id_photo_i_c", None),
    "AH": ("scope_id_no_i_c", None),
    "AI": ("scope_id_issuer_i_c", None),
    "AJ": ("scope_id_photo_i_c", None),
    "AK": ("other_id_type_i_c", None),
    "AL": ("other_id_no_i_c", None),
    "AM": ("other_id_issuer_i_c", None),
    "AN": ("pregnant_i_c", False),
    "AO": ("work_status_i_c", "NOT_PROVIDED"),
    "AP": ("observed_disability_i_c", None),
    "AQ": ("seeing_disability_i_c", None),
    "AR": ("hearing_disability_i_c", None),
    "AS": ("physical_disability_i_c", None),
    "AT": ("memory_disability_i_c", None),
    "AU": ("selfcare_disability_i_c", None),
    "AV": ("comms_disability_i_c", None),
    "AW": ("who_answers_phone_i_c", None),
    "AX": ("who_answers_alt_phone_i_c", None),
    "AY": ("primary_collector_id", ""),  # SPECIAL
    "AZ": ("alternate_collector_id", None),
    "BA": ("photo_i_f", None),
    "BB": ("id_type_i_f", None),
    "BC": ("other_id_type_i_f", None),
    "BD": ("id_no_i_f", None),
    "BE": ("seeing_disability_i_f", None),
    "BF": ("hearing_disability_i_f", None),
    "BG": ("physical_disability_i_f", None),
    "BH": ("memory_disability_i_f", None),
    "BI": ("selfcare_disability_i_f", None),
    "BJ": ("comms_disability_i_f", None),
    "BK": ("child_marital_status_i_f", None),
    "BL": ("marriage_age_i_f", None),
    "BM": ("school_enrolled_before_i_f", 1),
    "BN": ("school_enrolled_i_f", 0),
    "BO": ("school_type_i_f", None),
    "BP": ("years_in_school_i_f", None),
    "BQ": ("minutes_to_school_i_f", None),
}


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument(
            "amount",
            default=1,
            action="store",
            nargs="?",
            type=int,
        )

        parser.add_argument(
            "--seed",
            default=None,
            action="store",
            nargs="?",
            type=int,
        )

    def handle(self, *args, **options):
        amount = options["amount"]
        seed = options["seed"]
        print(f"Generating xlsx file ({amount}x HHs & INDs) with seed {seed}")

        generated_dir = os.path.join(settings.PROJECT_ROOT, "..", "generated")
        if not os.path.exists(generated_dir):
            os.makedirs(generated_dir)

        filepath = os.path.join(generated_dir, f"rdi_import_{amount}_hh_{amount}_ind_seed_{seed}.xlsx")
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.get_sheet_by_name(wb.get_sheet_names()[0]))

        households = wb.create_sheet("Households")
        for index, (_, (header, _)) in enumerate(household_header_mapping.items()):
            households.cell(row=1, column=index + 1).value = header
            households.cell(row=2, column=index + 1).value = "description"
        for count in range(amount):
            for index, (key, (_, value)) in enumerate(household_header_mapping.items()):
                if value is None:
                    continue
                if key == "G":
                    to_write = seed  # address is seed (for targeting filter by address)
                else:
                    to_write = value() if callable(value) else value
                households.cell(row=3 + count, column=index + 1).value = to_write
                if key == "A":
                    household_ids.append(to_write)

        individuals = wb.create_sheet("Individuals")
        for index, (_, (header, _)) in enumerate(individual_header_mapping.items()):
            individuals.cell(row=1, column=index + 1).value = header
            individuals.cell(row=2, column=index + 1).value = "description"
        for count in range(amount):
            for index, (key, (_, value)) in enumerate(individual_header_mapping.items()):
                if value is None:
                    continue
                if key in ["AY", "A"]:
                    to_write = household_ids[count]
                elif callable(value):
                    to_write = value()
                else:
                    to_write = value
                individuals.cell(row=3 + count, column=index + 1).value = to_write

        wb.save(filepath)
