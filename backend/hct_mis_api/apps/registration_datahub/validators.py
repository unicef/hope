import itertools
import logging
import re
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from operator import itemgetter
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Set, Tuple, Union
from zipfile import BadZipfile

from django.core import validators as django_core_validators

import openpyxl
import phonenumbers
from dateutil import parser
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from PIL.Image import Image

from hct_mis_api.apps.core.field_attributes.core_fields_attributes import (
    TYPE_SELECT_MANY,
    TYPE_SELECT_ONE,
    FieldFactory,
)
from hct_mis_api.apps.core.field_attributes.fields_types import (
    _DELIVERY_MECHANISM_DATA,
    _INDIVIDUAL,
    Scope,
)
from hct_mis_api.apps.core.kobo.common import (
    KOBO_FORM_INDIVIDUALS_COLUMN_NAME,
    get_field_name,
)
from hct_mis_api.apps.core.models import (
    BusinessArea,
    FlexibleAttribute,
    PeriodicFieldData,
)
from hct_mis_api.apps.core.utils import (
    SheetImageLoader,
    rename_dict_keys,
    serialize_flex_attributes,
)
from hct_mis_api.apps.core.validators import BaseValidator
from hct_mis_api.apps.geo.models import Area
from hct_mis_api.apps.household.models import (
    HEAD,
    NON_BENEFICIARY,
    ROLE_ALTERNATE,
    ROLE_PRIMARY,
)
from hct_mis_api.apps.payment.models import DeliveryMechanism, DeliveryMechanismData
from hct_mis_api.apps.program.models import Program
from hct_mis_api.apps.registration_data.models import KoboImportedSubmission
from hct_mis_api.apps.registration_datahub.tasks.utils import collectors_str_ids_to_list
from hct_mis_api.apps.registration_datahub.utils import (
    calculate_hash_for_kobo_submission,
    find_attachment_in_kobo,
)

logger = logging.getLogger(__name__)


class XlsxException(Exception):
    def __init__(self, errors: List) -> None:
        self.errors = errors


class XLSXValidator(BaseValidator):
    @classmethod
    def validate(cls, excluded_validators: Optional[Any] = None, *args: Any, **kwargs: Any) -> None:
        validate_methods: List[Callable] = [getattr(cls, m) for m in dir(cls) if m.startswith("validate_")]

        errors_list = []
        for method in validate_methods:
            errors = method(*args, **kwargs)
            errors_list.extend(errors)

        if errors_list:
            errors_list.sort(key=itemgetter("header"))
            raise XlsxException(errors_list)

    @classmethod
    def validate_file_extension(cls, *args: Any, **kwargs: Any) -> List:
        try:
            xlsx_file = kwargs["file"]
            file_suffix = Path(xlsx_file.name).suffix
            if file_suffix != ".xlsx":
                return [
                    {
                        "row_number": 1,
                        "header": f"{xlsx_file.name}",
                        "message": "Only .xlsx files are accepted for import",
                    }
                ]

            # Checking only extensions is not enough,
            # loading workbook to check if it is in fact true .xlsx file
            try:
                load_workbook(xlsx_file, data_only=True)
            except BadZipfile:
                return [{"row_number": 1, "header": f"{xlsx_file.name}", "message": "Invalid .xlsx file"}]

            return []
        except Exception as e:
            logger.exception(e)
            raise


class ImportDataInstanceValidator:
    business_area_code = None
    DOCUMENTS_ISSUING_COUNTRIES_MAPPING = {
        "birth_certificate_issuer_i_c": "birth_certificate_no_i_c",
        "drivers_license_issuer_i_c": "drivers_license_no_i_c",
        "electoral_card_issuer_i_c": "electoral_card_no_i_c",
        "national_id_issuer_i_c": "national_id_no_i_c",
        "national_passport_issuer_i_c": "national_passport_no_i_c",
        "tax_id_issuer_i_c": "tax_id_no_i_c",
        "other_id_issuer_i_c": "other_id_type_i_c",
        # identities
        "scope_id_issuer_i_c": "scope_id_no_i_c",
        "unhcr_id_issuer_i_c": "unhcr_id_no_i_c",
    }

    def __init__(self, program: Program) -> None:
        self.is_social_worker_program = program.is_social_worker_program
        self.all_fields = self.get_all_fields()
        self.delivery_mechanisms_xlsx_fields = DeliveryMechanismData.get_all_delivery_mechanisms_fields(
            by_xlsx_name=True
        )
        if self.is_social_worker_program:
            self.delivery_mechanisms_xlsx_fields = [f"pp_{field}" for field in self.delivery_mechanisms_xlsx_fields]

    def get_combined_attributes(self) -> Dict:
        scope_list = (
            [Scope.GLOBAL, Scope.XLSX, Scope.HOUSEHOLD_ID, Scope.DELIVERY_MECHANISM]
            if not self.is_social_worker_program
            else [Scope.XLSX_PEOPLE, Scope.DELIVERY_MECHANISM]
        )
        fields = FieldFactory.from_scopes(scope_list).apply_business_area()

        for field in fields:
            field["choices"] = [x.get("value") for x in field["choices"]]

        flex_attrs = self.serialize_flex_attributes()
        return {
            **fields.associated_with_household().to_dict_by("xlsx_field"),
            **flex_attrs["individuals"],
            **fields._associated_with([_INDIVIDUAL, _DELIVERY_MECHANISM_DATA]).to_dict_by("xlsx_field"),
            **flex_attrs["households"],
        }

    def serialize_flex_attributes(self) -> Dict:
        from hct_mis_api.apps.core.models import FlexibleAttribute

        flex_attributes = FlexibleAttribute.objects.prefetch_related("choices").all()

        result_dict = {
            "individuals": {},
            "households": {},
        }

        for attr in flex_attributes:
            associated_with = "Household" if attr.associated_with == 0 else "Individual"
            dict_key = associated_with.lower() + "s"

            result_dict[dict_key][attr.name] = {
                "id": attr.id,
                "type": attr.type,
                "name": attr.name,
                "xlsx_field": attr.name,
                "lookup": attr.name,
                "required": attr.required,
                "label": attr.label,
                "hint": attr.hint,
                "choices": attr.choices.values_list("name", flat=True),
                "associated_with": associated_with,
            }

        return result_dict

    def get_all_fields(self) -> Dict:
        try:
            return self.get_combined_attributes()
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def documents_validator(self, documents_numbers_dict: Dict, is_xlsx: bool = True) -> List:
        try:
            invalid_rows = []
            for key, values in documents_numbers_dict.items():
                key_name = "pp_" + key if self.is_social_worker_program else key
                if key == "other_id_no_i_c":
                    continue
                issuing_countries = values.get("issuing_countries")
                if not issuing_countries:
                    issuing_countries = [None] * len(values["validation_data"])
                if key == "other_id_type_i_c":
                    for name, value, validation_data, issuing_country in zip(
                        values["names"], values["numbers"], values["validation_data"], issuing_countries
                    ):
                        row_number = validation_data.get("row_number")
                        if not name and value:
                            error = {
                                "header": key_name,
                                "message": f"Name for {key_name} is required, when number is provided: no: {value}",
                            }
                            if is_xlsx is True:
                                error["row_number"] = row_number
                            invalid_rows.append(error)
                        if name and not value:
                            error = {
                                "header": key_name,
                                "message": f"Number for {key_name} is required, when name is provided",
                            }
                            if is_xlsx is True:
                                error["row_number"] = row_number
                            invalid_rows.append(error)
                        if (name or value) and not issuing_country:
                            error = {
                                "header": key_name,
                                "message": f"Issuing country for {key_name} is required, "
                                "when any document data are provided",
                            }
                            if is_xlsx is True:
                                error["row_number"] = row_number
                            invalid_rows.append(error)
                else:
                    for validation_data, value, issuing_country in zip_longest(
                        values["validation_data"], values["numbers"], issuing_countries
                    ):
                        row_number = (
                            validation_data.get("row_number") if isinstance(validation_data, dict) else validation_data
                        )
                        if value and not issuing_country:
                            error = {
                                "header": key_name,
                                "message": f"Issuing country for {key_name} is required, when any document data are provided",
                            }
                            if is_xlsx is True:
                                error["row_number"] = row_number
                            invalid_rows.append(error)
                        elif issuing_country and not value:
                            error = {
                                "header": key_name,
                                "message": f"Number for {key_name} is required, when issuing country is provided",
                            }
                            if is_xlsx is True:
                                error["row_number"] = row_number
                            invalid_rows.append(error)

            return invalid_rows
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def identity_validator(self, identities_numbers_dict: Dict, is_xlsx: bool = True) -> List[Dict[str, Any]]:
        try:
            invalid_rows = []
            for key, values in identities_numbers_dict.items():
                key_name = "pp_" + key if self.is_social_worker_program else key
                issuing_countries = values.get("issuing_countries")
                if not issuing_countries:
                    issuing_countries = [None] * len(values["validation_data"])
                for data_dict, value, issuing_country in zip_longest(
                    values["validation_data"], values["numbers"], issuing_countries
                ):
                    row_number = data_dict.get("row_number") if isinstance(data_dict, dict) else data_dict
                    if not value and not issuing_country:
                        continue
                    elif value and not issuing_country:
                        error = {
                            "header": key_name,
                            "message": f"Issuing country is required: partner: {values['partner']} no: {value}",
                        }
                        if is_xlsx is True:
                            error["row_number"] = row_number
                        invalid_rows.append(error)
                    elif issuing_country and not value:
                        error = {
                            "header": key_name,
                            "message": f"Number for {key_name} is required, when issuing country is provided",
                        }
                        if is_xlsx is True:
                            error["row_number"] = row_number
                        invalid_rows.append(error)

            return invalid_rows
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def delivery_mechanisms_validator(self, xlsx_delivery_mechanisms_dict: Dict) -> List[Dict[str, Any]]:
        delivery_mechanisms_to_required_fields_mapping = (
            DeliveryMechanism.get_delivery_mechanisms_to_xlsx_fields_mapping()
        )
        if self.is_social_worker_program:
            delivery_mechanisms_to_required_fields_mapping = {
                dm: [f"pp_{field}" for field in fields]
                for dm, fields in delivery_mechanisms_to_required_fields_mapping.items()
            }
        global_scope_xlsx_fields = list(
            FieldFactory.not_from_scope(Scope.DELIVERY_MECHANISM).to_dict_by("xlsx_field").keys()
        )

        try:
            all_rows_delivery_mechanisms_errors = []
            for row_number, data in xlsx_delivery_mechanisms_dict.items():
                delivery_mechanisms_errors = []
                delivery_mechanisms_fields_values_dict = defaultdict(dict)

                for delivery_mechanism_xlsx_field_name, value in data.items():
                    if not value:
                        continue
                    for dm, fields in delivery_mechanisms_to_required_fields_mapping.items():
                        if delivery_mechanism_xlsx_field_name in fields:
                            delivery_mechanisms_fields_values_dict[dm][delivery_mechanism_xlsx_field_name] = value

                dm_to_drop = []
                # drop delivery mechanism data validation for delivery mechanisms that contains only Scope.GLOBAL fields
                for dm, fields in delivery_mechanisms_fields_values_dict.items():  # type: ignore
                    # if all fields are Scope.GLOBAL, drop delivery mechanism data
                    if all([field in global_scope_xlsx_fields for field in fields.keys()]):
                        dm_to_drop.append(dm)
                for dm in dm_to_drop:
                    delivery_mechanisms_fields_values_dict.pop(dm)

                for dm, fields in delivery_mechanisms_to_required_fields_mapping.items():
                    if dm not in delivery_mechanisms_fields_values_dict:
                        continue

                    for field in fields:
                        if not delivery_mechanisms_fields_values_dict[dm].get(field, None):
                            delivery_mechanisms_errors.append(
                                {
                                    "row_number": row_number,
                                    "header": field,
                                    "message": f"Field {field} is required for delivery mechanism {dm}",
                                }
                            )
                all_rows_delivery_mechanisms_errors.extend(delivery_mechanisms_errors)

            return all_rows_delivery_mechanisms_errors
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise


class UploadXLSXInstanceValidator(ImportDataInstanceValidator):
    def __init__(self, program: Program) -> None:
        super().__init__(program)
        self.is_social_worker_program = program.is_social_worker_program
        self.head_of_household_count = defaultdict(int)
        self.combined_fields = self.get_combined_fields()
        self.household_ids = []

        self.errors = []
        self.delivery_mechanisms_errors = []
        self.pdu_flexible_attributes = FlexibleAttribute.objects.filter(
            type=FlexibleAttribute.PDU, program=program
        ).select_related("pdu_data")

    def get_combined_fields(self) -> Dict:
        core_fields = (
            FieldFactory.from_scopes([Scope.GLOBAL, Scope.XLSX, Scope.HOUSEHOLD_ID, Scope.DELIVERY_MECHANISM])
            if not self.is_social_worker_program
            else FieldFactory.from_scopes([Scope.XLSX_PEOPLE, Scope.DELIVERY_MECHANISM])
        )
        # TODO: update flex field for People
        flex_fields = serialize_flex_attributes()
        if self.is_social_worker_program:
            return {
                "people": {
                    **core_fields._associated_with([_INDIVIDUAL, _DELIVERY_MECHANISM_DATA]).to_dict_by("xlsx_field"),
                    **flex_fields["individuals"],
                },
            }
        else:
            return {
                "households": {
                    **core_fields.associated_with_household().to_dict_by("xlsx_field"),
                    **flex_fields["households"],
                },
                "individuals": {
                    **core_fields._associated_with([_INDIVIDUAL, _DELIVERY_MECHANISM_DATA]).to_dict_by("xlsx_field"),
                    **flex_fields["individuals"],
                },
            }

    def string_validator(self, value: Any, header: str, *args: Any, **kwargs: Any) -> Optional[bool]:
        try:
            if not self.required_validator(value, header, *args, **kwargs):
                return False
            if value is None:
                return True
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise
        return True

    def integer_validator(self, value: Any, header: str, *args: Any, **kwargs: Any) -> Optional[bool]:
        try:
            if not self.required_validator(value, header, *args, **kwargs):  # pragma: no cover
                return False

            if value is None:
                return True

            try:
                int(value)
                return True
            # need to use Exception because of how Graphene catches errors
            except Exception:
                return False
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def float_validator(self, value: Any, header: str, *args: Any, **kwargs: Any) -> bool:
        try:
            if not self.required_validator(value, header, *args, **kwargs):  # pragma: no cover
                return False
            if value is None:
                return True
            Decimal(value)
            return True
        except InvalidOperation:
            return False
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def decimal_validator(self, value: Any, header: str, *args: Any, **kwargs: Any) -> bool:
        try:
            if not self.required_validator(value, header, *args, **kwargs):  # pragma: no cover
                return False
            if value is None:
                return True
            Decimal(value)
            return True
        except InvalidOperation:
            return False
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def geolocation_validator(self, value: str, header: str, *args: Any, **kwargs: Any) -> bool:
        try:
            if not self.required_validator(value, header, *args, **kwargs):  # pragma: no cover
                return False
            if value is None:
                return True

            pattern = re.compile(r"^(-?\d+\.\d+?,\s*-?\d+\.\d+?)$")
            return bool(re.match(pattern, value))
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def date_validator(self, value: Any, header: str, *args: Any, **kwargs: Any) -> bool:
        try:
            if not self.required_validator(value, header, *args, **kwargs):  # pragma: no cover
                return False

            if not self.not_empty_validator(value):
                return True

            if self.integer_validator(value, header, *args, **kwargs):
                return False

            if isinstance(value, datetime):
                return True

            try:
                parser.parse(value)
            # need to use Exception because of how Graphene catches errors
            except Exception:
                return False
            return True
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def phone_validator(self, value: str, header: str, *args: Any, **kwargs: Any) -> bool:
        try:
            if not self.required_validator(value, header, *args, **kwargs):  # pragma: no cover
                return False
            if value is None:  # pragma: no cover
                return True

            try:
                phonenumbers.parse(value, None)
                return True
            except (phonenumbers.NumberParseException, TypeError):
                return False
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def choice_validator(self, value: str, header: str, *args: Any, **kwargs: Any) -> bool:
        try:
            field = self.all_fields.get(header)
            if field is None:
                return False

            if not self.required_validator(value, header, *args, **kwargs):  # pragma: no cover
                return False
            if value is None:
                return True

            choices = field["choices"]

            choice_type = self.all_fields[header]["type"]

            if choice_type == TYPE_SELECT_ONE:
                if isinstance(value, str):
                    return value.strip() in choices
                else:
                    if value not in choices:
                        str_value = str(value)
                        return str_value in choices
                return False

            elif choice_type == TYPE_SELECT_MANY:
                if isinstance(value, str):
                    if "," in value:
                        selected_choices = value.split(",")
                    elif ";" in value:
                        selected_choices = value.split(";")
                    else:
                        selected_choices = value.split(" ")
                else:
                    selected_choices = [value]
                for choice in selected_choices:
                    if isinstance(choice, str):
                        choice = choice.strip()
                        if choice in choices or choice.upper() in choices:
                            return True
                    if choice in choices:
                        return True
                return False

            return False
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def not_empty_validator(self, value: str, *args: Any, **kwargs: Any) -> bool:
        try:
            return not (value is None or value == "")
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def bool_validator(self, value: bool, header: str, *args: Any, **kwargs: Any) -> bool:
        try:
            if isinstance(value, bool):
                return True
            try:
                if self.all_fields[header]["required"] is False and (value is None or value == ""):
                    return True
            except KeyError:
                if value is None or value == "":
                    return True
            if type(value) is str:
                value = value.capitalize()
                if value in ("True", "False"):
                    return True
            return False  # pragma: no cover
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def required_validator(self, value: str, header: str, *args: Any, **kwargs: Any) -> bool:
        try:
            is_required = self.all_fields[header]["required"]
            is_not_empty = self.not_empty_validator(value)

            if is_required:
                return is_not_empty

            return True
        except KeyError:
            return True
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def image_validator(self, value: Image, header: str, cell: Cell, *args: Any, **kwargs: Any) -> Any:
        try:
            if self.required_validator(value, header, *args, **kwargs):  # pragma: no cover
                return True
            return self.image_loader.image_in(cell)  # pragma: no cover
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def rows_validator(self, sheet: Worksheet, business_area_slug: Optional[str] = None) -> None:
        try:
            first_row = sheet[1]
            combined_fields = {
                **self.combined_fields[sheet.title.lower()],
            }

            switch_dict: Dict[str, Callable] = {
                "ID": self.not_empty_validator,
                "STRING": self.string_validator,
                "INTEGER": self.integer_validator,
                "DECIMAL": self.float_validator,
                "BOOL": self.bool_validator,
                "DATE": self.date_validator,
                "DATETIME": self.date_validator,
                "SELECT_ONE": self.choice_validator,
                "SELECT_MANY": self.choice_validator,
                "PHONE_NUMBER": self.phone_validator,
                "GEOPOINT": self.geolocation_validator,
                "IMAGE": self.image_validator,
                "LIST_OF_IDS": self.integer_validator,
            }

            invalid_rows = []
            current_household_id = None

            identities_numbers = {
                "unhcr_id_no_i_c": {
                    "partner": "UNHCR",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "scope_id_no_i_c": {
                    "partner": "WFP",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
            }
            documents_numbers: Dict[str, Dict[str, Any]] = {
                "birth_certificate_no_i_c": {
                    "type": "BIRTH_CERTIFICATE",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "drivers_license_no_i_c": {
                    "type": "DRIVERS_LICENSE",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "electoral_card_no_i_c": {
                    "type": "ELECTORAL_CARD",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "national_id_no_i_c": {
                    "type": "NATIONAL_ID",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "national_passport_no_i_c": {
                    "type": "NATIONAL_PASSPORT",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "tax_id_no_i_c": {
                    "type": "TAX_ID",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "other_id_type_i_c": {
                    "type": "OTHER",
                    "names": [],
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "other_id_no_i_c": {},
            }

            delivery_mechanisms_data = defaultdict(dict)

            def has_value(cell: Cell) -> bool:
                if cell.value is None:
                    return False
                if isinstance(cell.value, str):
                    return cell.value.strip() != ""
                return True

            admin_area_code_tuples = []

            for row in sheet.iter_rows(min_row=3):
                # openpyxl keeps iterating on empty rows so need to omit empty rows
                if not any(has_value(cell) for cell in row):
                    continue
                row_number = 0
                for cell, header in zip(row, first_row):
                    current_field = combined_fields.get(header.value)
                    if not current_field:
                        continue

                    row_number = cell.row

                    value = cell.value
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)

                    household_id_can_be_empty = False
                    if header.value == "household_id":
                        current_household_id = value
                        if sheet.title == "Households":
                            self.household_ids.append(current_household_id)
                            self.head_of_household_count[current_household_id] = 0
                        else:
                            household_id_can_be_empty = True

                    if header.value == "relationship_i_c" and cell.value == "HEAD":
                        self.head_of_household_count[current_household_id] += 1

                    people_admin_columns = ("pp_admin1_i_c", "pp_admin2_i_c", "pp_admin3_i_c")
                    hh_admin_columns = ("admin1_h_c", "admin2_h_c", "admin3_h_c")
                    admin_columns_all = people_admin_columns + hh_admin_columns
                    if header.value in admin_columns_all:
                        if cell.value:
                            admin_area_code_tuples.append((row_number, header.value, cell.value))
                        # admin3 is not required for now
                        elif not cell.value and header.value not in ("admin3_h_c", "pp_admin3_i_c"):
                            invalid_rows.append(
                                {
                                    "row_number": row_number,
                                    "header": header.value,
                                    "message": f"{header.value.capitalize()} field cannot be null",
                                }
                            )

                    field_type = current_field["type"]
                    fn: Callable = switch_dict[field_type]

                    if (
                        fn(value, header.value, cell) is False
                        and household_id_can_be_empty is False
                        and header.value not in admin_columns_all
                    ):
                        message = (
                            f"Sheet: {sheet.title!r}, Unexpected value: "
                            f"{value} for type "
                            f"{field_type.replace('_', ' ').lower()} "
                            f"of field {header.value}"
                        )
                        invalid_rows.append({"row_number": cell.row, "header": header.value, "message": message})

                    header_value_doc = header.value[len("pp_") :] if header.value.startswith("pp_") else header.value
                    if header_value_doc in documents_numbers:
                        if header_value_doc == "other_id_type_i_c":
                            documents_numbers["other_id_type_i_c"]["names"].append(value)
                        elif header_value_doc == "other_id_no_i_c":
                            documents_numbers["other_id_type_i_c"]["numbers"].append(str(value) if value else None)
                        else:
                            documents_numbers[header_value_doc]["numbers"].append(str(value) if value else None)

                    if header_value_doc in self.DOCUMENTS_ISSUING_COUNTRIES_MAPPING.keys():
                        document_key = self.DOCUMENTS_ISSUING_COUNTRIES_MAPPING.get(header_value_doc)
                        documents_dict = documents_numbers
                        if document_key in identities_numbers.keys():
                            documents_dict = identities_numbers
                        if document_key:
                            documents_dict[document_key]["issuing_countries"].append(value)

                    if header_value_doc in identities_numbers:
                        identities_numbers[header_value_doc]["numbers"].append(str(value) if value else None)

                    if header.value in self.delivery_mechanisms_xlsx_fields:
                        delivery_mechanisms_data[row_number][header.value] = value

                if current_household_id and current_household_id not in self.household_ids:
                    message = f"Sheet: 'Individuals', There is no household with provided id: {current_household_id}"
                    invalid_rows.append({"row_number": row_number, "header": "relationship_i_c", "message": message})

                for header_value_doc in self.DOCUMENTS_ISSUING_COUNTRIES_MAPPING.values():
                    documents_or_identity_dict = (
                        identities_numbers if header_value_doc in identities_numbers.keys() else documents_numbers
                    )
                    documents_or_identity_dict[header_value_doc]["validation_data"].append({"row_number": row[0].row})
                self.errors.extend(self._validate_pdu(row, first_row, row_number))

            if sheet.title == "Individuals":
                for household_id, count in self.head_of_household_count.items():
                    if count == 0:
                        message = (
                            f"Sheet: 'Individuals', Household with id: {household_id} has to have a head of household"
                        )
                        invalid_rows.append({"row_number": 0, "header": "relationship_i_c", "message": message})
                    elif count > 1:
                        message = f"Sheet: 'Individuals', There are multiple head of households for household with id: {household_id}"
                        invalid_rows.append({"row_number": 0, "header": "relationship_i_c", "message": message})

            if sheet.title in ("Households", "People"):
                admin_area_invalid_rows = self.validate_admin_areas(admin_area_code_tuples, business_area_slug)
                if admin_area_invalid_rows:
                    invalid_rows.extend(admin_area_invalid_rows)

            invalid_doc_rows = []
            invalid_ident_rows = []
            invalid_delivery_mechanisms = []
            if sheet.title in ["Individuals", "People"]:
                invalid_doc_rows = self.documents_validator(documents_numbers)
                invalid_ident_rows = self.identity_validator(identities_numbers)
                invalid_delivery_mechanisms = self.delivery_mechanisms_validator(delivery_mechanisms_data)

            self.errors.extend([*invalid_rows, *invalid_doc_rows, *invalid_ident_rows])
            self.delivery_mechanisms_errors.extend(invalid_delivery_mechanisms)

        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def validate_admin_areas(
        self, admin_area_code_tuples: List[Tuple[int, str, str]], business_area_slug: Optional[str]
    ) -> List[Dict[str, Any]]:
        invalid_rows = []
        if admin_area_code_tuples:
            business_area_countries = BusinessArea.objects.get(slug=business_area_slug).countries.all()
            queryset = Area.objects.select_related("area_type")

            for code_tuple in admin_area_code_tuples:
                message = None
                row_number, header_name, p_code = code_tuple
                area = queryset.filter(p_code=p_code).first()
                if not area:
                    message = f"Sheet: 'Households': Area with code: {p_code} does not exist"
                elif area.area_type.country not in business_area_countries:
                    message = (
                        f"Sheet: 'Households': Admin Area: {p_code} unavailable in Business Area: {business_area_slug}"
                    )
                if message:
                    invalid_rows.append({"row_number": row_number, "header": header_name, "message": message})
        return invalid_rows

    def validate_file_with_template(self, wb: Workbook) -> None:
        try:
            combined_fields = self.combined_fields

            for name, fields in combined_fields.items():
                if name.capitalize() not in wb.sheetnames:  # pragma: no cover
                    self.errors.append(
                        {"row_number": 0, "header": "File", "message": f"Worksheet {name.capitalize()} does not exist."}
                    )
                    return

                sheet = wb[name.capitalize()]
                first_row = sheet[1]

                all_fields = list(fields.values())

                required_fields = {field["xlsx_field"] for field in all_fields if field["required"]}

                column_names = {cell.value for cell in first_row}

                columns_difference = required_fields.difference(column_names)

                if columns_difference:  # pragma: no cover
                    self.errors.extend(
                        [
                            {"row_number": 1, "header": col, "message": f"Missing column name {col}"}
                            for col in columns_difference
                        ]
                    )
                    return

        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def validate_file_extension(self, xlsx_file: Any) -> None:
        try:
            file_suffix = Path(xlsx_file.name).suffix
            if file_suffix != ".xlsx":
                self.errors.append(
                    {
                        "row_number": 1,
                        "header": f"{xlsx_file.name}",
                        "message": "Only .xlsx files are accepted for import",
                    }
                )
                return
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def validate_everything(
        self, xlsx_file: Any, business_area_slug: str
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        try:
            self.validate_file_extension(xlsx_file)
            if self.errors:
                return self.errors, self.delivery_mechanisms_errors
            try:
                wb = openpyxl.load_workbook(xlsx_file, data_only=True)
            except BadZipfile:
                return [
                    {"row_number": 1, "header": f"{xlsx_file.name}", "message": "Invalid .xlsx file"}
                ], self.delivery_mechanisms_errors

            self.validate_file_with_template(wb)
            if self.errors:  # pragma: no cover
                # return error if WS do not exist in the import file
                return self.errors, self.delivery_mechanisms_errors

            self.validate_index_id(wb)
            self.validate_collectors_size(wb)

            if self.is_social_worker_program:
                self.validate_people_collectors(wb)
                people_sheet = wb["People"]
                self.image_loader = SheetImageLoader(people_sheet)
                self.rows_validator(people_sheet, business_area_slug)
            else:
                self.validate_collectors(wb)
                individuals_sheet = wb["Individuals"]
                household_sheet = wb["Households"]
                self.image_loader = SheetImageLoader(household_sheet)
                self.rows_validator(household_sheet, business_area_slug)
                self.image_loader = SheetImageLoader(individuals_sheet)
                self.rows_validator(individuals_sheet)

            return self.errors, self.delivery_mechanisms_errors
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    @staticmethod
    def collector_column_validator(header: str, data_dict: Dict, household_ids: Set[str]) -> List[Dict[str, Any]]:
        try:
            is_primary_collector = header == "primary_collector_id"
            errors = []
            collectors_ids = []
            for row, cell in data_dict.items():
                if not cell.value:
                    continue
                list_of_ids = set(collectors_str_ids_to_list(cell.value) or [])
                contains_correct_ids = list_of_ids.issubset(household_ids)
                if not contains_correct_ids:
                    errors.append(
                        {
                            "row_number": row,
                            "header": header,
                            "message": "One or more ids are not attached to any household in the file.",
                        }
                    )
                collectors_ids.extend(list_of_ids)

            collectors_ids_set = set(collectors_ids)

            if is_primary_collector:
                household_ids_without_collectors = household_ids.difference(collectors_ids_set)
                errors.extend(
                    {
                        "row_number": 1,
                        "header": header,
                        "message": f"Household with id: {hh_id} does not have primary collector",
                    }
                    for hh_id in household_ids_without_collectors
                )

            ids_counter = Counter(collectors_ids)
            erroneous_collectors_ids = [item for item, count in ids_counter.items() if count > 1]
            message = "Household can contain only one primary and one alternate collector"
            errors.extend(
                {"row_number": 1, "header": header, "message": f"{message}, erroneous id: {hh_id}"}
                for hh_id in erroneous_collectors_ids
            )
            return errors
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def validate_collectors(self, wb: Workbook) -> None:
        try:
            individuals_sheet = wb["Individuals"]
            households_sheet = wb["Households"]
            first_row = individuals_sheet[1]

            household_ids = {
                str(int(cell.value)) if isinstance(cell.value, float) and cell.value.is_integer() else str(cell.value)
                for cell in households_sheet["A"][2:]
                if cell.value
            }

            primary_collectors_data = {}
            alternate_collectors_data = {}
            for cell in first_row:
                if cell.value == "primary_collector_id":
                    primary_collectors_data = {c.row: c for c in individuals_sheet[cell.column_letter][2:] if c.value}
                elif cell.value == "alternate_collector_id":
                    alternate_collectors_data = {c.row: c for c in individuals_sheet[cell.column_letter][2:] if c.value}

            self.errors.extend(
                self.collector_column_validator("primary_collector_id", primary_collectors_data, household_ids)
            )
            self.errors.extend(
                self.collector_column_validator(
                    "alternate_collector_id",
                    alternate_collectors_data,
                    household_ids,
                )
            )
            self.errors.extend(self.validate_collectors_unique(primary_collectors_data, alternate_collectors_data))

        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def validate_collectors_unique(self, primary_collectors: Dict, alternate_collectors: Dict) -> List[Dict[str, Any]]:
        try:
            errors = []

            for row in primary_collectors:
                primary_collector = set(
                    collectors_str_ids_to_list(cell.value if (cell := primary_collectors.get(row)) else None) or []
                )
                alternate_collector = set(
                    collectors_str_ids_to_list(cell.value if (cell := alternate_collectors.get(row)) else None) or []
                )
                if households_ids := primary_collector.intersection(alternate_collector):
                    errors.append(
                        {
                            "row_number": row,
                            "header": "Individuals",
                            "message": f"Individual from row: {row} cannot be the primary and the alternate collector for households: {', '.join(households_ids)} at the same time.",
                        }
                    )
            return errors
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def validate_index_id(self, wb: Workbook) -> None:
        try:
            if self.is_social_worker_program:
                people_sheet = wb["People"]
                header_row = people_sheet[1]
                index_id_col = 1  # by default
                for header in header_row:
                    if header.value == "pp_index_id":
                        index_id_col = int(header.column)

                index_ids = list(
                    people_sheet.iter_cols(min_col=index_id_col, max_col=index_id_col, min_row=3, values_only=True)
                )[0]
                duplicates = list(set([i for i in index_ids if index_ids.count(i) > 1 and i is not None]))
                if duplicates:
                    self.errors.append(
                        {
                            "row_number": 1,
                            "header": "People",
                            "message": f"There are duplicates with id(s): {duplicates}. Number have to be unique in the field pp_index_id.",
                        }
                    )
                    return

        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def validate_people_collectors(self, wb: Workbook) -> None:
        try:
            index_ids, primary_collector_ids, relationship_column = [], [], []
            people_sheet = wb["People"]
            first_row = people_sheet[1]

            for header in first_row:
                if header.value == "pp_index_id":
                    index_id_col = int(header.column)
                    index_ids = list(
                        people_sheet.iter_cols(min_col=index_id_col, max_col=index_id_col, min_row=3, values_only=True)
                    )[0]
                if header.value == "pp_primary_collector_id":
                    pr_collector_id_col = int(header.column)
                    primary_collector_ids = list(
                        people_sheet.iter_cols(
                            min_col=pr_collector_id_col, max_col=pr_collector_id_col, min_row=3, values_only=True
                        )
                    )[0]
                if header.value == "pp_relationship_i_c":
                    relationship_col = int(header.column)
                    relationship_column = list(
                        people_sheet.iter_cols(
                            min_col=relationship_col, max_col=relationship_col, min_row=3, values_only=True
                        )
                    )[0]
            pr_ids = [int(i) for i in primary_collector_ids if i is not None]
            for index_id, relationship, pr_col in itertools.zip_longest(
                index_ids, relationship_column, primary_collector_ids, fillvalue=None
            ):
                if relationship not in [HEAD, NON_BENEFICIARY] and index_id is not None:
                    self.errors.append(
                        {
                            "row_number": 1,
                            "header": "People",
                            "message": f"Invalid value in field 'pp_relationship_i_c' with index_id {index_id}. "
                            f"Value can be {HEAD} or {NON_BENEFICIARY}",
                        }
                    )
                if relationship == HEAD and index_id is not None and int(index_id) not in pr_ids:
                    self.errors.append(
                        {
                            "row_number": 1,
                            "header": "People",
                            "message": f"Individual with index_id {index_id} has to have an Primary collector.",
                        }
                    )
                if relationship == NON_BENEFICIARY and (pr_col is None):
                    self.errors.append(
                        {
                            "row_number": 1,
                            "header": "People",
                            "message": f"Invalid value in field 'pp_primary_collector_id' for Individual with index_id {index_id}. "
                            f"Value cannot be empty for relationship {NON_BENEFICIARY}",
                        }
                    )

        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def validate_collectors_size(self, wb: Workbook) -> None:
        try:
            if not self.is_social_worker_program:
                individuals_sheet = wb["Individuals"]
                households_sheet = wb["Households"]

                household_count = self._count_households(households_sheet)
                individuals_count = self._count_individuals(individuals_sheet)

                if household_count == 0:
                    self.errors.append(
                        {
                            "row_number": 1,
                            "header": "Households",
                            "message": "There aren't households in the file.",
                        }
                    )

                if individuals_count == 0:
                    self.errors.append(
                        {
                            "row_number": 1,
                            "header": "Individuals",
                            "message": "There aren't individuals in the file.",
                        }
                    )
            else:
                people_sheet = wb["People"]
                people_count = self._count_individuals(people_sheet)
                if people_count == 0:  # pragma: no cover
                    self.errors.append(
                        {
                            "row_number": 1,
                            "header": "People",
                            "message": "There aren't people in the file.",
                        }
                    )

        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def _validate_pdu(self, row: list[Any], header_row: list[Any], row_number: int) -> list:
        header_row = [header.value for header in header_row]
        validation_subtype_mapping = {
            PeriodicFieldData.DATE: self.date_validator,
            PeriodicFieldData.DECIMAL: self.decimal_validator,
            PeriodicFieldData.STRING: self.string_validator,
            PeriodicFieldData.BOOL: self.bool_validator,
        }
        errors = []
        for flexible_attribute in self.pdu_flexible_attributes:
            column_value = f"{flexible_attribute.name}_round_1_value"
            column_collection_date = f"{flexible_attribute.name}_round_1_collection_date"
            subtype = flexible_attribute.pdu_data.subtype
            if column_value not in header_row:
                continue
            column_value_index = header_row.index(column_value)
            column_collection_date_index = header_row.index(column_collection_date)
            value = row[column_value_index].value
            if value is None or value == "":
                continue
            collection_date = None
            is_collection_date_valid = True
            if column_collection_date_index >= 0:
                collection_date = row[column_collection_date_index].value
                is_collection_date_valid = self.date_validator(collection_date, column_collection_date)
            is_value_valid = validation_subtype_mapping[subtype](value, column_value)
            if not is_collection_date_valid:
                errors.append(
                    {
                        "row_number": row_number,
                        "header": column_collection_date,
                        "message": f"Invalid value {collection_date} for field {column_collection_date}",
                    }
                )
            if not is_value_valid:
                errors.append(
                    {
                        "row_number": row_number,
                        "header": column_value,
                        "message": f"Invalid value {value} for field {column_value}",
                    }
                )
        return errors

    def _count_individuals(self, individuals_sheet: Worksheet) -> int:
        first_row = individuals_sheet[1]
        individuals_count = 0
        for cell in first_row:
            if cell.value in ["full_name_i_c", "pp_full_name_i_c"]:
                for c in individuals_sheet[cell.column_letter][2:]:
                    if c.value:
                        individuals_count += 1
        return individuals_count

    def _count_households(self, households_sheet: Worksheet) -> int:
        household_count = 0
        for cell in households_sheet["A"][2:]:
            if cell.value:
                household_count += 1
        return household_count


class KoboProjectImportDataInstanceValidator(ImportDataInstanceValidator):
    def __init__(self, program: Program) -> None:
        super().__init__(program)
        self.combined_fields = self.get_combined_fields()
        self.expected_household_fields = self.get_expected_household_fields()
        self.expected_individuals_fields = self.get_expected_individuals_fields()

    def get_combined_fields(self) -> Dict[str, Dict]:
        core_fields = FieldFactory.from_scope(Scope.KOBO_IMPORT)
        flex_fields = serialize_flex_attributes()
        return {
            "households": {
                **core_fields.associated_with_household().to_dict_by("xlsx_field"),
                **flex_fields["households"],
            },
            "individuals": {
                **core_fields.associated_with_individual().to_dict_by("xlsx_field"),
                **flex_fields["individuals"],
            },
        }

    def get_expected_household_fields(self) -> Set:
        try:
            return {field["xlsx_field"] for field in self.combined_fields["households"].values() if field["required"]}
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def get_expected_individuals_fields(self) -> Set:
        try:
            return {field["xlsx_field"] for field in self.combined_fields["individuals"].values() if field["required"]}
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def standard_type_validator(self, value: str, field: str, field_type: str) -> Optional[str]:
        try:
            value_type_name = type(value).__name__

            if field_type == "INTEGER":
                try:
                    int(value)
                    return None
                except Exception:
                    return f"Invalid value {value} of type {value_type_name} for " f"field {field} of type int"
            elif field_type == "STRING":
                # everything from Kobo is string so cannot really validate it
                # only check phone number
                if field.startswith("phone_no"):
                    try:
                        phonenumbers.parse(value, None)
                    except (phonenumbers.NumberParseException, TypeError):
                        return f"Invalid phone number {value} for field {field}"
                return None

            elif field_type == "BOOL":
                # Important! if value == 0 or 1 it's also evaluated to True
                # checking for int values even tho Kobo returns everything as str
                # to no not break import if they start returning integers
                if value in ("True", "False", True, False, "0", "1", "TRUE", "FALSE", "true", "false"):
                    return None
                return f"Invalid value {value} of type {value_type_name} for " f"field {field} of type bool"
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise
        return None

    def image_validator(
        self, value: str, field: str, attachments: list[dict], *args: Any, **kwargs: Any
    ) -> Union[str, None]:
        try:
            if kwargs.get("skip_validate_pictures") is True:
                # skip validation if skip_validate_pictures=True
                return None
            allowed_extensions = django_core_validators.get_available_image_extensions()
            file_extension = value.split(".")[-1]

            if file_extension.lower() not in allowed_extensions:
                message = f"Specified image {value} for " f"field {field} is not a valid image file"
                return message

            message = f"Specified image {value} for field {field} is not in attachments"

            is_correct_attachment = find_attachment_in_kobo(attachments, value)

            # Kobo not always returns consent_sign_h_c in attachments, according to AB#168823 we should skip it
            if not is_correct_attachment and field == "consent_sign_h_c":
                return None
            is_valid_image = isinstance(value, str) and is_correct_attachment

            return None if is_valid_image else message
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def geopoint_validator(
        self, value: Optional[Sequence[Any]], field: str, *args: Any, **kwargs: Any
    ) -> Union[str, None]:
        message = f"Invalid geopoint {value} for field {field}"

        if not value or not isinstance(value, str):
            return message

        geopoint_to_list = value.split(" ")
        geopoint = " ".join(geopoint_to_list[:2])

        pattern = re.compile(r"^(-?\d+\.\d+? \s*-?\d+\.\d+?)$")
        is_valid_geopoint = bool(re.match(pattern, geopoint))

        return None if is_valid_geopoint else message

    def date_validator(self, value: str, field: str, *args: Any, **kwargs: Any) -> Union[str, None]:
        try:
            message = (
                f"Invalid datetime/date {value} for field {field}, "
                "accepted formats: datetime ISO 8601, date YYYY-MM-DD"
            )

            if not value:
                return message

            pattern_iso = re.compile(
                r"^(-?(?:[1-9][0-9]*)?[0-9]{4})-(1[0-2]|0[1-9])-"
                r"(3[01]|0[1-9]|[12][0-9])T(2[0-3]|[01][0-9]):([0-5][0-9]):"
                r"([0-5][0-9])(\.[0-9]+)?(Z|[+-](?:2[0-3]|[01][0-9]):"
                r"[0-5][0-9])?$"
            )

            matched = re.match(pattern_iso, value)

            if matched is None:
                pattern_date = re.compile(r"([12]\d{3}-(0[1-9]|1[0-2])-(0[1-9]|[12]\d|3[01]))")

                matched = re.match(pattern_date, value)

            return None if matched else message
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def choice_validator(self, value: str, field: str, *args: Any, **kwargs: Any) -> Union[str, None]:
        try:
            message = f"Invalid choice {value} for field {field}"
            if not value:
                return message

            found_field: Dict = self.all_fields[field]
            custom_validate_choices_method = found_field.get("custom_validate_choices")
            choices = found_field["choices"]
            choice_type = found_field["type"]

            if choice_type == TYPE_SELECT_ONE:
                if custom_validate_choices_method is not None:
                    return None if custom_validate_choices_method(value) is True else message

                is_in_choices = value in choices
                if is_in_choices is False:
                    # try uppercase version
                    uppercase_value = value.upper()
                    is_in_choices = uppercase_value in choices
                return None if is_in_choices else message

            elif choice_type == TYPE_SELECT_MANY:
                str_value = str(value)
                if "," in str_value:
                    selected_choices = str_value.split(",")
                elif ";" in str_value:
                    selected_choices = str_value.split(";")
                else:
                    selected_choices = str_value.split(" ")

                if custom_validate_choices_method is not None:
                    return None if custom_validate_choices_method(str_value) is True else message

                for choice in selected_choices:
                    choice = choice.strip()
                    if choice not in choices and choice.upper() not in choices:
                        return message
                return None

            return None
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def _get_field_type_error(
        self, field: str, value: Any, attachments: list, skip_validate_pictures: Optional[bool] = False
    ) -> Union[dict, None]:
        try:
            field_dict = self.all_fields.get(field)
            if field_dict is None:
                return None

            complex_types: Dict[str, Callable] = {
                "GEOPOINT": self.geopoint_validator,
                "IMAGE": self.image_validator,
                "DATE": self.date_validator,
                "SELECT_ONE": self.choice_validator,
                "SELECT_MANY": self.choice_validator,
            }
            field_type = field_dict["type"]
            complex_type_fn: Optional[Callable] = complex_types.get(field_type)

            if complex_type_fn:
                message = complex_type_fn(
                    field=field, value=value, attachments=attachments, skip_validate_pictures=skip_validate_pictures
                )
                if message is not None:
                    return {
                        "header": field,
                        "message": message,
                    }
            else:
                message = self.standard_type_validator(value, field, field_type)
                if message:
                    return {
                        "header": field,
                        "message": message,
                    }

            return None
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise

    def validate_everything(
        self, submissions: List, business_area: BusinessArea, skip_validate_pictures: Optional[bool] = False
    ) -> List:
        try:
            reduced_submissions: Sequence = rename_dict_keys(submissions, get_field_name)
            docs_and_identities_to_validate = []
            errors = []
            # have fun debugging this ;_;
            # thx
            # thx again

            identities_numbers = {
                "unhcr_id_no_i_c": {"partner": "UNHCR", "validation_data": [], "numbers": [], "issuing_countries": []},
                "scope_id_no_i_c": {"partner": "WFP", "validation_data": [], "numbers": [], "issuing_countries": []},
            }
            documents_numbers: Dict[str, Dict[str, Any]] = {
                "birth_certificate_no_i_c": {
                    "type": "BIRTH_CERTIFICATE",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "drivers_license_no_i_c": {
                    "type": "DRIVERS_LICENSE",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "electoral_card_no_i_c": {
                    "type": "ELECTORAL_CARD",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "national_id_no_i_c": {
                    "type": "NATIONAL_ID",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "national_passport_no_i_c": {
                    "type": "NATIONAL_PASSPORT",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "tax_id_no_i_c": {
                    "type": "TAX_ID",
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "other_id_type_i_c": {
                    "type": "OTHER",
                    "names": [],
                    "validation_data": [],
                    "numbers": [],
                    "issuing_countries": [],
                },
                "other_id_no_i_c": {},
            }
            kobo_asset_id = None
            if len(reduced_submissions) > 0:
                kobo_asset_id = reduced_submissions[0]["_xform_id_string"]
            all_saved_submissions = KoboImportedSubmission.objects.filter(kobo_asset_id=kobo_asset_id)
            if business_area.get_sys_option("ignore_amended_kobo_submissions"):
                all_saved_submissions = all_saved_submissions.filter(amended=False)
            all_saved_submissions = all_saved_submissions.values("kobo_submission_uuid", "kobo_submission_time")

            all_saved_submissions_dict = {}
            for submission in all_saved_submissions:
                item = all_saved_submissions_dict.get(str(submission["kobo_submission_uuid"]), [])
                item.append(submission["kobo_submission_time"].isoformat())
                all_saved_submissions_dict[str(submission["kobo_submission_uuid"])] = item
            household: Dict[str, Any]
            household_hash_list: List[str] = []
            for household in reduced_submissions:
                household_uuid = str(household.get("_uuid"))
                household_hash = calculate_hash_for_kobo_submission(household)
                submission_exists = household.get("_submission_time") in all_saved_submissions_dict.get(
                    household_uuid, []
                )
                submission_duplicate = household_hash in household_hash_list
                if submission_exists or submission_duplicate:
                    continue
                household_hash_list.append(household_hash)
                head_of_hh_counter = 0
                primary_collector_counter = 0
                alternate_collector_counter = 0
                expected_hh_fields = {
                    *self.expected_household_fields,
                }
                attachments = household.get("_attachments", [])
                hh_value: List[Dict]
                for hh_field, hh_value in household.items():
                    expected_hh_fields.discard(hh_field)
                    if hh_field == KOBO_FORM_INDIVIDUALS_COLUMN_NAME:
                        individual: Dict
                        for individual in hh_value:
                            expected_i_fields = {
                                *self.expected_individuals_fields,
                            }
                            current_individual_docs_and_identities = defaultdict(dict)
                            i_field: str
                            for i_field, i_value in individual.items():
                                if i_field in documents_numbers:
                                    if i_field == "other_id_type_i_c":
                                        documents_numbers["other_id_type_i_c"]["names"].append(i_value)
                                    elif i_field == "other_id_no_i_c":
                                        documents_numbers["other_id_type_i_c"]["validation_data"].append(
                                            {"value": i_value}
                                        )
                                        documents_numbers["other_id_type_i_c"]["numbers"].append(i_value)
                                    else:
                                        documents_numbers[i_field]["validation_data"].append({"value": i_value})
                                        documents_numbers[i_field]["numbers"].append(i_value)
                                if i_field in self.DOCUMENTS_ISSUING_COUNTRIES_MAPPING.keys():
                                    document_key = self.DOCUMENTS_ISSUING_COUNTRIES_MAPPING[i_field]
                                    documents_dict: Dict[str, Dict[str, Any]] = documents_numbers
                                    if document_key in identities_numbers.keys():
                                        documents_dict = identities_numbers
                                    documents_dict[document_key]["issuing_countries"].append(i_value)

                                if i_field in identities_numbers:
                                    identities_numbers[i_field]["numbers"].append(i_value)
                                    identities_numbers[i_field]["validation_data"].append({"value": i_value})

                                if i_field == "relationship_i_c" and i_value.upper() == "HEAD":
                                    head_of_hh_counter += 1
                                if i_field == "role_i_c":
                                    role = i_value.upper()
                                    if role == ROLE_PRIMARY:
                                        primary_collector_counter += 1
                                    elif role == ROLE_ALTERNATE:
                                        alternate_collector_counter += 1

                                expected_i_fields.discard(i_field)
                                error = self._get_field_type_error(
                                    i_field, i_value, attachments, skip_validate_pictures
                                )
                                if error:
                                    errors.append(error)

                            docs_and_identities_to_validate.append(current_individual_docs_and_identities)

                            i_expected_field_errors = [
                                {"header": field, "message": "Missing individual " f"required field {field}"}
                                for field in expected_i_fields
                            ]
                            errors.extend(i_expected_field_errors)

                        if head_of_hh_counter == 0:
                            errors.append(
                                {
                                    "header": "relationship_i_c",
                                    "message": "Household has to have a " "head of household",
                                }
                            )
                        if head_of_hh_counter > 1:
                            errors.append(
                                {
                                    "header": "relationship_i_c",
                                    "message": "Only one person can " "be a head of household",
                                }
                            )
                        if primary_collector_counter == 0:
                            errors.append(
                                {"header": "role_i_c", "message": "Household must have a " "primary collector"}
                            )
                        if primary_collector_counter > 1:
                            errors.append(
                                {"header": "role_i_c", "message": "Only one person can " "be a primary collector"}
                            )
                        if alternate_collector_counter > 1:
                            errors.append(
                                {"header": "role_i_c", "message": "Only one person can " "be a alternate collector"}
                            )
                    else:
                        error = self._get_field_type_error(hh_field, hh_value, attachments, skip_validate_pictures)
                        if error:
                            errors.append(error)
                hh_expected_field_errors = [
                    {"header": field, "message": f"Missing household required field {field}"}
                    for field in expected_hh_fields
                ]
                errors.extend(hh_expected_field_errors)

            document_errors = self.documents_validator(documents_numbers, is_xlsx=False)
            identities_errors = self.identity_validator(identities_numbers, is_xlsx=False)

            return [*errors, *document_errors, *identities_errors]
        except Exception as e:  # pragma: no cover
            logger.exception(e)
            raise
