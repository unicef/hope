from typing import Dict, List, Optional, Tuple

import openpyxl

from hct_mis_api.apps.core.field_attributes.core_fields_attributes import FieldFactory
from hct_mis_api.apps.core.field_attributes.fields_types import Scope
from hct_mis_api.apps.core.utils import serialize_flex_attributes
from hct_mis_api.apps.geo.models import Area


class TemplateFileGenerator:
    @classmethod
    def _create_workbook(cls) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_import_helper = wb.active
        ws_import_helper.title = "Import helper"
        wb.create_sheet("Households")
        wb.create_sheet("Individuals")
        wb.create_sheet("People")
        wb.create_sheet("Choices")

        wb = cls._add_import_helper(wb)

        return wb

    @classmethod
    def _handle_choices(cls, fields: Dict) -> List[List[str]]:
        rows: list[list[str]] = [["Field Name", "Label", "Value to be used in template"]]

        for field_name, field_value in fields.items():
            is_admin_level = field_name in ("admin1_h_c", "admin2_h_c")
            choices = field_value["choices"]
            if is_admin_level:
                choices = Area.get_admin_areas_as_choices(field_name[-5])
            if choices:
                for choice in field_value["choices"]:
                    row = [
                        field_name,
                        str(choice["label"]["English(EN)"]),
                        choice["value"],
                    ]
                    rows.append(row)

        return rows

    @classmethod
    def _handle_name_and_label_row(cls, fields: Dict) -> Tuple[List[str], List[str]]:
        names: List[str] = []
        labels: List[str] = []

        for field_name, field_value in fields.items():
            names.append(field_name)
            label = (
                f"{field_value['label']['English(EN)']}"
                f" - {field_value['type']}"
                f"{' - required' if field_value['required'] else ''}"
            )
            labels.append(label)

        return names, labels

    @classmethod
    def _add_template_columns(
        cls,
        wb: openpyxl.Workbook,
        business_area_slug: Optional[str] = None,
    ) -> openpyxl.Workbook:
        households_sheet_title = "Households"
        individuals_sheet_title = "Individuals"
        people_sheet_title = "People"

        ws_households = wb[households_sheet_title]
        ws_individuals = wb[individuals_sheet_title]
        ws_people = wb[people_sheet_title]
        ws_choices = wb["Choices"]

        flex_fields = serialize_flex_attributes()

        fields = FieldFactory.from_scopes(
            [Scope.GLOBAL, Scope.XLSX, Scope.HOUSEHOLD_ID, Scope.COLLECTOR]
        ).apply_business_area(business_area_slug=business_area_slug)
        people_fields = FieldFactory.from_scopes([Scope.XLSX_PEOPLE]).apply_business_area(
            business_area_slug=business_area_slug
        )

        households_fields = {
            **fields.associated_with_household().to_dict_by("xlsx_field"),
            **flex_fields[households_sheet_title.lower()],
        }

        individuals_fields = {
            **fields.associated_with_individual().to_dict_by("xlsx_field"),
            **flex_fields[individuals_sheet_title.lower()],
        }

        people_fields = {
            **people_fields.associated_with_individual().to_dict_by("xlsx_field"),
            **flex_fields[individuals_sheet_title.lower()],
        }

        households_rows = cls._handle_name_and_label_row(households_fields)
        individuals_rows = cls._handle_name_and_label_row(individuals_fields)
        people_rows = cls._handle_name_and_label_row(people_fields)

        for h_row, i_row, p_row in zip(households_rows, individuals_rows, people_rows):
            ws_households.append(h_row)
            ws_individuals.append(i_row)
            ws_people.append(p_row)

        choices = cls._handle_choices({**households_fields, **individuals_fields})
        for row in choices:
            ws_choices.append(row)

        return wb

    @classmethod
    def _add_import_helper(cls, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        ws_import_helper = wb["Import helper"]
        ws_import_helper.append(["Some data will be added soon"])
        # TODO: add info for Import helper
        return wb

    @classmethod
    def get_template_file(cls, business_area_slug: Optional[str] = None) -> openpyxl.Workbook:
        return cls._add_template_columns(cls._create_workbook(), business_area_slug)
