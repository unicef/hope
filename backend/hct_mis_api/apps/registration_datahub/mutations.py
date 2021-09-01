import json
import logging
import operator
import time
from io import BytesIO

from django.core.exceptions import ValidationError
from django.core.files import File
from django.db import transaction
from django.shortcuts import get_object_or_404

import graphene
import openpyxl
from graphene_file_upload.scalars import Upload

from hct_mis_api.apps.account.permissions import PermissionMutation, Permissions
from hct_mis_api.apps.activity_log.models import log_create
from hct_mis_api.apps.core.kobo.api import KoboAPI
from hct_mis_api.apps.core.kobo.common import count_population
from hct_mis_api.apps.core.models import BusinessArea
from hct_mis_api.apps.core.permissions import is_authenticated
from hct_mis_api.apps.core.scalars import BigInt
from hct_mis_api.apps.core.utils import (
    check_concurrency_version_in_mutation,
    decode_id_string,
)
from hct_mis_api.apps.core.validators import BaseValidator
from hct_mis_api.apps.registration_data.models import RegistrationDataImport
from hct_mis_api.apps.registration_data.schema import RegistrationDataImportNode
from hct_mis_api.apps.registration_datahub.celery_tasks import (
    merge_registration_data_import_task,
    rdi_deduplication_task,
    registration_kobo_import_task,
    registration_xlsx_import_task,
)
from hct_mis_api.apps.registration_datahub.models import (
    ImportData,
    RegistrationDataImportDatahub,
)
from hct_mis_api.apps.registration_datahub.schema import (
    ImportDataNode,
    KoboErrorNode,
    XlsxRowErrorNode,
)
from hct_mis_api.apps.registration_datahub.validators import (
    KoboProjectImportDataInstanceValidator,
    UploadXLSXInstanceValidator,
)
from hct_mis_api.apps.utils.mutations import ValidationErrorMutationMixin

logger = logging.getLogger(__name__)


@transaction.atomic(using="default")
@transaction.atomic(using="registration_datahub")
def create_registration_data_import_objects(registration_data_import_data, user, data_source):
    import_data_id = decode_id_string(registration_data_import_data.pop("import_data_id"))
    import_data_obj = ImportData.objects.get(id=import_data_id)

    business_area = BusinessArea.objects.get(slug=registration_data_import_data.pop("business_area_slug"))
    pull_pictures = registration_data_import_data.pop("pull_pictures", True)
    screen_beneficiary = registration_data_import_data.pop("screen_beneficiary", False)
    created_obj_datahub = RegistrationDataImportDatahub.objects.create(
        business_area_slug=business_area.slug,
        import_data=import_data_obj,
        **registration_data_import_data,
    )
    created_obj_hct = RegistrationDataImport(
        status=RegistrationDataImport.IMPORTING,
        imported_by=user,
        data_source=data_source,
        number_of_individuals=import_data_obj.number_of_individuals,
        number_of_households=import_data_obj.number_of_households,
        business_area=business_area,
        pull_pictures=pull_pictures,
        screen_beneficiary=screen_beneficiary,
        **registration_data_import_data,
    )
    created_obj_hct.full_clean()
    created_obj_hct.save()

    created_obj_datahub.hct_id = created_obj_hct.id
    created_obj_datahub.save()

    created_obj_hct.datahub_id = created_obj_datahub.id
    created_obj_hct.save()

    return (
        created_obj_datahub,
        created_obj_hct,
        import_data_obj,
        business_area,
    )


class RegistrationXlsxImportMutationInput(graphene.InputObjectType):
    import_data_id = graphene.ID()
    name = graphene.String()
    business_area_slug = graphene.String()
    screen_beneficiary = graphene.Boolean()


class RegistrationKoboImportMutationInput(graphene.InputObjectType):
    import_data_id = graphene.String()
    name = graphene.String()
    pull_pictures = graphene.Boolean()
    business_area_slug = graphene.String()
    screen_beneficiary = graphene.Boolean()


class RegistrationXlsxImportMutation(BaseValidator, PermissionMutation, ValidationErrorMutationMixin):
    registration_data_import = graphene.Field(RegistrationDataImportNode)

    class Arguments:
        registration_data_import_data = RegistrationXlsxImportMutationInput(required=True)

    @classmethod
    @transaction.atomic(using="default")
    @transaction.atomic(using="registration_datahub")
    @is_authenticated
    def processed_mutate(cls, root, info, registration_data_import_data):
        (
            created_obj_datahub,
            created_obj_hct,
            import_data_obj,
            business_area,
        ) = create_registration_data_import_objects(registration_data_import_data, info.context.user, "XLS")

        cls.has_permission(info, Permissions.RDI_IMPORT_DATA, business_area)

        if (
            created_obj_hct.should_check_against_sanction_list()
            and not business_area.should_check_against_sanction_list()
        ):
            raise ValidationError("Cannot check against sanction list")

        log_create(
            RegistrationDataImport.ACTIVITY_LOG_MAPPING, "business_area", info.context.user, None, created_obj_hct
        )
        registration_xlsx_import_task.delay(
            registration_data_import_id=str(created_obj_datahub.id),
            import_data_id=str(import_data_obj.id),
            business_area=str(business_area.id),
        )

        return RegistrationXlsxImportMutation(registration_data_import=created_obj_hct)


class RegistrationDeduplicationMutation(BaseValidator, PermissionMutation):
    ok = graphene.Boolean()

    class Arguments:
        registration_data_import_datahub_id = graphene.ID(required=True)
        version = BigInt(required=False)

    @classmethod
    def validate_object_status(cls, rdi_obj, *args, **kwargs):
        if rdi_obj.status != RegistrationDataImport.DEDUPLICATION_FAILED:
            logger.error(
                "Deduplication can only be called when Registration Data Import status is Deduplication Failed"
            )
            raise ValidationError(
                "Deduplication can only be called when Registration Data Import status is Deduplication Failed"
            )

    @classmethod
    @is_authenticated
    def mutate(cls, root, info, registration_data_import_datahub_id, **kwargs):
        old_rdi_obj = RegistrationDataImport.objects.get(datahub_id=registration_data_import_datahub_id)
        rdi_obj = RegistrationDataImport.objects.get(datahub_id=registration_data_import_datahub_id)
        check_concurrency_version_in_mutation(kwargs.get("version"), rdi_obj)
        cls.has_permission(info, Permissions.RDI_RERUN_DEDUPE, rdi_obj.business_area)

        cls.validate(rdi_obj=rdi_obj)

        rdi_obj.status = RegistrationDataImport.DEDUPLICATION
        rdi_obj.save()
        log_create(
            RegistrationDataImport.ACTIVITY_LOG_MAPPING, "business_area", info.context.user, old_rdi_obj, rdi_obj
        )
        rdi_deduplication_task.delay(registration_data_import_id=str(registration_data_import_datahub_id))

        return cls(ok=True)


class RegistrationKoboImportMutation(BaseValidator, PermissionMutation, ValidationErrorMutationMixin):
    registration_data_import = graphene.Field(RegistrationDataImportNode)

    class Arguments:
        registration_data_import_data = RegistrationKoboImportMutationInput(required=True)

    @classmethod
    def check_is_not_empty(cls, import_data_id):
        import_data = get_object_or_404(ImportData, id=decode_id_string(import_data_id))
        if import_data.number_of_households == 0 and import_data.number_of_individuals == 0:
            logger.error("Cannot import empty KoBo form")
            raise ValidationError("Cannot import empty KoBo form")

    @classmethod
    @transaction.atomic(using="default")
    @transaction.atomic(using="registration_datahub")
    @is_authenticated
    def processed_mutate(cls, root, info, registration_data_import_data):
        cls.check_is_not_empty(registration_data_import_data.import_data_id)

        (
            created_obj_datahub,
            created_obj_hct,
            import_data_obj,
            business_area,
        ) = create_registration_data_import_objects(registration_data_import_data, info.context.user, "KOBO")

        cls.has_permission(info, Permissions.RDI_IMPORT_DATA, business_area)

        if (
            created_obj_hct.should_check_against_sanction_list()
            and not business_area.should_check_against_sanction_list()
        ):
            raise ValidationError("Cannot check against sanction list")

        log_create(
            RegistrationDataImport.ACTIVITY_LOG_MAPPING, "business_area", info.context.user, None, created_obj_hct
        )
        registration_kobo_import_task.delay(
            registration_data_import_id=str(created_obj_datahub.id),
            import_data_id=str(import_data_obj.id),
            business_area=str(business_area.id),
        )

        return RegistrationXlsxImportMutation(registration_data_import=created_obj_hct)


class MergeRegistrationDataImportMutation(BaseValidator, PermissionMutation):
    registration_data_import = graphene.Field(RegistrationDataImportNode)

    class Arguments:
        id = graphene.ID(required=True)
        version = BigInt(required=False)

    @classmethod
    def validate_object_status(cls, *args, **kwargs):
        status = kwargs.get("status")
        if status != RegistrationDataImport.IN_REVIEW:
            logger.error("Only In Review Registration Data Import can be merged into Population")
            raise ValidationError("Only In Review Registration Data Import can be merged into Population")

    @classmethod
    @transaction.atomic(using="default")
    @transaction.atomic(using="registration_datahub")
    @is_authenticated
    def mutate(cls, root, info, id, **kwargs):
        decode_id = decode_id_string(id)
        old_obj_hct = RegistrationDataImport.objects.get(
            id=decode_id,
        )

        obj_hct = RegistrationDataImport.objects.get(
            id=decode_id,
        )
        check_concurrency_version_in_mutation(kwargs.get("version"), obj_hct)

        cls.has_permission(info, Permissions.RDI_MERGE_IMPORT, obj_hct.business_area)

        cls.validate(status=obj_hct.status)
        obj_hct.status = RegistrationDataImport.MERGING
        obj_hct.save()
        merge_registration_data_import_task.delay(registration_data_import_id=decode_id)

        log_create(
            RegistrationDataImport.ACTIVITY_LOG_MAPPING, "business_area", info.context.user, old_obj_hct, obj_hct
        )
        return MergeRegistrationDataImportMutation(obj_hct)


class UploadImportDataXLSXFile(PermissionMutation):
    import_data = graphene.Field(ImportDataNode)
    errors = graphene.List(XlsxRowErrorNode)

    class Arguments:
        file = Upload(required=True)
        business_area_slug = graphene.String(required=True)

    @classmethod
    @is_authenticated
    def mutate(cls, root, info, file, business_area_slug):

        cls.has_permission(info, Permissions.RDI_IMPORT_DATA, business_area_slug)
        errors = UploadXLSXInstanceValidator().validate_everything(file, business_area_slug)
        if errors:
            errors.sort(key=operator.itemgetter("row_number", "header"))
            return UploadImportDataXLSXFile(None, errors)

        wb = openpyxl.load_workbook(file)

        hh_sheet = wb["Households"]
        ind_sheet = wb["Individuals"]

        number_of_households = 0
        number_of_individuals = 0

        # Could just return max_row if openpyxl won't count empty rows too
        for row in hh_sheet.iter_rows(min_row=3):
            if not any([cell.value for cell in row]):
                continue
            number_of_households += 1

        for row in ind_sheet.iter_rows(min_row=3):

            if not any([cell.value for cell in row]):
                continue
            number_of_individuals += 1

        created = ImportData.objects.create(
            file=file,
            number_of_households=number_of_households,
            number_of_individuals=number_of_individuals,
        )

        return UploadImportDataXLSXFile(created, [])


class SaveKoboProjectImportDataMutation(PermissionMutation):
    import_data = graphene.Field(ImportDataNode)
    errors = graphene.List(KoboErrorNode)

    class Arguments:
        uid = Upload(required=True)
        business_area_slug = graphene.String(required=True)
        only_active_submissions = graphene.Boolean(required=True)

    @classmethod
    @is_authenticated
    def mutate(cls, root, info, uid, business_area_slug, only_active_submissions):
        cls.has_permission(info, Permissions.RDI_IMPORT_DATA, business_area_slug)

        kobo_api = KoboAPI(business_area_slug)

        submissions = kobo_api.get_project_submissions(uid, only_active_submissions)

        business_area = BusinessArea.objects.get(slug=business_area_slug)
        validator = KoboProjectImportDataInstanceValidator()
        errors = validator.validate_everything(submissions, business_area.name)

        if errors:
            errors.sort(key=operator.itemgetter("header"))
            return UploadImportDataXLSXFile(None, errors)

        number_of_households, number_of_individuals = count_population(submissions)

        import_file_name = f"project-uid-{uid}-{time.time()}.json"
        file = File(BytesIO(json.dumps(submissions).encode()), name=import_file_name)

        created = ImportData.objects.create(
            file=file,
            number_of_households=number_of_households,
            number_of_individuals=number_of_individuals,
        )

        return SaveKoboProjectImportDataMutation(created, [])


class DeleteRegistrationDataImport(graphene.Mutation):
    ok = graphene.Boolean()

    class Arguments:
        registration_data_import_id = graphene.String(required=True)

    @classmethod
    @is_authenticated
    def mutate(cls, root, info, **kwargs):
        decoded_id = decode_id_string(kwargs.get("registration_data_import_id"))
        rdi_obj = RegistrationDataImport.objects.get(id=decoded_id)
        rdi_obj.delete()

        log_create(RegistrationDataImport.ACTIVITY_LOG_MAPPING, "business_area", info.context.user, rdi_obj, None)
        return cls(ok=True)


class Mutations(graphene.ObjectType):
    upload_import_data_xlsx_file = UploadImportDataXLSXFile.Field()
    delete_registration_data_import = DeleteRegistrationDataImport.Field()
    registration_xlsx_import = RegistrationXlsxImportMutation.Field()
    registration_kobo_import = RegistrationKoboImportMutation.Field()
    save_kobo_import_data = SaveKoboProjectImportDataMutation.Field()
    merge_registration_data_import = MergeRegistrationDataImportMutation.Field()
    rerun_dedupe = RegistrationDeduplicationMutation.Field()
