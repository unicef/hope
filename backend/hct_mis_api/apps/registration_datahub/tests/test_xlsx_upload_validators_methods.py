import operator
from typing import Tuple
from unittest import mock

from django.conf import settings
from django.core.management import call_command

import openpyxl
import pytest
from parameterized import parameterized

from hct_mis_api.apps.core.base_test_case import APITestCase
from hct_mis_api.apps.core.field_attributes.core_fields_attributes import (
    get_core_fields_attributes,
)
from hct_mis_api.apps.core.fixtures import (
    create_afghanistan,
    create_pdu_flexible_attribute,
)
from hct_mis_api.apps.core.models import DataCollectingType, PeriodicFieldData
from hct_mis_api.apps.core.utils import SheetImageLoader
from hct_mis_api.apps.geo.fixtures import CountryFactory
from hct_mis_api.apps.payment.fixtures import generate_delivery_mechanisms
from hct_mis_api.apps.program.fixtures import get_program_with_dct_type_and_name
from hct_mis_api.apps.registration_datahub.validators import UploadXLSXInstanceValidator


class TestXLSXValidatorsMethods(APITestCase):
    databases = {"default"}
    fixtures = (f"{settings.PROJECT_ROOT}/apps/geo/fixtures/data.json",)

    FILES_DIR_PATH = f"{settings.PROJECT_ROOT}/apps/registration_datahub/tests/test_file"

    @classmethod
    def setUpTestData(cls) -> None:
        call_command("loadflexfieldsattributes")
        generate_delivery_mechanisms()

        cls.business_area = create_afghanistan()
        cls.program = get_program_with_dct_type_and_name()
        cls.social_worker_program = get_program_with_dct_type_and_name(dct_type=DataCollectingType.Type.SOCIAL)
        cls.country = CountryFactory()
        cls.business_area.countries.add(cls.country)

    def test_string_validator(self) -> None:
        validator = UploadXLSXInstanceValidator(self.program)
        self.assertTrue(validator.string_validator("Marek", "full_name_i_c"))

    def test_float_validator(self) -> None:
        validator = UploadXLSXInstanceValidator(self.program)
        self.assertFalse(validator.float_validator(None, "estimated_birth_date_i_c"))
        self.assertTrue(validator.float_validator(None, "age_at_registration"))
        self.assertTrue(validator.float_validator(1.1, "estimated_birth_date_i_c"))
        self.assertFalse(validator.float_validator("1.a1a", "estimated_birth_date_i_c"))

    def test_geolocation_validator(self) -> None:
        # test correct values:
        correct_values = (
            "1.1, 1.1",
            "0.0, 0.0",
            "54.1234252, 67.535232",
        )
        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value in correct_values:
            self.assertTrue(upload_xlsx_instance_validator.geolocation_validator(value, "hh_geopoint_h_c"))

        # test incorrect values:
        incorrect_values = (
            "1, 1, 1, 1",
            "0, 0",
            "52.124.124, 1241.242",
            "24.121a, bcd421.222",
        )
        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value in incorrect_values:
            self.assertFalse(upload_xlsx_instance_validator.geolocation_validator(value, "hh_geopoint_h_c"))

    def test_date_validator(self) -> None:
        # test correct values:
        correct_values = (
            "01-03-1994",
            "1-3-1994",
            "27-12-2020",
            "27/12/2020",
            "27.12.2020",
            "27.12.2020",
        )
        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value in correct_values:
            self.assertTrue(upload_xlsx_instance_validator.date_validator(value, "birth_date_i_c"))

        # test incorrect values:
        incorrect_values = (
            "13-13-1994",
            "213.22.2020",
            "qwerty",
            "24",
            "-24",
        )
        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value in incorrect_values:
            self.assertFalse(upload_xlsx_instance_validator.date_validator(value, "birth_date_i_c"))

    def test_integer_validator(self) -> None:
        # test correct values:
        correct_values = (
            "12",
            "0",
            0,
            12,
            12345,
            -12,
        )
        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value in correct_values:
            self.assertTrue(upload_xlsx_instance_validator.integer_validator(value, "size_h_c"))

        # test incorrect values:
        incorrect_values = (
            "13-13-1994",
            "213.22.2020",
            "qwerty",
            # 12.2345,
            "12,242",
        )

        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value in incorrect_values:
            self.assertFalse(upload_xlsx_instance_validator.integer_validator(value, "size_h_c"))

    def test_phone_validator(self) -> None:
        # test correct values:
        correct_values = (
            "+1-202-555-0172",
            "+44 1632 960852",
            "+1-613-555-0182",
            "+61 1900 654 321",
            "+36 55 979 922",
            "+353 20 915 8245",
            "+48 69 563 7300",
        )
        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value in correct_values:
            self.assertTrue(upload_xlsx_instance_validator.phone_validator(value, "phone_no_i_c"))

        # test incorrect values:
        incorrect_values: Tuple = (
            "123 123 123",  # no region code
            "13-13-1994",
            "213.22.2020",
            "qwerty",
            12.2345,
            "12,242",
            "12",
            12,
        )

        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value in incorrect_values:
            self.assertFalse(upload_xlsx_instance_validator.phone_validator(value, "phone_no_i_c"))

    def test_choice_validator(self) -> None:
        test_correct_values = (("REFUGEE", "residence_status_h_c"),)
        test_incorrect_values = (
            ("YES", "work_status"),
            ("OTHER", "work_status"),
            ("Hearing Problems", "disability"),
            ("Option 37", "assistance_type_h_f"),
        )
        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value, header in test_correct_values:
            self.assertTrue(upload_xlsx_instance_validator.choice_validator(value, header))

        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        for value, header in test_incorrect_values:
            self.assertFalse(upload_xlsx_instance_validator.choice_validator(value, header))

    def test_rows_validator_too_many_head_of_households(self) -> None:
        wb = openpyxl.load_workbook(
            f"{self.FILES_DIR_PATH}/error-xlsx.xlsx",
            data_only=True,
        )
        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
        upload_xlsx_instance_validator.rows_validator(wb["Households"])
        upload_xlsx_instance_validator.errors = []
        upload_xlsx_instance_validator.rows_validator(wb["Individuals"])
        expected = [
            {
                "row_number": 0,
                "header": "relationship_i_c",
                "message": "Sheet: 'Individuals', There are multiple head of households for household with id: 3",
            }
        ]
        self.assertEqual(expected, upload_xlsx_instance_validator.errors)

    def test_rows_validator(self) -> None:
        wb = openpyxl.load_workbook(
            f"{self.FILES_DIR_PATH}/invalid_rows.xlsx",
            data_only=True,
        )

        wb_valid = openpyxl.load_workbook(
            f"{self.FILES_DIR_PATH}/new_reg_data_import.xlsx",
            data_only=True,
        )

        invalid_file = (
            (
                wb["Households"],
                [
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 1 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 3,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 1, Option 2, Option "
                        "3 for type select many of field assistance_type_h_f",
                        "row_number": 4,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 13 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 5,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 3 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 6,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 1, Option 3 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 7,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 2, Option 3 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 8,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 2 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 9,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 1, Option 2, Option "
                        "4 for type select many of field assistance_type_h_f",
                        "row_number": 10,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 4 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 11,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 5 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 12,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 1, Option 4 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 13,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 2, Option 4 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 14,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 3 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 15,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 1, Option 2, Option "
                        "5 for type select many of field assistance_type_h_f",
                        "row_number": 16,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 6 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 17,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 7 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 18,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 1, Option 5 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 19,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 2, Option 5 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 20,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: 'Households', Unexpected value: Option 4 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 21,
                    },
                ],
            ),
            (
                wb["Individuals"],
                [
                    # TODO: fix this? (rebase issue?)
                    # {
                    #     "row_number": 4,
                    #     "header": "preferred_language_i_c",
                    #     "message": "Sheet: 'Individuals', Unexpected value: Test for type select one of field preferred_language_i_c",
                    # },
                    {
                        "row_number": 8,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', There is no household with provided id: TEXT",
                    },
                    {
                        "row_number": 29,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', There is no household with provided id: 52",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 34 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 35 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 36 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 37 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 38 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 39 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 40 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 41 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 42 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 43 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 44 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 45 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 46 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 47 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 48 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 49 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 50 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: 51 has to have a head of household",
                    },
                    {
                        "row_number": 0,
                        "header": "relationship_i_c",
                        "message": "Sheet: 'Individuals', Household with id: Some Text has to have a head of household",
                    },
                ],
            ),
        )
        valid_file = (
            (
                wb_valid["Households"],
                [],
            ),
            (
                wb_valid["Individuals"],
                [],
            ),
        )
        files = (invalid_file, valid_file)
        for file in files:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            for sheet, expected_values in file:
                upload_xlsx_instance_validator.image_loader = SheetImageLoader(sheet)
                upload_xlsx_instance_validator.errors = []
                upload_xlsx_instance_validator.rows_validator(sheet, self.business_area.slug)
                self.assertEqual(upload_xlsx_instance_validator.errors, expected_values)

    def test_validate_file_extension(self) -> None:
        file_path, expected_values = (
            f"{self.FILES_DIR_PATH}/" f"image.png",
            [{"row_number": 1, "message": "Only .xlsx files are accepted for import"}],
        )
        with open(file_path, "rb") as file:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            upload_xlsx_instance_validator.validate_file_extension(file)
            self.assertEqual(upload_xlsx_instance_validator.errors[0]["row_number"], expected_values[0]["row_number"])
            self.assertEqual(upload_xlsx_instance_validator.errors[0]["message"], expected_values[0]["message"])

            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            errors, delivery_mechanisms_errors = upload_xlsx_instance_validator.validate_everything(file, "afghanistan")
            self.assertEqual(errors[0]["row_number"], expected_values[0]["row_number"])
            self.assertEqual(errors[0]["message"], expected_values[0]["message"])
            self.assertEqual(delivery_mechanisms_errors, [])

    def test_validate_file_content_as_xlsx(self) -> None:
        file_path, expected_values = (
            f"{self.FILES_DIR_PATH}/" f"not_excel_file.xlsx",
            [{"row_number": 1, "message": "Invalid .xlsx file"}],
        )
        with open(file_path, "rb") as file:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            result, _ = upload_xlsx_instance_validator.validate_everything(file, "afghanistan")
            self.assertEqual(result[0]["row_number"], expected_values[0]["row_number"])
            self.assertEqual(result[0]["message"], expected_values[0]["message"])

    def test_validate_file_with_template(self) -> None:
        invalid_cols_file_path = f"{self.FILES_DIR_PATH}/new_reg_data_import.xlsx"
        with open(invalid_cols_file_path, "rb") as file:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            wb = openpyxl.load_workbook(file, data_only=True)
            upload_xlsx_instance_validator.validate_file_with_template(wb)
            errors = upload_xlsx_instance_validator.errors
            errors.sort(key=operator.itemgetter("row_number", "header"))
            self.assertEqual(errors, [])

    def test_required_validator(self) -> None:
        with mock.patch(
            "hct_mis_api.apps.registration_datahub.validators.UploadXLSXInstanceValidator.get_all_fields",
            lambda *args: {"test": {"required": True}},
        ):
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            result = upload_xlsx_instance_validator.required_validator(value="tak", header="test")
            self.assertTrue(result)

        with mock.patch(
            "hct_mis_api.apps.registration_datahub.validators.UploadXLSXInstanceValidator.get_all_fields",
            lambda *args: {"test": {"required": True}},
        ):
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            result = upload_xlsx_instance_validator.required_validator(value="", header="test")
            self.assertFalse(result)

        with mock.patch(
            "hct_mis_api.apps.registration_datahub.validators.UploadXLSXInstanceValidator.get_all_fields",
            lambda *args: {"test": {"required": False}},
        ):
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            result = upload_xlsx_instance_validator.required_validator(value="", header="test")
            self.assertTrue(result)

    def test_validate_empty_file(self) -> None:
        empty_file_path = f"{self.FILES_DIR_PATH}/empty_rdi.xlsx"
        wb = openpyxl.load_workbook(
            empty_file_path,
            data_only=True,
        )
        upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)

        expected_result = [
            {
                "header": "Households",
                "message": "There aren't households in the file.",
                "row_number": 1,
            },
            {
                "header": "Individuals",
                "message": "There aren't individuals in the file.",
                "row_number": 1,
            },
        ]

        upload_xlsx_instance_validator.validate_collectors_size(wb)
        self.assertEqual(upload_xlsx_instance_validator.errors, expected_result)

    def test_validate_collector_unique(self) -> None:
        file_path = f"{self.FILES_DIR_PATH}/test_collectors.xlsx"

        expected_result = [
            {
                "row_number": 3,
                "header": "Individuals",
                "message": "Individual from row: 3 cannot be the primary and the alternate collector for households: 992630574 at the same time.",
            },
            {
                "row_number": 4,
                "header": "Individuals",
                "message": "Individual from row: 4 cannot be the primary and the alternate collector for households: 853780211 at the same time.",
            },
        ]

        with open(file_path, "rb") as file:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            result, _ = upload_xlsx_instance_validator.validate_everything(file, "afghanistan")
        self.assertEqual(result, expected_result)

    def test_validate_incorrect_admin_area(self) -> None:
        file_path = f"{self.FILES_DIR_PATH}/invalid_area.xlsx"

        expected_result = [
            {
                "header": "admin1_h_c",
                "row_number": 3,
                "message": "Sheet: 'Households': Area with code: F-35 does not exist",
            },
            {
                "header": "admin2_h_c",
                "row_number": 3,
                "message": "Sheet: 'Households': Area with code: F-35 does not exist",
            },
            {
                "header": "admin1_h_c",
                "row_number": 4,
                "message": "Sheet: 'Households': Area with code: F-35 does not exist",
            },
            {
                "header": "admin2_h_c",
                "row_number": 4,
                "message": "Sheet: 'Households': Area with code: F-35 does not exist",
            },
            {
                "header": "admin1_h_c",
                "row_number": 6,
                "message": "Sheet: 'Households': Area with code: F-35 does not exist",
            },
            {
                "header": "admin2_h_c",
                "row_number": 6,
                "message": "Sheet: 'Households': Area with code: F-35 does not exist",
            },
        ]

        with open(file_path, "rb") as file:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            result, _ = upload_xlsx_instance_validator.validate_everything(file, "afghanistan")
        self.assertEqual(result, expected_result)

    def test_validate_people_sheet_invalid(self) -> None:
        file_path = f"{self.FILES_DIR_PATH}/rdi_people_test_invalid.xlsx"

        expected_result = [
            {
                "row_number": 1,
                "header": "People",
                "message": "There are duplicates with id(s): [1]. Number have to be unique in the field pp_index_id.",
            },
            {
                "row_number": 1,
                "header": "People",
                "message": "Invalid value in field 'pp_primary_collector_id' for Individual with index_id 1. Value cannot be empty for relationship NON_BENEFICIARY",
            },
            {
                "row_number": 1,
                "header": "People",
                "message": "Individual with index_id 1 has to have an Primary collector.",
            },
            {
                "row_number": 1,
                "header": "People",
                "message": "Individual with index_id 4 has to have an Primary collector.",
            },
            {
                "row_number": 1,
                "header": "People",
                "message": "Invalid value in field 'pp_relationship_i_c' with index_id 99. Value can be HEAD or NON_BENEFICIARY",
            },
            {
                "row_number": 3,
                "header": "pp_index_id",
                "message": "Sheet: 'People', Unexpected value: None for type integer of field pp_index_id",
            },
        ]
        with open(file_path, "rb") as file:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.social_worker_program)
            result, _ = upload_xlsx_instance_validator.validate_everything(file, "afghanistan")
        self.assertEqual(result, expected_result)

    def test_validate_people_sheet_valid(self) -> None:
        file_path = f"{self.FILES_DIR_PATH}/rdi_people_test.xlsx"

        with open(file_path, "rb") as file:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.social_worker_program)
            result, _ = upload_xlsx_instance_validator.validate_everything(file, "afghanistan")
        self.assertEqual(result, [])

    def test_validate_delivery_mechanism_data(self) -> None:
        file_path = f"{self.FILES_DIR_PATH}/rdi_import_3_hh_missing_required_delivery_fields.xlsx"
        with open(file_path, "rb") as file:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
            result, delivery_mechanisms_errors = upload_xlsx_instance_validator.validate_everything(file, "afghanistan")
        self.assertEqual(result, [])
        self.assertEqual(
            delivery_mechanisms_errors,
            [
                {
                    "header": "name_of_cardholder__atm_card_i_c",
                    "message": "Field name_of_cardholder__atm_card_i_c is required for delivery mechanism atm_card",
                    "row_number": 3,
                },
                {
                    "header": "name_of_cardholder__atm_card_i_c",
                    "message": "Field name_of_cardholder__atm_card_i_c is required for delivery mechanism atm_card",
                    "row_number": 14,
                },
                {
                    "header": "card_expiry_date__atm_card_i_c",
                    "message": "Field card_expiry_date__atm_card_i_c is required for delivery mechanism atm_card",
                    "row_number": 15,
                },
                {
                    "header": "name_of_cardholder__atm_card_i_c",
                    "message": "Field name_of_cardholder__atm_card_i_c is required for delivery mechanism atm_card",
                    "row_number": 15,
                },
            ],
        )

    @pytest.mark.skip("Fail on pipeline")
    def test_validate_delivery_mechanism_data_people(self) -> None:
        file_path = f"{self.FILES_DIR_PATH}/rdi_import_1_hh_10_people_missing_required_delivery_fields.xlsx"
        with open(file_path, "rb") as file:
            upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.social_worker_program)
            result, delivery_mechanisms_errors = upload_xlsx_instance_validator.validate_everything(file, "afghanistan")
        self.assertEqual(result, [])
        self.assertEqual(
            delivery_mechanisms_errors,
            [
                {
                    "header": "pp_card_expiry_date__atm_card_i_c",
                    "message": "Field pp_card_expiry_date__atm_card_i_c is required for delivery mechanism atm_card",
                    "row_number": 4,
                },
                {
                    "header": "pp_name_of_cardholder__atm_card_i_c",
                    "message": "Field pp_name_of_cardholder__atm_card_i_c is required for "
                    "delivery mechanism atm_card",
                    "row_number": 4,
                },
                {
                    "header": "pp_name_of_cardholder__atm_card_i_c",
                    "message": "Field pp_name_of_cardholder__atm_card_i_c is required for "
                    "delivery mechanism atm_card",
                    "row_number": 5,
                },
            ],
        )

    def test_validate_delivery_mechanism_data_global_fields_only_dropped(self) -> None:
        """
        Set full_name core field as required for payment.
        Validation ignores data that contains only global fields (no delivery mechanism specific fields).
        It's not possible to recognize which delivery mechanism to use if only global fields are provided.
        """
        self.maxDiff = None

        core_fields_attributes = get_core_fields_attributes().copy()
        full_name = [field for field in core_fields_attributes if field["name"] == "full_name"][0]
        full_name["required"] = False
        full_name["required_for_payment"] = True
        core_fields_attributes.append(full_name)

        with mock.patch(
            "hct_mis_api.apps.core.field_attributes.core_fields_attributes.get_core_fields_attributes",
            lambda: core_fields_attributes,
        ):
            file_path = f"{self.FILES_DIR_PATH}/rdi_import_3_hh_missing_required_delivery_fields.xlsx"
            with open(file_path, "rb") as file:
                upload_xlsx_instance_validator = UploadXLSXInstanceValidator(self.program)
                result, delivery_mechanisms_errors = upload_xlsx_instance_validator.validate_everything(
                    file, "afghanistan"
                )
            self.assertEqual(result, [])
            self.assertEqual(
                delivery_mechanisms_errors,
                [
                    {
                        "header": "name_of_cardholder__atm_card_i_c",
                        "message": "Field name_of_cardholder__atm_card_i_c is required for delivery "
                        "mechanism atm_card",
                        "row_number": 3,
                    },
                    {
                        "header": "name_of_cardholder__atm_card_i_c",
                        "message": "Field name_of_cardholder__atm_card_i_c is required for delivery "
                        "mechanism atm_card",
                        "row_number": 14,
                    },
                    {
                        "header": "card_expiry_date__atm_card_i_c",
                        "message": "Field card_expiry_date__atm_card_i_c is required for delivery mechanism atm_card",
                        "row_number": 15,
                    },
                    {
                        "header": "name_of_cardholder__atm_card_i_c",
                        "message": "Field name_of_cardholder__atm_card_i_c is required for delivery "
                        "mechanism atm_card",
                        "row_number": 15,
                    },
                ],
            )

    @parameterized.expand(
        [
            (PeriodicFieldData.STRING, ["Test", "2021-05-01"]),
            (PeriodicFieldData.DECIMAL, ["12.3", "2021-05-01"]),
            (PeriodicFieldData.BOOL, ["True", "2021-05-01"]),
            (PeriodicFieldData.DATE, ["1996-06-21", "2021-05-01"]),
        ]
    )
    def test_validate_pdu_string_valid(self, subtype: str, data_row: list) -> None:
        create_pdu_flexible_attribute(
            label="PDU Flex Attribute",
            subtype=subtype,
            number_of_rounds=1,
            rounds_names=["May"],
            program=self.program,
        )
        header_row = ["pdu_flex_attribute_round_1_value", "pdu_flex_attribute_round_1_collection_date"]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(header_row)
        sheet.append(data_row)
        validator = UploadXLSXInstanceValidator(self.program)
        errors = validator._validate_pdu(sheet[2], sheet[1], 3)
        self.assertEqual(errors, [])

    @parameterized.expand(
        [
            (PeriodicFieldData.DECIMAL, ["foo", "2021-05-01"]),
            (PeriodicFieldData.BOOL, ["foo", "2021-05-01"]),
            (PeriodicFieldData.DATE, ["foo", "2021-05-01"]),
        ]
    )
    def test_validate_pdu_string_value_error(self, subtype: str, data_row: list) -> None:
        create_pdu_flexible_attribute(
            label="PDU Flex Attribute",
            subtype=subtype,
            number_of_rounds=1,
            rounds_names=["May"],
            program=self.program,
        )
        header_row = ["pdu_flex_attribute_round_1_value", "pdu_flex_attribute_round_1_collection_date"]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(header_row)
        sheet.append(data_row)
        validator = UploadXLSXInstanceValidator(self.program)
        errors = validator._validate_pdu(sheet[2], sheet[1], 3)
        self.assertEqual(
            errors,
            [
                {
                    "row_number": 3,
                    "header": "pdu_flex_attribute_round_1_value",
                    "message": "Invalid value foo for field pdu_flex_attribute_round_1_value",
                }
            ],
        )

    def test_validate_pdu_wrong_collection_date(self) -> None:
        data_row = ["Test", "bar"]
        create_pdu_flexible_attribute(
            label="PDU Flex Attribute",
            subtype=PeriodicFieldData.STRING,
            number_of_rounds=1,
            rounds_names=["May"],
            program=self.program,
        )
        header_row = ["pdu_flex_attribute_round_1_value", "pdu_flex_attribute_round_1_collection_date"]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(header_row)
        sheet.append(data_row)
        validator = UploadXLSXInstanceValidator(self.program)
        errors = validator._validate_pdu(sheet[2], sheet[1], 3)
        self.assertEqual(
            errors,
            [
                {
                    "row_number": 3,
                    "header": "pdu_flex_attribute_round_1_collection_date",
                    "message": "Invalid value bar for field pdu_flex_attribute_round_1_collection_date",
                }
            ],
        )

    def test_validate_pdu_empty_row(self) -> None:
        data_row = ["Test", "bar"]
        create_pdu_flexible_attribute(
            label="PDU Flex Attribute",
            subtype=PeriodicFieldData.STRING,
            number_of_rounds=1,
            rounds_names=["May"],
            program=self.program,
        )
        header_row = []
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(header_row)
        sheet.append(data_row)
        validator = UploadXLSXInstanceValidator(self.program)
        errors = validator._validate_pdu(sheet[2], sheet[1], 3)
        self.assertEqual(errors, [])
