import json
import operator

from django.db import transaction

import openpyxl

from hct_mis_api.apps.registration_datahub.models import ImportData
from hct_mis_api.apps.registration_datahub.validators import UploadXLSXInstanceValidator


class ValidateXlsxImport:
    @transaction.atomic(using="default")
    @transaction.atomic(using="registration_datahub")
    def execute(self, import_data):
        import_data.status = ImportData.STATUS_RUNNING
        import_data.save()
        errors = UploadXLSXInstanceValidator().validate_everything(import_data.file, import_data.business_area_slug)
        if errors:
            errors.sort(key=operator.itemgetter("row_number", "header"))
            import_data.status = ImportData.STATUS_VALIDATION_ERROR
            import_data.validation_errors = json.dumps(errors)
        else:
            import_data.status = ImportData.STATUS_FINISHED

        wb = openpyxl.load_workbook(import_data.file)

        hh_sheet = wb["Households"]
        ind_sheet = wb["Individuals"]

        number_of_households = 0
        number_of_individuals = 0

        # Could just return max_row if openpyxl won't count empty rows too
        for row in hh_sheet.iter_rows(min_row=3):
            if not any([cell.value for cell in row]):
                continue
            number_of_households += 1

        for row in ind_sheet.iter_rows(min_row=3):

            if not any([cell.value for cell in row]):
                continue
            number_of_individuals += 1
        import_data.number_of_households = number_of_households
        import_data.number_of_individuals = number_of_individuals
        import_data.save()
        return {"import_data_id": import_data.id}
