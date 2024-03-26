import csv
import json
import os
import random
from argparse import ArgumentParser
from dataclasses import dataclass

import openpyxl
from faker import Faker

faker = Faker()


@dataclass
class Mapping:
    column_name: str
    type: int
    value: bool


def read_choices_from_csv(csv_file_path):
    choices_mapping = {}

    with open(csv_file_path, mode="r", encoding="utf-8") as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            if len(row) > 0:  # Ensures there is at least a key
                field_name = row[0].strip()
                choice_value = row[1].strip() if len(row) > 1 else ""  # Allows for an empty value
                if field_name in choices_mapping:
                    choices_mapping[field_name].append(choice_value)
                else:
                    choices_mapping[field_name] = [choice_value]
            else:
                print(f"Skipping invalid row: {row}")
    return choices_mapping


def get_random_choice_for_field(field, choices, default_value):
    return random.choice(choices[field]) if field in choices else default_value


def get_random_choice(column_name, type, value):
    if column_name[:3] == "pp_":
        real_column_name = column_name[3:]
        if real_column_name in choices:
            return random.choice(choices[real_column_name])
        real_column_name = real_column_name[:-3] + "h" + real_column_name[-2:]
        return random.choice(choices[real_column_name])
    return random.choice(choices[column_name])


this_file_path = os.path.dirname(os.path.abspath(__file__))
csv_filepath = f"{this_file_path}/choices.csv"
choices = read_choices_from_csv(csv_filepath)


def random_number(*args, **kwargs):
    return random.randint(1, 2**31)


def date(*args, **kwargs):
    return faker.date_between(start_date="-30y", end_date="today")


def name(*args, **kwargs):
    return faker.name()


def phone_number(*args, **kwargs):
    return faker.phone_number()


autoincrement_store = {}


def autoincrement(column_name, type, value):
    if column_name not in autoincrement_store:
        autoincrement_store[column_name] = -1
    autoincrement_store[column_name] += 1
    return autoincrement_store[column_name]


type_mapping = {
    "constant": lambda column_name, type, value: value,
    "autoincrement": autoincrement,
    "choice": get_random_choice,
    "age": lambda column_name, type, value: random.randint(0, 100),
    "date": date,
    "name": name,
    "phone_number": phone_number,
    "document_number": lambda column_name, type, value: faker.ssn(),
}


def load_header_mapping(file) -> list[Mapping]:
    with open(file, "r") as f:
        header_mapping = json.load(f, object_hook=lambda d: Mapping(**d))
    return header_mapping


def columns_to_json(header_mapping):
    columns = []
    for key, (column_name, value) in header_mapping.items():
        output_value = None
        if value == "SPECIAL":
            column_type = "special"
        elif value is None or type(value) == int or type(value) == float or type(value) == str or type(value) == bool:
            column_type = "constant"
            output_value = value
        elif callable(value) and value.__name__ == "<lambda>":
            column_type = "choice"
        elif callable(value) and value.__name__ == "random_number":
            column_type = "autoincrement"
        elif callable(value) and value.__name__ == "phone_number":
            column_type = "phone_number"
        elif callable(value) and value.__name__ == "name":
            column_type = "name"
        elif callable(value) and value.__name__ == "date":
            column_type = "date"
        else:
            print(value)
            raise
        columns.append({"column_name": column_name, "type": column_type, "value": output_value})
    print(json.dumps(columns, indent=2))


def main(amount: int, import_type: str, seed: int, output: str, extern_collector_amount: int):
    if seed:
        random.seed(seed)
    print(f"Generating xlsx file {amount} {import_type} with seed {seed}")
    generated_dir = output
    if not os.path.exists(generated_dir):
        os.makedirs(generated_dir)

    filepath = os.path.join(generated_dir, f"rdi_import_{amount}_{import_type}_seed_{seed}.xlsx".lower())
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Set column names for households, individuals and people in the first row
    if import_type == "people":
        handle_people(amount, wb, extern_collector_amount)
    if import_type == "households":
        handle_households(amount, wb, extern_collector_amount)

    wb.save(filepath)
    print(f"File saved to {filepath}")


def handle_households(amount, wb, extern_collector_amount):
    household_header_mapping = load_header_mapping(f"{this_file_path}/household_header_mapping.json")
    individual_header_mapping = load_header_mapping(f"{this_file_path}/individual_header_mapping.json")
    households_sheet = wb.create_sheet("Households")
    individuals_sheet = wb.create_sheet("Individuals")
    for index, mapping in enumerate(household_header_mapping, start=1):
        households_sheet.cell(row=1, column=index).value = mapping.column_name
    for index, mapping in enumerate(household_header_mapping, start=1):
        individuals_sheet.cell(row=1, column=index).value = mapping.column_name
    individual_row = 3  # Data starts from the third row for individuals
    for hh_row in range(3, amount + 3):
        individuals_amount = random.randint(1, 7)  # Random number of individuals 1 to 7
        # Fill in household data starting from the third row
        for index, mapping in enumerate(household_header_mapping, start=1):
            to_write = type_mapping[mapping.type](mapping.column_name, mapping.type, mapping.value)
            households_sheet.cell(row=hh_row, column=index).value = to_write

        # Generate and fill individuals for the current household
        handle_individuals(individual_header_mapping, individual_row, individuals_amount, individuals_sheet)


def handle_special_column_individuals(column_name, type, value, is_hoh, is_primary_collector, is_alternate_collector):
    if column_name == "household_id":
        return autoincrement_store["household_id"]
    if column_name == "relationship_i_c":
        if is_hoh:
            return "HEAD"
        else:
            choices_without_head = [choice for choice in choices["relationship_i_c"] if choice != "HEAD"]
            return random.choice(choices_without_head) if choices_without_head else "None"
    if column_name == "primary_collector_id" and is_primary_collector:
        return autoincrement_store["household_id"]
    if column_name == "alternate_collector_id" and is_alternate_collector:
        return autoincrement_store["household_id"]


def handle_individuals(
    individual_header_mapping: list[Mapping],
    individual_row,
    individuals_amount,
    individuals_sheet,
    extern_collector_amount,
):
    primary_collector_index = random.randint(0, individuals_amount - 1)
    alternate_collector_index = (
        None
        if individuals_amount == 1
        else random.choice([i for i in range(individuals_amount) if i != primary_collector_index])
    )

    for individual_index in range(individuals_amount):
        for index, mapping in enumerate(individual_header_mapping, start=1):
            if mapping.type == "special":
                to_write = handle_special_column_individuals(
                    mapping.column_name,
                    mapping.type,
                    mapping.value,
                    individual_index == 0,
                    individual_index == primary_collector_index,
                    individual_index == alternate_collector_index,
                )
            else:
                to_write = type_mapping[mapping.type](mapping.column_name, mapping.type, mapping.value)
            individuals_sheet.cell(row=individual_row, column=index).value = to_write
        individual_row += 1  # Increment row for the next individual


def handle_special_column_people(column_name, type, value, has_external_collector, collects_for_external_id):
    if column_name == "pp_primary_collector_id":
        if not has_external_collector and collects_for_external_id is None:
            return autoincrement_store["pp_index_id"]
        elif has_external_collector is True:
            return None
        else:
            return collects_for_external_id
    if column_name == "pp_relationship_i_c":
        return (
            "HEAD"
            if has_external_collector
            or collects_for_external_id is None
            or str(autoincrement_store["pp_index_id"])
            in ([] if isinstance(collects_for_external_id, int) else collects_for_external_id.split(","))
            else "NON_BENEFICIARY"
        )


def handle_people(amount, wb, extern_collector_amount):
    people_sheet = wb.create_sheet("People")
    people_header_mapping = load_header_mapping(f"{this_file_path}/people_header_mapping.json")
    for index, mapping in enumerate(people_header_mapping, start=1):
        people_sheet.cell(row=1, column=index).value = mapping.column_name
    people_row = 3  # Data starts from the third row for people
    collect_for_two_people = False
    if amount >= 3:
        collect_for_two_people = True
    for index in range(amount):
        has_external_collector = False
        if collect_for_two_people:
            has_external_collector = True
        if extern_collector_amount > 0 and not collect_for_two_people or collect_for_two_people and index == 1:
            has_external_collector = True
            extern_collector_amount -= 1
        people_row = handle_single_people_row(
            people_header_mapping, people_row, people_sheet, has_external_collector, None
        )

        if has_external_collector and not collect_for_two_people:
            collects_for_external_id = autoincrement_store["pp_index_id"]
            people_row = handle_single_people_row(
                people_header_mapping, people_row, people_sheet, False, collects_for_external_id
            )
        elif index == 1:
            collects_for_external_id = ",".join(("0", "1", "2"))
            people_row = handle_single_people_row(
                people_header_mapping, people_row, people_sheet, False, collects_for_external_id
            )
            collect_for_two_people = False


def handle_single_people_row(
    people_header_mapping, people_row, people_sheet, has_external_collector, collects_for_external_id
):
    for index, mapping in enumerate(people_header_mapping, start=1):
        if mapping.type == "special":
            to_write = handle_special_column_people(
                mapping.column_name,
                mapping.type,
                mapping.value,
                has_external_collector,
                collects_for_external_id,
            )
        else:
            to_write = type_mapping[mapping.type](mapping.column_name, mapping.type, mapping.value)
        people_sheet.cell(row=people_row, column=index).value = to_write
    people_row += 1
    return people_row


if __name__ == "__main__":
    parser = ArgumentParser(description="Generate a fake data XLSX file.")

    parser.add_argument("amount", type=int, help="Number of households or people to generate.")
    parser.add_argument(
        "--type", choices=["households", "people"], help="The type of data to process.", default="households"
    )
    parser.add_argument("--extern_collector_amount", type=int, help="Number of people to generate.", default=0)
    parser.add_argument("--seed", type=int, help="Seed for random number generator.", default=None)
    parser.add_argument("--output_dir", type=str, help="output file path", default="~/Documents/HOPE/RDI XLSX files")
    args = parser.parse_args()
    if args.extern_collector_amount > args.amount:
        parser.error("--extern_collector_amount cannot be greater than amount")
    main(args.amount, args.type, args.seed, args.output_dir, args.extern_collector_amount)
