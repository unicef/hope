from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from extras.test_utils.factories import (
    CurrencyFactory,
    DeliveryMechanismFactory,
    FinancialServiceProviderFactory,
    FinancialServiceProviderXlsxTemplateFactory,
    FspXlsxTemplatePerDeliveryMechanismFactory,
    PaymentFactory,
    PaymentHouseholdSnapshotFactory,
    PaymentPlanFactory,
    PaymentPlanGroupFactory,
)
from extras.test_utils.factories.program import ProgramCycleFactory, ProgramFactory
from extras.test_utils.selenium import HopeTestBrowser
from hope.models import BusinessArea, FinancialServiceProvider, Payment, PaymentPlan, Program

pytestmark = pytest.mark.django_db()

FSP_REFERENCE = "FSP-REFERENCE-E2E"
FSP_ROUTING_CODE = "FSP-ROUTING-E2E"
RECONCILIATION_NOTE = "RECONCILIATION-E2E"


@pytest.fixture
def fsp_extra_fields_payment_plan(business_area: BusinessArea) -> tuple[PaymentPlan, Payment]:
    program = ProgramFactory(
        name="FSP Extra Fields E2E Program",
        status=Program.ACTIVE,
        business_area=business_area,
    )
    program_cycle = ProgramCycleFactory(program=program)
    payment_plan_group = PaymentPlanGroupFactory(cycle=program_cycle, name="FSP Extra Fields E2E Group")
    delivery_mechanism = DeliveryMechanismFactory(
        code="fsp-extra-fields-e2e",
        name="FSP Extra Fields E2E",
        payment_gateway_id="fsp-extra-fields-e2e",
    )
    fsp = FinancialServiceProviderFactory(
        name="FSP Extra Fields E2E Provider",
        communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_XLSX,
    )
    fsp.delivery_mechanisms.add(delivery_mechanism)
    xlsx_template = FinancialServiceProviderXlsxTemplateFactory(
        name="FSP Extra Fields E2E Template",
        columns=["payment_id", "entitlement_quantity", "delivered_quantity"],
        core_fields=[],
        flex_fields=[],
        document_types=[],
    )
    FspXlsxTemplatePerDeliveryMechanismFactory(
        financial_service_provider=fsp,
        delivery_mechanism=delivery_mechanism,
        xlsx_template=xlsx_template,
    )
    currency = CurrencyFactory(code="USD", name="United States Dollar")
    payment_plan = PaymentPlanFactory(
        name="FSP Extra Fields E2E Payment Plan",
        program_cycle=program_cycle,
        payment_plan_group=payment_plan_group,
        business_area=business_area,
        status=PaymentPlan.Status.LOCKED_FSP,
        plan_type=PaymentPlan.PlanType.REGULAR,
        financial_service_provider=fsp,
        delivery_mechanism=delivery_mechanism,
        currency=currency,
        total_entitled_quantity=Decimal("125.00"),
        total_entitled_quantity_usd=Decimal("125.00"),
    )
    payment = PaymentFactory(
        parent=payment_plan,
        program=program,
        unicef_id="PAYMENT-FSP-EXTRA-FIELDS-E2E",
        status=Payment.STATUS_PENDING,
        currency=currency,
        delivery_type=delivery_mechanism,
        financial_service_provider=fsp,
        entitlement_quantity=Decimal("125.00"),
        entitlement_quantity_usd=Decimal("125.00"),
        delivered_quantity=None,
        delivered_quantity_usd=None,
    )
    payment.save(update_fields=["unicef_id"])
    PaymentHouseholdSnapshotFactory(payment=payment, snapshot_data={})
    return payment_plan, payment


def _confirm_workflow_action(browser: HopeTestBrowser, data_cy: str) -> None:
    browser.wait_for_element_clickable(f'[data-cy="{data_cy}"]').click()
    browser.wait_for_element_clickable('[data-cy="button-submit"]').click()


def _prepare_fsp_extra_fields_upload(template_path: str, upload_path: Path, payment_id: str) -> None:
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active
    headers = tuple(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    payment_row = tuple(next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    assert headers == ("payment_id",)
    assert payment_row == (payment_id,)
    worksheet.cell(row=1, column=2).value = "fsp_reference"
    worksheet.cell(row=1, column=3).value = "fsp_routing_code"
    worksheet.cell(row=2, column=2).value = FSP_REFERENCE
    worksheet.cell(row=2, column=3).value = FSP_ROUTING_CODE
    workbook.save(upload_path)


def _assert_export_and_prepare_reconciliation(
    export_path: str,
    reconciliation_path: Path,
    payment_id: str,
) -> None:
    workbook = openpyxl.load_workbook(export_path)
    worksheet = workbook.active
    headers = tuple(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    payment_row = tuple(next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    exported_payment = dict(zip(headers, payment_row, strict=True))
    assert headers == (
        "payment_id",
        "entitlement_quantity",
        "delivered_quantity",
        "fsp_reference",
        "fsp_routing_code",
    )
    assert exported_payment["payment_id"] == payment_id
    assert exported_payment["entitlement_quantity"] == 125
    assert exported_payment["delivered_quantity"] is None
    assert exported_payment["fsp_reference"] == FSP_REFERENCE
    assert exported_payment["fsp_routing_code"] == FSP_ROUTING_CODE
    worksheet.cell(row=2, column=headers.index("delivered_quantity") + 1).value = 125
    worksheet.cell(row=2, column=headers.index("fsp_reference") + 1).value = "RETURNED-FSP-VALUE"
    worksheet.cell(row=1, column=len(headers) + 1).value = "reconciliation_note"
    worksheet.cell(row=2, column=len(headers) + 1).value = RECONCILIATION_NOTE
    workbook.save(reconciliation_path)


def test_payment_fsp_extra_fields_full_xlsx_flow(
    login: HopeTestBrowser,
    business_area: BusinessArea,
    fsp_extra_fields_payment_plan: tuple[PaymentPlan, Payment],
    tmp_path: Path,
) -> None:
    payment_plan, payment = fsp_extra_fields_payment_plan
    program = payment_plan.program
    payment_plan_path = f"/{business_area.slug}/programs/{program.code}/payment-module/payment-plans/{payment_plan.id}"
    template_filename = f"payment_plan_{payment_plan.unicef_id}_fsp_extra_fields.xlsx"
    login.delete_downloaded_file_if_present(template_filename, browser=True)
    login.open(payment_plan_path)

    login.wait_for_element_clickable('[data-cy="button-fsp-extra-fields-template"]').click()
    login.assert_downloaded_file(template_filename, browser=True)
    template_path = login.get_path_of_downloaded_file(template_filename, browser=True)
    upload_path = tmp_path / "fsp_extra_fields_upload.xlsx"
    _prepare_fsp_extra_fields_upload(template_path, upload_path, payment.unicef_id)

    login.wait_for_element_clickable('[data-cy="button-fsp-extra-fields-upload"]').click()
    login.wait_for_element_visible('[data-cy="dialog-fsp-extra-fields-import"]')
    login.choose_file('[data-cy="dialog-fsp-extra-fields-import"] [data-cy="file-input"]', str(upload_path))
    login.wait_for_element_clickable('[data-cy="button-fsp-extra-fields-import-submit"]').click()
    login.wait_for_text("FSP extra fields import started.")

    payment.refresh_from_db()
    assert payment.extra_fields == {}
    assert payment.fsp_extra_fields == {
        "fsp_reference": FSP_REFERENCE,
        "fsp_routing_code": FSP_ROUTING_CODE,
    }

    login.open(payment_plan_path)
    login.wait_for_element_clickable('[data-cy="button-send-for-approval"]').click()
    login.wait_for_text("IN APPROVAL")
    _confirm_workflow_action(login, "button-approve")
    login.wait_for_text("IN AUTHORIZATION")
    _confirm_workflow_action(login, "button-authorize")
    login.wait_for_text("IN REVIEW")
    _confirm_workflow_action(login, "button-mark-as-released")
    login.wait_for_text("ACCEPTED")

    group = payment_plan.payment_plan_group
    group_path = f"/{business_area.slug}/programs/{program.code}/payment-module/groups/{group.id}"
    export_filename = f"payment_plan_group_{group.unicef_id}_payment_list_batch_1.xlsx"
    login.delete_downloaded_file_if_present(export_filename, browser=True)
    login.open(group_path)
    login.wait_for_element_clickable('[data-cy="button-delivery-export-xlsx-group"]').click()
    login.wait_for_element_visible('[data-cy="dialog-delivery-export-xlsx-group"]')
    login.wait_for_element_clickable('[data-cy="button-delivery-export-xlsx-group-submit"]').click()
    login.wait_for_text("Export started")
    login.open(group_path)
    login.wait_for_element_clickable('[data-cy="batch-download-link-1"]').click()

    login.assert_downloaded_file(export_filename, browser=True)
    export_path = login.get_path_of_downloaded_file(export_filename, browser=True)
    reconciliation_path = tmp_path / "fsp_extra_fields_reconciliation.xlsx"
    _assert_export_and_prepare_reconciliation(export_path, reconciliation_path, payment.unicef_id)

    login.wait_for_element_clickable('[data-cy="button-delivery-import-xlsx-group"]').click()
    login.wait_for_element_visible('[data-cy="dialog-delivery-import-xlsx-group"]')
    login.choose_file(
        '[data-cy="dialog-delivery-import-xlsx-group"] [data-cy="file-input"]',
        str(reconciliation_path),
    )
    login.wait_for_element_clickable('[data-cy="button-delivery-import-xlsx-group-submit"]').click()
    login.wait_for_text("Delivery reconciliation import started")

    payment.refresh_from_db()
    payment_plan.refresh_from_db()
    assert payment_plan.status == PaymentPlan.Status.FINISHED
    assert payment.delivered_quantity == Decimal("125.00")
    assert payment.extra_fields == {"reconciliation_note": RECONCILIATION_NOTE}
    assert payment.fsp_extra_fields == {
        "fsp_reference": FSP_REFERENCE,
        "fsp_routing_code": FSP_ROUTING_CODE,
    }

    login.open(f"/{business_area.slug}/programs/{program.code}/payment-module/payments/{payment.id}")
    login.wait_for_text(
        RECONCILIATION_NOTE,
        '[data-cy="label-Reconciliation Note"]',
        timeout=30,
    )
    login.wait_for_text(
        FSP_REFERENCE,
        '[data-cy="label-Fsp Reference"]',
        timeout=30,
    )
    login.wait_for_text(
        FSP_ROUTING_CODE,
        '[data-cy="label-Fsp Routing Code"]',
        timeout=30,
    )
    login.assert_text("Reconciliation Information: Extra Info")
    login.assert_text("FSP extra fields")
