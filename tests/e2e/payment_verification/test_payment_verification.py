from datetime import datetime
from decimal import Decimal
import os
from time import sleep

from dateutil.relativedelta import relativedelta
import openpyxl
import pytest
from selenium.webdriver.common.by import By

from e2e.page_object.grievance.details_grievance_page import GrievanceDetailsPage
from e2e.page_object.grievance.grievance_tickets import GrievanceTickets
from e2e.page_object.payment_verification.payment_record import PaymentRecord
from e2e.page_object.payment_verification.payment_verification import (
    PaymentVerification as PaymentVerificationComponent,
)
from e2e.page_object.payment_verification.payment_verification_details import (
    PaymentVerificationDetails,
)
from e2e.payment_module.test_payment_plans import find_file
from extras.test_utils.factories.core import DataCollectingTypeFactory
from extras.test_utils.factories.household import create_household
from extras.test_utils.factories.payment import (
    FinancialServiceProviderFactory,
    PaymentFactory,
    PaymentPlanFactory,
    PaymentVerificationFactory,
    PaymentVerificationPlanFactory,
    PaymentVerificationSummaryFactory,
    generate_delivery_mechanisms,
)
from extras.test_utils.factories.program import ProgramFactory
from extras.test_utils.factories.registration_data import RegistrationDataImportFactory
from hope.apps.account.models import User
from hope.apps.core.models import BusinessArea, DataCollectingType
from hope.apps.geo.models import Area
from hope.apps.payment.models import (
    DeliveryMechanism,
    Payment,
    PaymentPlan,
    PaymentVerification,
    PaymentVerificationPlan,
)
from hope.apps.program.models import BeneficiaryGroup, Program, ProgramCycle

pytestmark = pytest.mark.django_db()


@pytest.fixture
def active_program() -> Program:
    return get_program_with_dct_type_and_name("Active Program", "ACTI", status=Program.ACTIVE)


def get_program_with_dct_type_and_name(
    name: str,
    programme_code: str,
    dct_type: str = DataCollectingType.Type.STANDARD,
    status: str = Program.ACTIVE,
) -> Program:
    dct = DataCollectingTypeFactory(type=dct_type)
    beneficiary_group = BeneficiaryGroup.objects.filter(name="Main Menu").first()
    return ProgramFactory(
        name=name,
        programme_code=programme_code,
        start_date=datetime.now() - relativedelta(months=1),
        end_date=datetime.now() + relativedelta(months=1),
        data_collecting_type=dct,
        status=status,
        beneficiary_group=beneficiary_group,
    )


def create_program(
    name: str = "Test Program",
    dct_type: str = DataCollectingType.Type.STANDARD,
    beneficiary_group_name: str = "Main Menu",
) -> Program:
    dct = DataCollectingTypeFactory(type=dct_type)
    beneficiary_group = BeneficiaryGroup.objects.filter(name=beneficiary_group_name).first()
    return ProgramFactory(
        name=name,
        programme_code="1234",
        start_date=datetime.now() - relativedelta(months=1),
        end_date=datetime.now() + relativedelta(months=1),
        data_collecting_type=dct,
        status=Program.ACTIVE,
        cycle__title="First cycle for Test Program",
        cycle__status=ProgramCycle.DRAFT,
        cycle__start_date=datetime.now() - relativedelta(days=5),
        cycle__end_date=datetime.now() + relativedelta(days=5),
        beneficiary_group=beneficiary_group,
    )


@pytest.fixture
def social_worker_program() -> Program:
    return create_program(dct_type=DataCollectingType.Type.SOCIAL, beneficiary_group_name="People")


@pytest.fixture
def payment_verification_3() -> None:
    payment_verification_multiple_verification_plans(3)


def payment_verification_multiple_verification_plans(
    number_verification_plans: int,
) -> None:
    registration_data_import = RegistrationDataImportFactory(
        imported_by=User.objects.first(), business_area=BusinessArea.objects.first()
    )
    program = Program.objects.filter(name="Active Program").first()
    households = []
    for _ in range(number_verification_plans):
        household, _ = create_household(
            {
                "registration_data_import": registration_data_import,
                "admin2": Area.objects.order_by("?").first(),
                "program": program,
            },
            {"registration_data_import": registration_data_import},
        )
        households.append(household)

    payment_plan = PaymentPlanFactory(
        program_cycle=program.cycles.first(),
        status=PaymentPlan.Status.FINISHED,
        business_area=BusinessArea.objects.filter(slug="afghanistan").first(),
    )
    payments = [
        PaymentFactory(
            parent=payment_plan,
            business_area=BusinessArea.objects.first(),
            household=hh,
            head_of_household=household.head_of_household,
            entitlement_quantity=Decimal(21.36),
            delivered_quantity=Decimal(21.36),
            currency="PLN",
            status=Payment.STATUS_DISTRIBUTION_SUCCESS,
        )
        for hh in households
    ]

    PaymentVerificationSummaryFactory(payment_plan=payment_plan)

    for payment in payments:
        payment_verification_plan = PaymentVerificationPlanFactory(
            payment_plan=payment_plan,
            verification_channel=PaymentVerificationPlan.VERIFICATION_CHANNEL_MANUAL,
        )

        PaymentVerificationFactory(
            payment=payment,
            payment_verification_plan=payment_verification_plan,
            status=PaymentVerification.STATUS_PENDING,
        )


@pytest.fixture
def empty_payment_verification(social_worker_program: Program) -> None:
    registration_data_import = RegistrationDataImportFactory(
        imported_by=User.objects.first(), business_area=BusinessArea.objects.first()
    )
    program = Program.objects.filter(name="Active Program").first()
    household, individuals = create_household(
        {
            "registration_data_import": registration_data_import,
            "admin2": Area.objects.order_by("?").first(),
            "program": program,
        },
        {"registration_data_import": registration_data_import},
    )

    payment_plan = PaymentPlanFactory(
        program_cycle=program.cycles.first(),
        status=PaymentPlan.Status.FINISHED,
        business_area=BusinessArea.objects.filter(slug="afghanistan").first(),
    )
    PaymentFactory(
        parent=payment_plan,
        business_area=BusinessArea.objects.first(),
        household=household,
        head_of_household=household.head_of_household,
        entitlement_quantity=Decimal(21.36),
        delivered_quantity=Decimal(21.36),
        currency="PLN",
        status=Payment.STATUS_DISTRIBUTION_SUCCESS,
    )
    PaymentVerificationSummaryFactory(payment_plan=payment_plan)


@pytest.fixture
def add_payment_verification() -> PaymentVerification:
    return payment_verification_creator()


@pytest.fixture
def add_payment_verification_xlsx() -> PaymentVerification:
    return payment_verification_creator(channel=PaymentVerificationPlan.VERIFICATION_CHANNEL_XLSX)


def payment_verification_creator(
    channel: str = PaymentVerificationPlan.VERIFICATION_CHANNEL_MANUAL,
) -> PaymentVerification:
    generate_delivery_mechanisms()
    user = User.objects.first()
    business_area = BusinessArea.objects.first()
    registration_data_import = RegistrationDataImportFactory(imported_by=user, business_area=business_area)
    program = Program.objects.filter(name="Active Program").first()
    household, individuals = create_household(
        {
            "registration_data_import": registration_data_import,
            "admin2": Area.objects.order_by("?").first(),
            "program": program,
        },
        {"registration_data_import": registration_data_import},
    )

    dm_cash = DeliveryMechanism.objects.get(code="cash")
    fsp = FinancialServiceProviderFactory()
    fsp.delivery_mechanisms.set([dm_cash])

    payment_plan = PaymentPlanFactory(
        name="TEST",
        status=PaymentPlan.Status.FINISHED,
        program_cycle=program.cycles.first(),
        business_area=business_area,
        start_date=datetime.now() - relativedelta(months=1),
        end_date=datetime.now() + relativedelta(months=1),
        created_by=user,
        financial_service_provider=fsp,
        delivery_mechanism=dm_cash,
    )

    payment_plan.unicef_id = "PP-0000-00-1122334"
    payment_plan.save()
    PaymentVerificationSummaryFactory(
        payment_plan=payment_plan,
    )

    payment = PaymentFactory(
        parent=payment_plan,
        household=household,
        head_of_household=household.head_of_household,
        entitlement_quantity=21.36,
        delivered_quantity=21.36,
        currency="PLN",
        delivery_type=dm_cash,
        status=Payment.STATUS_DISTRIBUTION_SUCCESS,
    )
    payment_verification_plan = PaymentVerificationPlanFactory(
        payment_plan=payment_plan,
        verification_channel=channel,
    )
    return PaymentVerificationFactory(
        payment=payment,
        payment_verification_plan=payment_verification_plan,
        status=PaymentVerification.STATUS_PENDING,
    )


@pytest.fixture
def clear_downloaded_files(download_path: str) -> None:
    for file in os.listdir(download_path):
        os.remove(os.path.join(download_path, file))
    yield
    for file in os.listdir(download_path):
        os.remove(os.path.join(download_path, file))


@pytest.mark.usefixtures("login")
class TestSmokePaymentVerification:
    def test_smoke_payment_verification(
        self,
        active_program: Program,
        add_payment_verification: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")
        page_payment_verification.get_nav_payment_verification().click()
        assert "Payment Verification" in page_payment_verification.get_page_header_title().text
        assert "List of Payment Plans" in page_payment_verification.get_table_title().text
        assert "Payment Plan ID" in page_payment_verification.get_unicef_id().text
        assert "Verification Status" in page_payment_verification.get_verification_status().text
        assert "Total Amount" in page_payment_verification.get_total_delivered_quantity().text
        assert "Payment Disbursement Dates" in page_payment_verification.get_start_date().text
        assert "Programme Cycle Title" in page_payment_verification.get_cycle_title_header().text
        assert "Last Modified Date" in page_payment_verification.get_updated_at().text
        assert "PP-0000-00-1122334" in page_payment_verification.get_cash_plan_table_row().text
        assert "PENDING" in page_payment_verification.get_status_container().text
        assert active_program.cycles.first().title == page_payment_verification.get_cycle_title().text
        assert "Rows per page: 5 1–1 of 1" in page_payment_verification.get_table_pagination().text.replace("\n", " ")

    def test_smoke_payment_verification_details(
        self,
        active_program: Program,
        add_payment_verification: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")
        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()
        assert "Payment Plan PP-0000-00-1122334" in page_payment_verification_details.get_page_header_title().text
        assert "CREATE VERIFICATION PLAN" in page_payment_verification_details.get_button_new_plan().text
        assert "Payment Plan Details" in page_payment_verification_details.get_div_payment_plan_details().text
        assert "Active Program" in page_payment_verification_details.get_label_programme_name().text
        page_payment_verification_details.get_label_payment_records()
        page_payment_verification_details.get_label_start_date()
        page_payment_verification_details.get_label_end_date()
        page_payment_verification_details.get_table_label()
        assert "0%" in page_payment_verification_details.get_label_successful().text
        assert "0%" in page_payment_verification_details.get_label_erroneous().text
        assert "PENDING" in page_payment_verification_details.get_label_status().text
        assert "PENDING" in page_payment_verification_details.get_verification_plans_summary_status().text
        assert (
            "COMPLETION DATE -"
            in page_payment_verification_details.get_labelized_field_container_summary_completion_date().text.replace(
                "\n", " "
            )
        )
        assert "-" in page_payment_verification_details.get_label_completion_date().text
        assert (
            "NUMBER OF VERIFICATION PLANS 1"
            in page_payment_verification_details.get_labelized_field_container_summary_number_of_plans().text.replace(
                "\n", " "
            )
        )
        assert "1" in page_payment_verification_details.get_label_number_of_verification_plans().text
        assert "DELETE" in page_payment_verification_details.get_button_delete_plan().text
        assert "EDIT" in page_payment_verification_details.get_button_edit_plan().text
        assert "ACTIVATE" in page_payment_verification_details.get_button_activate_plan().text
        assert "PENDING" in page_payment_verification_details.get_label_status().text
        assert "PENDING" in page_payment_verification_details.get_verification_plan_status().text
        assert "MANUAL" in page_payment_verification_details.get_label_verification_channel().text

    def test_happy_path_payment_verification(
        self,
        active_program: Program,
        add_payment_verification: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
        page_payment_record: PaymentRecord,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")
        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()
        assert "1" in page_payment_verification_details.get_label_payment_records().text
        assert (datetime.now() - relativedelta(months=1)).strftime(
            "%-d %b %Y"
        ) in page_payment_verification_details.get_label_start_date().text
        assert (datetime.now() + relativedelta(months=1)).strftime(
            "%-d %b %Y"
        ) in page_payment_verification_details.get_label_end_date().text
        assert "Reconciliation Summary" in page_payment_verification_details.get_table_label().text
        payment_verification = add_payment_verification.payment_verification_plan
        assert (
            payment_verification.sampling.lower().replace("_", " ")
            in page_payment_verification_details.get_label_sampling().text.lower()
        )
        assert str(payment_verification.responded_count) in page_payment_verification_details.get_label_responded().text
        assert (
            str(payment_verification.received_with_problems_count)
            in page_payment_verification_details.get_label_received_with_issues().text
        )
        assert str(payment_verification.sample_size) in page_payment_verification_details.get_label_sample_size().text
        assert str(payment_verification.received_count) in page_payment_verification_details.get_label_received().text
        assert (
            str(payment_verification.not_received_count)
            in page_payment_verification_details.get_label_not_received().text
        )
        page_payment_verification_details.get_button_delete_plan().click()
        page_payment_verification_details.get_button_submit().click()
        try:
            page_payment_verification_details.get_button_new_plan().click()
        except BaseException:
            sleep(3)
            page_payment_verification_details.get_button_new_plan().click()
        page_payment_verification_details.get_button_submit().click()

        page_payment_verification_details.get_button_activate_plan().click()
        page_payment_verification_details.get_button_submit().click()

        page_payment_verification_details.get_rows()[0].find_elements(By.TAG_NAME, "a")[0].click()
        payment_record = Payment.objects.first()
        assert "Payment" in page_payment_record.get_page_header_title().text
        assert "VERIFY" in page_payment_record.get_button_ed_plan().text
        assert "DELIVERED FULLY" in page_payment_record.get_label_status()[0].text
        assert "DELIVERED FULLY" in page_payment_record.get_status_container().text
        assert payment_record.household.unicef_id in page_payment_record.get_label_household().text
        assert payment_record.parent.name in page_payment_record.get_label_target_population().text
        assert payment_record.parent.unicef_id in page_payment_record.get_label_distribution_modality().text
        assert payment_record.payment_verifications.first().status in page_payment_record.get_label_status()[1].text
        assert "PLN 0.00" in page_payment_record.get_label_amount_received().text
        assert payment_record.household.unicef_id in page_payment_record.get_label_household_id().text
        assert "21.36" in page_payment_record.get_label_entitlement_quantity().text
        assert "21.36" in page_payment_record.get_label_delivered_quantity().text
        assert "PLN" in page_payment_record.get_label_currency().text
        assert payment_record.delivery_type.name in page_payment_record.get_label_delivery_type().text
        assert payment_record.financial_service_provider.name in page_payment_record.get_label_fsp().text

        page_payment_record.get_button_ed_plan().click()

        page_payment_record.get_input_received_amount().click()
        page_payment_record.get_input_received_amount().send_keys("100")
        page_payment_record.get_button_submit().click()

        for _ in range(5):
            if "RECEIVED WITH ISSUES" in page_payment_record.get_label_status()[1].text:
                break
            sleep(1)
        assert "RECEIVED WITH ISSUES" in page_payment_record.get_label_status()[1].text
        page_payment_record.get_arrow_back().click()

        assert page_payment_record.wait_for_status_container("RECEIVED WITH ISSUES")
        assert page_payment_record.get_status_container().text == "RECEIVED WITH ISSUES"

    def test_payment_verification_successful_not_received(
        self,
        active_program: Program,
        add_payment_verification: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
        page_payment_record: PaymentRecord,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")
        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()
        page_payment_verification_details.get_button_activate_plan().click()
        page_payment_verification_details.get_button_submit().click()

        assert len(page_payment_verification_details.get_rows()) == 1
        page_payment_verification_details.scroll(execute=2)
        page_payment_verification_details.get_rows()[0].find_element(By.TAG_NAME, "a").click()
        page_payment_record.get_button_ed_plan().click()
        page_payment_record.get_choice_not_received().click()
        page_payment_record.get_button_submit().click()
        page_payment_record.get_arrow_back().click()

        assert page_payment_record.wait_for_status_container("NOT RECEIVED")

    def test_payment_verification_partially_successful_received_and_grievance_ticket(
        self,
        active_program: Program,
        add_payment_verification: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
        page_payment_record: PaymentRecord,
        page_grievance_tickets: GrievanceTickets,
        page_grievance_details_page: GrievanceDetailsPage,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")

        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()
        page_payment_verification_details.get_button_activate_plan().click()
        page_payment_verification_details.get_button_submit().click()

        assert len(page_payment_verification_details.get_rows()) == 1
        page_payment_verification_details.scroll(execute=2)
        page_payment_verification_details.get_rows()[0].find_element(By.TAG_NAME, "a").click()
        quantity = float(page_payment_record.get_label_delivered_quantity().text) - 1
        page_payment_record.get_button_ed_plan().click()
        page_payment_record.get_input_received_amount().send_keys(str(quantity))
        page_payment_record.get_button_submit().click()
        page_payment_record.get_arrow_back().click()

        assert page_payment_record.wait_for_status_container("RECEIVED WITH ISSUES")

        page_payment_verification_details.get_button_finish().click()
        page_payment_verification_details.get_button_submit().click()

        page_grievance_tickets.get_nav_grievance().click()
        page_grievance_tickets.get_tab_system_generated().click()
        assert len(page_grievance_tickets.get_ticket_list_row()) == 1
        page_grievance_tickets.get_ticket_list_row()[0].click()
        page_grievance_details_page.get_button_assign_to_me().click()
        page_grievance_details_page.get_button_set_in_progress().click()

        page_grievance_details_page.get_grievance_verify().click()
        page_grievance_details_page.get_input_new_received_amount().send_keys(str(quantity + 1))
        page_grievance_details_page.get_button_submit().click()

        page_grievance_details_page.get_button_send_for_approval().click()
        page_grievance_details_page.get_grievance_approve().click()
        page_grievance_details_page.get_button_confirm().click()

        page_grievance_details_page.get_button_close_ticket().click()
        page_grievance_details_page.get_button_confirm().click()

        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()

        assert page_payment_record.wait_for_status_container("RECEIVED")
        assert page_payment_record.get_status_container().text == "RECEIVED"

        page_grievance_tickets.scroll(execute=2)

    def test_payment_verification_by_payment_related_complaint(
        self,
        active_program: Program,
        add_payment_verification: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
        page_payment_record: PaymentRecord,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")

    @pytest.mark.xfail(reason="UNSTABLE")
    def test_payment_verification_xlsx_successful(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
        download_path: str,
        page_payment_record: PaymentRecord,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")

        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()
        page_payment_verification_details.get_button_activate_plan().click()
        page_payment_verification_details.get_button_submit().click()

        page_payment_verification_details.get_export_xlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        page_payment_verification_details.driver.refresh()

        page_payment_verification_details.get_download_xlsx().click()

        xlsx_file = find_file(".xlsx", number_of_ties=10, search_in_dir=download_path)
        wb1 = openpyxl.load_workbook(os.path.join(download_path, xlsx_file))
        ws1 = wb1.active
        for cell in ws1["N:N"]:
            if cell.row >= 2:
                ws1.cell(row=cell.row, column=3, value="YES")
                ws1.cell(
                    row=cell.row,
                    column=16,
                    value=ws1.cell(row=cell.row, column=15).value,
                )

        wb1.save(os.path.join(download_path, "new_" + xlsx_file))
        find_file("new_" + xlsx_file, number_of_ties=10, search_in_dir=download_path)
        page_payment_verification_details.get_import_xlsx().click()

        page_payment_verification_details.upload_file(
            os.path.abspath(os.path.join(download_path, "new_" + xlsx_file)),
            timeout=120,
        )
        page_payment_verification_details.get_button_import_entitlement().click()

        assert page_payment_record.wait_for_status_container("RECEIVED", timeout=60)
        assert page_payment_record.get_status_container().text == "RECEIVED"

    @pytest.mark.xfail(reason="UNSTABLE")
    def test_payment_verification_xlsx_partially_successful(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
        page_payment_record: PaymentRecord,
        download_path: str,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")

        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()
        page_payment_verification_details.get_button_activate_plan().click()
        page_payment_verification_details.get_button_submit().click()

        assert len(page_payment_verification_details.get_rows()) == 1
        page_payment_verification_details.scroll(execute=2)
        page_payment_verification_details.get_rows()[0].find_element(By.TAG_NAME, "a").click()
        quantity = page_payment_record.get_label_delivered_quantity().text
        page_payment_record.get_arrow_back().click()

        page_payment_verification_details.get_export_xlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        page_payment_verification_details.driver.refresh()

        page_payment_verification_details.get_download_xlsx().click()

        xlsx_file = find_file(".xlsx", number_of_ties=10, search_in_dir=download_path)
        wb1 = openpyxl.load_workbook(os.path.join(download_path, xlsx_file))
        ws1 = wb1.active
        for cell in ws1["N:N"]:
            if cell.row >= 2:
                ws1.cell(row=cell.row, column=3, value="YES")
                ws1.cell(row=cell.row, column=16, value=float(quantity) - 1.0)

        wb1.save(os.path.join(download_path, xlsx_file))
        page_payment_verification_details.get_import_xlsx().click()

        page_payment_verification_details.upload_file(
            os.path.abspath(os.path.join(download_path, xlsx_file)), timeout=120
        )

        page_payment_verification_details.get_button_import_entitlement().click()

        assert page_payment_record.wait_for_status_container("RECEIVED WITH ISSUES")

    def test_payment_verification_xlsx_not_received(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
        page_payment_record: PaymentRecord,
        download_path: str,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")

        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()
        page_payment_verification_details.get_button_activate_plan().click()
        page_payment_verification_details.get_button_submit().click()

        page_payment_verification_details.get_export_xlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        page_payment_verification_details.driver.refresh()

        page_payment_verification_details.get_download_xlsx().click()

        xlsx_file = find_file(".xlsx", number_of_ties=10, search_in_dir=download_path)
        wb1 = openpyxl.load_workbook(os.path.join(download_path, xlsx_file))
        ws1 = wb1.active
        for cell in ws1["N:N"]:
            if cell.row >= 2:
                ws1.cell(row=cell.row, column=3, value="NO")
                ws1.cell(row=cell.row, column=16, value=0)

        wb1.save(os.path.join(download_path, xlsx_file))
        page_payment_verification_details.get_import_xlsx().click()

        page_payment_verification_details.upload_file(
            os.path.abspath(os.path.join(download_path, xlsx_file)), timeout=120
        )

        page_payment_verification_details.get_button_import_entitlement().click()

        assert page_payment_record.wait_for_status_container("NOT RECEIVED")

    def test_payment_verification_discard(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
        page_payment_record: PaymentRecord,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")

        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()
        page_payment_verification_details.get_button_activate_plan().click()
        page_payment_verification_details.get_button_submit().click()
        page_payment_verification_details.get_button_discard().click()
        page_payment_verification_details.get_button_submit().click()

        page_payment_verification_details.get_button_activate_plan().click()
        page_payment_verification_details.get_button_submit().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        page_payment_verification_details.driver.refresh()

        page_payment_verification_details.get_export_xlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        page_payment_verification_details.driver.refresh()

        page_payment_verification_details.get_download_xlsx().click()

        page_payment_verification_details.get_button_discard().click()
        page_payment_verification_details.get_button_submit().click()

        page_payment_verification_details.check_alert("You cant discard if xlsx file was downloaded or imported")

    def test_payment_verification_xlsx_invalid(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PaymentVerification,
        page_payment_verification: PaymentVerificationComponent,
        page_payment_verification_details: PaymentVerificationDetails,
        page_payment_record: PaymentRecord,
    ) -> None:
        page_payment_verification.select_global_program_filter("Active Program")

        page_payment_verification.get_nav_payment_verification().click()
        page_payment_verification.get_cash_plan_table_row().click()
        page_payment_verification_details.get_button_activate_plan().click()
        page_payment_verification_details.get_button_submit().click()

        page_payment_verification_details.get_export_xlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        page_payment_verification_details.driver.refresh()

        page_payment_verification_details.get_download_xlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        page_payment_verification_details.driver.refresh()

        page_payment_verification_details.get_button_mark_as_invalid().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        page_payment_verification_details.driver.refresh()

        assert "INVALID" in page_payment_verification_details.get_verification_plan_status().text
