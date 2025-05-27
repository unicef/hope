import os
from tempfile import NamedTemporaryFile, _TemporaryFileWrapper
from typing import Any

import openpyxl
import pytest

from hct_mis_api.apps.core.fixtures import (DataCollectingTypeFactory,
                                            create_afghanistan)
from hct_mis_api.apps.core.models import (DataCollectingType,
                                          FlexibleAttribute, PeriodicFieldData)
from hct_mis_api.apps.household.fixtures import \
    create_household_and_individuals
from hct_mis_api.apps.household.models import Individual
from hct_mis_api.apps.periodic_data_update.fixtures import (
    PeriodicDataUpdateTemplateFactory, PeriodicDataUpdateUploadFactory)
from hct_mis_api.apps.periodic_data_update.models import (
    PeriodicDataUpdateTemplate, PeriodicDataUpdateUpload)
from hct_mis_api.apps.periodic_data_update.service.periodic_data_update_export_template_service import \
    PeriodicDataUpdateExportTemplateService
from hct_mis_api.apps.periodic_data_update.utils import (
    field_label_to_field_name, populate_pdu_with_null_values)
from hct_mis_api.apps.program.fixtures import ProgramFactory
from hct_mis_api.apps.program.models import BeneficiaryGroup, Program
from hct_mis_api.apps.registration_data.fixtures import \
    RegistrationDataImportFactory
from tests.selenium.page_object.programme_population.individuals import \
    Individuals
from tests.selenium.page_object.programme_population.periodic_data_update_templates import \
    PeriodicDatUpdateTemplates
from tests.selenium.page_object.programme_population.periodic_data_update_uploads import \
    PeriodicDataUpdateUploads

pytestmark = pytest.mark.django_db()


@pytest.fixture
def clear_downloaded_files(download_path: str) -> None:
    for file in os.listdir(download_path):
        os.remove(os.path.join(download_path, file))
    yield
    for file in os.listdir(download_path):
        os.remove(os.path.join(download_path, file))


@pytest.fixture
def program() -> Program:
    business_area = create_afghanistan()
    dct = DataCollectingTypeFactory(type=DataCollectingType.Type.STANDARD)
    beneficiary_group = BeneficiaryGroup.objects.filter(name="Main Menu").first()
    return ProgramFactory(
        name="Test Program",
        status=Program.ACTIVE,
        business_area=business_area,
        data_collecting_type=dct,
        beneficiary_group=beneficiary_group,
    )


@pytest.fixture
def individual(program: Program) -> Individual:
    business_area = create_afghanistan()
    rdi = RegistrationDataImportFactory()
    household, individuals = create_household_and_individuals(
        household_data={
            "business_area": business_area,
            "program_id": program.pk,
            "registration_data_import": rdi,
        },
        individuals_data=[
            {
                "business_area": business_area,
                "program_id": program.pk,
                "registration_data_import": rdi,
            },
        ],
    )
    return individuals[0]


@pytest.fixture
def string_attribute(program: Program) -> FlexibleAttribute:
    return create_flexible_attribute(
        label="Test String Attribute",
        subtype=FlexibleAttribute.STRING,
        number_of_rounds=1,
        rounds_names=["Test Round"],
        program=program,
    )


@pytest.fixture
def date_attribute(program: Program) -> FlexibleAttribute:
    return create_flexible_attribute(
        label="Test Date Attribute",
        subtype=FlexibleAttribute.DATE,
        number_of_rounds=1,
        rounds_names=["Test Round"],
        program=program,
    )


def create_flexible_attribute(
    label: str, subtype: str, number_of_rounds: int, rounds_names: list[str], program: Program
) -> FlexibleAttribute:
    name = field_label_to_field_name(label)
    flexible_attribute = FlexibleAttribute.objects.create(
        label={"English(EN)": label},
        name=name,
        type=FlexibleAttribute.PDU,
        associated_with=FlexibleAttribute.ASSOCIATED_WITH_INDIVIDUAL,
        program=program,
    )
    flexible_attribute.pdu_data = PeriodicFieldData.objects.create(
        subtype=subtype, number_of_rounds=number_of_rounds, rounds_names=rounds_names
    )
    flexible_attribute.save()
    return flexible_attribute


def add_pdu_data_to_xlsx(
    periodic_data_update_template: PeriodicDataUpdateTemplate, rows: list[list[Any]]
) -> _TemporaryFileWrapper:
    wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
    ws_pdu = wb[PeriodicDataUpdateExportTemplateService.PDU_SHEET]
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            ws_pdu.cell(row=row_index + 2, column=col_index + 7, value=value)
    tmp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)
    tmp_file.seek(0)
    return tmp_file


def prepare_xlsx_file(rounds_data: list, rows: list, program: Program) -> _TemporaryFileWrapper:
    periodic_data_update_template = PeriodicDataUpdateTemplate.objects.create(
        program=program,
        business_area=program.business_area,
        filters=dict(),
        rounds_data=rounds_data,
    )
    service = PeriodicDataUpdateExportTemplateService(periodic_data_update_template)
    service.generate_workbook()
    service.save_xlsx_file()
    tmp_file = add_pdu_data_to_xlsx(periodic_data_update_template, rows)
    tmp_file.seek(0)
    return tmp_file


@pytest.mark.usefixtures("login")
class TestPeriodicDataUpdateUpload:
    # @flaky(max_runs=5, min_passes=1)
    def test_periodic_data_update_upload_success(
        self,
        clear_downloaded_files: None,
        program: Program,
        individual: Individual,
        string_attribute: FlexibleAttribute,
        pageIndividuals: Individuals,
    ) -> None:
        populate_pdu_with_null_values(program, individual.flex_fields)
        individual.save()
        flexible_attribute = string_attribute
        tmp_file = prepare_xlsx_file(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["Test Value", "2021-05-02"]],
            program,
        )
        pageIndividuals.selectGlobalProgramFilter(program.name)
        pageIndividuals.getNavProgrammePopulation().click()
        pageIndividuals.getNavIndividuals().click()
        pageIndividuals.getTabPeriodicDataUpdates().click()
        pageIndividuals.getButtonImport().click()
        pageIndividuals.getDialogImport()
        assert "IMPORT" in pageIndividuals.getButtonImportSubmit().text
        pageIndividuals.upload_file(tmp_file.name)
        pageIndividuals.getButtonImportSubmit().click()
        pageIndividuals.getPduUpdates().click()
        periodic_data_update_upload = PeriodicDataUpdateUpload.objects.first()
        assert periodic_data_update_upload.status == PeriodicDataUpdateUpload.Status.SUCCESSFUL
        assert periodic_data_update_upload.error_message is None
        individual.refresh_from_db()
        assert individual.flex_fields[flexible_attribute.name]["1"]["value"] == "Test Value"
        assert individual.flex_fields[flexible_attribute.name]["1"]["collection_date"] == "2021-05-02"
        assert pageIndividuals.getUpdateStatus(periodic_data_update_upload.pk).text == "SUCCESSFUL"

    @pytest.mark.night
    def test_periodic_data_update_upload_form_error(
        self,
        clear_downloaded_files: None,
        program: Program,
        individual: Individual,
        date_attribute: FlexibleAttribute,
        pageIndividuals: Individuals,
    ) -> None:
        populate_pdu_with_null_values(program, individual.flex_fields)
        individual.save()
        flexible_attribute = date_attribute
        tmp_file = prepare_xlsx_file(
            [
                {
                    "field": flexible_attribute.name,
                    "round": 1,
                    "round_name": flexible_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
            [["Test Value", "2021-05-02"]],
            program,
        )
        pageIndividuals.selectGlobalProgramFilter(program.name)
        pageIndividuals.getNavProgrammePopulation().click()
        try:
            pageIndividuals.getNavIndividuals().click()
        except BaseException:
            pageIndividuals.getNavProgrammePopulation().click()
            pageIndividuals.getNavIndividuals().click()
        pageIndividuals.getTabPeriodicDataUpdates().click()
        pageIndividuals.getButtonImport().click()
        pageIndividuals.getDialogImport()
        pageIndividuals.upload_file(tmp_file.name)
        pageIndividuals.getButtonImportSubmit().click()
        pageIndividuals.getPduUpdates().click()
        pageIndividuals.getStatusContainer()
        periodic_data_update_upload = PeriodicDataUpdateUpload.objects.first()
        assert periodic_data_update_upload.status == PeriodicDataUpdateUpload.Status.FAILED
        assert pageIndividuals.getStatusContainer().text == "FAILED"
        assert pageIndividuals.getUpdateStatus(periodic_data_update_upload.pk).text == "FAILED"
        pageIndividuals.getUpdateDetailsBtn(periodic_data_update_upload.pk).click()
        error_text = "Row: 2\ntest_date_attribute__round_value\nEnter a valid date."
        assert pageIndividuals.getPduFormErrors().text == error_text

    @pytest.mark.skip("Unskip after fix: 214341")
    @pytest.mark.night
    def test_periodic_data_update_upload_error(
        self,
        clear_downloaded_files: None,
        program: Program,
        individual: Individual,
        string_attribute: FlexibleAttribute,
        pageIndividuals: Individuals,
    ) -> None:
        populate_pdu_with_null_values(program, individual.flex_fields)
        individual.save()
        periodic_data_update_template = PeriodicDataUpdateTemplate.objects.create(
            program=program,
            business_area=program.business_area,
            filters=dict(),
            rounds_data=[
                {
                    "field": string_attribute.name,
                    "round": 1,
                    "round_name": string_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
        )
        service = PeriodicDataUpdateExportTemplateService(periodic_data_update_template)
        service.generate_workbook()
        service.save_xlsx_file()
        wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
        del wb.custom_doc_props[PeriodicDataUpdateExportTemplateService.PROPERTY_ID_NAME]
        ws_meta = wb[PeriodicDataUpdateExportTemplateService.META_SHEET]
        ws_meta[PeriodicDataUpdateExportTemplateService.META_ID_ADDRESS] = "-1"
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            pageIndividuals.selectGlobalProgramFilter(program.name)
            pageIndividuals.getNavProgrammePopulation().click()
            pageIndividuals.getNavIndividuals().click()
            pageIndividuals.getTabPeriodicDataUpdates().click()
            pageIndividuals.getButtonImport().click()
            pageIndividuals.getDialogImport()
            pageIndividuals.upload_file(tmp_file.name)
            pageIndividuals.getButtonImportSubmit().click()
            error_text = pageIndividuals.getPduUploadError().text
            assert error_text == "Periodic Data Update Template with ID -1 not found"

    @pytest.mark.night
    def test_periodic_data_uploads_list(
        self,
        clear_downloaded_files: None,
        program: Program,
        string_attribute: FlexibleAttribute,
        pageIndividuals: Individuals,
        pagePeriodicDataUpdateTemplates: PeriodicDatUpdateTemplates,
        pagePeriodicDataUploads: PeriodicDataUpdateUploads,
    ) -> None:
        periodic_data_update_template = PeriodicDataUpdateTemplateFactory(
            program=program,
            business_area=program.business_area,
            status=PeriodicDataUpdateTemplate.Status.TO_EXPORT,
            filters=dict(),
            rounds_data=[
                {
                    "field": string_attribute.name,
                    "round": 1,
                    "round_name": string_attribute.pdu_data.rounds_names[0],
                    "number_of_records": 0,
                }
            ],
        )
        pdu_upload = PeriodicDataUpdateUploadFactory(
            template=periodic_data_update_template,
            status=PeriodicDataUpdateUpload.Status.SUCCESSFUL,
        )
        pageIndividuals.selectGlobalProgramFilter(program.name)
        pageIndividuals.getNavProgrammePopulation().click()
        try:
            pageIndividuals.getNavIndividuals().click()
        except BaseException:
            pageIndividuals.getNavProgrammePopulation().click()
            pageIndividuals.getNavIndividuals().click()
        pageIndividuals.getTabPeriodicDataUpdates().click()
        pagePeriodicDataUpdateTemplates.getPduUpdatesBtn().click()

        index = pdu_upload.id
        assert str(index) in pagePeriodicDataUploads.getUpdateId(index).text
        assert str(pdu_upload.template.id) in pagePeriodicDataUploads.getUpdateTemplate(index).text
        assert f"{pdu_upload.created_at:%-d %b %Y}" in pagePeriodicDataUploads.getUpdateCreatedAt(index).text
        assert pdu_upload.created_by.get_full_name() in pagePeriodicDataUploads.getUpdateCreatedBy(index).text
        assert "SUCCESSFUL" in pagePeriodicDataUploads.getUpdateStatus(index).text
