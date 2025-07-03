import os
from datetime import datetime
from decimal import Decimal
from time import sleep

import openpyxl
import pytest
from dateutil.relativedelta import relativedelta
from selenium.webdriver.common.by import By

from hct_mis_api.apps.account.models import User
from hct_mis_api.apps.core.fixtures import DataCollectingTypeFactory
from hct_mis_api.apps.core.models import BusinessArea, DataCollectingType
from hct_mis_api.apps.geo.models import Area
from hct_mis_api.apps.household.fixtures import create_household
from hct_mis_api.apps.payment.fixtures import (
    FinancialServiceProviderFactory,
    PaymentFactory,
    PaymentPlanFactory,
    PaymentVerificationFactory,
    PaymentVerificationPlanFactory,
    PaymentVerificationSummaryFactory,
    generate_delivery_mechanisms,
)
from hct_mis_api.apps.payment.models import DeliveryMechanism, Payment, PaymentPlan
from hct_mis_api.apps.payment.models import PaymentVerification as PV
from hct_mis_api.apps.payment.models import PaymentVerificationPlan
from hct_mis_api.apps.program.fixtures import ProgramFactory
from hct_mis_api.apps.program.models import BeneficiaryGroup, Program, ProgramCycle
from hct_mis_api.apps.registration_data.fixtures import RegistrationDataImportFactory
from tests.selenium.page_object.grievance.details_grievance_page import (
    GrievanceDetailsPage,
)
from tests.selenium.page_object.grievance.grievance_tickets import GrievanceTickets
from tests.selenium.page_object.payment_verification.payment_record import PaymentRecord
from tests.selenium.page_object.payment_verification.payment_verification import (
    PaymentVerification,
)
from tests.selenium.page_object.payment_verification.payment_verification_details import (
    PaymentVerificationDetails,
)
from tests.selenium.payment_module.test_payment_plans import find_file

pytestmark = pytest.mark.django_db()


@pytest.fixture
def active_program() -> Program:
    return get_program_with_dct_type_and_name("Active Program", "ACTI", status=Program.ACTIVE)


def get_program_with_dct_type_and_name(
    name: str, programme_code: str, dct_type: str = DataCollectingType.Type.STANDARD, status: str = Program.ACTIVE
) -> Program:
    dct = DataCollectingTypeFactory(type=dct_type)
    beneficiary_group = BeneficiaryGroup.objects.filter(name="Main Menu").first()
    program = ProgramFactory(
        name=name,
        programme_code=programme_code,
        start_date=datetime.now() - relativedelta(months=1),
        end_date=datetime.now() + relativedelta(months=1),
        data_collecting_type=dct,
        status=status,
        beneficiary_group=beneficiary_group,
    )
    return program


def create_program(
    name: str = "Test Program",
    dct_type: str = DataCollectingType.Type.STANDARD,
    beneficiary_group_name: str = "Main Menu",
) -> Program:
    dct = DataCollectingTypeFactory(type=dct_type)
    beneficiary_group = BeneficiaryGroup.objects.filter(name=beneficiary_group_name).first()
    yield ProgramFactory(
        name=name,
        programme_code="1234",
        start_date=datetime.now() - relativedelta(months=1),
        end_date=datetime.now() + relativedelta(months=1),
        data_collecting_type=dct,
        status=Program.ACTIVE,
        cycle__title="First cycle for Test Program",
        cycle__status=ProgramCycle.DRAFT,
        cycle__start_date=datetime.now() - relativedelta(days=5),
        cycle__end_date=datetime.now() + relativedelta(days=5),
        beneficiary_group=beneficiary_group,
    )


@pytest.fixture
def social_worker_program() -> Program:
    yield create_program(dct_type=DataCollectingType.Type.SOCIAL, beneficiary_group_name="People")


@pytest.fixture
def payment_verification_3() -> None:
    payment_verification_multiple_verification_plans(3)


def payment_verification_multiple_verification_plans(number_verification_plans: int) -> None:
    registration_data_import = RegistrationDataImportFactory(
        imported_by=User.objects.first(), business_area=BusinessArea.objects.first()
    )
    program = Program.objects.filter(name="Active Program").first()
    households = list()
    for _ in range(number_verification_plans):
        household, _ = create_household(
            {
                "registration_data_import": registration_data_import,
                "admin_area": Area.objects.order_by("?").first(),
                "program": program,
            },
            {"registration_data_import": registration_data_import},
        )
        households.append(household)

    payment_plan = PaymentPlanFactory(
        program_cycle=program.cycles.first(),
        status=PaymentPlan.Status.FINISHED,
        business_area=BusinessArea.objects.filter(slug="afghanistan").first(),
    )
    payments = list()
    for hh in households:
        payments.append(
            PaymentFactory(
                parent=payment_plan,
                business_area=BusinessArea.objects.first(),
                household=hh,
                head_of_household=household.head_of_household,
                entitlement_quantity=Decimal(21.36),
                delivered_quantity=Decimal(21.36),
                currency="PLN",
                status=Payment.STATUS_DISTRIBUTION_SUCCESS,
            )
        )

    PaymentVerificationSummaryFactory(payment_plan=payment_plan)

    for payment in payments:
        payment_verification_plan = PaymentVerificationPlanFactory(
            payment_plan=payment_plan,
            verification_channel=PaymentVerificationPlan.VERIFICATION_CHANNEL_MANUAL,
        )

        PaymentVerificationFactory(
            payment=payment,
            payment_verification_plan=payment_verification_plan,
            status=PV.STATUS_PENDING,
        )


@pytest.fixture
def empty_payment_verification(social_worker_program: Program) -> None:
    registration_data_import = RegistrationDataImportFactory(
        imported_by=User.objects.first(), business_area=BusinessArea.objects.first()
    )
    program = Program.objects.filter(name="Active Program").first()
    household, individuals = create_household(
        {
            "registration_data_import": registration_data_import,
            "admin_area": Area.objects.order_by("?").first(),
            "program": program,
        },
        {"registration_data_import": registration_data_import},
    )

    payment_plan = PaymentPlanFactory(
        program_cycle=program.cycles.first(),
        status=PaymentPlan.Status.FINISHED,
        business_area=BusinessArea.objects.filter(slug="afghanistan").first(),
    )
    PaymentFactory(
        parent=payment_plan,
        business_area=BusinessArea.objects.first(),
        household=household,
        head_of_household=household.head_of_household,
        entitlement_quantity=Decimal(21.36),
        delivered_quantity=Decimal(21.36),
        currency="PLN",
        status=Payment.STATUS_DISTRIBUTION_SUCCESS,
    )
    PaymentVerificationSummaryFactory(payment_plan=payment_plan)


@pytest.fixture
def add_payment_verification() -> PV:
    yield payment_verification_creator()


@pytest.fixture
def add_payment_verification_xlsx() -> PV:
    yield payment_verification_creator(channel=PaymentVerificationPlan.VERIFICATION_CHANNEL_XLSX)


def payment_verification_creator(channel: str = PaymentVerificationPlan.VERIFICATION_CHANNEL_MANUAL) -> PV:
    generate_delivery_mechanisms()
    user = User.objects.first()
    business_area = BusinessArea.objects.first()
    registration_data_import = RegistrationDataImportFactory(imported_by=user, business_area=business_area)
    program = Program.objects.filter(name="Active Program").first()
    household, individuals = create_household(
        {
            "registration_data_import": registration_data_import,
            "admin_area": Area.objects.order_by("?").first(),
            "program": program,
        },
        {"registration_data_import": registration_data_import},
    )

    dm_cash = DeliveryMechanism.objects.get(code="cash")
    fsp = FinancialServiceProviderFactory()
    fsp.delivery_mechanisms.set([dm_cash])

    payment_plan = PaymentPlanFactory(
        name="TEST",
        status=PaymentPlan.Status.FINISHED,
        program_cycle=program.cycles.first(),
        business_area=business_area,
        start_date=datetime.now() - relativedelta(months=1),
        end_date=datetime.now() + relativedelta(months=1),
        created_by=user,
        financial_service_provider=fsp,
        delivery_mechanism=dm_cash,
    )

    payment_plan.unicef_id = "PP-0000-00-1122334"
    payment_plan.save()
    PaymentVerificationSummaryFactory(
        payment_plan=payment_plan,
    )

    payment = PaymentFactory(
        parent=payment_plan,
        household=household,
        head_of_household=household.head_of_household,
        entitlement_quantity=21.36,
        delivered_quantity=21.36,
        currency="PLN",
        status=Payment.STATUS_DISTRIBUTION_SUCCESS,
    )
    payment_verification_plan = PaymentVerificationPlanFactory(
        payment_plan=payment_plan,
        verification_channel=channel,
    )
    pv = PaymentVerificationFactory(
        payment=payment,
        payment_verification_plan=payment_verification_plan,
        status=PV.STATUS_PENDING,
    )

    return pv


@pytest.fixture
def clear_downloaded_files(download_path: str) -> None:
    for file in os.listdir(download_path):
        os.remove(os.path.join(download_path, file))
    yield
    for file in os.listdir(download_path):
        os.remove(os.path.join(download_path, file))


@pytest.mark.usefixtures("login")
class TestSmokePaymentVerification:
    def test_smoke_payment_verification(
        self, active_program: Program, add_payment_verification: PV, pagePaymentVerification: PaymentVerification
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        assert "Payment Verification" in pagePaymentVerification.getPageHeaderTitle().text
        assert "List of Payment Plans" in pagePaymentVerification.getTableTitle().text
        assert "Payment Plan ID" in pagePaymentVerification.getUnicefid().text
        assert "Payment Plan Status" in pagePaymentVerification.getVerificationstatus().text
        assert "Total Amount" in pagePaymentVerification.getTotaldeliveredquantity().text
        assert "Payment Disbursement Dates" in pagePaymentVerification.getStartdate().text
        assert "Last Modified Date" in pagePaymentVerification.getUpdatedat().text
        assert "PP-0000-00-1122334" in pagePaymentVerification.getCashPlanTableRow().text
        assert "PENDING" in pagePaymentVerification.getStatusContainer().text
        assert "Rows per page: 5 1–1 of 1" in pagePaymentVerification.getTablePagination().text.replace("\n", " ")

    def test_smoke_payment_verification_details(
        self,
        active_program: Program,
        add_payment_verification: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        assert "Payment Plan PP-0000-00-1122334" in pagePaymentVerificationDetails.getPageHeaderTitle().text
        assert "CREATE VERIFICATION PLAN" in pagePaymentVerificationDetails.getButtonNewPlan().text
        assert "Payment Plan Details" in pagePaymentVerificationDetails.getDivPaymentPlanDetails().text
        assert "Active Program" in pagePaymentVerificationDetails.getLabelProgrammeName().text
        pagePaymentVerificationDetails.getLabelPaymentRecords()
        pagePaymentVerificationDetails.getLabelStartDate()
        pagePaymentVerificationDetails.getLabelEndDate()
        pagePaymentVerificationDetails.getTableLabel()
        assert "0%" in pagePaymentVerificationDetails.getLabelSuccessful().text
        assert "0%" in pagePaymentVerificationDetails.getLabelErroneous().text
        assert "PENDING" in pagePaymentVerificationDetails.getLabelStatus().text
        assert "PENDING" in pagePaymentVerificationDetails.getVerificationPlansSummaryStatus().text
        assert (
            "COMPLETION DATE -"
            in pagePaymentVerificationDetails.getLabelizedFieldContainerSummaryCompletionDate().text.replace("\n", " ")
        )
        assert "-" in pagePaymentVerificationDetails.getLabelCompletionDate().text
        assert (
            "NUMBER OF VERIFICATION PLANS 1"
            in pagePaymentVerificationDetails.getLabelizedFieldContainerSummaryNumberOfPlans().text.replace("\n", " ")
        )
        assert "1" in pagePaymentVerificationDetails.getLabelNumberOfVerificationPlans().text
        assert "DELETE" in pagePaymentVerificationDetails.getButtonDeletePlan().text
        assert "EDIT" in pagePaymentVerificationDetails.getButtonEditPlan().text
        assert "ACTIVATE" in pagePaymentVerificationDetails.getButtonActivatePlan().text
        assert "PENDING" in pagePaymentVerificationDetails.getLabelStatus().text
        assert "PENDING" in pagePaymentVerificationDetails.getVerificationPlanStatus().text
        assert "MANUAL" in pagePaymentVerificationDetails.getLabelVerificationChannel().text

    def test_happy_path_payment_verification(
        self,
        active_program: Program,
        add_payment_verification: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        assert "1" in pagePaymentVerificationDetails.getLabelPaymentRecords().text
        assert (datetime.now() - relativedelta(months=1)).strftime(
            "%-d %b %Y"
        ) in pagePaymentVerificationDetails.getLabelStartDate().text
        assert (datetime.now() + relativedelta(months=1)).strftime(
            "%-d %b %Y"
        ) in pagePaymentVerificationDetails.getLabelEndDate().text
        assert "Reconciliation Summary" in pagePaymentVerificationDetails.getTableLabel().text
        payment_verification = add_payment_verification.payment_verification_plan
        assert (
            payment_verification.sampling.lower().replace("_", " ")
            in pagePaymentVerificationDetails.getLabelSampling().text.lower()
        )
        assert str(payment_verification.responded_count) in pagePaymentVerificationDetails.getLabelResponded().text
        assert (
            str(payment_verification.received_with_problems_count)
            in pagePaymentVerificationDetails.getLabelReceivedWithIssues().text
        )
        assert str(payment_verification.sample_size) in pagePaymentVerificationDetails.getLabelSampleSize().text
        assert str(payment_verification.received_count) in pagePaymentVerificationDetails.getLabelReceived().text
        assert str(payment_verification.not_received_count) in pagePaymentVerificationDetails.getLabelNotReceived().text
        pagePaymentVerificationDetails.getButtonDeletePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()
        try:
            pagePaymentVerificationDetails.getButtonNewPlan().click()
        except BaseException:
            sleep(3)
            pagePaymentVerificationDetails.getButtonNewPlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        pagePaymentVerificationDetails.getRows()[0].find_elements(By.TAG_NAME, "a")[0].click()
        payment_record = Payment.objects.first()
        assert "Payment" in pagePaymentRecord.getPageHeaderTitle().text
        assert "VERIFY" in pagePaymentRecord.getButtonEdPlan().text
        assert "DELIVERED FULLY" in pagePaymentRecord.getLabelStatus()[0].text
        assert "DELIVERED FULLY" in pagePaymentRecord.getStatusContainer().text
        assert payment_record.household.unicef_id in pagePaymentRecord.getLabelHousehold().text
        assert payment_record.parent.name in pagePaymentRecord.getLabelTargetPopulation().text
        assert payment_record.parent.unicef_id in pagePaymentRecord.getLabelDistributionModality().text
        assert payment_record.payment_verifications.first().status in pagePaymentRecord.getLabelStatus()[1].text
        assert "PLN 0.00" in pagePaymentRecord.getLabelAmountReceived().text
        assert payment_record.household.unicef_id in pagePaymentRecord.getLabelHouseholdId().text
        assert "21.36" in pagePaymentRecord.getLabelEntitlementQuantity().text
        assert "21.36" in pagePaymentRecord.getLabelDeliveredQuantity().text
        assert "PLN" in pagePaymentRecord.getLabelCurrency().text
        assert payment_record.delivery_type.name in pagePaymentRecord.getLabelDeliveryType().text
        assert payment_record.financial_service_provider.name in pagePaymentRecord.getLabelFsp().text

        pagePaymentRecord.getButtonEdPlan().click()

        pagePaymentRecord.getInputReceivedamount().click()
        pagePaymentRecord.getInputReceivedamount().send_keys("100")
        pagePaymentRecord.getButtonSubmit().click()

        for _ in range(5):
            if "RECEIVED WITH ISSUES" in pagePaymentRecord.getLabelStatus()[1].text:
                break
            sleep(1)
        assert "RECEIVED WITH ISSUES" in pagePaymentRecord.getLabelStatus()[1].text
        pagePaymentRecord.getArrowBack().click()

        pagePaymentVerificationDetails.getButtonFinish().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        assert "Payment Plan" in pagePaymentVerificationDetails.getPageHeaderTitle().text
        for _ in range(5):
            if "FINISHED" in pagePaymentVerificationDetails.getLabelStatus().text:
                break
            sleep(1)
        else:
            raise AssertionError(f"Error: FINISHED in {pagePaymentVerificationDetails.getLabelStatus().text}")
        assert "FINISHED" in pagePaymentVerificationDetails.getVerificationPlanStatus().text
        assert "100%" in pagePaymentVerificationDetails.getLabelSuccessful().text

        pagePaymentRecord.getArrowBack().click()

        assert "FINISHED" in pagePaymentVerification.getCashPlanTableRow().text


@pytest.mark.usefixtures("login")
class TestPaymentVerification:
    @pytest.mark.parametrize(
        "channel",
        [
            "manual",
            "rapidpro",
            "xlsx",
        ],
    )
    def test_payment_verification_create_verification_plan_full_list(
        self,
        channel: str,
        active_program: Program,
        empty_payment_verification: None,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerification.getButtonNewPlan().click()
        pagePaymentVerification.getRadioVerificationChannel(channel).click()
        pagePaymentVerification.getButtonSubmit().click()
        assert "PENDING" in pagePaymentVerificationDetails.getVerificationPlansSummaryStatus().text
        assert "PENDING" in pagePaymentVerificationDetails.getVerificationPlanStatus().text
        assert channel.upper() in pagePaymentVerificationDetails.getLabelVerificationChannel().text
        assert "Full list" in pagePaymentVerificationDetails.getLabelSampling().text

    def test_payment_verification_create_verification_plan_random_sampling_manual(
        self,
        active_program: Program,
        empty_payment_verification: None,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerification.getButtonNewPlan().click()
        pagePaymentVerification.getTabRandomSampling().click()
        pagePaymentVerification.getInputAdmincheckbox().click()
        pagePaymentVerification.getInputAgecheckbox().click()
        pagePaymentVerification.getInputSexcheckbox().click()
        pagePaymentVerification.getButtonSubmit().click()
        assert "PENDING" in pagePaymentVerificationDetails.getVerificationPlansSummaryStatus().text
        assert "PENDING" in pagePaymentVerificationDetails.getVerificationPlanStatus().text
        assert "MANUAL" in pagePaymentVerificationDetails.getLabelVerificationChannel().text
        assert "Random sampling" in pagePaymentVerificationDetails.getLabelSampling().text
        assert "0" in pagePaymentVerificationDetails.getLabelSampleSize().text
        assert "1" in pagePaymentVerificationDetails.getLabelNumberOfVerificationPlans().text

    @pytest.mark.xfail(reason="Problem with deadlock during test - 202318")
    def test_payment_verification_records(
        self,
        active_program: Program,
        payment_verification_3: None,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        assert "3" in pagePaymentVerificationDetails.getLabelPaymentRecords().text
        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

    def test_payment_verification_delete(
        self,
        active_program: Program,
        payment_verification_3: None,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonDeletePlan()
        before_list_of_verification_plans = [i.text for i in pagePaymentVerificationDetails.getVerificationPlanPrefix()]
        pagePaymentVerificationDetails.deleteVerificationPlanByNumber(1)
        pagePaymentVerificationDetails.getButtonSubmit().click()
        pagePaymentVerificationDetails.checkAlert("Verification plan has been deleted.")
        for _ in range(50):
            if 2 == len(pagePaymentVerificationDetails.getVerificationPlanPrefix()):
                break
            sleep(0.1)
        else:
            raise AssertionError("Verification Plan was not deleted")
        assert before_list_of_verification_plans[1] not in pagePaymentVerificationDetails.getVerificationPlanPrefix()

    def test_payment_verification_edit(
        self,
        active_program: Program,
        payment_verification_3: None,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonEditPlan().click()
        pagePaymentVerification.getTabFullList().click()
        if "MANUAL" in pagePaymentVerificationDetails.getLabelVerificationChannel().text:
            pagePaymentVerification.getRadioVerificationChannel("xlsx").click()
            channel = "XLSX"
        else:
            pagePaymentVerification.getRadioVerificationChannel("manual").click()
            channel = "MANUAL"

        pagePaymentVerification.getButtonSubmit().click()
        assert "PENDING" in pagePaymentVerificationDetails.getVerificationPlansSummaryStatus().text
        assert "PENDING" in pagePaymentVerificationDetails.getVerificationPlanStatus().text
        for _ in range(50):
            if channel in pagePaymentVerificationDetails.getLabelVerificationChannel().text:
                break
        else:
            assert channel in pagePaymentVerificationDetails.getLabelVerificationChannel().text
        assert "Full list" in pagePaymentVerificationDetails.getLabelSampling().text

    def test_payment_verification_successful_received(
        self,
        active_program: Program,
        add_payment_verification: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        assert 1 == len(pagePaymentVerificationDetails.getRows())
        pagePaymentVerificationDetails.scroll(execute=2)
        pagePaymentVerificationDetails.getRows()[0].find_element(By.TAG_NAME, "a").click()
        quantity = pagePaymentRecord.getLabelDeliveredQuantity().text
        pagePaymentRecord.getButtonEdPlan().click()
        pagePaymentRecord.getInputReceivedamount().send_keys(quantity)
        pagePaymentRecord.getButtonSubmit().click()
        pagePaymentRecord.getArrowBack().click()

        assert pagePaymentRecord.waitForStatusContainer("RECEIVED")
        assert "RECEIVED" == pagePaymentRecord.getStatusContainer().text

    def test_payment_verification_successful_not_received(
        self,
        active_program: Program,
        add_payment_verification: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")
        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        assert 1 == len(pagePaymentVerificationDetails.getRows())
        pagePaymentVerificationDetails.scroll(execute=2)
        pagePaymentVerificationDetails.getRows()[0].find_element(By.TAG_NAME, "a").click()
        pagePaymentRecord.getButtonEdPlan().click()
        pagePaymentRecord.getChoiceNotReceived().click()
        pagePaymentRecord.getButtonSubmit().click()
        pagePaymentRecord.getArrowBack().click()

        assert pagePaymentRecord.waitForStatusContainer("NOT RECEIVED")

    def test_payment_verification_partially_successful_received_and_grievance_ticket(
        self,
        active_program: Program,
        add_payment_verification: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
        pageGrievanceTickets: GrievanceTickets,
        pageGrievanceDetailsPage: GrievanceDetailsPage,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")

        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        assert 1 == len(pagePaymentVerificationDetails.getRows())
        pagePaymentVerificationDetails.scroll(execute=2)
        pagePaymentVerificationDetails.getRows()[0].find_element(By.TAG_NAME, "a").click()
        quantity = float(pagePaymentRecord.getLabelDeliveredQuantity().text) - 1
        pagePaymentRecord.getButtonEdPlan().click()
        pagePaymentRecord.getInputReceivedamount().send_keys(str(quantity))
        pagePaymentRecord.getButtonSubmit().click()
        pagePaymentRecord.getArrowBack().click()

        assert pagePaymentRecord.waitForStatusContainer("RECEIVED WITH ISSUES")

        pagePaymentVerificationDetails.getButtonFinish().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        pageGrievanceTickets.getNavGrievance().click()
        pageGrievanceTickets.getTabSystemGenerated().click()
        assert 1 == len(pageGrievanceTickets.getTicketListRow())
        pageGrievanceTickets.getTicketListRow()[0].click()
        pageGrievanceDetailsPage.getButtonAssignToMe().click()
        pageGrievanceDetailsPage.getButtonSetInProgress().click()

        pageGrievanceDetailsPage.getGrievanceVerify().click()
        pageGrievanceDetailsPage.getInputNewReceivedAmount().send_keys(str(quantity + 1))
        pageGrievanceDetailsPage.getButtonSubmit().click()

        pageGrievanceDetailsPage.getButtonSendForApproval().click()
        pageGrievanceDetailsPage.getGrievanceApprove().click()
        pageGrievanceDetailsPage.getButtonConfirm().click()

        pageGrievanceDetailsPage.getButtonCloseTicket().click()
        pageGrievanceDetailsPage.getButtonConfirm().click()

        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()

        assert pagePaymentRecord.waitForStatusContainer("RECEIVED")
        assert "RECEIVED" == pagePaymentRecord.getStatusContainer().text

        pageGrievanceTickets.scroll(execute=2)
        pageGrievanceTickets.screenshot("0", file_path="./")

    def test_payment_verification_by_payment_related_complaint(
        self,
        active_program: Program,
        add_payment_verification: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")

    @pytest.mark.xfail(reason="UNSTABLE")
    def test_payment_verification_xlsx_successful(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        download_path: str,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")

        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        pagePaymentVerificationDetails.getExportXlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        pagePaymentVerificationDetails.driver.refresh()

        pagePaymentVerificationDetails.getDownloadXlsx().click()

        xlsx_file = find_file(".xlsx", number_of_ties=10, search_in_dir=download_path)
        wb1 = openpyxl.load_workbook(os.path.join(download_path, xlsx_file))
        ws1 = wb1.active
        for cell in ws1["N:N"]:
            if cell.row >= 2:
                ws1.cell(row=cell.row, column=3, value="YES")
                ws1.cell(row=cell.row, column=16, value=ws1.cell(row=cell.row, column=15).value)

        wb1.save(os.path.join(download_path, "new_" + xlsx_file))
        find_file("new_" + xlsx_file, number_of_ties=10, search_in_dir=download_path)
        pagePaymentVerificationDetails.getImportXlsx().click()

        pagePaymentVerificationDetails.upload_file(
            os.path.abspath(os.path.join(download_path, "new_" + xlsx_file)), timeout=120
        )
        pagePaymentVerificationDetails.getButtonImportEntitlement().click()

        assert pagePaymentRecord.waitForStatusContainer("RECEIVED", timeout=60)
        assert "RECEIVED" == pagePaymentRecord.getStatusContainer().text

    @pytest.mark.xfail(reason="UNSTABLE")
    def test_payment_verification_xlsx_partially_successful(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
        download_path: str,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")

        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        assert 1 == len(pagePaymentVerificationDetails.getRows())
        pagePaymentVerificationDetails.scroll(execute=2)
        pagePaymentVerificationDetails.getRows()[0].find_element(By.TAG_NAME, "a").click()
        quantity = pagePaymentRecord.getLabelDeliveredQuantity().text
        pagePaymentRecord.getArrowBack().click()

        pagePaymentVerificationDetails.getExportXlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        pagePaymentVerificationDetails.driver.refresh()

        pagePaymentVerificationDetails.getDownloadXlsx().click()

        xlsx_file = find_file(".xlsx", number_of_ties=10, search_in_dir=download_path)
        wb1 = openpyxl.load_workbook(os.path.join(download_path, xlsx_file))
        ws1 = wb1.active
        for cell in ws1["N:N"]:
            if cell.row >= 2:
                ws1.cell(row=cell.row, column=3, value="YES")
                ws1.cell(row=cell.row, column=16, value=float(quantity) - 1.0)

        wb1.save(os.path.join(download_path, xlsx_file))
        pagePaymentVerificationDetails.getImportXlsx().click()

        pagePaymentVerificationDetails.upload_file(os.path.abspath(os.path.join(download_path, xlsx_file)), timeout=120)

        pagePaymentVerificationDetails.getButtonImportEntitlement().click()

        assert pagePaymentRecord.waitForStatusContainer("RECEIVED WITH ISSUES")

    def test_payment_verification_xlsx_not_received(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
        download_path: str,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")

        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        pagePaymentVerificationDetails.getExportXlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        pagePaymentVerificationDetails.driver.refresh()

        pagePaymentVerificationDetails.getDownloadXlsx().click()

        xlsx_file = find_file(".xlsx", number_of_ties=10, search_in_dir=download_path)
        wb1 = openpyxl.load_workbook(os.path.join(download_path, xlsx_file))
        ws1 = wb1.active
        for cell in ws1["N:N"]:
            if cell.row >= 2:
                ws1.cell(row=cell.row, column=3, value="NO")
                ws1.cell(row=cell.row, column=16, value=0)

        wb1.save(os.path.join(download_path, xlsx_file))
        pagePaymentVerificationDetails.getImportXlsx().click()

        pagePaymentVerificationDetails.upload_file(os.path.abspath(os.path.join(download_path, xlsx_file)), timeout=120)

        pagePaymentVerificationDetails.getButtonImportEntitlement().click()

        assert pagePaymentRecord.waitForStatusContainer("NOT RECEIVED")

    def test_payment_verification_discard(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")

        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()
        pagePaymentVerificationDetails.getButtonDiscard().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        pagePaymentVerificationDetails.driver.refresh()

        pagePaymentVerificationDetails.getExportXlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        pagePaymentVerificationDetails.driver.refresh()

        pagePaymentVerificationDetails.getDownloadXlsx().click()

        pagePaymentVerificationDetails.getButtonDiscard().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        pagePaymentVerificationDetails.checkAlert("You cant discard if xlsx file was downloaded or imported")

    def test_payment_verification_xlsx_invalid(
        self,
        clear_downloaded_files: None,
        active_program: Program,
        add_payment_verification_xlsx: PV,
        pagePaymentVerification: PaymentVerification,
        pagePaymentVerificationDetails: PaymentVerificationDetails,
        pagePaymentRecord: PaymentRecord,
    ) -> None:
        pagePaymentVerification.selectGlobalProgramFilter("Active Program")

        pagePaymentVerification.getNavPaymentVerification().click()
        pagePaymentVerification.getCashPlanTableRow().click()
        pagePaymentVerificationDetails.getButtonActivatePlan().click()
        pagePaymentVerificationDetails.getButtonSubmit().click()

        pagePaymentVerificationDetails.getExportXlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        pagePaymentVerificationDetails.driver.refresh()

        pagePaymentVerificationDetails.getDownloadXlsx().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        pagePaymentVerificationDetails.driver.refresh()

        pagePaymentVerificationDetails.getButtonMarkAsInvalid().click()

        # ToDo: Workaround: Bug 220111
        sleep(2)
        pagePaymentVerificationDetails.driver.refresh()

        assert "INVALID" in pagePaymentVerificationDetails.getVerificationPlanStatus().text
