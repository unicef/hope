from io import BytesIO

from django.core.files.base import ContentFile

import openpyxl
import pytest

from hct_mis_api.apps.core.models import FlexibleAttribute
from hct_mis_api.apps.geo.models import Area, AreaType, Country
from hct_mis_api.apps.household.models import (
    FEMALE,
    MALE,
    Document,
    DocumentType,
    Individual,
)
from hct_mis_api.apps.payment.models import Account, AccountType, FinancialInstitution
from hct_mis_api.apps.program.models import Program
from hct_mis_api.apps.universal_update_script.models import UniversalUpdate
from hct_mis_api.apps.universal_update_script.universal_individual_update_service.universal_individual_update_service import (
    UniversalIndividualUpdateService,
)
from tests.extras.test_utils.factories.core import create_afghanistan
from tests.extras.test_utils.factories.household import create_household_and_individuals
from tests.extras.test_utils.factories.program import ProgramFactory

pytestmark = pytest.mark.django_db(transaction=True)


@pytest.fixture()
def poland() -> Country:
    return Country.objects.create(name="Poland", iso_code2="PL", iso_code3="POL", iso_num="616")


@pytest.fixture()
def germany() -> Country:
    return Country.objects.create(name="Germany", iso_code2="DE", iso_code3="DEU", iso_num="276")


@pytest.fixture()
def state(poland: Country) -> AreaType:
    return AreaType.objects.create(name="State", country=poland)


@pytest.fixture()
def district(poland: Country, state: AreaType) -> AreaType:
    return AreaType.objects.create(name="District", parent=state, country=poland)


@pytest.fixture()
def admin1(state: AreaType) -> Area:
    return Area.objects.create(name="Kabul", area_type=state, p_code="AF11")


@pytest.fixture()
def admin2(district: AreaType) -> Area:
    return Area.objects.create(name="Kabul1", area_type=district, p_code="AF1115")


@pytest.fixture()
def program(poland: Country, germany: Country) -> Program:
    business_area = create_afghanistan()
    business_area.countries.add(poland, germany)

    program = ProgramFactory(name="Test Program for Household", status=Program.ACTIVE, business_area=business_area)
    return program


@pytest.fixture
def account_type() -> AccountType:
    ac_mobile, _ = AccountType.objects.update_or_create(key="mobile", label="Mobile")
    return ac_mobile


@pytest.fixture
def flexible_attribute_individual() -> FlexibleAttribute:
    return FlexibleAttribute.objects.create(
        name="muac",
        type=FlexibleAttribute.INTEGER,
        associated_with=FlexibleAttribute.ASSOCIATED_WITH_INDIVIDUAL,
        label={"English(EN)": "Muac"},
    )


@pytest.fixture
def flexible_attribute_household() -> FlexibleAttribute:
    return FlexibleAttribute.objects.create(
        name="eggs",
        type=FlexibleAttribute.STRING,
        associated_with=FlexibleAttribute.ASSOCIATED_WITH_HOUSEHOLD,
        label={"English(EN)": "Eggs"},
    )


@pytest.fixture
def individual(
    program: Program,
    admin1: Area,
    admin2: Area,
    flexible_attribute_individual: FlexibleAttribute,
    flexible_attribute_household: FlexibleAttribute,
    account_type: AccountType,
) -> Individual:
    household, individuals = create_household_and_individuals(
        household_data={
            "unicef_id": "HH-20-0000.0002",
            "rdi_merge_status": "MERGED",
            "business_area": program.business_area,
            "program": program,
            "admin1": admin1,
            "size": 954,
            "returnee": True,
        },
        individuals_data=[
            {
                "unicef_id": "IND-00-0000.0011",
                "rdi_merge_status": "MERGED",
                "business_area": program.business_area,
                "sex": MALE,
                "phone_no": "+48555444333",
            },
        ],
    )

    ind = individuals[0]

    ind.flex_fields = {"muac": 0}
    ind.save()
    household.flex_fields = {"eggs": "OLD"}
    household.save()
    return ind


@pytest.fixture()
def wallet(individual: Individual, account_type: AccountType) -> Account:
    financial_institution = FinancialInstitution.objects.create(
        name="Test Financial Institution", type=FinancialInstitution.FinancialInstitutionType.TELCO
    )

    return Account.objects.create(
        account_type=account_type,
        financial_institution=financial_institution,
        individual=individual,
        number="1234567890",
        data={"number": "1234567890"},
        rdi_merge_status=Account.MERGED,
    )


@pytest.fixture
def document_national_id(individual: Individual, program: Program, poland: Country) -> Document:
    document_type = DocumentType.objects.create(key="national_id", label="National ID")
    return Document.objects.create(
        individual=individual,
        program=program,
        type=document_type,
        document_number="Test 123",
        rdi_merge_status=Document.MERGED,
        country=poland,
    )


@pytest.mark.elasticsearch
class TestUniversalIndividualUpdateService:
    def test_update_individual(
        self,
        individual: Individual,
        program: Program,
        admin1: Area,
        admin2: Area,
        document_national_id: Document,
        account_type: AccountType,
        wallet: Account,
    ) -> None:
        """
        This test generates file for individual update
        Then Changes manually individual household document data
        Then runs the update script based on the file generated and checks if data is updated to original

        In this way I'm checking if file is generated correctly and also if service can use the same file to update back
        :param individual:
        :param program:
        :return:
        """
        # create one more DeliveryMechanismConfig with empty account_type
        # DeliveryMechanismConfig.objects.get_or_create(
        #     fsp=FinancialServiceProviderFactory(),
        #     delivery_mechanism=DeliveryMechanism.objects.create(name="Test", code="test", account_type=None),
        #     required_fields=["phone_number"],
        # )
        # save old values
        given_name_old = individual.given_name
        sex_old = individual.sex
        birth_date_old = individual.birth_date
        phone_no_old = individual.phone_no
        address_old = individual.household.address
        admin1_old = individual.household.admin1
        size_old = individual.household.size
        returnee_old = individual.household.returnee
        muac_old = individual.flex_fields.get("muac")
        eggs_old = individual.household.flex_fields.get("eggs")
        wallet_number_old = wallet.data.get("number")
        document_number_old = document_national_id.document_number
        universal_update = UniversalUpdate(program=program)
        universal_update.unicef_ids = individual.unicef_id
        universal_update.individual_fields = ["given_name", "sex", "birth_date", "phone_no"]
        universal_update.individual_flex_fields_fields = ["muac"]
        universal_update.household_flex_fields_fields = ["eggs"]
        universal_update.household_fields = ["address", "admin1", "size", "returnee"]
        universal_update.save()
        universal_update.document_types.add(DocumentType.objects.first())
        universal_update.account_types.add(AccountType.objects.first())
        AccountType.objects.create(label="Cash", key="cash")
        service = UniversalIndividualUpdateService(universal_update)
        template_file = service.generate_xlsx_template()

        universal_update.refresh_from_db()
        content = template_file.getvalue()
        universal_update.update_file.save("template.xlsx", ContentFile(content))
        universal_update.save()
        universal_update.refresh_from_db()
        expected_generate_log = "Generating row 0 to 1\nGenerating Finished\n"
        assert universal_update.saved_logs == expected_generate_log
        # Change the data manually
        individual.given_name = "Test Name"
        individual.sex = FEMALE
        individual.birth_date = "1996-06-21"
        individual.phone_no = "+48555111111"
        individual.flex_fields = {"muac": 25}
        individual.save()
        household = individual.household
        household.address = "Wrocław"
        household.admin1 = None
        household.size = 100
        household.returnee = False
        household.flex_fields = {"eggs": "NEW"}
        household.save()
        document_national_id.document_number = "111"
        document_national_id.save()
        wallet.data["number"] = "0"
        wallet.number = "0"
        wallet.save()
        service = UniversalIndividualUpdateService(universal_update)
        universal_update.clear_logs()
        service.execute()
        universal_update.refresh_from_db()
        individual.refresh_from_db()
        document_national_id.refresh_from_db()
        wallet.refresh_from_db()
        expected_update_log = """Validating row 0 to 1 Indivduals
Validation successful
Updating row 0 to 1 Individuals
Deduplicating individuals Elasticsearch
Deduplicating documents
Update successful
"""
        assert universal_update.saved_logs == expected_update_log
        assert universal_update.saved_logs == universal_update.logs
        assert individual.given_name == given_name_old
        assert individual.sex == sex_old
        assert individual.birth_date == birth_date_old
        assert individual.phone_no == phone_no_old
        assert individual.household.address == address_old
        assert individual.household.admin1 == admin1_old
        assert document_national_id.document_number == document_number_old
        assert individual.household.size == size_old
        assert individual.household.returnee == returnee_old
        assert individual.flex_fields.get("muac") == muac_old
        assert individual.household.flex_fields.get("eggs") == eggs_old
        assert wallet.data.get("number") == wallet_number_old
        assert wallet.number == wallet_number_old

    def test_update_individual_empty_row(
        self,
        individual: Individual,
        program: Program,
        admin1: Area,
        admin2: Area,
        document_national_id: Document,
        account_type: AccountType,
        wallet: Account,
    ) -> None:
        # save old values
        given_name_old = individual.given_name
        sex_old = individual.sex
        birth_date_old = individual.birth_date
        phone_no_old = individual.phone_no
        address_old = individual.household.address
        admin1_old = individual.household.admin1
        size_old = individual.household.size
        returnee_old = individual.household.returnee
        muac_old = individual.flex_fields.get("muac")
        eggs_old = individual.household.flex_fields.get("eggs")
        wallet_number_old = wallet.data.get("phone_number")
        document_number_old = document_national_id.document_number
        universal_update = UniversalUpdate(program=program)
        universal_update.unicef_ids = individual.unicef_id
        universal_update.individual_fields = ["given_name", "sex", "birth_date", "phone_no"]
        universal_update.individual_flex_fields_fields = ["muac"]
        universal_update.household_flex_fields_fields = ["eggs"]
        universal_update.household_fields = ["address", "admin1", "size", "returnee"]
        universal_update.save()
        universal_update.document_types.add(DocumentType.objects.first())
        universal_update.account_types.add(AccountType.objects.first())
        service = UniversalIndividualUpdateService(universal_update)
        template_file = service.generate_xlsx_template()
        universal_update.refresh_from_db()
        content = template_file.getvalue()
        universal_update.update_file.save("template.xlsx", ContentFile(content))
        universal_update.save()
        universal_update.refresh_from_db()
        expected_generate_log = "Generating row 0 to 1\nGenerating Finished\n"
        assert universal_update.saved_logs == expected_generate_log
        # empty whole row xlsx template left only unicef id
        wb = openpyxl.load_workbook(universal_update.update_file.path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                if cell.column == 1:
                    continue
                cell.value = None
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.getvalue()
        universal_update.update_file.save("testing.xlsx", ContentFile(content))
        service = UniversalIndividualUpdateService(universal_update)
        universal_update.clear_logs()
        service.execute()
        universal_update.refresh_from_db()
        individual.refresh_from_db()
        document_national_id.refresh_from_db()
        wallet.refresh_from_db()
        expected_update_log = """Validating row 0 to 1 Indivduals
Validation successful
Updating row 0 to 1 Individuals
Deduplicating individuals Elasticsearch
Deduplicating documents
Update successful
"""
        assert universal_update.saved_logs == expected_update_log
        assert universal_update.saved_logs == universal_update.logs
        assert individual.given_name == given_name_old
        assert individual.sex == sex_old
        assert individual.birth_date == birth_date_old
        assert individual.phone_no == phone_no_old
        assert individual.household.address == address_old
        assert individual.household.admin1 == admin1_old
        assert document_national_id.document_number == document_number_old
        assert individual.household.size == size_old
        assert individual.household.returnee == returnee_old
        assert individual.flex_fields.get("muac") == muac_old
        assert individual.household.flex_fields.get("eggs") == eggs_old
        assert wallet.data.get("phone_number") == wallet_number_old

    def test_update_individual_invalid(
        self,
        individual: Individual,
        program: Program,
        admin1: Area,
        admin2: Area,
        document_national_id: Document,
        account_type: AccountType,
        wallet: Account,
    ) -> None:
        # save old values
        given_name_old = individual.given_name
        sex_old = individual.sex
        birth_date_old = individual.birth_date
        phone_no_old = individual.phone_no
        address_old = individual.household.address
        admin1_old = individual.household.admin1
        size_old = individual.household.size
        returnee_old = individual.household.returnee
        muac_old = individual.flex_fields.get("muac")
        eggs_old = individual.household.flex_fields.get("eggs")
        wallet_number_old = wallet.data.get("phone_number")
        document_number_old = document_national_id.document_number
        universal_update = UniversalUpdate(program=program)
        universal_update.unicef_ids = individual.unicef_id
        universal_update.individual_fields = ["given_name", "sex", "birth_date", "phone_no"]
        universal_update.individual_flex_fields_fields = ["muac"]
        universal_update.household_flex_fields_fields = ["eggs"]
        universal_update.household_fields = ["address", "admin1", "size", "returnee"]
        universal_update.save()
        universal_update.document_types.add(DocumentType.objects.first())
        universal_update.account_types.add(AccountType.objects.first())
        service = UniversalIndividualUpdateService(universal_update)
        template_file = service.generate_xlsx_template()
        universal_update.refresh_from_db()
        content = template_file.getvalue()
        universal_update.update_file.save("template.xlsx", ContentFile(content))
        universal_update.save()
        universal_update.refresh_from_db()
        expected_generate_log = "Generating row 0 to 1\nGenerating Finished\n"
        assert universal_update.saved_logs == expected_generate_log
        # put string in every collumn
        wb = openpyxl.load_workbook(universal_update.update_file.path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                if cell.column == 1:
                    continue
                cell.value = "TEST String"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.getvalue()
        universal_update.update_file.save("testing.xlsx", ContentFile(content))
        service = UniversalIndividualUpdateService(universal_update)
        universal_update.clear_logs()
        service.execute()
        universal_update.refresh_from_db()
        individual.refresh_from_db()
        document_national_id.refresh_from_db()
        wallet.refresh_from_db()
        expected_update_log = """Validating row 0 to 1 Indivduals
Validation failed
Row: 2 - Administrative area admin1 with p_code TEST String not found
Row: 2 - TEST String for column size is not a valid integer
Row: 2 - TEST String for column returnee is not a valid boolean allowed values are TRUE or FALSE
Row: 2 - Invalid value TEST String for column sex allowed values are ['MALE', 'FEMALE', 'OTHER', 'NOT_COLLECTED', 'NOT_ANSWERED']
Row: 2 - TEST String for column birth_date is not a valid date
Row: 2 - TEST String for column phone_no is not a valid phone number
Row: 2 - Country not found for field national_id_country_i_c and value TEST String
Row: 2 - Financial institution ID must be a number for field account__mobile__financial_institution_pk
"""
        assert universal_update.saved_logs == expected_update_log
        assert universal_update.saved_logs == universal_update.logs
        assert individual.given_name == given_name_old
        assert individual.sex == sex_old
        assert individual.birth_date == birth_date_old
        assert individual.phone_no == phone_no_old
        assert individual.household.address == address_old
        assert individual.household.admin1 == admin1_old
        assert document_national_id.document_number == document_number_old
        assert individual.household.size == size_old
        assert individual.household.returnee == returnee_old
        assert individual.flex_fields.get("muac") == muac_old
        assert individual.household.flex_fields.get("eggs") == eggs_old
        assert wallet.data.get("phone_number") == wallet_number_old

    def test_update_individual_empty_fields(
        self,
        individual: Individual,
        program: Program,
        admin1: Area,
        admin2: Area,
        document_national_id: Document,
        account_type: AccountType,
        wallet: Account,
    ) -> None:
        universal_update = UniversalUpdate(program=program)
        universal_update.unicef_ids = individual.unicef_id
        universal_update.save()
        service = UniversalIndividualUpdateService(universal_update)
        template_file = service.generate_xlsx_template()
        universal_update.refresh_from_db()
        content = template_file.getvalue()
        universal_update.update_file.save("template.xlsx", ContentFile(content))
        universal_update.save()
        universal_update.refresh_from_db()
        expected_generate_log = "Generating row 0 to 1\nGenerating Finished\n"
        assert universal_update.saved_logs == expected_generate_log
        # empty whole row xlsx template left only unicef id
        wb = openpyxl.load_workbook(universal_update.update_file.path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                if cell.column == 1:
                    continue
                cell.value = None
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.getvalue()
        universal_update.update_file.save("testing.xlsx", ContentFile(content))
        service = UniversalIndividualUpdateService(universal_update)
        universal_update.clear_logs()
        service.execute()
        universal_update.refresh_from_db()
        individual.refresh_from_db()
        document_national_id.refresh_from_db()
        wallet.refresh_from_db()
        expected_update_log = """Validating row 0 to 1 Indivduals
Validation successful
Updating row 0 to 1 Individuals
Deduplicating individuals Elasticsearch
Deduplicating documents
Update successful
"""
        assert universal_update.saved_logs == expected_update_log

    def test_accounts_validation(
        self,
        individual: Individual,
        program: Program,
        admin1: Area,
        admin2: Area,
        document_national_id: Document,
        account_type: AccountType,
        wallet: Account,
    ) -> None:
        universal_update = UniversalUpdate(program=program)
        universal_update.save()
        universal_update.document_types.add(DocumentType.objects.first())
        universal_update.account_types.add(AccountType.objects.first())
        service = UniversalIndividualUpdateService(universal_update)
        headers = ["unicef_id", "account__mobile__financial_institution_pk", "account__mobile__number"]
        row = (
            individual.unicef_id,
            wallet.financial_institution.id,
            wallet.number,
        )
        errors = service.validate_accounts(row, headers, individual, 1)
        assert errors == []
        row = (
            individual.unicef_id,
            None,  # Missing financial institution
            wallet.number,
        )
        errors = service.validate_accounts(row, headers, individual, 1)
        assert errors == [
            "Row: 1 - Financial institution ID must be provided for account type mobile if any other field is updated"
        ]
        row = (
            individual.unicef_id,
            None,  # Missing financial institution
            None,  # Missing account number
        )
        errors = service.validate_accounts(row, headers, individual, 1)
        assert errors == []
