from django.test import TestCase

import openpyxl
from freezegun import freeze_time

from hct_mis_api.apps.core.fixtures import create_afghanistan
from hct_mis_api.apps.core.utils import encode_id_base64
from hct_mis_api.apps.geo.fixtures import AreaFactory, AreaTypeFactory
from hct_mis_api.apps.grievance.fixtures import GrievanceTicketFactory
from hct_mis_api.apps.grievance.models import (
    GrievanceTicket,
    TicketComplaintDetails,
    TicketDeleteIndividualDetails,
    TicketIndividualDataUpdateDetails,
    TicketNeedsAdjudicationDetails,
    TicketNegativeFeedbackDetails,
    TicketPositiveFeedbackDetails,
    TicketReferralDetails,
    TicketSensitiveDetails,
    TicketSystemFlaggingDetails,
)
from hct_mis_api.apps.household.fixtures import create_household_and_individuals
from hct_mis_api.apps.household.models import FEMALE, MALE
from hct_mis_api.apps.payment.fixtures import PaymentFactory
from hct_mis_api.apps.payment.models import Payment
from hct_mis_api.apps.periodic_data_update.fixtures import (
    PeriodicDataUpdateTemplateFactory,
)
from hct_mis_api.apps.periodic_data_update.service.periodic_data_update_export_template_service import (
    PeriodicDataUpdateExportTemplateService,
)
from hct_mis_api.apps.program.fixtures import ProgramFactory
from hct_mis_api.apps.program.models import Program
from hct_mis_api.apps.registration_data.fixtures import RegistrationDataImportFactory
from hct_mis_api.apps.sanction_list.fixtures import SanctionListIndividualFactory
from hct_mis_api.apps.targeting.fixtures import TargetPopulationFactory


class TestPeriodicDataUpdateExportTemplateService(TestCase):
    rdi = None
    business_area = None
    program = None

    @classmethod
    def setUpTestData(cls) -> None:
        super().setUpTestData()
        cls.business_area = create_afghanistan()
        cls.program = ProgramFactory(name="Test Program", status=Program.DRAFT, business_area=cls.business_area)
        cls.rdi = RegistrationDataImportFactory()
        cls.household, cls.individuals = create_household_and_individuals(
            household_data={
                "business_area": cls.business_area,
                "program_id": cls.program.pk,
                "registration_data_import": cls.rdi,
            },
            individuals_data=[
                {
                    "business_area": cls.business_area,
                    "program_id": cls.program.pk,
                    "registration_data_import": cls.rdi,
                },
                {
                    "business_area": cls.business_area,
                    "program_id": cls.program.pk,
                    "registration_data_import": cls.rdi,
                },
            ],
        )
        cls.periodic_data_update_template = PeriodicDataUpdateTemplateFactory(
            program=cls.program,
            business_area=cls.business_area,
            rounds_data=[
                {
                    "field": "muac",
                    "round": 2,
                    "round_name": "May",
                    "number_of_records": 100,
                },
                {
                    "field": "month_worked",
                    "round": 4,
                    "round_name": "April",
                    "number_of_records": 58,
                },
            ],
        )

    def test_generate_workbook(
        self,
    ) -> None:
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        wb = service.generate_workbook()
        self.assertIsNotNone(wb)
        self.assertEqual(wb.sheetnames, [service.PDU_SHEET, service.META_SHEET])
        meta_sheet = wb[service.META_SHEET]
        pdu_sheet = wb[service.PDU_SHEET]
        self.assertEqual(meta_sheet["b1"].value, self.periodic_data_update_template.pk)
        self.assertEqual(wb.custom_doc_props["pdu_template_id"].value, str(self.periodic_data_update_template.pk))
        self.assertEqual(pdu_sheet.max_row, 3)

    def test_save_xlsx_file(self) -> None:
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        service.generate_workbook()
        service.save_xlsx_file()
        self.periodic_data_update_template.refresh_from_db()
        self.assertIsNotNone(self.periodic_data_update_template.file)
        wb = openpyxl.load_workbook(self.periodic_data_update_template.file.file.path)
        self.assertEqual(wb.sheetnames, [service.PDU_SHEET, service.META_SHEET])

    def test_generate_header(self) -> None:
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        test_header = [
            "individual__uuid",
            "individual_unicef_id",
            "first_name",
            "last_name",
            "muac__round_number",
            "muac__round_name",
            "muac__round_value",
            "muac__collection_date",
            "month_worked__round_number",
            "month_worked__round_name",
            "month_worked__round_value",
            "month_worked__collection_date",
        ]
        self.assertEqual(service._generate_header(), test_header)

    def test_generate_row_empty_flex_fields_individual(self) -> None:
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        individual = self.individuals[0]
        row = service._generate_row(individual)
        expected_row = [
            str(individual.pk),
            individual.unicef_id,
            individual.given_name,
            individual.family_name,
            2,
            "May",
            "",
            "",
            4,
            "April",
            "",
            "",
        ]
        self.assertEqual(row, expected_row)

    def test_generate_row_half_filled_flex_fields_individual(self) -> None:
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        individual = self.individuals[0]
        individual.flex_fields = {
            "muac": {
                "2": {
                    "value": 20,
                    "collection_date": "2021-05-01",
                }
            }
        }
        individual.save()
        row = service._generate_row(individual)
        expected_row = [
            str(individual.pk),
            individual.unicef_id,
            individual.given_name,
            individual.family_name,
            2,
            "May",
            "-",
            "-",
            4,
            "April",
            "",
            "",
        ]
        self.assertEqual(row, expected_row)

    def test_generate_row_fully_filled_flex_fields_individual(self) -> None:
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        individual = self.individuals[0]
        individual.flex_fields = {
            "muac": {
                "2": {
                    "value": 20,
                    "collection_date": "2021-05-01",
                }
            },
            "month_worked": {
                "4": {
                    "value": 20,
                    "collection_date": "2021-04-01",
                }
            },
        }
        individual.save()
        row = service._generate_row(individual)
        self.assertIsNone(row)

    def test_get_individuals_queryset_registration_data_import_id_filter(self) -> None:
        rdi1 = RegistrationDataImportFactory()
        individual = self.individuals[0]
        individual.registration_data_import = rdi1
        individual.save()
        self.periodic_data_update_template.filters = {
            "registration_data_import_id": encode_id_base64(str(rdi1.pk), "RegistrationDataImport")
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual)

    def test_get_individuals_queryset_target_population_id_filter(self) -> None:
        create_household_and_individuals(
            household_data={
                "business_area": self.business_area,
                "program_id": self.program.pk,
                "registration_data_import": self.rdi,
            },
            individuals_data=[
                {
                    "business_area": self.business_area,
                    "program_id": self.program.pk,
                    "registration_data_import": self.rdi,
                },
                {
                    "business_area": self.business_area,
                    "program_id": self.program.pk,
                    "registration_data_import": self.rdi,
                },
            ],
        )
        tp = TargetPopulationFactory()
        self.periodic_data_update_template.filters = {
            "target_population_id": encode_id_base64(str(tp.pk), "TargetPopulation")
        }
        tp.households.add(self.household)
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 2)
        self.assertEqual(set(queryset), set(self.individuals))

    def test_get_individuals_queryset_gender_filter(self) -> None:
        male = self.individuals[0]
        female = self.individuals[1]
        male.sex = MALE
        female.sex = FEMALE
        male.save()
        female.save()
        self.periodic_data_update_template.filters = {"gender": FEMALE}
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), female)

    @freeze_time("2024-07-12")
    def test_get_individuals_queryset_age_filter(self) -> None:
        individual32yo = self.individuals[0]
        individual29yo = self.individuals[1]
        individual32yo.birth_date = "1991-10-16"
        individual29yo.birth_date = "1994-10-16"
        individual32yo.save()
        individual29yo.save()
        self.periodic_data_update_template.filters = {"age": {"from": 30, "to": 32}}
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual32yo)

    def test_get_individuals_queryset_registration_date_filter(self) -> None:
        individual2023 = self.individuals[0]
        individual2020 = self.individuals[1]
        individual2023.first_registration_date = "2023-10-16"
        individual2020.first_registration_date = "2020-10-16"
        individual2023.save()
        individual2020.save()
        self.periodic_data_update_template.filters = {"registration_date": {"from": "2023-10-16", "to": "2023-10-16"}}
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual2023)

    def test_get_individuals_queryset_admin_filter(self) -> None:
        area_type_level_1 = AreaTypeFactory(
            name="State1",
            area_level=1,
        )
        area_type_level_2 = AreaTypeFactory(
            name="State2",
            area_level=2,
        )
        area1a = AreaFactory(name="City Test1", area_type=area_type_level_1, p_code="area1a")
        area2a = AreaFactory(name="City Test2", area_type=area_type_level_2, p_code="area2a", parent=area1a)
        area1b = AreaFactory(name="City Test1b", area_type=area_type_level_1, p_code="area1b")
        area2b = AreaFactory(name="City Test2b", area_type=area_type_level_2, p_code="area2b", parent=area1b)
        create_household_and_individuals(
            household_data={
                "business_area": self.business_area,
                "program_id": self.program.pk,
                "registration_data_import": self.rdi,
                "admin1": area1b,
                "admin2": area2b,
            },
            individuals_data=[
                {
                    "business_area": self.business_area,
                    "program_id": self.program.pk,
                    "registration_data_import": self.rdi,
                },
                {
                    "business_area": self.business_area,
                    "program_id": self.program.pk,
                    "registration_data_import": self.rdi,
                },
            ],
        )
        self.household.admin1 = area1a
        self.household.save()
        self.periodic_data_update_template.filters = {"admin1": [encode_id_base64(str(area1a.pk), "Area")]}
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 2)
        self.assertEqual(set(queryset), set(self.individuals))

        self.household.admin2 = area2a
        self.household.save()
        self.periodic_data_update_template.filters = {"admin2": [encode_id_base64(str(area2a.pk), "Area")]}
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 2)
        self.assertEqual(set(queryset), set(self.individuals))

    def test_get_individuals_queryset_has_grievance_ticket_referral_filter(self) -> None:
        individual_with_ticket = self.individuals[0]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_REFERRAL,
            status=GrievanceTicket.STATUS_NEW,
        )
        TicketReferralDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual_with_ticket)

    def test_get_individuals_queryset_has_grievance_ticket_referral_exclude_filter(self) -> None:
        individual_with_ticket = self.individuals[0]
        individual_without_ticket = self.individuals[1]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_REFERRAL,
            status=GrievanceTicket.STATUS_NEW,
        )
        TicketReferralDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": False,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual_without_ticket)

    def test_get_individuals_queryset_has_grievance_ticket_negative_feedback_filter(self) -> None:
        individual_with_ticket = self.individuals[0]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_POSITIVE_FEEDBACK,
            status=GrievanceTicket.STATUS_NEW,
        )
        TicketNegativeFeedbackDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual_with_ticket)

    def test_get_individuals_queryset_has_grievance_ticket_positive_feedback_filter(self) -> None:
        individual_with_ticket = self.individuals[0]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_POSITIVE_FEEDBACK,
            status=GrievanceTicket.STATUS_NEW,
        )
        TicketPositiveFeedbackDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual_with_ticket)

    def test_get_individuals_queryset_has_grievance_ticket_needs_adjudication_filter(self) -> None:
        create_household_and_individuals(
            household_data={
                "business_area": self.business_area,
                "program_id": self.program.pk,
                "registration_data_import": self.rdi,
            },
            individuals_data=[
                {
                    "business_area": self.business_area,
                    "program_id": self.program.pk,
                    "registration_data_import": self.rdi,
                },
                {
                    "business_area": self.business_area,
                    "program_id": self.program.pk,
                    "registration_data_import": self.rdi,
                },
            ],
        )
        individual_with_ticket = self.individuals[0]
        possible_duplicate = self.individuals[1]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_NEEDS_ADJUDICATION,
            status=GrievanceTicket.STATUS_NEW,
            issue_type=GrievanceTicket.ISSUE_TYPE_BIOGRAPHICAL_DATA_SIMILARITY,
        )
        details = TicketNeedsAdjudicationDetails.objects.create(
            ticket=grievance, golden_records_individual_id=individual_with_ticket.pk
        )
        details.possible_duplicates.add(possible_duplicate)
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 2)
        self.assertEqual(set(queryset), {individual_with_ticket, possible_duplicate})

    def test_get_individuals_queryset_has_grievance_ticket_system_flagging_filter(self) -> None:
        individual_with_ticket = self.individuals[0]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_SYSTEM_FLAGGING,
            status=GrievanceTicket.STATUS_NEW,
        )
        sanction_list_individual = SanctionListIndividualFactory()
        TicketSystemFlaggingDetails.objects.create(
            ticket=grievance,
            golden_records_individual_id=individual_with_ticket.pk,
            sanction_list_individual=sanction_list_individual,
        )
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual_with_ticket)

    def test_get_individuals_queryset_has_grievance_ticket_delete_individual_filter(self) -> None:
        individual_with_ticket = self.individuals[0]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_DATA_CHANGE,
            issue_type=GrievanceTicket.ISSUE_TYPE_DATA_CHANGE_DELETE_INDIVIDUAL,
            status=GrievanceTicket.STATUS_NEW,
        )
        TicketDeleteIndividualDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual_with_ticket)

    def test_get_individuals_queryset_has_grievance_ticket_individual_data_update_filter(self) -> None:
        individual_with_ticket = self.individuals[0]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_DATA_CHANGE,
            issue_type=GrievanceTicket.ISSUE_TYPE_INDIVIDUAL_DATA_CHANGE_DATA_UPDATE,
            status=GrievanceTicket.STATUS_NEW,
        )
        TicketIndividualDataUpdateDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual_with_ticket)

    def test_get_individuals_queryset_has_grievance_ticket_sensitive_filter(self) -> None:
        individual_with_ticket = self.individuals[0]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_SENSITIVE_GRIEVANCE,
            status=GrievanceTicket.STATUS_NEW,
            issue_type=GrievanceTicket.ISSUE_TYPE_CONFLICT_OF_INTEREST,
        )
        TicketSensitiveDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual_with_ticket)

    def test_get_individuals_queryset_has_grievance_ticket_complaint_filter(self) -> None:
        individual_with_ticket = self.individuals[0]
        grievance = GrievanceTicketFactory(
            business_area=self.business_area,
            category=GrievanceTicket.CATEGORY_GRIEVANCE_COMPLAINT,
            status=GrievanceTicket.STATUS_NEW,
            issue_type=GrievanceTicket.ISSUE_TYPE_OTHER_COMPLAINT,
        )
        TicketComplaintDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)
        self.periodic_data_update_template.filters = {
            "has_grievance_ticket": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), individual_with_ticket)

    def test_get_individuals_queryset_received_assistance_filter(self) -> None:
        household_without_payment, individuals_without_payment = create_household_and_individuals(
            household_data={
                "business_area": self.business_area,
                "program_id": self.program.pk,
                "registration_data_import": self.rdi,
            },
            individuals_data=[
                {
                    "business_area": self.business_area,
                    "program_id": self.program.pk,
                    "registration_data_import": self.rdi,
                },
                {
                    "business_area": self.business_area,
                    "program_id": self.program.pk,
                    "registration_data_import": self.rdi,
                },
            ],
        )
        PaymentFactory(household=self.household, status=Payment.STATUS_DISTRIBUTION_SUCCESS)
        self.periodic_data_update_template.filters = {
            "received_assistance": True,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 2)
        self.assertEqual(set(queryset), set(self.individuals))

        self.periodic_data_update_template.filters = {
            "received_assistance": False,
        }
        self.periodic_data_update_template.save()
        service = PeriodicDataUpdateExportTemplateService(self.periodic_data_update_template)
        queryset = service._get_individuals_queryset()
        self.assertEqual(queryset.count(), 2)
        self.assertEqual(set(queryset), set(individuals_without_payment))
