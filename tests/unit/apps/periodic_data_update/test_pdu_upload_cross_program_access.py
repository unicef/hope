"""The upload reads the template id out of the uploaded file, so it has to be scoped to the url path."""

from io import BytesIO
from typing import Any, Callable

from django.core.files.uploadedfile import SimpleUploadedFile
import openpyxl
import pytest
from rest_framework import status
from rest_framework.reverse import reverse
from rest_framework.test import APIClient

from extras.test_utils.factories import (
    BusinessAreaFactory,
    FlexibleAttributeForPDUFactory,
    HouseholdFactory,
    IndividualFactory,
    PDUXlsxTemplateFactory,
    PeriodicFieldDataFactory,
    ProgramFactory,
    UserFactory,
)
from hope.apps.account.permissions import Permissions
from hope.apps.periodic_data_update.service.periodic_data_update_export_template_service import (
    PDUXlsxExportTemplateService,
)
from hope.apps.periodic_data_update.utils import populate_pdu_with_null_values
from hope.models import (
    BusinessArea,
    FlexibleAttribute,
    Individual,
    PDUXlsxTemplate,
    PDUXlsxUpload,
    PeriodicFieldData,
    Program,
    User,
)

pytestmark = pytest.mark.django_db


def _pdu_field(program: Program) -> FlexibleAttribute:
    return FlexibleAttributeForPDUFactory(
        program=program,
        label="PDU Field",
        pdu_data=PeriodicFieldDataFactory(
            subtype=PeriodicFieldData.STRING, number_of_rounds=1, rounds_names=["January"]
        ),
    )


def _individual_with_empty_pdu(program: Program) -> Individual:
    individual = IndividualFactory(household=None, business_area=program.business_area, program=program)
    individual.household = HouseholdFactory(
        business_area=program.business_area, program=program, head_of_household=individual
    )
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()
    return individual


def _template_with_file(program: Program, field_name: str) -> PDUXlsxTemplate:
    template = PDUXlsxTemplateFactory(
        program=program,
        rounds_data=[{"field": field_name, "round": 1, "round_name": "January", "number_of_records": 1}],
    )
    service = PDUXlsxExportTemplateService(template)
    service.generate_workbook()
    service.save_xlsx_file()
    return template


def _filled_workbook(template: PDUXlsxTemplate, individual_uuid: str | None = None) -> SimpleUploadedFile:
    """The exported template with a value typed in, optionally re-pointed at another individual."""
    workbook = openpyxl.load_workbook(template.file.file)
    sheet = workbook[PDUXlsxExportTemplateService.PDU_SHEET]
    if individual_uuid is not None:
        sheet.cell(row=2, column=1, value=individual_uuid)
    sheet.cell(row=2, column=7, value="Positive")
    sheet.cell(row=2, column=8, value="2024-07-20")
    stream = BytesIO()
    workbook.save(stream)
    return SimpleUploadedFile(
        "file.xlsx",
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.fixture
def attacker_business_area() -> BusinessArea:
    return BusinessAreaFactory(slug="afghanistan", name="Afghanistan")


@pytest.fixture
def attacker_program(attacker_business_area: BusinessArea) -> Program:
    return ProgramFactory(business_area=attacker_business_area, name="Attacker Program")


@pytest.fixture
def victim_program() -> Program:
    return ProgramFactory(business_area=BusinessAreaFactory(slug="ukraine", name="Ukraine"), name="Victim Program")


@pytest.fixture
def victim_field(victim_program: Program) -> FlexibleAttribute:
    return _pdu_field(victim_program)


@pytest.fixture
def victim_individual(victim_program: Program, victim_field: FlexibleAttribute) -> Individual:
    return _individual_with_empty_pdu(victim_program)


@pytest.fixture
def attacker(
    attacker_business_area: BusinessArea,
    attacker_program: Program,
    create_user_role_with_permissions: Callable,
) -> User:
    user = UserFactory()
    create_user_role_with_permissions(user, [Permissions.PDU_UPLOAD], attacker_business_area, attacker_program)
    return user


@pytest.fixture
def api_client_attacker(api_client: Callable, attacker: User) -> APIClient:
    return api_client(attacker)


@pytest.fixture
def url_upload(attacker_business_area: BusinessArea, attacker_program: Program) -> str:
    return reverse(
        "api:periodic-data-update:periodic-data-update-uploads-upload",
        kwargs={"business_area_slug": attacker_business_area.slug, "program_code": attacker_program.code},
    )


def test_upload_with_template_of_other_business_area_is_denied(
    api_client_attacker: APIClient,
    url_upload: str,
    victim_program: Program,
    victim_field: FlexibleAttribute,
    victim_individual: Individual,
    django_capture_on_commit_callbacks: Any,
) -> None:
    victim_template = _template_with_file(victim_program, victim_field.name)
    upload_file = _filled_workbook(victim_template)

    with django_capture_on_commit_callbacks(execute=True):
        response = api_client_attacker.post(url_upload, {"file": upload_file}, format="multipart")

    assert response.status_code == status.HTTP_400_BAD_REQUEST, response.json()
    victim_individual.refresh_from_db()
    assert victim_individual.flex_fields[victim_field.name]["1"]["value"] is None
    assert not PDUXlsxUpload.objects.filter(template=victim_template).exists()


def test_upload_row_pointing_at_individual_of_other_business_area_is_rejected(
    api_client_attacker: APIClient,
    url_upload: str,
    attacker_program: Program,
    victim_field: FlexibleAttribute,
    victim_individual: Individual,
    django_capture_on_commit_callbacks: Any,
) -> None:
    attacker_field = _pdu_field(attacker_program)
    _individual_with_empty_pdu(attacker_program)
    attacker_template = _template_with_file(attacker_program, attacker_field.name)
    upload_file = _filled_workbook(attacker_template, individual_uuid=str(victim_individual.pk))

    with django_capture_on_commit_callbacks(execute=True):
        response = api_client_attacker.post(url_upload, {"file": upload_file}, format="multipart")

    assert response.status_code == status.HTTP_202_ACCEPTED, response.json()
    victim_individual.refresh_from_db()
    assert victim_individual.flex_fields[victim_field.name]["1"]["value"] is None
    assert PDUXlsxUpload.objects.get(template=attacker_template).status == PDUXlsxUpload.Status.FAILED
