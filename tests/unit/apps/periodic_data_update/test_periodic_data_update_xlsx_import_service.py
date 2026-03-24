"""Tests for periodic data update XLSX import service."""

import datetime
from io import BytesIO
import json
from tempfile import _TemporaryFileWrapper
from typing import Any
import uuid

from django.core.exceptions import ValidationError
from django.core.files import File
import openpyxl
import pytest

from extras.test_utils.factories import (
    BusinessAreaFactory,
    FlexibleAttributeForPDUFactory,
    HouseholdFactory,
    IndividualFactory,
    PeriodicFieldDataFactory,
    ProgramFactory,
    RegistrationDataImportFactory,
)
from hope.apps.periodic_data_update.service.periodic_data_update_export_template_service import (
    PDUXlsxExportTemplateService,
)
from hope.apps.periodic_data_update.service.periodic_data_update_import_service import (
    PDUXlsxImportService,
)
from hope.apps.periodic_data_update.utils import populate_pdu_with_null_values
from hope.models import BusinessArea, PDUXlsxTemplate, PDUXlsxUpload, PeriodicFieldData, Program, RegistrationDataImport

pytestmark = pytest.mark.django_db


def add_pdu_data_to_xlsx(
    periodic_data_update_template: PDUXlsxTemplate, rows: list[list[Any]]
) -> _TemporaryFileWrapper:
    wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
    ws_pdu = wb[PDUXlsxExportTemplateService.PDU_SHEET]
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            ws_pdu.cell(row=row_index + 2, column=col_index + 7, value=value)
    tmp_file = BytesIO()
    wb.save(tmp_file)
    tmp_file.seek(0)
    return tmp_file


@pytest.fixture
def business_area(db: Any) -> BusinessArea:
    return BusinessAreaFactory(slug="afghanistan", name="Afghanistan")


@pytest.fixture
def rdi(business_area: BusinessArea) -> RegistrationDataImport:
    return RegistrationDataImportFactory(business_area=business_area)


@pytest.fixture
def program(business_area: BusinessArea) -> Program:
    return ProgramFactory(business_area=business_area)


@pytest.fixture
def household(business_area: BusinessArea, program: Program, rdi: RegistrationDataImport) -> Any:
    individual = IndividualFactory(
        household=None,
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
    )
    household = HouseholdFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        head_of_household=individual,
    )
    individual.household = household
    individual.save()
    return household


@pytest.fixture
def individual(household: Any) -> Any:
    return household.head_of_household


@pytest.fixture
def string_attribute(program: Program) -> Any:
    pdu_data = PeriodicFieldDataFactory(
        subtype=PeriodicFieldData.STRING,
        number_of_rounds=1,
        rounds_names=["May"],
    )
    return FlexibleAttributeForPDUFactory(
        program=program,
        label="String Attribute",
        pdu_data=pdu_data,
    )


@pytest.fixture
def decimal_attribute(program: Program) -> Any:
    pdu_data = PeriodicFieldDataFactory(
        subtype=PeriodicFieldData.DECIMAL,
        number_of_rounds=1,
        rounds_names=["May"],
    )
    return FlexibleAttributeForPDUFactory(
        program=program,
        label="Decimal Attribute",
        pdu_data=pdu_data,
    )


@pytest.fixture
def boolean_attribute(program: Program) -> Any:
    pdu_data = PeriodicFieldDataFactory(
        subtype=PeriodicFieldData.BOOL,
        number_of_rounds=1,
        rounds_names=["May"],
    )
    return FlexibleAttributeForPDUFactory(
        program=program,
        label="Boolean Attribute",
        pdu_data=pdu_data,
    )


@pytest.fixture
def date_attribute(program: Program) -> Any:
    pdu_data = PeriodicFieldDataFactory(
        subtype=PeriodicFieldData.DATE,
        number_of_rounds=1,
        rounds_names=["May"],
    )
    return FlexibleAttributeForPDUFactory(
        program=program,
        label="Date Attribute",
        pdu_data=pdu_data,
    )


def prepare_test_data(
    program: Program,
    business_area: BusinessArea,
    rounds_data: list,
    rows: list,
) -> tuple:
    periodic_data_update_template = PDUXlsxTemplate.objects.create(
        program=program,
        business_area=business_area,
        filters={},
        rounds_data=rounds_data,
    )
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    service.generate_workbook()
    service.save_xlsx_file()
    tmp_file = add_pdu_data_to_xlsx(periodic_data_update_template, rows)
    periodic_data_update_upload = PDUXlsxUpload(
        template=periodic_data_update_template,
        created_by=periodic_data_update_template.created_by,
    )
    tmp_file.seek(0)
    periodic_data_update_upload.file.save("test.xlsx", File(tmp_file))
    periodic_data_update_upload.save()
    tmp_file.close()
    return periodic_data_update_template, periodic_data_update_upload


def test_import_data_string(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    string_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = string_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["Test Value", "2021-05-02"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.SUCCESSFUL
    assert periodic_data_update_upload.error_message is None
    individual.refresh_from_db()
    assert individual.flex_fields[flexible_attribute.name]["1"]["value"] == "Test Value"
    assert individual.flex_fields[flexible_attribute.name]["1"]["collection_date"] == "2021-05-02"


def test_import_data_decimal(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    decimal_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = decimal_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["20.456", "2021-05-02"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.SUCCESSFUL
    assert periodic_data_update_upload.error_message is None
    individual.refresh_from_db()
    assert individual.flex_fields[flexible_attribute.name]["1"]["value"] == 20.456
    assert individual.flex_fields[flexible_attribute.name]["1"]["collection_date"] == "2021-05-02"


def test_import_data_boolean(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    boolean_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = boolean_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [[True, "2021-05-02"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.SUCCESSFUL
    assert periodic_data_update_upload.error_message is None
    individual.refresh_from_db()
    assert individual.flex_fields[flexible_attribute.name]["1"]["value"] is True
    assert individual.flex_fields[flexible_attribute.name]["1"]["collection_date"] == "2021-05-02"


def test_import_data_boolean_1(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    boolean_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = boolean_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [[False, "2021-05-02"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.SUCCESSFUL
    assert periodic_data_update_upload.error_message is None
    individual.refresh_from_db()
    assert individual.flex_fields[flexible_attribute.name]["1"]["value"] is False
    assert individual.flex_fields[flexible_attribute.name]["1"]["collection_date"] == "2021-05-02"


def test_import_data_boolean_fail(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    boolean_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = boolean_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["Yes", "2021-05-02"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.FAILED
    errors = {
        "form_errors": [
            {
                "row": 2,
                "errors": {"boolean_attribute__round_value": [{"message": "Enter a boolean value", "code": "invalid"}]},
            }
        ],
        "non_form_errors": None,
    }
    assert json.loads(periodic_data_update_upload.error_message) == errors


def test_import_data_date(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["1996-06-21", "2021-05-02"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.SUCCESSFUL
    assert periodic_data_update_upload.error_message is None
    individual.refresh_from_db()
    assert individual.flex_fields[flexible_attribute.name]["1"]["value"] == "1996-06-21"
    assert individual.flex_fields[flexible_attribute.name]["1"]["collection_date"] == "2021-05-02"


def test_import_data_date_no_collection_date(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["1996-06-21"]],
    )
    periodic_data_update_template.created_at = datetime.datetime(2021, 3, 7)
    periodic_data_update_template.save()
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.SUCCESSFUL
    assert periodic_data_update_upload.error_message is None
    individual.refresh_from_db()
    assert individual.flex_fields[flexible_attribute.name]["1"]["value"] == "1996-06-21"
    assert individual.flex_fields[flexible_attribute.name]["1"]["collection_date"] == "2021-03-07"


def test_import_data_date_fail_string(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["bla bla", "2021-05-02"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.FAILED
    errors = {
        "form_errors": [
            {
                "row": 2,
                "errors": {"date_attribute__round_value": [{"message": "Enter a valid date.", "code": "invalid"}]},
            }
        ],
        "non_form_errors": None,
    }
    assert json.loads(periodic_data_update_upload.error_message) == errors


def test_import_data_date_fail_int(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [[58, "2021-05-02"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.FAILED
    errors = {
        "form_errors": [
            {
                "row": 2,
                "errors": {"date_attribute__round_value": [{"message": "Enter a date value", "code": "invalid"}]},
            }
        ],
        "non_form_errors": None,
    }
    assert json.loads(periodic_data_update_upload.error_message) == errors


def test_read_periodic_data_update_non_form_errors(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["2021-05-02", "2021-05-02"]],
    )
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    service.generate_workbook()
    service.save_xlsx_file()
    flexible_attribute.delete()
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.FAILED
    errors = {
        "form_errors": [],
        "non_form_errors": ["Some fields are missing in the flexible attributes"],
    }
    assert json.loads(periodic_data_update_upload.error_message) == errors


def test_import_data_date_format_correct(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [[datetime.datetime(2021, 5, 2), datetime.datetime(2021, 5, 2)]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()
    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.SUCCESSFUL
    assert periodic_data_update_upload.error_message is None
    individual.refresh_from_db()
    assert individual.flex_fields[flexible_attribute.name]["1"]["value"] == "2021-05-02"
    assert individual.flex_fields[flexible_attribute.name]["1"]["collection_date"] == "2021-05-02"


def test_read_periodic_data_update_template_object(
    program: Program,
    business_area: BusinessArea,
    string_attribute: Any,
) -> None:
    periodic_data_update_template = PDUXlsxTemplate.objects.create(
        program=program,
        business_area=business_area,
        filters={},
        rounds_data=[
            {
                "field": string_attribute.name,
                "round": 1,
                "round_name": string_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
    )
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    service.generate_workbook()
    service.save_xlsx_file()
    periodic_data_update_template_from_xlsx = PDUXlsxImportService.read_periodic_data_update_template_object(
        periodic_data_update_template.file.file
    )
    assert periodic_data_update_template_from_xlsx.pk == periodic_data_update_template.pk
    wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
    del wb.custom_doc_props[PDUXlsxExportTemplateService.PROPERTY_ID_NAME]
    tmp_file = BytesIO()
    wb.save(tmp_file)
    tmp_file.seek(0)
    periodic_data_update_template_from_xlsx = PDUXlsxImportService.read_periodic_data_update_template_object(tmp_file)
    assert periodic_data_update_template_from_xlsx.pk == periodic_data_update_template.pk
    wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
    del wb.custom_doc_props[PDUXlsxExportTemplateService.PROPERTY_ID_NAME]
    ws_meta = wb[PDUXlsxExportTemplateService.META_SHEET]
    ws_meta[PDUXlsxExportTemplateService.META_ID_ADDRESS] = ""
    tmp_file = BytesIO()
    wb.save(tmp_file)
    tmp_file.seek(0)
    with pytest.raises(ValidationError, match="Periodic Data Update Template ID is missing in the file"):
        PDUXlsxImportService.read_periodic_data_update_template_object(tmp_file)

    wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
    del wb.custom_doc_props[PDUXlsxExportTemplateService.PROPERTY_ID_NAME]
    ws_meta = wb[PDUXlsxExportTemplateService.META_SHEET]
    ws_meta[PDUXlsxExportTemplateService.META_ID_ADDRESS] = "abc"
    tmp_file = BytesIO()
    wb.save(tmp_file)
    tmp_file.seek(0)
    with pytest.raises(ValidationError, match="Periodic Data Update Template ID must be a number"):
        PDUXlsxImportService.read_periodic_data_update_template_object(tmp_file)

    wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
    del wb.custom_doc_props[PDUXlsxExportTemplateService.PROPERTY_ID_NAME]
    ws_meta = wb[PDUXlsxExportTemplateService.META_SHEET]
    ws_meta[PDUXlsxExportTemplateService.META_ID_ADDRESS] = True
    tmp_file = BytesIO()
    wb.save(tmp_file)
    tmp_file.seek(0)
    with pytest.raises(ValidationError, match="Periodic Data Update Template ID must be an integer"):
        PDUXlsxImportService.read_periodic_data_update_template_object(tmp_file)

    wb = openpyxl.load_workbook(periodic_data_update_template.file.file)
    del wb.custom_doc_props[PDUXlsxExportTemplateService.PROPERTY_ID_NAME]
    ws_meta = wb[PDUXlsxExportTemplateService.META_SHEET]
    ws_meta[PDUXlsxExportTemplateService.META_ID_ADDRESS] = -1
    tmp_file = BytesIO()
    wb.save(tmp_file)
    tmp_file.seek(0)
    with pytest.raises(ValidationError, match="Periodic Data Update Template with ID -1 not found"):
        PDUXlsxImportService.read_periodic_data_update_template_object(tmp_file)


def test_read_flexible_attributes(
    program: Program,
    business_area: BusinessArea,
    string_attribute: Any,
) -> None:
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": string_attribute.name,
                "round": 1,
                "round_name": string_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            },
            {
                "field": "Not existing field",
                "round": 1,
                "round_name": string_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            },
        ],
        [["1996-06-21", "2021-05-02"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service._open_workbook()
    with pytest.raises(ValidationError, match="Some fields are missing in the flexible attributes"):
        service._read_flexible_attributes()


def test_read_row(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["-", "-"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service._open_workbook()
    service._read_flexible_attributes()
    service._read_rows()


def test_import_cleaned_data(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["-", "-"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service._open_workbook()
    service._read_flexible_attributes()
    cleaned_data = {
        "individual__uuid": individual.id,
        "individual_unicef_id": individual.unicef_id,
        "first_name": "Debra",
        "last_name": "Taylor",
        "date_attribute__round_number": 2,
        "date_attribute__round_name": "May",
        "date_attribute__round_value": datetime.date(1996, 6, 21),
        "date_attribute__collection_date": datetime.date(2021, 5, 2),
    }
    with pytest.raises(
        ValidationError,
        match="Round number mismatch for field date_attribute. Expected: 1, Got: 2",
    ):
        service._import_cleaned_data(cleaned_data)
    not_existing_individual_id = uuid.uuid4()
    cleaned_data = {
        "individual__uuid": not_existing_individual_id,
        "individual_unicef_id": individual.unicef_id,
        "first_name": "Debra",
        "last_name": "Taylor",
        "date_attribute__round_number": 1,
        "date_attribute__round_name": "May",
        "date_attribute__round_value": datetime.date(1996, 6, 21),
        "date_attribute__collection_date": datetime.date(2021, 5, 2),
    }
    with pytest.raises(
        ValidationError,
        match=f"Individual not found for {individual.unicef_id}",
    ):
        service._import_cleaned_data(cleaned_data)
    individual.flex_fields = {"date_attribute": {"1": {"value": "1996-06-21", "collection_date": "2021-05-02"}}}
    individual.save()
    cleaned_data = {
        "individual__uuid": individual.id,
        "individual_unicef_id": individual.unicef_id,
        "first_name": "Debra",
        "last_name": "Taylor",
        "date_attribute__round_number": 1,
        "date_attribute__round_name": "May",
        "date_attribute__round_value": datetime.date(1996, 6, 21),
        "date_attribute__collection_date": datetime.date(2021, 5, 2),
    }
    with pytest.raises(
        ValidationError,
        match=f"Value already exists for field date_attribute for round 1 and individual {individual.unicef_id}",
    ):
        service._import_cleaned_data(cleaned_data)


def test_set_round_value(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["-", "-"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    service._open_workbook()
    service._read_flexible_attributes()
    service.set_round_value(individual, flexible_attribute.name, 1, "1996-06-21", "2021-05-02")
    assert individual.flex_fields[flexible_attribute.name]["1"]["value"] == "1996-06-21"
    assert individual.flex_fields[flexible_attribute.name]["1"]["collection_date"] == "2021-05-02"


def test_get_form_field_for_value(
    program: Program,
    business_area: BusinessArea,
    date_attribute: Any,
) -> None:
    flexible_attribute = date_attribute
    pdu_data = flexible_attribute.pdu_data
    PeriodicFieldData.objects.filter(id=pdu_data.id).update(subtype="INVALID")
    flexible_attribute.refresh_from_db()
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["-", "-"]],
    )
    service = PDUXlsxImportService(periodic_data_update_upload)
    with pytest.raises(
        ValidationError,
        match=f"Invalid subtype for field {flexible_attribute.name}",
    ):
        service._get_form_field_for_value(flexible_attribute)


def test_import_data_invalid_subtype_fail(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    boolean_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": boolean_attribute.name,
                "round": 1,
                "round_name": boolean_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            },
        ],
        [["True", "2021-05-02"]],
    )

    PeriodicFieldData.objects.filter(id=boolean_attribute.pdu_data.id).update(subtype="INVALID_SUBTYPE")

    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()

    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.FAILED
    error_data = json.loads(periodic_data_update_upload.error_message)
    assert [f"Invalid subtype for field {boolean_attribute.name}"] in error_data["non_form_errors"]


def test_import_data_round_mismatch_validation_error(
    program: Program,
    business_area: BusinessArea,
    individual: Any,
    date_attribute: Any,
) -> None:
    populate_pdu_with_null_values(program, individual.flex_fields)
    individual.save()

    flexible_attribute = date_attribute
    (
        periodic_data_update_template,
        periodic_data_update_upload,
    ) = prepare_test_data(
        program,
        business_area,
        [
            {
                "field": flexible_attribute.name,
                "round": 1,
                "round_name": flexible_attribute.pdu_data.rounds_names[0],
                "number_of_records": 0,
            }
        ],
        [["1996-06-21", "2021-05-02"]],
    )

    periodic_data_update_template.rounds_data[0]["round"] = 2
    periodic_data_update_template.save()

    service = PDUXlsxImportService(periodic_data_update_upload)
    service.import_data()

    assert periodic_data_update_upload.status == PDUXlsxUpload.Status.FAILED
    error_data = json.loads(periodic_data_update_upload.error_message)
    assert ["Round number mismatch for field date_attribute. Expected: 2, Got: 1"] in error_data["non_form_errors"]
