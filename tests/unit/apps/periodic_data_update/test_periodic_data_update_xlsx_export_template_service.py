"""Tests for PDU XLSX export template service."""

from typing import Any

from freezegun import freeze_time
import openpyxl
import pytest

from extras.test_utils.factories import (
    AreaFactory,
    AreaTypeFactory,
    BusinessAreaFactory,
    FlexibleAttributeForPDUFactory,
    GrievanceTicketFactory,
    HouseholdFactory,
    IndividualFactory,
    PaymentFactory,
    PaymentPlanFactory,
    PDUXlsxTemplateFactory,
    PeriodicFieldDataFactory,
    ProgramFactory,
    RegistrationDataImportFactory,
    SanctionListIndividualFactory,
)
from hope.apps.grievance.models import (
    GrievanceTicket,
    TicketComplaintDetails,
    TicketDeleteIndividualDetails,
    TicketIndividualDataUpdateDetails,
    TicketNeedsAdjudicationDetails,
    TicketNegativeFeedbackDetails,
    TicketPositiveFeedbackDetails,
    TicketReferralDetails,
    TicketSensitiveDetails,
    TicketSystemFlaggingDetails,
)
from hope.apps.household.const import FEMALE, MALE
from hope.apps.periodic_data_update.service.periodic_data_update_export_template_service import (
    PDUXlsxExportTemplateService,
)
from hope.models import BusinessArea, Individual, Payment, PeriodicFieldData, Program

pytestmark = pytest.mark.django_db


@pytest.fixture
def business_area() -> BusinessArea:
    return BusinessAreaFactory(slug="afghanistan", name="Afghanistan")


@pytest.fixture
def program(business_area: BusinessArea) -> Program:
    return ProgramFactory(name="Test Program", status=Program.DRAFT, business_area=business_area)


@pytest.fixture
def rdi(business_area: BusinessArea) -> Any:
    return RegistrationDataImportFactory(business_area=business_area)


@pytest.fixture
def individuals(business_area: BusinessArea, program: Program, rdi: Any) -> list[Individual]:
    individual1 = IndividualFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        household=None,
    )
    individual2 = IndividualFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        household=None,
    )
    return [individual1, individual2]


@pytest.fixture
def household(business_area: BusinessArea, program: Program, rdi: Any, individuals: list[Individual]) -> Any:
    household = HouseholdFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        head_of_household=individuals[0],
    )

    individuals[0].household = household
    individuals[0].save()
    individuals[1].household = household
    individuals[1].save()
    return household


@pytest.fixture
def pdu_field_muac(program: Program) -> Any:
    pdu_data = PeriodicFieldDataFactory(
        subtype=PeriodicFieldData.DECIMAL,
        number_of_rounds=5,
        rounds_names=["January", "February", "March", "April", "May"],
    )
    return FlexibleAttributeForPDUFactory(
        program=program,
        label="MUAC",
        pdu_data=pdu_data,
    )


@pytest.fixture
def pdu_field_month_worked(program: Program) -> Any:
    pdu_data = PeriodicFieldDataFactory(
        subtype=PeriodicFieldData.DECIMAL,
        number_of_rounds=5,
        rounds_names=["January", "February", "March", "April", "May"],
    )
    return FlexibleAttributeForPDUFactory(
        program=program,
        label="Month Worked",
        pdu_data=pdu_data,
    )


@pytest.fixture
def periodic_data_update_template(
    program: Program,
    pdu_field_muac: Any,
    pdu_field_month_worked: Any,
) -> Any:
    return PDUXlsxTemplateFactory(
        program=program,
        rounds_data=[
            {
                "field": pdu_field_muac.name,
                "round": 2,
                "round_name": "May",
                "number_of_records": 0,
            },
            {
                "field": pdu_field_month_worked.name,
                "round": 4,
                "round_name": "April",
                "number_of_records": 0,
            },
        ],
    )


def test_generate_workbook(
    periodic_data_update_template: Any,
    individuals: list[Individual],
    household: Any,
) -> None:
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    wb = service.generate_workbook()
    assert wb is not None
    assert wb.sheetnames == [service.PDU_SHEET, service.META_SHEET]
    meta_sheet = wb[service.META_SHEET]
    pdu_sheet = wb[service.PDU_SHEET]
    assert meta_sheet["b1"].value == periodic_data_update_template.pk
    assert wb.custom_doc_props["pdu_template_id"].value == str(periodic_data_update_template.pk)
    assert pdu_sheet.max_row == 3


def test_save_xlsx_file(periodic_data_update_template: Any) -> None:
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    service.generate_workbook()
    service.save_xlsx_file()
    periodic_data_update_template.refresh_from_db()
    assert periodic_data_update_template.file is not None
    wb = openpyxl.load_workbook(periodic_data_update_template.file.file.path)
    assert wb.sheetnames == [service.PDU_SHEET, service.META_SHEET]


def test_generate_header(periodic_data_update_template: Any) -> None:
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    test_header = [
        "individual__uuid",
        "individual_unicef_id",
        "first_name",
        "last_name",
        "muac__round_number",
        "muac__round_name",
        "muac__round_value",
        "muac__collection_date",
        "month_worked__round_number",
        "month_worked__round_name",
        "month_worked__round_value",
        "month_worked__collection_date",
    ]
    assert service._generate_header() == test_header


def test_generate_row_empty_flex_fields_individual(
    periodic_data_update_template: Any,
    individuals: list[Individual],
) -> None:
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    individual = individuals[0]
    row = service._generate_row(individual)
    expected_row = [
        str(individual.pk),
        individual.unicef_id,
        individual.given_name,
        individual.family_name,
        2,
        "May",
        "",
        "",
        4,
        "April",
        "",
        "",
    ]
    assert row == expected_row


def test_generate_row_half_filled_flex_fields_individual(
    periodic_data_update_template: Any,
    individuals: list[Individual],
) -> None:
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    individual = individuals[0]
    individual.flex_fields = {
        "muac": {
            "2": {
                "value": 20,
                "collection_date": "2021-05-01",
            }
        }
    }
    individual.save()
    row = service._generate_row(individual)
    expected_row = [
        str(individual.pk),
        individual.unicef_id,
        individual.given_name,
        individual.family_name,
        2,
        "May",
        "-",
        "-",
        4,
        "April",
        "",
        "",
    ]
    assert row == expected_row


def test_generate_row_fully_filled_flex_fields_individual(
    periodic_data_update_template: Any,
    individuals: list[Individual],
) -> None:
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    individual = individuals[0]
    individual.flex_fields = {
        "muac": {
            "2": {
                "value": 20,
                "collection_date": "2021-05-01",
            }
        },
        "month_worked": {
            "4": {
                "value": 20,
                "collection_date": "2021-04-01",
            }
        },
    }
    individual.save()
    row = service._generate_row(individual)
    assert row is None


def test_get_individuals_queryset_registration_data_import_id_filter(
    periodic_data_update_template: Any,
    individuals: list[Individual],
    business_area: BusinessArea,
) -> None:
    rdi1 = RegistrationDataImportFactory(business_area=business_area)
    individual = individuals[0]
    individual.registration_data_import = rdi1
    individual.save()
    periodic_data_update_template.filters = {"registration_data_import_id": str(rdi1.pk)}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual


def test_get_individuals_queryset_target_population_id_filter(
    periodic_data_update_template: Any,
    program: Program,
    business_area: BusinessArea,
    rdi: Any,
    individuals: list[Individual],
    household: Any,
) -> None:
    individual_other_1 = IndividualFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        household=None,
    )
    individual_other_2 = IndividualFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        household=None,
    )

    household2 = HouseholdFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        head_of_household=individual_other_1,
    )
    individual_other_1.household = household2
    individual_other_1.save()
    individual_other_2.household = household2
    individual_other_2.save()

    pp = PaymentPlanFactory()
    PaymentFactory(parent=pp, household=household, collector=individuals[0])
    periodic_data_update_template.filters = {"target_population_id": str(pp.pk)}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 2
    assert set(queryset) == set(individuals)


def test_get_individuals_queryset_gender_filter(
    periodic_data_update_template: Any,
    individuals: list[Individual],
) -> None:
    male = individuals[0]
    female = individuals[1]
    male.sex = MALE
    female.sex = FEMALE
    male.save()
    female.save()

    periodic_data_update_template.filters = {"gender": FEMALE}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == female


@freeze_time("2024-07-12")
def test_get_individuals_queryset_age_filter(
    periodic_data_update_template: Any,
    individuals: list[Individual],
) -> None:
    individual32yo = individuals[0]
    individual29yo = individuals[1]
    individual32yo.birth_date = "1991-10-16"
    individual29yo.birth_date = "1994-10-16"
    individual32yo.save()
    individual29yo.save()

    periodic_data_update_template.filters = {"age": {"from": 30, "to": 32}}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual32yo


def test_get_individuals_queryset_registration_date_filter(
    periodic_data_update_template: Any,
    individuals: list[Individual],
) -> None:
    individual2023 = individuals[0]
    individual2020 = individuals[1]
    individual2023.first_registration_date = "2023-10-16"
    individual2020.first_registration_date = "2020-10-16"
    individual2023.save()
    individual2020.save()

    periodic_data_update_template.filters = {"registration_date": {"from": "2023-10-16", "to": "2023-10-16"}}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual2023


def test_get_individuals_queryset_admin_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    program: Program,
    rdi: Any,
    individuals: list[Individual],
    household: Any,
) -> None:
    area_type_level_1 = AreaTypeFactory(name="State1", area_level=1)
    area_type_level_2 = AreaTypeFactory(name="State2", area_level=2)
    area1a = AreaFactory(name="City Test1", area_type=area_type_level_1, p_code="area1a")
    area2a = AreaFactory(
        name="City Test2",
        area_type=area_type_level_2,
        p_code="area2a",
        parent=area1a,
    )
    area1b = AreaFactory(name="City Test1b", area_type=area_type_level_1, p_code="area1b")
    area2b = AreaFactory(
        name="City Test2b",
        area_type=area_type_level_2,
        p_code="area2b",
        parent=area1b,
    )

    individual_other_1 = IndividualFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        household=None,
    )
    individual_other_2 = IndividualFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        household=None,
    )

    household2 = HouseholdFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        admin1=area1b,
        admin2=area2b,
        head_of_household=individual_other_1,
    )
    individual_other_1.household = household2
    individual_other_1.save()
    individual_other_2.household = household2
    individual_other_2.save()

    household.admin1 = area1a
    household.save()
    periodic_data_update_template.filters = {"admin1": [str(area1a.pk)]}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 2
    assert set(queryset) == set(individuals)

    household.admin2 = area2a
    household.save()
    periodic_data_update_template.filters = {"admin2": [str(area2a.pk)]}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 2
    assert set(queryset) == set(individuals)


def test_get_individuals_queryset_has_grievance_ticket_referral_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    individuals: list[Individual],
) -> None:
    individual_with_ticket = individuals[0]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_REFERRAL,
        status=GrievanceTicket.STATUS_NEW,
        issue_type=None,
    )
    TicketReferralDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)

    periodic_data_update_template.filters = {"has_grievance_ticket": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual_with_ticket


def test_get_individuals_queryset_has_grievance_ticket_referral_exclude_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    individuals: list[Individual],
) -> None:
    individual_with_ticket = individuals[0]
    individual_without_ticket = individuals[1]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_REFERRAL,
        status=GrievanceTicket.STATUS_NEW,
        issue_type=None,
    )
    TicketReferralDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)

    periodic_data_update_template.filters = {"has_grievance_ticket": False}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual_without_ticket


def test_get_individuals_queryset_has_grievance_ticket_negative_feedback_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    individuals: list[Individual],
) -> None:
    individual_with_ticket = individuals[0]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_POSITIVE_FEEDBACK,
        status=GrievanceTicket.STATUS_NEW,
        issue_type=None,
    )
    TicketNegativeFeedbackDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)

    periodic_data_update_template.filters = {"has_grievance_ticket": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual_with_ticket


def test_get_individuals_queryset_has_grievance_ticket_positive_feedback_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    individuals: list[Individual],
) -> None:
    individual_with_ticket = individuals[0]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_POSITIVE_FEEDBACK,
        status=GrievanceTicket.STATUS_NEW,
        issue_type=None,
    )
    TicketPositiveFeedbackDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)

    periodic_data_update_template.filters = {"has_grievance_ticket": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual_with_ticket


def test_get_individuals_queryset_has_grievance_ticket_needs_adjudication_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    program: Program,
    rdi: Any,
    individuals: list[Individual],
) -> None:
    possible_duplicate = IndividualFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        household=None,
    )

    household2 = HouseholdFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        head_of_household=possible_duplicate,
    )
    possible_duplicate.household = household2
    possible_duplicate.save()

    individual_with_ticket = individuals[0]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_NEEDS_ADJUDICATION,
        status=GrievanceTicket.STATUS_NEW,
        issue_type=GrievanceTicket.ISSUE_TYPE_BIOGRAPHICAL_DATA_SIMILARITY,
    )
    details = TicketNeedsAdjudicationDetails.objects.create(
        ticket=grievance, golden_records_individual_id=individual_with_ticket.pk
    )
    details.possible_duplicates.add(possible_duplicate)

    periodic_data_update_template.filters = {"has_grievance_ticket": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 2
    assert set(queryset) == {individual_with_ticket, possible_duplicate}


def test_get_individuals_queryset_has_grievance_ticket_sensitive_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    individuals: list[Individual],
) -> None:
    individual_with_ticket = individuals[0]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_SENSITIVE_GRIEVANCE,
        status=GrievanceTicket.STATUS_NEW,
        issue_type=GrievanceTicket.ISSUE_TYPE_CONFLICT_OF_INTEREST,
    )
    TicketSensitiveDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)

    periodic_data_update_template.filters = {"has_grievance_ticket": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual_with_ticket


def test_get_individuals_queryset_has_grievance_ticket_system_flagging_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    individuals: list[Individual],
) -> None:
    individual_with_ticket = individuals[0]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_SYSTEM_FLAGGING,
        status=GrievanceTicket.STATUS_NEW,
        issue_type=None,
    )
    sanction_list_individual = SanctionListIndividualFactory()
    TicketSystemFlaggingDetails.objects.create(
        ticket=grievance,
        golden_records_individual_id=individual_with_ticket.pk,
        sanction_list_individual=sanction_list_individual,
    )

    periodic_data_update_template.filters = {"has_grievance_ticket": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual_with_ticket


def test_get_individuals_queryset_has_grievance_ticket_delete_individual_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    individuals: list[Individual],
) -> None:
    individual_with_ticket = individuals[0]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_DATA_CHANGE,
        issue_type=GrievanceTicket.ISSUE_TYPE_DATA_CHANGE_DELETE_INDIVIDUAL,
        status=GrievanceTicket.STATUS_NEW,
    )
    TicketDeleteIndividualDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)

    periodic_data_update_template.filters = {"has_grievance_ticket": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual_with_ticket


def test_get_individuals_queryset_has_grievance_ticket_individual_data_update_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    individuals: list[Individual],
) -> None:
    individual_with_ticket = individuals[0]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_DATA_CHANGE,
        issue_type=GrievanceTicket.ISSUE_TYPE_INDIVIDUAL_DATA_CHANGE_DATA_UPDATE,
        status=GrievanceTicket.STATUS_NEW,
    )
    TicketIndividualDataUpdateDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)

    periodic_data_update_template.filters = {"has_grievance_ticket": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual_with_ticket


def test_get_individuals_queryset_has_grievance_ticket_complaint_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    individuals: list[Individual],
) -> None:
    individual_with_ticket = individuals[0]
    grievance = GrievanceTicketFactory(
        business_area=business_area,
        category=GrievanceTicket.CATEGORY_GRIEVANCE_COMPLAINT,
        status=GrievanceTicket.STATUS_NEW,
        issue_type=GrievanceTicket.ISSUE_TYPE_OTHER_COMPLAINT,
    )
    TicketComplaintDetails.objects.create(ticket=grievance, individual_id=individual_with_ticket.pk)

    periodic_data_update_template.filters = {"has_grievance_ticket": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 1
    assert queryset.first() == individual_with_ticket


def test_get_individuals_queryset_received_assistance_filter(
    periodic_data_update_template: Any,
    business_area: BusinessArea,
    program: Program,
    rdi: Any,
    individuals: list[Individual],
    household: Any,
) -> None:
    individual_without_payment_1 = IndividualFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        household=None,
    )
    individual_without_payment_2 = IndividualFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        household=None,
    )

    household_without_payment = HouseholdFactory(
        business_area=business_area,
        program=program,
        registration_data_import=rdi,
        head_of_household=individual_without_payment_1,
    )

    individual_without_payment_1.household = household_without_payment
    individual_without_payment_1.save()
    individual_without_payment_2.household = household_without_payment
    individual_without_payment_2.save()

    individuals_without_payment = [individual_without_payment_1, individual_without_payment_2]

    PaymentFactory(household=household, collector=individuals[0], status=Payment.STATUS_DISTRIBUTION_SUCCESS)

    periodic_data_update_template.filters = {"received_assistance": True}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 2
    assert set(queryset) == set(individuals)

    periodic_data_update_template.filters = {"received_assistance": False}
    periodic_data_update_template.save()
    service = PDUXlsxExportTemplateService(periodic_data_update_template)
    queryset = service._get_individuals_queryset()
    assert queryset.count() == 2
    assert set(queryset) == set(individuals_without_payment)
