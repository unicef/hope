"""Tests for CheckAgainstSanctionListTask helper methods."""

import base64
import datetime
import io
from unittest.mock import MagicMock

from openpyxl import load_workbook

from hope.apps.sanction_list.tasks.check_against_sanction_list import CheckAgainstSanctionListTask


def test_create_results_attachment_empty():
    task = CheckAgainstSanctionListTask()
    result = task._create_results_attachment({})
    # Decode the base64 result and load as a workbook
    wb = load_workbook(io.BytesIO(base64.b64decode(result)))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1  # header only
    assert rows[0][0] == "FIRST NAME"
    assert rows[0][1] == "SECOND NAME"
    assert rows[0][2] == "THIRD NAME"
    assert rows[0][3] == "FOURTH NAME"
    assert rows[0][4] == "DATE OF BIRTH"
    assert rows[0][5] == "ORIGINAL FILE ROW NUMBER"


def test_create_results_attachment_with_data():
    individual = MagicMock()
    individual.first_name = "John"
    individual.second_name = "Doe"
    individual.third_name = ""
    individual.fourth_name = ""
    individual.dates_of_birth.values_list.return_value = [datetime.date(1990, 5, 15)]

    task = CheckAgainstSanctionListTask()
    result = task._create_results_attachment({3: individual})
    wb = load_workbook(io.BytesIO(base64.b64decode(result)))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 2  # header + 1 data row
    assert rows[1][0] == "John"
    assert rows[1][1] == "Doe"
    assert rows[1][5] == 3
    assert "1990-05-15" in rows[1][4]
