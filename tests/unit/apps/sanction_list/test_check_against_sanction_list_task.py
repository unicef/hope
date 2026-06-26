import base64
import datetime
import io
import json
from typing import Any
from unittest.mock import patch

from constance.test import override_config
from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.template.loader import render_to_string
from django.test import override_settings
from django.utils import timezone
from freezegun import freeze_time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pytest

from extras.test_utils.factories import (
    SanctionListFactory,
    SanctionListIndividualDateOfBirthFactory,
    SanctionListIndividualFactory,
)
from hope.apps.sanction_list.tasks.check_against_sanction_list import (
    CheckAgainstSanctionListTask,
)
from hope.models import SanctionList, SanctionListIndividual, SanctionListIndividualDateOfBirth, UploadedXLSXFile

pytestmark = pytest.mark.django_db


def _build_xlsx(rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(("FIRST NAME", "SECOND NAME", "THIRD NAME", "FOURTH NAME", "DATE OF BIRTH"))
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _attachment_rows(mailjet_mock: Any) -> list[tuple]:
    attachment_b64 = mailjet_mock.return_value.attach_file.call_args.kwargs["attachment"]
    attachment_wb = load_workbook(io.BytesIO(base64.b64decode(attachment_b64)))
    return list(attachment_wb.active.iter_rows(values_only=True))


@pytest.fixture
def uploaded_file():
    return UploadedXLSXFile.objects.create(
        file=SimpleUploadedFile("test.xlsx", b"test"),
        associated_email="test_email@email.com",
    )


@pytest.fixture
def sanction_list() -> SanctionList:
    return SanctionListFactory()


@pytest.fixture
def john_doe(sanction_list: SanctionList) -> SanctionListIndividual:
    individual = SanctionListIndividualFactory(
        sanction_list=sanction_list,
        first_name="John",
        second_name="Doe",
        full_name="John Doe",
    )
    SanctionListIndividualDateOfBirthFactory(individual=individual, date=datetime.date(1980, 1, 1))
    return individual


@pytest.fixture
def make_uploaded_file(sanction_list: SanctionList) -> Any:
    def _make(rows: list[tuple]) -> UploadedXLSXFile:
        uploaded = UploadedXLSXFile.objects.create(
            file=SimpleUploadedFile("check.xlsx", _build_xlsx(rows)),
            associated_email="checker@example.com",
        )
        uploaded.selected_lists.add(sanction_list)
        return uploaded

    return _make


@patch("hope.apps.utils.celery_tasks.requests.post")
@patch("hope.apps.sanction_list.tasks.check_against_sanction_list.load_workbook")
@override_settings(EMAIL_SUBJECT_PREFIX="test")
@override_config(ENABLE_MAILJET=True)
@freeze_time("2024-01-10 01:01:01")
def test_sanction_list_email(
    mocked_load_workbook: Any,
    mocked_requests_post: Any,
    uploaded_file,
) -> None:
    class MockLoadWorkbook:
        class MockSheet:
            def iter_rows(self, min_row: int):
                return []

            def __getitem__(self, key: int):
                return []

        worksheets = [MockSheet()]

    mocked_load_workbook.return_value = MockLoadWorkbook()

    CheckAgainstSanctionListTask().execute(uploaded_file.id, "test.xlsx")

    # Build expected attachment
    attachment_wb = Workbook()
    attachment_ws = attachment_wb.active
    attachment_ws.title = "Sanction List Check Results"

    headers = (
        "FIRST NAME",
        "SECOND NAME",
        "THIRD NAME",
        "FOURTH NAME",
        "DATE OF BIRTH",
        "ORIGINAL FILE ROW NUMBER",
    )
    attachment_ws.append(headers)

    for i in range(1, len(headers) + 1):
        attachment_ws.column_dimensions[get_column_letter(i)].width = 30

    buffer = io.BytesIO()
    attachment_wb.save(buffer)
    buffer.seek(0)

    base64_encoded_content = base64.b64encode(buffer.getvalue()).decode("utf-8")

    original_file_name = "test.xlsx"
    subject = f"Sanction List Check - file: {original_file_name}, date: {timezone.now().strftime('%Y-%m-%d %I:%M %p')}"

    context = {
        "results": {},
        "results_count": 0,
        "file_name": original_file_name,
        "today_date": timezone.now(),
        "title": "Sanction List Check",
    }

    expected_payload = json.dumps(
        {
            "Messages": [
                {
                    "From": {
                        "Email": settings.DEFAULT_EMAIL,
                        "Name": settings.DEFAULT_EMAIL_DISPLAY,
                    },
                    "Subject": f"[test] {subject}",
                    "To": [{"Email": "test_email@email.com"}],
                    "Cc": [{"Email": settings.SANCTION_LIST_CC_MAIL}],
                    "HTMLPart": render_to_string(
                        "sanction_list/check_results.html",
                        context=context,
                    ),
                    "TextPart": render_to_string(
                        "sanction_list/check_results.txt",
                        context=context,
                    ),
                    "Attachments": [
                        {
                            "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            "Filename": f"{subject}.xlsx",
                            "Base64Content": base64_encoded_content,
                        }
                    ],
                }
            ]
        }
    )

    mocked_requests_post.assert_called_once()
    call_kwargs = mocked_requests_post.call_args
    assert call_kwargs[0][0] == "https://api.mailjet.com/v3.1/send"
    assert call_kwargs[1]["auth"] == (settings.MAILJET_API_KEY, settings.MAILJET_SECRET_KEY)
    assert call_kwargs[1]["timeout"] == 30

    actual_payload = json.loads(call_kwargs[1]["data"])
    expected = json.loads(expected_payload)

    # Compare everything except the binary xlsx attachment content
    actual_attachment = actual_payload["Messages"][0].pop("Attachments")
    expected_attachment = expected["Messages"][0].pop("Attachments")
    assert actual_payload == expected

    # Verify attachment metadata (not binary content — differs across openpyxl versions)
    assert len(actual_attachment) == 1
    assert actual_attachment[0]["ContentType"] == expected_attachment[0]["ContentType"]
    assert actual_attachment[0]["Filename"] == expected_attachment[0]["Filename"]
    assert actual_attachment[0]["Base64Content"]  # non-empty


def test_join_names_and_birthday_db():
    wb = Workbook()
    ws = wb.active

    individual = SanctionListIndividualFactory(
        first_name="FirstName",
        second_name="SecondName",
        third_name="ThirdName",
        fourth_name="FourthName",
    )

    SanctionListIndividualDateOfBirth.objects.create(
        individual=individual,
        date=datetime.date(1981, 1, 1),
    )
    SanctionListIndividualDateOfBirth.objects.create(
        individual=individual,
        date=datetime.date(1980, 2, 1),
    )

    results_dict = {2: individual}

    task = CheckAgainstSanctionListTask()
    task.join_names_and_birthday(ws, results_dict)

    rows = list(ws.iter_rows(values_only=True))

    assert rows[0][4] == "1980-02-01, 1981-01-01"
    assert rows[0][5] == 2


def test_execute_matches_two_name_row_with_date_cell(
    make_uploaded_file: Any,
    john_doe: SanctionListIndividual,
    mocker: Any,
    django_assert_num_queries: Any,
) -> None:
    mailjet_mock = mocker.patch("hope.apps.sanction_list.tasks.check_against_sanction_list.MailjetClient")
    uploaded = make_uploaded_file([("john", "doe", None, None, datetime.date(1980, 1, 1))])

    with django_assert_num_queries(7):
        CheckAgainstSanctionListTask().execute(uploaded.id, "check.xlsx")

    rows = _attachment_rows(mailjet_mock)
    assert len(rows) == 2
    assert rows[1][0] == "John"
    assert rows[1][1] == "Doe"
    assert rows[1][4] == "1980-01-01"
    assert rows[1][5] == 2
    assert mailjet_mock.call_args.kwargs["recipients"] == ["checker@example.com"]
    assert mailjet_mock.call_args.kwargs["ccs"] == [settings.SANCTION_LIST_CC_MAIL]
    mailjet_mock.return_value.send_email.assert_called_once_with()


def test_execute_matches_single_name_row_by_first_name(
    make_uploaded_file: Any,
    john_doe: SanctionListIndividual,
    mocker: Any,
) -> None:
    mailjet_mock = mocker.patch("hope.apps.sanction_list.tasks.check_against_sanction_list.MailjetClient")
    uploaded = make_uploaded_file([("john", None, None, None, None)])

    CheckAgainstSanctionListTask().execute(uploaded.id, "check.xlsx")

    rows = _attachment_rows(mailjet_mock)
    assert len(rows) == 2
    assert rows[1][0] == "John"
    assert rows[1][5] == 2


def test_execute_parses_date_of_birth_given_as_string(
    make_uploaded_file: Any,
    john_doe: SanctionListIndividual,
    mocker: Any,
) -> None:
    mailjet_mock = mocker.patch("hope.apps.sanction_list.tasks.check_against_sanction_list.MailjetClient")
    uploaded = make_uploaded_file([("john", "doe", None, None, "1980-01-01")])

    CheckAgainstSanctionListTask().execute(uploaded.id, "check.xlsx")

    rows = _attachment_rows(mailjet_mock)
    assert len(rows) == 2
    assert rows[1][0] == "John"
    assert rows[1][5] == 2


def test_execute_ignores_unparseable_date_of_birth(
    make_uploaded_file: Any,
    john_doe: SanctionListIndividual,
    mocker: Any,
) -> None:
    mailjet_mock = mocker.patch("hope.apps.sanction_list.tasks.check_against_sanction_list.MailjetClient")
    uploaded = make_uploaded_file([("john", "doe", None, None, "not-a-date")])

    CheckAgainstSanctionListTask().execute(uploaded.id, "check.xlsx")

    rows = _attachment_rows(mailjet_mock)
    assert len(rows) == 2
    assert rows[1][0] == "John"


def test_execute_skips_empty_rows_and_rows_without_names(
    make_uploaded_file: Any,
    john_doe: SanctionListIndividual,
    mocker: Any,
) -> None:
    mailjet_mock = mocker.patch("hope.apps.sanction_list.tasks.check_against_sanction_list.MailjetClient")
    uploaded = make_uploaded_file(
        [
            (None, None, None, None, None),
            (None, None, None, None, "1985-05-05"),
        ]
    )

    CheckAgainstSanctionListTask().execute(uploaded.id, "check.xlsx")

    rows = _attachment_rows(mailjet_mock)
    assert rows == [
        ("FIRST NAME", "SECOND NAME", "THIRD NAME", "FOURTH NAME", "DATE OF BIRTH", "ORIGINAL FILE ROW NUMBER")
    ]
    mailjet_mock.return_value.send_email.assert_called_once_with()


def test_execute_does_not_match_unrelated_names(
    make_uploaded_file: Any,
    john_doe: SanctionListIndividual,
    mocker: Any,
) -> None:
    mailjet_mock = mocker.patch("hope.apps.sanction_list.tasks.check_against_sanction_list.MailjetClient")
    uploaded = make_uploaded_file([("jane", "smith", None, None, None)])

    CheckAgainstSanctionListTask().execute(uploaded.id, "check.xlsx")

    rows = _attachment_rows(mailjet_mock)
    assert len(rows) == 1
