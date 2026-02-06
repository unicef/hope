from dataclasses import dataclass

from django.conf import settings
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
import pytest

from hope.apps.core.flex_fields_importer import FlexibleAttributeImporter


@dataclass
class MockCell:
    """Mock cell object that mimics the interface used by FlexibleAttributeImporter."""

    value: str | None


@pytest.fixture
def flex_importer():
    importer = FlexibleAttributeImporter()
    importer._reset_model_fields_variables()
    return importer


@pytest.fixture
def flex_workbook():
    filename = f"{settings.TESTS_ROOT}/apps/core/test_files/flex_init.xlsx"
    return load_workbook(filename=filename, data_only=True)


@pytest.fixture
def survey_sheet(flex_workbook):
    return flex_workbook["survey"]


@pytest.fixture
def choices_sheet(flex_workbook):
    return flex_workbook["choices"]


@pytest.mark.parametrize(
    ("object_type", "expected_fields"),
    [
        pytest.param(
            "attribute",
            [
                "choices",
                "is_removed",
                "id",
                "created_at",
                "updated_at",
                "name",
                "group",
                "type",
                "associated_with",
                "program",
                "pdu_data",
                "required",
                "label",
                "hint",
            ],
            id="attribute_type",
        ),
        pytest.param(
            "group",
            [
                "flex_attributes",
                "children",
                "id",
                "created_at",
                "updated_at",
                "is_removed",
                "name",
                "label",
                "required",
                "repeatable",
                "parent",
                "lft",
                "rght",
                "tree_id",
                "level",
            ],
            id="group_type",
        ),
        pytest.param(
            "choice",
            [
                "is_removed",
                "id",
                "created_at",
                "updated_at",
                "list_name",
                "name",
                "label",
                "flex_attributes",
            ],
            id="choice_type",
        ),
        pytest.param("Not Correct Arg", None, id="invalid_type"),
    ],
)
def test_get_model_fields_returns_expected_fields(flex_importer, object_type, expected_fields):
    result = flex_importer._get_model_fields(object_type)

    assert result == expected_fields


def test_assign_field_values_attribute_sets_expected_fields(flex_importer, survey_sheet):
    row_idx = 62
    row = survey_sheet[row_idx]

    type_value = survey_sheet.cell(row=row_idx, column=1).value
    name_value = survey_sheet.cell(row=row_idx, column=2).value
    label_value = survey_sheet.cell(row=row_idx, column=3).value
    required_value = survey_sheet.cell(row=row_idx, column=7).value

    flex_importer._assign_field_values(type_value, "type", "attribute", row, 61)
    flex_importer._assign_field_values(name_value, "name", "attribute", row, 61)
    flex_importer._assign_field_values(required_value, "required", "attribute", row, 61)
    flex_importer._assign_field_values(label_value, "label::English(EN)", "attribute", row, 61)
    flex_importer._assign_field_values(label_value, "label:English(EN)", "attribute", row, 61)

    assert flex_importer.object_fields_to_create == {
        "type": "INTEGER",
        "name": "dairy_h_f",
        "required": False,
    }
    assert flex_importer.json_fields_to_create == {"label": {"English(EN)": "Milk and dairy products: yoghurt, cheese"}}


def test_assign_field_values_attribute_empty_label_raises_validation_error(flex_importer, survey_sheet):
    row_idx = 62
    row = survey_sheet[row_idx]

    with pytest.raises(ValidationError, match="Row 61: English label cannot be empty"):
        flex_importer._assign_field_values("", "label::English(EN)", "attribute", row, 61)


def test_assign_field_values_attribute_empty_type_raises_validation_error(flex_importer, survey_sheet):
    row_idx = 62
    row = survey_sheet[row_idx]

    with pytest.raises(ValidationError, match="Row 61: Type is required"):
        flex_importer._assign_field_values("", "type", "attribute", row, 61)


def test_assign_field_values_attribute_empty_name_raises_validation_error(flex_importer, survey_sheet):
    row_idx = 62
    row = survey_sheet[row_idx]

    with pytest.raises(ValidationError, match="Row 61: Name is required"):
        flex_importer._assign_field_values("", "name", "attribute", row, 61)


def test_assign_field_values_group_sets_expected_fields(flex_importer, survey_sheet):
    row_idx = 5
    row = survey_sheet[row_idx]
    name_value = row[1].value
    required_value = row[6].value
    label_value = row[2].value

    flex_importer._assign_field_values(name_value, "name", "group", row, 4)
    flex_importer._assign_field_values(required_value, "required", "group", row, 4)
    flex_importer._assign_field_values(label_value, "label::English(EN)", "group", row, 4)

    assert flex_importer.object_fields_to_create == {
        "name": "consent",
        "required": False,
    }
    assert flex_importer.json_fields_to_create == {"label": {"English(EN)": "Consent"}}


def test_assign_field_values_group_empty_name_raises_validation_error(flex_importer, survey_sheet):
    row_idx = 5
    row = survey_sheet[row_idx]

    with pytest.raises(ValidationError, match="Row 61: Name is required"):
        flex_importer._assign_field_values("", "name", "group", row, 61)


def test_assign_field_values_choice_sets_expected_fields(flex_importer, choices_sheet):
    row_idx = 2
    row = choices_sheet[row_idx]
    list_name_value = choices_sheet.cell(row=row_idx, column=1).value
    name_value = choices_sheet.cell(row=row_idx, column=2).value
    label_value = choices_sheet.cell(row=row_idx, column=3).value

    flex_importer._assign_field_values(list_name_value, "list_name", "choice", row, 1)
    flex_importer._assign_field_values(name_value, "name", "choice", row, 1)
    flex_importer._assign_field_values(label_value, "label::English(EN)", "choice", row, 1)

    assert flex_importer.object_fields_to_create == {
        "list_name": "yes_no",
        "name": "1",
    }
    assert flex_importer.json_fields_to_create == {"label": {"English(EN)": "Yes"}}


def test_assign_field_values_choice_empty_label_raises_validation_error(flex_importer, choices_sheet):
    row_idx = 2
    row = choices_sheet[row_idx]

    with pytest.raises(ValidationError, match="Row 1: English label cannot be empty"):
        flex_importer._assign_field_values("", "label::English(EN)", "choice", row, 1)


def test_assign_field_values_choice_empty_list_name_raises_validation_error(flex_importer, choices_sheet):
    row_idx = 2
    row = choices_sheet[row_idx]

    with pytest.raises(ValidationError, match="Row 1: List Name is required"):
        flex_importer._assign_field_values("", "list_name", "choice", row, 1)


def test_assign_field_values_choice_empty_name_raises_validation_error(flex_importer, choices_sheet):
    row_idx = 2
    row = choices_sheet[row_idx]

    with pytest.raises(ValidationError, match="Row 1: Name is required"):
        flex_importer._assign_field_values("", "name", "choice", row, 1)


@pytest.mark.parametrize(
    ("row_data", "expected"),
    [
        pytest.param([MockCell("text"), MockCell("test_h_c")], False, id="core_field_h_c_suffix"),
        pytest.param([MockCell("text"), MockCell("test_i_c")], False, id="core_field_i_c_suffix"),
        pytest.param([MockCell("start"), MockCell("start")], False, id="excluded_start_field"),
        pytest.param([MockCell("end"), MockCell("end")], False, id="excluded_end_field"),
        pytest.param([MockCell("deviceid"), MockCell("deviceid")], False, id="excluded_deviceid_field"),
        pytest.param([MockCell("end_repeat"), MockCell("")], False, id="end_repeat_marker"),
        pytest.param([MockCell("end_group"), MockCell("")], False, id="end_group_marker"),
        pytest.param([MockCell("begin_group"), MockCell("test_group")], True, id="begin_group_allowed"),
    ],
)
def test_can_add_row_returns_expected_result(flex_importer, row_data, expected):
    flex_importer.current_group_tree = [None]

    result = flex_importer._can_add_row(row_data)

    assert result == expected


def test_get_list_of_field_choices_returns_expected_set(flex_importer, survey_sheet):
    result = flex_importer._get_list_of_field_choices(survey_sheet)

    expected = {
        "sex",
        "severity_of_disability",
        "marital_status",
        "id_type",
        "latrine",
        "disability",
        "assistance_type",
        "sufficient_water",
        "nationality",
        "child_marital_status",
        "assistance_source",
        "school_type",
        "status_head_of_hh",
        "admin2",
        "treatment_facility",
        "water_source",
        "admin1",
        "living_situation",
        "yes_no",
        "residence_status",
    }
    assert result == expected


@pytest.mark.parametrize(
    ("row_data", "expected"),
    [
        pytest.param([MockCell("text"), MockCell("first_name")], None, id="non_select_returns_none"),
        pytest.param(
            [MockCell("select_one test_group"), MockCell("test_group")],
            "test_group",
            id="select_one_returns_choice_name",
        ),
    ],
)
def test_get_field_choice_name_returns_expected_result(flex_importer, row_data, expected):
    result = flex_importer._get_field_choice_name(row_data)

    assert result == expected
