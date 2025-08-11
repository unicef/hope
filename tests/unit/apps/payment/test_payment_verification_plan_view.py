from io import BytesIO
from pathlib import Path
from typing import Any, List, Optional

from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from django.utils import timezone

import pytest
from extras.test_utils.factories.account import PartnerFactory, UserFactory
from extras.test_utils.factories.core import create_afghanistan
from extras.test_utils.factories.payment import (
    PaymentFactory,
    PaymentPlanFactory,
    PaymentVerificationFactory,
    PaymentVerificationPlanFactory,
)
from extras.test_utils.factories.program import ProgramFactory
from openpyxl import Workbook
from rest_framework import status

from hope.apps.account.permissions import Permissions
from hope.apps.payment.models import (
    Payment,
    PaymentPlan,
    PaymentVerification,
    PaymentVerificationPlan,
    PaymentVerificationSummary,
    build_summary,
)
from hope.apps.program.models import Program

pytestmark = pytest.mark.django_db


def generate_valid_xlsx_file(
    name: str = "unit_test.xlsx", worksheet_title_list: Optional[List[str]] = None
) -> SimpleUploadedFile:
    if worksheet_title_list is None:
        worksheet_title_list = ["Test"]
    wb = Workbook()
    for worksheet_title in worksheet_title_list:
        wb.create_sheet(title=worksheet_title)
    ws = wb.active
    ws["A1"] = "People"
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return SimpleUploadedFile(
        name, file_stream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


class TestPaymentVerificationViewSet:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()
        self.client = api_client(self.user)
        self.pp = PaymentPlanFactory(
            name="Payment Plan",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.FINISHED,
            created_by=self.user,
            created_at="2022-02-24",
        )
        PaymentVerificationSummary.objects.create(payment_plan=self.pp)
        build_summary(self.pp)
        self.payment_1 = PaymentFactory(
            pk="0329a41f-affd-4669-9e38-38ec2d6699b3",
            parent=self.pp,
            status=Payment.STATUS_SUCCESS,
            delivered_quantity=999,
            entitlement_quantity=112,
        )
        self.payment_2 = PaymentFactory(
            pk="299811ef-b123-427d-b77d-9fd5d1bc8946",
            parent=self.pp,
            status=Payment.STATUS_SUCCESS,
            delivered_quantity=111,
            entitlement_quantity=112,
        )
        self.pvp = PaymentVerificationPlanFactory(
            payment_plan=self.pp,
            sampling=PaymentVerificationPlan.SAMPLING_RANDOM,
            verification_channel=PaymentVerificationPlan.VERIFICATION_CHANNEL_MANUAL,
        )
        self.verification_1 = PaymentVerificationFactory(
            payment_verification_plan=self.pvp,
            payment=self.payment_1,
            status=PaymentVerification.STATUS_RECEIVED_WITH_ISSUES,
            status_date=timezone.now(),
        )
        self.verification_2 = PaymentVerificationFactory(
            payment_verification_plan=self.pvp,
            payment=self.payment_2,
            status=PaymentVerification.STATUS_RECEIVED_WITH_ISSUES,
        )
        url_kwargs = {
            "business_area_slug": self.afghanistan.slug,
            "program_slug": self.program_active.slug,
            "pk": str(self.pp.pk),
        }
        url_kwargs_id = {
            "business_area_slug": self.afghanistan.slug,
            "program_slug": self.program_active.slug,
            "pk": str(self.pp.pk),
            "verification_plan_id": str(self.pvp.pk),
        }
        self.url_list = reverse(
            "api:payments:payment-verifications-list",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
            },
        )
        self.url_details = reverse("api:payments:payment-verifications-detail", kwargs=url_kwargs)
        self.url_create = reverse(
            "api:payments:payment-verifications-create-payment-verification-plan", kwargs=url_kwargs
        )
        self.url_update = reverse(
            "api:payments:payment-verifications-update-payment-verification-plan", kwargs=url_kwargs_id
        )
        self.url_activate = reverse(
            "api:payments:payment-verifications-activate-payment-verification-plan", kwargs=url_kwargs_id
        )
        self.url_finish = reverse(
            "api:payments:payment-verifications-finish-payment-verification-plan", kwargs=url_kwargs_id
        )
        self.url_discard = reverse(
            "api:payments:payment-verifications-discard-payment-verification-plan", kwargs=url_kwargs_id
        )
        self.url_invalid = reverse(
            "api:payments:payment-verifications-invalid-payment-verification-plan", kwargs=url_kwargs_id
        )
        self.url_delete = reverse(
            "api:payments:payment-verifications-delete-payment-verification-plan", kwargs=url_kwargs_id
        )
        self.url_export_xlsx = reverse(
            "api:payments:payment-verifications-export-xlsx-payment-verification-plan", kwargs=url_kwargs_id
        )
        self.url_import_xlsx = reverse(
            "api:payments:payment-verifications-import-xlsx-payment-verification-plan", kwargs=url_kwargs_id
        )

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_VIEW_LIST], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_get_list(self, permissions: List, expected_status: int, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        response = self.client.get(self.url_list)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert len(resp_data["results"]) == 1
            pv = resp_data["results"][0]
            assert pv["verification_status"] == "PENDING"

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_VIEW_DETAILS], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_details(self, permissions: List, expected_status: int, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        response = self.client.get(self.url_details)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert resp_data["available_payment_records_count"] == 2
            assert resp_data["eligible_payments_count"] == 2
            assert resp_data["payment_verification_plans"][0]["status"] == PaymentVerificationPlan.STATUS_PENDING
            assert resp_data["payment_verification_plans"][0]["sampling"] == "Random sampling"
            assert (
                resp_data["payment_verification_plans"][0]["verification_channel"]
                == PaymentVerificationPlan.VERIFICATION_CHANNEL_MANUAL
            )
            assert resp_data["payment_verification_summary"]["status"] == "PENDING"

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_CREATE], status.HTTP_201_CREATED),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_create_pvp(self, permissions: List, expected_status: int, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        PaymentFactory(parent=self.pp, status=Payment.STATUS_SUCCESS, delivered_quantity=111, entitlement_quantity=112)
        response = self.client.post(
            self.url_create,
            {
                "sampling": "FULL_LIST",
                "full_list_arguments": {"excluded_admin_areas": []},
                "verification_channel": PaymentVerificationPlan.VERIFICATION_CHANNEL_XLSX,
                "rapid_pro_arguments": None,
                "random_sampling_arguments": None,
            },
            format="json",
        )
        assert response.status_code == expected_status
        if expected_status == status.HTTP_201_CREATED:
            assert response.status_code == status.HTTP_201_CREATED
            resp_data = response.json()
            assert "id" in resp_data
            assert len(resp_data["payment_verification_plans"]) == 2
            pvp = resp_data["payment_verification_plans"][1]
            assert pvp["status"] == PaymentVerificationPlan.STATUS_PENDING
            assert pvp["verification_channel"] == PaymentVerificationPlan.VERIFICATION_CHANNEL_XLSX
            assert pvp["sampling"] == "Full list"
            assert pvp["excluded_admin_areas_filter"] == []
            assert resp_data["payment_verification_summary"]["status"] == "PENDING"

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_UPDATE], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_update_pvp(self, permissions: List, expected_status: int, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        response = self.client.patch(
            self.url_update,
            {
                "sampling": "FULL_LIST",
                "full_list_arguments": {"excluded_admin_areas": []},
                "verification_channel": "MANUAL",
                "rapid_pro_arguments": None,
                "random_sampling_arguments": None,
            },
            format="json",
        )
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert len(resp_data["payment_verification_plans"]) == 1
            pvp = resp_data["payment_verification_plans"][0]
            assert pvp["status"] == PaymentVerificationPlan.STATUS_PENDING
            assert pvp["verification_channel"] == "MANUAL"
            assert pvp["sampling"] == "Full list"
            assert pvp["excluded_admin_areas_filter"] == []
            assert resp_data["payment_verification_summary"]["status"] == "PENDING"

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_ACTIVATE], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pvp_activate(
        self, permissions: List, expected_status: int, create_user_role_with_permissions: Any
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        response = self.client.post(self.url_activate, {"version": self.pvp.version}, format="json")
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert len(resp_data["payment_verification_plans"]) == 1
            assert resp_data["payment_verification_plans"][0]["status"] == PaymentVerificationPlan.STATUS_ACTIVE

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_FINISH], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pvp_finish(self, permissions: List, expected_status: int, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pvp.status = PaymentVerificationPlan.STATUS_ACTIVE
        self.pvp.save()
        response = self.client.post(self.url_finish, {"version": self.pvp.version}, format="json")
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert len(resp_data["payment_verification_plans"]) == 1
            assert resp_data["payment_verification_plans"][0]["status"] == PaymentVerificationPlan.STATUS_FINISHED

    def test_pvp_finish_validation_error(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user, [Permissions.PAYMENT_VERIFICATION_FINISH], self.afghanistan, self.program_active
        )
        response = self.client.post(self.url_finish, {"version": self.pvp.version}, format="json")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "You can finish only ACTIVE verification" in response.json()

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_DISCARD], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pvp_discard(self, permissions: List, expected_status: int, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pvp.status = PaymentVerificationPlan.STATUS_ACTIVE
        self.pvp.save()
        response = self.client.post(self.url_discard, {"version": self.pvp.version}, format="json")
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert len(resp_data["payment_verification_plans"]) == 1
            assert resp_data["payment_verification_plans"][0]["status"] == PaymentVerificationPlan.STATUS_PENDING

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_INVALID], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pvp_invalid(self, permissions: List, expected_status: int, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pvp.status = PaymentVerificationPlan.STATUS_ACTIVE
        self.pvp.verification_channel = PaymentVerificationPlan.VERIFICATION_CHANNEL_XLSX
        self.pvp.xlsx_file_imported = True
        self.pvp.save()
        response = self.client.post(self.url_invalid, {"version": self.pvp.version}, format="json")
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert len(resp_data["payment_verification_plans"]) == 1
            assert resp_data["payment_verification_plans"][0]["status"] == PaymentVerificationPlan.STATUS_INVALID

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_DELETE], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pvp_delete(self, permissions: List, expected_status: int, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        response = self.client.post(self.url_delete, {"version": self.pvp.version}, format="json")
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert len(resp_data["payment_verification_plans"]) == 0

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_EXPORT], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pvp_export_xlsx(
        self, permissions: List, expected_status: int, create_user_role_with_permissions: Any
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pvp.status = PaymentVerificationPlan.STATUS_ACTIVE
        self.pvp.verification_channel = PaymentVerificationPlan.VERIFICATION_CHANNEL_XLSX
        self.pvp.save()
        response = self.client.post(self.url_export_xlsx, {"version": self.pvp.version}, format="json")
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert len(resp_data["payment_verification_plans"]) == 1
            self.pvp.refresh_from_db()
            if not self.pvp.xlsx_file_exporting:
                assert self.pvp.has_xlsx_payment_verification_plan_file is True
            else:
                assert resp_data["payment_verification_plans"][0]["xlsx_file_exporting"] is True

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_IMPORT], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pvp_import_xlsx(
        self, permissions: List, expected_status: int, create_user_role_with_permissions: Any
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pvp.status = PaymentVerificationPlan.STATUS_ACTIVE
        self.pvp.verification_channel = PaymentVerificationPlan.VERIFICATION_CHANNEL_XLSX
        self.pvp.save()
        # file = generate_valid_xlsx_file(worksheet_title_list=["Payment Verifications", "Meta"])
        file = BytesIO(Path(f"{settings.TESTS_ROOT}/apps/payment/test_file/unordered_columns_1.xlsx").read_bytes())
        file.name = "unordered_columns_1.xlsx"
        response = self.client.post(
            self.url_import_xlsx, {"version": self.pvp.version, "file": file}, format="multipart"
        )
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert len(resp_data["payment_verification_plans"]) == 1
            assert resp_data["payment_verification_plans"][0]["xlsx_file_imported"] is True

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_VIEW_DETAILS], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_verifications_list(
        self, permissions: List, expected_status: int, create_user_role_with_permissions: Any
    ) -> None:
        url = reverse(
            "api:payments:verification-records-list",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
                "payment_verification_pk": str(self.pp.pk),
            },
        )
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        response = self.client.get(url)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()["results"]
            assert len(resp_data) == 2
            payment = resp_data[0]
            assert "id" in payment
            assert "verification" in payment
            assert "id" in payment["verification"]
            assert "status" in payment["verification"]
            assert "verification_channel" in payment["verification"]
            assert "received_amount" in payment["verification"]

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_VIEW_DETAILS], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_verification_details(
        self, permissions: List, expected_status: int, create_user_role_with_permissions: Any
    ) -> None:
        url = reverse(
            "api:payments:verification-records-detail",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
                "payment_verification_pk": str(self.pp.pk),
                "pk": str(self.payment_1.pk),
            },
        )
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        response = self.client.get(url)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert "verification" in resp_data
            assert "id" in resp_data["verification"]
            assert "status" in resp_data["verification"]
            assert "verification_channel" in resp_data["verification"]
            assert "received_amount" in resp_data["verification"]

    @pytest.mark.parametrize(
        "permissions, expected_status",
        [
            ([Permissions.PAYMENT_VERIFICATION_VERIFY, Permissions.PAYMENT_VERIFICATION_VIEW_LIST], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_update_verification(
        self, permissions: List, expected_status: int, create_user_role_with_permissions: Any
    ) -> None:
        url = reverse(
            "api:payments:verification-records-detail",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
                "payment_verification_pk": str(self.pp.id),
                "pk": str(self.payment_1.pk),
            },
        )
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pvp.status = PaymentVerificationPlan.STATUS_ACTIVE
        self.pvp.verification_channel = PaymentVerificationPlan.VERIFICATION_CHANNEL_MANUAL
        self.pvp.save()

        response = self.client.patch(
            url,
            {"version": self.verification_1.version, "received_amount": 123.22, "received": True},
            format="multipart",
        )
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()

            assert "id" in resp_data
            assert resp_data["verification"]["received_amount"] == "123.22"
