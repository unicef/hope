import datetime
from decimal import Decimal
from tempfile import NamedTemporaryFile
from types import SimpleNamespace
from unittest.mock import patch

from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
import openpyxl
import pytest

from extras.test_utils.factories import (
    BusinessAreaFactory,
    CurrencyFactory,
    DeliveryMechanismFactory,
    FileTempFactory,
    FinancialServiceProviderFactory,
    FinancialServiceProviderXlsxTemplateFactory,
    FspXlsxTemplatePerDeliveryMechanismFactory,
    HouseholdFactory,
    PaymentFactory,
    PaymentHouseholdSnapshotFactory,
    PaymentPlanFactory,
    PaymentVerificationFactory,
    PaymentVerificationPlanFactory,
    PaymentVerificationSummaryFactory,
    ProgramCycleFactory,
    ProgramFactory,
    UserFactory,
)
from hope.apps.payment.celery_tasks import import_follow_up_instruction_reconciliation_from_xlsx_async_task_action
from hope.apps.payment.xlsx.xlsx_follow_up_instruction_delivery_export_service import (
    XlsxFollowUpInstructionDeliveryExportService,
)
from hope.apps.payment.xlsx.xlsx_follow_up_instruction_reconciliation_import_service import (
    XlsxFollowUpInstructionReconciliationImportService,
)
from hope.models import FollowUpInstruction, Payment, PaymentPlan, PaymentVerification

pytestmark = pytest.mark.django_db


@pytest.fixture
def business_area():
    return BusinessAreaFactory(slug="afghanistan")


@pytest.fixture
def user():
    return UserFactory()


@pytest.fixture
def program(business_area):
    return ProgramFactory(business_area=business_area)


@pytest.fixture
def cycle(program):
    return ProgramCycleFactory(program=program)


@pytest.fixture
def currency():
    return CurrencyFactory(code="AFN")


@pytest.fixture
def delivery_mechanism():
    return DeliveryMechanismFactory(code="cash", name="Cash", payment_gateway_id="dm-cash")


@pytest.fixture
def fsp(delivery_mechanism):
    fsp = FinancialServiceProviderFactory()
    fsp.delivery_mechanisms.add(delivery_mechanism)
    return fsp


@pytest.fixture
def delivery_template(fsp, delivery_mechanism):
    template = FinancialServiceProviderXlsxTemplateFactory(
        columns=["payment_id", "entitlement_quantity", "delivered_quantity"],
        core_fields=[],
        flex_fields=[],
        document_types=[],
    )
    FspXlsxTemplatePerDeliveryMechanismFactory(
        financial_service_provider=fsp,
        delivery_mechanism=delivery_mechanism,
        xlsx_template=template,
    )
    return template


@pytest.fixture
def instruction(user, program, business_area):
    return FollowUpInstruction.objects.create(
        business_area=business_area,
        program=program,
        created_by=user,
        background_action_status=FollowUpInstruction.BackgroundActionStatus.XLSX_IMPORTING_RECONCILIATION,
    )


@pytest.fixture
def child_payment_plans(instruction, cycle, business_area, currency, delivery_mechanism, fsp):
    plan_one = PaymentPlanFactory(
        program_cycle=cycle,
        business_area=business_area,
        currency=currency,
        delivery_mechanism=delivery_mechanism,
        financial_service_provider=fsp,
        follow_up_instruction=instruction,
        plan_type=PaymentPlan.PlanType.FOLLOW_UP,
        status=PaymentPlan.Status.ACCEPTED,
        exchange_rate=Decimal("1.00"),
    )
    plan_two = PaymentPlanFactory(
        program_cycle=cycle,
        business_area=business_area,
        currency=currency,
        delivery_mechanism=delivery_mechanism,
        financial_service_provider=fsp,
        follow_up_instruction=instruction,
        plan_type=PaymentPlan.PlanType.FOLLOW_UP,
        status=PaymentPlan.Status.ACCEPTED,
        exchange_rate=Decimal("1.00"),
    )
    return plan_one, plan_two


@pytest.fixture
def instruction_payments(child_payment_plans, fsp, delivery_mechanism, currency):
    plan_one, plan_two = child_payment_plans
    shared_household = HouseholdFactory(business_area=plan_one.business_area, program=plan_one.program)
    payment_one = PaymentFactory(
        parent=plan_one,
        household=shared_household,
        collector=shared_household.head_of_household,
        head_of_household=shared_household.head_of_household,
        program=plan_one.program,
        currency=currency,
        delivery_type=delivery_mechanism,
        financial_service_provider=fsp,
        entitlement_quantity=Decimal("100.00"),
        entitlement_quantity_usd=Decimal("100.00"),
        delivered_quantity=Decimal("0.00"),
        status=Payment.STATUS_PENDING,
    )
    payment_two = PaymentFactory(
        parent=plan_two,
        household=shared_household,
        collector=shared_household.head_of_household,
        head_of_household=shared_household.head_of_household,
        program=plan_two.program,
        currency=currency,
        delivery_type=delivery_mechanism,
        financial_service_provider=fsp,
        entitlement_quantity=Decimal("40.00"),
        entitlement_quantity_usd=Decimal("40.00"),
        delivered_quantity=Decimal("0.00"),
        status=Payment.STATUS_PENDING,
    )
    second_household = HouseholdFactory(business_area=plan_two.business_area, program=plan_two.program)
    payment_three = PaymentFactory(
        parent=plan_two,
        household=second_household,
        collector=second_household.head_of_household,
        head_of_household=second_household.head_of_household,
        program=plan_two.program,
        currency=currency,
        delivery_type=delivery_mechanism,
        financial_service_provider=fsp,
        entitlement_quantity=Decimal("60.00"),
        entitlement_quantity_usd=Decimal("60.00"),
        delivered_quantity=Decimal("0.00"),
        status=Payment.STATUS_PENDING,
    )
    for payment in (payment_one, payment_two, payment_three):
        PaymentHouseholdSnapshotFactory(
            payment=payment,
            snapshot_data={
                "unicef_id": payment.household.unicef_id,
                "size": payment.household.size,
                "primary_collector": {
                    "unicef_id": payment.collector.unicef_id,
                    "full_name": payment.collector.full_name,
                },
                "alternate_collector": {},
            },
        )
    return payment_one, payment_two, payment_three


def _build_reconciliation_file(instruction: FollowUpInstruction) -> bytes:
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    household_col = headers.index("household_id") + 1
    delivered_col = headers.index("delivered_quantity") + 1
    values_by_household = {}
    for row_idx in range(2, worksheet.max_row + 1):
        household_id = str(worksheet.cell(row=row_idx, column=household_col).value)
        values_by_household[household_id] = row_idx

    household_ids = list(values_by_household.keys())
    worksheet.cell(row=values_by_household[household_ids[0]], column=delivered_col).value = "120.00"
    worksheet.cell(row=values_by_household[household_ids[1]], column=delivered_col).value = "0.00"

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        return tmp.read()


def test_reconciliation_import_happy_path_updates_child_payments_and_removes_stale_export_file(
    user,
    instruction,
    child_payment_plans,
    instruction_payments,
    delivery_template,
    django_capture_on_commit_callbacks,
):
    stale_export_file = FileTempFactory(file=ContentFile(b"stale export", name="stale-export.xlsx"))
    instruction.export_file = stale_export_file
    instruction.reconciliation_import_file = FileTempFactory(
        object_id=instruction.pk,
        created_by=user,
        file=ContentFile(_build_reconciliation_file(instruction), name="instruction-reconciliation.xlsx"),
    )
    instruction.save(
        update_fields=["export_file", "reconciliation_import_file", "background_action_status", "updated_at"]
    )

    job = SimpleNamespace(config={"follow_up_instruction_id": str(instruction.id)})

    with patch(
        "hope.apps.payment.services.payment_plan_services.PaymentPlanService.recalculate_signatures_in_batch"
    ) as recalculate_mock:
        with django_capture_on_commit_callbacks(execute=True):
            result = import_follow_up_instruction_reconciliation_from_xlsx_async_task_action(job)

    assert result is True
    instruction.refresh_from_db()
    for payment in instruction_payments:
        payment.refresh_from_db()
    for payment_plan in child_payment_plans:
        payment_plan.refresh_from_db()

    assert instruction.background_action_status is None
    assert instruction.export_file_id is None
    assert recalculate_mock.call_count == 2

    payment_one, payment_two, payment_three = instruction_payments
    assert payment_one.delivered_quantity == Decimal("100.00")
    assert payment_one.status == Payment.STATUS_DISTRIBUTION_SUCCESS
    assert payment_two.delivered_quantity == Decimal("20.00")
    assert payment_two.status == Payment.STATUS_DISTRIBUTION_PARTIAL
    assert payment_three.delivered_quantity == Decimal("0.00")
    assert payment_three.status == Payment.STATUS_NOT_DISTRIBUTED
    assert all(payment_plan.status == PaymentPlan.Status.FINISHED for payment_plan in child_payment_plans)


def test_reconciliation_import_validation_rejects_negative_household_total(
    instruction,
    instruction_payments,
    delivery_template,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    delivered_col = headers.index("delivered_quantity") + 1
    worksheet.cell(row=2, column=delivered_col).value = "-1"

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction,
            ContentFile(tmp.read(), name="negative-reconciliation.xlsx"),
        )

    service.open_workbook()
    service.validate()

    assert len(service.errors) == 1
    assert "cannot be below zero" in service.errors[0].message.lower()


def test_reconciliation_import_marks_xlsx_import_error_on_service_validation_errors(
    instruction,
    instruction_payments,
    delivery_template,
    user,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    delivered_col = headers.index("delivered_quantity") + 1
    worksheet.cell(row=2, column=delivered_col).value = "-1"

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        file_content = tmp.read()

    instruction.reconciliation_import_file = FileTempFactory(
        object_id=instruction.pk,
        created_by=user,
        file=ContentFile(file_content, name="invalid-reconciliation.xlsx"),
    )
    instruction.save(update_fields=["reconciliation_import_file", "updated_at"])

    job = SimpleNamespace(config={"follow_up_instruction_id": str(instruction.id)})

    with pytest.raises(ValidationError):
        import_follow_up_instruction_reconciliation_from_xlsx_async_task_action(job)

    instruction.refresh_from_db(fields=["background_action_status"])
    assert instruction.background_action_status == FollowUpInstruction.BackgroundActionStatus.XLSX_IMPORT_ERROR


def test_reconciliation_import_skips_export_file_cleanup_when_no_export_file(
    instruction,
    instruction_payments,
    child_payment_plans,
    delivery_template,
    user,
    django_capture_on_commit_callbacks,
):
    instruction.reconciliation_import_file = FileTempFactory(
        object_id=instruction.pk,
        created_by=user,
        file=ContentFile(_build_reconciliation_file(instruction), name="instruction-reconciliation.xlsx"),
    )
    instruction.save(update_fields=["reconciliation_import_file", "updated_at"])

    job = SimpleNamespace(config={"follow_up_instruction_id": str(instruction.id)})

    with patch("hope.apps.payment.services.payment_plan_services.PaymentPlanService.recalculate_signatures_in_batch"):
        with django_capture_on_commit_callbacks(execute=True):
            result = import_follow_up_instruction_reconciliation_from_xlsx_async_task_action(job)

    assert result is True
    instruction.refresh_from_db()
    assert instruction.export_file_id is None


def test_reconciliation_import_skips_export_file_cleanup_when_file_was_replaced(
    instruction,
    instruction_payments,
    child_payment_plans,
    delivery_template,
    user,
    django_capture_on_commit_callbacks,
):
    original_export_file = FileTempFactory(file=ContentFile(b"original export", name="original-export.xlsx"))
    replacement_export_file = FileTempFactory(file=ContentFile(b"new export", name="new-export.xlsx"))
    instruction.export_file = original_export_file
    instruction.reconciliation_import_file = FileTempFactory(
        object_id=instruction.pk,
        created_by=user,
        file=ContentFile(_build_reconciliation_file(instruction), name="instruction-reconciliation.xlsx"),
    )
    instruction.save(update_fields=["export_file", "reconciliation_import_file", "updated_at"])

    job = SimpleNamespace(config={"follow_up_instruction_id": str(instruction.id)})

    with patch("hope.apps.payment.services.payment_plan_services.PaymentPlanService.recalculate_signatures_in_batch"):
        with django_capture_on_commit_callbacks(execute=True):
            result = import_follow_up_instruction_reconciliation_from_xlsx_async_task_action(job)
            instruction.export_file = replacement_export_file
            instruction.save(update_fields=["export_file", "updated_at"])

    assert result is True
    instruction.refresh_from_db()
    assert instruction.export_file_id == replacement_export_file.pk


def test_validation_reports_error_for_missing_required_column(instruction):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "wrong_column"
    ws.cell(row=1, column=2).value = "delivered_quantity"

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="missing-header.xlsx")
        )

    service.open_workbook()
    service.validate()

    assert len(service.errors) == 1
    assert "required headers" in service.errors[0].message.lower()


def test_validation_reports_error_for_empty_household_id(
    instruction,
    instruction_payments,
    delivery_template,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    worksheet.cell(row=2, column=headers.index("household_id") + 1).value = ""

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="empty-household.xlsx")
        )

    service.open_workbook()
    service.validate()

    assert any("Household ID is required" in e.message for e in service.errors)


def test_validation_reports_error_for_unknown_household_id(
    instruction,
    instruction_payments,
    delivery_template,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    worksheet.cell(row=2, column=headers.index("household_id") + 1).value = "HH-UNKNOWN-9999"

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="unknown-household.xlsx")
        )

    service.open_workbook()
    service.validate()

    assert any("is not part of this Follow Up Instruction" in e.message for e in service.errors)


def test_validation_reports_error_for_duplicate_household_id(
    instruction,
    instruction_payments,
    delivery_template,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    household_col = headers.index("household_id") + 1
    first_household_id = worksheet.cell(row=2, column=household_col).value
    worksheet.cell(row=3, column=household_col).value = first_household_id

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="duplicate-household.xlsx")
        )

    service.open_workbook()
    service.validate()

    assert any("appears multiple times" in e.message for e in service.errors)


def test_validation_reports_error_for_date_as_delivered_quantity(
    instruction,
    instruction_payments,
    delivery_template,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    worksheet.cell(row=2, column=headers.index("delivered_quantity") + 1).value = datetime.date(2024, 1, 1)

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="date-quantity.xlsx")
        )

    service.open_workbook()
    service.validate()

    assert any("is not a valid number" in e.message for e in service.errors)


def test_validation_reports_error_when_delivered_exceeds_entitlement(
    instruction,
    instruction_payments,
    delivery_template,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    worksheet.cell(row=2, column=headers.index("delivered_quantity") + 1).value = "99999.00"

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="excess-quantity.xlsx")
        )

    service.open_workbook()
    service.validate()

    assert any("bigger than" in e.message.lower() for e in service.errors)


def test_validation_skips_row_when_delivered_quantity_is_empty(
    instruction,
    instruction_payments,
    delivery_template,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    delivered_col = headers.index("delivered_quantity") + 1
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_idx, column=delivered_col).value = None

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="empty-quantities.xlsx")
        )

    service.open_workbook()
    service.validate()

    assert service.household_updates == {}
    assert any("aren't any updates" in e.message for e in service.errors)


def test_validation_skips_blank_rows(
    instruction,
    instruction_payments,
    delivery_template,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    worksheet.cell(row=2, column=headers.index("delivered_quantity") + 1).value = "10.00"
    worksheet.insert_rows(2)

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="blank-row.xlsx")
        )

    service.open_workbook()
    service.validate()

    assert not service.errors


def test_validation_reports_error_when_no_changes_in_file(
    instruction,
    instruction_payments,
    delivery_template,
):
    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    delivered_col = headers.index("delivered_quantity") + 1
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_idx, column=delivered_col).value = "0"

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="no-changes.xlsx")
        )

    service.open_workbook()
    service.validate()

    assert any("aren't any updates" in e.message for e in service.errors)


def test_import_updates_payment_verification_to_received_when_full_amount_delivered(
    instruction,
    instruction_payments,
    child_payment_plans,
    delivery_template,
):
    payment_one = instruction_payments[0]
    plan_one = child_payment_plans[0]
    PaymentVerificationSummaryFactory(payment_plan=plan_one)
    pvp = PaymentVerificationPlanFactory(payment_plan=plan_one)
    pv = PaymentVerificationFactory(
        payment=payment_one,
        payment_verification_plan=pvp,
        status=PaymentVerification.STATUS_NOT_RECEIVED,
        received_amount=payment_one.entitlement_quantity,
    )

    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    household_col = headers.index("household_id") + 1
    delivered_col = headers.index("delivered_quantity") + 1
    for row_idx in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row=row_idx, column=household_col).value) == payment_one.household.unicef_id:
            worksheet.cell(row=row_idx, column=delivered_col).value = str(payment_one.entitlement_quantity)

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="full-delivery.xlsx")
        )

    service.open_workbook()
    service.validate()
    assert not service.errors
    service.import_payment_list()

    pv.refresh_from_db()
    assert pv.status == PaymentVerification.STATUS_RECEIVED


def test_import_updates_payment_verification_to_not_received_when_zero_delivered(
    instruction,
    instruction_payments,
    child_payment_plans,
    delivery_template,
):
    payment_three = instruction_payments[2]
    plan_two = child_payment_plans[1]
    payment_three.delivered_quantity = payment_three.entitlement_quantity
    payment_three.save(update_fields=["delivered_quantity"])
    PaymentVerificationSummaryFactory(payment_plan=plan_two)
    pvp = PaymentVerificationPlanFactory(payment_plan=plan_two)
    pv = PaymentVerificationFactory(
        payment=payment_three,
        payment_verification_plan=pvp,
        status=PaymentVerification.STATUS_RECEIVED,
        received_amount=payment_three.entitlement_quantity,
    )

    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    household_col = headers.index("household_id") + 1
    delivered_col = headers.index("delivered_quantity") + 1
    for row_idx in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row=row_idx, column=household_col).value) == payment_three.household.unicef_id:
            worksheet.cell(row=row_idx, column=delivered_col).value = "0"

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="zero-delivery.xlsx")
        )

    service.open_workbook()
    service.validate()
    assert not service.errors
    service.import_payment_list()

    pv.refresh_from_db()
    assert pv.status == PaymentVerification.STATUS_NOT_RECEIVED


def test_import_updates_payment_verification_to_received_with_issues_on_partial_delivery(
    instruction,
    instruction_payments,
    child_payment_plans,
    delivery_template,
):
    payment_one = instruction_payments[0]
    plan_one = child_payment_plans[0]
    PaymentVerificationSummaryFactory(payment_plan=plan_one)
    pvp = PaymentVerificationPlanFactory(payment_plan=plan_one)
    pv = PaymentVerificationFactory(
        payment=payment_one,
        payment_verification_plan=pvp,
        status=PaymentVerification.STATUS_RECEIVED,
        received_amount=payment_one.entitlement_quantity,
    )

    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    household_col = headers.index("household_id") + 1
    delivered_col = headers.index("delivered_quantity") + 1
    partial = payment_one.entitlement_quantity / 2
    for row_idx in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row=row_idx, column=household_col).value) == payment_one.household.unicef_id:
            worksheet.cell(row=row_idx, column=delivered_col).value = str(partial)

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="partial-delivery.xlsx")
        )

    service.open_workbook()
    service.validate()
    assert not service.errors
    service.import_payment_list()

    pv.refresh_from_db()
    assert pv.status == PaymentVerification.STATUS_RECEIVED_WITH_ISSUES


def test_import_skips_payment_verification_update_when_status_already_matches(
    instruction,
    instruction_payments,
    child_payment_plans,
    delivery_template,
):
    payment_one = instruction_payments[0]
    plan_one = child_payment_plans[0]
    PaymentVerificationSummaryFactory(payment_plan=plan_one)
    pvp = PaymentVerificationPlanFactory(payment_plan=plan_one)
    PaymentVerificationFactory(
        payment=payment_one,
        payment_verification_plan=pvp,
        status=PaymentVerification.STATUS_RECEIVED,
        received_amount=payment_one.entitlement_quantity,
    )

    workbook = XlsxFollowUpInstructionDeliveryExportService(instruction).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    household_col = headers.index("household_id") + 1
    delivered_col = headers.index("delivered_quantity") + 1
    for row_idx in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row=row_idx, column=household_col).value) == payment_one.household.unicef_id:
            worksheet.cell(row=row_idx, column=delivered_col).value = str(payment_one.entitlement_quantity)

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        service = XlsxFollowUpInstructionReconciliationImportService(
            instruction, ContentFile(tmp.read(), name="already-received.xlsx")
        )

    service.open_workbook()
    service.validate()
    service.import_payment_list()

    assert service.payment_verifications_to_save == []
