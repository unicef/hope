from decimal import Decimal
from types import SimpleNamespace
from typing import Any
from unittest.mock import Mock, patch

import openpyxl
import pytest

from extras.test_utils.factories import (
    BusinessAreaFactory,
    CurrencyFactory,
    DeliveryMechanismFactory,
    FileTempFactory,
    FinancialServiceProviderFactory,
    FinancialServiceProviderXlsxTemplateFactory,
    FspXlsxTemplatePerDeliveryMechanismFactory,
    HouseholdFactory,
    PaymentFactory,
    PaymentHouseholdSnapshotFactory,
    PaymentPlanFactory,
    ProgramCycleFactory,
    ProgramFactory,
    UserFactory,
)
from hope.apps.payment.celery_tasks import (
    create_follow_up_instruction_delivery_xlsx_async_task_action,
)
from hope.apps.payment.xlsx.xlsx_follow_up_instruction_base_export_service import (
    XlsxFollowUpInstructionBaseExportService,
)
from hope.models import FileTemp, FollowUpInstruction, Payment, PaymentPlan

pytestmark = pytest.mark.django_db


@pytest.fixture
def business_area():
    return BusinessAreaFactory(slug="afghanistan")


@pytest.fixture
def user():
    return UserFactory()


@pytest.fixture
def program(business_area):
    return ProgramFactory(business_area=business_area)


@pytest.fixture
def cycle(program):
    return ProgramCycleFactory(program=program)


@pytest.fixture
def currency():
    return CurrencyFactory(code="AFN")


@pytest.fixture
def delivery_mechanism():
    return DeliveryMechanismFactory(code="cash", name="Cash", payment_gateway_id="dm-cash")


@pytest.fixture
def fsp(delivery_mechanism):
    fsp = FinancialServiceProviderFactory()
    fsp.delivery_mechanisms.add(delivery_mechanism)
    return fsp


@pytest.fixture
def instruction(user, program, business_area):
    return FollowUpInstruction.objects.create(
        business_area=business_area,
        program=program,
        created_by=user,
        background_action_status=FollowUpInstruction.BackgroundActionStatus.XLSX_EXPORTING,
    )


@pytest.fixture
def child_payment_plans(instruction, cycle, business_area, currency, delivery_mechanism, fsp):
    plan_one = PaymentPlanFactory(
        program_cycle=cycle,
        business_area=business_area,
        currency=currency,
        delivery_mechanism=delivery_mechanism,
        financial_service_provider=fsp,
        follow_up_instruction=instruction,
        plan_type=PaymentPlan.PlanType.FOLLOW_UP,
        status=PaymentPlan.Status.ACCEPTED,
        exchange_rate=Decimal("1.00"),
    )
    plan_two = PaymentPlanFactory(
        program_cycle=cycle,
        business_area=business_area,
        currency=currency,
        delivery_mechanism=delivery_mechanism,
        financial_service_provider=fsp,
        follow_up_instruction=instruction,
        plan_type=PaymentPlan.PlanType.FOLLOW_UP,
        status=PaymentPlan.Status.ACCEPTED,
        exchange_rate=Decimal("1.00"),
    )
    return plan_one, plan_two


@pytest.fixture
def instruction_payments(child_payment_plans, fsp, delivery_mechanism, currency):
    plan_one, plan_two = child_payment_plans
    shared_household = HouseholdFactory(business_area=plan_one.business_area, program=plan_one.program)
    payment_one = PaymentFactory(
        parent=plan_one,
        household=shared_household,
        collector=shared_household.head_of_household,
        head_of_household=shared_household.head_of_household,
        program=plan_one.program,
        currency=currency,
        delivery_type=delivery_mechanism,
        financial_service_provider=fsp,
        entitlement_quantity=Decimal("100.00"),
        entitlement_quantity_usd=Decimal("100.00"),
        delivered_quantity=Decimal("0.00"),
        status=Payment.STATUS_PENDING,
    )
    payment_two = PaymentFactory(
        parent=plan_two,
        household=shared_household,
        collector=shared_household.head_of_household,
        head_of_household=shared_household.head_of_household,
        program=plan_two.program,
        currency=currency,
        delivery_type=delivery_mechanism,
        financial_service_provider=fsp,
        entitlement_quantity=Decimal("40.00"),
        entitlement_quantity_usd=Decimal("40.00"),
        delivered_quantity=Decimal("0.00"),
        status=Payment.STATUS_PENDING,
    )
    second_household = HouseholdFactory(business_area=plan_two.business_area, program=plan_two.program)
    payment_three = PaymentFactory(
        parent=plan_two,
        household=second_household,
        collector=second_household.head_of_household,
        head_of_household=second_household.head_of_household,
        program=plan_two.program,
        currency=currency,
        delivery_type=delivery_mechanism,
        financial_service_provider=fsp,
        entitlement_quantity=Decimal("60.00"),
        entitlement_quantity_usd=Decimal("60.00"),
        delivered_quantity=Decimal("0.00"),
        status=Payment.STATUS_PENDING,
    )
    for payment in (payment_one, payment_two, payment_three):
        PaymentHouseholdSnapshotFactory(
            payment=payment,
            snapshot_data={
                "unicef_id": payment.household.unicef_id,
                "size": payment.household.size,
                "primary_collector": {
                    "unicef_id": payment.collector.unicef_id,
                    "full_name": payment.collector.full_name,
                },
                "alternate_collector": {},
            },
        )
    return payment_one, payment_two, payment_three


@pytest.fixture
def delivery_template(fsp, delivery_mechanism):
    template = FinancialServiceProviderXlsxTemplateFactory(
        columns=["payment_id", "entitlement_quantity", "delivered_quantity"],
        core_fields=[],
        flex_fields=[],
        document_types=[],
    )
    FspXlsxTemplatePerDeliveryMechanismFactory(
        financial_service_provider=fsp,
        delivery_mechanism=delivery_mechanism,
        xlsx_template=template,
    )
    return template


def _load_exported_workbook(instruction: FollowUpInstruction) -> openpyxl.Workbook:
    assert instruction.export_file is not None
    instruction.export_file.file.open("rb")
    return openpyxl.load_workbook(instruction.export_file.file, data_only=True)


def _rows_by_household(worksheet: Any) -> dict[str, dict[str, Any]]:
    headers = [cell.value for cell in worksheet[1]]
    household_col = headers.index("household_id") + 1
    rows = {}
    for row_idx in range(2, worksheet.max_row + 1):
        household_id = str(worksheet.cell(row=row_idx, column=household_col).value)
        rows[household_id] = {
            header: worksheet.cell(row=row_idx, column=col_idx).value for col_idx, header in enumerate(headers, start=1)
        }
    return rows


def test_delivery_export_happy_path_aggregates_households(
    user,
    instruction,
    instruction_payments,
    delivery_template,
):
    job = SimpleNamespace(config={"follow_up_instruction_id": str(instruction.id), "user_id": str(user.id)})

    create_follow_up_instruction_delivery_xlsx_async_task_action(job)

    instruction.refresh_from_db()
    assert instruction.background_action_status is None
    workbook = _load_exported_workbook(instruction)
    headers = [cell.value for cell in workbook.active[1]]
    rows = _rows_by_household(workbook.active)

    assert headers == ["household_id", "entitlement_quantity", "delivered_quantity"]
    shared_household_id = instruction_payments[0].household.unicef_id
    assert rows[shared_household_id]["entitlement_quantity"] == Decimal("140.00")
    assert rows[shared_household_id]["delivered_quantity"] == Decimal("0.00")


class _MinimalExportService(XlsxFollowUpInstructionBaseExportService):
    filename_prefix = "test"

    def get_source_headers(self) -> list[str]:
        return ["household_id", "entitlement_quantity", "delivered_quantity"]

    def get_payment_row_data(self, payment: Any) -> dict[str, Any]:
        return {}


def test_get_representative_payment_plan_raises_when_no_child_plans(instruction: Any) -> None:
    with pytest.raises(ValueError, match="Follow Up Instruction has no child Payment Plans."):
        _MinimalExportService(instruction)


def test_prepare_headers_removes_household_unicef_id(instruction: Any, child_payment_plans: Any) -> None:
    class _ServiceWithUnicefId(_MinimalExportService):
        def get_source_headers(self) -> list[str]:
            return ["household_unicef_id", "entitlement_quantity"]

    service = _ServiceWithUnicefId(instruction)

    assert "household_unicef_id" not in service.headers
    assert "entitlement_quantity" in service.headers


def test_as_decimal_returns_zero_for_none() -> None:
    service = object.__new__(_MinimalExportService)

    assert service._as_decimal(None) == Decimal(0)


def test_as_decimal_returns_zero_for_empty_string() -> None:
    service = object.__new__(_MinimalExportService)

    assert service._as_decimal("") == Decimal(0)


def test_as_decimal_returns_decimal_unchanged() -> None:
    service = object.__new__(_MinimalExportService)

    assert service._as_decimal(Decimal("3.50")) == Decimal("3.50")


def test_as_decimal_converts_string_to_decimal() -> None:
    service = object.__new__(_MinimalExportService)

    assert service._as_decimal("2.25") == Decimal("2.25")


def test_merge_rows_fills_empty_non_summable_field_from_payment_row() -> None:
    service = object.__new__(_MinimalExportService)
    service.headers = ["household_id", "entitlement_quantity", "delivery_mechanism"]
    existing_row = {"household_id": "HH-01", "entitlement_quantity": Decimal(100), "delivery_mechanism": ""}
    new_row = {"household_id": "HH-01", "entitlement_quantity": Decimal(50), "delivery_mechanism": "Cash"}

    merged = service._merge_rows(existing_row, new_row)

    assert merged["delivery_mechanism"] == "Cash"


def test_save_xlsx_file_removes_existing_export_file(
    instruction: Any,
    child_payment_plans: Any,
    user: Any,
) -> None:
    old_file = FileTempFactory()
    instruction.export_file = old_file
    instruction.save(update_fields=["export_file"])
    old_file_id = old_file.pk
    service = _MinimalExportService(instruction)

    service.save_xlsx_file(user)

    instruction.refresh_from_db()
    assert not FileTemp.objects.filter(pk=old_file_id).exists()
    assert instruction.export_file is not None
    assert instruction.export_file.pk != old_file_id


@patch(
    "hope.apps.payment.xlsx.xlsx_follow_up_instruction_delivery_export_service"
    ".XlsxFollowUpInstructionDeliveryExportService.save_xlsx_file"
)
def test_delivery_export_marks_xlsx_export_error_on_exception(
    mock_save_xlsx_file: Mock,
    instruction,
    child_payment_plans,
    delivery_template,
    user,
) -> None:
    mock_save_xlsx_file.side_effect = Exception("export failed")
    job = SimpleNamespace(config={"follow_up_instruction_id": str(instruction.id), "user_id": str(user.id)})

    with pytest.raises(Exception, match="export failed"):
        create_follow_up_instruction_delivery_xlsx_async_task_action(job)

    instruction.refresh_from_db(fields=["background_action_status"])
    assert instruction.background_action_status == FollowUpInstruction.BackgroundActionStatus.XLSX_EXPORT_ERROR
