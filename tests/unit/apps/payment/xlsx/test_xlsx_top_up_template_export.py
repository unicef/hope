import io

import openpyxl
import pytest

from extras.test_utils.factories.household import HouseholdFactory
from extras.test_utils.factories.payment import PaymentFactory, PaymentPlanFactory
from hope.apps.payment.xlsx.xlsx_top_up_template_export_service import XlsxTopUpTemplateExportService
from hope.models import Payment, PaymentPlan


@pytest.fixture
def source_pp(db):
    return PaymentPlanFactory(plan_type=PaymentPlan.PlanType.REGULAR)


@pytest.fixture
def eligible_payment(source_pp):
    household = HouseholdFactory(
        program=source_pp.program_cycle.program,
        business_area=source_pp.business_area,
        size=4,
        village="Kabul Test Village",
    )
    return PaymentFactory(
        parent=source_pp,
        household=household,
        status=Payment.STATUS_DISTRIBUTION_SUCCESS,
    )


def _render(source_pp: PaymentPlan) -> openpyxl.Workbook:
    buffer = io.BytesIO()
    XlsxTopUpTemplateExportService(source_pp).write_to_stream(buffer)
    buffer.seek(0)
    return openpyxl.load_workbook(buffer, data_only=True)


def test_template_arrange_one_eligible_act_render_assert_household_unicef_id_first_column(source_pp, eligible_payment):
    wb = _render(source_pp)
    ws = wb.active

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header[0] == "household_unicef_id"


def test_template_arrange_one_eligible_act_render_assert_payment_id_column_absent(source_pp, eligible_payment):
    wb = _render(source_pp)
    ws = wb.active

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "payment_id" not in header


def test_template_arrange_one_eligible_act_render_assert_one_data_row(source_pp, eligible_payment):
    wb = _render(source_pp)
    ws = wb.active

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 1
    assert data_rows[0][0] == eligible_payment.household.unicef_id


def test_template_arrange_withdrawn_household_act_render_assert_skipped(source_pp):
    withdrawn_household = HouseholdFactory(
        program=source_pp.program_cycle.program,
        business_area=source_pp.business_area,
        withdrawn=True,
    )
    PaymentFactory(parent=source_pp, household=withdrawn_household, status=Payment.STATUS_DISTRIBUTION_SUCCESS)

    wb = _render(source_pp)
    ws = wb.active

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert data_rows == []


def test_template_arrange_already_topped_up_household_act_render_assert_skipped(source_pp, eligible_payment):
    top_up_pp = PaymentPlanFactory(
        plan_type=PaymentPlan.PlanType.TOP_UP,
        source_payment_plan=source_pp,
        program_cycle=source_pp.program_cycle,
    )
    PaymentFactory(parent=top_up_pp, household=eligible_payment.household, status=Payment.STATUS_PENDING)

    wb = _render(source_pp)
    ws = wb.active

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert data_rows == []


def test_template_arrange_failed_payment_act_render_assert_skipped(source_pp):
    household = HouseholdFactory(
        program=source_pp.program_cycle.program,
        business_area=source_pp.business_area,
    )
    PaymentFactory(parent=source_pp, household=household, status=Payment.STATUS_FORCE_FAILED)

    wb = _render(source_pp)
    ws = wb.active

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert data_rows == []


def test_template_arrange_one_eligible_act_render_assert_entitlement_column_empty(source_pp, eligible_payment):
    wb = _render(source_pp)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    amount_idx = header.index("entitlement_quantity")

    data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert data_row[amount_idx] == "" or data_row[amount_idx] is None
