from decimal import Decimal
from io import BytesIO
from typing import Any

from django.core.exceptions import ValidationError
import openpyxl
import pytest

from extras.test_utils.factories import (
    BusinessAreaFactory,
    CurrencyFactory,
    PaymentFactory,
    PaymentHouseholdSnapshotFactory,
    PaymentPlanFactory,
    PaymentPlanPurposeFactory,
    ProgramCycleFactory,
    ProgramFactory,
)
from hope.apps.payment.services.top_up_amount_service import TopUpAmountTemplateService, parse_top_up_amount_file
from hope.apps.payment.xlsx.xlsx_payment_plan_base_service import XlsxPaymentPlanBaseService
from hope.models import Payment, PaymentPlan, Program

pytestmark = pytest.mark.django_db

PAYMENT_ID_COLUMN = XlsxPaymentPlanBaseService.COLUMN_PAYMENT_ID
AMOUNT_COLUMN = XlsxPaymentPlanBaseService.COLUMN_ENTITLEMENT_QUANTITY


@pytest.fixture
def source_pp(db: Any) -> PaymentPlan:
    business_area = BusinessAreaFactory(slug="afghanistan")
    program = ProgramFactory(business_area=business_area, status=Program.ACTIVE)
    cycle = ProgramCycleFactory(program=program)
    purpose = PaymentPlanPurposeFactory()
    program.payment_plan_purposes.add(purpose)
    plan = PaymentPlanFactory(
        name="Standard PP",
        business_area=business_area,
        program_cycle=cycle,
        plan_type=PaymentPlan.PlanType.REGULAR,
        status=PaymentPlan.Status.ACCEPTED,
        currency=CurrencyFactory(code="USD"),
        payment_plan_purposes=[purpose],
    )
    for _ in range(2):
        payment = PaymentFactory(parent=plan, status=Payment.STATUS_PENDING)
        PaymentHouseholdSnapshotFactory(
            payment=payment,
            snapshot_data={"unicef_id": payment.household.unicef_id, "size": payment.household.size},
        )
    return plan


def _template(source_pp: PaymentPlan) -> Any:
    return TopUpAmountTemplateService(source_pp).generate_workbook()


def _as_file(workbook: Any) -> BytesIO:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _set_cell(worksheet: Any, row: int, column_name: str, value: Any) -> None:
    headers = [cell.value for cell in worksheet[1]]
    worksheet.cell(row=row, column=headers.index(column_name) + 1).value = value


def test_parse_top_up_amount_file_arrange_funded_row_act_parse_assert_amount_returned(
    source_pp: PaymentPlan,
) -> None:
    workbook = _template(source_pp)
    _set_cell(workbook.active, 2, AMOUNT_COLUMN, 12.5)
    funded_id = workbook.active.cell(row=2, column=1).value

    amounts = parse_top_up_amount_file(source_pp, _as_file(workbook))

    assert amounts == {funded_id: Decimal("12.50")}


def test_parse_top_up_amount_file_arrange_not_a_workbook_act_parse_assert_raises(
    source_pp: PaymentPlan,
) -> None:
    with pytest.raises(ValidationError, match="could not be read"):
        parse_top_up_amount_file(source_pp, BytesIO(b"definitely not an xlsx"))


def test_parse_top_up_amount_file_arrange_wrong_sheet_name_act_parse_assert_raises(
    source_pp: PaymentPlan,
) -> None:
    workbook = openpyxl.Workbook()
    workbook.active.title = "Some Other Sheet"

    with pytest.raises(ValidationError, match="not found in the uploaded file"):
        parse_top_up_amount_file(source_pp, _as_file(workbook))


def test_parse_top_up_amount_file_arrange_missing_amount_column_act_parse_assert_raises(
    source_pp: PaymentPlan,
) -> None:
    workbook = _template(source_pp)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    worksheet.cell(row=1, column=headers.index(AMOUNT_COLUMN) + 1).value = "something_else"

    with pytest.raises(ValidationError, match=f"Column '{AMOUNT_COLUMN}' is required"):
        parse_top_up_amount_file(source_pp, _as_file(workbook))


def test_parse_top_up_amount_file_arrange_non_numeric_amount_act_parse_assert_raises(
    source_pp: PaymentPlan,
) -> None:
    workbook = _template(source_pp)
    _set_cell(workbook.active, 2, AMOUNT_COLUMN, "not a number")

    with pytest.raises(ValidationError, match="Invalid amount"):
        parse_top_up_amount_file(source_pp, _as_file(workbook))


def test_parse_top_up_amount_file_arrange_duplicate_payment_row_act_parse_assert_raises(
    source_pp: PaymentPlan,
) -> None:
    workbook = _template(source_pp)
    worksheet = workbook.active
    first_payment_id = worksheet.cell(row=2, column=1).value
    _set_cell(worksheet, 3, PAYMENT_ID_COLUMN, first_payment_id)
    _set_cell(worksheet, 2, AMOUNT_COLUMN, 10)
    _set_cell(worksheet, 3, AMOUNT_COLUMN, 20)

    with pytest.raises(ValidationError, match="appears more than once"):
        parse_top_up_amount_file(source_pp, _as_file(workbook))


def test_parse_top_up_amount_file_arrange_blank_payment_id_row_act_parse_assert_row_skipped(
    source_pp: PaymentPlan,
) -> None:
    """A row with no payment id is padding, not an error — operators leave trailing blanks."""
    workbook = _template(source_pp)
    worksheet = workbook.active
    _set_cell(worksheet, 2, AMOUNT_COLUMN, 10)
    _set_cell(worksheet, 3, PAYMENT_ID_COLUMN, None)
    _set_cell(worksheet, 3, AMOUNT_COLUMN, 99)

    amounts = parse_top_up_amount_file(source_pp, _as_file(workbook))

    assert amounts == {worksheet.cell(row=2, column=1).value: Decimal("10.00")}
