from io import BytesIO
from typing import Any, Callable

from django.core.files.uploadedfile import SimpleUploadedFile
import openpyxl
import pytest
from rest_framework import status
from rest_framework.reverse import reverse

from extras.test_utils.factories import (
    BusinessAreaFactory,
    CurrencyFactory,
    PaymentFactory,
    PaymentHouseholdSnapshotFactory,
    PaymentPlanFactory,
    PaymentPlanPurposeFactory,
    ProgramCycleFactory,
    ProgramFactory,
    UserFactory,
)
from hope.apps.account.permissions import Permissions
from hope.apps.payment.services.top_up_amount_service import TopUpAmountTemplateService
from hope.apps.payment.xlsx.xlsx_payment_plan_base_service import XlsxPaymentPlanBaseService
from hope.models import Payment, PaymentPlan, Program

pytestmark = pytest.mark.django_db

DATES = {"dispersion_start_date": "2024-01-01", "dispersion_end_date": "2099-12-31"}


@pytest.fixture
def top_up_context(api_client: Callable, create_user_role_with_permissions: Any) -> dict[str, Any]:
    business_area = BusinessAreaFactory(slug="afghanistan")
    user = UserFactory()
    program = ProgramFactory(business_area=business_area, status=Program.ACTIVE)
    cycle = ProgramCycleFactory(program=program)
    purpose = PaymentPlanPurposeFactory()
    program.payment_plan_purposes.add(purpose)
    source_pp = PaymentPlanFactory(
        name="Standard PP",
        business_area=business_area,
        program_cycle=cycle,
        plan_type=PaymentPlan.PlanType.REGULAR,
        status=PaymentPlan.Status.ACCEPTED,
        currency=CurrencyFactory(code="USD"),
        payment_plan_purposes=[purpose],
    )
    payments = [PaymentFactory(parent=source_pp, status=Payment.STATUS_PENDING) for _ in range(3)]
    # The template is rendered from household snapshots, which a real Accepted plan always has.
    for payment in payments:
        PaymentHouseholdSnapshotFactory(
            payment=payment,
            snapshot_data={"unicef_id": payment.household.unicef_id, "size": payment.household.size},
        )
    create_user_role_with_permissions(user, [Permissions.PM_CREATE], business_area, program)
    url_kwargs = {
        "business_area_slug": business_area.slug,
        "program_code": program.code,
        "pk": source_pp.pk,
    }
    return {
        "source_pp": source_pp,
        "payments": payments,
        "client": api_client(user),
        "template_url": reverse("api:payments:payment-plans-top-up-amount-template", kwargs=url_kwargs),
        "create_url": reverse("api:payments:payment-plans-create-top-up", kwargs=url_kwargs),
    }


def _amount_file(source_pp: PaymentPlan, amounts_by_payment_id: dict[str, str | None]) -> SimpleUploadedFile:
    """Build a filled-in amount template. Payment ids missing from the mapping are left blank."""
    workbook = TopUpAmountTemplateService(source_pp).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    payment_id_column = headers.index(XlsxPaymentPlanBaseService.COLUMN_PAYMENT_ID) + 1
    amount_column = headers.index(XlsxPaymentPlanBaseService.COLUMN_ENTITLEMENT_QUANTITY) + 1
    for row_index in range(2, worksheet.max_row + 1):
        payment_id = worksheet.cell(row=row_index, column=payment_id_column).value
        worksheet.cell(row=row_index, column=amount_column).value = amounts_by_payment_id.get(str(payment_id))
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile("amounts.xlsx", buffer.read())


def test_top_up_amount_template_arrange_eligible_payments_act_get_assert_rows_without_amounts(
    top_up_context: dict[str, Any],
) -> None:
    response = top_up_context["client"].get(top_up_context["template_url"])

    assert response.status_code == status.HTTP_200_OK
    worksheet = openpyxl.load_workbook(BytesIO(response.content)).active
    headers = [cell.value for cell in worksheet[1]]
    amount_column = headers.index(XlsxPaymentPlanBaseService.COLUMN_ENTITLEMENT_QUANTITY) + 1
    assert worksheet.max_row == 4
    assert {worksheet.cell(row=row, column=amount_column).value for row in range(2, 5)} == {None}


def test_top_up_amount_template_arrange_no_eligible_payments_act_get_assert_400(
    top_up_context: dict[str, Any],
) -> None:
    top_up_context["source_pp"].payment_items.all().delete()

    response = top_up_context["client"].get(top_up_context["template_url"])

    assert response.status_code == status.HTTP_400_BAD_REQUEST


def test_create_top_up_arrange_amount_file_act_post_assert_only_funded_copied(
    top_up_context: dict[str, Any],
    django_capture_on_commit_callbacks: Any,
) -> None:
    funded, skipped, zeroed = top_up_context["payments"]
    upload = _amount_file(
        top_up_context["source_pp"],
        {funded.unicef_id: "40.00", skipped.unicef_id: None, zeroed.unicef_id: "0"},
    )

    with django_capture_on_commit_callbacks(execute=True):
        response = top_up_context["client"].post(
            top_up_context["create_url"], {**DATES, "file": upload}, format="multipart"
        )

    assert response.status_code == status.HTTP_201_CREATED
    top_up = PaymentPlan.objects.get(pk=response.json()["id"])
    assert list(top_up.payment_items.values_list("source_payment_id", flat=True)) == [funded.id]


def test_create_top_up_arrange_no_amount_given_act_post_assert_400(top_up_context: dict[str, Any]) -> None:
    response = top_up_context["client"].post(top_up_context["create_url"], DATES, format="json")

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert "fixed amount or an amount file" in str(response.json())


def test_create_top_up_arrange_both_fixed_and_file_act_post_assert_400(top_up_context: dict[str, Any]) -> None:
    upload = _amount_file(top_up_context["source_pp"], {top_up_context["payments"][0].unicef_id: "40.00"})

    response = top_up_context["client"].post(
        top_up_context["create_url"],
        {**DATES, "fixed_amount": "10.00", "file": upload},
        format="multipart",
    )

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert "not both" in str(response.json())


def test_create_top_up_arrange_amount_file_funding_nobody_act_post_assert_400(
    top_up_context: dict[str, Any],
) -> None:
    upload = _amount_file(top_up_context["source_pp"], {})

    response = top_up_context["client"].post(
        top_up_context["create_url"], {**DATES, "file": upload}, format="multipart"
    )

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert "funds nobody" in str(response.json())


def test_create_top_up_arrange_amount_file_with_foreign_payment_act_post_assert_400(
    top_up_context: dict[str, Any],
) -> None:
    workbook = TopUpAmountTemplateService(top_up_context["source_pp"]).generate_workbook()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    worksheet.cell(row=2, column=headers.index(XlsxPaymentPlanBaseService.COLUMN_PAYMENT_ID) + 1).value = "RCPT-NOPE"
    worksheet.cell(row=2, column=headers.index(XlsxPaymentPlanBaseService.COLUMN_ENTITLEMENT_QUANTITY) + 1).value = 5
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = top_up_context["client"].post(
        top_up_context["create_url"],
        {**DATES, "file": SimpleUploadedFile("amounts.xlsx", buffer.read())},
        format="multipart",
    )

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert "not eligible" in str(response.json())


def test_create_top_up_arrange_negative_amount_in_file_act_post_assert_400(top_up_context: dict[str, Any]) -> None:
    upload = _amount_file(top_up_context["source_pp"], {top_up_context["payments"][0].unicef_id: "-5"})

    response = top_up_context["client"].post(
        top_up_context["create_url"], {**DATES, "file": upload}, format="multipart"
    )

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert "Negative amount" in str(response.json())


def test_top_up_amount_template_arrange_follow_up_plan_act_get_assert_400(top_up_context: dict[str, Any]) -> None:
    source_pp = top_up_context["source_pp"]
    source_pp.plan_type = PaymentPlan.PlanType.FOLLOW_UP
    source_pp.save(update_fields=["plan_type"])

    response = top_up_context["client"].get(top_up_context["template_url"])

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert "No amount template" in str(response.json())


def test_top_up_amount_template_arrange_open_plan_act_get_assert_400(top_up_context: dict[str, Any]) -> None:
    source_pp = top_up_context["source_pp"]
    source_pp.status = PaymentPlan.Status.OPEN
    source_pp.save(update_fields=["status"])

    response = top_up_context["client"].get(top_up_context["template_url"])

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert "Accepted or Finished" in str(response.json())
