import datetime
import io
import os
import tempfile
from collections import namedtuple
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any, Tuple
from unittest.mock import patch
from zipfile import ZipFile

from django.conf import settings
from django.contrib.admin.options import get_content_type_for_model
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone

import pytz
from extras.test_utils.factories.account import PartnerFactory, UserFactory
from extras.test_utils.factories.core import create_afghanistan
from extras.test_utils.factories.household import (
    IndividualRoleInHouseholdFactory,
    create_household_and_individuals,
)
from extras.test_utils.factories.payment import (
    AccountFactory,
    FinancialServiceProviderFactory,
    FinancialServiceProviderXlsxTemplateFactory,
    FspXlsxTemplatePerDeliveryMechanismFactory,
    PaymentFactory,
    PaymentPlanFactory,
    PaymentVerificationFactory,
    PaymentVerificationPlanFactory,
    PaymentVerificationSummaryFactory,
    generate_delivery_mechanisms,
)
from extras.test_utils.factories.program import BeneficiaryGroupFactory, ProgramFactory
from extras.test_utils.factories.registration_data import RegistrationDataImportFactory
from extras.test_utils.factories.steficon import RuleCommitFactory, RuleFactory
from freezegun import freeze_time
from openpyxl import load_workbook
from parameterized import parameterized
from pytz import utc

from hct_mis_api.apps.account.permissions import Permissions
from hct_mis_api.apps.core.base_test_case import APITestCase
from hct_mis_api.apps.core.models import DataCollectingType, FileTemp
from hct_mis_api.apps.core.utils import (
    decode_id_string,
    decode_id_string_required,
    encode_id_base64,
    encode_id_base64_required,
)
from hct_mis_api.apps.household.models import ROLE_PRIMARY
from hct_mis_api.apps.payment.celery_tasks import (
    create_payment_plan_payment_list_xlsx_per_fsp,
    payment_plan_apply_engine_rule,
)
from hct_mis_api.apps.payment.models import (
    AccountType,
    DeliveryMechanism,
    FinancialServiceProvider,
    FinancialServiceProviderXlsxTemplate,
    Payment,
    PaymentPlan,
    PaymentVerification,
    PaymentVerificationPlan,
)
from hct_mis_api.apps.payment.services.payment_household_snapshot_service import (
    create_payment_plan_snapshot_data,
)
from hct_mis_api.apps.payment.xlsx.xlsx_payment_plan_per_fsp_import_service import (
    XlsxPaymentPlanImportPerFspService,
)
from hct_mis_api.apps.program.models import Program, ProgramCycle
from hct_mis_api.apps.steficon.models import Rule

if TYPE_CHECKING:
    from hct_mis_api.apps.household.models import Household, Individual

CREATE_PROGRAMME_MUTATION = """
mutation CreateProgram($programData: CreateProgramInput!) {
  createProgram(programData: $programData) {
    program {
      id
      cycles {
        edges {
          node {
            id
            status
          }
        }
      }
    }
  }
}
"""

UPDATE_PROGRAM_MUTATION = """
mutation UpdateProgram($programData: UpdateProgramInput!) {
  updateProgram(programData: $programData) {
    program {
      id
    }
  }
}
"""

CREATE_PAYMENT_PLAN_MUTATION = """
mutation CreatePaymentPlan($input: CreatePaymentPlanInput!) {
    createPaymentPlan(input: $input) {
        paymentPlan {
            id
            status
        }
    }
}
"""

FINALIZE_TARGET_POPULATION_MUTATION = """
mutation FinalizeTP($id: ID!) {
    finalizeTargetPopulation(id: $id) {
        targetPopulation {
            id
            status
        }
    }
}
"""

SET_STEFICON_RULE_MUTATION = """
mutation setSteficonRuleOnPaymentPlanPaymentList($paymentPlanId: ID!, $steficonRuleId: ID!, $version: BigInt) {
    setSteficonRuleOnPaymentPlanPaymentList(paymentPlanId: $paymentPlanId, steficonRuleId: $steficonRuleId, version: $version) {
        paymentPlan {
            unicefId
        }
    }
}
"""

OPEN_PAYMENT_PLAN_MUTATION = """
mutation OpenPaymentPlan($input: OpenPaymentPlanInput!) {
    openPaymentPlan(input: $input) {
        paymentPlan {
            id
            status
        }
    }
}
"""


PAYMENT_PLAN_ACTION_MUTATION = """
mutation ActionPaymentPlanMutation($input: ActionPaymentPlanInput!) {
    actionPaymentPlanMutation(input: $input) {
        paymentPlan {
            status
            id
        }
    }
}"""

CHOOSE_DELIVERY_MECHANISMS_MUTATION = """
mutation ChooseDeliveryMechanismsForPaymentPlan($input: ChooseDeliveryMechanismsForPaymentPlanInput!) {
    chooseDeliveryMechanismsForPaymentPlan(input: $input) {
        paymentPlan {
            id
            deliveryMechanism {
                order
                name
            }
        }
    }
}
"""

AVAILABLE_FSPS_FOR_DELIVERY_MECHANISMS_QUERY = """
query AvailableFspsForDeliveryMechanisms {
    availableFspsForDeliveryMechanisms {
        deliveryMechanism {
            name
            code
        }
        fsps {
            id
            name
        }
    }
}
"""

ASSIGN_FSPS_MUTATION = """
mutation AssignFspToDeliveryMechanism($paymentPlanId: ID!, $mappings: [FSPToDeliveryMechanismMappingInput!]!) {
    assignFspToDeliveryMechanism(input: {
        paymentPlanId: $paymentPlanId,
        mappings: $mappings
    }) {
        paymentPlan {
            id
            deliveryMechanism {
                name
                order
                fsp {
                    id
                }
            }
        }
    }
}
"""

EXPORT_XLSX_PER_FSP_MUTATION = """
mutation ExportXlsxPaymentPlanPaymentListPerFsp($paymentPlanId: ID!, $fspXlsxTemplateId: ID) {
    exportXlsxPaymentPlanPaymentListPerFsp(paymentPlanId: $paymentPlanId, fspXlsxTemplateId: $fspXlsxTemplateId) {
        paymentPlan {
            id
        }
    }
}
"""

EXPORT_XLSX_PER_FSP_MUTATION_AUTH_CODE = """
mutation ExportXlsxPaymentPlanPaymentListPerFsp($paymentPlanId: ID!, $fspXlsxTemplateId: ID) {
    exportXlsxPaymentPlanPaymentListPerFsp(paymentPlanId: $paymentPlanId, fspXlsxTemplateId: $fspXlsxTemplateId) {
        paymentPlan {
            status
            canCreateXlsxWithFspAuthCode
            hasPaymentListExportFile
            canRegenerateExportFilePerFsp
        }
    }
}
"""

IMPORT_XLSX_PER_FSP_MUTATION = """
mutation ImportXlsxPaymentPlanPaymentListPerFsp($paymentPlanId: ID!, $file: Upload!) {
    importXlsxPaymentPlanPaymentListPerFsp(paymentPlanId: $paymentPlanId, file: $file) {
        paymentPlan {
            id
        }
        errors {
            sheet
            coordinates
            message
        }
    }
}
"""

IMPORT_XLSX_PP_MUTATION = """
mutation importXlsxPPList($paymentPlanId: ID!, $file: Upload!) {
importXlsxPaymentPlanPaymentList(
  paymentPlanId: $paymentPlanId
  file: $file
) {
  paymentPlan {
    id
  }
  errors {
    sheet
    coordinates
    message
  }
}
}
"""


class TestPaymentPlanReconciliation(APITestCase):
    @classmethod
    def create_household_and_individual(cls, program: Program) -> Tuple["Household", "Individual"]:
        household, individuals = create_household_and_individuals(
            household_data={
                "registration_data_import": cls.registration_data_import,
                "business_area": cls.business_area,
                "program": program,
            },
            individuals_data=[{}],
        )
        IndividualRoleInHouseholdFactory(household=household, individual=individuals[0], role=ROLE_PRIMARY)
        return household, individuals[0]

    @classmethod
    def setUpTestData(cls) -> None:
        super().setUpTestData()
        cls.business_area = create_afghanistan()
        partner = PartnerFactory(name="Partner")
        cls.user = UserFactory.create(partner=partner)
        cls.all_necessary_permissions = [
            Permissions.PM_CREATE,
            Permissions.PM_VIEW_DETAILS,
            Permissions.PM_VIEW_LIST,
            Permissions.PM_IMPORT_XLSX_WITH_ENTITLEMENTS,
            Permissions.PM_APPLY_RULE_ENGINE_FORMULA_WITH_ENTITLEMENTS,
            Permissions.PROGRAMME_CREATE,
            Permissions.PROGRAMME_UPDATE,
            Permissions.PROGRAMME_ACTIVATE,
            Permissions.PM_PROGRAMME_CYCLE_CREATE,
            Permissions.PM_PROGRAMME_CYCLE_UPDATE,
            Permissions.TARGETING_CREATE,
            Permissions.TARGETING_LOCK,
            Permissions.TARGETING_SEND,
            Permissions.PM_LOCK_AND_UNLOCK,
            Permissions.PM_LOCK_AND_UNLOCK_FSP,
            Permissions.PM_SEND_FOR_APPROVAL,
            Permissions.PM_ACCEPTANCE_PROCESS_APPROVE,
            Permissions.PM_ACCEPTANCE_PROCESS_AUTHORIZE,
            Permissions.PM_ACCEPTANCE_PROCESS_FINANCIAL_REVIEW,
            Permissions.PM_IMPORT_XLSX_WITH_RECONCILIATION,
            Permissions.PM_DOWNLOAD_FSP_AUTH_CODE,
        ]
        cls.create_user_role_with_permissions(
            cls.user,
            cls.all_necessary_permissions,
            cls.business_area,
        )
        cls.program = ProgramFactory()
        cls.registration_data_import = RegistrationDataImportFactory(
            business_area=cls.business_area, program=cls.program
        )

        cls.household_1, cls.individual_1 = cls.create_household_and_individual(cls.program)
        cls.household_1.refresh_from_db()
        cls.household_2, cls.individual_2 = cls.create_household_and_individual(cls.program)
        cls.household_2.refresh_from_db()
        cls.household_3, cls.individual_3 = cls.create_household_and_individual(cls.program)
        cls.household_3.refresh_from_db()

        cls.data_collecting_type = DataCollectingType.objects.create(
            code="full", description="Full individual collected", active=True, type="STANDARD"
        )
        cls.data_collecting_type.limit_to.add(cls.business_area)
        generate_delivery_mechanisms()

    @patch("hct_mis_api.apps.payment.models.PaymentPlan.get_exchange_rate", return_value=2.0)
    def test_receiving_reconciliations_from_fsp(self, mock_get_exchange_rate: Any) -> None:
        beneficiary_group = BeneficiaryGroupFactory()
        create_programme_response = self.graphql_request(
            request_string=CREATE_PROGRAMME_MUTATION,
            context={"user": self.user},
            variables={
                "programData": {
                    "name": "NName",
                    "startDate": timezone.datetime(2022, 8, 24, tzinfo=utc),
                    "endDate": timezone.datetime(2022, 8, 31, tzinfo=utc),
                    "description": "desc",
                    "budget": "0.00",
                    "administrativeAreasOfImplementation": "",
                    "populationGoal": 0,
                    "frequencyOfPayments": "REGULAR",
                    "sector": "MULTI_PURPOSE",
                    "cashPlus": True,
                    "businessAreaSlug": self.business_area.slug,
                    "dataCollectingTypeCode": self.data_collecting_type.code,
                    "beneficiaryGroup": str(beneficiary_group.id),
                }
            },
        )
        program_id = create_programme_response["data"]["createProgram"]["program"]["id"]

        self.graphql_request(
            request_string=UPDATE_PROGRAM_MUTATION,
            context={"user": self.user},
            variables={
                "programData": {
                    "id": program_id,
                    "status": "ACTIVE",
                },
            },
        )

        program = Program.objects.get(id=decode_id_string_required(program_id))
        cycle = program.cycles.first()
        cycle.end_date = timezone.datetime(2022, 8, 24, tzinfo=utc).date()
        cycle.save()

        # create HH
        household_1, individual_1 = self.create_household_and_individual(program)
        household_1.refresh_from_db()
        household_2, individual_2 = self.create_household_and_individual(program)
        household_2.refresh_from_db()
        household_3, individual_3 = self.create_household_and_individual(program)
        household_3.refresh_from_db()

        account_type_bank = AccountType.objects.get(key="bank")
        dm_cash = DeliveryMechanism.objects.get(code="cash")
        dm_transfer = DeliveryMechanism.objects.get(code="transfer_to_account")

        for ind in [individual_1, individual_2, individual_3]:
            AccountFactory(individual=ind, account_type=account_type_bank)

        santander_fsp = FinancialServiceProviderFactory(
            name="Santander",
            distribution_limit=None,
            communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_XLSX,
        )
        santander_fsp.delivery_mechanisms.set([dm_cash, dm_transfer])
        FspXlsxTemplatePerDeliveryMechanismFactory(financial_service_provider=santander_fsp, delivery_mechanism=dm_cash)
        FspXlsxTemplatePerDeliveryMechanismFactory(
            financial_service_provider=santander_fsp, delivery_mechanism=dm_transfer
        )
        encoded_santander_fsp_id = encode_id_base64(santander_fsp.id, "FinancialServiceProvider")

        available_fsps_query_response = self.graphql_request(
            request_string=AVAILABLE_FSPS_FOR_DELIVERY_MECHANISMS_QUERY,
            context={"user": self.user},
        )
        assert "errors" not in available_fsps_query_response, available_fsps_query_response
        available_fsps_data = available_fsps_query_response["data"]["availableFspsForDeliveryMechanisms"]
        assert len(available_fsps_data) == 14

        dm_cash_resp = next((dm for dm in available_fsps_data if dm["deliveryMechanism"]["code"] == dm_cash.code), None)
        dm_transfer_resp = next(
            (dm for dm in available_fsps_data if dm["deliveryMechanism"]["code"] == dm_transfer.code), None
        )

        assert dm_cash_resp is not None
        assert dm_transfer_resp is not None
        assert dm_cash_resp["fsps"][0]["name"] == santander_fsp.name
        assert dm_cash_resp["deliveryMechanism"]["name"] == dm_cash.name
        assert dm_cash_resp["deliveryMechanism"]["code"] == dm_cash.code
        assert dm_transfer_resp["fsps"][0]["name"] == santander_fsp.name
        assert dm_transfer_resp["deliveryMechanism"]["name"] == dm_transfer.name
        assert dm_transfer_resp["deliveryMechanism"]["code"] == dm_transfer.code

        with patch(
            "hct_mis_api.apps.payment.services.payment_plan_services.transaction"
        ) as mock_prepare_payment_plan_task:
            create_payment_plan_response = self.graphql_request(
                request_string=CREATE_PAYMENT_PLAN_MUTATION,
                context={"user": self.user, "headers": {"Business-Area": self.business_area.slug}},
                variables={
                    "input": {
                        "name": "paymentPlanName",
                        "programCycleId": self.id_to_base64(cycle.id, "ProgramCycleNode"),
                        "excludedIds": "",
                        "targetingCriteria": {
                            "flagExcludeIfActiveAdjudicationTicket": False,
                            "flagExcludeIfOnSanctionList": False,
                            "rules": [
                                {
                                    "collectorsFiltersBlocks": [],
                                    "householdsFiltersBlocks": [],
                                    "householdIds": f"{household_1.unicef_id}, {household_2.unicef_id}, {household_3}",
                                    "individualIds": "",
                                    "individualsFiltersBlocks": [],
                                }
                            ],
                        },
                        "fspId": encoded_santander_fsp_id,
                        "deliveryMechanismCode": dm_cash.code,
                    },
                },
            )
            assert mock_prepare_payment_plan_task.on_commit.call_count == 1
            mock_prepare_payment_plan_task.on_commit.call_args[0][0]()  # call real func

        assert "errors" not in create_payment_plan_response, create_payment_plan_response
        assert create_payment_plan_response["data"]["createPaymentPlan"]["paymentPlan"]["status"] == "TP_OPEN"
        encoded_payment_plan_id = create_payment_plan_response["data"]["createPaymentPlan"]["paymentPlan"]["id"]
        payment_plan_id = decode_id_string(encoded_payment_plan_id)

        self.update_partner_access_to_program(self.user.partner, program)

        locked_pp_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "TP_LOCK",
                }
            },
        )
        self.assertEqual(locked_pp_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"], "TP_LOCKED")

        finalize_pp_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "DRAFT",
                }
            },
        )
        self.assertEqual(finalize_pp_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"], "DRAFT")

        # all cycles should have end_date before creation new one
        ProgramCycle.objects.filter(program_id=decode_id_string(program_id)).update(
            end_date=timezone.datetime(2022, 8, 25, tzinfo=utc).date(), title="NEW NEW NAME"
        )

        # OPEN PP
        open_payment_plan_response = self.graphql_request(
            request_string=OPEN_PAYMENT_PLAN_MUTATION,
            context={"user": self.user, "headers": {"Business-Area": self.business_area.slug}},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "dispersionStartDate": (timezone.now() - timedelta(days=1)).strftime("%Y-%m-%d"),
                    "dispersionEndDate": (timezone.now() + timedelta(days=1)).strftime("%Y-%m-%d"),
                    "currency": "USD",
                }
            },
        )
        assert "errors" not in open_payment_plan_response, open_payment_plan_response
        assert "OPEN" == open_payment_plan_response["data"]["openPaymentPlan"]["paymentPlan"]["status"]

        # check if Cycle is active
        assert ProgramCycle.objects.filter(title="NEW NEW NAME").first().status == "ACTIVE"

        payment_plan = PaymentPlan.objects.get(id=payment_plan_id)
        payment = PaymentFactory(
            parent=payment_plan,
            parent_split=payment_plan.splits.first(),
            business_area=self.business_area,
            household=self.household_1,
            collector=self.individual_1,
            delivery_type=dm_cash,
            entitlement_quantity=1000,
            entitlement_quantity_usd=100,
            delivered_quantity=None,
            delivered_quantity_usd=None,
            financial_service_provider=santander_fsp,
            currency="PLN",
            has_valid_wallet=True,
            delivery_date=None,
        )
        self.assertEqual(payment.entitlement_quantity, 1000)
        create_payment_plan_snapshot_data(payment_plan)

        lock_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "LOCK",
                }
            },
        )
        assert "errors" not in lock_payment_plan_response, lock_payment_plan_response
        assert lock_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"] == "LOCKED"

        rule = RuleFactory(name="Rule", type=Rule.TYPE_PAYMENT_PLAN)
        RuleCommitFactory(definition="result.value=Decimal('500')", rule=rule)

        self.assertEqual(payment_plan.background_action_status, None)

        with patch("hct_mis_api.apps.payment.mutations.payment_plan_apply_engine_rule") as mock:
            payment_plan.refresh_from_db()
            set_steficon_response = self.graphql_request(
                request_string=SET_STEFICON_RULE_MUTATION,
                context={"user": self.user},
                variables={
                    "paymentPlanId": encoded_payment_plan_id,
                    "steficonRuleId": encode_id_base64(rule.id, "Rule"),
                    "version": payment_plan.version,
                },
            )
            assert "errors" not in set_steficon_response, set_steficon_response
            assert mock.delay.call_count == 1
            call_args = mock.delay.call_args[0]
            payment_plan_apply_engine_rule(*call_args)

        payment_plan.refresh_from_db()
        self.assertEqual(payment_plan.background_action_status, None)

        payment.refresh_from_db()
        self.assertEqual(payment.entitlement_quantity, 500)

        lock_fsp_in_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "LOCK_FSP",
                }
            },
        )
        assert "errors" not in lock_fsp_in_payment_plan_response, lock_fsp_in_payment_plan_response
        self.assertEqual(
            lock_fsp_in_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "LOCKED_FSP",
        )

        send_for_approval_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "SEND_FOR_APPROVAL",
                }
            },
        )
        assert "errors" not in send_for_approval_payment_plan_response, send_for_approval_payment_plan_response
        self.assertEqual(
            send_for_approval_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "IN_APPROVAL",
        )

        approve_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "APPROVE",
                }
            },
        )
        assert "errors" not in approve_payment_plan_response, approve_payment_plan_response
        self.assertEqual(
            approve_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "IN_AUTHORIZATION",
        )

        authorize_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "AUTHORIZE",
                }
            },
        )
        assert "errors" not in authorize_payment_plan_response, authorize_payment_plan_response
        self.assertEqual(
            authorize_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "IN_REVIEW",
        )

        review_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "REVIEW",
                }
            },
        )
        assert "errors" not in review_payment_plan_response, review_payment_plan_response
        self.assertEqual(
            review_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "ACCEPTED",
        )

        payment_plan.refresh_from_db()
        self.assertEqual(payment_plan.background_action_status, None)

        with patch(
            "hct_mis_api.apps.payment.services.payment_plan_services.create_payment_plan_payment_list_xlsx_per_fsp"
        ) as mock_export:
            export_file_mutation = self.graphql_request(
                request_string=EXPORT_XLSX_PER_FSP_MUTATION,
                context={"user": self.user},
                variables={
                    "paymentPlanId": encoded_payment_plan_id,
                    "fspXlsxTemplateId": "",
                },
            )
            assert "errors" not in export_file_mutation, export_file_mutation
            assert mock_export.delay.call_count == 1
            call_args = mock_export.delay.call_args[0]
            create_payment_plan_payment_list_xlsx_per_fsp(*call_args)

        payment_plan.refresh_from_db()
        zip_file = payment_plan.export_file_per_fsp
        assert zip_file is not None

        with tempfile.TemporaryDirectory() as temp_dir:
            self.assertEqual(len(os.listdir(temp_dir)), 0)

            assert zip_file.file is not None
            with ZipFile(zip_file.file, "r") as zip_ref:
                self.assertEqual(len(zip_ref.namelist()), 1)
                zip_ref.extractall(temp_dir)

            self.assertEqual(len(os.listdir(temp_dir)), 1)

            file_name = os.listdir(temp_dir)[0]
            assert file_name.endswith(".xlsx")
            file_path = os.path.join(temp_dir, file_name)

            workbook = load_workbook(file_path)
            assert workbook.sheetnames == ["Santander"], workbook.sheetnames

            sheet = workbook["Santander"]
            assert sheet.max_row == 5, sheet.max_row

            self.assertEqual(sheet.cell(row=1, column=1).value, "payment_id")
            assert payment_plan.payment_items.count() == 4
            payment = payment_plan.eligible_payments.filter(household=household_1).first()
            # check if there is the same HH
            self.assertEqual(payment.household.unicef_id, household_1.unicef_id)

            self.assertEqual(sheet.cell(row=2, column=1).value, payment.unicef_id)  # unintuitive
            self.assertEqual(sheet.cell(row=1, column=2).value, "household_id")
            self.assertEqual(sheet.cell(row=2, column=2).value, household_1.unicef_id)
            self.assertEqual(sheet.cell(row=1, column=3).value, "household_size")
            self.assertEqual(sheet.cell(row=2, column=3).value, household_1.size)
            self.assertEqual(sheet.cell(row=1, column=4).value, "collector_name")
            self.assertEqual(sheet.cell(row=2, column=4).value, payment.collector.full_name)
            self.assertEqual(sheet.cell(row=1, column=5).value, "alternate_collector_full_name")
            self.assertEqual(sheet.cell(row=2, column=5).value, None)
            self.assertEqual(sheet.cell(row=1, column=6).value, "alternate_collector_given_name")
            self.assertEqual(sheet.cell(row=2, column=6).value, None)
            self.assertEqual(sheet.cell(row=1, column=7).value, "alternate_collector_family_name")
            self.assertEqual(sheet.cell(row=2, column=7).value, None)
            self.assertEqual(sheet.cell(row=1, column=8).value, "alternate_collector_middle_name")
            self.assertEqual(sheet.cell(row=2, column=8).value, None)
            self.assertEqual(sheet.cell(row=1, column=9).value, "alternate_collector_phone_no")
            self.assertEqual(sheet.cell(row=2, column=9).value, None)
            self.assertEqual(sheet.cell(row=1, column=10).value, "alternate_collector_document_numbers")
            self.assertEqual(sheet.cell(row=2, column=10).value, None)
            self.assertEqual(sheet.cell(row=1, column=11).value, "alternate_collector_sex")
            self.assertEqual(sheet.cell(row=2, column=11).value, None)
            self.assertEqual(sheet.cell(row=1, column=12).value, "payment_channel")
            self.assertEqual(sheet.cell(row=2, column=12).value, "Cash")
            self.assertEqual(sheet.cell(row=1, column=13).value, "fsp_name")
            self.assertEqual(sheet.cell(row=2, column=13).value, payment.financial_service_provider.name)
            self.assertEqual(sheet.cell(row=1, column=14).value, "currency")
            self.assertEqual(sheet.cell(row=2, column=14).value, payment.currency)
            self.assertEqual(sheet.cell(row=1, column=15).value, "entitlement_quantity")
            self.assertEqual(sheet.cell(row=2, column=15).value, payment.entitlement_quantity)
            self.assertEqual(sheet.cell(row=1, column=16).value, "entitlement_quantity_usd")
            self.assertEqual(sheet.cell(row=2, column=16).value, payment.entitlement_quantity_usd)
            self.assertEqual(sheet.cell(row=1, column=17).value, "delivered_quantity")
            self.assertEqual(sheet.cell(row=2, column=17).value, None)
            self.assertEqual(sheet.cell(row=1, column=18).value, "delivery_date")
            # self.assertEqual(sheet.cell(row=2, column=17).value, str(payment.delivery_date))

            payment.refresh_from_db()
            self.assertEqual(payment.entitlement_quantity, 500)
            self.assertEqual(payment.delivered_quantity, None)
            self.assertEqual(payment.status, Payment.STATUS_PENDING)
            self.assertEqual(payment_plan.is_reconciled, False)

            filled_file_name = "filled.xlsx"
            filled_file_path = os.path.join(temp_dir, filled_file_name)

            # update xls, delivered_quantity != entitlement_quantity
            sheet.cell(
                row=2, column=FinancialServiceProviderXlsxTemplate.DEFAULT_COLUMNS.index("delivered_quantity")
            ).value = 666
            workbook.save(filled_file_path)

            with open(filled_file_path, "rb") as file:
                uploaded_file = SimpleUploadedFile(filled_file_name, file.read())
                with patch("hct_mis_api.apps.payment.services.payment_plan_services.transaction") as mock_import:
                    import_mutation_response = self.graphql_request(
                        request_string=IMPORT_XLSX_PER_FSP_MUTATION,
                        context={"user": self.user},
                        variables={
                            "paymentPlanId": encoded_payment_plan_id,
                            "file": uploaded_file,
                        },
                    )
                    assert (
                        "errors" in import_mutation_response["data"]["importXlsxPaymentPlanPaymentListPerFsp"]
                    ), import_mutation_response
                    assert (
                        import_mutation_response["data"]["importXlsxPaymentPlanPaymentListPerFsp"]["errors"][0][
                            "message"
                        ]
                        == f"Payment {payment.unicef_id}: Delivered quantity 666.00 is bigger than Entitlement quantity 500.00"
                    ), import_mutation_response

            # update xls, delivered_quantity == entitlement_quantity
            sheet.cell(
                row=2, column=FinancialServiceProviderXlsxTemplate.DEFAULT_COLUMNS.index("delivered_quantity")
            ).value = payment.entitlement_quantity
            workbook.save(filled_file_path)

            with open(filled_file_path, "rb") as file:
                uploaded_file = SimpleUploadedFile(filled_file_name, file.read())
                with patch("hct_mis_api.apps.payment.services.payment_plan_services.transaction") as mock_import:
                    import_mutation_response = self.graphql_request(
                        request_string=IMPORT_XLSX_PER_FSP_MUTATION,
                        context={"user": self.user},
                        variables={
                            "paymentPlanId": encoded_payment_plan_id,
                            "file": uploaded_file,
                        },
                    )
                    assert (
                        import_mutation_response["data"]["importXlsxPaymentPlanPaymentListPerFsp"]["errors"] is None
                    ), import_mutation_response
                    assert mock_import.on_commit.call_count == 1
                    mock_import.on_commit.call_args[0][0]()  # call real func

            payment.refresh_from_db()
            payment.household.refresh_from_db()
            self.assertEqual(payment.entitlement_quantity, 500)
            self.assertEqual(payment.delivered_quantity, 500)
            self.assertEqual(payment.status, Payment.STATUS_DISTRIBUTION_SUCCESS)
            self.assertEqual(payment.household.total_cash_received, 500)
            self.assertEqual(payment.household.total_cash_received_usd, 250)
            self.assertEqual(payment_plan.eligible_payments.exclude(status__in=Payment.PENDING_STATUSES).count(), 1)
            self.assertEqual(payment_plan.eligible_payments.count(), 4)
            self.assertFalse(payment_plan.is_reconciled)

    @parameterized.expand(
        [
            (-1, Decimal(f"{round(100.00, 2):.2f}"), None, Payment.STATUS_ERROR),
            (0, Decimal(0), Decimal(0), Payment.STATUS_NOT_DISTRIBUTED),
            (
                400.10,
                Decimal(f"{round(400.23, 2):.2f}"),
                Decimal(f"{round(400.10, 2):.2f}"),
                Payment.STATUS_DISTRIBUTION_PARTIAL,
            ),
            (
                400.23,
                Decimal(f"{round(400.23, 2):.2f}"),
                Decimal(f"{round(400.23, 2):.2f}"),
                Payment.STATUS_DISTRIBUTION_SUCCESS,
            ),
            (
                500.00,
                Decimal(f"{round(500.00, 2):.2f}"),
                Decimal(f"{round(500.00, 2):.2f}"),
                Payment.STATUS_DISTRIBUTION_SUCCESS,
            ),
            (
                500,
                Decimal(f"{round(500.00, 2):.2f}"),
                Decimal(f"{round(500.00, 2):.2f}"),
                Payment.STATUS_DISTRIBUTION_SUCCESS,
            ),
            (600.00, Decimal(f"{round(100.00, 2):.2f}"), None, None),
            ("-1", Decimal(f"{round(100.00, 2):.2f}"), None, Payment.STATUS_ERROR),
            ("0", Decimal(0), Decimal(0), Payment.STATUS_NOT_DISTRIBUTED),
            (
                "400.10",
                Decimal(f"{round(400.23, 2):.2f}"),
                Decimal(f"{round(400.10, 2):.2f}"),
                Payment.STATUS_DISTRIBUTION_PARTIAL,
            ),
            (
                "400.23",
                Decimal(f"{round(400.23, 2):.2f}"),
                Decimal(f"{round(400.23, 2):.2f}"),
                Payment.STATUS_DISTRIBUTION_SUCCESS,
            ),
            (
                "500",
                Decimal(f"{round(500.00, 2):.2f}"),
                Decimal(f"{round(500.00, 2):.2f}"),
                Payment.STATUS_DISTRIBUTION_SUCCESS,
            ),
            (
                "500.00",
                Decimal(f"{round(500.00, 2):.2f}"),
                Decimal(f"{round(500.00, 2):.2f}"),
                Payment.STATUS_DISTRIBUTION_SUCCESS,
            ),
        ]
    )
    def test_receiving_payment_reconciliations_status(
        self,
        delivered_quantity: float,
        entitlement_quantity: Decimal,
        expected_delivered_quantity: Decimal,
        expected_status: str,
    ) -> None:
        service = XlsxPaymentPlanImportPerFspService(PaymentPlanFactory(created_by=self.user), None)  # type: ignore

        if not expected_status:
            with self.assertRaisesMessage(
                service.XlsxPaymentPlanImportPerFspServiceException,
                f"Invalid delivered_quantity {delivered_quantity} provided for payment_id xx",
            ):
                service._get_delivered_quantity_status_and_value(delivered_quantity, entitlement_quantity, "xx")

        else:
            status, value = service._get_delivered_quantity_status_and_value(
                delivered_quantity, entitlement_quantity, "xx"
            )
            self.assertEqual(status, expected_status)
            self.assertEqual(value, expected_delivered_quantity)

    def test_xlsx_payment_plan_import_per_fsp_service_import_row(self) -> None:
        pp = PaymentPlanFactory(
            status=PaymentPlan.Status.FINISHED,
            created_by=self.user,
        )
        pp.refresh_from_db()
        pvs = PaymentVerificationSummaryFactory()
        pvs.payment_plan = pp
        pvs.save()
        pvp = PaymentVerificationPlanFactory(
            payment_plan=pp,
            verification_channel=PaymentVerificationPlan.VERIFICATION_CHANNEL_MANUAL,
            status=PaymentVerificationPlan.STATUS_ACTIVE,
        )

        payment_1 = PaymentFactory(
            parent=PaymentPlan.objects.get(id=pp.id),
            business_area=self.business_area,
            household=self.household_1,
            collector=self.individual_1,
            delivery_type=None,
            entitlement_quantity=1111,
            entitlement_quantity_usd=100,
            delivered_quantity=1000,
            delivered_quantity_usd=99,
            financial_service_provider=None,
            currency="PLN",
        )
        payment_2 = PaymentFactory(
            parent=PaymentPlan.objects.get(id=pp.id),
            business_area=self.business_area,
            household=self.household_2,
            collector=self.individual_2,
            delivery_type=None,
            entitlement_quantity=2222,
            entitlement_quantity_usd=100,
            delivered_quantity=2000,
            delivered_quantity_usd=500,
            financial_service_provider=None,
            currency="PLN",
        )
        payment_3 = PaymentFactory(
            parent=PaymentPlan.objects.get(id=pp.id),
            business_area=self.business_area,
            household=self.household_3,
            collector=self.individual_3,
            delivery_type=None,
            entitlement_quantity=3333,
            entitlement_quantity_usd=300,
            delivered_quantity=3000,
            delivered_quantity_usd=290,
            financial_service_provider=None,
            currency="PLN",
        )
        verification_1 = PaymentVerificationFactory(
            payment_verification_plan=pvp,
            payment=payment_1,
            status=PaymentVerification.STATUS_RECEIVED_WITH_ISSUES,
            received_amount=999,
        )
        verification_2 = PaymentVerificationFactory(
            payment_verification_plan=pvp,
            payment=payment_2,
            status=PaymentVerification.STATUS_RECEIVED,
            received_amount=500,
        )
        verification_3 = PaymentVerificationFactory(
            payment_verification_plan=pvp,
            payment=payment_3,
            status=PaymentVerification.STATUS_PENDING,
            received_amount=None,
        )
        import_xlsx_service = XlsxPaymentPlanImportPerFspService(pp, io.BytesIO())
        import_xlsx_service.xlsx_headers = ["payment_id", "delivered_quantity", "delivery_date"]

        import_xlsx_service.payments_dict[str(payment_1.pk)] = payment_1
        import_xlsx_service.payments_dict[str(payment_2.pk)] = payment_2
        import_xlsx_service.payments_dict[str(payment_3.pk)] = payment_3
        row = namedtuple(
            "row",
            [
                "value",
            ],
        )

        import_xlsx_service._import_row(
            [row(str(payment_1.id)), row(999), row(pytz.utc.localize(datetime.datetime(2023, 5, 12)))], 1
        )
        import_xlsx_service._import_row(
            [row(str(payment_2.id)), row(100), row(pytz.utc.localize(datetime.datetime(2022, 12, 14)))], 1
        )
        import_xlsx_service._import_row(
            [row(str(payment_3.id)), row(2999), row(pytz.utc.localize(datetime.datetime(2021, 7, 25)))], 1
        )
        payment_1.save()
        payment_2.save()
        payment_3.save()
        # Update payment Verification
        PaymentVerification.objects.bulk_update(
            import_xlsx_service.payment_verifications_to_save, ("status", "status_date")
        )
        payment_1.refresh_from_db()
        payment_2.refresh_from_db()
        payment_3.refresh_from_db()
        verification_1.refresh_from_db()
        verification_2.refresh_from_db()
        verification_3.refresh_from_db()

        self.assertEqual(payment_1.delivered_quantity, 999)
        self.assertEqual(verification_1.received_amount, 999)
        self.assertEqual(verification_1.status, PaymentVerification.STATUS_RECEIVED)

        self.assertEqual(payment_2.delivered_quantity, 100)
        self.assertEqual(verification_2.received_amount, 500)
        self.assertEqual(verification_2.status, PaymentVerification.STATUS_RECEIVED_WITH_ISSUES)

        self.assertEqual(payment_3.delivered_quantity, 2999)
        self.assertEqual(verification_3.received_amount, None)
        self.assertEqual(verification_3.status, PaymentVerification.STATUS_PENDING)

    def test_payment_plan_is_fully_delivered(self) -> None:
        payment_plan = PaymentPlanFactory(
            status=PaymentPlan.Status.ACCEPTED,
            created_by=self.user,
        )
        for hh, ind in [
            (self.household_1, self.individual_1),
            (self.household_2, self.individual_2),
            (self.household_3, self.individual_3),
        ]:
            PaymentFactory(
                parent=payment_plan,
                business_area=self.business_area,
                household=hh,
                collector=ind,
                delivery_type=None,
                entitlement_quantity=999,
                entitlement_quantity_usd=10,
                delivered_quantity=999,
                delivered_quantity_usd=10,
                financial_service_provider=None,
                currency="PLN",
            )
        payment_plan.status_finished()
        payment_plan.save()
        payment_plan.refresh_from_db()
        self.assertTrue(
            all(
                [
                    payment.entitlement_quantity == payment.delivered_quantity
                    for payment in payment_plan.eligible_payments
                ]
            )
        )

    @freeze_time("2023-12-12")
    def test_follow_up_pp_entitlements_can_be_changed_with_steficon_rule(self) -> None:
        pp = PaymentPlanFactory(
            is_follow_up=True,
            status=PaymentPlan.Status.LOCKED,
            created_by=self.user,
        )
        pp.unicef_id = "PP-0060-23-00000002"
        pp.save()

        rule = RuleFactory(name="SomeRule")

        self.snapshot_graphql_request(
            request_string=SET_STEFICON_RULE_MUTATION,
            context={"user": self.user},
            variables={
                "paymentPlanId": encode_id_base64(pp.id, "PaymentPlan"),
                "steficonRuleId": encode_id_base64(rule.id, "Rule"),
                "version": pp.version,
            },
        )

    def test_apply_steficon_rule_with_wrong_payment_plan_status(self) -> None:
        payment_plan = PaymentPlanFactory(
            status=PaymentPlan.Status.OPEN,
            created_by=self.user,
        )
        rule = RuleFactory(name="SomeRule")

        self.snapshot_graphql_request(
            request_string=SET_STEFICON_RULE_MUTATION,
            context={"user": self.user},
            variables={
                "paymentPlanId": encode_id_base64(payment_plan.id, "PaymentPlan"),
                "steficonRuleId": encode_id_base64(rule.id, "Rule"),
                "version": payment_plan.version,
            },
        )

        payment_plan.status = PaymentPlan.Status.LOCKED
        payment_plan.background_action_status = PaymentPlan.BackgroundActionStatus.RULE_ENGINE_RUN
        payment_plan.save()
        payment_plan.refresh_from_db(fields=["status", "background_action_status"])

        self.assertEqual(payment_plan.status, PaymentPlan.Status.LOCKED)
        self.assertEqual(payment_plan.background_action_status, PaymentPlan.BackgroundActionStatus.RULE_ENGINE_RUN)
        self.snapshot_graphql_request(
            request_string=SET_STEFICON_RULE_MUTATION,
            context={"user": self.user},
            variables={
                "paymentPlanId": encode_id_base64(payment_plan.id, "PaymentPlan"),
                "steficonRuleId": encode_id_base64(rule.id, "Rule"),
                "version": payment_plan.version,
            },
        )

    def test_error_message_when_engine_rule_not_enabled_or_deprecated(self) -> None:
        payment_plan = PaymentPlanFactory(
            status=PaymentPlan.Status.LOCKED,
            created_by=self.user,
        )
        rule_not_enabled = RuleFactory(enabled=False)
        rule_deprecated = RuleFactory(deprecated=True)

        for rule in [rule_not_enabled, rule_deprecated]:
            self.snapshot_graphql_request(
                request_string=SET_STEFICON_RULE_MUTATION,
                context={"user": self.user},
                variables={
                    "paymentPlanId": encode_id_base64(payment_plan.id, "PaymentPlan"),
                    "steficonRuleId": encode_id_base64(rule.id, "Rule"),
                    "version": payment_plan.version,
                },
            )

    def test_follow_up_pp_entitlements_updated_with_file(self) -> None:
        content = Path(f"{settings.TESTS_ROOT}/apps/payment/test_file/pp_payment_list_valid.xlsx").read_bytes()
        pp = PaymentPlanFactory(
            is_follow_up=True,
            status=PaymentPlan.Status.LOCKED,
            created_by=self.user,
        )

        self.snapshot_graphql_request(
            request_string=IMPORT_XLSX_PP_MUTATION,
            context={"user": self.user},
            variables={
                "paymentPlanId": encode_id_base64(pp.id, "PaymentPlan"),
                "file": BytesIO(content),
            },
        )

    def test_correct_message_displayed_when_file_is_protected(self) -> None:
        content = Path(f"{settings.TESTS_ROOT}/apps/payment/test_file/import_file_protected.xlsx").read_bytes()
        dm_cash = DeliveryMechanism.objects.get(code="cash")
        financial_service_provider1 = FinancialServiceProviderFactory(
            communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_XLSX
        )
        financial_service_provider1.delivery_mechanisms.add(dm_cash)
        pp = PaymentPlanFactory(
            status=PaymentPlan.Status.ACCEPTED,
            created_by=self.user,
            financial_service_provider=financial_service_provider1,
            delivery_mechanism=dm_cash,
        )

        self.snapshot_graphql_request(
            request_string=IMPORT_XLSX_PER_FSP_MUTATION,
            context={"user": self.user},
            variables={
                "paymentPlanId": encode_id_base64(pp.id, "PaymentPlan"),
                "file": BytesIO(content),
            },
        )

    def test_import_with_wrong_payment_plan_status(self) -> None:
        payment_plan = PaymentPlanFactory(
            status=PaymentPlan.Status.OPEN,
            created_by=self.user,
        )

        self.assertEqual(payment_plan.status, PaymentPlan.Status.OPEN)
        self.snapshot_graphql_request(
            request_string=IMPORT_XLSX_PER_FSP_MUTATION,
            context={"user": self.user},
            variables={
                "paymentPlanId": encode_id_base64(payment_plan.id, "PaymentPlan"),
                "file": BytesIO(b"some data"),
            },
        )

    def test_export_xlsx_per_fsp_with_auth_code(self) -> None:
        dm_cash = DeliveryMechanism.objects.get(code="cash")
        fsp = FinancialServiceProviderFactory(
            communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_API, payment_gateway_id="ABC_333"
        )
        payment_plan = PaymentPlanFactory(
            status=PaymentPlan.Status.FINISHED,
            created_by=self.user,
            financial_service_provider=fsp,
            delivery_mechanism=dm_cash,
        )
        payment_1 = PaymentFactory(parent=payment_plan, fsp_auth_code="TestAuthCode")
        xlsx_template = FinancialServiceProviderXlsxTemplateFactory()
        variables = {
            "paymentPlanId": encode_id_base64_required(str(payment_plan.pk), "PaymentPlan"),
            "fspXlsxTemplateId": encode_id_base64_required(
                str(xlsx_template.pk), "FinancialServiceProviderXlsxTemplate"
            ),
        }
        self.snapshot_graphql_request(
            request_string=EXPORT_XLSX_PER_FSP_MUTATION_AUTH_CODE, context={"user": self.user}, variables=variables
        )
        # upd payment status
        payment_1.status = Payment.STATUS_SENT_TO_FSP
        payment_1.save()
        self.snapshot_graphql_request(
            request_string=EXPORT_XLSX_PER_FSP_MUTATION_AUTH_CODE, context={"user": self.user}, variables=variables
        )
        payment_plan.refresh_from_db()
        self.assertEqual(FileTemp.objects.filter(object_id=payment_plan.id).count(), 1)
        self.assertIsNotNone(payment_plan.export_file_per_fsp)
        self.assertEqual(
            payment_plan.export_file_per_fsp_id, FileTemp.objects.filter(object_id=payment_plan.id).first().pk
        )

    def test_export_xlsx_per_fsp_error_msg(self) -> None:
        fsp = FinancialServiceProviderFactory(
            communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_API, payment_gateway_id="ABC_aaa"
        )
        dm_cash = DeliveryMechanism.objects.get(code="cash")
        payment_plan = PaymentPlanFactory(
            status=PaymentPlan.Status.LOCKED,
            created_by=self.user,
            financial_service_provider=fsp,
            delivery_mechanism=dm_cash,
        )
        xlsx_template = FinancialServiceProviderXlsxTemplateFactory()
        variables = {
            "paymentPlanId": encode_id_base64_required(str(payment_plan.pk), "PaymentPlan"),
            "fspXlsxTemplateId": encode_id_base64_required(
                str(xlsx_template.pk), "FinancialServiceProviderXlsxTemplate"
            ),
        }
        # wrong Payment Plan status
        self.snapshot_graphql_request(
            request_string=EXPORT_XLSX_PER_FSP_MUTATION_AUTH_CODE, context={"user": self.user}, variables=variables
        )
        # upd payment plan status
        payment_plan.status = PaymentPlan.Status.ACCEPTED
        payment_plan.save()
        # no eligible payments
        self.snapshot_graphql_request(
            request_string=EXPORT_XLSX_PER_FSP_MUTATION_AUTH_CODE, context={"user": self.user}, variables=variables
        )
        # create temp file
        file = FileTemp.objects.create(
            object_id=payment_plan.pk,
            content_type=get_content_type_for_model(payment_plan),
            created=timezone.now(),
            file=ContentFile(b"Aaa", f"Test_File_ABC{payment_plan.pk}.xlsx"),
        )
        PaymentFactory(parent=payment_plan, fsp_auth_code="TestAuthCode", status=Payment.STATUS_SENT_TO_FSP)
        payment_plan.export_file_per_fsp = file
        payment_plan.save()
        # Payment Plan regenerate new file
        self.snapshot_graphql_request(
            request_string=EXPORT_XLSX_PER_FSP_MUTATION_AUTH_CODE, context={"user": self.user}, variables=variables
        )
        payment_plan.refresh_from_db()
        self.assertIsNotNone(payment_plan.export_file_per_fsp)
        self.assertNotEqual(str(payment_plan.export_file_per_fsp.pk), str(file.pk))
