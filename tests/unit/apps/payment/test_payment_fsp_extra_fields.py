from datetime import date
from io import BytesIO

import openpyxl
import pytest

from extras.test_utils.factories.payment import (
    FinancialServiceProviderXlsxTemplateFactory,
    PaymentFactory,
    PaymentHouseholdSnapshotFactory,
    PaymentPlanFactory,
)
from hope.apps.payment.api.serializers import PaymentDetailSerializer
from hope.apps.payment.xlsx.xlsx_payment_plan_delivery_export_service import (
    XlsxPaymentPlanDeliveryExportService,
)
from hope.apps.payment.xlsx.xlsx_payment_plan_delivery_import_service import (
    XlsxPaymentPlanDeliveryImportService,
)
from hope.apps.payment.xlsx.xlsx_payment_plan_fsp_extra_fields_export_service import (
    XlsxPaymentPlanFspExtraFieldsExportService,
)
from hope.apps.payment.xlsx.xlsx_payment_plan_fsp_extra_fields_import_service import (
    XlsxPaymentPlanFspExtraFieldsImportService,
)
from hope.models import PaymentPlan
from hope.one_time_scripts.migrate_payment_extras import migrate_payment_extras

pytestmark = pytest.mark.django_db


@pytest.fixture
def payment_plan():
    return PaymentPlanFactory(status=PaymentPlan.Status.LOCKED)


@pytest.fixture
def payments(payment_plan):
    first_payment = PaymentFactory(
        parent=payment_plan,
        program=payment_plan.program,
        unicef_id="PAYMENT-001",
        extras={
            "extra_fields": {"reconciliation_code": "REC-001"},
            "fsp_extra_fields": {
                "empty_field": "keep-empty",
                "keep": "keep-existing",
                "reference": "old-reference",
            },
        },
    )
    second_payment = PaymentFactory(
        parent=payment_plan,
        program=payment_plan.program,
        unicef_id="PAYMENT-002",
        extras={
            "extra_fields": {"reconciliation_code": "REC-002"},
            "fsp_extra_fields": {"other": "second-payment"},
        },
    )
    return first_payment, second_payment


@pytest.fixture
def fsp_extra_fields_import_file(payments):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["payment_id", "reference", "empty_field", "date_field"])
    worksheet.append([payments[0].unicef_id, "new-reference", None, date(2026, 7, 23)])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


@pytest.fixture
def payment_gateway_header_import_file(payments):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["payment_id", "amount"])
    worksheet.append([payments[0].unicef_id, 10])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


@pytest.fixture
def reconciliation_file(payments):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["payment_id", "delivered_quantity", "reference", "returned_code"])
    worksheet.append([payments[0].unicef_id, 10, "returned-reference", "RETURNED-001"])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


@pytest.fixture
def sparse_reconciliation_file(payments):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["payment_id", "delivered_quantity", "reference", "returned_code"])
    worksheet.append([payments[1].unicef_id, 10, "returned-reference", "RETURNED-002"])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


@pytest.fixture
def dynamic_system_header_import_file(payments, request):
    template_field, header = request.param
    FinancialServiceProviderXlsxTemplateFactory(**{template_field: [header]})
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["payment_id", header])
    worksheet.append([payments[0].unicef_id, "FSP-VALUE"])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream, header


@pytest.fixture
def account_system_header_import_file(payments):
    PaymentHouseholdSnapshotFactory(
        payment=payments[0],
        snapshot_data={
            "primary_collector": {
                "account_data": {
                    "financial_institution_name": "Example Bank",
                    "number": "123456",
                }
            }
        },
    )
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["payment_id", "number"])
    worksheet.append([payments[0].unicef_id, "FSP-VALUE"])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


@pytest.fixture
def migration_payments(payment_plan):
    flat_payment = PaymentFactory(
        parent=payment_plan,
        program=payment_plan.program,
        extras={"legacy": "value"},
    )
    empty_payment = PaymentFactory(
        parent=payment_plan,
        program=payment_plan.program,
        extras={},
    )
    namespaced_payment = PaymentFactory(
        parent=payment_plan,
        program=payment_plan.program,
        extras={"fsp_extra_fields": {"already": "namespaced"}},
    )
    return flat_payment, empty_payment, namespaced_payment


@pytest.fixture
def fsp_xlsx_template():
    return FinancialServiceProviderXlsxTemplateFactory(columns=["payment_id"])


def test_post_deployment_script_migrates_flat_extras_and_preserves_other_values(migration_payments, capsys):
    migrate_payment_extras(blocks_per_batch=1_000_000, sleep_seconds=0)

    migration_payments[0].refresh_from_db()
    migration_payments[1].refresh_from_db()
    migration_payments[2].refresh_from_db()
    assert migration_payments[0].extras == {"extra_fields": {"legacy": "value"}}
    assert migration_payments[1].extras == {}
    assert migration_payments[2].extras == {"fsp_extra_fields": {"already": "namespaced"}}
    assert "Remaining legacy Payment.extras rows: 0." in capsys.readouterr().out


def test_fsp_extra_fields_template_contains_only_fsp_fields(
    payment_plan,
    payments,
    django_assert_num_queries,
):
    with django_assert_num_queries(2):
        workbook = XlsxPaymentPlanFspExtraFieldsExportService(payment_plan).generate_workbook()

    rows = list(workbook.active.values)
    assert rows == [
        ("payment_id", "empty_field", "keep", "other", "reference"),
        ("PAYMENT-001", "keep-empty", "keep-existing", "", "old-reference"),
        ("PAYMENT-002", "", "", "second-payment", ""),
    ]
    assert "reconciliation_code" not in rows[0]


def test_manual_fsp_export_appends_fsp_extra_fields(
    payment_plan,
    payments,
    fsp_xlsx_template,
):
    service = XlsxPaymentPlanDeliveryExportService(payment_plan)
    headers = service.prepare_headers(fsp_xlsx_template)

    row = service.get_payment_row(payments[0])

    assert headers == ["payment_id", "empty_field", "keep", "other", "reference"]
    assert row == ["PAYMENT-001", "keep-empty", "keep-existing", "", "old-reference"]


def test_fsp_extra_fields_import_merges_non_empty_values(
    payment_plan,
    payments,
    fsp_extra_fields_import_file,
    django_assert_num_queries,
):
    old_signature = payments[0].signature_hash
    with django_assert_num_queries(1):
        service = XlsxPaymentPlanFspExtraFieldsImportService(payment_plan, fsp_extra_fields_import_file)
    service.open_workbook()
    service.validate()

    updated_count = service.import_payment_list()

    payments[0].refresh_from_db()
    payments[1].refresh_from_db()
    assert service.errors == []
    assert updated_count == 1
    assert payments[0].extra_fields == {"reconciliation_code": "REC-001"}
    assert payments[0].fsp_extra_fields == {
        "date_field": "2026-07-23T00:00:00",
        "empty_field": "keep-empty",
        "keep": "keep-existing",
        "reference": "new-reference",
    }
    assert payments[0].signature_hash != old_signature
    assert payments[1].fsp_extra_fields == {"other": "second-payment"}


def test_fsp_extra_fields_import_allows_payment_gateway_payload_header(
    payment_plan,
    payments,
    payment_gateway_header_import_file,
    django_assert_num_queries,
):
    service = XlsxPaymentPlanFspExtraFieldsImportService(payment_plan, payment_gateway_header_import_file)
    service.open_workbook()

    with django_assert_num_queries(2):
        service.validate()
    updated_count = service.import_payment_list()

    payments[0].refresh_from_db()
    assert service.errors == []
    assert updated_count == 1
    assert payments[0].fsp_extra_fields["amount"] == 10


@pytest.mark.parametrize(
    "dynamic_system_header_import_file",
    [
        ("core_fields", "beneficiary_reference"),
        ("flex_fields", "fsp_delivery_code"),
        ("document_types", "national_id"),
    ],
    indirect=True,
)
def test_fsp_extra_fields_import_rejects_dynamic_template_header(
    payment_plan,
    dynamic_system_header_import_file,
    django_assert_num_queries,
):
    import_file, header = dynamic_system_header_import_file
    service = XlsxPaymentPlanFspExtraFieldsImportService(payment_plan, import_file)
    service.open_workbook()

    with django_assert_num_queries(2):
        service.validate()

    assert len(service.errors) == 1
    assert service.errors[0].message == f"System-controlled headers cannot be imported: ['{header}']"


def test_fsp_extra_fields_import_rejects_expanded_account_header(
    payment_plan,
    account_system_header_import_file,
    django_assert_num_queries,
):
    service = XlsxPaymentPlanFspExtraFieldsImportService(payment_plan, account_system_header_import_file)
    service.open_workbook()

    with django_assert_num_queries(2):
        service.validate()

    assert len(service.errors) == 1
    assert service.errors[0].message == "System-controlled headers cannot be imported: ['number']"


def test_reconciliation_import_ignores_fsp_extra_field_columns(
    payment_plan,
    reconciliation_file,
):
    service = XlsxPaymentPlanDeliveryImportService(payment_plan, reconciliation_file)
    service.open_workbook()
    row = next(service.ws_payments.iter_rows(min_row=2))

    extras = service._get_extras_for_row(row)

    assert extras == {"returned_code": "RETURNED-001"}


def test_reconciliation_import_uses_plan_wide_fsp_header_ownership(
    payment_plan,
    sparse_reconciliation_file,
):
    service = XlsxPaymentPlanDeliveryImportService(payment_plan, sparse_reconciliation_file)
    service.open_workbook()
    row = next(service.ws_payments.iter_rows(min_row=2))

    extras = service._get_extras_for_row(row)

    assert extras == {"returned_code": "RETURNED-002"}


def test_payment_detail_serializer_separates_extra_field_namespaces(payments):
    serializer = PaymentDetailSerializer()

    assert serializer.get_extras(payments[0]) == {"reconciliation_code": "REC-001"}
    assert serializer.get_extra_fields(payments[0]) == {"reconciliation_code": "REC-001"}
    assert serializer.get_fsp_extra_fields(payments[0]) == {
        "empty_field": "keep-empty",
        "keep": "keep-existing",
        "reference": "old-reference",
    }
