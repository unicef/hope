import json
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, List
from unittest.mock import patch

import pytest
from django.conf import settings
from django.contrib.admin.options import get_content_type_for_model
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.test.utils import CaptureQueriesContext
from django.utils import timezone
from extras.test_utils.factories.account import (
    BusinessAreaFactory,
    PartnerFactory,
    UserFactory,
)
from extras.test_utils.factories.core import create_afghanistan
from extras.test_utils.factories.payment import (
    ApprovalFactory,
    ApprovalProcessFactory,
    DeliveryMechanismFactory,
    FinancialServiceProviderFactory,
    FinancialServiceProviderXlsxTemplateFactory,
    FspXlsxTemplatePerDeliveryMechanismFactory,
    PaymentFactory,
    PaymentPlanFactory,
    PaymentPlanSplitFactory,
)
from extras.test_utils.factories.program import ProgramCycleFactory, ProgramFactory
from extras.test_utils.factories.steficon import RuleCommitFactory
from openpyxl import Workbook
from rest_framework import status
from rest_framework.reverse import reverse
from test_utils.factories.household import create_household_and_individuals

from hope.apps.account.permissions import Permissions
from hope.models.core import FileTemp
from hope.apps.payment.api.views import PaymentPlanManagerialViewSet
from hope.apps.payment.models import (
    Approval,
    FinancialServiceProvider,
    Payment,
    PaymentPlan,
    PaymentPlanSplit,
)
from hope.models.program import Program, ProgramCycle
from hope.models.steficon import Rule
from hope.contrib.vision.models import FundsCommitmentGroup, FundsCommitmentItem

pytestmark = pytest.mark.django_db()


def generate_valid_xlsx_file(name: str = "unit_test.xlsx", worksheet_title: str = "Test") -> SimpleUploadedFile:
    wb = Workbook()
    wb.create_sheet(title=worksheet_title)
    ws = wb.active
    ws["A1"] = "People"
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return SimpleUploadedFile(
        name,
        file_stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class PaymentPlanTestMixin:
    def set_up(self, api_client: Callable, afghanistan: BusinessAreaFactory) -> None:
        self.partner = PartnerFactory(name="TestPartner")
        self.user = UserFactory(partner=self.partner)
        self.client = api_client(self.user)
        self.afghanistan = afghanistan
        self.program1 = ProgramFactory(business_area=self.afghanistan)
        self.program2 = ProgramFactory(business_area=self.afghanistan)
        self.payment_plan1 = PaymentPlanFactory(
            program_cycle=self.program1.cycles.first(),
            business_area=self.afghanistan,
            status=PaymentPlan.Status.IN_APPROVAL,
            created_by=self.user,
        )
        self.payment_plan2 = PaymentPlanFactory(
            program_cycle=self.program2.cycles.first(),
            business_area=self.afghanistan,
            status=PaymentPlan.Status.IN_APPROVAL,
            created_by=self.user,
        )
        self.payment_plan3 = PaymentPlanFactory(
            program_cycle=self.program2.cycles.first(),
            business_area=self.afghanistan,
            status=PaymentPlan.Status.OPEN,
            created_by=self.user,
        )
        self.payment_plan1.refresh_from_db()
        self.payment_plan2.refresh_from_db()
        self.payment_plan3.refresh_from_db()


class TestPaymentPlanManagerialList(PaymentPlanTestMixin):
    def set_up(self, api_client: Callable, afghanistan: BusinessAreaFactory) -> None:
        super().set_up(api_client, afghanistan)
        self.url = reverse(
            "api:payments:payment-plans-managerial-list",
            kwargs={"business_area_slug": self.afghanistan.slug},
        )

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([], status.HTTP_403_FORBIDDEN),
            ([Permissions.PAYMENT_VIEW_LIST_MANAGERIAL], status.HTTP_200_OK),
        ],
    )
    def test_list_payment_plans_permission(
        self,
        permissions: list,
        expected_status: str,
        api_client: Callable,
        afghanistan: BusinessAreaFactory,
        create_user_role_with_permissions: Callable,
    ) -> None:
        self.set_up(api_client, afghanistan)

        create_user_role_with_permissions(
            self.user,
            permissions,
            self.afghanistan,
            self.program1,
        )
        response = self.client.get(self.url)
        assert response.status_code == expected_status

    def test_list_payment_plans(
        self,
        api_client: Callable,
        afghanistan: BusinessAreaFactory,
        create_user_role_with_permissions: Callable,
        create_partner_role_with_permissions: Callable,
    ) -> None:
        def _test_list() -> Any:
            """
            Helper function to test list payment plans now and again to test caching
            """
            response = self.client.get(self.url)
            assert response.status_code == status.HTTP_200_OK
            response_json = response.json()["results"]
            assert len(response_json) == 2
            assert self.payment_plan1.unicef_id in [
                response_json[0]["unicef_id"],
                response_json[1]["unicef_id"],
            ]
            assert self.payment_plan2.unicef_id in [
                response_json[0]["unicef_id"],
                response_json[1]["unicef_id"],
            ]
            assert self.payment_plan3.unicef_id not in [
                response_json[0]["unicef_id"],
                response_json[1]["unicef_id"],
            ]
            return response

        self.set_up(api_client, afghanistan)
        create_user_role_with_permissions(
            self.user,
            [Permissions.PAYMENT_VIEW_LIST_MANAGERIAL],
            self.afghanistan,
            program=self.program1,
        )
        response = self.client.get(self.url)
        assert response.status_code == status.HTTP_200_OK
        response_json = response.json()["results"]
        assert len(response_json) == 1
        assert response_json[0]["unicef_id"] == self.payment_plan1.unicef_id

        create_partner_role_with_permissions(
            self.partner,
            [Permissions.PAYMENT_VIEW_LIST_MANAGERIAL],
            self.afghanistan,
            program=self.program2,
        )

        response = self.client.get(self.url)
        assert response.status_code == status.HTTP_200_OK
        response_json = response.json()["results"]
        assert len(response_json) == 2

        with CaptureQueriesContext(connection) as ctx:
            response = _test_list()
            etag = response.headers["etag"]

            assert json.loads(cache.get(etag)[0].decode("utf8")) == response.json()
            assert len(ctx.captured_queries) == 8  # on CI we have 22 here instead of 8 #FIXME

        # Test that reoccurring request use cached data
        with CaptureQueriesContext(connection) as ctx:
            response = _test_list()
            etag_second_call = response.headers["etag"]
            assert json.loads(cache.get(response.headers["etag"])[0].decode("utf8")) == response.json()
            assert etag_second_call == etag
            assert len(ctx.captured_queries) == 8

    def test_list_payment_plans_approval_process_data(
        self,
        api_client: Callable,
        afghanistan: BusinessAreaFactory,
        create_user_role_with_permissions: Callable,
    ) -> None:
        self.set_up(api_client, afghanistan)
        approval_process = ApprovalProcessFactory(
            payment_plan=self.payment_plan1,
            sent_for_approval_date=timezone.datetime(2021, 1, 1, 0, 0, 0, tzinfo=timezone.utc),
            sent_for_approval_by=self.user,
        )
        approval_approval = ApprovalFactory(approval_process=approval_process, type=Approval.APPROVAL)
        approval_authorization = ApprovalFactory(approval_process=approval_process, type=Approval.AUTHORIZATION)
        approval_release = ApprovalFactory(approval_process=approval_process, type=Approval.FINANCE_RELEASE)
        create_user_role_with_permissions(
            self.user,
            [Permissions.PAYMENT_VIEW_LIST_MANAGERIAL],
            self.afghanistan,
            self.program1,
        )
        response = self.client.get(self.url)
        assert response.status_code == status.HTTP_200_OK
        response_json = response.json()["results"]
        assert len(response_json) == 1
        assert response_json[0]["last_approval_process_date"] == approval_process.sent_for_approval_date.strftime(
            "%Y-%m-%dT%H:%M:%SZ"
        )
        assert response_json[0]["last_approval_process_by"] == str(approval_process.sent_for_approval_by)

        self.payment_plan1.status = PaymentPlan.Status.IN_AUTHORIZATION
        self.payment_plan1.save()
        response = self.client.get(self.url)
        assert response.status_code == status.HTTP_200_OK
        response_json = response.json()["results"]
        assert len(response_json) == 1

        assert response_json[0]["last_approval_process_date"] == approval_approval.created_at.strftime(
            "%Y-%m-%dT%H:%M:%S.%fZ"
        )
        assert response_json[0]["last_approval_process_by"] == str(approval_approval.created_by)

        self.payment_plan1.status = PaymentPlan.Status.IN_REVIEW
        self.payment_plan1.save()
        response = self.client.get(self.url)
        assert response.status_code == status.HTTP_200_OK
        response_json = response.json()["results"]
        assert len(response_json) == 1
        assert response_json[0]["last_approval_process_date"] == approval_authorization.created_at.strftime(
            "%Y-%m-%dT%H:%M:%S.%fZ"
        )
        assert response_json[0]["last_approval_process_by"] == str(approval_authorization.created_by)

        self.payment_plan1.status = PaymentPlan.Status.ACCEPTED
        self.payment_plan1.save()
        response = self.client.get(self.url)
        assert response.status_code == status.HTTP_200_OK
        response_json = response.json()["results"]
        assert len(response_json) == 1
        assert response_json[0]["last_approval_process_date"] == approval_release.created_at.strftime(
            "%Y-%m-%dT%H:%M:%S.%fZ"
        )
        assert response_json[0]["last_approval_process_by"] == str(approval_release.created_by)

    def _bulk_approve_action_response(self) -> Any:
        ApprovalProcessFactory(payment_plan=self.payment_plan1)
        ApprovalProcessFactory(payment_plan=self.payment_plan2)
        return self.client.post(
            reverse(
                "api:payments:payment-plans-managerial-bulk-action",
                kwargs={"business_area_slug": self.afghanistan.slug},
            ),
            data={
                "ids": [self.payment_plan1.id, self.payment_plan2.id],
                "action": PaymentPlan.Action.APPROVE.value,
                "comment": "Test comment",
            },
        )

    def test_bulk_action(
        self,
        api_client: Callable,
        afghanistan: BusinessAreaFactory,
        create_user_role_with_permissions: Callable,
        create_partner_role_with_permissions: Callable,
    ) -> None:
        self.set_up(api_client, afghanistan)
        create_user_role_with_permissions(
            self.user,
            [
                Permissions.PM_VIEW_LIST,
                Permissions.PM_ACCEPTANCE_PROCESS_APPROVE,
                Permissions.PAYMENT_VIEW_LIST_MANAGERIAL,
            ],
            self.afghanistan,
            self.program1,
        )
        create_partner_role_with_permissions(
            self.partner,
            [
                Permissions.PM_VIEW_LIST,
                Permissions.PM_ACCEPTANCE_PROCESS_APPROVE,
                Permissions.PAYMENT_VIEW_LIST_MANAGERIAL,
            ],
            self.afghanistan,
            self.program2,
        )
        response = self._bulk_approve_action_response()
        assert response.status_code == status.HTTP_204_NO_CONTENT
        self.payment_plan1.refresh_from_db()
        self.payment_plan2.refresh_from_db()
        assert self.payment_plan1.status == PaymentPlan.Status.IN_AUTHORIZATION
        assert self.payment_plan2.status == PaymentPlan.Status.IN_AUTHORIZATION

    def test_bulk_action_no_approve_permissions(
        self,
        api_client: Callable,
        afghanistan: BusinessAreaFactory,
        create_user_role_with_permissions: Callable,
        create_partner_role_with_permissions: Callable,
    ) -> None:
        self.set_up(api_client, afghanistan)
        create_user_role_with_permissions(
            self.user,
            [Permissions.PAYMENT_VIEW_LIST_MANAGERIAL],
            self.afghanistan,
            self.program1,
        )
        create_partner_role_with_permissions(
            self.partner,
            [Permissions.PAYMENT_VIEW_LIST_MANAGERIAL],
            self.afghanistan,
            self.program2,
        )
        response = self._bulk_approve_action_response()
        assert response.status_code == status.HTTP_403_FORBIDDEN
        self.payment_plan1.refresh_from_db()
        self.payment_plan2.refresh_from_db()
        assert self.payment_plan1.status == PaymentPlan.Status.IN_APPROVAL
        assert self.payment_plan2.status == PaymentPlan.Status.IN_APPROVAL

    @pytest.mark.parametrize(
        ("action_name", "result"),
        [
            (
                PaymentPlan.Action.APPROVE.name,
                Permissions.PM_ACCEPTANCE_PROCESS_APPROVE.name,
            ),
            (
                PaymentPlan.Action.AUTHORIZE.name,
                Permissions.PM_ACCEPTANCE_PROCESS_AUTHORIZE.name,
            ),
            (
                PaymentPlan.Action.REVIEW.name,
                Permissions.PM_ACCEPTANCE_PROCESS_FINANCIAL_REVIEW.name,
            ),
            ("Some other action name", None),
        ],
    )
    def test_get_action_permission(self, action_name: str, result: str) -> None:
        payment_plan_managerial_viewset = PaymentPlanManagerialViewSet()
        assert payment_plan_managerial_viewset._get_action_permission(action_name) == result


class TestPaymentPlanList:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()
        self.pp = PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.IN_APPROVAL,
            created_by=self.user,
        )
        # add TP
        self.tp = PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_OPEN,
            created_by=self.user,
        )
        self.pp_list_url = reverse(
            "api:payments:payment-plans-list",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
            },
        )
        self.pp_count_url = reverse(
            "api:payments:payment-plans-count",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
            },
        )
        self.client = api_client(self.user)

    def test_payment_plan_list_without_permissions(
        self,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_ACCEPTANCE_PROCESS_APPROVE],
            self.afghanistan,
            self.program_active,
        )
        response = self.client.get(self.pp_list_url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    @pytest.mark.parametrize(
        "permissions",
        [
            [Permissions.PM_VIEW_LIST],
        ],
    )
    def test_payment_plan_list_with_permissions(
        self,
        permissions: list,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(
            self.user,
            permissions,
            self.afghanistan,
            self.program_active,
        )
        response = self.client.get(self.pp_list_url)
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1

        response_count = self.client.get(self.pp_count_url)
        assert response_count.status_code == status.HTTP_200_OK
        assert response_count.json()["count"] == 1

        self.pp.refresh_from_db()
        payment_plan = response_data[0]
        assert str(self.pp.id) == payment_plan["id"]
        assert payment_plan["unicef_id"] == self.pp.unicef_id
        assert payment_plan["name"] == self.pp.name
        assert payment_plan["status"] == self.pp.status
        assert payment_plan["total_households_count"] == self.pp.total_households_count
        assert payment_plan["currency"] == self.pp.currency
        assert payment_plan["excluded_ids"] == self.pp.excluded_ids
        assert payment_plan["total_entitled_quantity"] == str(self.pp.total_entitled_quantity)
        assert payment_plan["total_delivered_quantity"] == str(self.pp.total_delivered_quantity)
        assert payment_plan["total_undelivered_quantity"] == str(self.pp.total_undelivered_quantity)
        assert payment_plan["dispersion_start_date"] == self.pp.dispersion_start_date.strftime("%Y-%m-%d")
        assert payment_plan["dispersion_end_date"] == self.pp.dispersion_end_date.strftime("%Y-%m-%d")
        assert payment_plan["is_follow_up"] == self.pp.is_follow_up
        assert payment_plan["follow_ups"] == []

    def test_payment_plan_caching(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_VIEW_LIST],
            self.afghanistan,
            self.program_active,
        )
        # first api call no cache
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.pp_list_url)
            assert response.status_code == status.HTTP_200_OK
            assert response.has_header("etag")
            etag = response.headers["etag"]
            assert json.loads(cache.get(etag)[0].decode("utf8")) == response.json()
            assert len(response.json()["results"]) == 1
            assert len(ctx.captured_queries) == 17

        # second call get from cache
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.pp_list_url)
            assert response.status_code == status.HTTP_200_OK
            assert response.has_header("etag")
            etag_second_call = response.headers["etag"]
            assert etag == etag_second_call
            assert len(ctx.captured_queries) == 6
        # upd PP
        self.pp.status = PaymentPlan.Status.IN_REVIEW
        self.pp.save()
        # PaymentPlan updated, invalidate cache... new call
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.pp_list_url)
            assert response.status_code == status.HTTP_200_OK
            assert response.has_header("etag")
            new_etag = response.headers["etag"]
            assert json.loads(cache.get(new_etag)[0].decode("utf8")) == response.json()
            assert len(response.json()["results"]) == 1
            assert len(ctx.captured_queries) == 11
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.pp_list_url)
            assert response.status_code == status.HTTP_200_OK
            assert response.has_header("etag")
            etag_second_call = response.headers["etag"]
            assert new_etag == etag_second_call
            assert len(ctx.captured_queries) == 6
        # add new PP cache invalidate... new call
        PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.OPEN,
            created_by=self.user,
        )
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.pp_list_url)
            assert response.status_code == status.HTTP_200_OK
            assert response.has_header("etag")
            etag = response.headers["etag"]
            assert json.loads(cache.get(etag)[0].decode("utf8")) == response.json()
            assert len(response.json()["results"]) == 2
            assert len(ctx.captured_queries) == 13
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.pp_list_url)
            assert response.status_code == status.HTTP_200_OK
            assert response.has_header("etag")
            etag_second_call = response.headers["etag"]
            assert etag == etag_second_call
            assert len(ctx.captured_queries) == 6

        # delete PP
        self.pp.delete()
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.pp_list_url)
            assert response.status_code == status.HTTP_200_OK
            assert response.has_header("etag")
            etag = response.headers["etag"]
            assert json.loads(cache.get(etag)[0].decode("utf8")) == response.json()
            assert len(response.json()["results"]) == 1
            assert len(ctx.captured_queries) == 11
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.pp_list_url)
            assert response.status_code == status.HTTP_200_OK
            assert response.has_header("etag")
            last_etag_second_call = response.headers["etag"]
            assert etag == last_etag_second_call
            assert len(ctx.captured_queries) == 6

        # upd TP no changes in cache
        self.tp.status = PaymentPlan.Status.TP_LOCKED
        self.tp.save()
        # cache will not be updated because just TPs list updated
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.pp_list_url)
            assert response.status_code == status.HTTP_200_OK
            assert len(response.json()["results"]) == 1
            assert response.has_header("etag")
            get_etag = response.headers["etag"]
            assert get_etag == last_etag_second_call
            assert len(ctx.captured_queries) == 6


class TestPaymentPlanDetail:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()
        self.pp = PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.IN_APPROVAL,
            created_by=self.user,
        )
        pp_id = str(self.pp.id)
        self.pp_detail_url = reverse(
            "api:payments:payment-plans-detail",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
                "pk": pp_id,
            },
        )
        self.client = api_client(self.user)

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_VIEW_DETAILS], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_payment_plan_detail_permissions(
        self,
        permissions: list,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)

        response = self.client.get(self.pp_detail_url)
        assert response.status_code == expected_status

    def test_payment_plan_detail(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_VIEW_DETAILS],
            self.afghanistan,
            self.program_active,
        )
        response = self.client.get(self.pp_detail_url)
        assert response.status_code == status.HTTP_200_OK

        payment_plan = response.json()
        self.pp.refresh_from_db()

        assert payment_plan["id"] == str(self.pp.id)
        assert payment_plan["unicef_id"] == self.pp.unicef_id
        assert payment_plan["name"] == self.pp.name
        assert payment_plan["status"] == self.pp.status
        assert payment_plan["total_households_count"] == self.pp.total_households_count
        assert payment_plan["currency"] == self.pp.currency
        assert payment_plan["excluded_ids"] == self.pp.excluded_ids
        assert payment_plan["total_entitled_quantity"] == str(self.pp.total_entitled_quantity)
        assert payment_plan["total_delivered_quantity"] == str(self.pp.total_delivered_quantity)
        assert payment_plan["total_undelivered_quantity"] == str(self.pp.total_undelivered_quantity)
        assert payment_plan["dispersion_start_date"] == self.pp.dispersion_start_date.strftime("%Y-%m-%d")
        assert payment_plan["dispersion_end_date"] == self.pp.dispersion_end_date.strftime("%Y-%m-%d")
        assert payment_plan["is_follow_up"] == self.pp.is_follow_up
        assert payment_plan["follow_ups"] == []
        assert payment_plan["created_by"] == f"{self.user.first_name} {self.user.last_name}"
        assert payment_plan["background_action_status"] is None
        assert payment_plan["start_date"] is None
        assert payment_plan["program"]["name"] == self.program_active.name
        assert payment_plan["has_payment_list_export_file"] is False
        assert payment_plan["has_fsp_delivery_mechanism_xlsx_template"] is False
        assert payment_plan["imported_file_name"] == ""
        assert payment_plan["payments_conflicts_count"] == 0
        assert payment_plan["delivery_mechanism"] is None
        assert payment_plan["bank_reconciliation_success"] == 0
        assert payment_plan["bank_reconciliation_error"] == 0
        assert payment_plan["can_create_payment_verification_plan"] is False
        assert payment_plan["available_payment_records_count"] == 0
        assert payment_plan["can_create_follow_up"] is False
        assert payment_plan["total_withdrawn_households_count"] == 0
        assert payment_plan["unsuccessful_payments_count"] == 0
        assert payment_plan["can_send_to_payment_gateway"] is False
        assert payment_plan["can_split"] is False
        assert payment_plan["total_households_count_with_valid_phone_no"] == 0
        assert payment_plan["can_create_xlsx_with_fsp_auth_code"] is False
        assert payment_plan["fsp_communication_channel"] == "XLSX"
        assert payment_plan["can_export_xlsx"] is False
        assert payment_plan["can_download_xlsx"] is False
        assert payment_plan["can_send_xlsx_password"] is False

    def test_has_fsp_delivery_mechanism_xlsx_template(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_VIEW_DETAILS],
            self.afghanistan,
            self.program_active,
        )
        xlsx_temp = FinancialServiceProviderXlsxTemplateFactory()
        dm = DeliveryMechanismFactory()
        fsp = FinancialServiceProviderFactory(
            name="Test FSP 1",
            communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_XLSX,
            vision_vendor_number=123,
        )
        fsp.xlsx_templates.set([xlsx_temp])
        FspXlsxTemplatePerDeliveryMechanismFactory(
            financial_service_provider=fsp,
            delivery_mechanism=dm,
            xlsx_template=xlsx_temp,
        )
        self.pp.delivery_mechanism = dm
        self.pp.financial_service_provider = fsp
        self.pp.save()

        response = self.client.get(self.pp_detail_url)
        assert response.status_code == status.HTTP_200_OK
        payment_plan = response.json()
        assert payment_plan["has_fsp_delivery_mechanism_xlsx_template"] is True

    def test_can_create_follow_up(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_VIEW_DETAILS],
            self.afghanistan,
            self.program_active,
        )
        self.pp.is_follow_up = True
        self.pp.save()

        response = self.client.get(self.pp_detail_url)
        assert response.status_code == status.HTTP_200_OK
        payment_plan = response.json()
        assert payment_plan["can_create_follow_up"] is False

    def test_get_can_split(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_VIEW_DETAILS],
            self.afghanistan,
            self.program_active,
        )
        PaymentPlanSplitFactory(payment_plan=self.pp, sent_to_payment_gateway=True)
        self.pp.status = PaymentPlan.Status.ACCEPTED
        self.pp.save()

        response = self.client.get(self.pp_detail_url)
        assert response.status_code == status.HTTP_200_OK
        payment_plan = response.json()
        assert payment_plan["can_split"] is False


class TestPaymentPlanFilter:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any, create_user_role_with_permissions: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()
        self.pp = PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.IN_APPROVAL,
            created_by=self.user,
        )
        self.pp_finished = PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.FINISHED,
            created_by=self.user,
        )
        self.list_url = reverse(
            "api:payments:payment-plans-list",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
            },
        )
        self.partner = PartnerFactory(name="ABC")
        self.user = UserFactory(partner=self.partner)
        self.client = api_client(self.user)
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_VIEW_LIST],
            self.afghanistan,
            self.program_active,
        )

    def test_filter_by_status(self) -> None:
        response = self.client.get(self.list_url, {"status": PaymentPlan.Status.FINISHED.value})
        self.pp_finished.refresh_from_db()
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["id"] == str(self.pp_finished.id)
        assert response_data[0]["status"] == "FINISHED"
        assert response_data[0]["name"] == self.pp_finished.name

        response = self.client.get(self.list_url, {"status": PaymentPlan.Status.IN_APPROVAL.value})
        self.pp.refresh_from_db()
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["id"] == str(self.pp.id)
        assert response_data[0]["status"] == "IN_APPROVAL"
        assert response_data[0]["name"] == self.pp.name

    def test_filter_by_program_cycle(self) -> None:
        new_pp = PaymentPlanFactory(
            name="TEST_ABC_123",
            business_area=self.afghanistan,
            program_cycle=ProgramCycleFactory(program=self.program_active),
            status=PaymentPlan.Status.ACCEPTED,
            created_by=self.user,
        )
        response = self.client.get(self.list_url, {"program_cycle": new_pp.program_cycle.id})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "TEST_ABC_123"
        assert response_data[0]["status"] == "ACCEPTED"

    def test_filter_by_search(self) -> None:
        new_pp = PaymentPlanFactory(
            name="TEST_ABC_999",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.ACCEPTED,
            created_by=self.user,
        )
        new_pp.refresh_from_db()
        # name
        response = self.client.get(self.list_url, {"search": "TEST_ABC"})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "TEST_ABC_999"
        assert response_data[0]["status"] == "ACCEPTED"
        # id
        response = self.client.get(self.list_url, {"search": str(new_pp.id)})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "TEST_ABC_999"
        # unicef_id
        response = self.client.get(self.list_url, {"search": new_pp.unicef_id})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "TEST_ABC_999"

    def test_filter_by_entitled_quantity(self) -> None:
        PaymentPlanFactory(
            name="PP_1",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.LOCKED_FSP,
            created_by=self.user,
            total_entitled_quantity=100,
        )
        PaymentPlanFactory(
            name="PP_2",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.LOCKED_FSP,
            created_by=self.user,
            total_entitled_quantity=200,
        )
        response = self.client.get(
            self.list_url,
            {"total_entitled_quantity__gte": "99", "total_entitled_quantity__lte": 201},
        )
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 2
        assert response_data[0]["name"] == "PP_1"
        assert response_data[1]["name"] == "PP_2"

        response = self.client.get(self.list_url, {"total_entitled_quantity__lte": 101})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "PP_1"

    def test_filter_by_dispersion_date(self) -> None:
        PaymentPlanFactory(
            name="PP_abc",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.LOCKED_FSP,
            created_by=self.user,
            dispersion_start_date="2022-02-24",
            dispersion_end_date="2022-03-03",
        )
        PaymentPlanFactory(
            name="PP_xyz",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.LOCKED_FSP,
            created_by=self.user,
            dispersion_start_date="2022-01-01",
            dispersion_end_date="2022-01-17",
        )
        response = self.client.get(self.list_url, {"start_date": "2022-02-23", "end_date": "2022-03-04"})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "PP_abc"

        response = self.client.get(self.list_url, {"end_date": "2022-01-18"})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "PP_xyz"

    def test_filter_by_is_follow_up(self) -> None:
        PaymentPlanFactory(
            name="NEW_FOLLOW_up",
            is_follow_up=True,
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.OPEN,
            created_by=self.user,
        )
        response = self.client.get(self.list_url, {"is_follow_up": True})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "NEW_FOLLOW_up"


class TestTargetPopulationList:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()
        self.tp = PaymentPlanFactory(
            name="Test new TP",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_OPEN,
            created_by=self.user,
        )
        # add PaymentPlan
        self.pp = PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.LOCKED_FSP,
            created_by=self.user,
        )
        self.tp_list_url = reverse(
            "api:payments:target-populations-list",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
            },
        )
        self.tp_count_url = reverse(
            "api:payments:target-populations-count",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
            },
        )
        self.client = api_client(self.user)

    def test_target_population_list_without_permissions(
        self,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_ACCEPTANCE_PROCESS_APPROVE],
            self.afghanistan,
            self.program_active,
        )
        response = self.client.get(self.tp_list_url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    @pytest.mark.parametrize(
        "permissions",
        [
            [Permissions.TARGETING_VIEW_LIST],
        ],
    )
    def test_target_population_list_with_permissions(
        self,
        permissions: list,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(
            self.user,
            permissions,
            self.afghanistan,
            self.program_active,
        )
        response = self.client.get(self.tp_list_url)
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 2

        response_count = self.client.get(self.tp_count_url)
        assert response_count.status_code == status.HTTP_200_OK
        assert response_count.json()["count"] == 2

        self.tp.refresh_from_db()
        tp = response_data[1]
        assert str(self.tp.id) == tp["id"]
        assert tp["name"] == "Test new TP"
        assert tp["status"] == self.tp.get_status_display().upper()
        assert tp["total_households_count"] == self.tp.total_households_count
        assert tp["total_individuals_count"] == self.tp.total_individuals_count
        assert tp["created_at"] == self.tp.created_at.isoformat().replace("+00:00", "Z")
        assert tp["updated_at"] == self.tp.updated_at.isoformat().replace("+00:00", "Z")
        assert tp["created_by"] == self.user.get_full_name()

    def test_target_population_caching(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.TARGETING_VIEW_LIST],
            self.afghanistan,
            self.program_active,
        )
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.tp_list_url)
            assert response.status_code == status.HTTP_200_OK

            etag = response.headers["etag"]
            assert json.loads(cache.get(etag)[0].decode("utf8")) == response.json()
            assert len(ctx.captured_queries) == 17

        # Test that reoccurring requests use cached data
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.tp_list_url)
            assert response.status_code == status.HTTP_200_OK

            etag_second_call = response.headers["etag"]
            assert json.loads(cache.get(response.headers["etag"])[0].decode("utf8")) == response.json()
            assert len(ctx.captured_queries) == 6
            assert etag_second_call == etag

        # After update, it does not use the cached data
        self.tp.status = PaymentPlan.Status.TP_PROCESSING
        self.tp.save()
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.tp_list_url)
            assert response.status_code == status.HTTP_200_OK

            etag_call_after_update = response.headers["etag"]
            assert json.loads(cache.get(response.headers["etag"])[0].decode("utf8")) == response.json()
            assert len(ctx.captured_queries) == 11

            assert etag_call_after_update != etag

        # Cached data again
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(self.tp_list_url)
            assert response.status_code == status.HTTP_200_OK

            etag_call_after_update_second_call = response.headers["etag"]
            assert json.loads(cache.get(response.headers["etag"])[0].decode("utf8")) == response.json()
            assert len(ctx.captured_queries) == 6
            assert etag_call_after_update_second_call == etag_call_after_update


class TestTargetPopulationDetail:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()
        self.tp = PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_LOCKED,
            created_by=self.user,
        )
        tp_id = str(self.tp.id)
        self.tp_detail_url = reverse(
            "api:payments:target-populations-detail",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
                "pk": tp_id,
            },
        )
        self.client = api_client(self.user)

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_VIEW_DETAILS], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_target_population_detail_permissions(
        self,
        permissions: list,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)

        response = self.client.get(self.tp_detail_url)
        assert response.status_code == expected_status

    def test_target_population_detail(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.TARGETING_VIEW_DETAILS],
            self.afghanistan,
            self.program_active,
        )
        response = self.client.get(self.tp_detail_url)
        assert response.status_code == status.HTTP_200_OK

        tp = response.json()
        self.tp.refresh_from_db()

        assert tp["id"] == str(self.tp.id)
        assert tp["name"] == self.tp.name
        assert tp["program_cycle"]["title"] == self.cycle.title
        assert tp["program"]["name"] == self.program_active.name
        assert tp["status"] == self.tp.status
        assert tp["total_households_count"] == self.tp.total_households_count
        assert tp["total_individuals_count"] == self.tp.total_individuals_count
        assert tp["created_by"] == f"{self.user.first_name} {self.user.last_name}"
        assert tp["background_action_status"] is None
        assert tp["male_children_count"] == self.tp.male_children_count
        assert tp["female_children_count"] == self.tp.female_children_count
        assert tp["male_adults_count"] == self.tp.male_adults_count
        assert tp["female_adults_count"] == self.tp.female_adults_count


class TestTargetPopulationFilter:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any, create_user_role_with_permissions: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()
        self.tp = PaymentPlanFactory(
            name="OPEN",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_OPEN,
            created_by=self.user,
            total_households_count=999,
        )
        self.tp_locked = PaymentPlanFactory(
            name="LOCKED",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_LOCKED,
            created_by=self.user,
            total_households_count=888,
        )
        self.tp_assigned = PaymentPlanFactory(
            name="Assigned TP",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.FINISHED,
            created_by=self.user,
            total_households_count=777,
        )
        self.list_url = reverse(
            "api:payments:target-populations-list",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
            },
        )
        self.partner = PartnerFactory(name="ABC")
        self.user = UserFactory(partner=self.partner)
        self.client = api_client(self.user)
        create_user_role_with_permissions(
            self.user,
            [Permissions.TARGETING_VIEW_LIST],
            self.afghanistan,
            self.program_active,
        )

    def test_filter_by_status(self) -> None:
        response = self.client.get(self.list_url, {"status": PaymentPlan.Status.TP_LOCKED.value})
        self.tp_locked.refresh_from_db()
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["id"] == str(self.tp_locked.id)
        assert response_data[0]["status"] == "LOCKED"
        assert response_data[0]["name"] == "LOCKED"

        response = self.client.get(self.list_url, {"status": "ASSIGNED"})
        self.tp_assigned.refresh_from_db()
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["id"] == str(self.tp_assigned.id)
        assert response_data[0]["name"] == "Assigned TP"

    def test_filter_by_program_cycle(self) -> None:
        new_tp = PaymentPlanFactory(
            name="TEST_ABC_QWOOL",
            business_area=self.afghanistan,
            program_cycle=ProgramCycleFactory(program=self.program_active),
            status=PaymentPlan.Status.TP_STEFICON_RUN,
            created_by=self.user,
        )
        response = self.client.get(self.list_url, {"program_cycle": new_tp.program_cycle.id})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "TEST_ABC_QWOOL"
        assert response_data[0]["status"] == "STEFICON RUN"

    def test_filter_by_number_of_hh(self) -> None:
        PaymentPlanFactory(
            name="PP_1",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_LOCKED,
            created_by=self.user,
            total_households_count=100,
        )
        PaymentPlanFactory(
            name="PP_2",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_LOCKED,
            created_by=self.user,
            total_households_count=200,
        )
        response = self.client.get(
            self.list_url,
            {"total_households_count__gte": "99", "total_households_count__lte": 201},
        )
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 2
        assert response_data[0]["name"] == "PP_2"
        assert response_data[1]["name"] == "PP_1"

        response = self.client.get(self.list_url, {"total_households_count__lte": 101})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "PP_1"

    def test_filter_by_created_date(self) -> None:
        tp_1 = PaymentPlanFactory(
            name="TP_Mmmmmmm",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.LOCKED_FSP,
            created_by=self.user,
            created_at="2022-02-24",
        )
        tp_2 = PaymentPlanFactory(
            name="TP_Uuuuu_Aaaaa",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.LOCKED_FSP,
            created_by=self.user,
            created_at="2022-01-01",
        )
        tp_1.created_at = "2022-02-24"
        tp_1.save()
        tp_2.created_at = "2022-01-01"
        tp_2.save()
        response = self.client.get(
            self.list_url,
            {"created_at__gte": "2022-02-23", "created_at__lte": "2022-03-04"},
        )
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "TP_Mmmmmmm"

        response = self.client.get(self.list_url, {"created_at__lte": "2022-01-18"})
        assert response.status_code == status.HTTP_200_OK
        response_data = response.json()["results"]
        assert len(response_data) == 1
        assert response_data[0]["name"] == "TP_Uuuuu_Aaaaa"


class TestTargetPopulationCreateUpdate:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()

        self.tp = PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_OPEN,
            created_by=self.user,
        )

        self.create_url = reverse(
            "api:payments:target-populations-list",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
            },
        )
        self.update_url = reverse(
            "api:payments:target-populations-detail",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
                "pk": self.tp.pk,
            },
        )
        self.client = api_client(self.user)
        self.rules = [
            {
                "individua_ids": "",
                "household_ids": "",
                "households_filters_blocks": [
                    {
                        "comparison_method": "RANGE",
                        "arguments": [1, 11],
                        "field_name": "size",
                        "flex_field_classification": "NOT_FLEX_FIELD",
                    }
                ],
                "individuals_filters_blocks": [],
                "collectors_filters_blocks": [],
            }
        ]

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_CREATE], status.HTTP_201_CREATED),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_create_payment_plan_success(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        data = {
            "name": "New Payment Plan",
            "program_cycle_id": self.cycle.id,
            "rules": self.rules,
            "excluded_ids": "IND-123",
            "exclusion_reason": "Just MMM Qwool Test",
            "flag_exclude_if_on_sanction_list": True,
            "flag_exclude_if_active_adjudication_ticket": False,
        }

        response = self.client.post(self.create_url, data, format="json")
        assert response.status_code == expected_status
        if response.status_code == status.HTTP_201_CREATED:
            assert response.data["name"] == data["name"]
            assert response.data["rules"][0]["households_filters_blocks"][0]["field_name"] == "size"
            assert response.data["exclusion_reason"] == data["exclusion_reason"]
            assert response.data["excluded_ids"] == data["excluded_ids"]
            assert response.data["flag_exclude_if_on_sanction_list"] is True
            assert response.data["flag_exclude_if_active_adjudication_ticket"] is False

    def test_create_payment_plan_invalid_data(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.TARGETING_CREATE],
            self.afghanistan,
            self.program_active,
        )
        invalid_data = {
            "name": "",
            "program_cycle_id": None,
            "rules": [],
        }
        response = self.client.post(self.create_url, invalid_data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "name" in response.data
        assert "program_cycle_id" in response.data

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_UPDATE], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_update_payment_plan_success(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)

        update_data = {
            "name": "Test with NEW NAME",
            "excluded_ids": "IND-999",
            "exclusion_reason": "123",
            "version": self.tp.version,
        }

        response = self.client.patch(self.update_url, update_data, format="json")

        assert response.status_code == expected_status
        if response.status_code == status.HTTP_200_OK:
            assert response.data["name"] == update_data["name"]
            assert response.data["exclusion_reason"] == update_data["exclusion_reason"]
            assert response.data["excluded_ids"] == update_data["excluded_ids"]

    def test_update_payment_plan_invalid_data(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.TARGETING_UPDATE],
            self.afghanistan,
            self.program_active,
        )
        invalid_update_data = {
            "name": "",
            "version": self.tp.version,
        }
        response = self.client.patch(self.update_url, invalid_update_data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "name" in response.data


class TestTargetPopulationActions:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()
        self.client = api_client(self.user)
        self.target_population = PaymentPlanFactory(
            name="TP_OPEN",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_OPEN,
            created_by=self.user,
            created_at="2022-02-24",
        )
        tp_id = self.target_population.pk
        url_kwargs = {
            "business_area_slug": self.afghanistan.slug,
            "program_slug": self.program_active.slug,
            "pk": tp_id,
        }
        self.url_lock = reverse("api:payments:target-populations-lock", kwargs=url_kwargs)
        self.url_unlock = reverse("api:payments:target-populations-unlock", kwargs=url_kwargs)
        self.url_rebuild = reverse("api:payments:target-populations-rebuild", kwargs=url_kwargs)
        self.url_mark_ready = reverse("api:payments:target-populations-mark-ready", kwargs=url_kwargs)
        self.url_copy = reverse("api:payments:target-populations-copy", kwargs=url_kwargs)
        self.url_apply_steficon = reverse("api:payments:target-populations-apply-engine-formula", kwargs=url_kwargs)

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_LOCK], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_lock(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)

        response = self.client.get(self.url_lock)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json() == {"message": "Target Population locked"}

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_UNLOCK], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_unlock(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.target_population.status = PaymentPlan.Status.TP_LOCKED
        self.target_population.save()

        response = self.client.get(self.url_unlock)
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json() == {"message": "Target Population unlocked"}

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_LOCK], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_rebuild(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)

        response = self.client.get(self.url_rebuild)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json() == {"message": "Target Population rebuilding"}

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            (
                [Permissions.TARGETING_CREATE, Permissions.TARGETING_SEND],
                status.HTTP_200_OK,
            ),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_mark_ready(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.target_population.financial_service_provider = FinancialServiceProviderFactory()
        self.target_population.delivery_mechanism = DeliveryMechanismFactory()
        self.target_population.status = PaymentPlan.Status.TP_LOCKED
        self.target_population.save()

        response = self.client.get(self.url_mark_ready)
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json() == {"message": "Target Population ready for Payment Plan"}

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_DUPLICATE], status.HTTP_201_CREATED),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_copy_tp(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        data = {"name": "Copied TP test 123", "program_cycle_id": self.cycle.pk}
        response = self.client.post(self.url_copy, data, format="json")

        assert response.status_code == expected_status
        if expected_status == status.HTTP_201_CREATED:
            assert response.status_code == status.HTTP_201_CREATED
            assert "id" in response.json()
            assert PaymentPlan.objects.filter(name="Copied TP test 123").count() == 1
            assert PaymentPlan.objects.all().count() == 2

    def test_copy_tp_validation_errors(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.TARGETING_DUPLICATE],
            self.afghanistan,
            self.program_active,
        )
        cycle = ProgramCycleFactory(program=self.program_active, status=ProgramCycle.ACTIVE, title="Cycle123")
        PaymentPlanFactory(
            name="Copied TP AGAIN",
            business_area=self.afghanistan,
            program_cycle=cycle,
        )
        data = {
            "name": "Copied TP AGAIN",
            "program_cycle_id": cycle.pk,
        }
        # TP with the same name already exists
        response = self.client.post(self.url_copy, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert (
            "Target Population with name: Copied TP AGAIN and program cycle: Cycle123 already exists." in response.data
        )

        # Cycle status is Finished
        cycle.status = ProgramCycle.FINISHED
        cycle.save()
        response_2 = self.client.post(self.url_copy, data, format="json")

        assert response_2.status_code == status.HTTP_400_BAD_REQUEST
        assert "Not possible to assign Finished Program Cycle to Targeting." in response_2.data

        # wrong input data to trigger validation error
        response_3 = self.client.post(self.url_copy, {}, format="json")
        assert response_3.status_code == status.HTTP_400_BAD_REQUEST
        assert "name" in response_3.data
        assert "program_cycle_id" in response_3.data

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_UPDATE], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_apply_engine_formula_tp(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        rule_for_tp = RuleCommitFactory(rule__type=Rule.TYPE_TARGETING, version=11).rule
        self.target_population.status = PaymentPlan.Status.TP_LOCKED
        self.target_population.save()
        data = {
            "engine_formula_rule_id": str(rule_for_tp.pk),
            "version": self.target_population.version,
        }
        response = self.client.post(self.url_apply_steficon, data, format="json")

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert "TARGETING" in resp_data["steficon_rule_targeting"]["rule"]["type"]

    def test_apply_engine_formula_tp_validation_errors(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.TARGETING_UPDATE],
            self.afghanistan,
            self.program_active,
        )
        rule_for_tp = RuleCommitFactory(rule__type=Rule.TYPE_TARGETING, rule__enabled=False, version=22).rule
        self.target_population.status = PaymentPlan.Status.TP_STEFICON_ERROR
        self.target_population.save()

        data = {
            "engine_formula_rule_id": rule_for_tp.pk,
            "version": self.target_population.version,
        }
        response = self.client.post(self.url_apply_steficon, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "This engine rule is not enabled or is deprecated." in response.data

        self.target_population.status = PaymentPlan.Status.TP_OPEN
        self.target_population.save()
        data = {
            "engine_formula_rule_id": rule_for_tp.pk,
            "version": self.target_population.version,
        }
        response_2 = self.client.post(self.url_apply_steficon, data, format="json")

        assert response_2.status_code == status.HTTP_400_BAD_REQUEST
        assert "Not allowed to run engine formula within status TP_OPEN." in response_2.data

        response_3 = self.client.post(self.url_apply_steficon, {}, format="json")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "engine_formula_rule_id" in response_3.data

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_REMOVE], status.HTTP_204_NO_CONTENT),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_tp_delete(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        tp = PaymentPlanFactory(
            name="TP_to_delete",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_OPEN,
            created_by=self.user,
        )
        tp_id = tp.pk
        delete_url = reverse(
            "api:payments:target-populations-detail",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
                "pk": tp_id,
            },
        )
        response = self.client.delete(delete_url)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_204_NO_CONTENT:
            assert response.status_code == status.HTTP_204_NO_CONTENT
            assert PaymentPlan.objects.filter(name="TP_to_delete").count() == 0
            assert PaymentPlan.all_objects.filter(name="TP_to_delete").count() == 1  # is_removed = True


class TestPendingPaymentsAction:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any) -> None:
        self.pending_payments_url_name = "api:payments:target-populations-pending-payments"

        self.afghanistan = create_afghanistan()
        self.program = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.partner = PartnerFactory(name="TestPartner")
        self.user = UserFactory(partner=self.partner)

        self.cycle = self.program.cycles.first()
        self.api_client = api_client(self.user)

        (
            self.household1,
            (
                self.individual1_1,
                self.individual1_2,
            ),
        ) = create_household_and_individuals(
            household_data={
                "program": self.program,
                "business_area": self.afghanistan,
            },
            individuals_data=[
                {
                    "program": self.program,
                    "business_area": self.afghanistan,
                },
                {
                    "program": self.program,
                    "business_area": self.afghanistan,
                },
            ],
        )

        self.target_population = PaymentPlanFactory(
            name="TP_OPEN",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.TP_OPEN,
            created_by=self.user,
            created_at="2022-02-24",
        )
        Payment.objects.create(
            household=self.household1,
            parent=self.target_population,
            business_area=self.afghanistan,
            collector_id=self.individual1_1.id,
            status_date="2022-02-24",
        )

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.TARGETING_VIEW_DETAILS], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pending_payments(
        self,
        permissions: list,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(
            user=self.user,
            permissions=permissions,
            business_area=self.afghanistan,
            program=self.program,
        )
        response = self.api_client.get(
            reverse(
                self.pending_payments_url_name,
                kwargs={
                    "business_area_slug": self.afghanistan.slug,
                    "program_slug": self.program.slug,
                    "pk": str(self.target_population.id),
                },
            )
        )

        assert response.status_code == expected_status


class TestPaymentPlanActions:
    @pytest.fixture(autouse=True)
    def setup(self, api_client: Any) -> None:
        self.afghanistan = create_afghanistan()
        self.partner = PartnerFactory(name="unittest")
        self.user = UserFactory(partner=self.partner)
        self.program_active = ProgramFactory(business_area=self.afghanistan, status=Program.ACTIVE)
        self.cycle = self.program_active.cycles.first()
        self.client = api_client(self.user)
        self.pp = PaymentPlanFactory(
            unicef_id="PP-0060-23-0.000.001",
            name="DRAFT PP",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.DRAFT,
            created_by=self.user,
            created_at="2022-02-24",
            currency="PLN",
        )
        pp_id = self.pp.pk
        url_kwargs = {
            "business_area_slug": self.afghanistan.slug,
            "program_slug": self.program_active.slug,
            "pk": pp_id,
        }
        self.url_kwargs_ba_program = {
            "business_area_slug": self.afghanistan.slug,
            "program_slug": self.program_active.slug,
        }
        self.url_list = reverse("api:payments:payment-plans-list", kwargs=self.url_kwargs_ba_program)
        self.url_lock = reverse("api:payments:payment-plans-lock", kwargs=url_kwargs)
        self.url_unlock = reverse("api:payments:payment-plans-unlock", kwargs=url_kwargs)
        self.url_exclude_hh = reverse("api:payments:payment-plans-exclude-beneficiaries", kwargs=url_kwargs)
        self.url_apply_steficon = reverse("api:payments:payment-plans-apply-engine-formula", kwargs=url_kwargs)
        self.url_lock_fsp = reverse("api:payments:payment-plans-lock-fsp", kwargs=url_kwargs)
        self.url_unlock_fsp = reverse("api:payments:payment-plans-unlock-fsp", kwargs=url_kwargs)
        self.url_export_entitlement_xlsx = reverse(
            "api:payments:payment-plans-entitlement-export-xlsx", kwargs=url_kwargs
        )
        self.url_import_entitlement_xlsx = reverse(
            "api:payments:payment-plans-entitlement-import-xlsx", kwargs=url_kwargs
        )
        self.url_send_for_approval = reverse("api:payments:payment-plans-send-for-approval", kwargs=url_kwargs)
        self.url_approval_process_reject = reverse("api:payments:payment-plans-reject", kwargs=url_kwargs)
        self.url_approval_process_approve = reverse("api:payments:payment-plans-approve", kwargs=url_kwargs)
        self.url_approval_process_authorize = reverse("api:payments:payment-plans-authorize", kwargs=url_kwargs)
        self.url_approval_process_mark_as_released = reverse(
            "api:payments:payment-plans-mark-as-released", kwargs=url_kwargs
        )
        self.url_send_to_payment_gate_way = reverse(
            "api:payments:payment-plans-send-to-payment-gateway", kwargs=url_kwargs
        )

        self.url_export_pdf_payment_plan_summary = reverse(
            "api:payments:payment-plans-export-pdf-payment-plan-summary",
            kwargs=url_kwargs,
        )
        self.url_generate_xlsx_with_auth_code = reverse(
            "api:payments:payment-plans-generate-xlsx-with-auth-code", kwargs=url_kwargs
        )
        self.url_send_xlsx_password = reverse("api:payments:payment-plans-send-xlsx-password", kwargs=url_kwargs)
        self.url_reconciliation_export_xlsx = reverse(
            "api:payments:payment-plans-reconciliation-export-xlsx", kwargs=url_kwargs
        )
        self.url_reconciliation_import_xlsx = reverse(
            "api:payments:payment-plans-reconciliation-import-xlsx", kwargs=url_kwargs
        )
        self.url_pp_split = reverse("api:payments:payment-plans-split", kwargs=url_kwargs)
        self.url_create_follow_up = reverse("api:payments:payment-plans-create-follow-up", kwargs=url_kwargs)
        self.url_funds_commitments = reverse("api:payments:payment-plans-assign-funds-commitments", kwargs=url_kwargs)

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_CREATE], status.HTTP_201_CREATED),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_create_pp(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        data = {
            "dispersion_start_date": "2025-02-01",
            "dispersion_end_date": "2099-03-01",
            "currency": "USD",
            "target_population_id": str(self.pp.id),
        }
        response = self.client.post(self.url_list, data, format="json")

        assert response.status_code == expected_status
        if expected_status == status.HTTP_201_CREATED:
            assert response.status_code == status.HTTP_201_CREATED
            resp_data = response.json()
            assert "id" in resp_data
            assert resp_data["currency"] == "USD"
            assert resp_data["status"] == "OPEN"

    def test_create_pp_validation_errors(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, [Permissions.PM_CREATE], self.afghanistan, self.program_active)
        response = self.client.post(self.url_list, {"target_population_id": str(self.pp.pk)}, format="json")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "dispersion_start_date" in response.json()
        assert "dispersion_end_date" in response.json()
        assert "currency" in response.json()

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_LOCK_AND_UNLOCK], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pp_lock(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.OPEN
        self.pp.save()
        PaymentFactory(parent=self.pp)
        response = self.client.get(self.url_lock)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json() == {"message": "Payment Plan locked"}

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_LOCK_AND_UNLOCK], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pp_unlock(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.LOCKED
        self.pp.save()

        response = self.client.get(self.url_unlock)
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json() == {"message": "Payment Plan unlocked"}

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_CREATE], status.HTTP_204_NO_CONTENT),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_payment_plan_delete(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        pp = PaymentPlanFactory(
            name="new pp for delete test",
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.OPEN,
            created_by=self.user,
        )
        pp_id = str(pp.pk)
        delete_url = reverse(
            "api:payments:payment-plans-detail",
            kwargs={
                "business_area_slug": self.afghanistan.slug,
                "program_slug": self.program_active.slug,
                "pk": pp_id,
            },
        )
        response = self.client.delete(delete_url)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_204_NO_CONTENT:
            assert response.status_code == status.HTTP_204_NO_CONTENT
            assert PaymentPlan.objects.filter(name="new pp for delete test").count() == 1
            assert PaymentPlan.objects.filter(name="new pp for delete test").first().status == PaymentPlan.Status.DRAFT

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            (
                [Permissions.PM_EXCLUDE_BENEFICIARIES_FROM_FOLLOW_UP_PP],
                status.HTTP_200_OK,
            ),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_exclude_beneficiaries(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.LOCKED
        self.pp.save()
        data = {
            "excluded_households_ids": ["HH-1", "HH-2"],
            "exclusion_reason": "Test Reason",
        }
        response = self.client.post(self.url_exclude_hh, data, format="json")

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert resp_data["background_action_status"] == "EXCLUDE_BENEFICIARIES"

    def test_exclude_beneficiaries_validation_errors(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_EXCLUDE_BENEFICIARIES_FROM_FOLLOW_UP_PP],
            self.afghanistan,
            self.program_active,
        )

        response_1 = self.client.post(self.url_exclude_hh, {"excluded_households_ids": ["HH-1"]}, format="json")
        assert response_1.status_code == status.HTTP_400_BAD_REQUEST
        assert "Beneficiary can be excluded only for 'Open' or 'Locked' status of Payment Plan" in response_1.data

        self.pp.status = PaymentPlan.Status.LOCKED
        self.pp.save()
        response_2 = self.client.post(
            self.url_exclude_hh,
            {"excluded_households_ids": ["HH-1", "HH-1"]},
            format="json",
        )
        assert response_2.status_code == status.HTTP_400_BAD_REQUEST
        assert "Duplicate IDs are not allowed." in response_2.data["excluded_households_ids"][0]

        response_3 = self.client.post(self.url_exclude_hh, {}, format="json")
        assert response_3.status_code == status.HTTP_400_BAD_REQUEST
        assert "excluded_households_ids" in response_3.json()

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            (
                [Permissions.PM_APPLY_RULE_ENGINE_FORMULA_WITH_ENTITLEMENTS],
                status.HTTP_200_OK,
            ),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_apply_engine_formula_pp(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        rule_for_pp = RuleCommitFactory(rule__type=Rule.TYPE_PAYMENT_PLAN, version=11).rule
        self.pp.status = PaymentPlan.Status.LOCKED
        self.pp.save()
        self.pp.refresh_from_db()
        data = {
            "engine_formula_rule_id": str(rule_for_pp.pk),
            "version": self.pp.version,
        }
        response = self.client.post(self.url_apply_steficon, data, format="json")

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            resp_data = response.json()
            assert "id" in resp_data
            assert "RULE_ENGINE_RUN" in resp_data["background_action_status"]

    def test_apply_engine_formula_tp_validation_errors(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_APPLY_RULE_ENGINE_FORMULA_WITH_ENTITLEMENTS],
            self.afghanistan,
            self.program_active,
        )
        rule_for_pp = RuleCommitFactory(rule__type=Rule.TYPE_PAYMENT_PLAN, rule__enabled=False, version=22).rule
        self.pp.status = PaymentPlan.Status.LOCKED
        self.pp.save()

        data = {
            "engine_formula_rule_id": str(rule_for_pp.pk),
            "version": self.pp.version,
        }
        response = self.client.post(self.url_apply_steficon, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "This engine rule is not enabled or is deprecated." in response.data

        self.pp.status = PaymentPlan.Status.TP_OPEN
        self.pp.save()
        data = {
            "engine_formula_rule_id": str(rule_for_pp.pk),
            "version": self.pp.version,
        }
        response_2 = self.client.post(self.url_apply_steficon, data, format="json")

        assert response_2.status_code == status.HTTP_400_BAD_REQUEST
        assert "Not allowed to run engine formula within status TP_OPEN." in response_2.data

        response_3 = self.client.post(self.url_apply_steficon, {}, format="json")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "engine_formula_rule_id" in response_3.data

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_LOCK_AND_UNLOCK_FSP], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pp_fsp_lock(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.LOCKED
        self.pp.financial_service_provider = FinancialServiceProviderFactory()
        self.pp.delivery_mechanism = DeliveryMechanismFactory()
        self.pp.save()
        PaymentFactory(parent=self.pp, entitlement_quantity=999)
        response = self.client.get(self.url_lock_fsp)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json() == {"message": "Payment Plan FSP locked"}

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_LOCK_AND_UNLOCK_FSP], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pp_fsp_unlock(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.LOCKED_FSP
        self.pp.save()

        response = self.client.get(self.url_unlock_fsp)
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json() == {"message": "Payment Plan FSP unlocked"}

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_VIEW_LIST], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pp_entitlement_export_xlsx(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.LOCKED
        self.pp.save()

        response = self.client.get(self.url_export_entitlement_xlsx)
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            self.pp.refresh_from_db()
            assert self.pp.has_export_file is True

    def test_pp_entitlement_export_xlsx_invalid_status(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(self.user, [Permissions.PM_VIEW_LIST], self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.OPEN
        self.pp.save()

        response = self.client.get(self.url_export_entitlement_xlsx)
        assert status.HTTP_400_BAD_REQUEST
        assert "You can only export Payment List for LOCKED Payment Plan" in response.data

    @patch("hope.apps.payment.models.PaymentPlan.get_exchange_rate", return_value=2.0)
    def test_pp_entitlement_import_xlsx(self, mock_exchange_rate: Any, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_IMPORT_XLSX_WITH_ENTITLEMENTS],
            self.afghanistan,
            self.program_active,
        )
        self.pp.status = PaymentPlan.Status.LOCKED
        self.pp.save()
        self.payment_1 = PaymentFactory(
            parent=self.pp,
            status=Payment.STATUS_PENDING,
            currency="PLN",
        )
        self.payment_2 = PaymentFactory(
            unicef_id="RCPT-0060-24-0.000.022",
            parent=self.pp,
            status=Payment.STATUS_PENDING,
            currency="PLN",
        )
        self.payment_1.unicef_id = "RCPT-0060-24-0.000.011"
        self.payment_1.save()
        self.payment_2.unicef_id = "RCPT-0060-24-0.000.022"
        self.payment_2.save()

        file = BytesIO(Path(f"{settings.TESTS_ROOT}/apps/payment/test_file/pp_entitlement_valid.xlsx").read_bytes())
        file.name = "pp_entitlement_valid.xlsx"
        response = self.client.post(self.url_import_entitlement_xlsx, {"file": file}, format="multipart")

        assert response.status_code == status.HTTP_200_OK
        if response.status_code == status.HTTP_200_OK:
            pp = response.json()
            assert pp["background_action_status"] == "XLSX_IMPORTING_ENTITLEMENTS"
            assert pp["imported_file_name"].startswith("pp_entitlement_valid") is True

    def test_pp_entitlement_import_xlsx_status_invalid(self, create_user_role_with_permissions: Any) -> None:
        self.pp.status = PaymentPlan.Status.OPEN
        self.pp.save()
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_IMPORT_XLSX_WITH_ENTITLEMENTS],
            self.afghanistan,
            self.program_active,
        )
        test_file = SimpleUploadedFile("test.xlsx", b"123", content_type="application/vnd.ms-excel")
        response = self.client.post(self.url_import_entitlement_xlsx, {"file": test_file}, format="multipart")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "User can only import for LOCKED Payment Plan" in response.data[0]

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_SEND_FOR_APPROVAL], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_send_for_approval(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.LOCKED_FSP
        self.pp.save()
        response = self.client.get(self.url_send_for_approval)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json()["status"] == "IN_APPROVAL"

    @pytest.mark.parametrize(
        ("permissions", "expected_status", "payment_plan_status"),
        [
            (
                [Permissions.PM_ACCEPTANCE_PROCESS_APPROVE, Permissions.PM_VIEW_LIST],
                status.HTTP_200_OK,
                PaymentPlan.Status.IN_APPROVAL,
            ),
            (
                [Permissions.PM_ACCEPTANCE_PROCESS_AUTHORIZE, Permissions.PM_VIEW_LIST],
                status.HTTP_200_OK,
                PaymentPlan.Status.IN_AUTHORIZATION,
            ),
            (
                [
                    Permissions.PM_ACCEPTANCE_PROCESS_FINANCIAL_REVIEW,
                    Permissions.PM_VIEW_LIST,
                ],
                status.HTTP_200_OK,
                PaymentPlan.Status.IN_REVIEW,
            ),
            ([], status.HTTP_403_FORBIDDEN, PaymentPlan.Status.IN_APPROVAL),
        ],
    )
    def test_approval_process_reject(
        self,
        permissions: List,
        expected_status: int,
        payment_plan_status: PaymentPlan.Status,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        ApprovalProcessFactory(payment_plan=self.pp)
        self.pp.status = payment_plan_status
        self.pp.save()
        response = self.client.post(self.url_approval_process_reject, {"comment": "test123"}, format="json")
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json()["status"] == "LOCKED_FSP"

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            (
                [Permissions.PM_ACCEPTANCE_PROCESS_APPROVE],
                status.HTTP_200_OK,
            ),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_approval_process_approve(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        ApprovalProcessFactory(payment_plan=self.pp)
        self.pp.status = PaymentPlan.Status.IN_APPROVAL
        self.pp.save()
        response = self.client.post(self.url_approval_process_approve, {"comment": "test123"}, format="json")
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json()["status"] == "IN_AUTHORIZATION"

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            (
                [Permissions.PM_ACCEPTANCE_PROCESS_AUTHORIZE],
                status.HTTP_200_OK,
            ),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_approval_process_authorize(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        ApprovalProcessFactory(payment_plan=self.pp)
        self.pp.status = PaymentPlan.Status.IN_AUTHORIZATION
        self.pp.save()
        response = self.client.post(self.url_approval_process_authorize, {"comment": "test123"}, format="json")
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json()["status"] == "IN_REVIEW"

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            (
                [Permissions.PM_ACCEPTANCE_PROCESS_FINANCIAL_REVIEW],
                status.HTTP_200_OK,
            ),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_approval_process_mark_as_released(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        ApprovalProcessFactory(payment_plan=self.pp)
        self.pp.status = PaymentPlan.Status.IN_REVIEW
        self.pp.save()
        response = self.client.post(
            self.url_approval_process_mark_as_released,
            {"comment": "test123"},
            format="json",
        )
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json()["status"] == "ACCEPTED"

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_SEND_TO_PAYMENT_GATEWAY], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_pp_send_to_payment_gateway(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        fsp = FinancialServiceProviderFactory(
            communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_API,
            payment_gateway_id="123",
        )
        PaymentPlanSplitFactory(payment_plan=self.pp)
        self.pp.status = PaymentPlan.Status.ACCEPTED
        self.pp.financial_service_provider = fsp
        self.pp.save()
        PaymentFactory(parent=self.pp)
        response = self.client.get(self.url_send_to_payment_gate_way)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.json()["status"] == "ACCEPTED"

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_DOWNLOAD_FSP_AUTH_CODE], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_generate_xlsx_with_auth_code(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        fsp_xlsx_template_id = FinancialServiceProviderXlsxTemplateFactory().pk
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        test_file = FileTemp.objects.create(
            object_id=self.pp.pk,
            content_type=get_content_type_for_model(self.pp),
            created_by=self.user,
        )
        fsp = FinancialServiceProviderFactory(
            communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_API,
            payment_gateway_id="123",
        )
        PaymentPlanSplitFactory(payment_plan=self.pp)
        self.pp.status = PaymentPlan.Status.IN_APPROVAL
        self.pp.financial_service_provider = fsp
        self.pp.save()
        response = self.client.post(
            self.url_generate_xlsx_with_auth_code,
            {"fsp_xlsx_template_id": fsp_xlsx_template_id},
            format="json",
        )

        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_400_BAD_REQUEST
            assert (
                "Payment List Per FSP export is only available for ACCEPTED or FINISHED Payment Plans." in response.data
            )

            self.pp.status = PaymentPlan.Status.ACCEPTED
            self.pp.export_file_per_fsp = test_file
            self.pp.save()
            response_2 = self.client.post(
                self.url_generate_xlsx_with_auth_code,
                {"fsp_xlsx_template_id": fsp_xlsx_template_id},
                format="json",
            )

            assert response_2.status_code == status.HTTP_400_BAD_REQUEST
            assert "Export failed: Payment Plan already has created exported file." in response_2.data

            self.pp.export_file_per_fsp = None
            self.pp.save()
            payment = PaymentFactory(parent=self.pp, status=Payment.STATUS_PENDING)
            response_3 = self.client.post(
                self.url_generate_xlsx_with_auth_code,
                {"fsp_xlsx_template_id": fsp_xlsx_template_id},
                format="json",
            )

            assert response_3.status_code == status.HTTP_400_BAD_REQUEST
            assert (
                "Export failed: There could be not Pending Payments and FSP communication channel should be set to API."
                in response_3.data
            )

            # success
            payment.status = Payment.STATUS_SENT_TO_PG
            payment.save()
            response_ok = self.client.post(
                self.url_generate_xlsx_with_auth_code,
                {"fsp_xlsx_template_id": fsp_xlsx_template_id},
                format="json",
            )

            assert response_ok.status_code == status.HTTP_200_OK

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_SEND_XLSX_PASSWORD], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_send_xlsx_password(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.ACCEPTED
        self.pp.save()

        response = self.client.get(self.url_send_xlsx_password)
        assert response.status_code == expected_status

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_VIEW_LIST], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_reconciliation_export_xlsx(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.ACCEPTED
        self.pp.save()
        PaymentFactory(parent=self.pp, status=Payment.STATUS_PENDING)

        response = self.client.get(self.url_reconciliation_export_xlsx)
        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_200_OK
            assert "id" in response.data

            self.pp.eligible_payments.delete()
            response_1 = self.client.get(self.url_reconciliation_export_xlsx)
            assert response_1.status_code == status.HTTP_400_BAD_REQUEST
            assert "Export failed: The Payment List is empty." in response_1.data

            self.pp.status = PaymentPlan.Status.IN_APPROVAL
            self.pp.save()
            response_2 = self.client.get(self.url_reconciliation_export_xlsx)
            assert response_2.status_code == status.HTTP_400_BAD_REQUEST
            assert (
                "Payment List Per FSP export is only available for ACCEPTED or FINISHED Payment Plans."
                in response_2.data
            )

    def test_pp_reconciliation_import_xlsx_invalid(self, create_user_role_with_permissions: Any) -> None:
        self.pp.status = PaymentPlan.Status.OPEN
        self.pp.save()
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_IMPORT_XLSX_WITH_RECONCILIATION],
            self.afghanistan,
            self.program_active,
        )
        test_file = SimpleUploadedFile("test.xlsx", b"123", content_type="application/vnd.ms-excel")
        response = self.client.post(self.url_reconciliation_import_xlsx, {"file": test_file}, format="multipart")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert (
            "Payment List Per FSP export is only available for ACCEPTED or FINISHED Payment Plans." in response.data[0]
        )

        fsp_api = FinancialServiceProviderFactory(
            communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_API,
            payment_gateway_id="123",
        )
        self.pp.status = PaymentPlan.Status.ACCEPTED
        self.pp.financial_service_provider = fsp_api
        self.pp.save()
        response_2 = self.client.post(self.url_reconciliation_import_xlsx, {"file": test_file}, format="multipart")
        assert response_2.status_code == status.HTTP_400_BAD_REQUEST
        assert (
            "Only for FSP with Communication Channel XLSX can be imported reconciliation manually."
            in response_2.data[0]
        )

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_SPLIT], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_split(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        fsp = FinancialServiceProviderFactory(
            communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_API,
            payment_gateway_id="123",
        )
        split = PaymentPlanSplitFactory(payment_plan=self.pp, sent_to_payment_gateway=True)
        self.pp.status = PaymentPlan.Status.IN_APPROVAL
        self.pp.financial_service_provider = fsp
        self.pp.save()
        data = {"payments_no": 1, "split_type": PaymentPlanSplit.SplitType.BY_RECORDS}
        response = self.client.post(self.url_pp_split, data, format="json")

        if expected_status == status.HTTP_200_OK:
            assert response.status_code == status.HTTP_400_BAD_REQUEST
            assert "Payment plan is already sent to payment gateway" in response.data

            split.sent_to_payment_gateway = False
            split.save()
            response_2 = self.client.post(self.url_pp_split, data, format="json")
            assert response_2.status_code == status.HTTP_400_BAD_REQUEST
            assert "Payment plan must be accepted to make a split" in response_2.data

            self.pp.status = PaymentPlan.Status.ACCEPTED
            self.pp.save()
            self.pp.eligible_payments.delete()
            response_3 = self.client.post(
                self.url_pp_split,
                {"split_type": PaymentPlanSplit.SplitType.BY_RECORDS},
                format="json",
            )
            assert response_3.status_code == status.HTTP_400_BAD_REQUEST
            assert "Payment Number is required for split by records" in response_3.data

            fsp_api = FinancialServiceProviderFactory(
                communication_channel=FinancialServiceProvider.COMMUNICATION_CHANNEL_API,
                payment_gateway_id="123",
            )
            PaymentFactory.create_batch(
                51,
                parent=self.pp,
                status=Payment.STATUS_PENDING,
                financial_service_provider=fsp_api,
            )
            response_4 = self.client.post(self.url_pp_split, data, format="json")
            assert response_4.status_code == status.HTTP_400_BAD_REQUEST
            assert "Cannot split Payment Plan into more than 50 parts" in response_4.data

            # success
            response_ok = self.client.post(
                self.url_pp_split,
                {
                    "payments_no": 30,
                    "split_type": PaymentPlanSplit.SplitType.BY_RECORDS,
                },
                format="json",
            )
            assert response_ok.status_code == status.HTTP_200_OK
            assert "id" in response_ok.data

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_EXPORT_PDF_SUMMARY], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_export_pdf_payment_plan_summary(
        self,
        permissions: List,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.LOCKED
        self.pp.save()
        PaymentFactory(parent=self.pp)
        response = self.client.get(self.url_export_pdf_payment_plan_summary)

        assert response.status_code == expected_status
        if expected_status == status.HTTP_200_OK:
            assert "id" in response.json()

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_CREATE], status.HTTP_201_CREATED),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_create_follow_up(
        self,
        permissions: list,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        PaymentFactory(parent=self.pp, status=Payment.STATUS_FORCE_FAILED)

        response = self.client.post(
            self.url_create_follow_up,
            {
                "dispersion_start_date": "2024-01-01",
                "dispersion_end_date": "2026-01-01",
            },
            format="json",
        )
        assert response.status_code == expected_status

        if expected_status == status.HTTP_201_CREATED:
            assert "id" in response.json()
            assert response.json()["is_follow_up"] is True
            assert "id" in response.json()["source_payment_plan"]
            assert response.json()["name"] == "DRAFT PP Follow Up"
            assert response.json()["dispersion_start_date"] == "2024-01-01"
            assert response.json()["dispersion_end_date"] == "2026-01-01"
            assert response.json()["currency"] == self.pp.currency

    @pytest.mark.parametrize(
        ("permissions", "expected_status"),
        [
            ([Permissions.PM_ASSIGN_FUNDS_COMMITMENTS], status.HTTP_200_OK),
            ([], status.HTTP_403_FORBIDDEN),
        ],
    )
    def test_assign_funds_commitments(
        self,
        permissions: list,
        expected_status: int,
        create_user_role_with_permissions: Any,
    ) -> None:
        create_user_role_with_permissions(self.user, permissions, self.afghanistan, self.program_active)
        self.pp.status = PaymentPlan.Status.IN_REVIEW
        self.pp.save()

        group = FundsCommitmentGroup.objects.create()

        funds_commitment_item = FundsCommitmentItem.objects.create(
            funds_commitment_group_id=group.pk,
            office=self.afghanistan,
            rec_serial_number="999",
            payment_plan=None,
        )
        assert funds_commitment_item.payment_plan is None

        response = self.client.post(
            self.url_funds_commitments,
            {"fund_commitment_items_ids": ["999"]},
            format="json",
        )
        assert response.status_code == expected_status

        if expected_status == status.HTTP_200_OK:
            assert "id" in response.json()
            funds_commitment_item.refresh_from_db()
            assert funds_commitment_item.payment_plan_id == self.pp.pk

    def test_assign_funds_commitments_validation_errors(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_ASSIGN_FUNDS_COMMITMENTS],
            self.afghanistan,
            self.program_active,
        )

        response = self.client.post(
            self.url_funds_commitments,
            {"fund_commitment_items_ids": ["333"]},
            format="json",
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Payment plan must be in review" in response.json()

        self.pp.status = PaymentPlan.Status.IN_REVIEW
        self.pp.save()
        other_pp = PaymentPlanFactory(
            business_area=self.afghanistan,
            program_cycle=self.cycle,
            status=PaymentPlan.Status.DRAFT,
            created_by=self.user,
        )
        group = FundsCommitmentGroup.objects.create()
        FundsCommitmentItem.objects.create(
            funds_commitment_group_id=group.pk,
            office=None,
            rec_serial_number="333",
            payment_plan=other_pp,
        )

        response = self.client.post(
            self.url_funds_commitments,
            {"fund_commitment_items_ids": ["333"]},
            format="json",
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Chosen Funds Commitments are already assigned to different Payment Plan" in response.json()

        FundsCommitmentItem.objects.create(
            funds_commitment_group_id=group.pk,
            office=None,
            rec_serial_number="2355",
            payment_plan=None,
        )
        response = self.client.post(
            self.url_funds_commitments,
            {"fund_commitment_items_ids": ["2355"]},
            format="json",
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Chosen Funds Commitments have wrong Business Area" in response.json()

    def test_fsp_xlsx_template_list(self, create_user_role_with_permissions: Any) -> None:
        create_user_role_with_permissions(
            self.user,
            [Permissions.PM_EXPORT_XLSX_FOR_FSP],
            self.afghanistan,
            self.program_active,
        )

        xlsx_1 = FinancialServiceProviderXlsxTemplateFactory(name="XLSX_1")
        xlsx_2 = FinancialServiceProviderXlsxTemplateFactory(name="XLSX_2")
        xlsx_3 = FinancialServiceProviderXlsxTemplateFactory(name="Other BA")

        financial_service_provider = FinancialServiceProviderFactory()
        financial_service_provider.allowed_business_areas.set([self.afghanistan])
        fsp_2 = FinancialServiceProviderFactory()

        financial_service_provider.xlsx_templates.set([xlsx_1, xlsx_2])
        fsp_2.xlsx_templates.set([xlsx_3])

        response = self.client.get(
            reverse(
                "api:payments:payment-plans-fsp-xlsx-template-list",
                kwargs=self.url_kwargs_ba_program,
            ),
            {"fund_commitment_items_ids": ["333"]},
            format="json",
        )
        assert response.status_code == status.HTTP_200_OK
        assert len(response.json()) == 2
        assert response.json()[0]["name"] == "XLSX_1"
        assert response.json()[1]["name"] == "XLSX_2"
