import operator
from pathlib import Path
from typing import Any
from unittest import mock

import openpyxl
import pytest

from extras.test_utils.factories.core import (
    BeneficiaryGroupFactory,
    BusinessAreaFactory,
    DataCollectingTypeFactory,
    FlexibleAttributeFactory,
    FlexibleAttributeForPDUFactory,
    PeriodicFieldDataFactory,
)
from extras.test_utils.factories.geo import AreaFactory, AreaTypeFactory, CountryFactory
from extras.test_utils.factories.program import ProgramFactory
from hope.apps.core.field_attributes.core_fields_attributes import TYPE_SELECT_MANY, TYPE_SELECT_ONE
from hope.apps.core.utils import SheetImageLoader
from hope.apps.registration_data.validators import (
    KoboProjectImportDataInstanceValidator,
    UploadXLSXInstanceValidator,
)
from hope.models import DataCollectingType, FlexibleAttribute, PeriodicFieldData

pytestmark = [pytest.mark.django_db, pytest.mark.usefixtures("mock_elasticsearch")]

FILES_DIR = Path(__file__).resolve().parent / "test_file"


@pytest.fixture
def business_area() -> Any:
    return BusinessAreaFactory(name="Afghanistan", long_name="Afghanistan")


@pytest.fixture
def countries(business_area: Any) -> dict[str, Any]:
    afghanistan = CountryFactory(
        name="Afghanistan",
        short_name="Afghanistan",
        iso_code2="AF",
        iso_code3="AFG",
        iso_num="0004",
    )
    CountryFactory(name="Isle of Man", short_name="Isle of Man", iso_code2="IM", iso_code3="IMN", iso_num="0833")
    CountryFactory(name="Poland", short_name="Poland", iso_code2="PL", iso_code3="POL", iso_num="0616")
    CountryFactory(name="Palestine", short_name="Palestine", iso_code2="PS", iso_code3="PSE", iso_num="0275")
    CountryFactory(name="San Marino", short_name="San Marino", iso_code2="SM", iso_code3="SMR", iso_num="0674")
    CountryFactory(
        name="Saint Vincent",
        short_name="Saint Vincent",
        iso_code2="VC",
        iso_code3="VCT",
        iso_num="0670",
    )
    business_area.countries.add(afghanistan)
    return {"afghanistan": afghanistan}


@pytest.fixture
def afghanistan_admin_areas(countries: dict[str, Any]) -> list[Any]:
    area_type = AreaTypeFactory(country=countries["afghanistan"], name="admin", area_level=1)
    return [
        AreaFactory(p_code="AF29", area_type=area_type),
        AreaFactory(p_code="AF2401", area_type=area_type),
        AreaFactory(p_code="AF02", area_type=area_type),
        AreaFactory(p_code="AF1524", area_type=area_type),
        AreaFactory(p_code="AF31", area_type=area_type),
        AreaFactory(p_code="AF0619", area_type=area_type),
        AreaFactory(p_code="AF0201", area_type=area_type),
        AreaFactory(p_code="AF11", area_type=area_type),
        AreaFactory(p_code="AF1115", area_type=area_type),
    ]


@pytest.fixture
def assistance_type_flex_attribute() -> Any:
    return FlexibleAttributeFactory(
        name="assistance_type_h_f",
        type=FlexibleAttribute.SELECT_MANY,
        associated_with=FlexibleAttribute.ASSOCIATED_WITH_HOUSEHOLD,
        label={"English(EN)": "Assistance Type"},
    )


@pytest.fixture
def program(
    business_area: Any,
    countries: dict[str, Any],
    afghanistan_admin_areas: list[Any],
    assistance_type_flex_attribute: Any,
) -> Any:
    return ProgramFactory(business_area=business_area)


@pytest.fixture
def social_worker_program(
    business_area: Any,
    countries: dict[str, Any],
    afghanistan_admin_areas: list[Any],
) -> Any:
    data_collecting_type = DataCollectingTypeFactory(type=DataCollectingType.Type.SOCIAL)
    beneficiary_group = BeneficiaryGroupFactory(master_detail=False)
    return ProgramFactory(
        business_area=business_area,
        data_collecting_type=data_collecting_type,
        beneficiary_group=beneficiary_group,
    )


@pytest.fixture
def pdu_attribute_factory(program: Any):
    def _create(subtype: str) -> Any:
        pdu_data = PeriodicFieldDataFactory(subtype=subtype, number_of_rounds=1, rounds_names=["May"])
        return FlexibleAttributeForPDUFactory(
            label="PDU Flex Attribute",
            program=program,
            pdu_data=pdu_data,
        )

    return _create


def test_string_validator(program: Any, countries: dict[str, Any], afghanistan_admin_areas: list[Any]) -> None:
    validator = UploadXLSXInstanceValidator(program)
    assert validator.string_validator("Marek", "full_name_i_c")


@pytest.mark.parametrize(
    ("value", "header", "expected"),
    [
        (None, "estimated_birth_date_i_c", False),
        (None, "age_at_registration", True),
        (1.1, "estimated_birth_date_i_c", True),
        ("1.a1a", "estimated_birth_date_i_c", False),
    ],
)
def test_float_validator(program: Any, value: Any, header: str, expected: bool) -> None:
    validator = UploadXLSXInstanceValidator(program)
    assert validator.float_validator(value, header) is expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("1.1, 1.1", True),
        ("0.0, 0.0", True),
        ("54.1234252, 67.535232", True),
        ("1, 1, 1, 1", False),
        ("0, 0", False),
        ("52.124.124, 1241.242", False),
        ("24.121a, bcd421.222", False),
    ],
)
def test_geolocation_validator(program: Any, value: str, expected: bool) -> None:
    validator = UploadXLSXInstanceValidator(program)
    assert validator.geolocation_validator(value, "hh_geopoint_h_c") is expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("01-03-1994", True),
        ("1-3-1994", True),
        ("27-12-2020", True),
        ("27/12/2020", True),
        ("27.12.2020", True),
        ("13-13-1994", False),
        ("213.22.2020", False),
        ("qwerty", False),
        ("24", False),
        ("-24", False),
    ],
)
def test_date_validator(program: Any, value: str, expected: bool) -> None:
    validator = UploadXLSXInstanceValidator(program)
    assert validator.date_validator(value, "birth_date_i_c") is expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("12", True),
        ("0", True),
        (0, True),
        (12, True),
        (12345, True),
        (-12, True),
        ("13-13-1994", False),
        ("213.22.2020", False),
        ("qwerty", False),
        ("12,242", False),
    ],
)
def test_integer_validator(program: Any, value: Any, expected: bool) -> None:
    validator = UploadXLSXInstanceValidator(program)
    assert validator.integer_validator(value, "size_h_c") is expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("+1-202-555-0172", True),
        ("+44 1632 960852", True),
        ("+1-613-555-0182", True),
        ("+61 1900 654 321", True),
        ("+36 55 979 922", True),
        ("+353 20 915 8245", True),
        ("+48 69 563 7300", True),
        ("123 123 123", False),
        ("13-13-1994", False),
        ("213.22.2020", False),
        ("qwerty", False),
        (12.2345, False),
        ("12,242", False),
        ("12", False),
        (12, False),
    ],
)
def test_phone_validator(program: Any, value: Any, expected: bool) -> None:
    validator = UploadXLSXInstanceValidator(program)
    assert validator.phone_validator(value, "phone_no_i_c") is expected


@pytest.mark.parametrize(
    ("value", "header", "expected"),
    [
        ("REFUGEE", "residence_status_h_c", True),
        ("GOVERNMENT_PARTNER,UNICEF", "consent_sharing_h_c", True),
        ("GOVERNMENT_PARTNER, UNICEF", "consent_sharing_h_c", True),
        ("GOVERNMENT_PARTNER;UNICEF", "consent_sharing_h_c", True),
        ("GOVERNMENT_PARTNER; UNICEF", "consent_sharing_h_c", True),
        ("GOVERNMENT_PARTNER UNICEF", "consent_sharing_h_c", True),
        ("YES", "work_status", False),
        ("OTHER", "work_status", False),
        ("Hearing Problems", "disability", False),
        ("Wrong Option", "assistance_type_h_f", False),
        ("YES", "consent_h_c", False),
    ],
)
def test_choice_validator(
    program: Any,
    assistance_type_flex_attribute: Any,
    value: str,
    header: str,
    expected: bool,
) -> None:
    validator = UploadXLSXInstanceValidator(program)
    assert validator.choice_validator(value, header) is expected


def test_import_choice_validator_empty_value_returns_message(program: Any) -> None:
    validator = KoboProjectImportDataInstanceValidator(program)
    validator.all_fields = {"field_one": {"type": TYPE_SELECT_ONE, "choices": ["A"]}}

    assert validator.choice_validator("", "field_one") == "Invalid choice  for field field_one"


def test_import_choice_validator_handler_none_returns_none(program: Any) -> None:
    validator = KoboProjectImportDataInstanceValidator(program)
    validator.all_fields = {"field_one": {"type": "UNKNOWN", "choices": ["A"]}}

    assert validator.choice_validator("A", "field_one") is None


def test_import_choice_validator_custom_select_one(program: Any) -> None:
    validator = KoboProjectImportDataInstanceValidator(program)
    validator.all_fields = {
        "field_one": {
            "type": TYPE_SELECT_ONE,
            "choices": ["YES", "NO"],
            "custom_validate_choices": lambda v: v == "YES",
        }
    }

    assert validator.choice_validator("YES", "field_one") is None
    assert validator.choice_validator("NO", "field_one") == "Invalid choice NO for field field_one"


def test_import_choice_validator_custom_select_many(program: Any) -> None:
    validator = KoboProjectImportDataInstanceValidator(program)
    validator.all_fields = {
        "field_many": {
            "type": TYPE_SELECT_MANY,
            "choices": ["A", "B"],
            "custom_validate_choices": lambda v: v == "A",
        }
    }

    assert validator.choice_validator("A", "field_many") is None
    assert validator.choice_validator("B", "field_many") == "Invalid choice B for field field_many"


def test_import_choice_validator_select_many_invalid_choice(program: Any) -> None:
    validator = KoboProjectImportDataInstanceValidator(program)
    validator.all_fields = {
        "field_many": {
            "type": TYPE_SELECT_MANY,
            "choices": ["A", "B"],
        }
    }

    assert validator.choice_validator("C", "field_many") == "Invalid choice C for field field_many"


def test_rows_validator_too_many_head_of_households(
    program: Any,
) -> None:
    wb = openpyxl.load_workbook(
        FILES_DIR / "error-xlsx.xlsx",
        data_only=True,
    )
    validator = UploadXLSXInstanceValidator(program)
    validator.rows_validator(wb["Households"])
    validator.errors = []
    validator.rows_validator(wb["Individuals"])
    expected = [
        {
            "row_number": 0,
            "header": "relationship_i_c",
            "message": "Sheet: 'Individuals', There are multiple head of households for household with id: 3",
        }
    ]
    assert expected == validator.errors


def test_rows_validator(
    program: Any,
    business_area: Any,
) -> None:
    wb = openpyxl.load_workbook(
        FILES_DIR / "invalid_rows.xlsx",
        data_only=True,
    )

    wb_valid = openpyxl.load_workbook(
        FILES_DIR / "new_reg_data_import.xlsx",
        data_only=True,
    )

    invalid_file = (
        (
            wb["Households"],
            [
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 1 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 3,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 1, Option 2, Option "
                    "3 for type select many of field assistance_type_h_f",
                    "row_number": 4,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 13 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 5,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 3 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 6,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 1, Option 3 for type "
                    "select many of field assistance_type_h_f",
                    "row_number": 7,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 2, Option 3 for type "
                    "select many of field assistance_type_h_f",
                    "row_number": 8,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 2 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 9,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 1, Option 2, Option "
                    "4 for type select many of field assistance_type_h_f",
                    "row_number": 10,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 4 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 11,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 5 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 12,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 1, Option 4 for type "
                    "select many of field assistance_type_h_f",
                    "row_number": 13,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 2, Option 4 for type "
                    "select many of field assistance_type_h_f",
                    "row_number": 14,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 3 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 15,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 1, Option 2, Option "
                    "5 for type select many of field assistance_type_h_f",
                    "row_number": 16,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 6 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 17,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 7 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 18,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 1, Option 5 for type "
                    "select many of field assistance_type_h_f",
                    "row_number": 19,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 2, Option 5 for type "
                    "select many of field assistance_type_h_f",
                    "row_number": 20,
                },
                {
                    "header": "assistance_type_h_f",
                    "message": "Sheet: 'Households', Unexpected value: Option 4 for type select "
                    "many of field assistance_type_h_f",
                    "row_number": 21,
                },
            ],
        ),
        (
            wb["Individuals"],
            [
                {
                    "row_number": 4,
                    "header": "preferred_language_i_c",
                    "message": "Sheet: 'Individuals', Unexpected value: TestInvalid for "
                    "type select one of field preferred_language_i_c",
                },
                {
                    "row_number": 8,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', There is no household with provided id: TEXT",
                },
                {
                    "row_number": 29,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', There is no household with provided id: 52",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 34 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 35 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 36 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 37 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 38 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 39 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 40 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 41 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 42 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 43 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 44 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 45 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 46 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 47 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 48 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 49 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 50 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: 51 has to have a head of household",
                },
                {
                    "row_number": 0,
                    "header": "relationship_i_c",
                    "message": "Sheet: 'Individuals', Household with id: Some Text has to have a head of household",
                },
            ],
        ),
    )
    valid_file = (
        (
            wb_valid["Households"],
            [],
        ),
        (
            wb_valid["Individuals"],
            [],
        ),
    )
    files = (invalid_file, valid_file)
    for file in files:
        validator = UploadXLSXInstanceValidator(program)
        for sheet, expected_values in file:
            validator.image_loader = SheetImageLoader(sheet)
            validator.errors = []
            validator.rows_validator(sheet, business_area.slug)
            assert validator.errors == expected_values


def test_validate_file_extension(program: Any) -> None:
    file_path = FILES_DIR / "image.png"
    expected_values = [{"row_number": 1, "message": "Only .xlsx files are accepted for import"}]
    with open(file_path, "rb") as file:
        validator = UploadXLSXInstanceValidator(program)
        validator.validate_file_extension(file)
        assert validator.errors[0]["row_number"] == expected_values[0]["row_number"]
        assert validator.errors[0]["message"] == expected_values[0]["message"]

        validator = UploadXLSXInstanceValidator(program)
        errors = validator.validate_everything(file, "afghanistan")
        assert errors[0]["row_number"] == expected_values[0]["row_number"]
        assert errors[0]["message"] == expected_values[0]["message"]


def test_validate_file_content_as_xlsx(program: Any) -> None:
    file_path = FILES_DIR / "not_excel_file.xlsx"
    expected_values = [{"row_number": 1, "message": "Invalid .xlsx file"}]
    with open(file_path, "rb") as file:
        validator = UploadXLSXInstanceValidator(program)
        result = validator.validate_everything(file, "afghanistan")
        assert result[0]["row_number"] == expected_values[0]["row_number"]
        assert result[0]["message"] == expected_values[0]["message"]


def test_validate_file_with_template(program: Any) -> None:
    invalid_cols_file_path = FILES_DIR / "new_reg_data_import.xlsx"
    with open(invalid_cols_file_path, "rb") as file:
        validator = UploadXLSXInstanceValidator(program)
        wb = openpyxl.load_workbook(file, data_only=True)
        validator.validate_file_with_template(wb)
        errors = validator.errors
        errors.sort(key=operator.itemgetter("row_number", "header"))
        assert errors == []


def test_required_validator(program: Any) -> None:
    with mock.patch(
        "hope.apps.registration_data.validators.UploadXLSXInstanceValidator.get_all_fields",
        return_value={"test": {"required": True}},
    ):
        validator = UploadXLSXInstanceValidator(program)
        result = validator.required_validator(value="tak", header="test")
        assert result

    with mock.patch(
        "hope.apps.registration_data.validators.UploadXLSXInstanceValidator.get_all_fields",
        return_value={"test": {"required": True}},
    ):
        validator = UploadXLSXInstanceValidator(program)
        result = validator.required_validator(value="", header="test")
        assert not result

    with mock.patch(
        "hope.apps.registration_data.validators.UploadXLSXInstanceValidator.get_all_fields",
        return_value={"test": {"required": False}},
    ):
        validator = UploadXLSXInstanceValidator(program)
        result = validator.required_validator(value="", header="test")
        assert result


def test_validate_empty_file(program: Any) -> None:
    empty_file_path = FILES_DIR / "empty_rdi.xlsx"
    wb = openpyxl.load_workbook(
        empty_file_path,
        data_only=True,
    )
    validator = UploadXLSXInstanceValidator(program)

    expected_result = [
        {
            "header": "Households",
            "message": "There aren't households in the file.",
            "row_number": 1,
        },
        {
            "header": "Individuals",
            "message": "There aren't individuals in the file.",
            "row_number": 1,
        },
    ]

    validator.validate_collectors_size(wb)
    assert validator.errors == expected_result


def test_validate_collector_unique(program: Any) -> None:
    file_path = FILES_DIR / "test_collectors.xlsx"

    expected_result = [
        {
            "row_number": 3,
            "header": "Individuals",
            "message": "Individual from row: 3 cannot be the primary and the alternate collector for households: "
            "992630574 at the same time.",
        },
        {
            "row_number": 4,
            "header": "Individuals",
            "message": "Individual from row: 4 cannot be the primary and the alternate collector for households: "
            "853780211 at the same time.",
        },
    ]

    with open(file_path, "rb") as file:
        validator = UploadXLSXInstanceValidator(program)
        result = validator.validate_everything(file, "afghanistan")
    assert result == expected_result


def test_validate_incorrect_admin_area(program: Any) -> None:
    file_path = FILES_DIR / "invalid_area.xlsx"

    expected_result = [
        {
            "header": "admin1_h_c",
            "row_number": 3,
            "message": "Sheet: 'Households': Area with code: F-35 does not exist",
        },
        {
            "header": "admin2_h_c",
            "row_number": 3,
            "message": "Sheet: 'Households': Area with code: F-35 does not exist",
        },
        {
            "header": "admin1_h_c",
            "row_number": 4,
            "message": "Sheet: 'Households': Area with code: F-35 does not exist",
        },
        {
            "header": "admin2_h_c",
            "row_number": 4,
            "message": "Sheet: 'Households': Area with code: F-35 does not exist",
        },
        {
            "header": "admin1_h_c",
            "row_number": 6,
            "message": "Sheet: 'Households': Area with code: F-35 does not exist",
        },
        {
            "header": "admin2_h_c",
            "row_number": 6,
            "message": "Sheet: 'Households': Area with code: F-35 does not exist",
        },
    ]

    with open(file_path, "rb") as file:
        validator = UploadXLSXInstanceValidator(program)
        result = validator.validate_everything(file, "afghanistan")
    assert result == expected_result


def test_validate_people_sheet_invalid(
    social_worker_program: Any,
) -> None:
    file_path = FILES_DIR / "rdi_people_test_invalid.xlsx"

    expected_result = [
        {
            "row_number": 1,
            "header": "People",
            "message": "There are duplicates with id(s): [1]. Number have to be unique in the field pp_index_id.",
        },
        {
            "row_number": 1,
            "header": "People",
            "message": "Invalid value in field 'pp_primary_collector_id' for Individual with index_id 1. "
            "Value cannot be empty for relationship NON_BENEFICIARY",
        },
        {
            "row_number": 1,
            "header": "People",
            "message": "Individual with index_id 1 has be collector for somebody.",
        },
        {
            "row_number": 1,
            "header": "People",
            "message": "Invalid value in field 'pp_relationship_i_c' with index_id 99. Value can be HEAD or"
            " NON_BENEFICIARY",
        },
        {
            "row_number": 3,
            "header": "pp_index_id",
            "message": "Sheet: 'People', Unexpected value: None for type integer of field pp_index_id",
        },
    ]
    with open(file_path, "rb") as file:
        validator = UploadXLSXInstanceValidator(social_worker_program)
        result = validator.validate_everything(file, "afghanistan")
    assert result == expected_result


def test_validate_people_sheet_valid(
    social_worker_program: Any,
) -> None:
    file_path = FILES_DIR / "rdi_people_test.xlsx"

    with open(file_path, "rb") as file:
        validator = UploadXLSXInstanceValidator(social_worker_program)
        result = validator.validate_everything(file, "afghanistan")
    assert result == []


@pytest.mark.parametrize(
    ("subtype", "data_row"),
    [
        (PeriodicFieldData.STRING, ["Test", "2021-05-01"]),
        (PeriodicFieldData.DECIMAL, ["12.3", "2021-05-01"]),
        (PeriodicFieldData.BOOL, ["True", "2021-05-01"]),
        (PeriodicFieldData.DATE, ["1996-06-21", "2021-05-01"]),
    ],
)
def test_validate_pdu_string_valid(
    program: Any,
    pdu_attribute_factory: Any,
    subtype: str,
    data_row: list,
) -> None:
    pdu_attribute_factory(subtype)
    header_row = [
        "pdu_flex_attribute_round_1_value",
        "pdu_flex_attribute_round_1_collection_date",
    ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(header_row)
    sheet.append(data_row)
    validator = UploadXLSXInstanceValidator(program)
    errors = validator._validate_pdu(sheet[2], sheet[1], 3)
    assert errors == []


@pytest.mark.parametrize(
    ("subtype", "data_row"),
    [
        (PeriodicFieldData.DECIMAL, ["foo", "2021-05-01"]),
        (PeriodicFieldData.BOOL, ["foo", "2021-05-01"]),
        (PeriodicFieldData.DATE, ["foo", "2021-05-01"]),
    ],
)
def test_validate_pdu_string_value_error(
    program: Any,
    pdu_attribute_factory: Any,
    subtype: str,
    data_row: list,
) -> None:
    pdu_attribute_factory(subtype)
    header_row = [
        "pdu_flex_attribute_round_1_value",
        "pdu_flex_attribute_round_1_collection_date",
    ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(header_row)
    sheet.append(data_row)
    validator = UploadXLSXInstanceValidator(program)
    errors = validator._validate_pdu(sheet[2], sheet[1], 3)
    assert errors == [
        {
            "row_number": 3,
            "header": "pdu_flex_attribute_round_1_value",
            "message": "Invalid value foo for field pdu_flex_attribute_round_1_value",
        }
    ]


def test_validate_pdu_wrong_collection_date(
    program: Any,
    pdu_attribute_factory: Any,
) -> None:
    data_row = ["Test", "bar"]
    pdu_attribute_factory(PeriodicFieldData.STRING)
    header_row = [
        "pdu_flex_attribute_round_1_value",
        "pdu_flex_attribute_round_1_collection_date",
    ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(header_row)
    sheet.append(data_row)
    validator = UploadXLSXInstanceValidator(program)
    errors = validator._validate_pdu(sheet[2], sheet[1], 3)
    assert errors == [
        {
            "row_number": 3,
            "header": "pdu_flex_attribute_round_1_collection_date",
            "message": "Invalid value bar for field pdu_flex_attribute_round_1_collection_date",
        }
    ]


def test_validate_pdu_empty_row(
    program: Any,
    pdu_attribute_factory: Any,
) -> None:
    data_row = ["Test", "bar"]
    pdu_attribute_factory(PeriodicFieldData.STRING)
    header_row = []
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(header_row)
    sheet.append(data_row)
    validator = UploadXLSXInstanceValidator(program)
    errors = validator._validate_pdu(sheet[2], sheet[1], 3)
    assert errors == []
